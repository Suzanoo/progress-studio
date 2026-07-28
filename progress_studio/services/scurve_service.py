from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


@dataclass(frozen=True)
class SCurveData:
    dates: tuple[date | datetime, ...]
    plan: tuple[float, ...]
    actual: tuple[float | None, ...]


class SCurveService:
    """Build chart-ready cumulative values from activity rows in the main sheet.

    The generated workbook intentionally contains Excel formulas for WBS and S-curve
    summaries. openpyxl does not calculate those formulas, so the desktop preview
    calculates the same weighted project curve directly from activity Amount and
    weekly Plan/Actual input cells.
    """

    HEADER_ROW = 4
    WEEK_LABEL_ROW = 3
    FIRST_DATA_ROW = 5

    def read(self, workbook_path: Path, sheet_name: str = "main") -> SCurveData:
        path = workbook_path.expanduser().resolve()
        if not path.is_file():
            raise FileNotFoundError(f"Progress workbook not found: {path}")

        wb = load_workbook(path, data_only=False, read_only=False)
        try:
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"Worksheet not found: {sheet_name}")
            ws = wb[sheet_name]
            headers = self._header_map(ws)
            required = {"row type", "p/a", "amount"}
            missing = sorted(required - headers.keys())
            if missing:
                raise ValueError("Required columns not found: " + ", ".join(missing))

            week_cols = self._week_columns(ws)
            dates = tuple(ws.cell(self.HEADER_ROW, col).value for col in week_cols)
            valid_dates = tuple(value for value in dates if isinstance(value, (date, datetime)))
            if len(valid_dates) != len(week_cols):
                raise ValueError("Weekly date headers are incomplete in the main sheet.")

            row_type_col = headers["row type"]
            pa_col = headers["p/a"]
            amount_col = headers["amount"]

            plan_weighted = [0.0] * len(week_cols)
            actual_weighted = [0.0] * len(week_cols)
            actual_present = [False] * len(week_cols)
            total_amount = 0.0

            for row in range(self.FIRST_DATA_ROW, ws.max_row + 1):
                if self._key(ws.cell(row, row_type_col).value) != "activity":
                    continue
                if self._key(ws.cell(row, pa_col).value) != "p":
                    continue
                amount = self._number(ws.cell(row, amount_col).value)
                if amount is None or amount <= 0:
                    continue
                total_amount += amount
                actual_row = row + 1

                for index, col in enumerate(week_cols):
                    plan_value = self._number(ws.cell(row, col).value)
                    if plan_value is not None:
                        plan_weighted[index] += amount * plan_value

                    if actual_row <= ws.max_row and self._key(ws.cell(actual_row, pa_col).value) == "a":
                        actual_value = self._number(ws.cell(actual_row, col).value)
                        if actual_value is not None:
                            actual_weighted[index] += amount * actual_value
                            actual_present[index] = True

            if total_amount <= 0:
                raise ValueError("No activity Amount values were found for the S-curve preview.")

            plan_weekly = [value / total_amount * 100.0 for value in plan_weighted]
            actual_weekly = [value / total_amount * 100.0 for value in actual_weighted]
            plan = self._cumulative(plan_weekly)

            actual_values: list[float | None] = []
            running = 0.0
            latest_actual = max((i for i, present in enumerate(actual_present) if present), default=-1)
            for index, value in enumerate(actual_weekly):
                if index > latest_actual:
                    actual_values.append(None)
                else:
                    running += value
                    actual_values.append(running if any(actual_present[: index + 1]) else None)

            return SCurveData(valid_dates, tuple(plan), tuple(actual_values))
        finally:
            wb.close()

    def _header_map(self, ws) -> dict[str, int]:
        result: dict[str, int] = {}
        for col in range(1, ws.max_column + 1):
            key = self._key(ws.cell(self.HEADER_ROW, col).value)
            if key:
                result[key] = col
        return result

    def _week_columns(self, ws) -> list[int]:
        result: list[int] = []
        for col in range(1, ws.max_column + 1):
            label = str(ws.cell(self.WEEK_LABEL_ROW, col).value or "").strip().upper()
            if label.startswith("W") and label[1:].isdigit():
                result.append(col)
        if not result:
            raise ValueError("Weekly timescale was not found in the main sheet.")
        return result

    @staticmethod
    def _key(value: object) -> str:
        return "" if value is None else str(value).strip().lower()

    @staticmethod
    def _number(value: object) -> float | None:
        if isinstance(value, bool) or value is None:
            return None
        try:
            return float(value)
        except (TypeError, ValueError):
            return None

    @staticmethod
    def _cumulative(values: list[float]) -> list[float]:
        result: list[float] = []
        running = 0.0
        for value in values:
            running += value
            result.append(running)
        return result
