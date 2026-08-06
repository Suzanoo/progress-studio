from pathlib import Path

from openpyxl import load_workbook

from progress_studio.config import WORKBOOK_SCHEMA
from progress_studio.domain.progress import ProgressBuildResult
from progress_studio.infrastructure.excel.activity_data_theme import apply_activity_data_wbs_hierarchy
from progress_studio.infrastructure.excel.dashboard_workbook import build_dashboard
from progress_studio.infrastructure.excel.progress_workbook import (
    find_sheet,
    prepare_progress_and_scurve,
)


class ProgressService:
    @staticmethod
    def _project_name(ws) -> str | None:
        """Read the project label from the first Project Summary plan row."""
        headers = {
            str(ws.cell(4, col).value or "").strip().lower(): col
            for col in range(1, ws.max_column + 1)
        }
        row_type_col = headers.get("row type")
        description_col = headers.get("description")
        pa_col = headers.get("p/a")
        if row_type_col is None or description_col is None:
            return None
        for row in range(5, ws.max_row + 1):
            row_type = str(ws.cell(row, row_type_col).value or "").strip().lower()
            pa = str(ws.cell(row, pa_col).value or "").strip().upper() if pa_col else "P"
            if row_type == "project summary" and pa == "P":
                value = str(ws.cell(row, description_col).value or "").strip()
                return value or None
        return None

    def build(self, input_file: Path, output_file: Path) -> ProgressBuildResult:
        wb = load_workbook(input_file)
        try:
            ws = find_sheet(wb, WORKBOOK_SCHEMA.main_sheet)
            values = prepare_progress_and_scurve(wb, ws)

            # The workbook is already useful before BOQ mapping. Apply the same
            # Activity Data hierarchy theme used by mapped exports and create the
            # Dashboard as soon as progress/progress_table become available.
            apply_activity_data_wbs_hierarchy(ws)
            build_dashboard(wb, project_name=self._project_name(ws))

            wb.calculation.calcMode = "auto"
            wb.calculation.fullCalcOnLoad = True
            wb.calculation.forceFullCalc = True
            output_file.parent.mkdir(parents=True, exist_ok=True)
            wb.save(output_file)
        finally:
            wb.close()
        return ProgressBuildResult(*values)
