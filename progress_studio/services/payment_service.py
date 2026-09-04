from __future__ import annotations

from pathlib import Path
import os
import shutil
import tempfile

from openpyxl import load_workbook

from progress_studio.domain.payment_models import (
    PaymentInputData,
    PaymentInputResult,
    PaymentInputValidation,
    PaymentLineRenderResult,
    PaymentMultiLineRenderResult,
    PaymentPositionResult,
    PaymentPreparationResult,
    PaymentSnapshotResult,
    PaymentWorkbookValidation,
)
from progress_studio.infrastructure.excel.final_workbook_policy import finalize_workbook
from progress_studio.infrastructure.excel.main_dataset_workbook_adapter import (
    main_dataset_from_workbook,
)
from progress_studio.infrastructure.excel.okd_workbook import (
    build_progress_table_from_source,
)
from progress_studio.infrastructure.excel.payment_breakdown_workbook import (
    render_payment_breakdown,
)
from progress_studio.infrastructure.excel.rebuild_workbook_reader import (
    RebuildWorkbookReader,
)
from progress_studio.infrastructure.excel.traditional_overlay_workbook import (
    reassert_traditional_overlay_transparency,
)
from progress_studio.infrastructure.excel.xlsx_package_preservation import (
    restore_opaque_workbook_parts,
)
from progress_studio.infrastructure.excel.xlsx_package_validator import (
    validate_xlsx_tables,
)
from progress_studio.infrastructure.excel.payment_input_reader import (
    PaymentInputSparseReader,
)
from progress_studio.infrastructure.excel.payment_input_workbook import (
    PaymentInputWorkbook,
)
from progress_studio.infrastructure.excel.payment_line_renderer import (
    PaymentLineRenderer,
)
from progress_studio.infrastructure.excel.payment_progress_index import (
    ActivityProgressIndexReader,
)
from progress_studio.infrastructure.excel.payment_workbook import (
    PaymentWorkbookError,
    PaymentWorkbookSnapshotter,
)
from progress_studio.services.payment_breakdown_adapter import (
    MainDatasetPaymentBreakdownAdapter,
    PaymentBreakdownDatasetSnapshot,
)
from progress_studio.services.payment_position_engine import PaymentPositionEngine


class PaymentService:
    def __init__(
        self,
        snapshotter: PaymentWorkbookSnapshotter | None = None,
        payment_input: PaymentInputWorkbook | None = None,
        payment_reader: PaymentInputSparseReader | None = None,
        progress_index_reader: ActivityProgressIndexReader | None = None,
        position_engine: PaymentPositionEngine | None = None,
        line_renderer: PaymentLineRenderer | None = None,
        breakdown_adapter: MainDatasetPaymentBreakdownAdapter | None = None,
    ) -> None:
        self.snapshotter = snapshotter or PaymentWorkbookSnapshotter()
        self.payment_reader = payment_reader or PaymentInputSparseReader()
        self.payment_input = payment_input or PaymentInputWorkbook(
            reader=self.payment_reader
        )
        self.progress_index_reader = (
            progress_index_reader or ActivityProgressIndexReader()
        )
        self.position_engine = position_engine or PaymentPositionEngine()
        self.line_renderer = line_renderer or PaymentLineRenderer()
        self.breakdown_adapter = (
            breakdown_adapter or MainDatasetPaymentBreakdownAdapter()
        )

    def prepare_embedded_payment_input(
        self,
        source_workbook: Path,
        output_workbook: Path,
        periods: int | None = None,
    ) -> dict[str, int]:
        """Prepare/reconcile Payment Input only; never rebuild Payment or Progress views."""
        source = Path(source_workbook)
        output = Path(output_workbook)
        if not source.is_file():
            raise PaymentWorkbookError(f"Workbook was not found: {source}")
        if output.resolve() != source.resolve():
            output.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(source, output)

        preserved = None
        try:
            preserved = self.payment_reader.read(source)
        except PaymentWorkbookError:
            preserved = None

        wb = load_workbook(output)
        try:
            if "main" not in wb.sheetnames:
                raise PaymentWorkbookError("Worksheet 'main' was not found.")
            stats = self.payment_input.embed(
                wb,
                preserved=preserved,
                periods=periods,
            )
            finalize_workbook(wb, mode="snapshot", include_guide=True)
            wb.save(output)
            return stats
        finally:
            wb.close()

    def prepare_payment_breakdown(
        self,
        source_workbook: Path,
        output_workbook: Path,
    ) -> PaymentBreakdownDatasetSnapshot:
        """Create Payment-Breakdown without rebuilding or mutating existing features.

        The workflow mirrors the proven EV rebuild boundary:
        - read current `main` through the sparse rebuild reader;
        - copy the source workbook to a temporary output;
        - mutate only Payment-Breakdown in one normal workbook pass;
        - preserve opaque OOXML parts that openpyxl cannot round-trip reliably;
        - atomically replace the requested output.

        Existing Earned Value, Dashboard, Payment, overlays, user sheets and
        external-link caches remain owned by their original workbook parts.
        """
        source = Path(source_workbook).expanduser().resolve()
        output = Path(output_workbook).expanduser().resolve()

        if not source.is_file():
            raise PaymentWorkbookError(f"Workbook was not found: {source}")
        if output.suffix.lower() not in {".xlsx", ".xlsm"}:
            raise PaymentWorkbookError(
                "Payment Breakdown output must use .xlsx or .xlsm."
            )
        if source == output:
            raise PaymentWorkbookError(
                "Payment Breakdown output must be a new workbook path."
            )

        try:
            dataset = RebuildWorkbookReader().read_main_dataset(source)
        except Exception as exc:
            raise PaymentWorkbookError(str(exc)) from exc

        snapshot = self.breakdown_adapter.derive(dataset)

        output.parent.mkdir(parents=True, exist_ok=True)
        fd, temp_name = tempfile.mkstemp(
            prefix=f".{output.stem}.pb4.",
            suffix=output.suffix,
            dir=output.parent,
        )
        os.close(fd)
        temp_path = Path(temp_name)
        keep_vba = source.suffix.lower() == ".xlsm"

        try:
            shutil.copy2(source, temp_path)
            wb = load_workbook(
                temp_path,
                read_only=False,
                data_only=False,
                keep_vba=keep_vba,
                keep_links=True,
            )
            try:
                if "main" not in wb.sheetnames:
                    raise PaymentWorkbookError(
                        "Worksheet 'main' was not found."
                    )

                render_payment_breakdown(wb, snapshot)

                # Same presentation guard used by Earned Value.  PB owns no
                # existing drawings and therefore must not rebuild them.
                reassert_traditional_overlay_transparency(wb)
                wb.save(temp_path)
            finally:
                wb.close()

            # openpyxl does not model every drawing/shape and cached external-link
            # record used by an existing Progress Studio workbook.  PB owns none
            # of those package parts, so restore them byte-for-byte from source.
            restore_opaque_workbook_parts(source, temp_path)

            validate_xlsx_tables(temp_path)
            os.replace(temp_path, output)
            return snapshot
        except Exception:
            temp_path.unlink(missing_ok=True)
            raise

    def rebuild_embedded_workbook(
        self,
        source_workbook: Path,
        output_workbook: Path,
        periods: int | None = None,
    ) -> PaymentMultiLineRenderResult | None:
        """Rebuild the one-workbook Payment workflow.
        Persistent: main + Payment Input.
        Generated/replaced: progress_table + Payment.
        """
        source = Path(source_workbook)
        output = Path(output_workbook)
        if not source.is_file():
            raise PaymentWorkbookError(f"Workbook was not found: {source}")
        if output.resolve() != source.resolve():
            output.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(source, output)

        preserved = None
        try:
            preserved = self.payment_reader.read(source)
        except PaymentWorkbookError:
            preserved = None

        wb = load_workbook(output)
        try:
            if "main" not in wb.sheetnames:
                raise PaymentWorkbookError("Worksheet 'main' was not found.")

            # Payment rebuild intentionally refreshes only the two generated snapshots:
            # progress_table here, Payment after the workbook is saved.
            build_progress_table_from_source(wb, wb["main"])
            if "progress_table" in wb.sheetnames:
                wb["progress_table"].sheet_state = "hidden"

            self.payment_input.embed(
                wb,
                preserved=preserved,
                periods=periods,
            )

            # Payment is generated; remove stale content before saving the source
            # used by the line renderer.
            if "Payment" in wb.sheetnames:
                del wb["Payment"]

            # Save the intermediate workbook without finalizing it. The Payment
            # renderer owns the one final policy pass once Payment has been added.
            wb.save(output)
        finally:
            wb.close()

        try:
            return self.render_payment_backbones(output, output, output)
        except PaymentWorkbookError as exc:
            if "no resolved requirements" not in str(exc).lower():
                raise

            # No Payment sheet will be generated, so finalize the intermediate
            # workbook here exactly once before returning it to the user.
            final_wb = load_workbook(output)
            try:
                finalize_workbook(
                    final_wb,
                    mode="snapshot",
                    include_guide=True,
                )
                final_wb.save(output)
            finally:
                final_wb.close()
            return None

    def validate_workbook(
        self,
        workbook: Path,
    ) -> PaymentWorkbookValidation:
        return self.snapshotter.validate(Path(workbook))

    def create_payment_snapshot(
        self,
        source: Path,
        output: Path,
    ) -> PaymentSnapshotResult:
        return self.snapshotter.create_snapshot(
            Path(source),
            Path(output),
        )

    def create_fake_payment_input(
        self,
        source: Path,
        output: Path,
        periods: int,
    ) -> PaymentInputResult:
        return self.payment_input.create(
            Path(source),
            Path(output),
            periods,
        )

    def validate_payment_input(
        self,
        payment_workbook: Path,
        progress_workbook: Path | None = None,
    ) -> PaymentInputValidation:
        progress = (
            Path(progress_workbook)
            if progress_workbook is not None
            else None
        )
        return self.payment_input.validate(
            Path(payment_workbook),
            progress,
        )

    def read_payment_requirements(
        self,
        payment_workbook: Path,
    ) -> PaymentInputData:
        return self.payment_reader.read(Path(payment_workbook))

    def prepare_payment_input(
        self,
        progress_workbook: Path,
        payment_workbook: Path,
    ) -> PaymentPreparationResult:
        """Validate and resolve the uploaded Payment Input with one read per workbook."""
        payment = self.payment_reader.read(Path(payment_workbook))
        progress = self.progress_index_reader.read(
            Path(progress_workbook)
        )
        progress_ids = set(progress.activities)
        matched = sum(
            1
            for activity_id in payment.activity_ids
            if activity_id in progress_ids
        )
        validation = PaymentInputValidation(
            workbook=payment.workbook,
            payment_sheet=payment.sheet,
            payment_periods=len(payment.periods),
            activity_rows=len(payment.activity_ids),
            matched_activities=matched,
            missing_activities=len(payment.activity_ids) - matched,
            populated_requirements=payment.populated_requirements,
        )
        positions = self.position_engine.resolve(payment, progress)
        return PaymentPreparationResult(
            validation=validation,
            positions=positions,
        )

    def prepare_payment_positions(
        self,
        progress_workbook: Path,
        payment_workbook: Path,
    ) -> PaymentPositionResult:
        payment = self.payment_reader.read(Path(payment_workbook))
        progress = self.progress_index_reader.read(
            Path(progress_workbook)
        )
        return self.position_engine.resolve(payment, progress)

    def render_payment_backbones(
        self,
        progress_workbook: Path,
        payment_workbook: Path,
        output_workbook: Path,
        period_ids: tuple[str, ...] | None = None,
    ) -> PaymentMultiLineRenderResult:
        """Render all populated periods by default, or an explicit subset when requested."""
        prepared = self.prepare_payment_input(
            Path(progress_workbook),
            Path(payment_workbook),
        )
        by_id = {
            period.period_id: period
            for period in prepared.positions.periods
        }

        if period_ids is None:
            selected = [
                period
                for period in prepared.positions.periods
                if period.points
            ]
        else:
            selected = []
            for period_id in period_ids:
                period = by_id.get(period_id)
                if period is None:
                    raise PaymentWorkbookError(
                        f"Payment period {period_id} was not found in Payment Input."
                    )
                if period.points:
                    selected.append(period)

        if not selected:
            raise PaymentWorkbookError(
                "Payment Input has no resolved requirements to render."
            )

        return self.line_renderer.render_periods(
            Path(progress_workbook),
            Path(output_workbook),
            tuple(selected),
        )

    def render_single_payment_line(
        self,
        progress_workbook: Path,
        payment_workbook: Path,
        output_workbook: Path,
        period_id: str = "P01",
    ) -> PaymentLineRenderResult:
        """Resolve and render one payment period at its planned eligible boundary."""
        prepared = self.prepare_payment_input(
            Path(progress_workbook),
            Path(payment_workbook),
        )
        period = next(
            (
                item
                for item in prepared.positions.periods
                if item.period_id == period_id
            ),
            None,
        )
        if period is None:
            raise PaymentWorkbookError(
                f"Payment period {period_id} was not found in Payment Input."
            )
        if not period.points:
            raise PaymentWorkbookError(
                f"{period_id} has no resolved requirements to render."
            )
        return self.line_renderer.render_single_period(
            Path(progress_workbook),
            Path(output_workbook),
            period,
        )
