from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from collections.abc import Callable
import shutil
import tempfile

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from progress_studio.config import WORKBOOK_SCHEMA
from progress_studio.domain.schedule_source import ScheduleSource
from progress_studio.infrastructure.excel.import_workbook_writer import ImportWorkbookWriter
from progress_studio.services.amount_service import AmountService
from progress_studio.services.distribution_service import DistributionService
from progress_studio.services.okd_service import OkdService
from progress_studio.services.progress_service import ProgressService
from progress_studio.services.schedule_workbook_service import ScheduleWorkbookService
from progress_studio.services.timescale_service import TimescaleService


@dataclass(frozen=True, slots=True)
class WorkbookGenerationResult:
    output_file: Path
    wbs_count: int
    activity_count: int
    generated_distribution_count: int
    missing_date_count: int


class WorkbookGenerationService:
    """Generate a fresh Progress workbook from any normalized schedule source."""

    def __init__(self) -> None:
        self.writer = ImportWorkbookWriter()
        self.schedule = ScheduleWorkbookService()
        self.timescale = TimescaleService()
        self.amount = AmountService()
        self.progress = ProgressService()
        self.distribution = DistributionService()
        self.okd = OkdService()

    @staticmethod
    def _write_amount_mapping(workbook_file: Path, amounts: dict[str, float]) -> None:
        wb = load_workbook(workbook_file)
        try:
            if WORKBOOK_SCHEMA.mapping_sheet in wb.sheetnames:
                del wb[WORKBOOK_SCHEMA.mapping_sheet]
            ws = wb.create_sheet(WORKBOOK_SCHEMA.mapping_sheet)
            ws.append(["Activity ID", "WBS", "Description", "Amount", "Status"])
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="4472C4")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            main = wb[WORKBOOK_SCHEMA.main_sheet]
            from progress_studio.infrastructure.excel.amount_workbook import collect_schedule_rows
            for item in collect_schedule_rows(main):
                if str(item["row_type"]).lower() in {"project summary", "wbs"}:
                    ws.append(["", item["wbs"], item["description"], None, "PARENT"])
                    continue
                activity_id = str(item["activity_id"] or "").strip().upper()
                value = float(amounts.get(activity_id, 0.0))
                ws.append([activity_id, item["wbs"], item["description"], value, "Mapped"])
                ws.cell(ws.max_row, 4).number_format = '#,##0.00'
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = f"A1:E{ws.max_row}"
            wb.save(workbook_file)
        finally:
            wb.close()

    def generate(
        self,
        source: ScheduleSource,
        output_file: Path,
        *,
        cutoff_day: str = "Friday",
        distribution_method: str = "auto",
        amounts: dict[str, float] | None = None,
        progress_callback: Callable[[str, str, bool], None] | None = None,
    ) -> WorkbookGenerationResult:
        output_file = Path(output_file)

        def report(step: str, message: str, complete: bool = False) -> None:
            if progress_callback is not None:
                progress_callback(step, message, complete)

        report("read", "Reading working schedule...")
        rows = source.activities()
        report("read", f"Read {len(rows):,} schedule rows.", True)
        amounts = {key.strip().upper(): float(value) for key, value in (amounts or {}).items()}
        with tempfile.TemporaryDirectory(prefix="progress-studio-generate-") as temp_dir_text:
            temp_dir = Path(temp_dir_text)
            imported = temp_dir / "01_imported.xlsx"
            scheduled = temp_dir / "02_schedule.xlsx"
            timescaled = temp_dir / "03_timescale.xlsx"
            amount_mapped = temp_dir / "04_amount.xlsx"
            progress = temp_dir / "05_progress.xlsx"
            distributed = temp_dir / "06_distributed.xlsx"

            report("main", "Building main schedule...")
            self.writer.write(imported, Path("working-tree"), source.project_name, rows)
            self.schedule.prepare(imported, scheduled)
            report("main", "Main schedule built.", True)

            report("timescale", "Building timescale...")
            self.timescale.build(scheduled, timescaled, cutoff_day)
            report("timescale", "Timescale built.", True)

            report("mapping", "Applying mapped amounts...")
            self._write_amount_mapping(timescaled, amounts)
            self.amount.apply_mapping(timescaled, amount_mapped)
            report("mapping", "Mapped amounts applied.", True)

            report("progress", "Building progress sheets, Activity Data theme, and Dashboard...")
            self.progress.build(amount_mapped, progress)
            report("progress", "Progress sheets, Activity Data theme, and Dashboard built.", True)

            report("distribution", "Generating plan distribution...")
            distribution = self.distribution.generate(
                progress,
                distributed,
                method=distribution_method,
                debug=False,
            )
            report("distribution", "Plan distribution generated.", True)

            report("okd", "Building OKD sheets...")
            self.okd.build(distributed, distributed)
            report("okd", "OKD sheets built.", True)

            report("finalize", "Writing final workbook...")
            output_file.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(distributed, output_file)

        return WorkbookGenerationResult(
            output_file=output_file,
            wbs_count=sum(row.is_summary for row in rows),
            activity_count=sum(not row.is_summary for row in rows),
            generated_distribution_count=distribution.generated,
            missing_date_count=distribution.skipped_no_dates,
        )
