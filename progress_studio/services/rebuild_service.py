from __future__ import annotations

from pathlib import Path
import os
import shutil
import tempfile

from openpyxl import load_workbook

from progress_studio.infrastructure.excel.rebuild_workbook_reader import (
    RebuildWorkbookReadError,
    RebuildWorkbookReader,
)

from progress_studio.infrastructure.excel.okd_workbook import (
    build_progress_views_from_source,
)
from progress_studio.infrastructure.excel.monthly_main_workbook import (
    build_monthly_main_view,
)
from progress_studio.infrastructure.excel.dashboard_workbook import build_dashboard
from progress_studio.infrastructure.excel.live_dashboard_workbook import build_live_dashboard
from progress_studio.infrastructure.excel.live_monthly_workbook import build_live_monthly_view
from progress_studio.infrastructure.excel.calculation_policy import (
    configure_incremental_excel_recalculation,
)
from progress_studio.infrastructure.excel.xlsx_package_validator import validate_xlsx_tables
from progress_studio.infrastructure.excel.workbook_visibility import apply_final_sheet_visibility
from progress_studio.infrastructure.excel.workbook_protection import apply_final_sheet_protection
from progress_studio.services.payment_service import PaymentService
from progress_studio.services.monthly_cache_deriver import MonthlyCacheDeriver

from progress_studio.domain.rebuild_models import (
    RebuildMode,
    RebuildSheetContract,
    RebuildWorkbookAnalysis,
    ProgressRebuildResult,
    PaymentRebuildResult,
    LiveProgressRebuildResult,
)


class RebuildContractError(ValueError):
    """Raised when an input workbook cannot satisfy the standalone rebuild contract."""


DEFAULT_REBUILD_CONTRACT = RebuildSheetContract(
    source_of_truth=("main", "Payment Input"),
    preserve=("main", "Payment Input"),
    generated_progress=(
        "main_monthly",
        "progress",
        "progress_table",
        "Dashboard_Data",
        "Dashboard",
    ),
    generated_payment=("Payment",),
    internal_preserve=(
        "Info",
        "Timescale Info",
        "Amount Mapping",
        "Distribution Report",
        "ProgressStudio Extensions",
        "BOQ Activity Mapping",
        "Mapping Summary",
        "Rebuild Audit",
    ),
)


class WorkbookRebuildEngine:
    """Standalone rebuild boundary.

    This engine deliberately knows nothing about XML, BOQ files, .progressstudio
    sessions, .boqstudio sessions, mapping allocations, or WorkingScheduleTree.

    `main` in the supplied workbook is authoritative. Payment mode additionally
    requires the embedded `Payment Input` sheet.

    MS-RB1 only analyzes and plans ownership. It does not regenerate sheets yet.
    """

    MAIN_SHEET = "main"
    PAYMENT_INPUT_SHEET = "Payment Input"

    def __init__(
        self,
        contract: RebuildSheetContract | None = None,
        reader: RebuildWorkbookReader | None = None,
        payment_service: PaymentService | None = None,
    ) -> None:
        self.contract = contract or DEFAULT_REBUILD_CONTRACT
        self.reader = reader or RebuildWorkbookReader()
        self.payment_service = payment_service or PaymentService()

    def analyze(
        self,
        workbook_path: Path,
        mode: RebuildMode | str,
    ) -> RebuildWorkbookAnalysis:
        path = Path(workbook_path).expanduser().resolve()
        rebuild_mode = self._mode(mode)
        self._validate_path(path)

        try:
            probe = self.reader.probe(path)
        except RebuildWorkbookReadError as exc:
            raise RebuildContractError(str(exc)) from exc

        if probe.activity_count <= 0:
            raise RebuildContractError(
                "Worksheet 'main' contains no valid Plan Activity rows."
            )

        payment_input_present = self.PAYMENT_INPUT_SHEET in probe.sheet_names
        if rebuild_mode is RebuildMode.PAYMENT and not payment_input_present:
            raise RebuildContractError(
                "Rebuild Payment requires embedded worksheet 'Payment Input'."
            )

        generated = self.contract.generated_for(rebuild_mode)
        existing_generated = tuple(name for name in generated if name in probe.sheet_names)
        missing_generated = tuple(name for name in generated if name not in probe.sheet_names)

        preserve_names = set(self.contract.preserve) | set(self.contract.internal_preserve)
        preserve_present = tuple(
            name for name in probe.sheet_names if name in preserve_names
        )

        known = (
            set(self.contract.preserve)
            | set(self.contract.internal_preserve)
            | set(self.contract.generated_progress)
            | set(self.contract.generated_payment)
        )
        unknown = tuple(name for name in probe.sheet_names if name not in known)

        return RebuildWorkbookAnalysis(
            workbook=path,
            mode=rebuild_mode,
            main_sheet=self.MAIN_SHEET,
            main_rows=probe.main_rows,
            main_columns=probe.main_columns,
            activity_count=probe.activity_count,
            payment_input_present=payment_input_present,
            existing_generated_sheets=existing_generated,
            missing_generated_sheets=missing_generated,
            preserve_sheets_present=preserve_present,
            unknown_sheets=unknown,
            contract=self.contract,
        )


    def rebuild_progress(
        self,
        source_workbook: Path,
        output_workbook: Path,
        *,
        project_name: str | None = None,
    ) -> ProgressRebuildResult:
        """Delete and rebuild only Progress-generated sheets from current ``main``.

        MS-RB2 execution contract:
        - ``main`` is authoritative and preserved exactly as loaded.
        - ``Payment Input`` and ``Payment`` are not touched.
        - Internal metadata and unknown user sheets are preserved.
        - Only Progress-owned generated sheets are replaced.
        - Output is written atomically so a failed rebuild cannot corrupt input.
        """
        source = Path(source_workbook).expanduser().resolve()
        output = Path(output_workbook).expanduser().resolve()

        analysis = self.analyze(source, RebuildMode.PROGRESS)
        if output.suffix.lower() not in {".xlsx", ".xlsm"}:
            raise RebuildContractError(
                "Rebuild output must use .xlsx or .xlsm."
            )

        output.parent.mkdir(parents=True, exist_ok=True)
        fd, temp_name = tempfile.mkstemp(
            prefix=f".{output.stem}.rb2.",
            suffix=output.suffix,
            dir=output.parent,
        )
        os.close(fd)
        temp_path = Path(temp_name)

        keep_vba = source.suffix.lower() == ".xlsm"
        try:
            shutil.copy2(source, temp_path)
            wb = load_workbook(temp_path, read_only=False, data_only=False, keep_vba=keep_vba)
            values_wb = load_workbook(source, read_only=False, data_only=True, keep_vba=keep_vba)
            try:
                if self.MAIN_SHEET not in wb.sheetnames:
                    raise RebuildContractError(
                        "Worksheet 'main' disappeared during rebuild preparation."
                    )

                preserved_payment_sheet = "Payment" in wb.sheetnames
                preserved_payment_input_sheet = "Payment Input" in wb.sheetnames

                # Hard contract: delete only Progress-generated sheets. Builders are
                # deterministic and may also remove their own target, but deleting
                # here makes ownership explicit and prevents stale sheet remnants.
                for sheet_name in self.contract.generated_progress:
                    if sheet_name in wb.sheetnames:
                        del wb[sheet_name]

                main = wb[self.MAIN_SHEET]
                values_main = values_wb[self.MAIN_SHEET]
                (
                    activity_count,
                    week_count,
                    progress_table_rows,
                    checked_cells,
                ) = build_progress_views_from_source(
                    wb,
                    main,
                    snapshot_progress=True,
                    value_source=values_main,
                )

                if "progress_table" in wb.sheetnames:
                    wb["progress_table"].sheet_state = "hidden"

                monthly_periods = build_monthly_main_view(
                    wb,
                    source_sheet=self.MAIN_SHEET,
                    target_sheet="main_monthly",
                    require_timescale=True,
                    snapshot=True,
                    value_source=values_main,
                )

                build_dashboard(
                    wb,
                    project_name=project_name or output.stem,
                )

                # Re-assert visibility/support contract after builders.
                if "progress_table" in wb.sheetnames:
                    wb["progress_table"].sheet_state = "hidden"
                if "Dashboard_Data" in wb.sheetnames:
                    wb["Dashboard_Data"].sheet_state = "hidden"

                configure_incremental_excel_recalculation(wb)
                apply_final_sheet_visibility(wb)
                apply_final_sheet_protection(wb)
                wb.save(temp_path)
            finally:
                wb.close()
                values_wb.close()

            validate_xlsx_tables(temp_path)
            os.replace(temp_path, output)

            return ProgressRebuildResult(
                source_workbook=source,
                output_workbook=output,
                activity_count=activity_count,
                week_count=week_count,
                progress_table_rows=progress_table_rows,
                progress_table_checked_cells=checked_cells,
                monthly_periods=monthly_periods,
                rebuilt_sheets=self.contract.generated_progress,
                preserved_payment_sheet=preserved_payment_sheet,
                preserved_payment_input_sheet=preserved_payment_input_sheet,
            )
        except Exception:
            temp_path.unlink(missing_ok=True)
            raise


    def rebuild_live_progress(
        self,
        source_workbook: Path,
        output_workbook: Path,
        *,
        project_name: str | None = None,
    ) -> LiveProgressRebuildResult:
        """LW-7 one-pass Live Progress writer.

        Read path: sparse OOXML -> MainDataset.
        Write path: one mutable openpyxl workbook -> one save.
        No second data_only workbook is opened.
        """
        source = Path(source_workbook).expanduser().resolve()
        output = Path(output_workbook).expanduser().resolve()
        analysis = self.analyze(source, RebuildMode.PROGRESS)
        if output.suffix.lower() not in {".xlsx", ".xlsm"}:
            raise RebuildContractError("Rebuild output must use .xlsx or .xlsm.")

        dataset = self.reader.read_main_dataset(source)
        monthly_cache = MonthlyCacheDeriver().derive(dataset)

        output.parent.mkdir(parents=True, exist_ok=True)
        fd, temp_name = tempfile.mkstemp(
            prefix=f".{output.stem}.lw7.",
            suffix=output.suffix,
            dir=output.parent,
        )
        os.close(fd)
        temp_path = Path(temp_name)
        keep_vba = source.suffix.lower() == ".xlsm"

        try:
            shutil.copy2(source, temp_path)
            wb = load_workbook(
                temp_path,
                read_only=False,
                data_only=False,
                keep_vba=keep_vba,
            )
            try:
                preserved_payment_sheet = "Payment" in wb.sheetnames
                preserved_payment_input_sheet = "Payment Input" in wb.sheetnames

                # Live owns only these generated Progress views.
                for sheet_name in (
                    "main_monthly",
                    "progress",
                    "progress_table",
                    "Dashboard_Data",
                    "Dashboard",
                ):
                    if sheet_name in wb.sheetnames:
                        del wb[sheet_name]

                monthly_periods = build_live_monthly_view(
                    wb,
                    dataset,
                    monthly_cache,
                    source_sheet=self.MAIN_SHEET,
                    target_sheet="main_monthly",
                )
                build_live_dashboard(
                    wb,
                    dataset,
                    project_name=project_name or output.stem,
                )

                if "Dashboard_Data" in wb.sheetnames:
                    wb["Dashboard_Data"].sheet_state = "hidden"

                apply_final_sheet_visibility(wb)
                apply_final_sheet_protection(wb)
                wb.save(temp_path)
            finally:
                wb.close()

            validate_xlsx_tables(temp_path)
            os.replace(temp_path, output)
            return LiveProgressRebuildResult(
                source_workbook=source,
                output_workbook=output,
                activity_count=len(dataset.activities),
                week_count=len(dataset.periods),
                monthly_periods=monthly_periods,
                dashboard_rows=(len(dataset.rows) * 2),
                rebuilt_sheets=("main_monthly", "Dashboard_Data", "Dashboard"),
                preserved_payment_sheet=preserved_payment_sheet,
                preserved_payment_input_sheet=preserved_payment_input_sheet,
            )
        except Exception:
            temp_path.unlink(missing_ok=True)
            raise

    def rebuild_payment(
        self,
        source_workbook: Path,
        output_workbook: Path,
    ) -> PaymentRebuildResult:
        """Replace only ``Payment`` from current ``main`` + ``Payment Input``.

        MS-RB4 deliberately does not rebuild or reconcile any Progress-generated
        sheet. This makes Payment-only edits cheap and keeps the Progress workbook
        exactly as the user last rebuilt it.
        """
        source = Path(source_workbook).expanduser().resolve()
        output = Path(output_workbook).expanduser().resolve()

        analysis = self.analyze(source, RebuildMode.PAYMENT)
        source_probe = self.reader.probe(source)
        if output.suffix.lower() not in {".xlsx", ".xlsm"}:
            raise RebuildContractError("Rebuild output must use .xlsx or .xlsm.")

        output.parent.mkdir(parents=True, exist_ok=True)
        fd, temp_name = tempfile.mkstemp(
            prefix=f".{output.stem}.rb4.",
            suffix=output.suffix,
            dir=output.parent,
        )
        os.close(fd)
        temp_path = Path(temp_name)

        try:
            shutil.copy2(source, temp_path)

            # The existing Payment renderer already follows the correct ownership
            # contract: it reads main + Payment Input, removes stale Payment, and
            # writes a new Payment sheet while leaving every other sheet untouched.
            rendered = self.payment_service.render_payment_backbones(
                temp_path,
                temp_path,
                temp_path,
            )

            visibility_wb = load_workbook(temp_path, read_only=False, data_only=False)
            try:
                apply_final_sheet_visibility(visibility_wb)
                apply_final_sheet_protection(visibility_wb)
                visibility_wb.save(temp_path)
            finally:
                visibility_wb.close()

            validate_xlsx_tables(temp_path)
            os.replace(temp_path, output)

            return PaymentRebuildResult(
                source_workbook=source,
                output_workbook=output,
                rendered_periods=rendered.rendered_periods,
                rendered_points=rendered.rendered_points,
                period_ids=rendered.period_ids,
                rebuilt_sheets=self.contract.generated_payment,
                progress_generated_preserved=tuple(
                    name for name in self.contract.generated_progress
                    if name in source_probe.sheet_names
                ),
            )
        except Exception:
            temp_path.unlink(missing_ok=True)
            raise

    def generated_sheets_for(self, mode: RebuildMode | str) -> tuple[str, ...]:
        """Return the only sheets a future rebuild execution may replace."""
        return self.contract.generated_for(self._mode(mode))

    @staticmethod
    def _mode(mode: RebuildMode | str) -> RebuildMode:
        if isinstance(mode, RebuildMode):
            return mode
        try:
            return RebuildMode(str(mode).strip().lower())
        except ValueError as exc:
            raise RebuildContractError(
                "Rebuild mode must be 'progress' or 'payment'."
            ) from exc

    @staticmethod
    def _validate_path(path: Path) -> None:
        if not path.is_file():
            raise RebuildContractError(f"Workbook was not found: {path}")
        if path.suffix.lower() not in {".xlsx", ".xlsm"}:
            raise RebuildContractError(
                "Standalone rebuild accepts an Excel .xlsx or .xlsm workbook."
            )

