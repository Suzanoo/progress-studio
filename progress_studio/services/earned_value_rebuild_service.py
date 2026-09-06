from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
import os
import shutil
import tempfile
from typing import Callable
import zipfile
from xml.etree import ElementTree as ET

from openpyxl import load_workbook

from progress_studio.domain.earned_value import EarnedValueResult
from progress_studio.infrastructure.excel.earned_value_input_reader import (
    EarnedValueInputWorkbookReader,
    EarnedValueWorkbookInputError,
)
from progress_studio.infrastructure.excel.earned_value_workbook import (
    EARNED_VALUE_SHEET,
    render_earned_value_sheet,
)
from progress_studio.infrastructure.excel.rebuild_workbook_reader import (
    RebuildWorkbookReader,
    RebuildWorkbookReadError,
)
from progress_studio.infrastructure.excel.traditional_overlay_workbook import (
    reassert_traditional_overlay_transparency,
)
from progress_studio.services.earned_value_deriver import (
    EarnedValueDeriver,
    EarnedValueInputError,
)


_MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"


class EarnedValueRebuildError(ValueError):
    """Raised when a workbook cannot satisfy the EV workspace contract."""


@dataclass(frozen=True, slots=True)
class EarnedValueRebuildAnalysis:
    workbook: Path
    cutoff_date: datetime
    activity_count: int
    boq_count: int
    allocation_count: int
    project_bac: float
    existing_earned_value_sheet: bool


@dataclass(frozen=True, slots=True)
class EarnedValueRebuildResult:
    source_workbook: Path
    output_workbook: Path
    cutoff_date: datetime
    activity_count: int
    boq_count: int
    allocation_count: int
    project_bac: float
    refreshed_sheet: str


def _as_datetime(value: object) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    return None


class EarnedValueRebuildService:
    """Standalone EV extension used by the Rebuild workspace.

    EV reads current ``main`` progress values and embedded BOQ/mapping provenance.
    A current reporting control may seed the initial EV view date, but it is not
    required as a post-rebuild calculation authority. The service mutates only
    EV-owned workbook presentation:
    ``Earned Value`` and ``EV_Data``.

    The source workbook is copied to a temporary output, opened once as a normal
    workbook, EV-owned sheets are added/refreshed, the proven overlay transparency
    guard is re-applied, and the workbook is saved once. Existing Progress,
    Payment, Dashboard and user sheets are not rebuilt by this service.
    """

    def __init__(
        self,
        *,
        input_reader: EarnedValueInputWorkbookReader | None = None,
        main_reader: RebuildWorkbookReader | None = None,
        deriver: EarnedValueDeriver | None = None,
        renderer: Callable[[object, EarnedValueResult], None] = render_earned_value_sheet,
    ) -> None:
        self.input_reader = input_reader or EarnedValueInputWorkbookReader()
        self.main_reader = main_reader or RebuildWorkbookReader()
        self.deriver = deriver or EarnedValueDeriver()
        self.renderer = renderer

    def analyze(self, workbook_path: Path) -> EarnedValueRebuildAnalysis:
        path = self._validate_path(workbook_path)
        result, cutoff, activity_count, boq_count, allocation_count = self._derive(path)
        return EarnedValueRebuildAnalysis(
            workbook=path,
            cutoff_date=cutoff,
            activity_count=activity_count,
            boq_count=boq_count,
            allocation_count=allocation_count,
            project_bac=float(result.project_bac),
            existing_earned_value_sheet=self._has_sheet(path, EARNED_VALUE_SHEET),
        )

    def generate(
        self,
        source_workbook: Path,
        output_workbook: Path,
    ) -> EarnedValueRebuildResult:
        source = self._validate_path(source_workbook)
        output = Path(output_workbook).expanduser().resolve()
        if output.suffix.lower() not in {".xlsx", ".xlsm"}:
            raise EarnedValueRebuildError(
                "Earned Value output must use .xlsx or .xlsm."
            )
        if source == output:
            raise EarnedValueRebuildError(
                "Earned Value output must be a new workbook path."
            )

        result, cutoff, activity_count, boq_count, allocation_count = self._derive(source)

        output.parent.mkdir(parents=True, exist_ok=True)
        fd, temp_name = tempfile.mkstemp(
            prefix=f".{output.stem}.ev5.",
            suffix=output.suffix,
            dir=output.parent,
        )
        os.close(fd)
        temp_path = Path(temp_name)
        keep_vba = source.suffix.lower() == ".xlsm"

        try:
            shutil.copy2(source, temp_path)
            workbook = load_workbook(
                temp_path,
                read_only=False,
                data_only=False,
                keep_vba=keep_vba,
            )
            try:
                # EV owns only its two sheets. Existing Progress/Payment/Dashboard
                # builders are deliberately not called here.
                self.renderer(workbook, result)

                # Progress Studio already carries this guard for normal workbook
                # round-trips. Re-assert presentation only; do not rebuild charts,
                # series, cutoff logic or any existing worksheet data.
                reassert_traditional_overlay_transparency(workbook)

                workbook.save(temp_path)
            finally:
                workbook.close()

            os.replace(temp_path, output)
        except Exception:
            temp_path.unlink(missing_ok=True)
            raise

        return EarnedValueRebuildResult(
            source_workbook=source,
            output_workbook=output,
            cutoff_date=cutoff,
            activity_count=activity_count,
            boq_count=boq_count,
            allocation_count=allocation_count,
            project_bac=float(result.project_bac),
            refreshed_sheet=EARNED_VALUE_SHEET,
        )

    def _derive(
        self,
        path: Path,
    ) -> tuple[EarnedValueResult, datetime, int, int, int]:
        try:
            embedded = self.input_reader.read(path)
            dataset = self.main_reader.read_main_dataset(path)
            cutoff = self._initial_view_date(path, dataset)
            result = self.deriver.derive(
                dataset,
                embedded.boq_rows,
                embedded.allocations,
                cutoff_date=cutoff,
            )
        except (
            EarnedValueWorkbookInputError,
            RebuildWorkbookReadError,
            EarnedValueInputError,
            KeyError,
            ValueError,
        ) as exc:
            if isinstance(exc, EarnedValueRebuildError):
                raise
            raise EarnedValueRebuildError(str(exc)) from exc

        activity_count = len(tuple(dataset.activities))
        if activity_count <= 0:
            raise EarnedValueRebuildError(
                "Worksheet 'main' contains no valid Plan Activity rows."
            )
        if not result.project_points:
            raise EarnedValueRebuildError(
                "Earned Value has no reporting periods to render."
            )
        return (
            result,
            cutoff,
            activity_count,
            len(embedded.boq_rows),
            len(embedded.allocations),
        )

    @classmethod
    def _initial_view_date(cls, path: Path, dataset) -> datetime:
        """Resolve the initial EV view without making Dashboard cutoff authoritative.

        Prefer the latest month containing real Actual progress in ``main`` and
        select that month's final reporting period. Existing Dashboard/main cutoff
        controls are only a fallback seed. If neither exists, use the latest
        project reporting period.
        """
        periods = tuple(getattr(dataset, "periods", ()) or ())
        rows = tuple(getattr(dataset, "rows", ()) or ())
        actual_rows = tuple(
            row for row in rows
            if str(getattr(row, "pa", "") or "").strip().upper() == "A"
            and str(getattr(row, "activity_id", "") or "").strip()
        )

        latest_actual: datetime | None = None
        for period in periods:
            reporting_date = _as_datetime(getattr(period, "reporting_date", None))
            if reporting_date is None:
                continue
            has_actual = False
            for row in actual_rows:
                try:
                    value = row.period_value(period.column)
                except (AttributeError, TypeError):
                    value = None
                if value not in (None, ""):
                    try:
                        has_actual = abs(float(value)) > 1e-12
                    except (TypeError, ValueError):
                        has_actual = False
                if has_actual:
                    break
            if has_actual and (latest_actual is None or reporting_date > latest_actual):
                latest_actual = reporting_date

        if latest_actual is not None:
            same_month = [
                _as_datetime(getattr(period, "reporting_date", None))
                for period in periods
                if _as_datetime(getattr(period, "reporting_date", None)) is not None
                and _as_datetime(getattr(period, "reporting_date", None)).year == latest_actual.year
                and _as_datetime(getattr(period, "reporting_date", None)).month == latest_actual.month
            ]
            if same_month:
                return max(same_month)
            return latest_actual

        saved_view_seed = cls._read_cutoff(path)
        if saved_view_seed is not None:
            return saved_view_seed

        reporting_dates = [
            value
            for period in periods
            if (value := _as_datetime(getattr(period, "reporting_date", None))) is not None
        ]
        if reporting_dates:
            return max(reporting_dates)

        raise EarnedValueRebuildError(
            "Earned Value requires at least one valid reporting date in worksheet 'main'."
        )

    @staticmethod
    def _read_cutoff(path: Path) -> datetime | None:
        keep_vba = path.suffix.lower() == ".xlsm"
        workbook = load_workbook(
            path,
            read_only=True,
            data_only=True,
            keep_vba=keep_vba,
        )
        try:
            # Earned Value!M3 is never used as a rebuild input. Existing Progress
            # controls are only fallback seeds for the initial EV view date.
            if "Dashboard" in workbook.sheetnames:
                cutoff = _as_datetime(workbook["Dashboard"]["K5"].value)
                if cutoff is not None:
                    return cutoff

            if "main" in workbook.sheetnames:
                ws = workbook["main"]
                max_row = min(ws.max_row, 80)
                max_col = min(ws.max_column, 30)
                for row in range(1, max_row + 1):
                    for col in range(1, max_col + 1):
                        value = str(ws.cell(row, col).value or "").strip().lower()
                        if value != "cutoff date":
                            continue
                        if col < max_col:
                            cutoff = _as_datetime(ws.cell(row, col + 1).value)
                            if cutoff is not None:
                                return cutoff
            return None
        finally:
            workbook.close()

    @staticmethod
    def _has_sheet(path: Path, sheet_name: str) -> bool:
        """Inspect workbook.xml directly; no extra mutable workbook pass."""
        try:
            with zipfile.ZipFile(path, "r") as archive:
                workbook_xml = archive.read("xl/workbook.xml")
        except (OSError, KeyError, zipfile.BadZipFile):
            return False

        root = ET.fromstring(workbook_xml)
        sheets = root.find(f"{{{_MAIN_NS}}}sheets")
        if sheets is None:
            return False
        return any(
            sheet.attrib.get("name") == sheet_name
            for sheet in sheets.findall(f"{{{_MAIN_NS}}}sheet")
        )

    @staticmethod
    def _validate_path(workbook_path: Path) -> Path:
        path = Path(workbook_path).expanduser().resolve()
        if not path.is_file():
            raise EarnedValueRebuildError(
                f"Earned Value workbook was not found: {path}"
            )
        if path.suffix.lower() not in {".xlsx", ".xlsm"}:
            raise EarnedValueRebuildError(
                "Select a Progress Studio .xlsx or .xlsm workbook."
            )
        return path
