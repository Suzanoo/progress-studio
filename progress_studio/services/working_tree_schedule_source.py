from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from progress_studio.domain.activity import Activity
from progress_studio.domain.working_tree import WorkingNodeKind, WorkingScheduleTree, WorkingTreeNode
from progress_studio.infrastructure.excel.amount_workbook import find_header
from progress_studio.services.schedule_service import ScheduleService


@dataclass(frozen=True, slots=True)
class _ActivityMetadata:
    task_id: int | None = None
    uid: int | None = None
    plan_start: datetime | None = None
    plan_finish: datetime | None = None
    actual_start: datetime | None = None
    actual_finish: datetime | None = None
    percent_complete: float | None = None
    physical_percent_complete: float | None = None
    total_slack_minutes: float | None = None


class WorkingTreeScheduleSource:
    """Adapt the editable working tree to the existing workbook generator.

    Workbook-origin activities retain their schedule metadata. User-created
    activities intentionally use blank dates and are skipped safely by the
    distribution engine until the user completes them in ``main``.
    """

    def __init__(
        self,
        progress_file: Path,
        nodes: list[WorkingTreeNode],
        amounts: dict[str, float],
    ) -> None:
        self.progress_file = Path(progress_file)
        self.tree = WorkingScheduleTree(nodes)
        self.amounts = {key.strip().upper(): float(value) for key, value in amounts.items()}
        self._project_name, self._metadata = self._read_source_metadata()

    @property
    def project_name(self) -> str:
        return self._project_name


    @staticmethod
    def _date(value: object) -> datetime | None:
        if isinstance(value, datetime):
            return value
        if isinstance(value, str):
            text = value.strip()
            for pattern in ("%Y-%m-%d", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S"):
                try:
                    return datetime.strptime(text, pattern)
                except ValueError:
                    continue
        return None

    @staticmethod
    def _number(value: object) -> float | None:
        if value in (None, ""):
            return None
        try:
            return float(value)
        except (TypeError, ValueError):
            return None

    def _read_source_metadata(self) -> tuple[str, dict[str, _ActivityMetadata]]:
        wb = load_workbook(self.progress_file, data_only=False, read_only=False)
        try:
            ws = wb["main"]
            header_row, headers = find_header(
                ws,
                ["Row Type", "Description", "Activity ID", "Plan Start", "Plan Finish"],
            )
            row_type_col = headers["row type"]
            activity_col = headers["activity id"]
            pa_col = headers.get("p/a")
            description_col = headers["description"]
            project_name = self.progress_file.stem
            metadata: dict[str, _ActivityMetadata] = {}
            for row in range(header_row + 1, ws.max_row + 1):
                row_type = str(ws.cell(row, row_type_col).value or "").strip().lower()
                pa = str(ws.cell(row, pa_col).value or "P").strip().upper() if pa_col else "P"
                if row_type == "project summary" and pa == "P":
                    project_name = str(ws.cell(row, description_col).value or project_name).strip()
                if row_type != "activity" or pa != "P":
                    continue
                activity_id = str(ws.cell(row, activity_col).value or "").strip().upper()
                if not activity_id:
                    continue
                total_float_hours = self._number(ws.cell(row, headers.get("total float (hr)", 0)).value) if headers.get("total float (hr)") else None
                metadata[activity_id] = _ActivityMetadata(
                    task_id=ws.cell(row, headers["task id"]).value if headers.get("task id") else None,
                    uid=ws.cell(row, headers["uid"]).value if headers.get("uid") else None,
                    plan_start=self._date(ws.cell(row, headers["plan start"]).value),
                    plan_finish=self._date(ws.cell(row, headers["plan finish"]).value),
                    actual_start=self._date(ws.cell(row, headers["actual start"]).value) if headers.get("actual start") else None,
                    actual_finish=self._date(ws.cell(row, headers["actual finish"]).value) if headers.get("actual finish") else None,
                    percent_complete=(self._number(ws.cell(row, headers["% complete"]).value) or 0.0) * 100.0 if headers.get("% complete") and self._number(ws.cell(row, headers["% complete"]).value) is not None else None,
                    physical_percent_complete=(self._number(ws.cell(row, headers["physical %"]).value) or 0.0) * 100.0 if headers.get("physical %") and self._number(ws.cell(row, headers["physical %"]).value) is not None else None,
                    total_slack_minutes=total_float_hours * 60.0 if total_float_hours is not None else None,
                )
            return project_name or self.progress_file.stem, metadata
        finally:
            wb.close()

    def activities(self) -> list[Activity]:
        rows: list[Activity] = []
        source_order = 0
        for depth, node in self.tree.walk():
            if node.deleted:
                continue
            source_order += 1
            if node.kind is WorkingNodeKind.WBS:
                rows.append(
                    Activity(
                        source_order=source_order,
                        task_id=None,
                        uid=None,
                        activity_id="",
                        name=node.name,
                        wbs=node.code,
                        outline_level=depth,
                        is_summary=True,
                        plan_start=None,
                        plan_finish=None,
                        actual_start=None,
                        actual_finish=None,
                        percent_complete=None,
                        physical_percent_complete=None,
                        total_slack_minutes=None,
                        amount=None,
                    )
                )
                continue

            key = (node.source_activity_id or node.code).strip().upper()
            meta = self._metadata.get(key, _ActivityMetadata())
            rows.append(
                Activity(
                    source_order=source_order,
                    task_id=meta.task_id,
                    uid=meta.uid,
                    activity_id=node.code,
                    name=node.name,
                    wbs=self.tree.get(node.parent_id).code if node.parent_id and self.tree.get(node.parent_id) else "",
                    outline_level=depth,
                    is_summary=False,
                    plan_start=meta.plan_start,
                    plan_finish=meta.plan_finish,
                    actual_start=meta.actual_start,
                    actual_finish=meta.actual_finish,
                    percent_complete=meta.percent_complete,
                    physical_percent_complete=meta.physical_percent_complete,
                    total_slack_minutes=meta.total_slack_minutes,
                    amount=self.amounts.get(node.code.strip().upper(), 0.0),
                )
            )
        ScheduleService().roll_up_summary_dates(rows)
        return rows
