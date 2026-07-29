from __future__ import annotations

from pathlib import Path
import shutil

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from progress_studio.domain.mapping_models import (
    ActivityRow as ActivityRecord,
    AllocationRecord,
    BOQRow as BOQRecord,
)
from progress_studio.infrastructure.excel.mapping_reader import BOQSheetReader, ProgressActivityReader


def _headers(ws, row: int = 1) -> dict[str, int]:
    return {
        str(ws.cell(row, col).value or "").strip().lower(): col
        for col in range(1, ws.max_column + 1)
        if str(ws.cell(row, col).value or "").strip()
    }


class BOQMappingService:
    """Read mapping inputs efficiently and export percentage allocations."""

    def __init__(
        self,
        activity_reader: ProgressActivityReader | None = None,
        boq_reader: BOQSheetReader | None = None,
    ) -> None:
        self.activity_reader = activity_reader or ProgressActivityReader()
        self.boq_reader = boq_reader or BOQSheetReader()

    def read_activities(self, progress_file: Path) -> list[ActivityRecord]:
        return self.activity_reader.read(progress_file)

    def list_boq_sheets(self, boq_file: Path) -> list[str]:
        return self.boq_reader.list_sheets(boq_file)

    def read_boq(self, boq_file: Path, sheet_name: str) -> list[BOQRecord]:
        return self.boq_reader.read(boq_file, sheet_name)

    def export(
        self,
        progress_file: Path,
        output_file: Path,
        boq_rows: list[BOQRecord],
        allocations: list[AllocationRecord],
    ) -> Path:
        output_file.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(progress_file, output_file)
        wb = load_workbook(output_file)
        try:
            ws_amount = wb["Amount Mapping"]
            h = _headers(ws_amount)
            aid_col = h["activity id"]
            amount_col = h["amount"]
            status_col = h.get("status")

            by_key = {row.key: row for row in boq_rows}
            totals: dict[str, float] = {}
            for allocation in allocations:
                row = by_key.get(allocation.boq_key)
                if row:
                    allocated = row.amount * allocation.share_percent / 100.0
                    totals[allocation.activity_id] = totals.get(allocation.activity_id, 0.0) + allocated

            for r in range(2, ws_amount.max_row + 1):
                aid = str(ws_amount.cell(r, aid_col).value or "").strip()
                if not aid:
                    continue
                ws_amount.cell(r, amount_col).value = totals.get(aid, 0.0)
                ws_amount.cell(r, amount_col).number_format = "#,##0.00"
                if status_col:
                    ws_amount.cell(r, status_col).value = "MAPPED" if aid in totals else "UNMAPPED"

            if "BOQ Activity Mapping" in wb.sheetnames:
                del wb["BOQ Activity Mapping"]
            ws = wb.create_sheet("BOQ Activity Mapping")
            headers = [
                "Activity ID", "BOQ Key", "Source Sheet", "Source Row",
                "WBS-2", "WBS-3", "WBS-4", "BOQ Description", "BOQ Amount",
                "Share %", "Allocated Amount",
            ]
            ws.append(headers)
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="4472C4")
                cell.alignment = Alignment(horizontal="center")

            for allocation in allocations:
                row = by_key.get(allocation.boq_key)
                if not row:
                    continue
                allocated = row.amount * allocation.share_percent / 100.0
                ws.append([
                    allocation.activity_id, row.key, row.source_sheet, row.source_row,
                    row.wbs2, row.wbs3, row.wbs4, row.description, row.amount,
                    allocation.share_percent / 100.0, allocated,
                ])
                current = ws.max_row
                ws.cell(current, 9).number_format = "#,##0.00"
                ws.cell(current, 10).number_format = "0.00%"
                ws.cell(current, 11).number_format = "#,##0.00"

            ws.freeze_panes = "A2"
            ws.auto_filter.ref = f"A1:K{max(1, ws.max_row)}"
            widths = {
                "A": 16, "B": 30, "C": 22, "D": 12, "E": 22, "F": 28,
                "G": 28, "H": 60, "I": 18, "J": 12, "K": 18,
            }
            for col, width in widths.items():
                ws.column_dimensions[col].width = width
            wb.calculation.calcMode = "auto"
            wb.calculation.fullCalcOnLoad = True
            wb.calculation.forceFullCalc = True
            wb.save(output_file)
            return output_file
        finally:
            wb.close()
