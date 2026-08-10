from __future__ import annotations

from copy import copy

from openpyxl.styles import Protection
from openpyxl.workbook.protection import WorkbookProtection

from progress_studio.config.workbook_protection import WORKBOOK_SHEET_PASSWORD


EDITABLE_MAIN_PLAN_HEADERS = {
    "wbs",
    "description",
    "activity id",
    "outline level",
    "plan start",
    "plan finish",
    "amount",
    "% complete",
}

EDITABLE_MAIN_ACTUAL_HEADERS = {
    "actual start",
    "actual finish",
    "% complete",
    "physical %",
}

GENERATED_READ_ONLY = {
    "main_monthly",
    "Payment",
    "Dashboard",
    "progress",
    "progress_table",
    "Dashboard_Data",
}


def _set_unlocked(cell) -> None:
    old = copy(cell.protection)
    cell.protection = Protection(
        locked=False,
        hidden=old.hidden,
    )


def _find_headers(ws) -> tuple[int, dict[str, int]]:
    required = {"row type", "p/a"}
    for row in range(1, min(ws.max_row, 30) + 1):
        headers: dict[str, int] = {}
        for col in range(1, ws.max_column + 1):
            text = str(ws.cell(row, col).value or "").strip().lower()
            if text:
                headers[text] = col
        if required.issubset(headers):
            return row, headers
    return 4, {}


def _protect_sheet(ws) -> None:
    p = ws.protection
    p.sheet = True
    p.password = WORKBOOK_SHEET_PASSWORD

    # Keep normal Excel usage practical.
    # In sheetProtection XML, False means the action is not blocked.
    p.selectLockedCells = False
    p.selectUnlockedCells = False
    p.formatCells = False
    p.formatColumns = False
    p.formatRows = False
    p.sort = False
    p.autoFilter = False

    # Structural editing is allowed only on main below.
    p.insertRows = True
    p.deleteRows = True
    p.insertColumns = True
    p.deleteColumns = True


def _protect_main(ws) -> None:
    """Protect formulas/headers while keeping schedule editing practical."""
    header_row, headers = _find_headers(ws)
    _protect_sheet(ws)

    # main is the authoritative editing sheet, so row operations remain available.
    ws.protection.insertRows = False
    ws.protection.deleteRows = False

    row_type_col = headers.get("row type")
    pa_col = headers.get("p/a")

    fixed_end = headers.get("xml amount", 0)
    if fixed_end <= 0:
        fixed_end = max(headers.values(), default=0)

    for row in range(header_row + 1, ws.max_row + 1):
        row_type = (
            str(ws.cell(row, row_type_col).value or "").strip().lower()
            if row_type_col else ""
        )
        pa = (
            str(ws.cell(row, pa_col).value or "").strip().upper()
            if pa_col else ""
        )

        # Only Activity rows are intended for direct weekly input.
        is_activity = row_type == "activity"

        editable_headers = (
            EDITABLE_MAIN_PLAN_HEADERS
            if pa == "P"
            else EDITABLE_MAIN_ACTUAL_HEADERS
            if pa == "A"
            else set()
        )

        for header in editable_headers:
            col = headers.get(header)
            if not col:
                continue
            cell = ws.cell(row, col)
            if not (isinstance(cell.value, str) and cell.value.startswith("=")):
                _set_unlocked(cell)

        # Plan/Actual Activity timescale cells are editable when they are not
        # formulas. This preserves the existing Excel workflow for manual progress
        # and payment-alignment adjustments while keeping WBS/project rollups locked.
        if is_activity and pa in {"P", "A"} and fixed_end < ws.max_column:
            for col in range(fixed_end + 1, ws.max_column + 1):
                cell = ws.cell(row, col)
                if not (isinstance(cell.value, str) and cell.value.startswith("=")):
                    _set_unlocked(cell)




def _protect_dashboard(ws) -> None:
    _protect_sheet(ws)

    # Interactive controls stay editable while KPI/chart/formula/layout cells stay locked.
    # G5 = Weekly/Monthly, K5 = Cutoff Date, P37 = Activity Status Focus.
    _set_unlocked(ws["G5"])
    _set_unlocked(ws["K5"])
    _set_unlocked(ws["P37"])

def _protect_payment_input(ws) -> None:
    _protect_sheet(ws)

    # Payment Input hierarchy is generated/reconciled from main. Only ACT payment
    # percentages are user-editable.
    header_row = 6
    first_data_row = 8
    first_payment_col = 5

    for row in range(first_data_row, ws.max_row + 1):
        if str(ws.cell(row, 1).value or "").strip().upper() != "ACT":
            continue
        for col in range(first_payment_col, ws.max_column + 1):
            _set_unlocked(ws.cell(row, col))


def apply_final_sheet_protection(workbook) -> tuple[str, ...]:
    """Apply lightweight, deterministic protection to the final workbook.

    Workbook structure itself is intentionally NOT protected.
    """
    protected: list[str] = []

    for ws in workbook.worksheets:
        if ws.title == "main":
            _protect_main(ws)
        elif ws.title == "Payment Input":
            _protect_payment_input(ws)
        elif ws.title == "Dashboard":
            _protect_dashboard(ws)
        else:
            _protect_sheet(ws)

        protected.append(ws.title)

    # Do not protect workbook structure; Rebuild can replace sheets freely and
    # advanced Excel users can still manage normal hidden public data sheets.
    if workbook.security is None:
        workbook.security = WorkbookProtection(
            lockStructure=False,
            lockWindows=False,
        )
    else:
        workbook.security.lockStructure = False
        workbook.security.lockWindows = False

    return tuple(protected)
