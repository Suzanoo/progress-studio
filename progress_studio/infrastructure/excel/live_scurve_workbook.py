from __future__ import annotations

from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from progress_studio.domain.main_dataset import MainDataset


PROGRESS_SHEET = "progress"


def _rows(ws, dataset: MainDataset) -> dict[str, int]:
    headers = {name: col for name, col in dataset.headers}
    row_type_col = headers.get("row type")
    pa_col = headers.get("p/a")
    if row_type_col is None or pa_col is None:
        return {}

    found: dict[str, int] = {}
    for row in range(dataset.header_row + 1, ws.max_row + 1):
        row_type = str(ws.cell(row, row_type_col).value or "").strip().lower()
        pa = str(ws.cell(row, pa_col).value or "").strip().upper()
        if row_type == "project summary":
            if pa == "P":
                found["project_plan"] = row
            elif pa == "A":
                found["project_actual"] = row
        elif row_type == "s-curve":
            if pa == "P":
                found["plan"] = row
            elif pa == "AP":
                found["acc_plan"] = row
            elif pa == "A":
                found["actual"] = row
            elif pa == "AA":
                found["acc_actual"] = row
    return found


def normalize_weekly_scurve_source_contract(
    workbook,
    dataset: MainDataset,
    *,
    sheet_name: str = "main",
) -> bool:
    """LW-11.1: keep ``main`` independent from Dashboard controls.

    ``main`` is the editable source of truth. Its four S-Curve rows are rebuilt
    from Project Summary Plan/Actual rows only; no Dashboard cutoff reference is
    allowed here.  Cutoff ownership belongs to the lightweight ``progress``
    calculation contract.
    """
    if sheet_name not in workbook.sheetnames or not dataset.periods:
        return False

    ws = workbook[sheet_name]
    rows = _rows(ws, dataset)
    required = {"project_plan", "project_actual", "plan", "acc_plan", "actual", "acc_actual"}
    if not required.issubset(rows):
        return False

    first_col = dataset.periods[0].column
    first_letter = get_column_letter(first_col)

    for period in dataset.periods:
        col = period.column
        letter = get_column_letter(col)
        ws.cell(rows["plan"], col).value = f"={letter}{rows['project_plan']}"
        ws.cell(rows["acc_plan"], col).value = (
            f'=IF(COUNT(${first_letter}{rows["plan"]}:{letter}{rows["plan"]})=0,"",'
            f'SUM(${first_letter}{rows["plan"]}:{letter}{rows["plan"]}))'
        )
        ws.cell(rows["actual"], col).value = (
            f'=IF(COUNT({letter}{rows["project_actual"]})=0,"",'
            f'{letter}{rows["project_actual"]})'
        )
        ws.cell(rows["acc_actual"], col).value = (
            f'=IF(COUNT(${first_letter}{rows["actual"]}:{letter}{rows["actual"]})=0,"",'
            f'SUM(${first_letter}{rows["actual"]}:{letter}{rows["actual"]}))'
        )
    return True


def build_live_progress_contract(
    workbook,
    dataset: MainDataset,
    *,
    source_sheet: str = "main",
    target_sheet: str = PROGRESS_SHEET,
    cutoff_ref: str = "Dashboard!$K$5",
) -> int:
    """Build the lightweight LW-11 S-Curve history contract.

    Contract columns are deliberately tiny: Date, Plan and Actual. Both Plan and
    Actual expose their complete cumulative history from ``main``.  Cutoff is a
    presentation concern owned by ``Dashboard_Data``; changing Dashboard cutoff
    must never recalculate or rewrite the underlying Actual curve.

    ``cutoff_ref`` is retained only for API compatibility with LW-11.1 callers.
    """
    del cutoff_ref
    if source_sheet not in workbook.sheetnames:
        raise ValueError(f"Progress source worksheet was not found: {source_sheet}")
    if not dataset.periods:
        raise ValueError("Weekly periods were not found in MainDataset.")

    normalize_weekly_scurve_source_contract(workbook, dataset, sheet_name=source_sheet)
    source = workbook[source_sheet]
    rows = _rows(source, dataset)

    # Legacy compatibility: some early Live fixtures/workbooks have no explicit
    # S-Curve rows. Prefer a live Project Summary cumulative source when present;
    # otherwise use the already-derived tiny cache as a static compatibility
    # fallback. Current Progress Studio workbooks always take the S-Curve path.
    project_plan = rows.get("project_plan")
    project_actual = rows.get("project_actual")
    if project_plan is not None and project_actual is None:
        pa_col = {name: col for name, col in dataset.headers}.get("p/a")
        candidate = project_plan + 1
        if pa_col is not None and candidate <= source.max_row:
            if str(source.cell(candidate, pa_col).value or "").strip().upper() == "A":
                project_actual = candidate

    has_scurve = {"acc_plan", "acc_actual"}.issubset(rows)
    has_project_summary = project_plan is not None and project_actual is not None
    fallback_cache = None
    if not has_scurve and not has_project_summary:
        from progress_studio.services.progress_cache_deriver import ProgressCacheDeriver
        fallback_cache = ProgressCacheDeriver().derive(dataset)

    if target_sheet in workbook.sheetnames:
        del workbook[target_sheet]
    ws = workbook.create_sheet(target_sheet)
    ws.append(["Date", "Plan", "Actual"])

    source_ref = "'" + source_sheet.replace("'", "''") + "'"
    first_letter = get_column_letter(dataset.periods[0].column)
    for out_row, period in enumerate(dataset.periods, start=2):
        letter = get_column_letter(period.column)
        ws.cell(out_row, 1, period.reporting_date)
        if has_scurve:
            plan_source = f'{source_ref}!{letter}{rows["acc_plan"]}'
            actual_source = f'{source_ref}!{letter}{rows["acc_actual"]}'
            ws.cell(out_row, 2, f'={plan_source}')
            ws.cell(
                out_row,
                3,
                f'=IF({actual_source}="","",{actual_source})',
            )
        elif has_project_summary:
            plan_range = f'{source_ref}!${first_letter}${project_plan}:{letter}${project_plan}'
            actual_range = f'{source_ref}!${first_letter}${project_actual}:{letter}${project_actual}'
            ws.cell(out_row, 2, f'=IF(COUNT({plan_range})=0,"",SUM({plan_range}))')
            ws.cell(
                out_row,
                3,
                f'=IF(COUNT({actual_range})=0,"",SUM({actual_range}))',
            )
        else:
            point = fallback_cache.points[out_row - 2]
            ws.cell(out_row, 2, point.plan_cumulative)
            actual = point.actual_cumulative
            ws.cell(
                out_row,
                3,
                "" if actual is None else actual,
            )
        ws.cell(out_row, 1).number_format = "dd/mm/yyyy"
        ws.cell(out_row, 2).number_format = "0.00%"
        ws.cell(out_row, 3).number_format = "0.00%"

    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:C{ws.max_row}"
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.sheet_state = "hidden"
    return len(dataset.periods)


# Backward-compatible entry point retained for callers/tests from LW-10.0.3.
# Its old behavior (putting cutoff logic in main) is intentionally retired.
def apply_weekly_scurve_cutoff_contract(
    workbook,
    dataset: MainDataset,
    *,
    sheet_name: str = "main",
    cutoff_ref: str = "Dashboard!$K$5",
) -> bool:
    del cutoff_ref
    return normalize_weekly_scurve_source_contract(
        workbook,
        dataset,
        sheet_name=sheet_name,
    )
