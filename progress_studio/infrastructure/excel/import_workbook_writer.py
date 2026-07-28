from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from progress_studio.config import WORKBOOK_SCHEMA
from progress_studio.domain import Activity, ActivityWbsSequencer


class ImportWorkbookWriter:
    def write(self, output_file: Path, source_file: Path, project_name: str, rows: list[Activity]) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = WORKBOOK_SCHEMA.main_sheet
        headers = [
            "Row Type", "WBS", "Description", "Activity ID", "Task ID", "UID",
            "Outline Level", "Plan Start", "Plan Finish", "Actual Start", "Actual Finish",
            "% Complete", "Physical %", "Total Float (hr)", "XML Amount",
        ]
        ws.append(headers)
        header_fill = PatternFill("solid", fgColor="1F4E78")
        wbs_fill = PatternFill("solid", fgColor="F4B183")
        thin = Side(style="thin", color="D9D9D9")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        activity_wbs = ActivityWbsSequencer()
        current_parent_wbs = ""

        for excel_row, row in enumerate(rows, start=2):
            total_float_hours = row.total_slack_minutes / 60.0 if row.total_slack_minutes is not None else None
            if row.is_summary:
                current_parent_wbs = str(row.wbs or "").strip()
                display_wbs = current_parent_wbs
            else:
                parent_wbs = current_parent_wbs or str(row.wbs or "").strip()
                display_wbs = activity_wbs.next_code(parent_wbs, fallback=row.activity_id)

            ws.append([
                "WBS" if row.is_summary else "Activity",
                display_wbs,
                row.name,
                row.activity_id if not row.is_summary else None,
                row.task_id, row.uid, row.outline_level,
                row.plan_start, row.plan_finish, row.actual_start, row.actual_finish,
                row.percent_complete / 100.0 if row.percent_complete is not None else None,
                row.physical_percent_complete / 100.0 if row.physical_percent_complete is not None else None,
                total_float_hours,
                row.amount if not row.is_summary else None,
            ])
            ws.row_dimensions[excel_row].outlineLevel = min(max(row.outline_level - 1, 0), 7)
            ws.cell(excel_row, 3).alignment = Alignment(indent=min(max(row.outline_level - 1, 0), 15), vertical="center")
            for col in range(1, len(headers) + 1):
                ws.cell(excel_row, col).border = border
                if row.is_summary:
                    ws.cell(excel_row, col).fill = wbs_fill
                    ws.cell(excel_row, col).font = Font(bold=True)

        for row_no in range(2, ws.max_row + 1):
            for col in (8, 9, 10, 11):
                ws.cell(row_no, col).number_format = "dd/mm/yyyy"
            for col in (12, 13):
                ws.cell(row_no, col).number_format = "0.00%"
            ws.cell(row_no, 14).number_format = "0.00"
            ws.cell(row_no, 15).number_format = '#,##0.00'

        widths = {"A": 12, "B": 16, "C": 48, "D": 14, "E": 11, "F": 11, "G": 14,
                  "H": 19, "I": 19, "J": 19, "K": 19, "L": 13, "M": 13, "N": 16, "O": 16}
        for column, width in widths.items():
            ws.column_dimensions[column].width = width
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:N{ws.max_row}"
        ws.sheet_properties.outlinePr.summaryBelow = False
        ws.sheet_view.showGridLines = False

        info = wb.create_sheet(WORKBOOK_SCHEMA.info_sheet)
        info.append(["Project", project_name])
        info.append(["Source", str(source_file)])
        info.append(["WBS / Summary Rows", sum(row.is_summary for row in rows)])
        info.append(["Activity Rows", sum(not row.is_summary for row in rows)])
        info.column_dimensions["A"].width = 22
        info.column_dimensions["B"].width = 70
        output_file.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_file)
