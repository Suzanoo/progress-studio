from __future__ import annotations

from collections import defaultdict
from pathlib import Path
import os
import shutil
import tempfile
from copy import copy
from collections.abc import Callable

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from progress_studio.domain.export_models import ExportResult, ExportValidation
from progress_studio.domain.mapping_models import ActivityRow, AllocationRecord, BOQRow, SupplementalWBS
from progress_studio.domain.working_tree import WorkingNodeKind, WorkingNodeOrigin, WorkingTreeNode, WorkingScheduleTree
from progress_studio.infrastructure.excel.xlsx_package_validator import validate_xlsx_tables
from progress_studio.infrastructure.excel.workbook_visibility import apply_final_sheet_visibility
from progress_studio.infrastructure.excel.amount_workbook import find_header, normalize_header
from progress_studio.infrastructure.excel.mapping_reader import validate_progress_workbook_contract
from progress_studio.infrastructure.excel.calculation_policy import configure_incremental_excel_recalculation
from progress_studio.infrastructure.excel.activity_data_theme import apply_activity_data_wbs_hierarchy
from progress_studio.infrastructure.excel.dashboard_workbook import build_dashboard
from progress_studio.infrastructure.excel.monthly_main_workbook import build_monthly_main_view
from progress_studio.infrastructure.excel.okd_workbook import OKDExportError, build_progress_views_from_source
from progress_studio.infrastructure.excel.worksheet_filters import configure_filter_buttons
from progress_studio.infrastructure.excel.edited_workbook_migrator import migrate_edited_main_into_workbook
from progress_studio.infrastructure.excel.payment_input_reader import PaymentInputSparseReader
from progress_studio.infrastructure.excel.payment_input_workbook import PaymentInputWorkbook
from progress_studio.infrastructure.excel.payment_workbook import PaymentWorkbookError
from progress_studio.services.payment_service import PaymentService
from progress_studio.services.working_tree_schedule_source import WorkingTreeScheduleSource
from progress_studio.services.workbook_generation_service import WorkbookGenerationService

CURRENCY_FORMAT = '#,##0.00'
PERCENT_FORMAT = '0.00%'
MAPPING_SHEET = 'BOQ Activity Mapping'
SUMMARY_SHEET = 'Mapping Summary'
EXTENSION_SHEET = 'ProgressStudio Extensions'


def _headers(ws, row: int = 1) -> dict[str, int]:
    return {
        str(ws.cell(row, col).value or '').strip().lower(): col
        for col in range(1, ws.max_column + 1)
        if str(ws.cell(row, col).value or '').strip()
    }


def _safe_remove_sheet(workbook, name: str) -> None:
    if name in workbook.sheetnames:
        del workbook[name]


def _style_header(ws, row: int = 1) -> None:
    for cell in ws[row]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='4472C4')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


def _atomic_replace(temp_file: Path, output_file: Path) -> None:
    output_file.parent.mkdir(parents=True, exist_ok=True)
    os.replace(temp_file, output_file)


class MappedWorkbookExporter:
    """Create a final mapped workbook without reading data from GUI widgets."""

    def export(
        self,
        progress_file: Path,
        output_file: Path,
        boq_rows: list[BOQRow],
        allocations: list[AllocationRecord],
        validation: ExportValidation,
        *,
        activities: list[ActivityRow] | None = None,
        supplemental_wbs: list[SupplementalWBS] | None = None,
        working_tree_nodes: list[WorkingTreeNode] | None = None,
        overwrite: bool = False,
        progress_callback: Callable[[str, str, bool], None] | None = None,
        edited_workbook: Path | None = None,
    ) -> ExportResult:
        progress_file = Path(progress_file).resolve()
        output_file = Path(output_file).resolve()
        if not progress_file.is_file():
            raise ValueError(f'Progress workbook was not found: {progress_file}')
        if progress_file == output_file:
            raise ValueError('Export output must be different from the loaded Progress workbook.')
        if output_file.exists() and not overwrite:
            raise FileExistsError(f'Export file already exists: {output_file}')
        if output_file.suffix.lower() != '.xlsx':
            raise ValueError('Export filename must use the .xlsx extension.')
        if edited_workbook is not None and output_file == Path(edited_workbook).expanduser().resolve():
            raise ValueError('Rebuild output must be different from the edited input workbook.')

        source_workbook = load_workbook(progress_file, read_only=False, data_only=False)
        try:
            validate_progress_workbook_contract(source_workbook)
            can_generate = (
                bool(working_tree_nodes)
                and self._supports_main_rebuild(source_workbook)
                and "Timescale Info" in source_workbook.sheetnames
            )
        finally:
            source_workbook.close()

        output_file.parent.mkdir(parents=True, exist_ok=True)
        fd, temp_name = tempfile.mkstemp(
            prefix=f'.{output_file.stem}.', suffix='.tmp.xlsx', dir=output_file.parent
        )
        os.close(fd)
        temp_file = Path(temp_name)
        try:
            totals = self._allocation_totals(boq_rows, allocations)
            activities = activities or []
            if can_generate:
                source = WorkingTreeScheduleSource(progress_file, working_tree_nodes or [], totals)
                cutoff_day = self._read_cutoff_day(progress_file)
                generation = WorkbookGenerationService().generate(
                    source,
                    temp_file,
                    cutoff_day=cutoff_day,
                    distribution_method="auto",
                    amounts=totals,
                    progress_callback=progress_callback,
                )
                amount_rows = generation.activity_count
            else:
                shutil.copy2(progress_file, temp_file)
                amount_rows = 0

            if progress_callback is not None and not can_generate:
                for step, message in (("read", "Source workbook loaded."), ("main", "Main schedule prepared."), ("timescale", "Existing timescale preserved."), ("mapping", "Mapped amounts prepared."), ("progress", "Existing progress sheets preserved."), ("distribution", "Existing distribution preserved."), ("okd", "Existing OKD sheets preserved."), ("monthly", "Monthly main view prepared.")):
                    progress_callback(step, message, True)
            # Payment Input is persistent user data. Read it sparsely from the
            # edited workbook (or current source) before rebuilding generated sheets.
            preserved_payment = None
            payment_source = Path(edited_workbook).expanduser().resolve() if edited_workbook is not None else progress_file
            try:
                preserved_payment = PaymentInputSparseReader().read(payment_source)
            except PaymentWorkbookError:
                preserved_payment = None

            workbook = load_workbook(temp_file)
            migration = None
            payment_reconcile = None
            try:
                validate_progress_workbook_contract(workbook)
                if not can_generate:
                    if working_tree_nodes and self._supports_main_rebuild(workbook):
                        amount_rows = self._rebuild_main_from_working_tree(
                            workbook, working_tree_nodes, totals
                        )
                    else:
                        amount_rows = self._write_main_amounts(
                            workbook, totals, validation.allocated_amount
                        )
                self._write_amount_mapping(workbook, totals)
                self._write_extension_sheet(workbook, activities, totals, supplemental_wbs or [])
                mapping_rows = self._write_mapping_sheet(workbook, boq_rows, allocations)
                if edited_workbook is not None:
                    if progress_callback is not None:
                        progress_callback("migrate", "Reading user edits from existing workbook...", False)
                    migration = migrate_edited_main_into_workbook(workbook, edited_workbook)
                    if progress_callback is not None:
                        progress_callback(
                            "migrate",
                            f"Migrated {migration.matched_activity_count}/{migration.source_activity_count} activities.",
                            True,
                        )
                self._write_summary_sheet(workbook, validation, progress_file.name, output_file.name)
                apply_activity_data_wbs_hierarchy(workbook['main'])
                # Rebuild project progress and the value-only activity snapshot from
                # the final mapped main sheet. This is the refresh boundary for
                # progress_table: user edits remain in main until the next rebuild.
                # Minimal legacy workbooks used by the compatibility path may not
                # contain a weekly timescale; leave their existing views untouched.
                try:
                    build_progress_views_from_source(workbook, workbook['main'])
                except OKDExportError as exc:
                    if "Weekly timescale not found" not in str(exc):
                        raise

                # progress_table is a generated snapshot/support dataset. Keep it
                # available to Dashboard/engine but out of the user's visible tabs.
                if "progress_table" in workbook.sheetnames:
                    workbook["progress_table"].sheet_state = "hidden"

                # Payment Input is persistent and reconciled; Payment itself is a
                # generated snapshot that will be replaced after this save.
                try:
                    payment_reconcile = PaymentInputWorkbook().embed(
                        workbook,
                        preserved=preserved_payment,
                    )
                except PaymentWorkbookError:
                    payment_reconcile = None

                build_monthly_main_view(workbook, require_timescale=False)
                build_dashboard(workbook, project_name=output_file.stem)
                configure_incremental_excel_recalculation(workbook)
                apply_final_sheet_visibility(workbook)
                workbook.save(temp_file)
                if progress_callback is not None:
                    progress_callback("finalize", "Workbook finalized.", True)
            finally:
                workbook.close()

            # Payment is disposable/generated: always recreate it from the final
            # main + embedded Payment Input. This also replaces any stale Payment sheet.
            if payment_reconcile is not None:
                try:
                    PaymentService().render_payment_backbones(
                        temp_file,
                        temp_file,
                        temp_file,
                    )
                    if progress_callback is not None:
                        progress_callback(
                            "payment",
                            f"Payment rebuilt from {payment_reconcile['periods']} periods.",
                            True,
                        )
                except PaymentWorkbookError:
                    # A valid embedded Payment Input may intentionally have no
                    # populated requirements yet; leave Payment absent until later.
                    pass

            validate_xlsx_tables(temp_file)
            _atomic_replace(temp_file, output_file)
            return ExportResult(output_file, validation, amount_rows, mapping_rows, migration)
        except Exception:
            temp_file.unlink(missing_ok=True)
            raise




    @staticmethod
    def _read_cutoff_day(progress_file: Path) -> str:
        """Reuse the source workbook cutoff when available."""
        workbook = load_workbook(progress_file, data_only=True, read_only=True)
        try:
            if "Timescale Info" in workbook.sheetnames:
                ws = workbook["Timescale Info"]
                for row in ws.iter_rows(min_row=1, max_col=2, values_only=True):
                    if str(row[0] or "").strip().lower() == "cutoff day":
                        return str(row[1] or "Friday")
            return "Friday"
        finally:
            workbook.close()

    @staticmethod
    def _supports_main_rebuild(workbook) -> bool:
        if 'main' not in workbook.sheetnames:
            return False
        try:
            find_header(
                workbook['main'],
                ['Row Type', 'WBS', 'Description', 'P/A', 'Activity ID',
                 'Outline Level', 'Plan Start', 'Plan Finish', 'Amount'],
            )
            return True
        except ValueError:
            # Legacy/minimal test workbooks keep the MS6 amount-update path.
            return False

    @staticmethod
    def _copy_row_style(ws, source_row: int, target_row: int) -> None:
        for col in range(1, ws.max_column + 1):
            source = ws.cell(source_row, col)
            target = ws.cell(target_row, col)
            if source.has_style:
                target._style = copy(source._style)
            if source.number_format:
                target.number_format = source.number_format
            target.alignment = copy(source.alignment)
            target.font = copy(source.font)
            target.fill = copy(source.fill)
            target.border = copy(source.border)
            target.protection = copy(source.protection)
        ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height

    @classmethod
    def _rebuild_main_from_working_tree(
        cls, workbook, nodes: list[WorkingTreeNode], totals: dict[str, float]
    ) -> int:
        """Rebuild only the editable `main` schedule tree.

        User-created activities intentionally receive blank Plan Start/Finish
        cells. This is a supported draft state: the user completes dates in the
        exported main sheet before regenerating downstream progress sheets.
        """
        ws = workbook['main']
        header_row, headers = find_header(
            ws, ['Row Type', 'WBS', 'Description', 'P/A', 'Activity ID', 'Outline Level', 'Plan Start', 'Plan Finish', 'Amount']
        )
        row_type_col = headers['row type']
        wbs_col = headers['wbs']
        desc_col = headers['description']
        pa_col = headers['p/a']
        activity_col = headers['activity id']
        outline_col = headers['outline level']
        amount_col = headers['amount']

        original_activity_rows: dict[str, tuple[int, int]] = {}
        original_wbs_rows: dict[tuple[tuple[str, str], ...], tuple[int, int]] = {}
        path_stack: list[tuple[str, str]] = []
        project_plan = project_actual = None
        wbs_template = activity_template = None
        for row in range(header_row + 1, ws.max_row + 1):
            pa = str(ws.cell(row, pa_col).value or '').strip().upper()
            if pa != 'P':
                continue
            row_type = normalize_header(ws.cell(row, row_type_col).value)
            if row_type == 'project summary':
                project_plan, project_actual = row, row + 1
            elif row_type == 'wbs':
                wbs_template = wbs_template or row
                level = int(ws.cell(row, outline_col).value or 0)
                code = str(ws.cell(row, wbs_col).value or '').strip()
                name = str(ws.cell(row, desc_col).value or '').strip()
                path_stack = path_stack[: max(level - 1, 0)]
                path_stack.append((code, name))
                original_wbs_rows[tuple(path_stack)] = (row, row + 1)
            elif row_type == 'activity':
                activity_template = activity_template or row
                activity_id = str(ws.cell(row, activity_col).value or '').strip().upper()
                if activity_id:
                    original_activity_rows[activity_id] = (row, row + 1)

        if project_plan is None or project_actual is None or wbs_template is None or activity_template is None:
            raise ValueError('The main worksheet does not contain reusable Project/WBS/Activity row templates.')

        # Snapshot source cells before deleting rows.
        max_col = ws.max_column
        source_values = {
            row: [ws.cell(row, col).value for col in range(1, max_col + 1)]
            for row in range(header_row + 1, ws.max_row + 1)
        }
        source_styles = {}
        for key, row in [('project_p', project_plan), ('project_a', project_actual), ('wbs_p', wbs_template), ('wbs_a', wbs_template + 1), ('act_p', activity_template), ('act_a', activity_template + 1)]:
            source_styles[key] = [copy(ws.cell(row, col)._style) for col in range(1, max_col + 1)]

        project_p_values = list(source_values[project_plan])
        project_a_values = list(source_values[project_actual])
        ws.delete_rows(header_row + 1, ws.max_row - header_row)

        def append_pair(kind: str, plan_values: list, actual_values: list) -> tuple[int, int]:
            plan_row = ws.max_row + 1
            ws.append(plan_values)
            actual_row = ws.max_row + 1
            ws.append(actual_values)
            for col in range(1, max_col + 1):
                ws.cell(plan_row, col)._style = copy(source_styles[f'{kind}_p'][col - 1])
                ws.cell(actual_row, col)._style = copy(source_styles[f'{kind}_a'][col - 1])
            return plan_row, actual_row

        project_plan_row, project_actual_row = append_pair('project', project_p_values, project_a_values)
        tree = WorkingScheduleTree(nodes)
        node_rows: list[tuple[int, WorkingTreeNode, int, int]] = []
        activity_counter_by_parent: dict[str, int] = defaultdict(int)

        for depth, node in tree.walk():
            if node.deleted:
                continue
            if node.kind is WorkingNodeKind.WBS:
                source_pair = original_wbs_rows.get(node.source_path)
                if source_pair:
                    plan_values = list(source_values[source_pair[0]])
                    actual_values = list(source_values[source_pair[1]])
                else:
                    plan_values = [None] * max_col
                    actual_values = [None] * max_col
                plan_values[row_type_col - 1] = 'WBS'
                plan_values[wbs_col - 1] = node.code
                plan_values[desc_col - 1] = node.name
                plan_values[pa_col - 1] = 'P'
                plan_values[activity_col - 1] = None
                plan_values[outline_col - 1] = depth
                actual_values[row_type_col - 1] = None
                actual_values[wbs_col - 1] = None
                actual_values[desc_col - 1] = None
                actual_values[pa_col - 1] = 'A'
                actual_values[activity_col - 1] = None
                actual_values[outline_col - 1] = None
                if node.origin is WorkingNodeOrigin.USER_CREATED:
                    for col_name in ('plan start', 'plan finish', 'actual start', 'actual finish'):
                        if col_name in headers:
                            plan_values[headers[col_name] - 1] = None
                    for col in range(18, max_col + 1):
                        plan_values[col - 1] = None
                        actual_values[col - 1] = None
                pr, ar = append_pair('wbs', plan_values, actual_values)
            else:
                source_pair = original_activity_rows.get(node.source_activity_id.strip().upper())
                if source_pair:
                    plan_values = list(source_values[source_pair[0]])
                    actual_values = list(source_values[source_pair[1]])
                else:
                    plan_values = [None] * max_col
                    actual_values = [None] * max_col
                parent = tree.get(node.parent_id) if node.parent_id else None
                parent_code = parent.code if parent else ''
                activity_counter_by_parent[node.parent_id or ''] += 1
                activity_wbs = f'{parent_code}.{activity_counter_by_parent[node.parent_id or ""]}' if parent_code else str(activity_counter_by_parent[node.parent_id or ''])
                plan_values[row_type_col - 1] = 'Activity'
                plan_values[wbs_col - 1] = activity_wbs
                plan_values[desc_col - 1] = node.name
                plan_values[pa_col - 1] = 'P'
                plan_values[activity_col - 1] = node.code
                plan_values[outline_col - 1] = depth
                plan_values[amount_col - 1] = totals.get(node.code.strip().upper(), 0.0)
                actual_values[row_type_col - 1] = None
                actual_values[wbs_col - 1] = None
                actual_values[desc_col - 1] = None
                actual_values[pa_col - 1] = 'A'
                actual_values[activity_col - 1] = node.code
                actual_values[outline_col - 1] = None
                actual_values[amount_col - 1] = f'={get_column_letter(amount_col)}{ws.max_row + 1}'
                if node.origin is WorkingNodeOrigin.USER_CREATED:
                    # Draft dates are deliberately blank and downstream weekly
                    # distributions are not generated in this milestone.
                    for col_name in ('plan start', 'plan finish', 'actual start', 'actual finish'):
                        if col_name in headers:
                            plan_values[headers[col_name] - 1] = None
                            actual_values[headers[col_name] - 1] = None
                    for col in range(18, max_col + 1):
                        plan_values[col - 1] = None
                        actual_values[col - 1] = None
                pr, ar = append_pair('act', plan_values, actual_values)
            node_rows.append((depth, node, pr, ar))

        amount_letter = get_column_letter(amount_col)
        row_type_letter = get_column_letter(row_type_col)
        pa_letter = get_column_letter(pa_col)
        for index in range(len(node_rows) - 1, -1, -1):
            depth, node, plan_row, actual_row = node_rows[index]
            if node.kind is WorkingNodeKind.ACTIVITY:
                ws.cell(actual_row, amount_col).value = f'={amount_letter}{plan_row}'
                continue
            end_row = plan_row
            for child_depth, _child, child_plan, child_actual in node_rows[index + 1:]:
                if child_depth <= depth:
                    break
                end_row = child_actual
            ws.cell(plan_row, amount_col).value = 0 if end_row <= plan_row else (
                f'=SUMIFS(${amount_letter}${plan_row + 1}:${amount_letter}${end_row},'
                f'${row_type_letter}${plan_row + 1}:${row_type_letter}${end_row},"Activity",'
                f'${pa_letter}${plan_row + 1}:${pa_letter}${end_row},"P")'
            )
            ws.cell(actual_row, amount_col).value = f'={amount_letter}{plan_row}'

        last_row = ws.max_row
        ws.cell(project_plan_row, amount_col).value = (
            f'=SUMIFS(${amount_letter}${project_plan_row + 2}:${amount_letter}${last_row},'
            f'${row_type_letter}${project_plan_row + 2}:${row_type_letter}${last_row},"Activity",'
            f'${pa_letter}${project_plan_row + 2}:${pa_letter}${last_row},"P")'
        )
        ws.cell(project_actual_row, amount_col).value = f'={amount_letter}{project_plan_row}'
        configure_filter_buttons(
            ws,
            header_row=header_row,
            last_row=last_row,
            last_col=max_col,
            visible_columns={row_type_col, pa_col},
        )
        return sum(1 for _d, node, _p, _a in node_rows if node.kind is WorkingNodeKind.ACTIVITY)

    @staticmethod
    def _write_extension_sheet(workbook, activities: list[ActivityRow], totals: dict[str, float], supplemental_wbs: list[SupplementalWBS]) -> None:
        _safe_remove_sheet(workbook, EXTENSION_SHEET)
        supplemental = [row for row in activities if row.is_supplemental]
        if not supplemental and not supplemental_wbs:
            return
        ws = workbook.create_sheet(EXTENSION_SHEET)
        ws.append([
            'Node Type', 'Origin', 'Parent WBS', 'Supplemental WBS Code', 'Supplemental WBS Name',
            'Activity ID', 'Activity Name', 'Mapped Amount', 'Export Contract',
        ])
        _style_header(ws)
        for node in supplemental_wbs:
            ws.append(['WBS', 'Progress Studio', ' / '.join(code for code, _ in node.parent_path), node.code, node.name, '', '', 0.0, 'Editable extension hierarchy'])
        for row in supplemental:
            ws.append([
                'Activity', 'Progress Studio', row.parent_wbs, row.supplemental_wbs_code,
                row.supplemental_wbs_name, row.activity_id, row.description,
                totals.get(row.activity_id, 0.0),
                'Extension only - not inserted into source schedule hierarchy',
            ])
            ws.cell(ws.max_row, 8).number_format = CURRENCY_FORMAT
        ws.freeze_panes = 'A2'
        for index, width in enumerate((14, 18, 22, 24, 32, 18, 45, 18, 55), start=1):
            ws.column_dimensions[ws.cell(1, index).column_letter].width = width

    @staticmethod
    def _allocation_totals(boq_rows, allocations) -> dict[str, float]:
        by_key = {row.key: row for row in boq_rows}
        totals: dict[str, float] = defaultdict(float)
        for allocation in allocations:
            row = by_key.get(allocation.boq_key)
            if row is None:
                raise ValueError(f'Allocation references missing BOQ item: {allocation.boq_key}')
            totals[allocation.activity_id.strip().upper()] += (
                row.amount * allocation.share_percent / 100.0
            )
        return dict(totals)

    @staticmethod
    def _write_main_amounts(workbook, totals: dict[str, float], expected_total: float) -> int:
        ws = workbook['main']
        header_row, headers = find_header(
            ws, ['Row Type', 'Activity ID', 'P/A', 'Outline Level', 'Amount']
        )
        row_type_col = headers['row type']
        activity_col = headers['activity id']
        pa_col = headers['p/a']
        outline_col = headers['outline level']
        amount_col = headers['amount']

        plan_rows: list[dict[str, object]] = []
        activity_rows: dict[str, int] = {}
        for row_index in range(header_row + 1, ws.max_row + 1):
            if str(ws.cell(row_index, pa_col).value or '').strip().upper() != 'P':
                continue
            row_type = normalize_header(ws.cell(row_index, row_type_col).value)
            if row_type not in {'project summary', 'wbs', 'activity'}:
                continue
            try:
                level = int(ws.cell(row_index, outline_col).value or 0)
            except (TypeError, ValueError):
                raise ValueError(f'Invalid Outline Level in main worksheet row {row_index}.')
            plan_rows.append({'row': row_index, 'row_type': row_type, 'level': level})
            if row_type == 'activity':
                activity_id = str(ws.cell(row_index, activity_col).value or '').strip().upper()
                if not activity_id:
                    raise ValueError(f'Blank Activity ID in main worksheet row {row_index}.')
                if activity_id in activity_rows:
                    raise ValueError(f'Duplicate Activity ID in main worksheet: {activity_id}')
                activity_rows[activity_id] = row_index

        missing = sorted(set(totals) - set(activity_rows))
        if missing:
            preview = ', '.join(missing[:10])
            suffix = '' if len(missing) <= 10 else f' (+{len(missing) - 10} more)'
            raise ValueError('Mapped Activity IDs were not found in main worksheet: ' + preview + suffix)

        for activity_id, row_index in activity_rows.items():
            ws.cell(row_index, amount_col).value = totals.get(activity_id, 0.0)
            ws.cell(row_index, amount_col).number_format = CURRENCY_FORMAT
            if row_index + 1 <= ws.max_row and str(ws.cell(row_index + 1, pa_col).value or '').strip().upper() == 'A':
                ws.cell(row_index + 1, amount_col).value = None

        amount_letter = get_column_letter(amount_col)
        row_type_letter = get_column_letter(row_type_col)
        pa_letter = get_column_letter(pa_col)
        for index in range(len(plan_rows) - 1, -1, -1):
            item = plan_rows[index]
            row_index = int(item['row'])
            row_type = str(item['row_type'])
            level = int(item['level'])
            if row_type not in {'wbs', 'project summary'}:
                continue
            end_row = row_index
            for descendant in plan_rows[index + 1:]:
                if int(descendant['level']) <= level:
                    break
                end_row = int(descendant['row'])
            ws.cell(row_index, amount_col).value = 0 if end_row <= row_index else (
                f'=SUMIFS(${amount_letter}${row_index + 1}:${amount_letter}${end_row},'
                f'${row_type_letter}${row_index + 1}:${row_type_letter}${end_row},"Activity",'
                f'${pa_letter}${row_index + 1}:${pa_letter}${end_row},"P")'
            )
            ws.cell(row_index, amount_col).number_format = CURRENCY_FORMAT
            if row_index + 1 <= ws.max_row and str(ws.cell(row_index + 1, pa_col).value or '').strip().upper() == 'A':
                ws.cell(row_index + 1, amount_col).value = None

        written_total = sum(float(totals.get(activity_id, 0.0)) for activity_id in activity_rows)
        if abs(written_total - expected_total) > 0.01:
            raise ValueError(
                'Export reconciliation failed: main Activity Amount total '
                f'{written_total:,.2f} does not match allocated amount {expected_total:,.2f}.'
            )
        return len(activity_rows)

    @staticmethod
    def _write_amount_mapping(workbook, totals: dict[str, float]) -> None:
        if 'Amount Mapping' not in workbook.sheetnames:
            raise ValueError("Worksheet 'Amount Mapping' was not found in the Progress workbook.")
        ws = workbook['Amount Mapping']
        headers = _headers(ws)
        missing = [name for name in ('activity id', 'amount') if name not in headers]
        if missing:
            raise ValueError('Missing Amount Mapping columns: ' + ', '.join(missing))
        activity_col = headers['activity id']
        amount_col = headers['amount']
        status_col = headers.get('status')
        for row_index in range(2, ws.max_row + 1):
            activity_id = str(ws.cell(row_index, activity_col).value or '').strip().upper()
            if not activity_id:
                continue
            ws.cell(row_index, amount_col).value = totals.get(activity_id, 0.0)
            ws.cell(row_index, amount_col).number_format = CURRENCY_FORMAT
            if status_col:
                ws.cell(row_index, status_col).value = 'MAPPED' if totals.get(activity_id, 0.0) > 0 else 'UNMAPPED'

    @staticmethod
    def _write_mapping_sheet(workbook, boq_rows, allocations) -> int:
        _safe_remove_sheet(workbook, MAPPING_SHEET)
        ws = workbook.create_sheet(MAPPING_SHEET)
        headers = [
            'Activity ID', 'BOQ Key', 'Source Sheet', 'Source Row',
            'WBS-2', 'WBS-3', 'WBS-4', 'BOQ Description', 'BOQ Amount',
            'Share %', 'Allocated Amount', 'Mapping ID', 'BOQ ID',
        ]
        ws.append(headers)
        _style_header(ws)
        by_key = {row.key: row for row in boq_rows}
        for index, allocation in enumerate(sorted(allocations, key=lambda x: (x.activity_id, x.boq_key)), start=1):
            row = by_key.get(allocation.boq_key)
            if row is None:
                raise ValueError(f'Allocation references missing BOQ item: {allocation.boq_key}')
            allocated = row.amount * allocation.share_percent / 100.0
            mapping_id = f'MAP-{index:06d}'
            ws.append([
                allocation.activity_id, row.key, row.source_sheet, row.source_row,
                row.wbs2, row.wbs3, row.wbs4, row.description, row.amount,
                allocation.share_percent / 100.0, allocated, mapping_id, row.stable_id or row.key,
            ])
            current = ws.max_row
            ws.cell(current, 9).number_format = CURRENCY_FORMAT
            ws.cell(current, 10).number_format = PERCENT_FORMAT
            ws.cell(current, 11).number_format = CURRENCY_FORMAT

        ws.freeze_panes = 'A2'
        widths = [16, 32, 22, 12, 22, 28, 28, 60, 18, 12, 18, 16, 20]
        for index, width in enumerate(widths, start=1):
            ws.column_dimensions[ws.cell(1, index).column_letter].width = width
        if ws.max_row >= 2:
            table = Table(displayName='BOQActivityMappingTable', ref=f'A1:M{ws.max_row}')
            table.tableStyleInfo = TableStyleInfo(
                name='TableStyleMedium2', showFirstColumn=False, showLastColumn=False,
                showRowStripes=True, showColumnStripes=False,
            )
            ws.add_table(table)
        return len(allocations)

    @staticmethod
    def _write_summary_sheet(workbook, validation, source_name: str, output_name: str) -> None:
        _safe_remove_sheet(workbook, SUMMARY_SHEET)
        ws = workbook.create_sheet(SUMMARY_SHEET, 0)
        ws['A1'] = 'Progress Studio Mapping Reconciliation'
        ws['A1'].font = Font(size=14, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill('solid', fgColor='1F4E78')
        ws.merge_cells('A1:B1')
        rows = [
            ('Source workbook', source_name),
            ('Export workbook', output_name),
            ('Export status', 'Complete' if validation.is_complete else 'Partial'),
            ('Activities', validation.activity_count),
            ('BOQ items', validation.boq_count),
            ('Allocation records', validation.allocation_count),
            ('Mapped activities', validation.mapped_activity_count),
            ('Fully allocated BOQ items', validation.full_boq_count),
            ('Partially allocated BOQ items', validation.partial_boq_count),
            ('Unmapped BOQ items', validation.unmapped_boq_count),
            ('Total BOQ amount', validation.total_boq_amount),
            ('Allocated amount', validation.allocated_amount),
            ('Remaining amount', validation.remaining_amount),
            ('Allocated percent', validation.allocated_percent / 100.0),
        ]
        for row_index, (label, value) in enumerate(rows, start=3):
            ws.cell(row_index, 1).value = label
            ws.cell(row_index, 1).font = Font(bold=True)
            ws.cell(row_index, 2).value = value
        for row_index in (13, 14, 15):
            ws.cell(row_index, 2).number_format = CURRENCY_FORMAT
        ws.cell(16, 2).number_format = PERCENT_FORMAT
        ws.column_dimensions['A'].width = 34
        ws.column_dimensions['B'].width = 28
        ws.freeze_panes = 'A3'
