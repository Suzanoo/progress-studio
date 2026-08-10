from __future__ import annotations

from dataclasses import dataclass
from datetime import date
from pathlib import Path
from openpyxl import load_workbook

from progress_studio.domain.rebuild_models import EditedWorkbookMigrationResult
from progress_studio.infrastructure.excel.amount_workbook import find_header, normalize_header, to_amount
from progress_studio.infrastructure.excel.okd_workbook import OKDExportError, as_date, find_week_columns

AUDIT_SHEET = "Rebuild Audit"

@dataclass(slots=True)
class _ActivitySnapshot:
    activity_id: str
    description: str
    plan_start: date | None
    plan_finish: date | None
    amount: float | None
    plan_values: dict[date, float | None]
    actual_values: dict[date, float | None]

    @property
    def signature(self):
        return (" ".join(self.description.lower().split()), self.plan_start, self.plan_finish)

def _numeric_or_blank(value):
    if value in (None, ""):
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    return None


def _find_source_main_sheet(workbook):
    if "main" in workbook.sheetnames:
        return workbook["main"]
    required = ["Row Type", "Description", "P/A", "Activity ID", "Plan Start", "Plan Finish", "Amount"]
    matches = []
    for ws in workbook.worksheets:
        try:
            header_row, _headers = find_header(ws, required)
            if header_row == 4:
                find_week_columns(ws, workbook.epoch)
                matches.append(ws)
        except (ValueError, OKDExportError):
            # Only worksheets that satisfy both the main headers and weekly timescale contract qualify.
            continue
    if len(matches) == 1:
        return matches[0]
    if not matches:
        raise ValueError("Edited workbook does not contain a Progress Studio main schedule worksheet.")
    names = ", ".join(ws.title for ws in matches)
    raise ValueError(f"Edited workbook contains multiple possible main schedule worksheets: {names}")

def _read_activity_snapshots(path: Path) -> list[_ActivitySnapshot]:
    wb = load_workbook(path, data_only=False, read_only=False)
    try:
        ws = _find_source_main_sheet(wb)
        header_row, h = find_header(ws, ["Row Type", "Description", "P/A", "Activity ID", "Plan Start", "Plan Finish", "Amount"])
        if header_row != 4:
            raise ValueError("Edited workbook main sheet uses an unsupported header layout.")
        weeks = find_week_columns(ws, wb.epoch)
        rows = []
        for r in range(header_row + 1, ws.max_row + 1):
            if normalize_header(ws.cell(r, h["row type"]).value) != "activity":
                continue
            if str(ws.cell(r, h["p/a"]).value or "").strip().upper() != "P":
                continue
            aid = str(ws.cell(r, h["activity id"]).value or "").strip().upper()
            if not aid:
                continue
            ar = r + 1
            if ar > ws.max_row or str(ws.cell(ar, h["p/a"]).value or "").strip().upper() != "A":
                raise ValueError(f"Edited workbook Activity {aid} is missing its Actual row.")
            rows.append(_ActivitySnapshot(
                activity_id=aid,
                description=str(ws.cell(r, h["description"]).value or "").strip(),
                plan_start=as_date(ws.cell(r, h["plan start"]).value, wb.epoch),
                plan_finish=as_date(ws.cell(r, h["plan finish"]).value, wb.epoch),
                amount=to_amount(ws.cell(r, h["amount"]).value),
                plan_values={d: _numeric_or_blank(ws.cell(r, c).value) for c, d in weeks},
                actual_values={d: _numeric_or_blank(ws.cell(ar, c).value) for c, d in weeks},
            ))
        if not rows:
            raise ValueError("Edited workbook contains no Activity rows in main.")
        return rows
    finally:
        wb.close()

def migrate_edited_main_into_workbook(workbook, edited_path: Path) -> EditedWorkbookMigrationResult:
    edited_path = Path(edited_path).expanduser().resolve()
    if not edited_path.is_file() or edited_path.suffix.lower() != ".xlsx":
        raise ValueError("Select a valid edited .xlsx workbook.")
    source_rows = _read_activity_snapshots(edited_path)
    ws = workbook["main"]
    header_row, h = find_header(ws, ["Row Type", "Description", "P/A", "Activity ID", "Plan Start", "Plan Finish", "Amount"])
    if header_row != 4:
        raise ValueError("Rebuilt workbook main sheet uses an unsupported header layout.")
    week_by_date = {d: c for c, d in find_week_columns(ws, workbook.epoch)}
    by_id = {}
    by_sig = {}
    for r in range(header_row + 1, ws.max_row + 1):
        if normalize_header(ws.cell(r, h["row type"]).value) != "activity" or str(ws.cell(r, h["p/a"]).value or "").strip().upper() != "P":
            continue
        aid = str(ws.cell(r, h["activity id"]).value or "").strip().upper()
        if aid:
            by_id[aid] = r
        sig = (
            " ".join(str(ws.cell(r, h["description"]).value or "").strip().lower().split()),
            as_date(ws.cell(r, h["plan start"]).value, workbook.epoch),
            as_date(ws.cell(r, h["plan finish"]).value, workbook.epoch),
        )
        by_sig.setdefault(sig, []).append(r)

    id_count = sig_count = amount_count = plan_count = actual_count = 0
    unmatched, ambiguous, used = [], [], set()
    for src in source_rows:
        target = by_id.get(src.activity_id)
        kind = "id" if target is not None else ""
        if target is None:
            candidates = [r for r in by_sig.get(src.signature, []) if r not in used]
            if len(candidates) == 1:
                target, kind = candidates[0], "sig"
            elif len(candidates) > 1:
                ambiguous.append(src.activity_id); continue
            else:
                unmatched.append(src.activity_id); continue
        if target in used:
            ambiguous.append(src.activity_id); continue
        used.add(target)
        id_count += kind == "id"
        sig_count += kind == "sig"
        actual_row = target + 1
        if src.amount is not None:
            ws.cell(target, h["amount"]).value = src.amount
            amount_count += 1
        for d, value in src.plan_values.items():
            c = week_by_date.get(d)
            if c is not None:
                ws.cell(target, c).value = value
                plan_count += 1
        for d, value in src.actual_values.items():
            c = week_by_date.get(d)
            if c is not None:
                ws.cell(actual_row, c).value = value
                actual_count += 1
    result = EditedWorkbookMigrationResult(
        len(source_rows), len(by_id), id_count + sig_count, id_count, sig_count,
        tuple(unmatched), tuple(ambiguous), amount_count, plan_count, actual_count,
    )
    if AUDIT_SHEET in workbook.sheetnames:
        del workbook[AUDIT_SHEET]
    audit = workbook.create_sheet(AUDIT_SHEET)
    audit.sheet_state = "visible"
    audit.append(["Rebuild from Edited Workbook", edited_path.name])
    audit.append(["Source activities", result.source_activity_count])
    audit.append(["Target activities", result.target_activity_count])
    audit.append(["Matched activities", result.matched_activity_count])
    audit.append(["Matched by Activity ID", result.matched_by_activity_id])
    audit.append(["Matched by Description + Plan Dates", result.matched_by_signature])
    audit.append(["Amount cells migrated", result.amount_cells_migrated])
    audit.append(["Plan weekly cells migrated", result.plan_cells_migrated])
    audit.append(["Actual weekly cells migrated", result.actual_cells_migrated])
    audit.append(["Unmatched activities", len(result.unmatched_activity_ids)])
    audit.append(["Ambiguous activities", len(result.ambiguous_activity_ids)])
    if result.unmatched_activity_ids:
        audit.append([])
        audit.append(["Unmatched Activity IDs"])
        for activity_id in result.unmatched_activity_ids:
            audit.append([activity_id])
    if result.ambiguous_activity_ids:
        audit.append([])
        audit.append(["Ambiguous Activity IDs"])
        for activity_id in result.ambiguous_activity_ids:
            audit.append([activity_id])
    audit.column_dimensions["A"].width = 38
    audit.column_dimensions["B"].width = 34
    return result
