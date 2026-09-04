from __future__ import annotations

from datetime import date, datetime
from math import isclose

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from progress_studio.services.payment_breakdown_adapter import (
    PaymentBreakdownDatasetSnapshot,
)


PAYMENT_BREAKDOWN_SHEET = "Payment-Breakdown"

_FONT = "Aptos"
_NAVY = "1F4E78"
_GREEN = "70AD47"
_LIGHT_BLUE = "D9EAF7"
_LIGHT_AMBER = "FFF2CC"
_LIGHT_RED = "FCE4D6"
_BORDER = "D9E2F3"
_TEXT = "1F1F1F"
_MUTED = "666666"
_RED = "FF0000"
_WHITE = "FFFFFF"
_PROGRESS_TOLERANCE = 1e-6


def _fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def _border() -> Border:
    side = Side(style="thin", color=_BORDER)
    return Border(left=side, right=side, top=side, bottom=side)


def _display_period(period) -> str:
    reporting = period.reporting_date
    if isinstance(reporting, datetime):
        reporting = reporting.date()
    if isinstance(reporting, date):
        return f"{period.key}\n{reporting:%d-%b-%y}"
    return period.key


def _is_active_fraction(value: float) -> bool:
    """Red text only for 0% < value < 100%.

    Values numerically equal to 100% within the PB tolerance are treated as
    complete and keep the normal text colour.
    """
    return value > _PROGRESS_TOLERANCE and not isclose(
        value,
        1.0,
        rel_tol=0.0,
        abs_tol=_PROGRESS_TOLERANCE,
    ) and value < 1.0


def _write_progress_values(ws, row: int, start_col: int, values) -> None:
    for offset, raw in enumerate(values):
        value = float(raw)
        cell = ws.cell(row, start_col + offset, value)
        cell.number_format = "0.00%"
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _border()
        cell.font = Font(
            name=_FONT,
            size=10,
            color=_RED if _is_active_fraction(value) else _TEXT,
        )


def render_payment_breakdown(
    workbook,
    snapshot: PaymentBreakdownDatasetSnapshot,
    *,
    sheet_name: str = PAYMENT_BREAKDOWN_SHEET,
):
    """Replace and render the Payment-Breakdown derived worksheet.

    PB-3 is intentionally a renderer only.  It does not save the workbook,
    modify `main`, or participate in Rebuild ownership yet.
    """
    old_index = None
    if sheet_name in workbook.sheetnames:
        old_index = workbook.sheetnames.index(sheet_name)
        workbook.remove(workbook[sheet_name])

    if old_index is None:
        ws = workbook.create_sheet(sheet_name)
    else:
        ws = workbook.create_sheet(sheet_name, old_index)

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "F4"

    period_start_col = 6
    period_count = len(snapshot.periods)
    last_col = max(5, period_start_col + period_count - 1)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    title = ws.cell(1, 1, "Payment Breakdown")
    title.fill = _fill(_NAVY)
    title.font = Font(name=_FONT, size=15, bold=True, color=_WHITE)
    title.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    ws.cell(
        2,
        1,
        (
            f"Derived exact-name groups: {len(snapshot.activities)}  |  "
            f"Eligible source activities: {snapshot.eligible_source_count}  |  "
            f"Skipped source activities: {len(snapshot.skipped_activity_ids)}"
        ),
    )
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
    ws.cell(2, 1).font = Font(name=_FONT, size=9, color=_MUTED)
    ws.cell(2, 1).alignment = Alignment(horizontal="left")

    current_row = 4
    headers = ("Activity Name", "WBS", "Activity ID", "Amount", "Progress Type")

    for derived in snapshot.activities:
        block_header_row = current_row
        for col, label in enumerate(headers, start=1):
            cell = ws.cell(block_header_row, col, label)
            cell.fill = _fill(_GREEN)
            cell.font = Font(name=_FONT, size=10, bold=True, color=_WHITE)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _border()

        for index, period in enumerate(snapshot.periods, start=period_start_col):
            cell = ws.cell(block_header_row, index, _display_period(period))
            cell.fill = _fill(_GREEN)
            cell.font = Font(name=_FONT, size=9, bold=True, color=_WHITE)
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )
            cell.border = _border()

        ws.row_dimensions[block_header_row].height = 32
        current_row += 1

        for source in derived.source_activities:
            progress_row = current_row
            cumulative_row = current_row + 1

            values = (
                derived.activity_name,
                source.wbs or "",
                source.activity_id,
                source.amount,
                "Activity Progress",
            )
            for col, value in enumerate(values, start=1):
                cell = ws.cell(progress_row, col, value)
                cell.border = _border()
                cell.font = Font(name=_FONT, size=10, color=_TEXT)
                cell.alignment = Alignment(vertical="center")

            ws.cell(progress_row, 4).number_format = "#,##0.00"
            _write_progress_values(
                ws,
                progress_row,
                period_start_col,
                source.period_progress,
            )

            cumulative_values = (
                "",
                source.wbs or "",
                source.activity_id,
                source.amount,
                "Activity Cumulative",
            )
            for col, value in enumerate(cumulative_values, start=1):
                cell = ws.cell(cumulative_row, col, value)
                cell.fill = _fill(_LIGHT_AMBER)
                cell.border = _border()
                cell.font = Font(name=_FONT, size=10, color=_TEXT)
                cell.alignment = Alignment(vertical="center")

            ws.cell(cumulative_row, 4).number_format = "#,##0.00"
            _write_progress_values(
                ws,
                cumulative_row,
                period_start_col,
                source.cumulative_progress,
            )
            for col in range(period_start_col, period_start_col + period_count):
                ws.cell(cumulative_row, col).fill = _fill(_LIGHT_AMBER)

            current_row += 2

        combined_progress_row = current_row
        combined_cumulative_row = current_row + 1

        combined = (
            derived.activity_name,
            "",
            "",
            derived.total_amount,
            "Combined Progress",
        )
        for col, value in enumerate(combined, start=1):
            cell = ws.cell(combined_progress_row, col, value)
            cell.fill = _fill(_LIGHT_BLUE)
            cell.border = _border()
            cell.font = Font(name=_FONT, size=10, bold=True, color=_TEXT)
            cell.alignment = Alignment(vertical="center")
        ws.cell(combined_progress_row, 4).number_format = "#,##0.00"
        _write_progress_values(
            ws,
            combined_progress_row,
            period_start_col,
            derived.period_progress,
        )
        for col in range(period_start_col, period_start_col + period_count):
            ws.cell(combined_progress_row, col).fill = _fill(_LIGHT_BLUE)
            ws.cell(combined_progress_row, col).font = Font(
                name=_FONT,
                size=10,
                bold=True,
                color=_RED if _is_active_fraction(ws.cell(combined_progress_row, col).value) else _TEXT,
            )

        combined_cumulative = (
            derived.activity_name,
            "",
            "",
            derived.total_amount,
            "Combined Cumulative",
        )
        for col, value in enumerate(combined_cumulative, start=1):
            cell = ws.cell(combined_cumulative_row, col, value)
            cell.fill = _fill(_LIGHT_RED)
            cell.border = _border()
            cell.font = Font(name=_FONT, size=10, bold=True, color=_TEXT)
            cell.alignment = Alignment(vertical="center")
        ws.cell(combined_cumulative_row, 4).number_format = "#,##0.00"
        _write_progress_values(
            ws,
            combined_cumulative_row,
            period_start_col,
            derived.cumulative_progress,
        )
        for col in range(period_start_col, period_start_col + period_count):
            ws.cell(combined_cumulative_row, col).fill = _fill(_LIGHT_RED)
            ws.cell(combined_cumulative_row, col).font = Font(
                name=_FONT,
                size=10,
                bold=True,
                color=_RED if _is_active_fraction(ws.cell(combined_cumulative_row, col).value) else _TEXT,
            )

        current_row += 3

    widths = {
        1: 34,
        2: 14,
        3: 14,
        4: 16,
        5: 22,
    }
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width
    for col in range(period_start_col, period_start_col + period_count):
        ws.column_dimensions[get_column_letter(col)].width = 12

    return ws
