from __future__ import annotations

import copy
from datetime import date, datetime, timedelta
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError as exc:
    raise SystemExit(
        "openpyxl was not found.\nInstall it with: pip install openpyxl"
    ) from exc


DEFAULT_SHEET = "main"
WEEKDAY_NAMES = {
    1: "Monday",
    2: "Tuesday",
    3: "Wednesday",
    4: "Thursday",
    5: "Friday",
    6: "Saturday",
    7: "Sunday",
}

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
YEAR_FILL = PatternFill("solid", fgColor="44546A")
MONTH_FILL = PatternFill("solid", fgColor="D9EAF7")
WEEK_FILL = PatternFill("solid", fgColor="D9E1F2")
DATE_FILL = PatternFill("solid", fgColor="FFFFFF")
PROJECT_FILL = PatternFill("solid", fgColor="17365D")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN_GRAY = Side(style="thin", color="D9E1F2")
HEADER_BORDER = Border(
    left=THIN_GRAY,
    right=THIN_GRAY,
    top=THIN_GRAY,
    bottom=THIN_GRAY,
)


def normalize_header(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip().lower()


def parse_cutoff_day(value: str | int) -> int:
    aliases = {
        "mon": 1,
        "monday": 1,
        "tue": 2,
        "tues": 2,
        "tuesday": 2,
        "wed": 3,
        "wednesday": 3,
        "thu": 4,
        "thur": 4,
        "thurs": 4,
        "thursday": 4,
        "fri": 5,
        "friday": 5,
        "sat": 6,
        "saturday": 6,
        "sun": 7,
        "sunday": 7,
    }

    text = str(value).strip().lower()
    if text.isdigit():
        number = int(text)
        if number in WEEKDAY_NAMES:
            return number

    if text in aliases:
        return aliases[text]

    raise ValueError("Cutoff day must be 1-7 or a weekday name such as Fri.")


def ask_cutoff_day() -> int:
    print("Weekly Cutoff Day")
    print("1 = Monday")
    print("2 = Tuesday")
    print("3 = Wednesday")
    print("4 = Thursday")
    print("5 = Friday")
    print("6 = Saturday")
    print("7 = Sunday")

    while True:
        try:
            return parse_cutoff_day(input("Select : "))
        except ValueError as exc:
            print(f"ERROR: {exc}")


def resolve_input_files(input_path: Path) -> list[Path]:
    if not input_path.exists():
        raise FileNotFoundError(f"Input not found: {input_path}")

    if input_path.is_file():
        if input_path.suffix.lower() != ".xlsx":
            raise ValueError(f"Input file must be .xlsx: {input_path}")
        return [input_path]

    if input_path.is_dir():
        files = sorted(
            path
            for path in input_path.iterdir()
            if path.is_file()
            and path.suffix.lower() == ".xlsx"
            and not path.name.startswith("~$")
        )
        if not files:
            raise ValueError(f"No .xlsx files found in folder: {input_path}")
        return files

    raise ValueError(f"Input is not a supported file or folder: {input_path}")


def resolve_output_folder(output_path: Path) -> Path:
    if output_path.exists() and not output_path.is_dir():
        raise ValueError(f"--output must be a folder: {output_path}")
    output_path.mkdir(parents=True, exist_ok=True)
    return output_path


def find_sheet(workbook, sheet_name: str):
    if sheet_name in workbook.sheetnames:
        return workbook[sheet_name]

    if len(workbook.sheetnames) == 1:
        return workbook[workbook.sheetnames[0]]

    raise ValueError(
        f"Worksheet not found: '{sheet_name}' "
        f"(available worksheets: {', '.join(workbook.sheetnames)})"
    )


def get_header_map(ws, header_row: int = 1) -> dict[str, int]:
    result: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        header = normalize_header(ws.cell(header_row, col).value)
        if header:
            result[header] = col
    return result


def to_date(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def align_on_or_before(value: date, cutoff_day: int) -> date:
    # Python weekday: Monday=0 ... Sunday=6
    target = cutoff_day - 1
    delta = (value.weekday() - target) % 7
    return value - timedelta(days=delta)


def align_on_or_after(value: date, cutoff_day: int) -> date:
    target = cutoff_day - 1
    delta = (target - value.weekday()) % 7
    return value + timedelta(days=delta)


def collect_plan_range(ws, header_map: dict[str, int]) -> tuple[date, date]:
    required = ["plan start", "plan finish"]
    missing = [name for name in required if name not in header_map]
    if missing:
        raise ValueError("Required columns not found: " + ", ".join(missing))

    plan_start_col = header_map["plan start"]
    plan_finish_col = header_map["plan finish"]

    starts: list[date] = []
    finishes: list[date] = []

    for row in range(2, ws.max_row + 1):
        start = to_date(ws.cell(row, plan_start_col).value)
        finish = to_date(ws.cell(row, plan_finish_col).value)
        if start is not None:
            starts.append(start)
        if finish is not None:
            finishes.append(finish)

    if not starts or not finishes:
        raise ValueError("No Plan Start / Plan Finish values were found for the timescale.")

    return min(starts), max(finishes)


def generate_week_dates(
    earliest_start: date,
    latest_finish: date,
    cutoff_day: int,
    margin_weeks: int,
) -> list[date]:
    margin_start = earliest_start - timedelta(weeks=margin_weeks)
    margin_finish = latest_finish + timedelta(weeks=margin_weeks)

    first_week = align_on_or_before(margin_start, cutoff_day)
    last_week = align_on_or_after(margin_finish, cutoff_day)

    result: list[date] = []
    current = first_week
    while current <= last_week:
        result.append(current)
        current += timedelta(days=7)

    return result


def copy_header_style(source, target) -> None:
    target._style = copy.copy(source._style)
    target.font = copy.copy(source.font)
    target.fill = copy.copy(source.fill)
    target.border = copy.copy(source.border)
    target.alignment = copy.copy(source.alignment)
    target.protection = copy.copy(source.protection)
    target.number_format = source.number_format


def write_info_sheet(
    wb,
    cutoff_day: int,
    margin_weeks: int,
    earliest_start: date,
    latest_finish: date,
    week_dates: list[date],
) -> None:
    if "Timescale Info" in wb.sheetnames:
        del wb["Timescale Info"]

    ws = wb.create_sheet("Timescale Info")
    rows = [
        ["Setting", "Value"],
        ["Cutoff Day", WEEKDAY_NAMES[cutoff_day]],
        ["Margin Before", f"{margin_weeks} Weeks"],
        ["Margin After", f"{margin_weeks} Weeks"],
        ["Earliest Plan Start", earliest_start],
        ["Latest Plan Finish", latest_finish],
        ["First Week", week_dates[0]],
        ["Last Week", week_dates[-1]],
        ["Total Weeks", len(week_dates)],
    ]

    for row in rows:
        ws.append(row)

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = HEADER_BORDER

    for row in range(2, ws.max_row + 1):
        ws.cell(row, 1).font = Font(bold=True)
        ws.cell(row, 1).fill = PatternFill("solid", fgColor="D9EAF7")
        ws.cell(row, 1).border = HEADER_BORDER
        ws.cell(row, 2).border = HEADER_BORDER

    for row in (5, 6, 7, 8):
        ws.cell(row, 2).number_format = "dd/mm/yyyy"

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 20
    ws.freeze_panes = "A2"


def get_project_name(wb, input_file: Path) -> str:
    if "Info" in wb.sheetnames:
        info = wb["Info"]
        for row in range(1, info.max_row + 1):
            if normalize_header(info.cell(row, 1).value) == "project":
                value = info.cell(row, 2).value
                if value not in (None, ""):
                    return str(value).strip()

    name = input_file.stem
    for suffix in ("_organized_plan_actual", "_plan_actual", "_organized"):
        if name.lower().endswith(suffix):
            name = name[: -len(suffix)]
            break
    return name or "Project"


def collect_project_rollup(ws, header_map: dict[str, int]) -> dict[str, date | None]:
    pa_col = header_map.get("p/a")
    ps_col = header_map["plan start"]
    pf_col = header_map["plan finish"]
    acs_col = header_map.get("actual start")
    acf_col = header_map.get("actual finish")

    plan_starts: list[date] = []
    plan_finishes: list[date] = []
    actual_starts: list[date] = []
    actual_finishes: list[date] = []

    for row in range(2, ws.max_row + 1):
        pa = str(ws.cell(row, pa_col).value or "").strip().upper() if pa_col else ""
        if pa in ("", "P"):
            value = to_date(ws.cell(row, ps_col).value)
            if value: plan_starts.append(value)
            value = to_date(ws.cell(row, pf_col).value)
            if value: plan_finishes.append(value)
        if pa == "A":
            if acs_col:
                value = to_date(ws.cell(row, acs_col).value)
                if value: actual_starts.append(value)
            if acf_col:
                value = to_date(ws.cell(row, acf_col).value)
                if value: actual_finishes.append(value)

    return {
        "plan_start": min(plan_starts) if plan_starts else None,
        "plan_finish": max(plan_finishes) if plan_finishes else None,
        "actual_start": min(actual_starts) if actual_starts else None,
        "actual_finish": max(actual_finishes) if actual_finishes else None,
    }


def merge_group_headers(ws, row: int, first_col: int, week_dates: list[date], key_fn, label_fn, fill, font) -> None:
    group_start = first_col
    current_key = key_fn(week_dates[0])
    for index in range(1, len(week_dates) + 1):
        at_end = index == len(week_dates)
        next_key = None if at_end else key_fn(week_dates[index])
        if at_end or next_key != current_key:
            group_end = first_col + index - 1
            if group_end > group_start:
                ws.merge_cells(start_row=row, start_column=group_start, end_row=row, end_column=group_end)
            cell = ws.cell(row, group_start)
            cell.value = label_fn(week_dates[index - 1])
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = HEADER_BORDER
            for col in range(group_start, group_end + 1):
                c = ws.cell(row, col)
                c.fill = fill
                c.font = font
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = HEADER_BORDER
            group_start = group_end + 1
            current_key = next_key


def rebuild_outline_levels(ws, header_row: int, first_data_row: int) -> None:
    """Rebuild Excel row outlines from Outline Level values after inserting rows.

    openpyxl moves cells during insert_rows but does not always move RowDimension.
    Clear old outlines and rebuild them to match the current rows.
    Project Summary uses level 0 and existing data is shifted down one level.
    This allows the Project row to collapse the entire WBS hierarchy.
    """
    header_map = get_header_map(ws, header_row)
    outline_col = header_map.get("outline level")
    pa_col = header_map.get("p/a")
    row_type_col = header_map.get("row type")

    # Clear metadata that may remain attached to incorrect rows after insert_rows.
    for row in range(1, ws.max_row + 1):
        dim = ws.row_dimensions[row]
        dim.outlineLevel = 0
        dim.hidden = False
        dim.collapsed = False

    if outline_col is None:
        return

    previous_plan_level: int | None = None

    for row in range(first_data_row, ws.max_row + 1):
        row_type = normalize_header(ws.cell(row, row_type_col).value) if row_type_col else ""
        pa = str(ws.cell(row, pa_col).value or "").strip().upper() if pa_col else ""

        if row_type == "project summary":
            excel_level = 0
            previous_plan_level = 0
        elif pa == "P":
            raw_level = ws.cell(row, outline_col).value
            try:
                source_level = int(raw_level)
            except (TypeError, ValueError):
                source_level = 1

            # XML: top WBS=1, child WBS=2, Activity=3...
            # Excel: Project=0, top WBS=1, child WBS=2, Activity=3...
            excel_level = min(max(source_level, 1), 7)
            previous_plan_level = excel_level
        elif pa == "A" and previous_plan_level is not None:
            # The Actual row must stay in the same group as its paired Plan row.
            excel_level = previous_plan_level
        else:
            excel_level = 0
            previous_plan_level = None

        ws.row_dimensions[row].outlineLevel = excel_level

    # Place the summary row above its detail rows.
    ws.sheet_properties.outlinePr.summaryBelow = False
    ws.sheet_properties.outlinePr.applyStyles = True


def insert_project_summary(ws, header_map: dict[str, int], project_name: str, rollup: dict[str, date | None], first_data_row: int) -> None:
    ws.insert_rows(first_data_row, 2)

    cols = {
        "row_type": header_map.get("row type"),
        "description": header_map.get("description"),
        "pa": header_map.get("p/a"),
        "outline": header_map.get("outline level"),
        "plan_start": header_map.get("plan start"),
        "plan_finish": header_map.get("plan finish"),
        "actual_start": header_map.get("actual start"),
        "actual_finish": header_map.get("actual finish"),
    }

    plan_row = first_data_row
    actual_row = first_data_row + 1
    if cols["row_type"]: ws.cell(plan_row, cols["row_type"]).value = "Project Summary"
    if cols["description"]: ws.cell(plan_row, cols["description"]).value = project_name
    if cols["pa"]:
        ws.cell(plan_row, cols["pa"]).value = "P"
        ws.cell(actual_row, cols["pa"]).value = "A"
    if cols["outline"]: ws.cell(plan_row, cols["outline"]).value = 0
    if cols["plan_start"]: ws.cell(plan_row, cols["plan_start"]).value = rollup["plan_start"]
    if cols["plan_finish"]: ws.cell(plan_row, cols["plan_finish"]).value = rollup["plan_finish"]
    if cols["actual_start"]: ws.cell(actual_row, cols["actual_start"]).value = rollup["actual_start"]
    if cols["actual_finish"]: ws.cell(actual_row, cols["actual_finish"]).value = rollup["actual_finish"]

    for row in (plan_row, actual_row):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row, col)
            cell.fill = PROJECT_FILL
            cell.font = Font(color="FFFFFF", bold=True)
            cell.border = HEADER_BORDER
            cell.alignment = Alignment(vertical="center")
        ws.row_dimensions[row].height = 21
        ws.row_dimensions[row].outlineLevel = 0

    for col_name in ("plan_start", "plan_finish", "actual_start", "actual_finish"):
        col = cols[col_name]
        if col:
            ws.cell(plan_row if col_name.startswith("plan") else actual_row, col).number_format = "dd/mm/yyyy"


def ensure_amount_column(ws) -> int:
    """
    Create the Amount column before building the timescale.

    Do this in Script 03 before weekly columns and column grouping exist.
    This lets Script 04 write Amount values without inserting a column later.
    It also preserves timescale grouping.
    """
    header_map = get_header_map(ws, 1)

    if "amount" in header_map:
        return header_map["amount"]

    total_float_col = header_map.get("total float (hr)")
    if total_float_col is None:
        raise ValueError("The Total Float (hr) column was not found for Amount placement.")

    amount_col = total_float_col
    ws.insert_cols(amount_col, 1)

    # Copy style and width from the shifted Total Float column.
    source_col = amount_col + 1
    for row in range(1, ws.max_row + 1):
        copy_header_style(ws.cell(row, source_col), ws.cell(row, amount_col))

    source_letter = get_column_letter(source_col)
    target_letter = get_column_letter(amount_col)
    ws.column_dimensions[target_letter].width = max(16, ws.column_dimensions[source_letter].width or 0)

    ws.cell(1, amount_col).value = "Amount"
    ws.cell(1, amount_col).alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )

    # Amount remains blank in Script 03; Script 04 maps it from BOQ data.
    for row in range(2, ws.max_row + 1):
        ws.cell(row, amount_col).value = None
        ws.cell(row, amount_col).number_format = '#,##0.00'

    return amount_col



def create_amount_mapping_sheet(wb, ws, header_row: int = 4, first_data_row: int = 5) -> int:
    """Create the Amount Mapping worksheet for convenient amount entry.

    This worksheet contains one Plan activity row per Activity ID, and Script 04
    writes Amount values back to the main worksheet by matching Activity ID.
    """
    sheet_name = "Amount Mapping"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    mapping = wb.create_sheet(sheet_name)
    headers = ["Activity ID", "WBS", "Description", "Amount", "Status"]
    mapping.append(headers)

    header_map = get_header_map(ws, header_row)
    required = ["row type", "activity id", "description", "p/a", "amount"]
    missing = [name for name in required if name not in header_map]
    if missing:
        raise ValueError("Columns required for Amount Mapping were not found: " + ", ".join(missing))

    row_type_col = header_map["row type"]
    activity_id_col = header_map["activity id"]
    description_col = header_map["description"]
    pa_col = header_map["p/a"]
    amount_col = header_map["amount"]
    wbs_col = header_map.get("wbs")

    count = 0
    for row in range(first_data_row, ws.max_row + 1):
        row_type = normalize_header(ws.cell(row, row_type_col).value)
        pa = str(ws.cell(row, pa_col).value or "").strip().upper()
        if row_type != "activity" or pa != "P":
            continue

        activity_id = ws.cell(row, activity_id_col).value
        if activity_id in (None, ""):
            continue

        amount = ws.cell(row, amount_col).value
        mapping.append([
            activity_id,
            ws.cell(row, wbs_col).value if wbs_col else None,
            ws.cell(row, description_col).value,
            amount,
            "Ready" if amount not in (None, "") else "Waiting",
        ])
        count += 1

    for cell in mapping[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = HEADER_BORDER

    for row in range(2, mapping.max_row + 1):
        mapping.cell(row, 4).number_format = '#,##0.00'
        for col in range(1, 6):
            mapping.cell(row, col).border = HEADER_BORDER
            mapping.cell(row, col).alignment = Alignment(vertical="center")

    mapping.column_dimensions["A"].width = 16
    mapping.column_dimensions["B"].width = 14
    mapping.column_dimensions["C"].width = 52
    mapping.column_dimensions["D"].width = 18
    mapping.column_dimensions["E"].width = 14
    mapping.freeze_panes = "A2"
    mapping.auto_filter.ref = f"A1:E{mapping.max_row}"

    mapping.sheet_properties.pageSetUpPr.fitToPage = True
    mapping.page_setup.fitToWidth = 1
    mapping.page_setup.fitToHeight = 0
    return count

def add_weekly_timescale(
    input_file: Path,
    output_file: Path,
    sheet_name: str,
    cutoff_day: int,
    margin_weeks: int,
) -> tuple[date, date, list[date]]:
    wb = load_workbook(input_file)
    ws = find_sheet(wb, sheet_name)

    # Create Amount before the timescale so Script 04 does not insert columns.
    ensure_amount_column(ws)
    header_map = get_header_map(ws, 1)
    earliest_start, latest_finish = collect_plan_range(ws, header_map)
    rollup = collect_project_rollup(ws, header_map)
    project_name = get_project_name(wb, input_file)
    week_dates = generate_week_dates(earliest_start, latest_finish, cutoff_day, margin_weeks)

    original_max_col = ws.max_column
    original_max_row = ws.max_row

    # Add three header rows for Year, Month, and Week; move the original header to row 4.
    ws.insert_rows(1, 3)
    first_timescale_col = original_max_col + 1

    # Activity data headers occupy the top three rows and the original column header is on row 4.
    ws.merge_cells(start_row=1, start_column=1, end_row=3, end_column=original_max_col)
    data_header = ws.cell(1, 1)
    data_header.value = "Activity Data"
    data_header.fill = HEADER_FILL
    data_header.font = HEADER_FONT
    data_header.alignment = Alignment(horizontal="center", vertical="center")
    data_header.border = HEADER_BORDER

    for col in range(1, original_max_col + 1):
        cell = ws.cell(4, col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = HEADER_BORDER

    # Week number and week-ending date.
    for index, week_date in enumerate(week_dates, start=1):
        col = first_timescale_col + index - 1
        week_cell = ws.cell(3, col)
        date_cell = ws.cell(4, col)
        week_cell.value = f"W{index}"
        date_cell.value = week_date
        week_cell.fill = WEEK_FILL
        week_cell.font = Font(bold=True, color="000000")
        date_cell.fill = DATE_FILL
        date_cell.font = Font(bold=True, color="000000")
        for cell in (week_cell, date_cell):
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = HEADER_BORDER
        date_cell.number_format = "dd/mm/yy"
        ws.column_dimensions[get_column_letter(col)].width = 11

    merge_group_headers(ws, 1, first_timescale_col, week_dates, lambda d: d.year, lambda d: str(d.year), YEAR_FILL, Font(color="FFFFFF", bold=True))
    merge_group_headers(ws, 2, first_timescale_col, week_dates, lambda d: (d.year, d.month), lambda d: d.strftime("%B"), MONTH_FILL, Font(color="000000", bold=True))

    # Project Summary appears above all WBS rows.
    insert_project_summary(ws, header_map, project_name, rollup, first_data_row=5)

    # Important: insert_rows does not reliably move RowDimension.
    # Rebuild outlines from the Outline Level column to match actual rows.
    rebuild_outline_levels(ws, header_row=4, first_data_row=5)

    mapping_count = create_amount_mapping_sheet(wb, ws, header_row=4, first_data_row=5)

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 22
    ws.row_dimensions[4].height = 32
    ws.freeze_panes = ws.cell(5, first_timescale_col)

    if ws.auto_filter.ref:
        ws.auto_filter.ref = f"A4:{get_column_letter(original_max_col)}{original_max_row + 5}"

    write_info_sheet(wb, cutoff_day, margin_weeks, earliest_start, latest_finish, week_dates)
    info = wb["Timescale Info"]
    info.append(["Project Summary", project_name])
    info.append(["Header Levels", "Year / Month / Week / Date"])
    info.append(["Amount Input Option 1", "Enter Amount directly in the main worksheet and skip Script 04."])
    info.append(["Amount Input Option 2", "Enter Amount in Amount Mapping and run Script 04."])
    info.append(["Mapping Activities", mapping_count])

    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_file)
    wb.close()
    return earliest_start, latest_finish, week_dates

