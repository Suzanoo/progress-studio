
from __future__ import annotations

from copy import copy
from collections import OrderedDict
from datetime import date, datetime

from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from progress_studio.domain.main_dataset import MainDataset
from progress_studio.domain.monthly_cache import MonthlyCache
from progress_studio.infrastructure.excel.timescale_workbook import (
    DATE_FILL,
    HEADER_BORDER,
    MONTH_FILL,
    WEEK_FILL,
    YEAR_FILL,
)
from progress_studio.infrastructure.excel.progress_workbook import (
    add_progress_conditional_formatting,
    clear_progress_conditional_formatting,
    clear_timescale_direct_fills,
    SCURVE_PLAN_FILL,
    SCURVE_ACTUAL_FILL,
    WBS_PLAN_FILL,
    WBS_ACTUAL_FILL,
)


def _as_date(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def _is_weekly_display_label(value: object) -> bool:
    text = str(value or "").strip().upper()
    return text == "X" or (len(text) > 1 and text.startswith("W") and text[1:].isdigit())


def _display_month_buckets(source) -> list[tuple[tuple[int, int], list[int]]]:
    """Return physical monthly buckets from the copied weekly display timescale.

    Live Rebuild's MainDataset intentionally contains reporting Wn periods only.
    The worksheet still owns display-only X margin columns, so the live monthly
    writer must discover the physical display range from the worksheet rather
    than treating dataset.periods[0] as the left edge.
    """

    grouped: OrderedDict[tuple[int, int], list[int]] = OrderedDict()
    for col in range(1, source.max_column + 1):
        label = source.cell(3, col).value
        reporting_date = _as_date(source.cell(4, col).value)
        if reporting_date is None or not _is_weekly_display_label(label):
            continue
        grouped.setdefault((reporting_date.year, reporting_date.month), []).append(col)
    return list(grouped.items())



def build_live_monthly_view(
    workbook,
    dataset: MainDataset,
    cache: MonthlyCache,
    *,
    source_sheet: str = "main",
    target_sheet: str = "main_monthly",
) -> int:
    """LW-10.0 Full Live Monthly baseline.

    Every monthly timescale cell links directly to the corresponding weekly range
    in `main`. This intentionally maximizes live behavior before later LW-10.x
    milestones cut formula complexity/volume based on measured pain points.
    """
    if source_sheet not in workbook.sheetnames:
        raise ValueError(f"Monthly source worksheet was not found: {source_sheet}")
    if not dataset.periods:
        raise ValueError("Weekly periods were not found in MainDataset.")

    source = workbook[source_sheet]
    if target_sheet in workbook.sheetnames:
        del workbook[target_sheet]

    monthly = workbook.copy_worksheet(source)
    monthly.title = target_sheet
    source_index = workbook.worksheets.index(source)
    workbook._sheets.remove(monthly)
    workbook._sheets.insert(source_index + 1, monthly)
    monthly.cell(1, 1).value = "Activity Data — Monthly View"

    display_buckets = _display_month_buckets(source)
    if not display_buckets:
        raise ValueError("Weekly display timescale was not found on the main worksheet.")

    first_timescale_col = display_buckets[0][1][0]
    reporting_by_month = {
        (period.reporting_date.year, period.reporting_date.month): period
        for period in cache.periods
        if period.reporting_date is not None
    }

    for merged in list(monthly.merged_cells.ranges):
        if merged.max_col >= first_timescale_col:
            monthly.unmerge_cells(str(merged))

    reporting_template_col = dataset.periods[0].column
    template_styles = {
        row: copy(monthly.cell(row, reporting_template_col)._style)
        for row in range(5, monthly.max_row + 1)
    }
    template_formats = {
        row: monthly.cell(row, reporting_template_col).number_format
        for row in range(5, monthly.max_row + 1)
    }

    monthly.delete_cols(
        first_timescale_col,
        monthly.max_column - first_timescale_col + 1,
    )

    # Year/month/date grammar.  One physical column per calendar month.
    # Reporting months use the cache's M1..Mn identity/source columns; months
    # containing display-only X weeks only remain a single X monthly margin.
    rendered_periods: list[tuple[object | None, list[int]]] = []
    for index, ((year, month), weekly_cols) in enumerate(display_buckets, start=1):
        col = first_timescale_col + index - 1
        reporting_period = reporting_by_month.get((year, month))
        is_reporting = reporting_period is not None
        reporting = (
            reporting_period.reporting_date
            if reporting_period is not None
            else _as_date(source.cell(4, weekly_cols[-1]).value)
        )
        source_columns = (
            list(reporting_period.source_columns)
            if reporting_period is not None
            else list(weekly_cols)
        )
        rendered_periods.append((reporting_period, source_columns))

        monthly.cell(2, col).value = date(year, month, 1).strftime("%B")
        monthly.cell(2, col).fill = MONTH_FILL
        monthly.cell(2, col).font = Font(color="000000", bold=True)
        monthly.cell(2, col).alignment = Alignment(horizontal="center", vertical="center")
        monthly.cell(2, col).border = HEADER_BORDER

        monthly.cell(3, col).value = reporting_period.key if is_reporting else "X"
        monthly.cell(3, col).fill = WEEK_FILL
        monthly.cell(3, col).font = Font(color="000000", bold=True)
        monthly.cell(3, col).alignment = Alignment(horizontal="center", vertical="center")
        monthly.cell(3, col).border = HEADER_BORDER

        monthly.cell(4, col).value = reporting
        monthly.cell(4, col).fill = DATE_FILL
        monthly.cell(4, col).font = Font(color="000000", bold=True)
        monthly.cell(4, col).alignment = Alignment(horizontal="center", vertical="center")
        monthly.cell(4, col).border = HEADER_BORDER
        monthly.cell(4, col).number_format = "dd/mm/yy"
        monthly.column_dimensions[get_column_letter(col)].width = 12

        monthly.cell(1, col).value = str(year)
        monthly.cell(1, col).fill = YEAR_FILL
        monthly.cell(1, col).font = Font(color="FFFFFF", bold=True)
        monthly.cell(1, col).alignment = Alignment(horizontal="center", vertical="center")
        monthly.cell(1, col).border = HEADER_BORDER

    source_ref = "'" + source_sheet.replace("'", "''") + "'"
    by_source_row = {row.source_row: row for row in cache.rows}
    for source_row, cached in by_source_row.items():
        if source_row > monthly.max_row:
            continue
        for index, (period, source_columns) in enumerate(rendered_periods, start=0):
            col = first_timescale_col + index
            cell = monthly.cell(source_row, col)
            if source_row in template_styles:
                cell._style = copy(template_styles[source_row])
                cell.number_format = template_formats[source_row]

            # X-only months are display canvas.  They stay physically present in
            # main_monthly but never acquire progress formulas or reporting data.
            if period is None:
                cell.value = ""
                continue

            first_week = get_column_letter(source_columns[0])
            last_week = get_column_letter(source_columns[-1])
            source_range = (
                f"{source_ref}!{first_week}{source_row}:{last_week}{source_row}"
            )

            row_type = cached.row_type.strip().lower()
            pa = cached.pa.strip().upper()

            if row_type == "s-curve":
                # LW-11.2: main_monthly remains an Activity Data view only.
                # It has no Dashboard/cutoff ownership. Chart monthly data is
                # adapted from the authoritative `progress` contract instead.
                if pa in {"AP", "AA"}:
                    cell.value = f"={source_ref}!{last_week}{source_row}"
                else:
                    cell.value = f'=IF(COUNT({source_range})=0,"",SUM({source_range}))'
            else:
                # Full-live baseline intentionally uses the same straightforward
                # formula for Project/WBS/Activity Plan and Actual rows.
                cell.value = f'=IF(COUNT({source_range})=0,"",SUM({source_range}))'

    # Match main exactly: blank timescale cells have no fill; populated cells
    # are colored by the same Project/WBS/Activity Plan/Actual CF rules.
    monthly_timescale_cols = list(
        range(first_timescale_col, first_timescale_col + len(display_buckets))
    )
    header_columns = {name: col for name, col in dataset.headers}
    required = ("row type", "activity id", "p/a", "outline level")
    if monthly_timescale_cols and all(name in header_columns for name in required):
        clear_timescale_direct_fills(
            monthly,
            monthly_timescale_cols,
            monthly.max_row,
        )
        clear_progress_conditional_formatting(monthly, monthly_timescale_cols)
        add_progress_conditional_formatting(
            monthly,
            monthly_timescale_cols,
            header_columns["row type"],
            header_columns["activity id"],
            header_columns["p/a"],
            header_columns["outline level"],
            monthly.max_row,
        )

        # S-Curve summary rows sit outside the Project/WBS/Activity CF grammar.
        # Paint their timescale explicitly with the same palette used by main.
        pa_col = header_columns["p/a"]
        row_type_col = header_columns["row type"]
        scurve_fills = {
            "P": SCURVE_PLAN_FILL,
            "AP": WBS_PLAN_FILL,
            "A": SCURVE_ACTUAL_FILL,
            "AA": WBS_ACTUAL_FILL,
        }
        for row in range(dataset.header_row + 1, monthly.max_row + 1):
            row_type = str(monthly.cell(row, row_type_col).value or "").strip().lower()
            if row_type != "s-curve":
                continue
            pa = str(monthly.cell(row, pa_col).value or "").strip().upper()
            fill = scurve_fills.get(pa)
            if fill is None:
                continue
            for col in monthly_timescale_cols:
                monthly.cell(row, col).fill = copy(fill)

    try:
        monthly.data_validations.dataValidation = []
    except AttributeError:
        pass
    monthly.freeze_panes = monthly.cell(5, first_timescale_col)
    monthly.sheet_properties.outlinePr.summaryBelow = False
    monthly.sheet_properties.outlinePr.applyStyles = True
    monthly.sheet_view.showGridLines = source.sheet_view.showGridLines
    return len(display_buckets)
