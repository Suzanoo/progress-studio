from __future__ import annotations

from datetime import date, datetime, timedelta

from openpyxl.chart import LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.error_bar import ErrorBars
from openpyxl.chart.text import RichText
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.drawing.text import CharacterProperties, Paragraph, ParagraphProperties, RichTextProperties
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

from progress_studio.domain.main_dataset import MainDataset
from progress_studio.infrastructure.excel.dashboard_workbook import BLUE, GREEN, DATA_SHEET

WEEKLY_OVERLAY_NAME = "PS_CURVE_OVERLAY_WEEKLY"
MONTHLY_OVERLAY_NAME = "PS_CURVE_OVERLAY_MONTHLY"
OVERLAY_TOP_ROW = 5
CUTOFF_LABEL_FILL = "1F4E78"
CUTOFF_VALUE_FILL = "D9EAF7"
OVERLAY_MARKER_SIZE = 7
OVERLAY_LABEL_FORMAT = "0.0%"
OVERLAY_LABEL_FONT_SIZE = 700  # DrawingML uses 1/100 pt -> 7 pt.
CUTOFF_LABEL_FONT_SIZE = 1000  # 10 pt: cutoff must remain legible over the schedule.
PLAN_LABEL_TEXT = "1F4E79"
PLAN_LABEL_FILL = "DDEBF7"
PLAN_LABEL_BORDER = "9CC2E5"
ACTUAL_LABEL_TEXT = "385723"
ACTUAL_LABEL_FILL = "E2F0D9"
ACTUAL_LABEL_BORDER = "A9D18E"
CUTOFF_RED = "C00000"
CUTOFF_LABEL_BG = "FCE4D6"
CUTOFF_LABEL_BORDER = "C00000"


def ensure_overlay_visible_actual_columns(
    workbook,
    *,
    weekly_cutoff_ref: str,
    monthly_cutoff_ref: str,
) -> None:
    """Build independent cutoff-aware Actual helpers and cutoff-line helpers.

    Dashboard, main, and main_monthly deliberately own separate cutoff state.
    The traditional overlays never read Dashboard!K5 after their initial values
    are seeded; each view masks Actual and renders its red cutoff line from its
    own local cutoff cell.
    """
    if DATA_SHEET not in workbook.sheetnames:
        raise ValueError("Dashboard_Data is required before building traditional overlays.")
    ws = workbook[DATA_SHEET]
    ws["P1"] = "Weekly Actual Visible"
    ws["Q1"] = "Monthly Actual Visible"
    ws["R1"] = "Weekly Cutoff Line"
    ws["S1"] = "Monthly Cutoff Line"
    for row in range(2, ws.max_row + 1):
        next_row = row + 1
        # Actual is cumulative. Keep a visible 0% baseline before the first
        # reported Actual, carry the latest cumulative value through blank source
        # cells, then mask everything after this sheet's cutoff.
        ws.cell(
            row, 16,
            f'=IF(A{row}="",NA(),IF(A{row}>{weekly_cutoff_ref},NA(),'
            f'IF(C{row}<>"",C{row},IFERROR(LOOKUP(2,1/($C$2:C{row}<>""),$C$2:C{row}),0))))'
        )
        ws.cell(
            row, 17,
            f'=IF(D{row}="",NA(),IF(D{row}>{monthly_cutoff_ref},NA(),'
            f'IF(F{row}<>"",F{row},IFERROR(LOOKUP(2,1/($F$2:F{row}<>""),$F$2:F{row}),0))))'
        )
        # A single 1.0 point plus a full-height negative Y error bar produces
        # a vertical cutoff line without VBA or movable worksheet shapes.
        ws.cell(row, 18, f'=IF(A{row}="",NA(),IF(AND(A{row}<={weekly_cutoff_ref},OR(A{next_row}="",A{next_row}>{weekly_cutoff_ref})),1,NA()))')
        ws.cell(
            row, 19,
            f'=IF(D{row}="",NA(),IF(AND(D{row}<={monthly_cutoff_ref},OR(D{next_row}="",D{next_row}>{monthly_cutoff_ref})),1,NA()))'
        )
        ws.cell(row, 16).number_format = "0.00%"
        ws.cell(row, 17).number_format = "0.00%"
        ws.cell(row, 18).number_format = "0%"
        ws.cell(row, 19).number_format = "0%"



def _as_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None

def _project_dates(dataset: MainDataset):
    """Return the authoritative project window as date-only values.

    Prefer the Project Summary plan row when it exists. Activity timestamps often
    carry an 08:00 time component while reporting cutoffs are midnight. Comparing
    those raw datetimes shifts a same-day project start into the following period.
    """
    summary_rows = tuple(
        row for row in dataset.rows
        if row.row_type.strip().lower() == "project summary" and row.pa.strip().upper() == "P"
    )
    rows = summary_rows or dataset.activities or tuple(
        row for row in dataset.rows if row.plan_start is not None or row.plan_finish is not None
    )
    starts = [_as_date(row.plan_start) for row in rows if row.plan_start is not None]
    finishes = [_as_date(row.plan_finish) for row in rows if row.plan_finish is not None]
    starts = [value for value in starts if value is not None]
    finishes = [value for value in finishes if value is not None]
    return (min(starts) if starts else None, max(finishes) if finishes else None)


def _weekly_project_window(data_ws, dataset: MainDataset) -> tuple[int, int, int, int]:
    """Return project-only helper rows and their physical weekly columns.

    ``Dashboard_Data`` already owns the authoritative reporting/calculation
    window.  Use its nonblank Plan rows as the source of truth instead of
    recomputing inclusion from Project Start/Finish.  This preserves the final
    reporting week that overlaps Project Finish while excluding display-only
    margins on both sides.
    """
    def _weekly_date(row: int):
        # Dashboard_Data!A can be a lightweight formula link to ``progress`` in
        # rebuilt workbooks.  Column J is the literal weekly reporting/cutoff
        # list owned by the Dashboard contract, so prefer it whenever present.
        # This keeps overlay geometry independent from Excel formula caches.
        return _as_date(data_ws.cell(row, 10).value) or _as_date(data_ws.cell(row, 1).value)

    helper_rows = [
        row for row in range(2, data_ws.max_row + 1)
        if _weekly_date(row) is not None
        and data_ws.cell(row, 2).value not in (None, "")
    ]
    if not helper_rows:
        first_col = dataset.periods[0].column if dataset.periods else 1
        last_col = dataset.periods[-1].column if dataset.periods else first_col
        return 2, max(2, data_ws.max_row), first_col, last_col

    first_helper, last_helper = helper_rows[0], helper_rows[-1]
    first_date = _weekly_date(first_helper)
    last_date = _weekly_date(last_helper)

    date_to_col = {
        _as_date(period.reporting_date): period.column
        for period in dataset.periods
        if _as_date(period.reporting_date) is not None
    }
    first_col = date_to_col.get(first_date)
    last_col = date_to_col.get(last_date)
    if first_col is None or last_col is None:
        # Fall back to the closest physical reporting columns without changing
        # the helper source window.
        period_dates = [
            (_as_date(period.reporting_date), period.column)
            for period in dataset.periods
            if _as_date(period.reporting_date) is not None
        ]
        if first_col is None and first_date is not None:
            first_col = next((col for value, col in period_dates if value >= first_date), period_dates[0][1])
        if last_col is None and last_date is not None:
            last_col = next((col for value, col in period_dates if value >= last_date), period_dates[-1][1])
    return first_helper, last_helper, int(first_col), int(last_col)


def _monthly_project_window(data_ws, dataset: MainDataset, monthly_ws=None) -> tuple[int, int, int, int]:
    """Return project-only monthly helper rows and physical month columns.

    Monthly reporting can legitimately finish in the month *after* the raw
    Project Finish date because the final weekly reporting period overlaps the
    finish.  Therefore the last nonblank Monthly Plan helper, not the calendar
    month of Project Finish, owns the chart boundary.
    """
    helper_rows = [
        row for row in range(2, data_ws.max_row + 1)
        if _as_date(data_ws.cell(row, 4).value) is not None
        and data_ws.cell(row, 5).value not in (None, "")
    ]
    if not helper_rows:
        fallback_col = dataset.periods[0].column if dataset.periods else 1
        return 2, 2, fallback_col, fallback_col

    first_helper, last_helper = helper_rows[0], helper_rows[-1]
    first_date = _as_date(data_ws.cell(first_helper, 4).value)
    last_date = _as_date(data_ws.cell(last_helper, 4).value)
    first_key = (first_date.year, first_date.month)
    last_key = (last_date.year, last_date.month)

    # RN-3 introduces X display-margin weeks.  ``dataset.periods`` now contains
    # reporting Wn periods only, so its first physical weekly column is no
    # longer the left edge of ``main_monthly``.  Monthly chart geometry must be
    # mapped against the monthly worksheet itself; otherwise the anchor shifts
    # right by the number of pre-project X columns/months.
    if monthly_ws is not None:
        month_to_col: dict[tuple[int, int], int] = {}
        for col in range(1, monthly_ws.max_column + 1):
            value = _as_date(monthly_ws.cell(4, col).value)
            if value is None:
                continue
            month_to_col.setdefault((value.year, value.month), col)
        first_col = month_to_col.get(first_key)
        last_col = month_to_col.get(last_key)
        if first_col is not None and last_col is not None:
            return first_helper, last_helper, first_col, max(first_col, last_col)

    # Compatibility fallback for callers/tests that do not provide the monthly
    # worksheet.  This preserves the pre-RN-3 behavior but is not authoritative
    # for X-margin geometry.
    first_timescale_col = dataset.periods[0].column if dataset.periods else 1
    if not dataset.periods:
        return first_helper, last_helper, first_timescale_col, first_timescale_col

    full_month_keys: list[tuple[int, int]] = []
    for period in dataset.periods:
        value = _as_date(period.reporting_date)
        if value is None:
            continue
        key = (value.year, value.month)
        if not full_month_keys or full_month_keys[-1] != key:
            full_month_keys.append(key)

    first_pos = next((i for i, key in enumerate(full_month_keys) if key == first_key), 0)
    last_pos = next((i for i, key in enumerate(full_month_keys) if key == last_key), len(full_month_keys) - 1)
    if last_pos < first_pos:
        last_pos = first_pos
    return first_helper, last_helper, first_timescale_col + first_pos, first_timescale_col + last_pos


def _build_explicit_overlay_series_sources(
    data_ws,
    *,
    weekly_first: int,
    weekly_last: int,
    monthly_first: int,
    monthly_last: int,
) -> tuple[tuple[int, int, int, int, int, int], tuple[int, int, int, int, int, int]]:
    """Build small chart-only helper series with an explicit leading (0, 0).

    Physical chart geometry uses N schedule cells, while a right-edge marker
    layout needs N+1 data points. Dashboard_Data's project-only source can begin
    at row 2, so relying on a pre-project source row is not robust. These helper
    columns always create one synthetic zero point followed by live formulas back
    to the canonical Plan/Actual/Cutoff helpers. No business data is duplicated.
    """
    # Weekly: T:W. Monthly: X:AA. These columns are internal Dashboard_Data only.
    weekly_cols = (20, 21, 22, 23)
    monthly_cols = (24, 25, 26, 27)
    headers = (
        (20, "Weekly Overlay Date"),
        (21, "Weekly Overlay Plan"),
        (22, "Weekly Overlay Actual"),
        (23, "Weekly Overlay Cutoff"),
        (24, "Monthly Overlay Date"),
        (25, "Monthly Overlay Plan"),
        (26, "Monthly Overlay Actual"),
        (27, "Monthly Overlay Cutoff"),
    )
    for col, label in headers:
        data_ws.cell(1, col, label)

    # Clear stale helper values from a previous rebuild before writing the new set.
    clear_to = max(data_ws.max_row, 3 + (weekly_last - weekly_first + 1), 3 + (monthly_last - monthly_first + 1))
    for row in range(2, clear_to + 1):
        for col in range(20, 28):
            data_ws.cell(row, col).value = None

    # In Live/Snapshot rebuilds column A may be a formula (``=progress!A...``),
    # while column J intentionally stores the same reporting dates as literal
    # values for validation.  Use that literal date for the synthetic (0, 0)
    # anchor so the first point always exists before the first real period.
    weekly_first_date = (
        _as_date(data_ws.cell(weekly_first, 10).value)
        or _as_date(data_ws.cell(weekly_first, 1).value)
    )
    weekly_anchor_date = (weekly_first_date - timedelta(days=7)) if weekly_first_date else None
    data_ws.cell(2, 20, weekly_anchor_date)
    data_ws.cell(2, 21, 0)
    data_ws.cell(2, 22, 0)
    data_ws.cell(2, 23, "=NA()")
    data_ws.cell(2, 21).number_format = data_ws.cell(2, 22).number_format = "0.00%"
    weekly_row = 3
    for source_row in range(weekly_first, weekly_last + 1):
        weekly_date = (
            _as_date(data_ws.cell(source_row, 10).value)
            or data_ws.cell(source_row, 1).value
        )
        data_ws.cell(weekly_row, 20, weekly_date)
        data_ws.cell(weekly_row, 20).number_format = "dd/mm/yyyy"
        data_ws.cell(weekly_row, 21, f"=B{source_row}")
        data_ws.cell(weekly_row, 22, f"=P{source_row}")
        data_ws.cell(weekly_row, 23, f"=R{source_row}")
        for col in (21, 22, 23):
            data_ws.cell(weekly_row, col).number_format = "0.00%"
        weekly_row += 1
    weekly_bounds = (2, weekly_row - 1, *weekly_cols)

    monthly_first_date = _as_date(data_ws.cell(monthly_first, 4).value)
    monthly_anchor_date = None
    if monthly_first_date:
        monthly_anchor_date = monthly_first_date.replace(day=1) - timedelta(days=1)
    data_ws.cell(2, 24, monthly_anchor_date)
    data_ws.cell(2, 24).number_format = "mmmm yyyy"
    data_ws.cell(2, 25, 0)
    data_ws.cell(2, 26, 0)
    data_ws.cell(2, 27, "=NA()")
    data_ws.cell(2, 25).number_format = data_ws.cell(2, 26).number_format = "0.00%"
    monthly_row = 3
    for source_row in range(monthly_first, monthly_last + 1):
        data_ws.cell(monthly_row, 24, data_ws.cell(source_row, 4).value)
        data_ws.cell(monthly_row, 24).number_format = "mmmm yyyy"
        data_ws.cell(monthly_row, 25, f"=E{source_row}")
        data_ws.cell(monthly_row, 26, f"=Q{source_row}")
        data_ws.cell(monthly_row, 27, f"=S{source_row}")
        for col in (25, 26, 27):
            data_ws.cell(monthly_row, col).number_format = "0.00%"
        monthly_row += 1
    monthly_bounds = (2, monthly_row - 1, *monthly_cols)
    return weekly_bounds, monthly_bounds

def _scurve_plan_row(dataset: MainDataset) -> int:
    candidates = [
        row.row_number for row in dataset.rows
        if row.row_type.strip().lower() == "s-curve"
        and row.description.strip().lower() == "plan"
    ]
    if candidates:
        return min(candidates)
    any_scurve = [row.row_number for row in dataset.rows if row.row_type.strip().lower() == "s-curve"]
    if any_scurve:
        return min(any_scurve)
    return max((row.row_number for row in dataset.rows), default=dataset.header_row + 2) + 1


def _remove_data_validations_for_cell(ws, coordinate: str) -> None:
    """Remove stale single-cell validations that target ``coordinate``.

    Rebuild can run on workbooks created by older LW revisions.  Those files
    may already contain a local cutoff dropdown in the legacy Activity Data
    columns, or an earlier validation on the current M cell.  Keeping both
    produces duplicate cutoff selectors in Excel.
    """
    validations = list(ws.data_validations.dataValidation)
    for validation in validations:
        try:
            targets_cell = coordinate in validation.cells
        except TypeError:
            targets_cell = False
        if targets_cell:
            ws.data_validations.dataValidation.remove(validation)


def _clear_legacy_cutoff_control(ws, dataset: MainDataset, *, row: int) -> None:
    """Remove the pre-LW-12.4.1 cutoff selector from its old columns."""
    legacy_label_col = dataset.header_column("description") or 3
    legacy_value_col = dataset.header_column("amount") or max(4, legacy_label_col + 1)
    if (legacy_label_col, legacy_value_col) == (12, 13):
        return

    legacy_label = ws.cell(row, legacy_label_col)
    if str(legacy_label.value or "").strip() != "Cutoff Date":
        return

    legacy_value = ws.cell(row, legacy_value_col)
    _remove_data_validations_for_cell(ws, legacy_value.coordinate)
    for cell in (legacy_label, legacy_value):
        cell.value = None
        cell.comment = None
        cell.fill = PatternFill(fill_type=None)
        cell.font = Font()
        cell.alignment = Alignment()
        cell.number_format = "General"


def _add_cutoff_control(
    ws,
    dataset: MainDataset,
    *,
    row: int,
    list_col: str,
    list_last_row: int,
    initial_value,
    display_format: str,
) -> str:
    """Add a compact independent cutoff selector in columns L:M.

    Column M is intentionally the editable value cell on both traditional
    views.  Keeping the control out of the timescale footprint makes it easy
    to reach even when the transparent overlay is visible.
    """
    label_col = 12  # L
    value_col = 13  # M
    _clear_legacy_cutoff_control(ws, dataset, row=row)
    label = ws.cell(row, label_col)
    value = ws.cell(row, value_col)
    _remove_data_validations_for_cell(ws, value.coordinate)
    label.value = "Cutoff Date"
    label.fill = PatternFill("solid", fgColor=CUTOFF_LABEL_FILL)
    label.font = Font(color="FFFFFF", bold=True)
    label.alignment = Alignment(horizontal="right", vertical="center")
    value.value = initial_value
    value.number_format = display_format
    value.fill = PatternFill("solid", fgColor=CUTOFF_VALUE_FILL)
    value.font = Font(color="1F1F1F", bold=True)
    value.alignment = Alignment(horizontal="center", vertical="center")
    value.comment = Comment(
        "This cutoff belongs to this sheet only. Press F9 or Save after changing it to recalculate the overlay.",
        "Progress Studio",
    )
    validation = DataValidation(
        type="list",
        formula1=f'=INDIRECT("{DATA_SHEET}!${list_col}$2:${list_col}${max(2, list_last_row)}")',
        allow_blank=False,
    )
    validation.promptTitle = "Cutoff Date"
    validation.prompt = "Choose this sheet's reporting cutoff date."
    validation.errorTitle = "Invalid Cutoff"
    validation.error = "Select a date from the list."
    validation.errorStyle = "stop"
    validation.showInputMessage = True
    validation.showErrorMessage = True
    ws.add_data_validation(validation)
    validation.add(value)
    return f"'{ws.title}'!${get_column_letter(value_col)}${row}"


def _cutoff_proxy_ref(workbook, source_ref: str, *, column: int, label: str) -> str:
    """Return a lightweight workbook-name proxy for a local cutoff control.

    Dashboard_Data must not build a large direct dependency on ``main``. A
    workbook defined name keeps the helper formulas source-neutral without adding
    extra live formulas to the tiny ``progress`` adapter. ``column`` is retained
    in the signature for backward test/readability compatibility.
    """
    del column
    name = "PS_WEEKLY_OVERLAY_CUTOFF" if "weekly" in label.lower() else "PS_MONTHLY_OVERLAY_CUTOFF"
    if name in workbook.defined_names:
        del workbook.defined_names[name]
    workbook.defined_names.add(DefinedName(name, attr_text=source_ref))
    return name


def _cutoff_label_text_properties() -> RichText:
    """Use a larger bold red font for the cutoff tag than curve values."""
    run = CharacterProperties(sz=CUTOFF_LABEL_FONT_SIZE, b=True, solidFill=CUTOFF_RED)
    paragraph = Paragraph(pPr=ParagraphProperties(defRPr=run))
    return RichText(bodyPr=RichTextProperties(), p=[paragraph])


def _label_text_properties(text_color: str) -> RichText:
    """Compact 7 pt series-tinted label text for a dense traditional overlay."""
    run = CharacterProperties(sz=OVERLAY_LABEL_FONT_SIZE, solidFill=text_color)
    paragraph = Paragraph(pPr=ParagraphProperties(defRPr=run))
    return RichText(bodyPr=RichTextProperties(), p=[paragraph])


def _label_graphical_properties(fill_color: str, border_color: str) -> GraphicalProperties:
    """Pale opaque tag background so values remain readable over schedule bars."""
    props = GraphicalProperties(solidFill=fill_color)
    props.line.solidFill = border_color
    props.line.width = 6350  # ~0.5 pt; enough separation without a heavy box.
    return props


def _overlay_chart(
    *,
    data_ws,
    date_col: int,
    plan_col: int,
    actual_col: int,
    cutoff_col: int | None = None,
    first_row: int,
    last_row: int,
    cutoff_label_format: str | None = None,
) -> LineChart:
    chart = LineChart()
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 1
    chart.y_axis.majorUnit = 0.25
    chart.y_axis.numFmt = "0%"
    chart.y_axis.title = None
    chart.x_axis.title = None
    chart.x_axis.tickLblPos = "none"
    chart.legend = None
    # Period-end geometry uses the existing reporting point immediately before
    # the project window as a chart-only start anchor.  Plan may be blank in
    # that pre-project cell, so Excel must render that one blank as zero.
    # Cutoff-masked Actual remains #N/A outside its visible window and therefore
    # still renders as a gap.
    chart.display_blanks = "gap"
    chart.y_axis.majorGridlines = None

    # Both chart and plot areas must be transparent so the underlying
    # timescale bars remain visible through the traditional overlay.
    chart.graphical_properties = GraphicalProperties(
        noFill=True, ln=LineProperties(noFill=True)
    )
    chart.plot_area.graphicalProperties = GraphicalProperties(
        noFill=True, ln=LineProperties(noFill=True)
    )

    plan = Reference(data_ws, min_col=plan_col, max_col=plan_col, min_row=first_row, max_row=last_row)
    actual = Reference(data_ws, min_col=actual_col, max_col=actual_col, min_row=first_row, max_row=last_row)
    cats = Reference(data_ws, min_col=date_col, min_row=first_row, max_row=last_row)
    chart.add_data(plan, titles_from_data=False)
    chart.add_data(actual, titles_from_data=False)
    chart.set_categories(cats)
    chart.series[0].tx = SeriesLabel(v="Plan")
    chart.series[1].tx = SeriesLabel(v="Actual")

    label_styles = (
        (BLUE, "t", PLAN_LABEL_TEXT, PLAN_LABEL_FILL, PLAN_LABEL_BORDER),
        (GREEN, "b", ACTUAL_LABEL_TEXT, ACTUAL_LABEL_FILL, ACTUAL_LABEL_BORDER),
    )
    for series, (color, position, text_color, fill_color, border_color) in zip(
        chart.series[:2], label_styles
    ):
        series.graphicalProperties.line.solidFill = color
        series.graphicalProperties.line.width = 19050
        series.marker.symbol = "circle"
        series.marker.size = OVERLAY_MARKER_SIZE
        series.marker.graphicalProperties.solidFill = color
        series.marker.graphicalProperties.line.solidFill = color
        # LW-12.3.2: keep every value, but turn labels into compact tinted tags.
        # Plan sits above its curve; Actual sits below so nearby series do not
        # compete for the same vertical space.  Pale backgrounds preserve
        # readability while keeping the schedule bars visible around the tags.
        series.dLbls = DataLabelList(
            showVal=True,
            showCatName=False,
            showSerName=False,
            showLegendKey=False,
            numFmt=OVERLAY_LABEL_FORMAT,
            dLblPos=position,
            spPr=_label_graphical_properties(fill_color, border_color),
            txPr=_label_text_properties(text_color),
        )

    if cutoff_col is not None:
        cutoff = Reference(data_ws, min_col=cutoff_col, max_col=cutoff_col, min_row=first_row, max_row=last_row)
        chart.add_data(cutoff, titles_from_data=False)
        # The cutoff series is appended after the initial category assignment.
        # Re-apply the same categories so its category-name data label has a
        # real date/category reference without introducing helper data.
        chart.set_categories(cats)
        cutoff_series = chart.series[2]
        cutoff_series.tx = SeriesLabel(v="Cutoff")
        cutoff_series.graphicalProperties.line.noFill = True
        cutoff_series.marker.symbol = "none"
        cutoff_series.errBars = ErrorBars(
            errDir="y",
            errBarType="minus",
            errValType="fixedVal",
            noEndCap=True,
            val=1,
            spPr=GraphicalProperties(
                ln=LineProperties(solidFill=CUTOFF_RED, w=19050, prstDash="dash")
            ),
        )
        # Only the single non-#N/A cutoff point receives a label.  Category
        # name contributes the reporting date, so the visible tag reads
        # "Cutoff <date>" without worksheet shapes or VBA.
        cutoff_series.dLbls = DataLabelList(
            showVal=False,
            showCatName=True,
            showSerName=True,
            showLegendKey=False,
            dLblPos="t",
            separator=" ",
            numFmt=cutoff_label_format,
            spPr=_label_graphical_properties(CUTOFF_LABEL_BG, CUTOFF_LABEL_BORDER),
            txPr=_cutoff_label_text_properties(),
        )
    return chart


def _responsive_anchor(*, first_col: int, last_col: int, top_row: int, bottom_row: int) -> TwoCellAnchor:
    """Create an Excel 'Move and size with cells' anchor over the schedule grid.

    AnchorMarker uses zero-based row/column indexes. The right/bottom marker is
    placed one cell beyond the intended range so the chart spans the complete
    Project Start -> Project Finish timescale and schedule height.
    """
    start = AnchorMarker(col=max(0, first_col - 1), row=max(0, top_row - 1))
    end = AnchorMarker(col=max(first_col, last_col), row=max(top_row, bottom_row))
    return TwoCellAnchor(editAs="twoCell", _from=start, to=end)



def reassert_traditional_overlay_transparency(workbook) -> None:
    """Restore transparent chart/plot areas after an openpyxl round-trip.

    openpyxl preserves the outer chart-space noFill when loading an existing
    workbook, but drops ``plotArea/spPr`` from traditional overlays on the next
    save. Payment-only rebuilds intentionally preserve Progress views, so the
    shared final policy reasserts this renderer-owned presentation property
    without rebuilding the chart or its series.
    """
    for sheet_name in ("main", "main_monthly"):
        if sheet_name not in workbook.sheetnames:
            continue
        for chart in workbook[sheet_name]._charts:
            chart.graphical_properties = GraphicalProperties(
                noFill=True, ln=LineProperties(noFill=True)
            )
            chart.plot_area.graphicalProperties = GraphicalProperties(
                noFill=True, ln=LineProperties(noFill=True)
            )

def build_traditional_overlays(workbook, dataset: MainDataset) -> tuple[bool, bool]:
    """LW-12.4: independent-cutoff, project-bounded responsive overlays."""
    if not dataset.periods:
        return False, False
    if DATA_SHEET not in workbook.sheetnames:
        raise ValueError("Dashboard_Data is required before building traditional overlays.")
    data_ws = workbook[DATA_SHEET]

    scurve_plan_row = _scurve_plan_row(dataset)
    control_row = max(dataset.header_row + 1, scurve_plan_row - 1)
    bottom_anchor_row = max(OVERLAY_TOP_ROW + 1, scurve_plan_row - 1)

    weekly_list_last = max(2, len(dataset.periods) + 1)
    monthly_list_last = max(
        2,
        1 + sum(1 for r in range(2, data_ws.max_row + 1) if data_ws.cell(r, 11).value not in (None, "")),
    )

    dashboard_cutoff = workbook["Dashboard"]["K5"].value if "Dashboard" in workbook.sheetnames else None
    weekly_dates = [data_ws.cell(r, 10).value for r in range(2, weekly_list_last + 1) if data_ws.cell(r, 10).value not in (None, "")]
    monthly_dates = [data_ws.cell(r, 11).value for r in range(2, monthly_list_last + 1) if data_ws.cell(r, 11).value not in (None, "")]
    dashboard_day = _as_date(dashboard_cutoff)
    weekly_initial = next(
        (d for d in weekly_dates if _as_date(d) == dashboard_day),
        weekly_dates[-1] if weekly_dates else dashboard_cutoff,
    )
    eligible_monthly = [
        d for d in monthly_dates
        if dashboard_day is not None and _as_date(d) is not None and _as_date(d) <= dashboard_day
    ]
    monthly_initial = eligible_monthly[-1] if eligible_monthly else (monthly_dates[-1] if monthly_dates else dashboard_cutoff)

    weekly_cutoff_ref = None
    monthly_cutoff_ref = None
    if "main" in workbook.sheetnames:
        weekly_cutoff_ref = _add_cutoff_control(
            workbook["main"], dataset, row=control_row, list_col="J", list_last_row=weekly_list_last,
            initial_value=weekly_initial, display_format="dd/mm/yyyy",
        )
    if "main_monthly" in workbook.sheetnames:
        monthly_cutoff_ref = _add_cutoff_control(
            workbook["main_monthly"], dataset, row=control_row, list_col="K", list_last_row=monthly_list_last,
            initial_value=monthly_initial, display_format="mmm yyyy",
        )

    if weekly_cutoff_ref is None:
        weekly_cutoff_ref = "Dashboard!$K$5"
    if monthly_cutoff_ref is None:
        monthly_cutoff_ref = "Dashboard!$K$5"

    weekly_cutoff_ref = _cutoff_proxy_ref(
        workbook, weekly_cutoff_ref, column=6, label="weekly_overlay_cutoff"
    )
    monthly_cutoff_ref = _cutoff_proxy_ref(
        workbook, monthly_cutoff_ref, column=7, label="monthly_overlay_cutoff"
    )
    ensure_overlay_visible_actual_columns(
        workbook,
        weekly_cutoff_ref=weekly_cutoff_ref,
        monthly_cutoff_ref=monthly_cutoff_ref,
    )

    # Remove prior overlay charts when rebuilding the same workbook.
    for sheet_name in ("main", "main_monthly"):
        if sheet_name in workbook.sheetnames:
            workbook[sheet_name]._charts = []

    weekly_first, weekly_last, weekly_first_col, weekly_last_col = _weekly_project_window(data_ws, dataset)
    monthly_first, monthly_last, monthly_first_col, monthly_last_col = _monthly_project_window(
        data_ws, dataset, workbook["main_monthly"] if "main_monthly" in workbook.sheetnames else None
    )

    # Build chart-only sources with an explicit leading (0, 0). This keeps
    # right-edge marker geometry correct even when Dashboard_Data starts exactly
    # at the first project reporting period and has no pre-project source row.
    weekly_series, monthly_series = _build_explicit_overlay_series_sources(
        data_ws,
        weekly_first=weekly_first,
        weekly_last=weekly_last,
        monthly_first=monthly_first,
        monthly_last=monthly_last,
    )
    weekly_chart_first, weekly_chart_last, weekly_date_col, weekly_plan_col, weekly_actual_col, weekly_cutoff_col = weekly_series
    monthly_chart_first, monthly_chart_last, monthly_date_col, monthly_plan_col, monthly_actual_col, monthly_cutoff_col = monthly_series

    weekly_added = False
    if "main" in workbook.sheetnames:
        chart = _overlay_chart(
            data_ws=data_ws,
            date_col=weekly_date_col,
            plan_col=weekly_plan_col,
            actual_col=weekly_actual_col,
            cutoff_col=weekly_cutoff_col,
            first_row=weekly_chart_first,
            last_row=weekly_chart_last,
            cutoff_label_format="dd/mm/yyyy",
        )
        chart.anchor = _responsive_anchor(
            first_col=weekly_first_col,
            last_col=weekly_last_col,
            top_row=OVERLAY_TOP_ROW,
            bottom_row=bottom_anchor_row,
        )
        workbook["main"].add_chart(chart)
        weekly_added = True

    monthly_added = False
    if "main_monthly" in workbook.sheetnames:
        chart = _overlay_chart(
            data_ws=data_ws,
            date_col=monthly_date_col,
            plan_col=monthly_plan_col,
            actual_col=monthly_actual_col,
            cutoff_col=monthly_cutoff_col,
            first_row=monthly_chart_first,
            last_row=monthly_chart_last,
            cutoff_label_format="mmmm yyyy",
        )
        chart.anchor = _responsive_anchor(
            first_col=monthly_first_col,
            last_col=monthly_last_col,
            top_row=OVERLAY_TOP_ROW,
            bottom_row=bottom_anchor_row,
        )
        workbook["main_monthly"].add_chart(chart)
        monthly_added = True

    return weekly_added, monthly_added
