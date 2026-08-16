
from __future__ import annotations

from openpyxl.styles import Alignment, Font, PatternFill, Side, Border

GUIDE_SHEET = "README"

NAVY = "17365D"
BLUE = "D9EAF7"
GREEN = "E2F0D9"
GRAY = "F3F6F8"
WHITE = "FFFFFF"
MUTED = "667085"
BORDER = Border(bottom=Side(style="thin", color="D9E1F2"))


def build_workbook_guide(workbook) -> None:
    """Create a short, user-facing workbook guide as the first worksheet."""
    if GUIDE_SHEET in workbook.sheetnames:
        del workbook[GUIDE_SHEET]
    ws = workbook.create_sheet(GUIDE_SHEET, 0)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A6"

    widths = {"A": 3, "B": 18, "C": 22, "D": 22, "E": 22, "F": 22, "G": 3}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.merge_cells("B2:F2")
    ws["B2"] = "PROGRESS STUDIO — WORKBOOK GUIDE"
    ws["B2"].font = Font(size=18, bold=True, color=NAVY)

    ws.merge_cells("B3:F3")
    ws["B3"] = "Normal progress update:  main → edit Actual → F9 / Save for formulas • Rebuild for generated snapshots"
    ws["B3"].fill = PatternFill("solid", fgColor=GREEN)
    ws["B3"].font = Font(size=11, bold=True, color=NAVY)
    ws["B3"].alignment = Alignment(vertical="center")
    ws.row_dimensions[3].height = 28

    sections = [
        ("1  UPDATE PROGRESS", "Open main. Update the Actual row in the correct weekly columns. "
         "For a normal progress update, do not edit the workbook structure."),
        ("2  RECALCULATE", "After editing, press F9 or save the workbook to recalculate Excel formulas. "
         "Generated snapshots/caches are not rebuilt by Excel; use Progress Studio Rebuild for those."),
        ("3  DASHBOARD", "Open Dashboard. Select Weekly or Monthly, then select the Cutoff Date. "
         "KPI, S-Curve and Activity Progress follow the selected period."),
        ("4  MONTHLY", "main_monthly is linked to weekly data in main. "
         "Do not type progress directly into main_monthly. F9 / Save refreshes formulas; use Rebuild when a generated view must be regenerated."),
        ("5  PAYMENT", "Payment Input contains payment requirements. Payment shows payment lines. "
         "Use Progress Studio → Rebuild Payment when payment configuration changes."),
    ]
    row = 6
    for title, body in sections:
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        ws.cell(row, 2, title)
        ws.cell(row, 2).fill = PatternFill("solid", fgColor=BLUE)
        ws.cell(row, 2).font = Font(bold=True, color=NAVY)
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
        ws.cell(row, 4, body)
        ws.cell(row, 4).alignment = Alignment(wrap_text=True, vertical="top")
        for cell in ws[row][1:6]:
            cell.border = BORDER
        ws.row_dimensions[row].height = 44
        row += 2

    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    ws.cell(row, 2, "EDIT / VIEW / DO NOT EDIT")
    ws.cell(row, 2).fill = PatternFill("solid", fgColor=NAVY)
    ws.cell(row, 2).font = Font(bold=True, color=WHITE)
    row += 1

    rules = [
        ("EDIT", "main → Actual weekly progress"),
        ("VIEW", "Dashboard • main_monthly • Payment"),
        ("DO NOT EDIT", "main_monthly formulas • Dashboard formulas • Payment rendering"),
    ]
    for label, body in rules:
        ws.cell(row, 2, label)
        ws.cell(row, 2).font = Font(bold=True, color=NAVY)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
        ws.cell(row, 3, body)
        ws.cell(row, 3).alignment = Alignment(wrap_text=True)
        for cell in ws[row][1:6]:
            cell.border = BORDER
        row += 1

    row += 1
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    ws.cell(row, 2, "WHEN TO USE PROGRESS STUDIO REBUILD")
    ws.cell(row, 2).fill = PatternFill("solid", fgColor=GRAY)
    ws.cell(row, 2).font = Font(bold=True, color=NAVY)
    row += 1
    ws.merge_cells(start_row=row, start_column=2, end_row=row+1, end_column=6)
    ws.cell(row, 2, "Use Rebuild for structural changes: Add/Delete Activity, change Activity ID, "
         "change WBS structure or project timescale, major Plan/rebaseline changes, "
         "or Payment Input / Payment regeneration.")
    ws.cell(row, 2).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[row].height = 32

    row += 3
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    ws.cell(row, 2, "Performance note: Manual Calculation keeps editing fast. F9 / Save recalculates Excel formulas; Rebuild regenerates Python-owned snapshots.")
    ws.cell(row, 2).font = Font(italic=True, color=MUTED)
    ws.cell(row, 2).alignment = Alignment(wrap_text=True)
    ws.row_dimensions[row].height = 26

    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.print_area = f"B2:F{row}"
