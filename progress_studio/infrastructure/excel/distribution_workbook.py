from __future__ import annotations

import argparse
import sys
from collections import Counter
from datetime import date, datetime, timedelta
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.datetime import from_excel
except ImportError as exc:
    raise SystemExit(
        "openpyxl is not installed. Run: pip install openpyxl"
    ) from exc

from progress_studio.infrastructure.excel.calculation_policy import configure_incremental_excel_recalculation
from progress_studio.infrastructure.excel.styles import normalize_argb
from progress_studio.services.distribution import (
    AutoDecision,
    decide_distribution,
    get_distribution,
    load_rules,
)


SCRIPT_VERSION = "3.0.0-auto-distribution"
DEFAULT_SHEET = "main"
DEFAULT_HEADER_ROW = 4
DEFAULT_WEEK_ROW = 3
PERCENT_FORMAT = "0.00%"


class PlanDistributionError(RuntimeError):
    pass


def normalize(value: object) -> str:
    return " ".join(str(value or "").strip().lower().split())


def parse_date(value: object, epoch) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        try:
            converted = from_excel(value, epoch=epoch)
            return converted.date() if isinstance(converted, datetime) else converted
        except (TypeError, ValueError, OverflowError):
            return None

    text = str(value).strip()
    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            parsed = datetime.strptime(text, fmt).date()
            return parsed.replace(year=parsed.year - 543) if parsed.year > 2400 else parsed
        except ValueError:
            continue
    return None


def get_header_map(ws, header_row: int) -> dict[str, int]:
    headers: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        key = normalize(ws.cell(header_row, col).value)
        if key:
            headers[key] = col
    return headers


def find_column(
    headers: dict[str, int],
    aliases: tuple[str, ...],
    *,
    required: bool = False,
) -> int | None:
    for alias in aliases:
        key = normalize(alias)
        if key in headers:
            return headers[key]
    if required:
        raise PlanDistributionError(
            f"Missing required column. Expected one of: {', '.join(aliases)}"
        )
    return None


def cell_value(ws, row: int, col: int | None) -> object:
    return ws.cell(row, col).value if col else ""


def calculation_week_columns(
    weeks: list[tuple[int, date]],
    schedule_start: date,
    schedule_finish: date,
) -> list[tuple[int, date]]:
    """Return only weekly periods that overlap the real schedule window.

    Timescale margin columns remain visible in the workbook, but are display-only
    and must never participate in Plan distribution/progress calculations.
    """
    result: list[tuple[int, date]] = []
    for index, item in enumerate(weeks):
        col, cutoff = item
        period_start = cutoff - timedelta(days=6)
        if cutoff < schedule_start or period_start > schedule_finish:
            continue
        result.append((col, cutoff))
    return result


def find_week_columns(ws, week_row: int, header_row: int, epoch) -> list[tuple[int, date]]:
    weeks: list[tuple[int, date]] = []
    for col in range(1, ws.max_column + 1):
        label = str(ws.cell(week_row, col).value or "").strip().upper()
        if not (label.startswith("W") and label[1:].isdigit()):
            continue
        cutoff = parse_date(ws.cell(header_row, col).value, epoch)
        if cutoff is not None:
            weeks.append((col, cutoff))
    if not weeks:
        raise PlanDistributionError(
            "Weekly timescale was not found. Expected W1, W2, ... and cutoff dates."
        )
    return weeks


def overlap_days(a_start: date, a_finish: date, p_start: date, p_finish: date) -> int:
    start = max(a_start, p_start)
    finish = min(a_finish, p_finish)
    return 0 if start > finish else (finish - start).days + 1


def active_periods(
    activity_start: date,
    activity_finish: date,
    weeks: list[tuple[int, date]],
) -> list[tuple[int, int]]:
    periods: list[tuple[int, int]] = []
    for index, (col, cutoff) in enumerate(weeks):
        period_start = (
            cutoff - timedelta(days=6)
            if index == 0
            else weeks[index - 1][1] + timedelta(days=1)
        )
        days = overlap_days(activity_start, activity_finish, period_start, cutoff)
        if days > 0:
            periods.append((col, days))
    return periods


def distribute(
    activity_start: date,
    activity_finish: date,
    weeks: list[tuple[int, date]],
    method: str,
) -> dict[int, float]:
    if activity_finish < activity_start:
        raise PlanDistributionError(
            f"Plan Finish {activity_finish} is before Plan Start {activity_start}"
        )

    periods = active_periods(activity_start, activity_finish, weeks)
    if not periods:
        return {}

    curve = get_distribution(method).generator(len(periods))
    duration_factors = [days / 7.0 for _, days in periods]
    combined = [weight * factor for weight, factor in zip(curve, duration_factors)]
    total = sum(combined)

    if total <= 0:
        combined = curve
        total = sum(combined)

    normalized = [value / total for value in combined]
    normalized[-1] = max(0.0, 1.0 - sum(normalized[:-1])) if len(normalized) > 1 else 1.0

    return {col: value for (col, _), value in zip(periods, normalized)}


def clear_plan_cells(ws, row: int, week_columns: list[int]) -> None:
    for col in week_columns:
        ws.cell(row, col).value = None
        ws.cell(row, col).number_format = PERCENT_FORMAT


def create_report_sheet(wb, rows: list[dict[str, object]], mode: str) -> None:
    sheet_name = "Distribution Report"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    title_fill = PatternFill("solid", fgColor=normalize_argb("1F4E78"))
    header_fill = PatternFill("solid", fgColor=normalize_argb("D9EAF7"))
    source_fills = {
        "activity_code": PatternFill("solid", fgColor=normalize_argb("C6EFCE")),
        "wbs": PatternFill("solid", fgColor=normalize_argb("DDEBF7")),
        "activity_name": PatternFill("solid", fgColor=normalize_argb("FFF2CC")),
        "default": PatternFill("solid", fgColor=normalize_argb("F4CCCC")),
        "manual": PatternFill("solid", fgColor=normalize_argb("E2F0D9")),
    }

    ws["A1"] = "PLAN DISTRIBUTION REPORT"
    ws["A1"].font = Font(bold=True, color=normalize_argb("FFFFFF"), size=14)
    ws["A1"].fill = title_fill
    ws.merge_cells("A1:I1")

    ws["A3"] = "Mode"
    ws["B3"] = "AUTO" if mode == "auto" else get_distribution(mode).name
    ws["A4"] = "Activities"
    ws["B4"] = len(rows)

    counts = Counter(str(row["distribution"]) for row in rows)
    matched = sum(1 for row in rows if row["source"] != "default")
    defaults = sum(1 for row in rows if row["source"] == "default")

    summary_row = 3
    for method in ("flat", "front", "back", "bell"):
        ws.cell(summary_row, 4, get_distribution(method).name)
        ws.cell(summary_row, 5, counts.get(method, 0))
        summary_row += 1
    ws["G3"] = "Matched Rules"
    ws["H3"] = matched
    ws["G4"] = "Default Applied"
    ws["H4"] = defaults

    headers = [
        "Activity ID",
        "Activity Name",
        "WBS",
        "Plan Start",
        "Plan Finish",
        "Distribution",
        "Source",
        "Matched Rule",
        "Reason",
    ]
    header_row = 7
    for col, value in enumerate(headers, start=1):
        cell = ws.cell(header_row, col, value)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for index, row in enumerate(rows, start=header_row + 1):
        values = [
            row["activity_id"],
            row["activity_name"],
            row["wbs"],
            row["plan_start"],
            row["plan_finish"],
            get_distribution(str(row["distribution"])).name,
            str(row["source"]).replace("_", " ").title(),
            row["matched_rule"],
            row["reason"],
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(index, col, value)

        source = str(row["source"])
        ws.cell(index, 7).fill = source_fills.get(source, source_fills["manual"])

    ws.freeze_panes = "A8"
    ws.auto_filter.ref = f"A7:I{max(header_row, ws.max_row)}"
    widths = [16, 42, 32, 14, 14, 18, 18, 22, 52]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    for row in ws.iter_rows(min_row=8, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def update_info_sheet(wb, method: str, report_rows: list[dict[str, object]]) -> None:
    ws = wb["Info"] if "Info" in wb.sheetnames else wb.create_sheet("Info")
    start_row = ws.max_row + 2 if ws.max_row else 1
    ws.cell(start_row, 1, "Plan Distribution Mode")
    ws.cell(start_row, 2, "AUTO" if method == "auto" else get_distribution(method).name)
    ws.cell(start_row + 1, 1, "Distribution Report")
    ws.cell(start_row + 1, 2, "See sheet: Distribution Report")
    if method == "auto":
        defaults = sum(1 for row in report_rows if row["source"] == "default")
        ws.cell(start_row + 2, 1, "Auto Default Applied")
        ws.cell(start_row + 2, 2, defaults)


def generate_plan_distribution(
    input_file: Path,
    output_file: Path,
    *,
    method: str,
    rules_file: Path | None,
    sheet_name: str,
    header_row: int,
    week_row: int,
    debug: bool,
) -> tuple[int, int, int, Counter]:
    wb = load_workbook(input_file)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise PlanDistributionError(f"Worksheet not found: {sheet_name}")

    ws = wb[sheet_name]
    headers = get_header_map(ws, header_row)

    row_type_col = find_column(headers, ("Row Type",), required=True)
    pa_col = find_column(headers, ("P/A",), required=True)
    plan_start_col = find_column(headers, ("Plan Start",), required=True)
    plan_finish_col = find_column(headers, ("Plan Finish",), required=True)
    activity_id_col = find_column(headers, ("Activity ID",), required=True)
    activity_name_col = find_column(
        headers,
        ("Activity Name", "Activity Description", "Description", "Task Name"),
    )
    wbs_col = find_column(
        headers,
        ("WBS", "WBS Name", "WBS Path", "WBS-1", "WBS 1"),
    )

    display_weeks = find_week_columns(ws, week_row, header_row, wb.epoch)
    week_columns = [col for col, _ in display_weeks]

    activity_windows: list[tuple[date, date]] = []
    for row in range(header_row + 1, ws.max_row + 1):
        if normalize(cell_value(ws, row, row_type_col)) != "activity":
            continue
        if normalize(cell_value(ws, row, pa_col)) != "p":
            continue
        start = parse_date(cell_value(ws, row, plan_start_col), wb.epoch)
        finish = parse_date(cell_value(ws, row, plan_finish_col), wb.epoch)
        if start is not None and finish is not None:
            activity_windows.append((start, finish))
    if not activity_windows:
        raise PlanDistributionError("No valid activity Plan Start / Plan Finish values were found.")

    schedule_start = min(start for start, _ in activity_windows)
    schedule_finish = max(finish for _, finish in activity_windows)
    weeks = calculation_week_columns(display_weeks, schedule_start, schedule_finish)
    if not weeks:
        raise PlanDistributionError("No calculation weeks overlap the schedule window.")

    rules = load_rules(rules_file) if method == "auto" else None

    generated = 0
    skipped_no_dates = 0
    skipped_outside = 0
    report_rows: list[dict[str, object]] = []
    method_counts: Counter = Counter()

    for row in range(header_row + 1, ws.max_row + 1):
        row_type = normalize(cell_value(ws, row, row_type_col))
        pa_value = normalize(cell_value(ws, row, pa_col))
        activity_id = str(cell_value(ws, row, activity_id_col) or "").strip()

        if row_type != "activity" or pa_value != "p" or not activity_id:
            continue

        clear_plan_cells(ws, row, week_columns)

        plan_start = parse_date(cell_value(ws, row, plan_start_col), wb.epoch)
        plan_finish = parse_date(cell_value(ws, row, plan_finish_col), wb.epoch)
        activity_name = str(cell_value(ws, row, activity_name_col) or "").strip()
        wbs = str(cell_value(ws, row, wbs_col) or "").strip()

        if plan_start is None or plan_finish is None:
            skipped_no_dates += 1
            continue

        if method == "auto":
            decision: AutoDecision = decide_distribution(
                activity_code=activity_id,
                wbs=wbs,
                activity_name=activity_name,
                rules=rules or {},
            )
            selected_method = decision.distribution
            source = decision.source
            matched_rule = decision.matched_rule
            reason = decision.reason
        else:
            selected_method = method
            source = "manual"
            matched_rule = get_distribution(method).name
            reason = "User-selected distribution applied to all activities"

        allocations = distribute(plan_start, plan_finish, weeks, selected_method)
        if not allocations:
            skipped_outside += 1
            continue

        for col, value in allocations.items():
            cell = ws.cell(row, col)
            cell.value = value
            cell.number_format = PERCENT_FORMAT

        generated += 1
        method_counts[selected_method] += 1
        report_rows.append(
            {
                "activity_id": activity_id,
                "activity_name": activity_name,
                "wbs": wbs,
                "plan_start": plan_start,
                "plan_finish": plan_finish,
                "distribution": selected_method,
                "source": source,
                "matched_rule": matched_rule,
                "reason": reason,
            }
        )

        if debug:
            print(
                f"GENERATED: {activity_id} | {selected_method} | "
                f"{source} | {matched_rule or 'DEFAULT'}"
            )

    create_report_sheet(wb, report_rows, method)
    update_info_sheet(wb, method, report_rows)

    configure_incremental_excel_recalculation(wb)

    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_file)
    wb.close()

    return generated, skipped_no_dates, skipped_outside, method_counts


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Generate weekly Plan progress distributions."
    )
    parser.add_argument("--input", type=Path, required=True)
    parser.add_argument("--output", type=Path)
    parser.add_argument(
        "--method",
        default="flat",
        choices=["flat", "front", "back", "bell", "auto"],
    )
    parser.add_argument(
        "--rules",
        type=Path,
        help="Optional custom distribution_rules.json for Auto mode",
    )
    parser.add_argument("--sheet", default=DEFAULT_SHEET)
    parser.add_argument("--header-row", type=int, default=DEFAULT_HEADER_ROW)
    parser.add_argument("--week-row", type=int, default=DEFAULT_WEEK_ROW)
    parser.add_argument("--debug", action="store_true")
    return parser


def main() -> int:
    args = build_parser().parse_args()
    input_file = args.input.expanduser().resolve()
    output_file = (
        args.output.expanduser().resolve()
        if args.output
        else input_file.with_name(
            f"{input_file.stem}_{args.method}_plan{input_file.suffix}"
        )
    )

    try:
        if not input_file.is_file():
            raise PlanDistributionError(f"Input file not found: {input_file}")
        if input_file.suffix.lower() != ".xlsx":
            raise PlanDistributionError("Only .xlsx files are supported")

        generated, no_dates, outside, counts = generate_plan_distribution(
            input_file,
            output_file,
            method=args.method,
            rules_file=args.rules,
            sheet_name=args.sheet,
            header_row=args.header_row,
            week_row=args.week_row,
            debug=args.debug,
        )

        print(f"SCRIPT VERSION          : {SCRIPT_VERSION}")
        print(f"MODE                    : {args.method.upper()}")
        print(f"INPUT                   : {input_file}")
        print(f"OUTPUT                  : {output_file}")
        print(f"PLAN GENERATED          : {generated}")
        print(f"SKIPPED - NO DATES      : {no_dates}")
        print(f"SKIPPED - OUTSIDE RANGE : {outside}")
        for key in ("flat", "front", "back", "bell"):
            print(f"{get_distribution(key).name.upper():24}: {counts.get(key, 0)}")
        print("DONE")
        return 0

    except (PlanDistributionError, KeyError, ValueError, json.JSONDecodeError) as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1
    except PermissionError:
        print(
            "ERROR: Unable to save. Close the Excel workbook and run again.",
            file=sys.stderr,
        )
        return 1
    except Exception as exc:
        print(f"UNEXPECTED ERROR: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
