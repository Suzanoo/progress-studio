from __future__ import annotations

import shutil
import tempfile
from copy import copy
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Border, Side
from PIL import Image, ImageDraw, ImageFont

from progress_studio.config.payment_theme import PaymentLineTheme, load_payment_line_theme
from progress_studio.domain.payment_models import (
    PaymentLineRenderResult,
    PaymentMultiLineRenderResult,
    PaymentResolvedPeriod,
    PaymentResolvedPoint,
)
from progress_studio.infrastructure.excel.payment_workbook import PaymentWorkbookError


class PaymentLineRenderer:
    """Cell-based vertical-backbone Payment renderer.

    Each Payment period gets one vertical backbone at its planned eligible
    position: the latest resolved requirement boundary among that period's sparse
    Activity points. The Payment Date supplied in the input workbook is retained
    only as legacy/reference metadata and never controls backbone placement.
    Line geometry remains cell-border based. A single lightweight, cell-anchored
    badge image is added per Payment only for the visible label.
    """

    MAIN_SHEET = "main"
    PAYMENT_SHEET = "Payment"
    HEADER_ROW = 4

    def __init__(self, theme: PaymentLineTheme | None = None) -> None:
        self.theme = theme or load_payment_line_theme()

    def render_periods(
        self,
        source_workbook: Path,
        output_workbook: Path,
        periods: tuple[PaymentResolvedPeriod, ...],
    ) -> PaymentMultiLineRenderResult:
        source = Path(source_workbook)
        output = Path(output_workbook)
        active = tuple(period for period in periods if period.points)
        if not active:
            raise PaymentWorkbookError("No resolved Payment points were available to render.")

        output.parent.mkdir(parents=True, exist_ok=True)
        keep_vba = source.suffix.lower() == ".xlsm"
        with tempfile.NamedTemporaryFile(
            prefix="payment_backbone_", suffix=output.suffix, dir=output.parent, delete=False
        ) as handle:
            temp_path = Path(handle.name)

        colors: list[tuple[str, str]] = []
        rendered_points = 0
        try:
            with tempfile.TemporaryDirectory(prefix="payment_labels_") as label_dir:
                wb = load_workbook(source, keep_vba=keep_vba)
                try:
                    if self.MAIN_SHEET not in wb.sheetnames:
                        raise PaymentWorkbookError("Worksheet 'main' was not found.")
                    if self.PAYMENT_SHEET in wb.sheetnames:
                        wb.remove(wb[self.PAYMENT_SHEET])

                    main = wb[self.MAIN_SHEET]
                    payment = wb.copy_worksheet(main)
                    payment.title = self.PAYMENT_SHEET
                    # Keep the user-facing Payment pair together: main -> Payment Input -> Payment.
                    if "Payment Input" in wb.sheetnames:
                        wb._sheets.remove(payment)
                        input_index = wb.sheetnames.index("Payment Input")
                        wb._sheets.insert(input_index + 1, payment)
                    payment.freeze_panes = main.freeze_panes
                    payment.sheet_view.showGridLines = main.sheet_view.showGridLines
                    payment.auto_filter.ref = main.auto_filter.ref

                    timeline = self._timescale_boundaries(payment)
                    if not timeline:
                        raise PaymentWorkbookError("Weekly timescale dates were not found on the main sheet.")

                    for period in active:
                        color = self.theme.colors.get(period.period_id, self.theme.fallback_color)
                        line = Side(style=self.theme.line_style, color=color)
                        endpoint = Side(style=self.theme.endpoint_style, color=color)
                        # The backbone is an output, not an input: it sits at the
                        # latest resolved requirement boundary for this payment.
                        backbone = max(self._boundary(point) for point in period.points)
                        self._paint_header_marker(payment, period, backbone, line)
                        self._paint_vertical_backbone(payment, period, backbone, line, endpoint)
                        self._add_payment_label(
                            payment,
                            period,
                            backbone,
                            color,
                            Path(label_dir),
                        )
                        colors.append((period.period_id, color))
                        rendered_points += len(period.points)

                    wb.save(temp_path)
                finally:
                    wb.close()
            shutil.move(str(temp_path), str(output))
        except PaymentWorkbookError:
            raise
        except Exception as exc:
            raise PaymentWorkbookError(f"Payment lines could not be rendered: {exc}") from exc
        finally:
            temp_path.unlink(missing_ok=True)

        return PaymentMultiLineRenderResult(
            source_workbook=source,
            output_workbook=output,
            payment_sheet=self.PAYMENT_SHEET,
            period_ids=tuple(period.period_id for period in active),
            rendered_points=rendered_points,
            rendered_periods=len(active),
            colors=tuple(colors),
        )

    def render_single_period(
        self,
        source_workbook: Path,
        output_workbook: Path,
        period: PaymentResolvedPeriod,
    ) -> PaymentLineRenderResult:
        """Compatibility wrapper: single period rendered with the backbone engine."""
        result = self.render_periods(source_workbook, output_workbook, (period,))
        color = result.colors[0][1]
        return PaymentLineRenderResult(
            source_workbook=result.source_workbook,
            output_workbook=result.output_workbook,
            payment_sheet=result.payment_sheet,
            period_id=period.period_id,
            rendered_points=result.rendered_points,
            color=color,
        )

    def _paint_vertical_backbone(
        self,
        ws,
        period: PaymentResolvedPeriod,
        backbone_boundary: int,
        line: Side,
        endpoint: Side,
    ) -> None:
        points = tuple(sorted(period.points, key=lambda p: p.activity_row))
        if not points:
            return
        first_row = points[0].activity_row
        last_row = points[-1].activity_row
        self._vertical_boundary(ws, first_row, last_row, backbone_boundary, line)

        for point in points:
            target_boundary = self._boundary(point)
            self._horizontal_boundary(ws, point.activity_row, backbone_boundary, target_boundary, line)
            # A stronger one-row cap marks the actual resolved % target without
            # writing into the weekly Plan cells.
            self._vertical_boundary(ws, point.activity_row, point.activity_row, target_boundary, endpoint)

    def _paint_header_marker(
        self,
        ws,
        period: PaymentResolvedPeriod,
        boundary: int,
        line: Side,
    ) -> None:
        """Carry the backbone through the timescale header and attach a lightweight note."""
        self._vertical_boundary(ws, 1, self.HEADER_ROW, boundary, line)
        marker_col = boundary if boundary <= ws.max_column else ws.max_column
        marker_cell = ws.cell(self.HEADER_ROW, marker_col)
        eligible_text = (
            period.planned_eligible_date.isoformat()
            if period.planned_eligible_date else "unresolved"
        )
        controllers = ", ".join(period.controlling_activity_ids) or "n/a"
        legacy_text = period.payment_date.isoformat() if period.payment_date else "none"
        note = (
            f"{period.period_id} Planned Eligible Date: {eligible_text}\n"
            f"Controlling Activity: {controllers}\n"
            f"Input Payment Date (reference only): {legacy_text}"
        )
        if marker_cell.comment is not None and marker_cell.comment.text:
            note = marker_cell.comment.text + "\n\n" + note
        marker_cell.comment = Comment(note, "Progress Studio")


    def _add_payment_label(
        self,
        ws,
        period: PaymentResolvedPeriod,
        boundary: int,
        color: str,
        label_dir: Path,
    ) -> None:
        """Add one small floating badge per Payment; line geometry stays cell-based."""
        date_text = (
            period.planned_eligible_date.strftime("%d-%b-%y")
            if period.planned_eligible_date else "unresolved"
        )
        label_text = f"{period.period_id} | {date_text}"

        label = self.theme.label
        width, height = label.width_px, label.height_px
        image = Image.new("RGBA", (width, height), (255, 255, 255, 0))
        draw = ImageDraw.Draw(image)
        draw.rounded_rectangle(
            (0, 0, width - 1, height - 1),
            radius=label.corner_radius_px,
            fill=f"#{color}",
        )
        try:
            font = ImageFont.truetype("arialbd.ttf", label.font_size)
        except OSError:
            try:
                font = ImageFont.load_default(size=label.font_size)
            except TypeError:
                font = ImageFont.load_default()

        bbox = draw.textbbox((0, 0), label_text, font=font)
        text_w = bbox[2] - bbox[0]
        text_h = bbox[3] - bbox[1]
        draw.text(
            ((width - text_w) / 2, (height - text_h) / 2 - bbox[1]),
            label_text,
            fill=f"#{label.text_color}",
            font=font,
        )

        label_path = label_dir / f"{period.period_id}_label.png"
        image.save(label_path, optimize=True)

        badge = XLImage(str(label_path))
        badge.width = width
        badge.height = height
        marker_col = min(max(boundary, 1), ws.max_column)
        anchor_col = min(
            max(marker_col + label.anchor_column_offset, 1),
            ws.max_column,
        )
        badge.anchor = ws.cell(label.anchor_row, anchor_col).coordinate
        ws.add_image(badge)

    def _timescale_boundaries(self, ws) -> tuple[tuple[int, date], ...]:
        result: list[tuple[int, date]] = []
        for col in range(1, ws.max_column + 1):
            value = ws.cell(self.HEADER_ROW, col).value
            if isinstance(value, datetime):
                result.append((col, value.date()))
            elif isinstance(value, date):
                result.append((col, value))
        return tuple(result)


    @staticmethod
    def _boundary(point: PaymentResolvedPoint) -> int:
        return point.timescale_column if point.boundary_edge == "left" else point.timescale_column + 1

    @staticmethod
    def _replace_border(cell, *, left=None, right=None, top=None, bottom=None) -> None:
        old = copy(cell.border)
        cell.border = Border(
            left=left if left is not None else old.left,
            right=right if right is not None else old.right,
            top=top if top is not None else old.top,
            bottom=bottom if bottom is not None else old.bottom,
            diagonal=old.diagonal,
            diagonal_direction=old.diagonal_direction,
            diagonalUp=old.diagonalUp,
            diagonalDown=old.diagonalDown,
            outline=old.outline,
            vertical=old.vertical,
            horizontal=old.horizontal,
        )

    def _vertical_boundary(self, ws, row1: int, row2: int, boundary: int, line: Side) -> None:
        start, end = sorted((row1, row2))
        if boundary <= ws.max_column:
            for row in range(start, end + 1):
                self._replace_border(ws.cell(row, boundary), left=line)
        else:
            col = max(boundary - 1, 1)
            for row in range(start, end + 1):
                self._replace_border(ws.cell(row, col), right=line)

    def _horizontal_boundary(self, ws, row: int, boundary1: int, boundary2: int, line: Side) -> None:
        if boundary1 == boundary2:
            return
        left = min(boundary1, boundary2)
        right = max(boundary1, boundary2)
        for col in range(left, right):
            if 1 <= col <= ws.max_column:
                self._replace_border(ws.cell(row, col), bottom=line)
