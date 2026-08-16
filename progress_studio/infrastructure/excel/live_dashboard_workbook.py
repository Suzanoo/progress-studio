
from __future__ import annotations

from collections import OrderedDict
from datetime import date, datetime

from openpyxl.chart import LineChart, Reference
from openpyxl.chart.axis import DateAxis
from openpyxl.chart.legend import LegendEntry
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.datavalidation import DataValidation

from progress_studio.domain.activity_table import ActivityTableModel
from progress_studio.domain.main_dataset import MainDataset
from progress_studio.domain.progress_cache import ProgressCache
from progress_studio.infrastructure.excel.dashboard_workbook import (
    AMBER, BLUE, DATA_SHEET, DASHBOARD_SHEET, GREEN,
    LIGHT_AMBER, LIGHT_BLUE, LIGHT_GRAY, LIGHT_GREEN, LIGHT_RED,
    MUTED, NAVY, RED, WHITE, _FONT, _LAYOUT,
    _merge_title, _remove, _solid, _style_box, _thin_border, _add_kpi_icon,
)
from progress_studio.services.activity_table_deriver import ActivityTableDeriver
from progress_studio.services.progress_cache_deriver import ProgressCacheDeriver
from progress_studio.infrastructure.excel.live_scurve_workbook import build_live_progress_contract


def _as_date(value: date | datetime | None) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    return value


def _default_cutoff(cache: ProgressCache) -> date | None:
    latest_actual = None
    latest_any = None
    for point in cache.points:
        point_date = _as_date(point.reporting_date)
        if point_date is not None:
            latest_any = point_date
            if point.actual_cumulative is not None:
                latest_actual = point_date
    return latest_actual or latest_any


def _project_dates(dataset: MainDataset) -> tuple[date | None, date | None]:
    starts = [row.plan_start.date() for row in dataset.activities if row.plan_start is not None]
    finishes = [row.plan_finish.date() for row in dataset.activities if row.plan_finish is not None]
    return (min(starts) if starts else None, max(finishes) if finishes else None)


def _monthly_points(cache: ProgressCache):
    grouped: OrderedDict[tuple[int, int], object] = OrderedDict()
    for point in cache.points:
        if point.reporting_date is None:
            continue
        key = (point.reporting_date.year, point.reporting_date.month)
        grouped[key] = point
    return list(grouped.values())



def _find_scurve_rows(ws, dataset: MainDataset) -> tuple[int | None, int | None]:
    """Return Acc.Plan (AP) and Acc.Actual (AA) row numbers."""
    header_columns = {name: col for name, col in dataset.headers}
    row_type_col = header_columns.get("row type")
    pa_col = header_columns.get("p/a")
    if row_type_col is None or pa_col is None:
        return None, None

    acc_plan = None
    acc_actual = None
    for row in range(dataset.header_row + 1, ws.max_row + 1):
        row_type = str(ws.cell(row, row_type_col).value or "").strip().lower()
        pa = str(ws.cell(row, pa_col).value or "").strip().upper()
        if row_type != "s-curve":
            continue
        if pa == "AP":
            acc_plan = row
        elif pa == "AA":
            acc_actual = row
    return acc_plan, acc_actual


def _build_live_data_sheet(workbook, dataset: MainDataset, cache: ProgressCache) -> None:
    """LW-11.2/11.3 thin Weekly/Monthly selector over ``progress``.

    ``progress`` owns the complete S-Curve history only. Dashboard_Data owns the
    lightweight presentation mask: Weekly uses every progress row; Monthly uses
    the last reporting point in each calendar month; Actual points after cutoff
    return ``#N/A`` so Excel does not plot them. Actual values themselves are
    never recalculated because cutoff changed.
    """
    _remove(workbook, DATA_SHEET)
    ws = workbook.create_sheet(DATA_SHEET)
    ws.append([
        "Weekly Date", "Weekly Plan", "Weekly Actual",
        "Monthly Date", "Monthly Plan", "Monthly Actual",
        "Selected Date", "Selected Plan", "Selected Actual",
        "Weekly Cutoff", "Monthly Cutoff",
        "Selected Actual Raw", "Marker Date", "Cutoff Plan Marker", "Cutoff Actual Marker",
    ])

    if "progress" not in workbook.sheetnames:
        build_live_progress_contract(workbook, dataset)
    progress = workbook["progress"]

    weekly_count = max(0, progress.max_row - 1)
    month_last_rows: OrderedDict[tuple[int, int], tuple[int, date]] = OrderedDict()
    for idx, period in enumerate(dataset.periods, start=2):
        reporting = _as_date(period.reporting_date)
        if reporting is not None:
            month_last_rows[(reporting.year, reporting.month)] = (idx, reporting)
    monthly_points = list(month_last_rows.values())

    max_rows = max(weekly_count, len(monthly_points), 1)
    for idx in range(max_rows):
        row = idx + 2

        if idx < weekly_count:
            progress_row = idx + 2
            ws.cell(row, 1, f"='progress'!A{progress_row}")
            ws.cell(row, 2, f"='progress'!B{progress_row}")
            ws.cell(row, 3, f"='progress'!C{progress_row}")
            ws.cell(row, 10, f"='progress'!A{progress_row}")

        if idx < len(monthly_points):
            progress_row, reporting = monthly_points[idx]
            ws.cell(row, 4, reporting)
            ws.cell(row, 5, f"='progress'!B{progress_row}")
            ws.cell(row, 6, f"='progress'!C{progress_row}")
            ws.cell(row, 11, reporting)

        # LW-11.3.2: renderer-only cutoff mask. ``progress`` keeps complete
        # Actual history; only this selected chart series is cutoff-aware.
        # Do not let an empty Monthly source cell become Excel serial date 0.
        # A direct IF(..., Dn) over a blank Dn is evaluated as 0 by Excel; when
        # the chart auto-treats dates as a date axis this stretches the axis back
        # to 1899 and squeezes all real monthly points into a vertical line.
        ws.cell(
            row, 7,
            f'=IF(Dashboard!$G$5="Weekly",'
            f'IF(A{row}="","",A{row}),'
            f'IF(D{row}="","",D{row}))',
        )
        ws.cell(
            row, 8,
            f'=IF(G{row}="",NA(),IF(Dashboard!$G$5="Weekly",'
            f'IF(B{row}="",NA(),B{row}),IF(E{row}="",NA(),E{row})))',
        )
        # Raw selected Actual is deliberately separate from the chart mask.
        # KPI formulas read this error-free helper; only column I contains #N/A.
        ws.cell(
            row, 12,
            f'=IF(G{row}="","",IF(Dashboard!$G$5="Weekly",C{row},F{row}))',
        )
        ws.cell(
            row, 9,
            f'=IF(G{row}="",NA(),IF(G{row}>Dashboard!$K$5,NA(),'
            f'IF(L{row}="",NA(),L{row})))',
        )

    # LW-11.3.5: marker helpers are a *single-point* range, not a full
    # weekly/monthly series filled with #N/A. Excel can retain DataLabel objects
    # for hidden/error points, which caused dozens of labels to appear. Keeping
    # the marker source physically one row long guarantees one label per marker.
    ws["M2"] = "=Dashboard!$K$5"
    ws["N2"] = (
        f'=IFERROR(SUMIFS($H$2:$H${max_rows + 1},$G$2:$G${max_rows + 1},Dashboard!$K$5),NA())'
    )
    ws["O2"] = (
        f'=IFERROR(IF(ABS(SUMIFS($L$2:$L${max_rows + 1},$G$2:$G${max_rows + 1},Dashboard!$K$5)-'
        f'SUMIFS($H$2:$H${max_rows + 1},$G$2:$G${max_rows + 1},Dashboard!$K$5))>0.0000001,'
        f'SUMIFS($L$2:$L${max_rows + 1},$G$2:$G${max_rows + 1},Dashboard!$K$5),NA()),NA())'
    )

    for row in range(2, ws.max_row + 1):
        for col in (1, 4, 7, 10, 11, 13):
            ws.cell(row, col).number_format = "dd/mm/yyyy"
        for col in (2, 3, 5, 6, 8, 9, 12, 14, 15):
            ws.cell(row, col).number_format = "0.00%"
    ws.column_dimensions["J"].hidden = True
    ws.column_dimensions["K"].hidden = True
    ws.sheet_state = "hidden"

def _kpi_box(ws, title_range: str, value_range: str, title: str, value, fill: str, color: str, number_format: str = "0.00%", *, icon: str | None = None) -> None:
    ws.merge_cells(title_range)
    title_cell = ws[title_range.split(":")[0]]
    title_cell.value = title
    title_cell.font = Font(name=_FONT, bold=True, color=color, size=9)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(value_range)
    cell = ws[value_range.split(":")[0]]
    cell.value = value
    cell.number_format = number_format
    cell.fill = _solid(fill)
    cell.font = Font(name=_FONT, bold=True, color=color, size=15)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    for rg in (title_range, value_range):
        for row in ws[rg]:
            for c in row:
                c.fill = _solid(fill)
                c.border = _thin_border()
    if icon:
        _add_kpi_icon(ws, value_range.split(":")[0], icon)


def _write_activity_section(ws, model: ActivityTableModel, dataset: MainDataset) -> None:
    """Direct-to-main activity table: O(rows) formulas, no activity×time cache."""
    _merge_title(ws, "B37:M37", "ACTIVITY PROGRESS", 11)
    headers = ["WBS", "Activity", "Type", "Total", "Amount", "Progress", "Variance", "Status"]
    starts = ["B", "C", "F", "H", "J", "L", "N", "P"]
    ends = ["B", "E", "G", "I", "K", "M", "O", "Q"]
    for start, end, header in zip(starts, ends, headers):
        ws.merge_cells(f"{start}38:{end}38")
        cell = ws[f"{start}38"]
        cell.value = header
        cell.fill = _solid(NAVY)
        cell.font = Font(name=_FONT, bold=True, color=WHITE, size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in ws[f"{start}38:{end}38"]:
            for c in row:
                c.border = _thin_border()

    if not dataset.periods:
        return
    first_col = dataset.periods[0].column
    last_col = dataset.periods[-1].column
    from openpyxl.utils import get_column_letter
    first_letter = get_column_letter(first_col)
    last_letter = get_column_letter(last_col)
    header_row = dataset.header_row

    output_row = 39
    pair_index = 0
    for item in model.rows:
        row = output_row
        is_plan = item.type_label == "Plan"
        if is_plan:
            ws[f"B{row}"] = item.wbs if item.row_type != "project summary" else "PROJECT"
            ws.merge_cells(f"C{row}:E{row}")
            ws[f"C{row}"] = item.activity
        else:
            ws.merge_cells(f"C{row}:E{row}")

        ws.merge_cells(f"F{row}:G{row}")
        ws[f"F{row}"] = item.type_label
        ws.merge_cells(f"H{row}:I{row}")
        ws[f"H{row}"] = item.total
        ws.merge_cells(f"J{row}:K{row}")
        ws.merge_cells(f"L{row}:M{row}")
        ws.merge_cells(f"N{row}:O{row}")
        ws.merge_cells(f"P{row}:Q{row}")

        source_row = item.source_plan_row if is_plan else item.source_actual_row
        if source_row is not None:
            ws[f"L{row}"] = (
                f'=IFERROR(SUMIFS(main!${first_letter}${source_row}:'
                f'${last_letter}${source_row},'
                f'main!${first_letter}${header_row}:${last_letter}${header_row},'
                f'"<="&$K$5),0)'
            )
        else:
            ws[f"L{row}"] = 0

        # Amount follows Plan total x selected progress. This avoids reading a
        # second generated worksheet and uses only one formula per visual row.
        plan_row = row if is_plan else row - 1
        ws[f"J{row}"] = f'=IFERROR($H${plan_row}*L{row},0)' if not is_plan else f'=IFERROR(H{row}*L{row},0)'
        if is_plan:
            ws[f"N{row}"] = ""
            # Pair-filter contract: Plan gets the same Actual-derived status value
            # so Excel filtering keeps both rows together. It is hidden visually.
            ws[f"P{row}"] = (
                f'=IF(AND(L{row}<=0,L{row+1}<=0),"Not Due",'
                f'IF(AND(L{row}>0,L{row+1}<=0),"No Progress",'
                f'IF(L{row+1}>=1,"Complete",'
                f'IF(L{row+1}<L{row},"Behind","On Track"))))'
            )
        else:
            ws[f"N{row}"] = f'=IFERROR(L{row}-L{row-1},0)'
            ws[f"P{row}"] = (
                f'=IF(AND(L{row-1}<=0,L{row}<=0),"Not Due",'
                f'IF(AND(L{row-1}>0,L{row}<=0),"No Progress",'
                f'IF(L{row}>=1,"Complete",'
                f'IF(L{row}<L{row-1},"Behind","On Track"))))'
            )

        base_fill = WHITE if pair_index % 2 == 0 else LIGHT_GRAY
        if item.row_type in {"project summary", "wbs"}:
            base_fill = LIGHT_BLUE if item.outline_level <= 1 else "F5F8FC"
        for row_cells in ws[f"B{row}:Q{row}"]:
            for cell in row_cells:
                cell.border = _thin_border()
                cell.fill = _solid(base_fill)
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                if item.row_type in {"project summary", "wbs"}:
                    cell.font = Font(name=_FONT, bold=True, color=NAVY, size=9)

        pa_fill = LIGHT_BLUE if is_plan else LIGHT_GREEN
        pa_color = BLUE if is_plan else GREEN
        ws[f"F{row}"].fill = _solid(pa_fill)
        ws[f"F{row}"].font = Font(name=_FONT, bold=True, color=pa_color, size=9)
        ws[f"L{row}"].fill = _solid(pa_fill)
        ws[f"L{row}"].font = Font(name=_FONT, bold=True, color=pa_color, size=9)
        ws.row_dimensions[row].outlineLevel = min(max(item.outline_level, 0), 7)
        ws.row_dimensions[row].height = 22
        ws[f"H{row}"].number_format = "#,##0.00"
        ws[f"J{row}"].number_format = "#,##0.00"
        ws[f"L{row}"].number_format = "0.00%"
        ws[f"N{row}"].number_format = "0.00%;[Red]-0.00%;0.00%"
        if is_plan:
            # Keep filter value but make Status invisible on Plan row.
            ws[f"P{row}"].font = Font(name=_FONT, color=base_fill, size=9)
            ws[f"C{row}"].alignment = Alignment(
                vertical="center",
                wrap_text=True,
                indent=min(max(item.outline_level, 0), 7),
            )
        output_row += 1
        if not is_plan:
            pair_index += 1

    ws.sheet_properties.outlinePr.summaryBelow = False
    last_activity_row = max(38, output_row - 1)
    # Keep the Activity Table visually clean: only Status gets an Excel filter.
    ws.auto_filter.ref = f"P38:P{last_activity_row}"
    ws.print_area = f"B2:Q{max(56, last_activity_row)}"


def build_live_dashboard(
    workbook,
    dataset: MainDataset,
    *,
    project_name: str | None = None,
    cutoff: date | datetime | None = None,
) -> None:
    """LW-8 interactive Live dashboard with tiny formulas only."""
    cache = ProgressCacheDeriver().derive(dataset)
    cutoff_date = _as_date(cutoff) or _default_cutoff(cache)
    activity_model = ActivityTableDeriver().derive(dataset, cutoff=cutoff_date)
    # LW-11.1: progress is the sole curve calculation contract.
    build_live_progress_contract(workbook, dataset)
    _build_live_data_sheet(workbook, dataset, cache)

    _remove(workbook, DASHBOARD_SHEET)
    ws = workbook.create_sheet(DASHBOARD_SHEET, 0)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A7"

    widths = {
        "A": 3, "B": 14, "C": 14, "D": 14, "E": 14, "F": 14, "G": 14,
        "H": 14, "I": 14, "J": 14, "K": 14, "L": 14, "M": 14,
        "N": 11, "O": 11, "P": 12, "Q": 12,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.merge_cells("B2:M2")
    ws["B2"] = _LAYOUT["title"]
    ws["B2"].font = Font(name=_FONT, size=20, bold=True, color=NAVY)

    _merge_title(ws, "B4:M4", "PROJECT INFORMATION", 11)
    _style_box(ws, "B5:M6", LIGHT_GRAY)
    start_date, finish_date = _project_dates(dataset)
    project_value = float(cache.total_amount or 0.0)

    # Keep the established View/Cutoff control cells stable (G5/K5). Extra
    # project information lives on row 6 so older workbook formulas and user
    # muscle-memory do not move just because the information panel grew.
    ws["B5"] = "Project"
    ws.merge_cells("C5:D5")
    ws["C5"] = project_name or dataset.workbook_name
    ws["F5"] = "View"
    ws.merge_cells("G5:H5")
    ws["G5"] = "Weekly"
    ws["J5"] = "Cutoff Date"
    ws.merge_cells("K5:M5")
    ws["K5"] = cutoff_date
    ws["K5"].number_format = "dd/mm/yyyy"

    ws["B6"] = "Project Start"
    ws.merge_cells("C6:D6")
    ws["C6"] = start_date
    ws["C6"].number_format = "dd/mm/yyyy"
    ws["F6"] = "Project Finish"
    ws.merge_cells("G6:H6")
    ws["G6"] = finish_date
    ws["G6"].number_format = "dd/mm/yyyy"
    ws["J6"] = "Project Value"
    ws.merge_cells("K6:M6")
    ws["K6"] = project_value
    ws["K6"].number_format = "#,##0.00"

    view_validation = DataValidation(type="list", formula1='"Weekly,Monthly"', allow_blank=False)
    ws.add_data_validation(view_validation)
    view_validation.add(ws["G5"])

    data_ws = workbook[DATA_SHEET]
    weekly_count = len(cache.points)
    monthly_count = sum(1 for row in range(2, data_ws.max_row + 1) if data_ws.cell(row, 11).value not in (None, ""))
    cutoff_validation = DataValidation(
        type="list",
        formula1=(
            f'=INDIRECT(IF($G$5="Weekly","{DATA_SHEET}!$J$2:$J${max(2,weekly_count+1)}",'
            f'"{DATA_SHEET}!$K$2:$K${max(2,monthly_count+1)}"))'
        ),
        allow_blank=False,
    )
    ws.add_data_validation(cutoff_validation)
    cutoff_validation.add(ws["K5"])

    # KPI formulas read only the tiny selected-view cache.
    last_data_row = max(2, data_ws.max_row)
    # The cutoff dropdown contains exact reporting dates, so KPI values can use
    # simple SUMIFS against the selected-view helpers. This avoids LOOKUP over
    # the chart's #N/A mask and keeps KPI cards fully numeric/recalc-friendly.
    plan_formula = (
        f'=IFERROR(SUMIFS(Dashboard_Data!$H$2:$H${last_data_row},'
        f'Dashboard_Data!$G$2:$G${last_data_row},$K$5),0)'
    )
    actual_formula = (
        f'=IFERROR(SUMIFS(Dashboard_Data!$L$2:$L${last_data_row},'
        f'Dashboard_Data!$G$2:$G${last_data_row},$K$5),0)'
    )
    duration_days = max(0, (finish_date - start_date).days) if start_date and finish_date else 0
    schedule_formula = (
        '=IF(ROUND(E10-B10,4)=0,"ON SCHEDULE"&CHAR(10)&"0.00%",'
        'IF(E10<B10,"DELAY"&CHAR(10)&TEXT(E10-B10,"+0.00%;-0.00%;0.00%"),'
        '"AHEAD"&CHAR(10)&TEXT(E10-B10,"+0.00%;-0.00%;0.00%")))'
    )
    impact_formula = f'=ROUND(ABS(E10-B10)*{duration_days},0)&" Days"'

    _merge_title(ws, "B8:M8", "KPI SUMMARY", 11)
    _kpi_box(ws, "B9:D9", "B10:D12", "PLANNED PROGRESS", plan_formula, LIGHT_BLUE, BLUE, icon="planned")
    _kpi_box(ws, "E9:G9", "E10:G12", "ACTUAL PROGRESS", actual_formula, LIGHT_GREEN, GREEN, icon="actual")
    _kpi_box(ws, "H9:J9", "H10:J12", "SCHEDULE STATUS", schedule_formula, LIGHT_AMBER, AMBER, "General", icon="schedule")
    ws["H10"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    _kpi_box(ws, "K9:M9", "K10:M12", "TIME IMPACT", impact_formula, LIGHT_RED, RED, "General", icon="time_impact")

    _merge_title(ws, "B15:M15", "S-CURVE — PLAN VS ACTUAL", 11)
    _style_box(ws, "B16:M34", WHITE)
    chart = LineChart()
    chart.height = float(_LAYOUT["chart_height"])
    chart.width = float(_LAYOUT["chart_width"])
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = float(_LAYOUT.get("chart_y_max", 1.10))
    chart.y_axis.numFmt = "0%"
    chart.legend.position = "t"
    chart.display_blanks = "gap"

    # Match the normal Dashboard renderer. Live rebuild is a data-ownership mode,
    # not a second visual theme.
    grid = GraphicalProperties()
    grid.line.solidFill = "E5E7EB"
    grid.line.width = 9000
    chart.y_axis.majorGridlines.graphicalProperties = grid
    axis_line = GraphicalProperties()
    axis_line.line.solidFill = "B8C2CC"
    axis_line.line.width = 9000
    chart.y_axis.spPr = axis_line
    date_axis = DateAxis()
    date_axis.number_format = "mmm-yy"
    date_axis.majorTimeUnit = "days"
    date_axis.title = "Date"
    date_axis.spPr = axis_line
    chart.x_axis = date_axis
    data = Reference(data_ws, min_col=8, max_col=9, min_row=1, max_row=last_data_row)
    cats = Reference(data_ws, min_col=7, min_row=2, max_row=last_data_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    if len(chart.series) >= 2:
        chart.series[0].graphicalProperties.line.solidFill = BLUE
        chart.series[1].graphicalProperties.line.solidFill = GREEN
        # Keep the two main curves clean: no point markers along the full line.
        chart.series[0].marker.symbol = "none"
        chart.series[1].marker.symbol = "none"

    # Two marker-only series expose the current cutoff positions without labels.
    # KPI cards carry the numeric values, so the chart stays visually clean.
    marker_data = Reference(data_ws, min_col=14, max_col=15, min_row=1, max_row=2)
    marker_cats = Reference(data_ws, min_col=13, min_row=2, max_row=2)
    chart.add_data(marker_data, titles_from_data=True)
    # add_data() applies the chart's existing category range to new series in
    # some Excel/openpyxl combinations. Force both marker series to the single
    # dynamic cutoff date cell so the marker source remains one physical point.
    for marker_series in chart.series[2:4]:
        marker_series.cat = None
    chart.set_categories(cats)
    from openpyxl.chart.data_source import AxDataSource, NumRef
    marker_cat_ref = f"'{DATA_SHEET}'!$M$2:$M$2"
    for marker_series in chart.series[2:4]:
        marker_series.cat = AxDataSource(numRef=NumRef(f=marker_cat_ref))
    for idx, color in ((2, BLUE), (3, GREEN)):
        if len(chart.series) <= idx:
            continue
        series = chart.series[idx]
        series.graphicalProperties.line.noFill = True
        series.marker.symbol = "circle"
        series.marker.size = 7
        series.marker.graphicalProperties.solidFill = color
        series.marker.graphicalProperties.line.solidFill = color

    # Marker-only helper series are useful on the plot but add noise to the
    # legend; keep the legend focused on the two actual curves.
    if chart.legend is not None:
        chart.legend.legendEntry = [LegendEntry(idx=2, delete=True), LegendEntry(idx=3, delete=True)]

    # Minimal axes: percent scale is self-explanatory; dates remain the only
    # horizontal context. Excel will thin date labels automatically as needed.
    chart.y_axis.majorUnit = 0.25
    chart.y_axis.title = "Progress (%)"
    ws.add_chart(chart, "B16")

    ws.merge_cells("B35:M35")
    ws["B35"] = "Plan = full baseline • Actual history stays fixed • Cutoff masks visible Actual points"
    ws["B35"].font = Font(name=_FONT, size=9, italic=True, color=MUTED)

    _write_activity_section(ws, activity_model, dataset)
