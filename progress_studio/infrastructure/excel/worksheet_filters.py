from __future__ import annotations

from openpyxl.utils import get_column_letter
from openpyxl.worksheet.filters import FilterColumn


def configure_filter_buttons(
    ws,
    *,
    header_row: int,
    last_row: int,
    last_col: int,
    visible_columns: set[int],
    first_col: int = 1,
) -> None:
    """Keep AutoFilter behavior while exposing buttons only on useful columns.

    ``visible_columns`` contains 1-based worksheet column indexes. Excel stores
    FilterColumn ``colId`` values relative to the AutoFilter range, so every
    non-visible column receives ``showButton=False`` rather than shrinking the
    filter range and breaking row filtering across the sheet.
    """
    if last_col < first_col or last_row < header_row:
        return

    ws.auto_filter.ref = (
        f"{get_column_letter(first_col)}{header_row}:"
        f"{get_column_letter(last_col)}{last_row}"
    )
    ws.auto_filter.filterColumn = []
    for column in range(first_col, last_col + 1):
        if column in visible_columns:
            continue
        ws.auto_filter.filterColumn.append(
            FilterColumn(colId=column - first_col, showButton=False)
        )
