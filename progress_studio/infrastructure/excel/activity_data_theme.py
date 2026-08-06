from __future__ import annotations

from copy import copy
from openpyxl.styles import Font, PatternFill

from progress_studio.infrastructure.excel.export_theme import (
    ActivityDataPalette,
    DEFAULT_ACTIVITY_DATA_PALETTE,
)


def _normalized(value: object) -> str:
    return str(value or "").strip().lower()


def _as_level(value: object) -> int | None:
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _activity_data_last_column(ws, header_row: int) -> int:
    """Return the column before the first weekly timescale date.

    The weekly dates are stored on the same row as the Activity Data headers.
    Keeping this boundary explicit guarantees that this formatter never changes
    timescale fills, borders, or bars.
    """
    for col in range(1, ws.max_column + 1):
        value = ws.cell(header_row, col).value
        if hasattr(value, "year") and hasattr(value, "month") and hasattr(value, "day"):
            return col - 1
    return ws.max_column


def apply_activity_data_wbs_hierarchy(
    ws,
    *,
    header_row: int = 4,
    first_data_row: int = 5,
    palette: ActivityDataPalette = DEFAULT_ACTIVITY_DATA_PALETTE,
) -> None:
    """Apply a four-step WBS color hierarchy to Activity Data only.

    WBS levels 1 through 4 use progressively lighter fills. Levels deeper than
    4 intentionally reuse the level-4 fill so the palette remains readable and
    predictable. Project Summary, Activity rows, borders, and all timescale
    cells are left unchanged. Actual rows inherit their paired WBS Plan style.
    """
    headers = {
        _normalized(ws.cell(header_row, col).value): col
        for col in range(1, ws.max_column + 1)
        if _normalized(ws.cell(header_row, col).value)
    }
    row_type_col = headers.get("row type")
    outline_col = headers.get("outline level")
    pa_col = headers.get("p/a")
    if row_type_col is None or outline_col is None:
        return

    last_data_col = _activity_data_last_column(ws, header_row)
    level_fills = {
        1: PatternFill("solid", fgColor=palette.wbs_level_1_fill),
        2: PatternFill("solid", fgColor=palette.wbs_level_2_fill),
        3: PatternFill("solid", fgColor=palette.wbs_level_3_fill),
        4: PatternFill("solid", fgColor=palette.wbs_level_4_fill),
    }

    previous_wbs_level: int | None = None
    for row in range(first_data_row, ws.max_row + 1):
        row_type = _normalized(ws.cell(row, row_type_col).value)
        pa = str(ws.cell(row, pa_col).value or "").strip().upper() if pa_col else ""

        if row_type == "wbs":
            level = _as_level(ws.cell(row, outline_col).value)
            previous_wbs_level = level
        elif pa == "A" and previous_wbs_level is not None:
            level = previous_wbs_level
        else:
            previous_wbs_level = None
            continue

        if level is None or level < 1:
            continue
        fill = level_fills[min(level, 4)]

        for col in range(1, last_data_col + 1):
            cell = ws.cell(row, col)
            cell.fill = copy(fill)
            cell.font = copy(cell.font)
            cell.font = Font(
                name=cell.font.name,
                size=cell.font.size,
                bold=True,
                italic=cell.font.italic,
                vertAlign=cell.font.vertAlign,
                underline=cell.font.underline,
                strike=cell.font.strike,
                color=palette.font_color,
            )
