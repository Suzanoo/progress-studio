from __future__ import annotations

import argparse
import sys
import traceback
from collections import defaultdict
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.styles import Alignment, Font, PatternFill, Protection
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.workbook.defined_name import DefinedName
except ImportError as exc:
    raise SystemExit(
        "openpyxl was not found.\nInstall it with: pip install openpyxl"
    ) from exc

from progress_studio.infrastructure.excel.calculation_policy import configure_incremental_excel_recalculation
from progress_studio.infrastructure.excel.export_theme import DEFAULT_TIMESCALE_PALETTE
from progress_studio.infrastructure.excel.worksheet_filters import configure_filter_buttons


SCRIPT_VERSION = "6.4-wbs-level-color-hierarchy"
DEFAULT_SHEET = "main"
HEADER_ROW = 4
FIRST_DATA_ROW = 5
CURRENCY_FORMAT = '#,##0.00'
PERCENT_FORMAT = '0.00%'

def solid_fill(rgb: str) -> PatternFill:
    """Create an openpyxl-safe solid fill using 8-digit aRGB."""
    argb = str(rgb).strip().replace("#", "").upper()

    if len(argb) == 6:
        argb = "FF" + argb

    if len(argb) != 8:
        raise ValueError(
            f"Invalid Excel color: {rgb!r}. "
            "Expected #RRGGBB, RRGGBB, or AARRGGBB."
        )

    return PatternFill(
        fill_type="solid",
        start_color=argb,
        end_color=argb,
        fgColor=argb,
        bgColor=argb,
    )


# Export colors are configured in export_theme.py.
_TIMESCALE = DEFAULT_TIMESCALE_PALETTE
ACTIVITY_PLAN_FILL = solid_fill(_TIMESCALE.activity_plan_fill)
ACTIVITY_ACTUAL_FILL = solid_fill(_TIMESCALE.activity_actual_fill)
PROJECT_PLAN_FILL = solid_fill(_TIMESCALE.project_plan_fill)
PROJECT_ACTUAL_FILL = solid_fill(_TIMESCALE.project_actual_fill)
WBS_LEVEL1_PLAN_FILL = solid_fill(_TIMESCALE.wbs_level_1_plan_fill)
WBS_LEVEL1_ACTUAL_FILL = solid_fill(_TIMESCALE.wbs_level_1_actual_fill)
WBS_LEVEL2_PLAN_FILL = solid_fill(_TIMESCALE.wbs_level_2_plan_fill)
WBS_LEVEL2_ACTUAL_FILL = solid_fill(_TIMESCALE.wbs_level_2_actual_fill)

# Backward-compatible aliases used by S-Curve styling.
WBS_PLAN_FILL = WBS_LEVEL2_PLAN_FILL
WBS_ACTUAL_FILL = WBS_LEVEL2_ACTUAL_FILL
SCURVE_PLAN_FILL = solid_fill(_TIMESCALE.scurve_plan_fill)
SCURVE_ACTUAL_FILL = solid_fill(_TIMESCALE.scurve_actual_fill)
SCURVE_ACC_FILL = solid_fill(_TIMESCALE.scurve_acc_fill)


def normalize_header(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip().lower()


def normalize_activity_id(value: object) -> str:
    """Use the Activity ID directly from P6, such as A1000, without a project prefix."""
    if value is None:
        return ""
    return str(value).strip().upper()


def find_sheet(workbook, sheet_name: str):
    if sheet_name in workbook.sheetnames:
        return workbook[sheet_name]
    raise ValueError(
        f"Worksheet not found: '{sheet_name}' "
        f"(available worksheets: {', '.join(workbook.sheetnames)})"
    )


def find_header_row_and_columns(
    ws,
    activity_id_header: str,
    amount_header: str,
    search_rows: int = 30,
) -> tuple[int, int, int]:
    """Find the BOQ header row without requiring it to be the first row."""
    activity_key = normalize_header(activity_id_header)
    amount_key = normalize_header(amount_header)

    for row in range(1, min(ws.max_row, search_rows) + 1):
        header_map: dict[str, int] = {}
        for col in range(1, ws.max_column + 1):
            value = normalize_header(ws.cell(row, col).value)
            if value:
                header_map[value] = col

        if activity_key in header_map and amount_key in header_map:
            return row, header_map[activity_key], header_map[amount_key]

    raise ValueError(
        f"BOQ headers '{activity_id_header}' and '{amount_header}' were not found "
        f"within {search_rows} first rows of worksheet '{ws.title}'"
    )


def read_boq_amounts(
    boq_file: Path,
    boq_sheet_name: str | None,
    activity_id_header: str,
    amount_header: str,
) -> tuple[dict[str, float], int, int]:
    wb = load_workbook(boq_file, data_only=True, read_only=False)

    if boq_sheet_name:
        if boq_sheet_name not in wb.sheetnames:
            wb.close()
            raise ValueError(f"BOQ worksheet not found: '{boq_sheet_name}'")
        candidate_sheets = [wb[boq_sheet_name]]
    else:
        candidate_sheets = [wb[name] for name in wb.sheetnames]

    selected_ws = None
    header_row = activity_col = amount_col = None

    for ws in candidate_sheets:
        try:
            header_row, activity_col, amount_col = find_header_row_and_columns(
                ws,
                activity_id_header,
                amount_header,
            )
            selected_ws = ws
            break
        except ValueError:
            continue

    if selected_ws is None or header_row is None:
        wb.close()
        raise ValueError(
            "No BOQ worksheet contains the required Activity ID and Amount columns."
        )

    amounts: dict[str, float] = defaultdict(float)
    mapped_rows = 0
    skipped_rows = 0

    for row in range(header_row + 1, selected_ws.max_row + 1):
        activity_id = normalize_activity_id(
            selected_ws.cell(row, activity_col).value
        )
        raw_amount = selected_ws.cell(row, amount_col).value

        if not activity_id:
            skipped_rows += 1
            continue

        try:
            amount = float(raw_amount)
        except (TypeError, ValueError):
            skipped_rows += 1
            continue

        amounts[activity_id] += amount
        mapped_rows += 1

    wb.close()
    return dict(amounts), mapped_rows, skipped_rows


def get_header_map(ws) -> dict[str, int]:
    result: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        header = normalize_header(ws.cell(HEADER_ROW, col).value)
        if header:
            result[header] = col
    return result


def require_columns(header_map: dict[str, int], required: list[str]) -> None:
    missing = [name for name in required if name.lower() not in header_map]
    if missing:
        raise ValueError("Required columns not found: " + ", ".join(missing))


def build_plan_rows(
    ws,
    row_type_col: int,
    pa_col: int,
    outline_level_col: int,
) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []

    for row in range(HEADER_ROW + 1, ws.max_row + 1):
        if normalize_header(ws.cell(row, pa_col).value) != "p":
            continue

        row_type = normalize_header(ws.cell(row, row_type_col).value)
        if row_type not in {"project summary", "wbs", "activity"}:
            continue

        raw_level = ws.cell(row, outline_level_col).value
        try:
            level = int(raw_level)
        except (TypeError, ValueError):
            continue

        rows.append({"row": row, "row_type": row_type, "level": level})

    return rows


def find_descendant_end(rows: list[dict[str, object]], index: int) -> int:
    """Return the last Plan row below the parent based on Outline Level."""
    current_level = int(rows[index]["level"])
    end_row = int(rows[index]["row"])

    for item in rows[index + 1 :]:
        if int(item["level"]) <= current_level:
            break
        end_row = int(item["row"])

    return end_row


def find_descendant_physical_end(
    ws,
    rows: list[dict[str, object]],
    index: int,
    pa_col: int,
) -> int:
    """
    Return the actual last descendant row, including the last child Actual row.

    Data uses adjacent Plan/Actual rows, while plan_rows contains Plan rows only.
    Calling find_descendant_end() directly would exclude the last Actual row.
    This would omit one activity from WBS Actual calculations.
    """
    plan_end = find_descendant_end(rows, index)
    actual_candidate = plan_end + 1

    if (
        actual_candidate <= ws.max_row
        and normalize_header(ws.cell(actual_candidate, pa_col).value) == "a"
    ):
        return actual_candidate

    return plan_end


def find_timescale_columns(ws, header_map: dict[str, int]) -> list[int]:
    """Find the timescale from row 3 week labels and row 4 dates."""
    result: list[int] = []

    for col in range(1, ws.max_column + 1):
        week_label = str(ws.cell(3, col).value or "").strip().upper()
        if week_label.startswith("W") and week_label[1:].isdigit():
            result.append(col)

    if not result:
        raise ValueError("Weekly timescale from Script 03 was not found.")
    return result


def activity_plan_rows(plan_rows: list[dict[str, object]]) -> list[int]:
    return [
        int(item["row"])
        for item in plan_rows
        if str(item["row_type"]) == "activity"
    ]


def set_activity_amounts(
    ws,
    plan_rows: list[dict[str, object]],
    activity_id_col: int,
    pa_col: int,
    amount_col: int,
    boq_amounts: dict[str, float],
) -> tuple[int, list[str]]:
    mapped = 0
    missing_ids: list[str] = []
    amount_letter = get_column_letter(amount_col)

    for item in plan_rows:
        if str(item["row_type"]) != "activity":
            continue

        plan_row = int(item["row"])
        activity_id = normalize_activity_id(ws.cell(plan_row, activity_id_col).value)

        if activity_id and activity_id in boq_amounts:
            ws.cell(plan_row, amount_col).value = boq_amounts[activity_id]
            mapped += 1
        else:
            ws.cell(plan_row, amount_col).value = 0
            missing_ids.append(activity_id or f"<blank row {plan_row}>")

        # Actual rows reference Plan amounts for weighted progress formulas.
        # The values are hidden to keep the worksheet clean.
        actual_row = plan_row + 1
        if (
            actual_row <= ws.max_row
            and normalize_header(ws.cell(actual_row, pa_col).value) == "a"
        ):
            ws.cell(actual_row, amount_col).value = f"={amount_letter}{plan_row}"
            ws.cell(actual_row, amount_col).number_format = ";;;"

    return mapped, missing_ids



def prepare_activity_amount_refs(
    ws,
    plan_rows: list[dict[str, object]],
    pa_col: int,
    amount_col: int,
) -> tuple[int, int]:
    """Preserve user-entered Amount values and make Actual rows reference Plan rows.

    Return (activities_with_amount, activities_without_amount).
    """
    amount_letter = get_column_letter(amount_col)
    with_amount = 0
    without_amount = 0
    for item in plan_rows:
        if str(item["row_type"]) != "activity":
            continue
        plan_row = int(item["row"])
        value = ws.cell(plan_row, amount_col).value
        if value in (None, ""):
            without_amount += 1
        else:
            with_amount += 1
            ws.cell(plan_row, amount_col).number_format = CURRENCY_FORMAT

        actual_row = plan_row + 1
        if actual_row <= ws.max_row and normalize_header(ws.cell(actual_row, pa_col).value) == "a":
            # Actual Amount is calculated later, after % Complete formulas exist.
            ws.cell(actual_row, amount_col).value = None
            ws.cell(actual_row, amount_col).number_format = CURRENCY_FORMAT
    return with_amount, without_amount

def rollup_amounts(
    ws,
    plan_rows: list[dict[str, object]],
    row_type_col: int,
    pa_col: int,
    amount_col: int,
) -> tuple[int, int]:
    amount_letter = get_column_letter(amount_col)
    row_type_letter = get_column_letter(row_type_col)
    pa_letter = get_column_letter(pa_col)
    wbs_count = 0
    project_count = 0

    for index in range(len(plan_rows) - 1, -1, -1):
        item = plan_rows[index]
        row = int(item["row"])
        row_type = str(item["row_type"])

        if row_type not in {"wbs", "project summary"}:
            continue

        end_row = find_descendant_end(plan_rows, index)
        if end_row <= row:
            ws.cell(row, amount_col).value = 0
        else:
            ws.cell(row, amount_col).value = (
                f'=SUMIFS(${amount_letter}${row + 1}:${amount_letter}${end_row},'
                f'${row_type_letter}${row + 1}:${row_type_letter}${end_row},'
                f'"Activity",'
                f'${pa_letter}${row + 1}:${pa_letter}${end_row},"P")'
            )

        # Parent Actual Amount is calculated later from descendant Activity Actual rows.
        actual_row = row + 1
        if (
            actual_row <= ws.max_row
            and normalize_header(ws.cell(actual_row, pa_col).value) == "a"
        ):
            ws.cell(actual_row, amount_col).value = None
            ws.cell(actual_row, amount_col).number_format = CURRENCY_FORMAT

        if row_type == "wbs":
            wbs_count += 1
        else:
            project_count += 1

    return wbs_count, project_count


def add_progress_formulas(
    ws,
    plan_rows: list[dict[str, object]],
    timescale_cols: list[int],
    row_type_col: int,
    activity_id_col: int,
    pa_col: int,
    amount_col: int,
) -> None:
    """
    Activity: User input
    WBS/Project: amount-weighted weekly progress from descendant activities.

    Use the full range including the last activity Actual row.
    Calculate the denominator directly from activity rows rather than parent Amount values.
    """
    amount_letter = get_column_letter(amount_col)
    row_type_letter = get_column_letter(row_type_col)
    activity_id_letter = get_column_letter(activity_id_col)
    pa_letter = get_column_letter(pa_col)

    for index, item in enumerate(plan_rows):
        row = int(item["row"])
        row_type = str(item["row_type"])

        if row_type == "activity":
            for current_row in (row, row + 1):
                if current_row > ws.max_row:
                    continue
                for col in timescale_cols:
                    ws.cell(current_row, col).number_format = PERCENT_FORMAT
                    ws.cell(current_row, col).protection = Protection(locked=False)
            continue

        if row_type not in {"wbs", "project summary"}:
            continue

        end_row = find_descendant_physical_end(
            ws, plan_rows, index, pa_col
        )
        if end_row <= row:
            continue

        # Each parent uses two adjacent rows: Plan then Actual.
        # Children therefore begin after the parent Actual row, not at row + 1.
        # Example: WBS Plan=13, WBS Actual=14, first child=15.
        first_child_row = row + 2

        if first_child_row > end_row:
            continue

        for target_row, pa_code in ((row, "P"), (row + 1, "A")):
            if target_row > ws.max_row:
                continue

            for col in timescale_cols:
                week_letter = get_column_letter(col)

                # Actual rows have no Row Type in the Script 02 workbook.
                # They have Activity ID and P/A="A", so Activity ID identifies the activity.
                if pa_code == "A":
                    activity_test = (
                        f'--(${activity_id_letter}${first_child_row}:'
                        f'${activity_id_letter}${end_row}<>"")'
                    )
                    count_formula = (
                        f'COUNTIFS('
                        f'${activity_id_letter}${first_child_row}:${activity_id_letter}${end_row},"<>",'
                        f'${pa_letter}${first_child_row}:${pa_letter}${end_row},"A",'
                        f'{week_letter}${first_child_row}:{week_letter}${end_row},"<>")'
                    )
                else:
                    activity_test = (
                        f'--(${row_type_letter}${first_child_row}:'
                        f'${row_type_letter}${end_row}="Activity")'
                    )
                    count_formula = (
                        f'COUNTIFS('
                        f'${row_type_letter}${first_child_row}:${row_type_letter}${end_row},"Activity",'
                        f'${pa_letter}${first_child_row}:${pa_letter}${end_row},"P",'
                        f'{week_letter}${first_child_row}:{week_letter}${end_row},"<>")'
                    )

                # Weekly progress must always be weighted by the full Plan Amount.
                # For Actual rows the full amount is on the preceding Plan row, while
                # the visible Actual Amount column contains earned value.
                if pa_code == "A":
                    weight_range = (
                        f'${amount_letter}${first_child_row - 1}:'
                        f'${amount_letter}${end_row - 1}'
                    )
                else:
                    weight_range = (
                        f'${amount_letter}${first_child_row}:'
                        f'${amount_letter}${end_row}'
                    )

                weighted_sum = (
                    f'SUMPRODUCT('
                    f'{activity_test},'
                    f'--(${pa_letter}${first_child_row}:${pa_letter}${end_row}="{pa_code}"),'
                    f'{weight_range},'
                    f'{week_letter}${first_child_row}:{week_letter}${end_row})'
                )

                weight_sum = (
                    f'SUMPRODUCT('
                    f'{activity_test},'
                    f'--(${pa_letter}${first_child_row}:${pa_letter}${end_row}="{pa_code}"),'
                    f'{weight_range})'
                )

                formula = (
                    f'=IF({count_formula}=0,"",'
                    f'IFERROR({weighted_sum}/{weight_sum},""))'
                )

                ws.cell(target_row, col).value = formula
                ws.cell(target_row, col).number_format = PERCENT_FORMAT
                ws.cell(target_row, col).protection = Protection(locked=True)


def add_percent_complete_formulas(
    ws,
    plan_rows: list[dict[str, object]],
    timescale_cols: list[int],
    pa_col: int,
    percent_complete_col: int,
) -> None:
    """Calculate % Complete as the sum of weekly progress values for each row."""
    first_week = get_column_letter(timescale_cols[0])
    last_week = get_column_letter(timescale_cols[-1])

    for item in plan_rows:
        plan_row = int(item["row"])
        for row in (plan_row, plan_row + 1):
            if row > ws.max_row:
                continue
            pa_value = normalize_header(ws.cell(row, pa_col).value)
            if pa_value not in {"p", "a"}:
                continue

            week_range = f"{first_week}{row}:{last_week}{row}"
            ws.cell(row, percent_complete_col).value = (
                f'=IF(COUNT({week_range})=0,"",SUM({week_range}))'
            )
            ws.cell(row, percent_complete_col).number_format = PERCENT_FORMAT
            ws.cell(row, percent_complete_col).protection = Protection(locked=True)


def add_actual_amount_formulas(
    ws,
    plan_rows: list[dict[str, object]],
    activity_id_col: int,
    pa_col: int,
    amount_col: int,
    percent_complete_col: int,
) -> None:
    """Calculate earned Actual Amount for activities and roll it up to parents.

    Activity Actual Amount = Plan Amount x Actual % Complete.
    WBS/Project Actual Amount = sum of descendant Activity Actual Amounts.
    """
    activity_id_letter = get_column_letter(activity_id_col)
    pa_letter = get_column_letter(pa_col)
    amount_letter = get_column_letter(amount_col)
    percent_letter = get_column_letter(percent_complete_col)

    # Activities first so parent SUMIFS formulas can roll them up.
    for item in plan_rows:
        if str(item["row_type"]) != "activity":
            continue
        plan_row = int(item["row"])
        actual_row = plan_row + 1
        if (
            actual_row > ws.max_row
            or normalize_header(ws.cell(actual_row, pa_col).value) != "a"
        ):
            continue
        ws.cell(actual_row, amount_col).value = (
            f'=IF({percent_letter}{actual_row}="","",'
            f'{amount_letter}{plan_row}*{percent_letter}{actual_row})'
        )
        ws.cell(actual_row, amount_col).number_format = CURRENCY_FORMAT
        ws.cell(actual_row, amount_col).protection = Protection(locked=True)

    # Roll up from deepest parent to Project Summary.
    for index in range(len(plan_rows) - 1, -1, -1):
        item = plan_rows[index]
        row_type = str(item["row_type"])
        if row_type not in {"wbs", "project summary"}:
            continue
        plan_row = int(item["row"])
        actual_row = plan_row + 1
        if (
            actual_row > ws.max_row
            or normalize_header(ws.cell(actual_row, pa_col).value) != "a"
        ):
            continue
        end_row = find_descendant_physical_end(ws, plan_rows, index, pa_col)
        first_child_row = plan_row + 2
        if first_child_row > end_row:
            ws.cell(actual_row, amount_col).value = 0
        else:
            ws.cell(actual_row, amount_col).value = (
                f'=SUMIFS(${amount_letter}${first_child_row}:${amount_letter}${end_row},'
                f'${activity_id_letter}${first_child_row}:${activity_id_letter}${end_row},"<>",'
                f'${pa_letter}${first_child_row}:${pa_letter}${end_row},"A")'
            )
        ws.cell(actual_row, amount_col).number_format = CURRENCY_FORMAT
        ws.cell(actual_row, amount_col).protection = Protection(locked=True)


def add_percent_complete_warning(
    ws,
    percent_complete_col: int,
    last_progress_row: int,
) -> None:
    """Color the cell red when % Complete exceeds 100%."""
    col_letter = get_column_letter(percent_complete_col)
    target = f"{col_letter}{FIRST_DATA_ROW}:{col_letter}{last_progress_row}"

    # Clear only existing conditional formatting affecting the % Complete column.
    try:
        for key in list(ws.conditional_formatting._cf_rules):
            remove_rule = False
            for cell_range in key.sqref.ranges:
                if cell_range.min_col <= percent_complete_col <= cell_range.max_col:
                    remove_rule = True
                    break
            if remove_rule:
                del ws.conditional_formatting._cf_rules[key]
    except AttributeError:
        pass

    red_fill = solid_fill("FFC7CE")
    dark_red_font = Font(color="FF9C0006", bold=True)
    ws.conditional_formatting.add(
        target,
        FormulaRule(
            formula=[f'AND({col_letter}{FIRST_DATA_ROW}<>"",{col_letter}{FIRST_DATA_ROW}>1)'],
            fill=red_fill,
            font=dark_red_font,
            stopIfTrue=True,
        ),
    )


def clear_timescale_direct_fills(
    ws,
    timescale_cols: list[int],
    last_progress_row: int,
) -> None:
    """
    Clear inherited fills only within the weekly timescale.
    Keep blank cells unfilled and let conditional formatting control colors.
    """
    no_fill = PatternFill()
    for row in range(FIRST_DATA_ROW, last_progress_row + 1):
        for col in timescale_cols:
            ws.cell(row, col).fill = no_fill


def clear_progress_conditional_formatting(ws, timescale_cols: list[int]) -> None:
    """Clear existing weekly-timescale conditional formatting before rebuilding it."""
    start_col = timescale_cols[0]
    end_col = timescale_cols[-1]

    try:
        for key in list(ws.conditional_formatting._cf_rules):
            remove_rule = False
            for cell_range in key.sqref.ranges:
                if not (
                    cell_range.max_col < start_col
                    or cell_range.min_col > end_col
                ):
                    remove_rule = True
                    break
            if remove_rule:
                del ws.conditional_formatting._cf_rules[key]
    except AttributeError:
        pass


def add_progress_conditional_formatting(
    ws,
    timescale_cols: list[int],
    row_type_col: int,
    activity_id_col: int,
    pa_col: int,
    outline_level_col: int,
    last_progress_row: int,
) -> None:
    start_letter = get_column_letter(timescale_cols[0])
    end_letter = get_column_letter(timescale_cols[-1])
    row_type_letter = get_column_letter(row_type_col)
    activity_id_letter = get_column_letter(activity_id_col)
    pa_letter = get_column_letter(pa_col)
    outline_level_letter = get_column_letter(outline_level_col)
    target = f"{start_letter}{FIRST_DATA_ROW}:{end_letter}{last_progress_row}"
    first_week = start_letter

    rules = [
        # Activity Plan / Actual: light colors
        FormulaRule(
            formula=[f'AND(${row_type_letter}{FIRST_DATA_ROW}="Activity",${pa_letter}{FIRST_DATA_ROW}="P",{first_week}{FIRST_DATA_ROW}<>"")'],
            fill=ACTIVITY_PLAN_FILL,
            stopIfTrue=True,
        ),
        FormulaRule(
            formula=[f'AND(${activity_id_letter}{FIRST_DATA_ROW}<>"",${pa_letter}{FIRST_DATA_ROW}="A",{first_week}{FIRST_DATA_ROW}<>"")'],
            fill=ACTIVITY_ACTUAL_FILL,
            stopIfTrue=True,
        ),
        # Project summary: darkest band.
        FormulaRule(
            formula=[f'AND(${row_type_letter}{FIRST_DATA_ROW}="Project Summary",${pa_letter}{FIRST_DATA_ROW}="P",{first_week}{FIRST_DATA_ROW}<>"")'],
            fill=PROJECT_PLAN_FILL,
            font=Font(color="FFFFFFFF", bold=True),
            stopIfTrue=True,
        ),
        FormulaRule(
            formula=[
                f'AND(${row_type_letter}{FIRST_DATA_ROW - 1}="Project Summary",'
                f'${pa_letter}{FIRST_DATA_ROW}="A",{first_week}{FIRST_DATA_ROW}<>"")'
            ],
            fill=PROJECT_ACTUAL_FILL,
            font=Font(color="FFFFFFFF", bold=True),
            stopIfTrue=True,
        ),
        # WBS level 1: parent WBS (1, 2, 3, ...).
        FormulaRule(
            formula=[
                f'AND(${row_type_letter}{FIRST_DATA_ROW}="WBS",'
                f'${outline_level_letter}{FIRST_DATA_ROW}=1,'
                f'${pa_letter}{FIRST_DATA_ROW}="P",{first_week}{FIRST_DATA_ROW}<>"")'
            ],
            fill=WBS_LEVEL1_PLAN_FILL,
            font=Font(color="FFFFFFFF", bold=True),
            stopIfTrue=True,
        ),
        FormulaRule(
            formula=[
                f'AND(${row_type_letter}{FIRST_DATA_ROW - 1}="WBS",'
                f'${outline_level_letter}{FIRST_DATA_ROW - 1}=1,'
                f'${pa_letter}{FIRST_DATA_ROW}="A",{first_week}{FIRST_DATA_ROW}<>"")'
            ],
            fill=WBS_LEVEL1_ACTUAL_FILL,
            font=Font(color="FFFFFFFF", bold=True),
            stopIfTrue=True,
        ),
        # WBS level 2 and deeper: lighter than the parent WBS.
        FormulaRule(
            formula=[
                f'AND(${row_type_letter}{FIRST_DATA_ROW}="WBS",'
                f'${outline_level_letter}{FIRST_DATA_ROW}>=2,'
                f'${pa_letter}{FIRST_DATA_ROW}="P",{first_week}{FIRST_DATA_ROW}<>"")'
            ],
            fill=WBS_LEVEL2_PLAN_FILL,
            font=Font(color="FFFFFFFF", bold=True),
            stopIfTrue=True,
        ),
        FormulaRule(
            formula=[
                f'AND(${row_type_letter}{FIRST_DATA_ROW - 1}="WBS",'
                f'${outline_level_letter}{FIRST_DATA_ROW - 1}>=2,'
                f'${pa_letter}{FIRST_DATA_ROW}="A",{first_week}{FIRST_DATA_ROW}<>"")'
            ],
            fill=WBS_LEVEL2_ACTUAL_FILL,
            font=Font(color="FFFFFFFF", bold=True),
            stopIfTrue=True,
        ),
    ]

    for rule in rules:
        ws.conditional_formatting.add(target, rule)


def add_activity_input_validation(
    ws,
    plan_rows: list[dict[str, object]],
    timescale_cols: list[int],
) -> None:
    """Limit activity input to 0%-100% per week."""
    validation = DataValidation(
        type="decimal",
        operator="between",
        formula1="0",
        formula2="1",
        allow_blank=True,
    )
    validation.error = "Enter a percentage from 0% to 100%."
    validation.errorTitle = "Invalid Progress"
    validation.prompt = "Enter weekly progress such as 5% or 0.05."
    validation.promptTitle = "Activity Progress"
    validation.showErrorMessage = True
    validation.showInputMessage = True
    ws.add_data_validation(validation)

    start_letter = get_column_letter(timescale_cols[0])
    end_letter = get_column_letter(timescale_cols[-1])
    for row in activity_plan_rows(plan_rows):
        validation.add(f"{start_letter}{row}:{end_letter}{row}")
        validation.add(f"{start_letter}{row + 1}:{end_letter}{row + 1}")


def remove_existing_scurve_rows(ws, row_type_col: int) -> None:
    rows_to_delete: list[int] = []
    for row in range(FIRST_DATA_ROW, ws.max_row + 1):
        if normalize_header(ws.cell(row, row_type_col).value) == "s-curve":
            rows_to_delete.append(row)
    for row in reversed(rows_to_delete):
        ws.delete_rows(row, 1)


def add_scurve_rows(
    ws,
    timescale_cols: list[int],
    header_map: dict[str, int],
    project_plan_row: int,
) -> dict[str, int]:
    row_type_col = header_map["row type"]
    description_col = header_map["description"]
    pa_col = header_map["p/a"]

    start_row = ws.max_row + 2
    labels = [
        ("Plan", "P"),
        ("Acc. Plan", "AP"),
        ("Actual", "A"),
        ("Acc. Actual", "AA"),
    ]
    result: dict[str, int] = {}

    for offset, (label, code) in enumerate(labels):
        row = start_row + offset
        result[label] = row
        ws.cell(row, row_type_col).value = "S-Curve"
        ws.cell(row, description_col).value = label
        ws.cell(row, pa_col).value = code
        ws.cell(row, description_col).font = Font(bold=True)
        ws.cell(row, description_col).alignment = Alignment(horizontal="left")

    project_actual_row = project_plan_row + 1
    first_week_col = timescale_cols[0]

    for col in timescale_cols:
        col_letter = get_column_letter(col)
        plan_row = result["Plan"]
        acc_plan_row = result["Acc. Plan"]
        actual_row = result["Actual"]
        acc_actual_row = result["Acc. Actual"]

        ws.cell(plan_row, col).value = f"={col_letter}{project_plan_row}"
        ws.cell(actual_row, col).value = f"={col_letter}{project_actual_row}"

        # Use cumulative SUM from the first week through the current week.
        # Show blank when the Actual range contains no numbers.
        # This avoids #VALUE! even when the first Actual week is blank.
        first_letter = get_column_letter(first_week_col)

        ws.cell(acc_plan_row, col).value = (
            f'=IF(COUNT(${first_letter}{plan_row}:{col_letter}{plan_row})=0,"",'
            f'SUM(${first_letter}{plan_row}:{col_letter}{plan_row}))'
        )
        last_letter = get_column_letter(timescale_cols[-1])

        ws.cell(acc_actual_row, col).value = (
            # 1) Before the first Actual value: blank.
            # 2) Within the Actual range: show cumulative values.
            # 3) After the latest Actual value: blank.
            f'=IF(COUNT(${first_letter}{actual_row}:{col_letter}{actual_row})=0,"",'
            f'IF(COUNT({col_letter}{actual_row}:${last_letter}{actual_row})=0,"",'
            f'SUM(${first_letter}{actual_row}:{col_letter}{actual_row})))'
        )

        for row in result.values():
            ws.cell(row, col).number_format = PERCENT_FORMAT
            ws.cell(row, col).protection = Protection(locked=True)

        ws.cell(plan_row, col).fill = SCURVE_PLAN_FILL
        ws.cell(acc_plan_row, col).fill = WBS_PLAN_FILL
        ws.cell(actual_row, col).fill = SCURVE_ACTUAL_FILL
        ws.cell(acc_actual_row, col).fill = WBS_ACTUAL_FILL

    for label, row in result.items():
        ws.row_dimensions[row].height = 20
        ws.row_dimensions[row].outlineLevel = 0
        ws.cell(row, description_col).fill = (
            WBS_PLAN_FILL if "Plan" in label else WBS_ACTUAL_FILL
        )
        ws.cell(row, description_col).font = Font(color="FFFFFFFF", bold=True)

    return result


def replace_defined_name(wb, name: str, reference: str) -> None:
    try:
        if name in wb.defined_names:
            del wb.defined_names[name]
    except Exception:
        pass

    defined_name = DefinedName(name, attr_text=reference)
    try:
        wb.defined_names.add(defined_name)
    except AttributeError:
        wb.defined_names.append(defined_name)


def add_named_ranges(
    wb,
    ws,
    timescale_cols: list[int],
    scurve_rows: dict[str, int],
    amount_col: int,
    last_activity_row: int,
) -> None:
    sheet_name = ws.title.replace("'", "''")
    first_week = get_column_letter(timescale_cols[0])
    last_week = get_column_letter(timescale_cols[-1])
    amount_letter = get_column_letter(amount_col)

    refs = {
        "PlanRange": f"'{sheet_name}'!${first_week}${scurve_rows['Plan']}:${last_week}${scurve_rows['Plan']}",
        "AccPlan": f"'{sheet_name}'!${first_week}${scurve_rows['Acc. Plan']}:${last_week}${scurve_rows['Acc. Plan']}",
        "ActualRange": f"'{sheet_name}'!${first_week}${scurve_rows['Actual']}:${last_week}${scurve_rows['Actual']}",
        "AccActual": f"'{sheet_name}'!${first_week}${scurve_rows['Acc. Actual']}:${last_week}${scurve_rows['Acc. Actual']}",
        "AmountRange": f"'{sheet_name}'!${amount_letter}${FIRST_DATA_ROW}:${amount_letter}${last_activity_row}",
    }
    for name, reference in refs.items():
        replace_defined_name(wb, name, reference)


def prepare_progress_and_scurve(wb, ws) -> tuple[int, int, int, int, int]:
    header_map = get_header_map(ws)
    require_columns(
        header_map,
        [
            "row type",
            "description",
            "activity id",
            "p/a",
            "outline level",
            "amount",
            "% complete",
            "total float (hr)",
        ],
    )

    row_type_col = header_map["row type"]
    pa_col = header_map["p/a"]
    outline_level_col = header_map["outline level"]
    amount_col = header_map["amount"]
    activity_id_col = header_map["activity id"]
    percent_complete_col = header_map["% complete"]

    remove_existing_scurve_rows(ws, row_type_col)
    plan_rows = build_plan_rows(ws, row_type_col, pa_col, outline_level_col)
    timescale_cols = find_timescale_columns(ws, header_map)

    with_amount, without_amount = prepare_activity_amount_refs(
        ws, plan_rows, pa_col, amount_col
    )
    wbs_count, project_count = rollup_amounts(
        ws, plan_rows, row_type_col, pa_col, amount_col
    )
    add_progress_formulas(
        ws,
        plan_rows,
        timescale_cols,
        row_type_col,
        activity_id_col,
        pa_col,
        amount_col,
    )
    add_percent_complete_formulas(
        ws,
        plan_rows,
        timescale_cols,
        pa_col,
        percent_complete_col,
    )
    add_actual_amount_formulas(
        ws,
        plan_rows,
        activity_id_col,
        pa_col,
        amount_col,
        percent_complete_col,
    )

    project_plan_row = next(
        int(item["row"])
        for item in plan_rows
        if str(item["row_type"]) == "project summary"
    )
    last_activity_data_row = max(
        int(item["row"]) + (1 if str(item["row_type"]) == "activity" else 0)
        for item in plan_rows
    )

    clear_timescale_direct_fills(
        ws, timescale_cols, last_activity_data_row
    )
    clear_progress_conditional_formatting(ws, timescale_cols)
    add_progress_conditional_formatting(
        ws,
        timescale_cols,
        row_type_col,
        activity_id_col,
        pa_col,
        outline_level_col,
        last_activity_data_row,
    )
    add_percent_complete_warning(
        ws, percent_complete_col, last_activity_data_row
    )
    add_activity_input_validation(ws, plan_rows, timescale_cols)

    scurve_rows = add_scurve_rows(ws, timescale_cols, header_map, project_plan_row)
    add_named_ranges(
        wb, ws, timescale_cols, scurve_rows, amount_col, last_activity_data_row
    )

    ws.freeze_panes = ws.cell(FIRST_DATA_ROW, timescale_cols[0])
    configure_filter_buttons(
        ws,
        header_row=HEADER_ROW,
        last_row=last_activity_data_row,
        last_col=ws.max_column,
        visible_columns={row_type_col, pa_col},
    )

    return with_amount, without_amount, wbs_count, project_count, len(timescale_cols)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="05_prepare_progress_scurve.py",
        description=(
            "Use existing Amount values to create WBS weighted progress, "
            "Plan/Actual styling, validation, and S-Curve summaries."
        ),
    )
    parser.add_argument("--input", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    parser.add_argument("--sheet", default=DEFAULT_SHEET)
    parser.add_argument(
        "--debug",
        action="store_true",
        help="Show the full traceback when an error occurs.",
    )
    return parser


def main() -> int:
    args = build_parser().parse_args()
    input_file = args.input.expanduser().resolve()
    output_file = args.output.expanduser().resolve()

    if not input_file.exists() or input_file.suffix.lower() != ".xlsx":
        print(f"ERROR: Input .xlsx file not found: {input_file}", file=sys.stderr)
        return 1
    if output_file.suffix.lower() != ".xlsx":
        print("ERROR: Output must end with .xlsx.", file=sys.stderr)
        return 1

    try:
        wb = load_workbook(input_file)
        ws = find_sheet(wb, args.sheet)
        with_amount, without_amount, wbs_count, project_count, week_count = (
            prepare_progress_and_scurve(wb, ws)
        )

        configure_incremental_excel_recalculation(wb)
        output_file.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_file)
        wb.close()
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        if args.debug:
            traceback.print_exc()
        return 1

    print(f"SCRIPT VERSION         : {SCRIPT_VERSION}")
    print(f"INPUT                  : {input_file}")
    print(f"OUTPUT                 : {output_file}")
    print(f"ACTIVITIES WITH AMOUNT : {with_amount}")
    print(f"ACTIVITIES NO AMOUNT   : {without_amount}")
    print(f"WBS ROLLUPS            : {wbs_count}")
    print(f"PROJECT ROLLUPS        : {project_count}")
    print(f"WEEKLY COLUMNS         : {week_count}")
    print("S-CURVE ROWS           : Plan / Acc. Plan / Actual / Acc. Actual")
    print("NAMED RANGES           : PlanRange / AccPlan / ActualRange / AccActual / AmountRange")
    if without_amount:
        print("WARNING                : Activities without Amount have no weight in progress calculations.")
    print("Completed.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
