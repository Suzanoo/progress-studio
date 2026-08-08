from __future__ import annotations

from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from progress_studio.infrastructure.excel.calculation_policy import configure_incremental_excel_recalculation
from progress_studio.infrastructure.excel.styles import HEADER_FONT, normalize_argb
from progress_studio.domain.amount import (
    AmountApplicationResult,
    AmountMappingResult,
    AmountSourceDecision,
)

HEADER_ROW = 4
CURRENCY_FORMAT = '#,##0.00'


def normalize_header(value: object) -> str:
    return "" if value is None else str(value).strip().lower()


def normalize_activity_id(value: object) -> str:
    return "" if value is None else str(value).strip().upper()


def to_amount(value: object) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).strip().replace(",", ""))
    except ValueError:
        return None


def header_map(ws, row: int) -> dict[str, int]:
    return {
        normalize_header(ws.cell(row, col).value): col
        for col in range(1, ws.max_column + 1)
        if normalize_header(ws.cell(row, col).value)
    }


def find_header(ws, required: list[str], search_rows: int = 30) -> tuple[int, dict[str, int]]:
    wanted = [normalize_header(item) for item in required]
    for row in range(1, min(ws.max_row, search_rows) + 1):
        found = header_map(ws, row)
        if all(item in found for item in wanted):
            return row, found
    raise ValueError(f"Headers not found: {', '.join(required)} in worksheet '{ws.title}'")


def decide_amount_source(values: list[float | None]) -> AmountSourceDecision:
    use_xml = any(value is not None for value in values)
    return AmountSourceDecision(use_xml, "XML" if use_xml else "PLACEHOLDER")


def collect_schedule_rows(ws) -> list[dict[str, object]]:
    row, headers = find_header(ws, ["Row Type", "Description"])
    row_type_col = headers["row type"]
    description_col = headers["description"]
    activity_col = headers.get("activity id")
    wbs_col = headers.get("wbs")
    amount_col = headers.get("xml amount") or headers.get("amount")
    outline_col = headers.get("outline level")
    pa_col = headers.get("p/a")
    result: list[dict[str, object]] = []
    seen: set[str] = set()
    for r in range(row + 1, ws.max_row + 1):
        row_type = str(ws.cell(r, row_type_col).value or "").strip()
        if not row_type:
            continue
        if pa_col:
            pa = str(ws.cell(r, pa_col).value or "").strip().upper()
            if pa and pa != "P":
                continue
        activity_id = str(ws.cell(r, activity_col).value or "").strip() if activity_col else ""
        if row_type.lower() == "activity":
            if not activity_id or activity_id in seen:
                continue
            seen.add(activity_id)
        raw_amount = ws.cell(r, amount_col).value if amount_col else None
        amount = to_amount(raw_amount)
        result.append({
            "row_type": row_type,
            "wbs": str(ws.cell(r, wbs_col).value or "").strip() if wbs_col else "",
            "description": str(ws.cell(r, description_col).value or "").strip(),
            "activity_id": activity_id,
            "amount": amount,
            "outline_level": int(ws.cell(r, outline_col).value or 0) if outline_col else 0,
        })
    return result


def rebuild_amount_mapping(workbook_path: Path, *, main_sheet: str, mapping_sheet: str, placeholder: float) -> AmountMappingResult:
    wb = load_workbook(workbook_path)
    ws_main = wb[main_sheet]
    rows = collect_schedule_rows(ws_main)
    activities = [r for r in rows if str(r["row_type"]).lower() == "activity"]
    decision = decide_amount_source([r["amount"] for r in activities])
    if mapping_sheet in wb.sheetnames:
        del wb[mapping_sheet]
    ws = wb.create_sheet(mapping_sheet)
    ws.append(["Activity ID", "WBS", "Description", "Amount", "Status"])
    for cell in ws[1]:
        cell.font = Font(bold=True, color=HEADER_FONT)
        cell.fill = PatternFill("solid", fgColor=normalize_argb("4472C4"))
        cell.alignment = Alignment(horizontal="center", vertical="center")
    total = 0.0
    for item in rows:
        row_type = str(item["row_type"]).lower()
        if row_type in {"project summary", "wbs"}:
            ws.append(["", item["wbs"], item["description"], None, "PARENT"])
            for col in range(1, 6):
                ws.cell(ws.max_row, col).fill = PatternFill("solid", fgColor=normalize_argb("E7E6E6"))
                ws.cell(ws.max_row, col).font = Font(bold=True)
            continue
        amount = float(item["amount"] or 0.0) if decision.use_xml_amounts else float(placeholder)
        total += amount
        ws.append([item["activity_id"], item["wbs"], item["description"], amount, decision.source_label])
        ws.cell(ws.max_row, 4).number_format = CURRENCY_FORMAT
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:E{ws.max_row}"
    for col, width in {"A": 18, "B": 22, "C": 60, "D": 18, "E": 16}.items():
        ws.column_dimensions[col].width = width
    configure_incremental_excel_recalculation(wb)
    wb.save(workbook_path)
    wb.close()
    return AmountMappingResult(len(activities), total, decision.source_label)


def read_amounts(ws, activity_header: str = "Activity ID", amount_header: str = "Amount"):
    row, headers = find_header(ws, [activity_header, amount_header])
    ac, mc = headers[normalize_header(activity_header)], headers[normalize_header(amount_header)]
    amounts: dict[str, float] = defaultdict(float)
    used = skipped = 0
    duplicates: list[str] = []
    seen: set[str] = set()
    for r in range(row + 1, ws.max_row + 1):
        aid = normalize_activity_id(ws.cell(r, ac).value)
        amount = to_amount(ws.cell(r, mc).value)
        if not aid or amount is None:
            skipped += 1
            continue
        if aid in seen and aid not in duplicates:
            duplicates.append(aid)
        seen.add(aid)
        amounts[aid] += amount
        used += 1
    return dict(amounts), used, skipped, duplicates


def apply_amount_mapping(input_file: Path, output_file: Path, *, main_sheet: str, mapping_sheet: str) -> AmountApplicationResult:
    wb = load_workbook(input_file, data_only=False)
    ws = wb[main_sheet]
    source_amounts, used, skipped, duplicates = read_amounts(wb[mapping_sheet])
    headers = header_map(ws, HEADER_ROW)
    required = ["row type", "activity id", "p/a", "outline level", "amount"]
    missing_headers = [h for h in required if h not in headers]
    if missing_headers:
        wb.close()
        raise ValueError("Required columns not found: " + ", ".join(missing_headers))
    rtc, aic, pac, olc, amc = [headers[h] for h in required]
    plan_rows: list[dict[str, object]] = []
    for r in range(HEADER_ROW + 1, ws.max_row + 1):
        if str(ws.cell(r, pac).value or "").strip().upper() != "P":
            continue
        rt = normalize_header(ws.cell(r, rtc).value)
        if rt not in {"project summary", "wbs", "activity"}:
            continue
        try:
            level = int(ws.cell(r, olc).value)
        except (TypeError, ValueError):
            continue
        plan_rows.append({"row": r, "row_type": rt, "level": level})
    mapped = wbs_count = project_count = 0
    unmapped: list[str] = []
    for item in plan_rows:
        r, rt = int(item["row"]), str(item["row_type"])
        if rt != "activity":
            continue
        aid = normalize_activity_id(ws.cell(r, aic).value)
        if aid in source_amounts:
            ws.cell(r, amc).value = source_amounts[aid]
            mapped += 1
        elif to_amount(ws.cell(r, amc).value) is None:
            unmapped.append(aid or f"<blank row {r}>")
        if r + 1 <= ws.max_row and str(ws.cell(r + 1, pac).value or "").strip().upper() == "A":
            ws.cell(r + 1, amc).value = None
    amount_letter = get_column_letter(amc)
    row_type_letter = get_column_letter(rtc)
    pa_letter = get_column_letter(pac)
    for index in range(len(plan_rows) - 1, -1, -1):
        item = plan_rows[index]
        r, rt, level = int(item["row"]), str(item["row_type"]), int(item["level"])
        if rt not in {"wbs", "project summary"}:
            continue
        end = r
        for descendant in plan_rows[index + 1:]:
            if int(descendant["level"]) <= level:
                break
            end = int(descendant["row"])
        ws.cell(r, amc).value = 0 if end <= r else (
            f'=SUMIFS(${amount_letter}${r + 1}:${amount_letter}${end},'
            f'${row_type_letter}${r + 1}:${row_type_letter}${end},"Activity",'
            f'${pa_letter}${r + 1}:${pa_letter}${end},"P")'
        )
        wbs_count += rt == "wbs"
        project_count += rt == "project summary"
        if r + 1 <= ws.max_row and str(ws.cell(r + 1, pac).value or "").strip().upper() == "A":
            ws.cell(r + 1, amc).value = None
    for r in range(HEADER_ROW + 1, ws.max_row + 1):
        ws.cell(r, amc).number_format = CURRENCY_FORMAT
    if ws.auto_filter.ref:
        ws.auto_filter.ref = f"A{HEADER_ROW}:{get_column_letter(ws.max_column)}{ws.max_row}"
    map_ws = wb[mapping_sheet]
    _, map_headers = find_header(map_ws, ["Activity ID", "Amount"])
    status_col = map_headers.get("status")
    for r in range(2, map_ws.max_row + 1):
        aid = normalize_activity_id(map_ws.cell(r, map_headers["activity id"]).value)
        if aid in source_amounts:
            map_ws.cell(r, map_headers["amount"]).value = source_amounts[aid]
            if status_col:
                map_ws.cell(r, status_col).value = "Mapped"
        elif status_col:
            map_ws.cell(r, status_col).value = "Ready" if to_amount(map_ws.cell(r, map_headers["amount"]).value) is not None else "Waiting"
    configure_incremental_excel_recalculation(wb)
    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_file)
    wb.close()
    return AmountApplicationResult(mapped, tuple(unmapped), int(wbs_count), int(project_count), used, skipped, tuple(duplicates), f"worksheet {mapping_sheet}")
