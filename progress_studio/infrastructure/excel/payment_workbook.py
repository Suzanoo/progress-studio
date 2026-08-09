from __future__ import annotations

import re
import shutil
import tempfile
from datetime import date, datetime, timedelta
from pathlib import Path, PurePosixPath
from xml.etree import ElementTree as ET
from zipfile import ZIP_DEFLATED, ZipFile, ZipInfo

from progress_studio.domain.payment_models import PaymentSnapshotResult, PaymentWorkbookValidation


MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
DOC_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PKG_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
CONTENT_NS = "http://schemas.openxmlformats.org/package/2006/content-types"

ET.register_namespace("", MAIN_NS)
ET.register_namespace("r", DOC_REL_NS)
ET.register_namespace("", PKG_REL_NS)


class PaymentWorkbookError(ValueError):
    pass


class PaymentWorkbookSnapshotter:
    """Fast package-level snapshot of exported ``main`` into ``Payment``.

    MS-PAY1 intentionally does not recalculate or render anything. Copying the
    worksheet XML keeps the Payment sheet visually identical to main while
    avoiding a full openpyxl load/save of a large exported workbook.
    """

    MAIN_SHEET = "main"
    PAYMENT_SHEET = "Payment"
    ACTIVITY_ID_HEADERS = {"activity id", "activity_id", "activityid", "act id", "act. id"}

    def validate(self, workbook_path: Path) -> PaymentWorkbookValidation:
        path = Path(workbook_path)
        if not path.exists():
            raise PaymentWorkbookError(f"Workbook was not found: {path}")
        if path.suffix.lower() not in {".xlsx", ".xlsm"}:
            raise PaymentWorkbookError("Select a Progress Studio .xlsx or .xlsm workbook.")

        try:
            with ZipFile(path, "r") as package:
                sheet_map = self._sheet_map(package)
                if self.MAIN_SHEET not in sheet_map:
                    raise PaymentWorkbookError("Worksheet 'main' was not found. Select an exported Progress Studio workbook.")
                main_part = sheet_map[self.MAIN_SHEET]
                root = ET.fromstring(package.read(main_part))
                max_row, max_column = self._dimensions(root)
                shared_strings = self._shared_strings(package)
                activity_rows = self._count_activity_rows(root, shared_strings)
                if activity_rows <= 0:
                    raise PaymentWorkbookError("No activity rows were found in the 'main' worksheet.")
                project_start, project_finish = self._project_dates(root, shared_strings)
                default_periods = self._default_periods(project_start, project_finish)
                return PaymentWorkbookValidation(
                    path, self.MAIN_SHEET, activity_rows, max_row, max_column,
                    project_start, project_finish, default_periods,
                )
        except PaymentWorkbookError:
            raise
        except Exception as exc:
            raise PaymentWorkbookError(f"Workbook cannot be opened: {exc}") from exc

    def create_snapshot(self, source: Path, output: Path) -> PaymentSnapshotResult:
        """Create a valid Excel snapshot by copying ``main`` inside the workbook.

        The previous package-level worksheet XML clone could leave two worksheets
        pointing at the same table/drawing relationship parts. Excel repairs that
        situation on open.  For Payment we favour correctness: copy the workbook,
        let openpyxl clone the worksheet objects, and save atomically.
        """
        from openpyxl import load_workbook

        validation = self.validate(source)
        source_path = Path(source)
        output_path = Path(output)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        keep_vba = source_path.suffix.lower() == ".xlsm"
        with tempfile.NamedTemporaryFile(prefix="payment_", suffix=output_path.suffix, dir=output_path.parent, delete=False) as handle:
            temp_path = Path(handle.name)
        replaced = False
        try:
            wb = load_workbook(source_path, keep_vba=keep_vba)
            try:
                if self.PAYMENT_SHEET in wb.sheetnames:
                    wb.remove(wb[self.PAYMENT_SHEET])
                    replaced = True
                main = wb[self.MAIN_SHEET]
                payment = wb.copy_worksheet(main)
                payment.title = self.PAYMENT_SHEET
                # copy_worksheet intentionally omits a few view/filter settings.
                payment.freeze_panes = main.freeze_panes
                payment.sheet_view.showGridLines = main.sheet_view.showGridLines
                payment.auto_filter.ref = main.auto_filter.ref
                wb.save(temp_path)
            finally:
                wb.close()
            shutil.move(str(temp_path), str(output_path))
        except Exception as exc:
            raise PaymentWorkbookError(f"Payment snapshot could not be created: {exc}") from exc
        finally:
            temp_path.unlink(missing_ok=True)

        return PaymentSnapshotResult(source_path, output_path, self.PAYMENT_SHEET, replaced, validation.activity_rows)

    @staticmethod
    def _write_package(src: ZipFile, output: Path, replacements: dict[str, bytes], remove: set[str]) -> None:
        # Write beside the destination then atomically replace, so an interrupted
        # snapshot never leaves the requested output path half-written.
        with tempfile.NamedTemporaryFile(prefix="payment_", suffix=output.suffix, dir=output.parent, delete=False) as handle:
            temp_path = Path(handle.name)
        try:
            with ZipFile(temp_path, "w") as dst:
                existing = set(src.namelist())
                for info in src.infolist():
                    if info.filename in remove:
                        continue
                    data = replacements.get(info.filename, src.read(info.filename))
                    new_info = ZipInfo(info.filename, date_time=info.date_time)
                    new_info.compress_type = info.compress_type
                    new_info.comment = info.comment
                    new_info.extra = info.extra
                    new_info.internal_attr = info.internal_attr
                    new_info.external_attr = info.external_attr
                    new_info.create_system = info.create_system
                    dst.writestr(new_info, data)
                for name, data in replacements.items():
                    if name not in existing:
                        dst.writestr(name, data, compress_type=ZIP_DEFLATED)
            shutil.move(str(temp_path), str(output))
        finally:
            temp_path.unlink(missing_ok=True)

    @staticmethod
    def _sheet_map(package: ZipFile) -> dict[str, str]:
        workbook = ET.fromstring(package.read("xl/workbook.xml"))
        rels = ET.fromstring(package.read("xl/_rels/workbook.xml.rels"))
        rel_map = {
            rel.attrib["Id"]: rel.attrib["Target"]
            for rel in rels.findall(f"{{{PKG_REL_NS}}}Relationship")
        }
        sheets = workbook.find(f"{{{MAIN_NS}}}sheets")
        if sheets is None:
            return {}
        result = {}
        for sheet in sheets:
            rid = sheet.attrib.get(f"{{{DOC_REL_NS}}}id")
            if rid and rid in rel_map:
                result[sheet.attrib["name"]] = PaymentWorkbookSnapshotter._normalise_sheet_target(rel_map[rid])
        return result

    @staticmethod
    def _normalise_sheet_target(target: str) -> str:
        target = target.lstrip("/")
        return target if target.startswith("xl/") else "xl/" + target

    @staticmethod
    def _worksheet_rels_part(sheet_part: str) -> str:
        p = PurePosixPath(sheet_part)
        return str(p.parent / "_rels" / f"{p.name}.rels")

    @staticmethod
    def _next_relationship_id(rels_root: ET.Element) -> str:
        numbers = []
        for rel in rels_root.findall(f"{{{PKG_REL_NS}}}Relationship"):
            match = re.fullmatch(r"rId(\d+)", rel.attrib.get("Id", ""))
            if match:
                numbers.append(int(match.group(1)))
        return f"rId{max(numbers, default=0) + 1}"

    @staticmethod
    def _next_sheet_part(package: ZipFile) -> str:
        numbers = []
        for name in package.namelist():
            match = re.fullmatch(r"xl/worksheets/sheet(\d+)\.xml", name)
            if match:
                numbers.append(int(match.group(1)))
        return f"xl/worksheets/sheet{max(numbers, default=0) + 1}.xml"

    @staticmethod
    def _dimensions(root: ET.Element) -> tuple[int, int]:
        dimension = root.find(f"{{{MAIN_NS}}}dimension")
        ref = dimension.attrib.get("ref", "A1") if dimension is not None else "A1"
        end = ref.split(":")[-1]
        match = re.fullmatch(r"([A-Z]+)(\d+)", end)
        if not match:
            return 0, 0
        col_letters, row_text = match.groups()
        col = 0
        for ch in col_letters:
            col = col * 26 + (ord(ch) - 64)
        return int(row_text), col

    @staticmethod
    def _shared_strings(package: ZipFile) -> list[str]:
        if "xl/sharedStrings.xml" not in package.namelist():
            return []
        root = ET.fromstring(package.read("xl/sharedStrings.xml"))
        strings = []
        for si in root.findall(f"{{{MAIN_NS}}}si"):
            strings.append("".join(node.text or "" for node in si.iter(f"{{{MAIN_NS}}}t")))
        return strings

    def activity_ids(self, workbook_path: Path) -> list[str]:
        path = Path(workbook_path)
        with ZipFile(path, "r") as package:
            sheet_map = self._sheet_map(package)
            if self.MAIN_SHEET not in sheet_map:
                raise PaymentWorkbookError("Worksheet 'main' was not found.")
            root = ET.fromstring(package.read(sheet_map[self.MAIN_SHEET]))
            return self._activity_ids(root, self._shared_strings(package))

    def payment_tree_rows(self, workbook_path: Path) -> list[dict]:
        """Read the lightweight WBS/activity hierarchy and weekly Plan values from main.

        Only Plan-side WBS/Activity rows are returned. WBS rows carry hierarchy context;
        Activity rows additionally carry the weekly incremental Plan values used to
        generate sparse fake payment requirements.
        """
        path = Path(workbook_path)
        with ZipFile(path, "r") as package:
            sheet_map = self._sheet_map(package)
            if self.MAIN_SHEET not in sheet_map:
                raise PaymentWorkbookError("Worksheet 'main' was not found.")
            root = ET.fromstring(package.read(sheet_map[self.MAIN_SHEET]))
            shared = self._shared_strings(package)

        sheet_data = root.find(f"{{{MAIN_NS}}}sheetData")
        if sheet_data is None:
            return []
        rows = list(sheet_data.findall(f"{{{MAIN_NS}}}row"))
        header_row = None
        cols: dict[str, int] = {}
        for row in rows[:30]:
            values = {}
            for cell in row.findall(f"{{{MAIN_NS}}}c"):
                m = re.match(r"([A-Z]+)", cell.attrib.get("r", ""))
                if m:
                    values[self._cell_text(cell, shared).strip().lower()] = self._letters_to_col(m.group(1))
            if "row type" in values and "wbs" in values and "description" in values and "activity id" in values:
                header_row = int(row.attrib.get("r", "0"))
                cols = values
                break
        if header_row is None:
            return []

        date_row = next((r for r in rows if int(r.attrib.get("r", "0")) == header_row), None)
        week_dates: dict[int, date] = {}
        if date_row is not None:
            for cell in date_row.findall(f"{{{MAIN_NS}}}c"):
                m = re.match(r"([A-Z]+)", cell.attrib.get("r", ""))
                if not m:
                    continue
                col = self._letters_to_col(m.group(1))
                d = self._excel_date(cell, shared)
                if d:
                    week_dates[col] = d

        outline_col = cols.get("outline level")
        pa_col = cols.get("p/a")
        result: list[dict] = []
        for row in rows:
            row_num = int(row.attrib.get("r", "0"))
            if row_num <= header_row:
                continue
            cells: dict[int, ET.Element] = {}
            for cell in row.findall(f"{{{MAIN_NS}}}c"):
                m = re.match(r"([A-Z]+)", cell.attrib.get("r", ""))
                if m:
                    cells[self._letters_to_col(m.group(1))] = cell

            row_type = self._cell_text(cells.get(cols["row type"]), shared).strip().lower() if cells.get(cols["row type"]) is not None else ""
            if row_type not in {"wbs", "activity"}:
                continue
            pa = self._cell_text(cells.get(pa_col), shared).strip().upper() if pa_col and cells.get(pa_col) is not None else "P"
            if pa and pa != "P":
                continue

            wbs = self._cell_text(cells.get(cols["wbs"]), shared).strip() if cells.get(cols["wbs"]) is not None else ""
            name = self._cell_text(cells.get(cols["description"]), shared).strip() if cells.get(cols["description"]) is not None else ""
            outline = 0
            if outline_col and cells.get(outline_col) is not None:
                try:
                    outline = int(float(self._cell_text(cells.get(outline_col), shared).strip() or 0))
                except ValueError:
                    outline = 0

            if row_type == "wbs":
                result.append({
                    "row_type": "WBS", "wbs": wbs, "activity_id": "",
                    "activity_name": name, "outline_level": outline, "weekly_plan": (),
                })
                continue

            activity_id = self._cell_text(cells.get(cols["activity id"]), shared).strip() if cells.get(cols["activity id"]) is not None else ""
            if not activity_id:
                continue
            weekly = []
            for col, d in sorted(week_dates.items()):
                cell = cells.get(col)
                if cell is None:
                    continue
                raw = self._cell_text(cell, shared).strip()
                if not raw:
                    continue
                try:
                    value = float(raw)
                except ValueError:
                    continue
                if value != 0:
                    weekly.append((d, value))
            result.append({
                "row_type": "ACT", "wbs": wbs, "activity_id": activity_id,
                "activity_name": name, "outline_level": outline, "weekly_plan": tuple(weekly),
            })
        return result

    def activity_plan_profiles(self, workbook_path: Path) -> list[dict]:
        """Backward-compatible activity-only view used by existing payment code/tests."""
        return [row for row in self.payment_tree_rows(workbook_path) if row["row_type"] == "ACT"]

    @staticmethod
    def _letters_to_col(letters: str) -> int:
        value = 0
        for ch in letters:
            value = value * 26 + ord(ch) - 64
        return value

    @classmethod
    def _count_activity_rows(cls, root: ET.Element, shared_strings: list[str]) -> int:
        return len(cls._activity_ids(root, shared_strings))

    @classmethod
    def _activity_ids(cls, root: ET.Element, shared_strings: list[str]) -> list[str]:
        sheet_data = root.find(f"{{{MAIN_NS}}}sheetData")
        if sheet_data is None:
            return []
        activity_col = None
        header_row = None
        for row in sheet_data.findall(f"{{{MAIN_NS}}}row"):
            row_num = int(row.attrib.get("r", "0"))
            if row_num > 30:
                break
            for cell in row.findall(f"{{{MAIN_NS}}}c"):
                value = cls._cell_text(cell, shared_strings).strip().lower()
                if value in cls.ACTIVITY_ID_HEADERS:
                    match = re.match(r"[A-Z]+", cell.attrib.get("r", ""))
                    if match:
                        activity_col = match[0]
                        header_row = row_num
                    break
            if activity_col:
                break
        if not activity_col or header_row is None:
            return []

        activity_ids: list[str] = []
        seen: set[str] = set()
        for row in sheet_data.findall(f"{{{MAIN_NS}}}row"):
            row_num = int(row.attrib.get("r", "0"))
            if row_num <= header_row:
                continue
            for cell in row.findall(f"{{{MAIN_NS}}}c"):
                ref = cell.attrib.get("r", "")
                match = re.match(r"[A-Z]+", ref)
                if match and match[0] == activity_col:
                    text = cls._cell_text(cell, shared_strings).strip()
                    if text and text not in seen:
                        seen.add(text)
                        activity_ids.append(text)
                    break
        return activity_ids


    @classmethod
    def _project_dates(cls, root: ET.Element, shared_strings: list[str]) -> tuple[date | None, date | None]:
        sheet_data = root.find(f"{{{MAIN_NS}}}sheetData")
        if sheet_data is None:
            return None, None

        header_row = None
        columns: dict[str, str] = {}
        for row in sheet_data.findall(f"{{{MAIN_NS}}}row"):
            row_num = int(row.attrib.get("r", "0"))
            if row_num > 30:
                break
            values: dict[str, str] = {}
            for cell in row.findall(f"{{{MAIN_NS}}}c"):
                match = re.match(r"[A-Z]+", cell.attrib.get("r", ""))
                if not match:
                    continue
                values[cls._cell_text(cell, shared_strings).strip().lower()] = match[0]
            if "row type" in values and "plan start" in values and "plan finish" in values:
                header_row = row_num
                columns = values
                break
        if header_row is None:
            return None, None

        start_values: list[date] = []
        finish_values: list[date] = []
        for row in sheet_data.findall(f"{{{MAIN_NS}}}row"):
            row_num = int(row.attrib.get("r", "0"))
            if row_num <= header_row:
                continue
            cells = {
                re.match(r"[A-Z]+", cell.attrib.get("r", ""))[0]: cell
                for cell in row.findall(f"{{{MAIN_NS}}}c")
                if re.match(r"[A-Z]+", cell.attrib.get("r", ""))
            }
            row_type_cell = cells.get(columns["row type"])
            row_type = cls._cell_text(row_type_cell, shared_strings).strip().lower() if row_type_cell is not None else ""
            if row_type not in {"project summary", "activity"}:
                continue
            start = cls._excel_date(cells.get(columns["plan start"]), shared_strings)
            finish = cls._excel_date(cells.get(columns["plan finish"]), shared_strings)
            if start:
                start_values.append(start)
            if finish:
                finish_values.append(finish)
            if row_type == "project summary" and start and finish:
                return start, finish
        return (min(start_values) if start_values else None, max(finish_values) if finish_values else None)

    @classmethod
    def _excel_date(cls, cell: ET.Element | None, shared_strings: list[str]) -> date | None:
        if cell is None:
            return None
        text = cls._cell_text(cell, shared_strings).strip()
        if not text:
            return None
        try:
            serial = float(text)
        except ValueError:
            try:
                return datetime.fromisoformat(text).date()
            except ValueError:
                return None
        return (datetime(1899, 12, 30) + timedelta(days=serial)).date()

    @staticmethod
    def _default_periods(start: date | None, finish: date | None) -> int:
        if start is None or finish is None or finish <= start:
            return 1
        return max((finish.year - start.year) * 12 + finish.month - start.month, 1)

    @staticmethod
    def _cell_text(cell: ET.Element, shared_strings: list[str]) -> str:
        cell_type = cell.attrib.get("t")
        if cell_type == "inlineStr":
            inline = cell.find(f"{{{MAIN_NS}}}is")
            if inline is None:
                return ""
            return "".join(node.text or "" for node in inline.iter(f"{{{MAIN_NS}}}t"))
        value = cell.find(f"{{{MAIN_NS}}}v")
        if value is None or value.text is None:
            return ""
        if cell_type == "s":
            try:
                return shared_strings[int(value.text)]
            except (ValueError, IndexError):
                return ""
        return value.text
