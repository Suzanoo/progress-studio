from __future__ import annotations

from datetime import date, datetime, timedelta
import json
import re
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.axis import DateAxis
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.image import Image as XLImage
from openpyxl.formatting.rule import DataBarRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from progress_studio.infrastructure.excel.calculation_policy import configure_incremental_excel_recalculation

DASHBOARD_SHEET = "Dashboard"
DATA_SHEET = "Dashboard_Data"
PROGRESS_SHEET = "progress"
TABLE_SHEET = "progress_table"

NAVY = "17365D"
BLUE = "2F75B5"
GREEN = "70AD47"
RED = "C00000"
AMBER = "BF9000"
LIGHT_BLUE = "EAF2F8"
LIGHT_GREEN = "EAF4E3"
LIGHT_RED = "FCE8E6"
LIGHT_AMBER = "FFF4D6"
LIGHT_GRAY = "F3F4F6"
BORDER = "D9DEE7"
TEXT = "1F2937"
MUTED = "667085"
WHITE = "FFFFFF"


def _load_dashboard_theme() -> dict:
    path = Path(__file__).resolve().parents[2] / "config" / "dashboard_theme.json"
    defaults = {
        "font": "Aptos",
        "colors": {
            "navy": NAVY, "blue": BLUE, "green": GREEN, "red": RED,
            "amber": AMBER, "light_blue": LIGHT_BLUE, "light_green": LIGHT_GREEN,
            "light_red": LIGHT_RED, "light_amber": LIGHT_AMBER,
            "light_gray": LIGHT_GRAY, "border": BORDER, "text": TEXT,
            "muted": MUTED, "white": WHITE,
        },
        "layout": {"title": "PROGRESS STUDIO DASHBOARD", "default_view": "Monthly", "chart_height": 8.0, "chart_width": 20.5, "chart_y_max": 1.10},
        "icons": {"enabled": True, "size": 32, "planned": "planned.png", "actual": "actual.png", "schedule": "schedule.png", "time_impact": "time_impact.png"},
    }
    try:
        loaded = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, ValueError, TypeError):
        return defaults
    result = defaults.copy()
    result["colors"] = {**defaults["colors"], **loaded.get("colors", {})}
    result["layout"] = {**defaults["layout"], **loaded.get("layout", {})}
    result["icons"] = {**defaults["icons"], **loaded.get("icons", {})}
    result["font"] = loaded.get("font", defaults["font"])
    return result


_THEME = _load_dashboard_theme()
_FONT = _THEME["font"]
_COLORS = _THEME["colors"]
_LAYOUT = _THEME["layout"]
_ICONS = _THEME["icons"]
NAVY = _COLORS["navy"]
BLUE = _COLORS["blue"]
GREEN = _COLORS["green"]
RED = _COLORS["red"]
AMBER = _COLORS["amber"]
LIGHT_BLUE = _COLORS["light_blue"]
LIGHT_GREEN = _COLORS["light_green"]
LIGHT_RED = _COLORS["light_red"]
LIGHT_AMBER = _COLORS["light_amber"]
LIGHT_GRAY = _COLORS["light_gray"]
BORDER = _COLORS["border"]
TEXT = _COLORS["text"]
MUTED = _COLORS["muted"]
WHITE = _COLORS["white"]


def _solid(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def _thin_border() -> Border:
    side = Side(style="thin", color=BORDER)
    return Border(left=side, right=side, top=side, bottom=side)


def _as_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
    return None


def _as_number(value):
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).strip().replace(",", ""))
    except (TypeError, ValueError):
        return None


def _normalise_header(value) -> str:
    return "".join(ch for ch in str(value or "").strip().lower() if ch.isalnum())


def _progress_columns(progress_ws) -> dict[str, int]:
    aliases = {
        "project_start": {"projectstart"},
        "project_finish": {"projectfinish"},
        "date": {"weekstart", "weeklydate", "date", "period", "week"},
        "plan": {"plan", "weeklyplan", "planned", "plannedprogress"},
        "actual": {"actual", "weeklyactual", "actualprogress"},
    }
    result: dict[str, int] = {}
    for col in range(1, progress_ws.max_column + 1):
        key = _normalise_header(progress_ws.cell(1, col).value)
        for role, names in aliases.items():
            if key in names and role not in result:
                result[role] = col
    missing = [role for role in ("date", "plan", "actual") if role not in result]
    if missing:
        headers = [str(progress_ws.cell(1, c).value or "") for c in range(1, progress_ws.max_column + 1)]
        raise RuntimeError(
            "Dashboard data source is missing required progress columns "
            f"{missing}. Found headers: {headers}"
        )
    return result



def _remove(workbook, name: str) -> None:
    if name in workbook.sheetnames:
        del workbook[name]


def _resolve_reference(workbook, value):
    if not isinstance(value, str) or not value.startswith("="):
        return value
    match = re.fullmatch(r"='?([^'!]+)'?!\$?([A-Z]+)\$?(\d+)", value.strip())
    if not match:
        return value
    sheet_name, column, row = match.groups()
    if sheet_name not in workbook.sheetnames:
        return value
    return workbook[sheet_name][f"{column}{row}"].value


def _progress_rows(workbook, progress_ws) -> list[tuple[int, date]]:
    """Return reporting periods that overlap the real project window only.

    ``progress`` intentionally keeps the visible +/- timescale margin from ``main``.
    Dashboard reporting must not treat those display-only rows as calculation or
    chart periods.  A weekly cutoff is retained when its seven-day reporting
    interval overlaps Project Start..Project Finish; this also preserves the final
    reporting week when Project Finish falls before that week's cutoff date.
    """
    columns = _progress_columns(progress_ws)

    project_start = project_finish = None
    project_start_col = columns.get("project_start")
    project_finish_col = columns.get("project_finish")
    if project_start_col is not None and project_finish_col is not None:
        for row in range(2, progress_ws.max_row + 1):
            if project_start is None:
                project_start = _as_date(
                    _resolve_reference(workbook, progress_ws.cell(row, project_start_col).value)
                )
            if project_finish is None:
                project_finish = _as_date(
                    _resolve_reference(workbook, progress_ws.cell(row, project_finish_col).value)
                )
            if project_start is not None and project_finish is not None:
                break

    result: list[tuple[int, date]] = []
    for row in range(2, progress_ws.max_row + 1):
        raw_date = _resolve_reference(workbook, progress_ws.cell(row, columns["date"]).value)
        week_date = _as_date(raw_date)
        if week_date is None:
            continue
        if project_start is not None and project_finish is not None:
            period_start = week_date - timedelta(days=6)
            if week_date < project_start or period_start > project_finish:
                continue
        result.append((row, week_date))
    if not result:
        raise RuntimeError(
            "Dashboard generation failed: progress sheet contains no usable project reporting dates."
        )
    return result


def _monthly_groups(rows: list[tuple[int, date]]) -> list[tuple[date, list[int]]]:
    """Group weekly reporting rows by month using the last real cutoff date.

    Monthly reporting must stay on the same reporting calendar as Weekly.  Using
    a synthetic calendar month-end (for example 31-Aug when reporting is every
    Friday) creates a cutoff date that has no source data.  The monthly period is
    therefore represented by the final available weekly cutoff inside that month.
    """
    groups: dict[tuple[int, int], list[tuple[int, date]]] = {}
    for row, value in rows:
        groups.setdefault((value.year, value.month), []).append((row, value))
    return [
        (entries[-1][1], [row for row, _ in entries])
        for _, entries in sorted(groups.items())
    ]


def _progress_percent_value(progress_ws, row: int, col: int) -> float | None:
    """Return progress 0..100 percent-points as Excel chart fraction 0..1."""
    numeric = _as_number(progress_ws.cell(row, col).value)
    if numeric is None:
        return None
    return numeric / 100.0


def _source_percent_formula(progress_ws, row: int, col: int) -> str | float | None:
    value = _progress_percent_value(progress_ws, row, col)
    if value is not None:
        return value
    column_letter = progress_ws.cell(1, col).column_letter
    return f'''=IF('{PROGRESS_SHEET}'!{column_letter}{row}="","",'{PROGRESS_SHEET}'!{column_letter}{row}/100)'''


def _build_data_sheet(workbook, progress_ws) -> None:
    _remove(workbook, DATA_SHEET)
    ws = workbook.create_sheet(DATA_SHEET)
    columns = _progress_columns(progress_ws)
    rows = _progress_rows(workbook, progress_ws)
    months = _monthly_groups(rows)

    ws.append(["Weekly Date", "Weekly Plan", "Weekly Actual", "Monthly Date", "Monthly Plan", "Monthly Actual", "Period", "Plan", "Actual"])
    for output_row, (source_row, week_date) in enumerate(rows, start=2):
        ws.cell(output_row, 1, week_date)
        ws.cell(output_row, 2, _source_percent_formula(progress_ws, source_row, columns["plan"]))
        ws.cell(output_row, 3, _source_percent_formula(progress_ws, source_row, columns["actual"]))
        ws.cell(output_row, 1).number_format = "dd/mm/yyyy"
        ws.cell(output_row, 2).number_format = "0.00%"
        ws.cell(output_row, 3).number_format = "0.00%"

    for output_row, (month_date, source_rows) in enumerate(months, start=2):
        last_row = source_rows[-1]
        ws.cell(output_row, 4, month_date)
        # progress is cumulative; monthly Plan is simply the last weekly cutoff
        # inside that month.
        ws.cell(
            output_row,
            5,
            _source_percent_formula(progress_ws, last_row, columns["plan"]),
        )

        # Actual is cumulative and remains live in the hybrid progress contract.
        # Use one contiguous progress range for the month, so Excel can pick the
        # last nonblank weekly Actual without the old multi-area LOOKUP error.
        actual_letter = progress_ws.cell(1, columns["actual"]).column_letter
        first_actual_row = source_rows[0]
        last_actual_row = source_rows[-1]
        actual_range = (
            f"'{PROGRESS_SHEET}'!{actual_letter}{first_actual_row}:"
            f"{actual_letter}{last_actual_row}"
        )
        ws.cell(
            output_row,
            6,
            f'=IF(COUNT({actual_range})=0,"",'
            f'LOOKUP(2,1/({actual_range}<>""),{actual_range})/100)',
        )
        ws.cell(output_row, 4).number_format = "mmm-yyyy"
        ws.cell(output_row, 5).number_format = "0.00%"
        ws.cell(output_row, 6).number_format = "0.00%"

    display_count = max(len(rows), len(months))
    for output_row in range(2, display_count + 2):
        weekly_row = output_row if output_row <= len(rows) + 1 else None
        monthly_row = output_row if output_row <= len(months) + 1 else None
        weekly_date = f"A{weekly_row}" if weekly_row else '""'
        weekly_plan = f"B{weekly_row}" if weekly_row else '""'
        weekly_actual = f"C{weekly_row}" if weekly_row else '""'
        monthly_date = f"D{monthly_row}" if monthly_row else '""'
        monthly_plan = f"E{monthly_row}" if monthly_row else '""'
        monthly_actual = f"F{monthly_row}" if monthly_row else '""'
        ws.cell(output_row, 7, f'=IF(Dashboard!$G$5="Weekly",{weekly_date},{monthly_date})')
        # Chart helpers use #N/A outside their valid series range.  Excel treats
        # formula-empty strings as zero in some chart modes, which previously made
        # Plan plunge from 100% to 0% at the first display-margin period.
        ws.cell(
            output_row,
            8,
            f'=IF(G{output_row}="",NA(),IF(Dashboard!$G$5="Weekly",'
            f'IF({weekly_plan}="",NA(),{weekly_plan}),IF({monthly_plan}="",NA(),{monthly_plan})))',
        )
        # Actual is a renderer-only cutoff series: show the cumulative value (or
        # zero before progress starts) through the selected cutoff, then stop.
        ws.cell(
            output_row,
            9,
            f'=IF(OR(G{output_row}="",G{output_row}>Dashboard!$K$5),NA(),'
            f'IF(Dashboard!$G$5="Weekly",IF({weekly_actual}="",0,{weekly_actual}),'
            f'IF({monthly_actual}="",0,{monthly_actual})))',
        )
        ws.cell(output_row, 7).number_format = "dd/mm/yyyy"
        ws.cell(output_row, 8).number_format = "0.00%"
        ws.cell(output_row, 9).number_format = "0.00%"

    # Dedicated validation lists keep the Cutoff Date dropdown aligned with the
    # selected reporting view.  Monthly is intentionally a subset of Weekly so
    # every monthly cutoff is also a valid weekly reporting date.
    ws["J1"] = "Weekly Cutoff"
    ws["K1"] = "Monthly Cutoff"
    for output_row, (_, week_date) in enumerate(rows, start=2):
        ws.cell(output_row, 10, week_date)
        ws.cell(output_row, 10).number_format = "dd/mm/yyyy"
    for output_row, (month_date, _) in enumerate(months, start=2):
        ws.cell(output_row, 11, month_date)
        ws.cell(output_row, 11).number_format = "dd/mm/yyyy"
    ws.column_dimensions["J"].hidden = True
    ws.column_dimensions["K"].hidden = True

    if ws.max_row < 2 or ws["A2"].value is None:
        raise RuntimeError("Dashboard generation failed: Dashboard_Data contains no weekly rows.")
    ws.sheet_state = "hidden"


def _merge_title(ws, cell_range: str, value: str, size: int = 12) -> None:
    ws.merge_cells(cell_range)
    cell = ws[cell_range.split(":")[0]]
    cell.value = value
    cell.font = Font(name=_FONT, bold=True, size=size, color=TEXT)
    cell.alignment = Alignment(vertical="center")


def _style_box(ws, cell_range: str, fill: str = WHITE) -> None:
    for row in ws[cell_range]:
        for cell in row:
            cell.fill = _solid(fill)
            cell.border = _thin_border()


def _add_kpi_icon(ws, anchor: str, icon_name: str) -> None:
    if not bool(_ICONS.get("enabled", True)):
        return
    filename = _ICONS.get(icon_name)
    if not filename:
        return
    icon_path = Path(__file__).resolve().parents[2] / "assets" / "dashboard" / "icons" / str(filename)
    if not icon_path.exists():
        return
    image = XLImage(str(icon_path))
    size = int(_ICONS.get("size", 32))
    image.width = size
    image.height = size
    ws.add_image(image, anchor)


def _kpi(ws, title_range: str, value_range: str, title: str, formula: str, fill: str, color: str, number_format: str = "0.00%", *, icon: str | None = None) -> None:
    _style_box(ws, f"{title_range.split(':')[0]}:{value_range.split(':')[1]}", fill)
    ws.merge_cells(title_range)
    ws.merge_cells(value_range)
    title_cell = ws[title_range.split(":")[0]]
    value_cell = ws[value_range.split(":")[0]]
    title_cell.value = title
    title_cell.font = Font(name=_FONT, size=9, color=MUTED)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    value_cell.value = formula
    value_cell.font = Font(name=_FONT, size=18, bold=True, color=color)
    value_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    value_cell.number_format = number_format
    if icon:
        _add_kpi_icon(ws, value_range.split(":")[0], icon)




def _date_axis_for_line_chart(chart: LineChart, *, title: str) -> DateAxis:
    """Return a DateAxis with reciprocal axis ids valid for Excel OOXML.

    LineChart starts with horizontal axId=10 and value axId=100.  A fresh
    openpyxl DateAxis uses axId=500 but the existing value axis still points
    at 10, which serializes a dangling ``crossAx`` reference and makes Excel
    repair the chart on open.  Keep the original LineChart axis ids and bind
    both axes explicitly.
    """
    date_axis = DateAxis()
    date_axis.axId = 10
    date_axis.crossAx = 100
    date_axis.number_format = "mmm-yy"
    date_axis.majorTimeUnit = "days"
    date_axis.title = title

    chart.y_axis.axId = 100
    chart.y_axis.crossAx = 10
    chart.x_axis = date_axis
    return date_axis

def _build_dashboard_sheet(workbook, project_name: str | None = None) -> None:
    _remove(workbook, DASHBOARD_SHEET)
    ws = workbook.create_sheet(DASHBOARD_SHEET, 0)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A7"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.outlinePr.summaryBelow = True

    # Simple, wide dashboard canvas. Four KPI cards fit on one line.
    widths = {
        "A": 3, "B": 14, "C": 14, "D": 14,
        "E": 14, "F": 14, "G": 14,
        "H": 14, "I": 14, "J": 14,
        "K": 14, "L": 14, "M": 14,
        "N": 11, "O": 11, "P": 12, "Q": 12,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for row in range(1, 90):
        ws.row_dimensions[row].height = 20

    ws.merge_cells("B2:M2")
    ws["B2"] = _LAYOUT["title"]
    ws["B2"].font = Font(name=_FONT, size=20, bold=True, color=NAVY)
    ws["B2"].alignment = Alignment(vertical="center")
    ws.row_dimensions[2].height = 32

    _merge_title(ws, "B4:M4", "PROJECT INFORMATION", 11)
    _style_box(ws, "B5:M6", LIGHT_GRAY)
    ws["B5"] = "Project"
    ws.merge_cells("C5:D5")
    ws["C5"] = project_name or "Project"
    ws["C5"].font = Font(name=_FONT, bold=True, color=TEXT)

    ws["F5"] = "View"
    ws.merge_cells("G5:H5")
    default_view = str(_LAYOUT.get("default_view", "Monthly")).title()
    if default_view not in {"Weekly", "Monthly"}:
        default_view = "Monthly"
    ws["G5"] = default_view
    ws["J5"] = "Cutoff Date"
    ws.merge_cells("K5:M5")
    data_ws = workbook[DATA_SHEET]
    weekly_cutoffs = [data_ws.cell(row, 10).value for row in range(2, data_ws.max_row + 1) if data_ws.cell(row, 10).value]
    monthly_cutoffs = [data_ws.cell(row, 11).value for row in range(2, data_ws.max_row + 1) if data_ws.cell(row, 11).value]
    initial_cutoffs = monthly_cutoffs if default_view == "Monthly" else weekly_cutoffs
    ws["K5"] = initial_cutoffs[-1] if initial_cutoffs else (weekly_cutoffs[-1] if weekly_cutoffs else None)
    ws["K5"].number_format = "dd/mm/yyyy"
    ws["G5"].font = Font(name=_FONT, bold=True, color=NAVY)
    ws["K5"].font = Font(name=_FONT, bold=True, color=NAVY)
    ws["G5"].alignment = Alignment(horizontal="center")
    ws["K5"].alignment = Alignment(horizontal="center")

    ws["B6"] = "Data source"
    ws.merge_cells("C6:H6")
    ws["C6"] = "Project live from progress • Activity snapshot from progress_table"
    ws["C6"].font = Font(name=_FONT, size=9, color=MUTED)
    ws["J6"] = "Chart rule"
    ws.merge_cells("K6:M6")
    ws["K6"] = "Plan: full baseline  |  Actual: to cutoff"
    ws["K6"].font = Font(name=_FONT, size=9, color=MUTED)

    view_validation = DataValidation(type="list", formula1='"Weekly,Monthly"', allow_blank=False)
    view_validation.error = "Choose Weekly or Monthly."
    view_validation.errorTitle = "Invalid reporting view"
    view_validation.prompt = "Switch the dashboard chart between Weekly and Monthly."
    view_validation.promptTitle = "Reporting View"
    view_validation.showInputMessage = True
    view_validation.showErrorMessage = True
    ws.add_data_validation(view_validation)
    view_validation.add(ws["G5"])

    last_week_row = max(2, len(weekly_cutoffs) + 1)
    last_month_row = max(2, len(monthly_cutoffs) + 1)
    cutoff_validation = DataValidation(
        type="list",
        # Excel data validation cannot directly use a cross-sheet range chosen
        # by IF. INDIRECT keeps the dropdown dynamic without VBA/macros.
        formula1=(
            f'=INDIRECT(IF($G$5="Weekly","{DATA_SHEET}!$J$2:$J${last_week_row}",'
            f'"{DATA_SHEET}!$K$2:$K${last_month_row}"))'
        ),
        allow_blank=False,
    )
    cutoff_validation.prompt = "Select the reporting cutoff date. KPI and Actual use this date."
    cutoff_validation.promptTitle = "Cutoff Date"
    cutoff_validation.showInputMessage = True
    ws.add_data_validation(cutoff_validation)
    cutoff_validation.add(ws["K5"])

    # Dashboard contract: S-Curve/KPIs come from the stable ``progress`` snapshot.
    # Dashboard_Data is only a thin Weekly/Monthly view adapter. Activity tables
    # continue to read ``progress_table`` directly.
    _merge_title(ws, "B8:M8", "KPI SUMMARY", 11)
    progress_ws = workbook[PROGRESS_SHEET]
    progress_cols = _progress_columns(progress_ws)
    progress_last_row = max(2, progress_ws.max_row)
    date_letter = progress_ws.cell(1, progress_cols["date"]).column_letter
    plan_letter = progress_ws.cell(1, progress_cols["plan"]).column_letter
    actual_letter = progress_ws.cell(1, progress_cols["actual"]).column_letter
    progress_ref = f"'{PROGRESS_SHEET}'"
    plan_kpi = (
        f'=IFERROR(LOOKUP(2,1/(({progress_ref}!${date_letter}$2:${date_letter}${progress_last_row}<=$K$5)*'
        f'({progress_ref}!${plan_letter}$2:${plan_letter}${progress_last_row}<>"")),'
        f'{progress_ref}!${plan_letter}$2:${plan_letter}${progress_last_row})/100,0)'
    )
    actual_kpi = (
        f'=IFERROR(LOOKUP(2,1/(({progress_ref}!${date_letter}$2:${date_letter}${progress_last_row}<=$K$5)*'
        f'({progress_ref}!${actual_letter}$2:${actual_letter}${progress_last_row}<>"")),'
        f'{progress_ref}!${actual_letter}$2:${actual_letter}${progress_last_row})/100,0)'
    )
    schedule_formula = (
        '=IF(E10=B10,"ON SCHEDULE",'
        'IF(E10<B10,"DELAY "&TEXT(B10-E10,"0.00%"),'
        '"AHEAD "&TEXT(E10-B10,"0.00%")))'
    )
    # progress!A/B are the authoritative project baseline dates. Convert KPI 3
    # variance to calendar days across the fixed baseline duration and round to
    # whole days.
    time_impact_formula = (
        f'=ROUND(ABS(E10-B10)*MAX(0,{progress_ref}!$B$2-{progress_ref}!$A$2),0)&" Days"'
    )
    _kpi(ws, "B9:D9", "B10:D12", "PLANNED PROGRESS", plan_kpi, LIGHT_BLUE, BLUE, icon="planned")
    _kpi(ws, "E9:G9", "E10:G12", "ACTUAL PROGRESS", actual_kpi, LIGHT_GREEN, GREEN, icon="actual")
    _kpi(ws, "H9:J9", "H10:J12", "SCHEDULE STATUS", schedule_formula, LIGHT_AMBER, AMBER, "General", icon="schedule")
    _kpi(ws, "K9:M9", "K10:M12", "TIME IMPACT", time_impact_formula, LIGHT_RED, RED, "General", icon="time_impact")

    # Schedule/Time Impact cards follow the KPI 3 condition visually.
    for target in ("H9:J12", "K9:M12"):
        ws.conditional_formatting.add(
            target,
            FormulaRule(formula=['$E$10<$B$10'], fill=_solid(LIGHT_RED), font=Font(name=_FONT, bold=True, color=RED)),
        )
        ws.conditional_formatting.add(
            target,
            FormulaRule(formula=['$E$10>$B$10'], fill=_solid(LIGHT_GREEN), font=Font(name=_FONT, bold=True, color=GREEN)),
        )
        ws.conditional_formatting.add(
            target,
            FormulaRule(formula=['$E$10=$B$10'], fill=_solid(LIGHT_BLUE), font=Font(name=_FONT, bold=True, color=BLUE)),
        )
    for row in range(9, 13):
        ws.row_dimensions[row].height = 24
    ws.row_dimensions[10].height = 28
    ws.row_dimensions[11].height = 28
    ws.row_dimensions[12].height = 28

    # S-Curve: Plan is full baseline; Actual stops at selected cutoff.
    _merge_title(ws, "B15:M15", "S-CURVE — PLAN VS ACTUAL", 11)
    _style_box(ws, "B16:M34", WHITE)
    chart = LineChart()
    chart.title = None
    chart.style = int(_LAYOUT.get("chart_style", 2))
    chart.height = float(_LAYOUT["chart_height"])
    chart.width = float(_LAYOUT["chart_width"])
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = float(_LAYOUT.get("chart_y_max", 1.10))
    chart.y_axis.majorUnit = float(_LAYOUT.get("chart_major_unit", 0.2))
    chart.y_axis.numFmt = "0%"
    chart.y_axis.title = "Progress"
    # A date axis is essential because the selector helper has Weekly capacity
    # even when Monthly is selected. Blank trailing Monthly helper rows must not
    # consume equal-width text categories and squeeze the curve to the left.
    _date_axis_for_line_chart(chart, title="Period")
    chart.legend.position = "t"
    chart.display_blanks = "gap"

    # Make the plotting area lighter and cleaner than the default Excel chart.
    grid = GraphicalProperties()
    grid.line.solidFill = _COLORS.get("chart_grid", "E5E7EB")
    grid.line.width = 9000
    chart.y_axis.majorGridlines.graphicalProperties = grid
    axis_line = GraphicalProperties()
    axis_line.line.solidFill = _COLORS.get("chart_axis", "B8C2CC")
    axis_line.line.width = 9000
    chart.y_axis.spPr = axis_line
    chart.x_axis.spPr = axis_line

    max_rows = min(250, workbook[DATA_SHEET].max_row)
    data = Reference(workbook[DATA_SHEET], min_col=8, max_col=9, min_row=1, max_row=max_rows)
    cats = Reference(workbook[DATA_SHEET], min_col=7, min_row=2, max_row=max_rows)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    if len(chart.series) >= 2:
        chart.series[0].graphicalProperties.line.solidFill = BLUE
        chart.series[0].graphicalProperties.line.width = int(_LAYOUT.get("plan_line_width", 26000))
        chart.series[1].graphicalProperties.line.solidFill = GREEN
        chart.series[1].graphicalProperties.line.width = int(_LAYOUT.get("actual_line_width", 26000))
        if bool(_LAYOUT.get("chart_markers", False)):
            chart.series[0].marker.symbol = "circle"
            chart.series[0].marker.size = int(_LAYOUT.get("marker_size", 4))
            chart.series[1].marker.symbol = "circle"
            chart.series[1].marker.size = int(_LAYOUT.get("marker_size", 4))
    ws.add_chart(chart, "B16")

    ws.merge_cells("B35:M35")
    ws["B35"] = "Plan curve = full baseline duration    •    Actual curve = selected cutoff date"
    ws["B35"].font = Font(name=_FONT, size=9, italic=True, color=MUTED)
    ws["B35"].alignment = Alignment(horizontal="left", vertical="center")

    # Activity Progress keeps the OKD 2-row Plan/Actual contract.
    # Status filtering is native Excel AutoFilter on the Status column only.
    # Both Plan and Actual rows carry the same filter value so a pair never
    # splits when the user filters Behind / On Track / Complete / Not Started.
    _merge_title(ws, "B37:Q37", "ACTIVITY PROGRESS", 11)

    headers = [
        "WBS", "Activity", "Type", "Total",
        "Amount", "Progress", "Variance", "Status",
    ]
    starts = ["B", "C", "F", "H", "J", "L", "N", "P"]
    ends = ["B", "E", "G", "I", "K", "M", "O", "Q"]
    for start_col, end_col, header in zip(starts, ends, headers):
        ws.merge_cells(f"{start_col}38:{end_col}38")
        cell = ws[f"{start_col}38"]
        cell.value = header
        cell.fill = _solid(NAVY)
        cell.font = Font(name=_FONT, bold=True, color=WHITE, size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        for row_cells in ws[f"{start_col}38:{end_col}38"]:
            for c in row_cells:
                c.border = _thin_border()

    source = workbook[TABLE_SHEET]
    kind_col = None
    for col in range(1, source.max_column + 1):
        if str(source.cell(1, col).value or "").strip().lower() == "_kind":
            kind_col = col
            break

    # The timescale occupies F through the column immediately before _Kind.
    # Keep Dashboard formulas bounded to those real reporting columns instead
    # of using a broad sentinel range such as F:ZZ.
    last_timescale_col = get_column_letter((kind_col - 1) if kind_col else source.max_column)

    output_row = 39
    source_row = 2
    pair_index = 0
    while source_row <= source.max_row:
        plan_row = source_row
        actual_row = source_row + 1 if source_row + 1 <= source.max_row else source_row
        kind = str(source.cell(plan_row, kind_col).value or "").strip().lower() if kind_col else ""
        wbs_code = str(source.cell(plan_row, 1).value or "").strip()
        if not kind:
            kind = "project" if wbs_code.upper() == "PROJECT" else "activity"
        wbs_depth = 0 if wbs_code.upper() == "PROJECT" else max(0, len([part for part in wbs_code.split(".") if part]) - 1)

        if kind == "project":
            outline_level = 0
        else:
            parts = [part for part in wbs_code.split(".") if part]
            outline_level = min(max(len(parts), 1), 7)

        for pa_label, src_row, is_plan in (("Plan", plan_row, True), ("Actual", actual_row, False)):
            row = output_row
            if is_plan:
                ws[f"B{row}"] = f"='{TABLE_SHEET}'!A{plan_row}"
                ws.merge_cells(f"C{row}:E{row}")
                ws[f"C{row}"] = f"='{TABLE_SHEET}'!B{plan_row}"
            else:
                ws.merge_cells(f"C{row}:E{row}")

            ws.merge_cells(f"F{row}:G{row}")
            ws[f"F{row}"] = pa_label
            ws.merge_cells(f"H{row}:I{row}")
            if is_plan:
                ws[f"H{row}"] = f"='{TABLE_SHEET}'!C{plan_row}"
            ws.merge_cells(f"J{row}:K{row}")
            # progress_table is already the authoritative weekly activity source.
            # Sum only reporting columns whose header date is on/before the
            # selected Dashboard cutoff. SUMIFS recalculates immediately when
            # K5 changes and avoids re-deriving progress from the S-curve.
            progress_formula = (
                f'=IFERROR(SUMIFS(\'{TABLE_SHEET}\'!$F{src_row}:${last_timescale_col}{src_row},'
                f'\'{TABLE_SHEET}\'!$F$1:${last_timescale_col}$1,"<="&$K$5)/100,0)'
            )
            ws[f"L{row}"] = progress_formula
            ws.merge_cells(f"L{row}:M{row}")
            ws[f"J{row}"] = f'=IFERROR(\'{TABLE_SHEET}\'!$C${plan_row}*L{row},0)'

            ws.merge_cells(f"N{row}:O{row}")
            ws.merge_cells(f"P{row}:Q{row}")
            if is_plan:
                ws[f"N{row}"] = ""
                # Native Excel filter needs the same status value on both rows
                # of a Plan/Actual pair.  Keep the Plan-row status visually
                # hidden below while preserving the filter value.
                actual_output_row = row + 1
                ws[f"P{row}"] = (
                    f'=IF(L{row}<=0,"Not Started",'
                    f'IF(L{row}>=1,"Complete",'
                    f'IF(L{actual_output_row}<L{row},"Behind","On Track")))'
                )
            else:
                plan_output_row = row - 1
                ws[f"N{row}"] = f'=IFERROR(L{row}-L{plan_output_row},0)'
                ws[f"P{row}"] = (
                    f'=IF(L{plan_output_row}<=0,"Not Started",'
                    f'IF(L{plan_output_row}>=1,"Complete",'
                    f'IF(L{row}<L{plan_output_row},"Behind","On Track")))'
                )

            base_fill = WHITE if pair_index % 2 == 0 else LIGHT_GRAY
            if kind in {"project", "wbs"}:
                base_fill = LIGHT_BLUE if kind == "project" or wbs_depth == 0 else ("EEF4FA" if wbs_depth == 1 else "F5F8FC")

            for row_cells in ws[f"B{row}:Q{row}"]:
                for cell in row_cells:
                    cell.border = _thin_border()
                    cell.fill = _solid(base_fill)
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
                    if kind in {"project", "wbs"}:
                        cell.font = Font(name=_FONT, bold=True, color=NAVY, size=9)

            if is_plan:
                ws[f"C{row}"].alignment = Alignment(vertical="center", wrap_text=True, indent=min(wbs_depth if kind in {"project", "wbs"} else wbs_depth + 1, 4))

            # Plan/Actual are a visual pair. Type and progress use the same blue/green
            # language as the OKD application without introducing extra business logic.
            pa_fill = LIGHT_BLUE if is_plan else LIGHT_GREEN
            pa_color = BLUE if is_plan else GREEN
            ws[f"F{row}"].fill = _solid(pa_fill)
            ws[f"F{row}"].font = Font(name=_FONT, bold=True, color=pa_color, size=9)
            ws[f"F{row}"].alignment = Alignment(horizontal="center", vertical="center")
            ws[f"L{row}"].fill = _solid(pa_fill)
            ws[f"L{row}"].font = Font(name=_FONT, bold=True, color=pa_color, size=9)
            if is_plan:
                # Keep the native-filter value without displaying duplicate
                # status text on the Plan row.
                ws[f"P{row}"].font = Font(name=_FONT, color=base_fill, size=9)
            else:
                ws[f"N{row}"].fill = _solid(LIGHT_GREEN)
                ws[f"P{row}"].fill = _solid(LIGHT_GREEN)
                ws[f"N{row}"].font = Font(name=_FONT, bold=True, color=GREEN, size=9)
                ws[f"P{row}"].font = Font(name=_FONT, bold=True, color=GREEN, size=9)
                ws[f"N{row}"].alignment = Alignment(horizontal="right", vertical="center")
                ws[f"P{row}"].alignment = Alignment(horizontal="center", vertical="center")

            ws.row_dimensions[row].outlineLevel = outline_level
            ws.row_dimensions[row].hidden = False
            ws.row_dimensions[row].collapsed = False
            ws.row_dimensions[row].height = 22

            ws[f"H{row}"].number_format = "#,##0.00"
            ws[f"J{row}"].number_format = "#,##0.00"
            ws[f"L{row}"].number_format = "0.00%"
            ws[f"N{row}"].number_format = "0.00%;[Red]-0.00%;0.00%"
            output_row += 1

        source_row += 2
        pair_index += 1

    # Summary rows sit above their detail rows, matching the main-sheet outline.
    # Both Plan and Actual rows receive the same outline level so a pair never splits.
    ws.sheet_properties.outlinePr.summaryBelow = False
    ws.sheet_properties.outlinePr.applyStyles = True
    ws.sheet_properties.outlinePr.showOutlineSymbols = True

    # Native Excel filter on Status only.  P38:P... is intentional: the
    # Activity Table remains visually clean and the Plan/Actual pair values
    # match, so filtering never separates the pair.
    last_activity_row = max(38, output_row - 1)
    ws.auto_filter.ref = f"P38:P{last_activity_row}"

    ws.print_area = f"B2:Q{max(56, output_row - 1)}"


def build_dashboard(workbook, *, project_name: str | None = None) -> None:
    """Create the simple vertical dashboard as a separate first worksheet."""
    if PROGRESS_SHEET not in workbook.sheetnames or TABLE_SHEET not in workbook.sheetnames:
        return
    _build_data_sheet(workbook, workbook[PROGRESS_SHEET])
    _build_dashboard_sheet(workbook, project_name=project_name)


def build_dashboard_file(input_file: Path, output_file: Path, *, project_name: str | None = None) -> None:
    workbook = load_workbook(input_file, data_only=False)
    try:
        build_dashboard(workbook, project_name=project_name)
        configure_incremental_excel_recalculation(workbook)
        workbook.save(output_file)
    finally:
        workbook.close()
