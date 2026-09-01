from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from progress_studio.domain.mapping_models import AllocationRecord, BOQRow


MAPPING_SHEET = "BOQ Activity Mapping"
SUMMARY_SHEET = "Mapping Summary"


class EarnedValueWorkbookInputError(ValueError):
    """Raised when embedded workbook data is not safe to use for Earned Value."""


@dataclass(frozen=True, slots=True)
class EmbeddedEarnedValueInputs:
    """BOQ and allocation inputs reconstructed from a mapped Progress Studio workbook."""

    workbook: Path
    boq_rows: tuple[BOQRow, ...]
    allocations: tuple[AllocationRecord, ...]


class EarnedValueInputWorkbookReader:
    """Read BOQ provenance embedded by ``MappedWorkbookExporter``.

    The mapping sheet stores Share % as an Excel fraction (for example 0.30 for
    30%), while ``AllocationRecord.share_percent`` uses percentage points
    (30.0). This reader owns that boundary conversion.

    Mapping completeness is intentionally validated from ``Mapping Summary``
    before reconstruction. An unmapped BOQ item has no row in
    ``BOQ Activity Mapping`` and therefore cannot be recovered safely from that
    sheet alone.
    """

    REQUIRED_MAPPING_HEADERS = (
        "Activity ID",
        "BOQ Key",
        "Source Sheet",
        "Source Row",
        "WBS-2",
        "WBS-3",
        "WBS-4",
        "BOQ Description",
        "BOQ Amount",
        "Share %",
        "BOQ ID",
    )

    def read(self, workbook_path: Path) -> EmbeddedEarnedValueInputs:
        path = Path(workbook_path).expanduser().resolve()
        if not path.is_file():
            raise EarnedValueWorkbookInputError(
                f"Earned Value workbook was not found: {path}"
            )
        if path.suffix.lower() not in {".xlsx", ".xlsm"}:
            raise EarnedValueWorkbookInputError(
                "Select a Progress Studio .xlsx or .xlsm workbook."
            )

        try:
            workbook = load_workbook(
                path,
                read_only=True,
                data_only=True,
            )
        except Exception as exc:
            raise EarnedValueWorkbookInputError(
                f"Earned Value workbook cannot be opened: {exc}"
            ) from exc

        try:
            self._validate_mapping_complete(workbook)
            boq_rows, allocations = self._read_mapping(workbook)
        finally:
            workbook.close()

        return EmbeddedEarnedValueInputs(
            workbook=path,
            boq_rows=boq_rows,
            allocations=allocations,
        )

    def _validate_mapping_complete(self, workbook) -> None:
        if SUMMARY_SHEET not in workbook.sheetnames:
            raise EarnedValueWorkbookInputError(
                f"Worksheet '{SUMMARY_SHEET}' was not found. "
                "Earned Value requires embedded mapping reconciliation."
            )

        ws = workbook[SUMMARY_SHEET]
        summary: dict[str, object] = {}
        for row in ws.iter_rows(min_col=1, max_col=2, values_only=True):
            label = self._text(row[0]).lower()
            if label:
                summary[label] = row[1]

        required = (
            "boq items",
            "fully allocated boq items",
            "partially allocated boq items",
            "unmapped boq items",
            "allocated percent",
        )
        missing = [name for name in required if name not in summary]
        if missing:
            raise EarnedValueWorkbookInputError(
                "Mapping Summary is missing required fields: " + ", ".join(missing)
            )

        boq_count = self._int(summary["boq items"], "BOQ items")
        full_count = self._int(
            summary["fully allocated boq items"], "Fully allocated BOQ items"
        )
        partial_count = self._int(
            summary["partially allocated boq items"], "Partially allocated BOQ items"
        )
        unmapped_count = self._int(
            summary["unmapped boq items"], "Unmapped BOQ items"
        )
        allocated_fraction = self._fraction(
            summary["allocated percent"], "Allocated percent"
        )

        if (
            boq_count < 0
            or full_count < 0
            or partial_count < 0
            or unmapped_count < 0
        ):
            raise EarnedValueWorkbookInputError(
                "Mapping Summary contains negative reconciliation counts."
            )

        if full_count + partial_count + unmapped_count != boq_count:
            raise EarnedValueWorkbookInputError(
                "Mapping Summary reconciliation counts are inconsistent."
            )

        tolerance = 0.0001  # 0.01 percentage point
        complete = (
            boq_count > 0
            and full_count == boq_count
            and partial_count == 0
            and unmapped_count == 0
            and abs(allocated_fraction - 1.0) <= tolerance
        )
        if not complete:
            raise EarnedValueWorkbookInputError(
                "Earned Value requires 100% BOQ mapping: "
                f"{full_count}/{boq_count} fully allocated, "
                f"{partial_count} partial, {unmapped_count} unmapped, "
                f"{allocated_fraction * 100.0:.2f}% allocated."
            )

    def _read_mapping(
        self,
        workbook,
    ) -> tuple[tuple[BOQRow, ...], tuple[AllocationRecord, ...]]:
        if MAPPING_SHEET not in workbook.sheetnames:
            raise EarnedValueWorkbookInputError(
                f"Worksheet '{MAPPING_SHEET}' was not found."
            )

        ws = workbook[MAPPING_SHEET]
        rows = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows)
        except StopIteration as exc:
            raise EarnedValueWorkbookInputError(
                f"Worksheet '{MAPPING_SHEET}' is empty."
            ) from exc

        headers = {
            self._text(value).lower(): index
            for index, value in enumerate(header_row)
            if self._text(value)
        }
        missing = [
            name
            for name in self.REQUIRED_MAPPING_HEADERS
            if name.lower() not in headers
        ]
        if missing:
            raise EarnedValueWorkbookInputError(
                "BOQ Activity Mapping is missing required columns: "
                + ", ".join(missing)
            )

        boq_by_key: dict[str, BOQRow] = {}
        allocations: list[AllocationRecord] = []

        for sheet_row, values in enumerate(rows, start=2):
            activity_id = self._cell(values, headers, "activity id").strip().upper()
            boq_key = self._cell(values, headers, "boq key").strip()

            if not activity_id and not boq_key:
                continue
            if not activity_id or not boq_key:
                raise EarnedValueWorkbookInputError(
                    f"Incomplete mapping identity at {MAPPING_SHEET} row {sheet_row}."
                )

            source_sheet = self._cell(values, headers, "source sheet").strip()
            source_row = self._required_int(
                self._value(values, headers, "source row"),
                f"{MAPPING_SHEET}!Source Row row {sheet_row}",
            )
            wbs2 = self._cell(values, headers, "wbs-2").strip()
            wbs3 = self._cell(values, headers, "wbs-3").strip()
            wbs4 = self._cell(values, headers, "wbs-4").strip()
            description = self._cell(values, headers, "boq description").strip()
            amount = self._required_float(
                self._value(values, headers, "boq amount"),
                f"{MAPPING_SHEET}!BOQ Amount row {sheet_row}",
            )
            stable_id = self._cell(values, headers, "boq id").strip() or boq_key
            share_fraction = self._required_float(
                self._value(values, headers, "share %"),
                f"{MAPPING_SHEET}!Share % row {sheet_row}",
            )

            if amount < 0.0:
                raise EarnedValueWorkbookInputError(
                    f"Negative BOQ Amount at {MAPPING_SHEET} row {sheet_row}."
                )
            if share_fraction <= 0.0 or share_fraction > 1.0:
                raise EarnedValueWorkbookInputError(
                    f"Share % at {MAPPING_SHEET} row {sheet_row} "
                    "must be greater than 0% and no more than 100%."
                )

            boq = BOQRow(
                key=boq_key,
                source_sheet=source_sheet,
                source_row=source_row,
                wbs2=wbs2,
                wbs3=wbs3,
                wbs4=wbs4,
                description=description,
                amount=amount,
                stable_id=stable_id,
            )
            prior = boq_by_key.get(boq_key)
            if prior is not None and prior != boq:
                raise EarnedValueWorkbookInputError(
                    f"Conflicting embedded BOQ metadata for BOQ Key '{boq_key}'."
                )
            boq_by_key[boq_key] = boq

            allocations.append(
                AllocationRecord(
                    boq_key=boq_key,
                    activity_id=activity_id,
                    share_percent=share_fraction * 100.0,
                )
            )

        if not boq_by_key:
            raise EarnedValueWorkbookInputError(
                f"Worksheet '{MAPPING_SHEET}' contains no mapping records."
            )

        allocation_totals: dict[str, float] = {}
        for allocation in allocations:
            allocation_totals[allocation.boq_key] = (
                allocation_totals.get(allocation.boq_key, 0.0)
                + allocation.share_percent
            )

        tolerance = 0.01  # percentage point; same EV-1 contract
        bad = [
            (key, percent)
            for key, percent in allocation_totals.items()
            if abs(percent - 100.0) > tolerance
        ]
        if bad:
            key, percent = bad[0]
            raise EarnedValueWorkbookInputError(
                "Embedded allocation rows do not reconcile to 100% for "
                f"BOQ Key '{key}': {percent:.2f}%."
            )

        return (
            tuple(boq_by_key.values()),
            tuple(allocations),
        )

    @staticmethod
    def _text(value: object) -> str:
        return str(value or "").strip()

    @staticmethod
    def _value(values: tuple[object, ...], headers: dict[str, int], name: str) -> object:
        index = headers[name]
        return values[index] if index < len(values) else None

    @classmethod
    def _cell(
        cls,
        values: tuple[object, ...],
        headers: dict[str, int],
        name: str,
    ) -> str:
        return cls._text(cls._value(values, headers, name))

    @staticmethod
    def _required_float(value: object, label: str) -> float:
        try:
            return float(value)
        except (TypeError, ValueError) as exc:
            raise EarnedValueWorkbookInputError(
                f"{label} must be numeric."
            ) from exc

    @staticmethod
    def _required_int(value: object, label: str) -> int:
        try:
            number = float(value)
        except (TypeError, ValueError) as exc:
            raise EarnedValueWorkbookInputError(
                f"{label} must be an integer."
            ) from exc
        if not number.is_integer():
            raise EarnedValueWorkbookInputError(
                f"{label} must be an integer."
            )
        return int(number)

    @classmethod
    def _int(cls, value: object, label: str) -> int:
        return cls._required_int(value, f"Mapping Summary '{label}'")

    @classmethod
    def _fraction(cls, value: object, label: str) -> float:
        number = cls._required_float(value, f"Mapping Summary '{label}'")
        if number < 0.0 or number > 1.0:
            raise EarnedValueWorkbookInputError(
                f"Mapping Summary '{label}' must be between 0% and 100%."
            )
        return number
