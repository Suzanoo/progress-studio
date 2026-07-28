from __future__ import annotations

from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from progress_studio.domain.mapping_models import ActivityRow, BOQRow


def _header_map(values: Iterable[object]) -> dict[str, int]:
    result: dict[str, int] = {}
    for index, value in enumerate(values):
        name = str(value or "").strip().lower()
        if name:
            result[name] = index
    return result


def _text(value: object) -> str:
    return str(value or "").strip()


def _number(value: object) -> float:
    if value in (None, ""):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(",", "").strip())
    except (TypeError, ValueError):
        return 0.0


def _wbs_parent_child(wbs: str) -> tuple[str, str]:
    value = wbs.strip()
    if not value:
        return "", ""
    parts = [part for part in value.split(".") if part]
    if len(parts) <= 1:
        return "", value
    return ".".join(parts[:-1]), value


def _ancestor_codes(wbs: str) -> tuple[str, ...]:
    parts = [part for part in wbs.strip().split(".") if part]
    return tuple(".".join(parts[:index]) for index in range(1, len(parts)))


class ProgressActivityReader:
    def read(self, path: Path) -> list[ActivityRow]:
        workbook = load_workbook(path, read_only=True, data_only=False)
        try:
            if "Amount Mapping" not in workbook.sheetnames:
                raise ValueError("Worksheet 'Amount Mapping' was not found in the Progress workbook.")
            sheet = workbook["Amount Mapping"]
            rows = sheet.iter_rows(values_only=True)
            try:
                headers = _header_map(next(rows))
            except StopIteration as exc:
                raise ValueError("The Amount Mapping worksheet is empty.") from exc

            required = ("activity id", "wbs", "description")
            missing = [name for name in required if name not in headers]
            if missing:
                raise ValueError("Missing Progress columns: " + ", ".join(missing))

            result: list[ActivityRow] = []
            wbs_names: dict[str, str] = {}
            for values in rows:
                activity_id = _text(values[headers["activity id"]])
                wbs = _text(values[headers["wbs"]])
                description = _text(values[headers["description"]])

                # WBS rows are already present in Amount Mapping. Keep their
                # names while streaming so Activity rows can carry a compact
                # hierarchy without reopening or rescanning the workbook.
                if not activity_id:
                    if wbs and description:
                        wbs_names[wbs] = description
                    continue

                parent, child = _wbs_parent_child(wbs)
                path = tuple(
                    (code, wbs_names.get(code, code))
                    for code in _ancestor_codes(wbs)
                )
                result.append(
                    ActivityRow(
                        activity_id=activity_id,
                        parent_wbs=parent,
                        child_wbs=child,
                        description=description,
                        wbs_path=path,
                    )
                )
            return result
        finally:
            workbook.close()


class BOQSheetReader:
    def list_sheets(self, path: Path) -> list[str]:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            return list(workbook.sheetnames)
        finally:
            workbook.close()

    def read(self, path: Path, sheet_name: str) -> list[BOQRow]:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"Worksheet {sheet_name!r} was not found in the BOQ workbook.")
            sheet = workbook[sheet_name]
            rows = sheet.iter_rows(values_only=True)
            try:
                headers = _header_map(next(rows))
            except StopIteration as exc:
                raise ValueError(f"Worksheet {sheet_name!r} is empty.") from exc

            required = ("wbs-2", "wbs-3", "wbs-4", "description", "amount")
            missing = [name for name in required if name not in headers]
            if missing:
                raise ValueError("Missing BOQ columns: " + ", ".join(missing))

            source_sheet_index = headers.get("source sheet")
            source_row_index = headers.get("source row")
            result: list[BOQRow] = []
            for excel_row, values in enumerate(rows, start=2):
                amount = _number(values[headers["amount"]])
                if amount <= 0:
                    continue
                source_sheet = (
                    _text(values[source_sheet_index]) if source_sheet_index is not None else sheet_name
                ) or sheet_name
                raw_source_row = values[source_row_index] if source_row_index is not None else excel_row
                try:
                    source_row = int(raw_source_row)
                except (TypeError, ValueError):
                    source_row = excel_row
                key = f"{source_sheet}|{source_row}|{excel_row}"
                result.append(
                    BOQRow(
                        key=key,
                        source_sheet=source_sheet,
                        source_row=source_row,
                        wbs2=_text(values[headers["wbs-2"]]),
                        wbs3=_text(values[headers["wbs-3"]]),
                        wbs4=_text(values[headers["wbs-4"]]),
                        description=_text(values[headers["description"]]),
                        amount=amount,
                    )
                )
            return result
        finally:
            workbook.close()
