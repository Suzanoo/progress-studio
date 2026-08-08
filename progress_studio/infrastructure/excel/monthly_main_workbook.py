from __future__ import annotations

from collections import OrderedDict
from copy import copy
from datetime import date, datetime

from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from progress_studio.infrastructure.excel.calculation_policy import configure_incremental_excel_recalculation
from progress_studio.infrastructure.excel.progress_workbook import (
    add_progress_conditional_formatting,
    clear_progress_conditional_formatting,
    clear_timescale_direct_fills,
)
from progress_studio.infrastructure.excel.timescale_workbook import (
    DATE_FILL,
    HEADER_BORDER,
    MONTH_FILL,
    WEEK_FILL,
    YEAR_FILL,
)
from progress_studio.infrastructure.excel.worksheet_filters import configure_filter_buttons

HEADER_ROW = 4
FIRST_DATA_ROW = 5


def _normalize(value: object) -> str:
    return str(value or "").strip().lower()


def _as_date(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def _header_map(ws) -> dict[str, int]:
    return {
        _normalize(ws.cell(HEADER_ROW, col).value): col
        for col in range(1, ws.max_column + 1)
        if _normalize(ws.cell(HEADER_ROW, col).value)
    }


def _timescale_columns(ws) -> list[int]:
    return [
        col
        for col in range(1, ws.max_column + 1)
        if _as_date(ws.cell(HEADER_ROW, col).value) is not None
    ]


def _month_buckets(ws, timescale_cols: list[int]) -> list[tuple[tuple[int, int], list[int]]]:
    grouped: OrderedDict[tuple[int, int], list[int]] = OrderedDict()
    for col in timescale_cols:
        value = _as_date(ws.cell(HEADER_ROW, col).value)
        if value is None:
            continue
        grouped.setdefault((value.year, value.month), []).append(col)
    return list(grouped.items())


def _remove_timescale_merges(ws, first_col: int) -> None:
    for merged in list(ws.merged_cells.ranges):
        if merged.max_col >= first_col:
            ws.unmerge_cells(str(merged))


def _write_year_headers(ws, first_col: int, buckets: list[tuple[tuple[int, int], list[int]]]) -> None:
    if not buckets:
        return
    start = first_col
    current_year = buckets[0][0][0]
    for index in range(1, len(buckets) + 1):
        at_end = index == len(buckets)
        next_year = None if at_end else buckets[index][0][0]
        if at_end or next_year != current_year:
            end = first_col + index - 1
            if end > start:
                ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
            cell = ws.cell(1, start)
            cell.value = str(current_year)
            for col in range(start, end + 1):
                target = ws.cell(1, col)
                target.fill = YEAR_FILL
                target.font = Font(color="FFFFFF", bold=True)
                target.alignment = Alignment(horizontal="center", vertical="center")
                target.border = HEADER_BORDER
            start = end + 1
            current_year = next_year


def _quote_sheet(name: str) -> str:
    return "'" + name.replace("'", "''") + "'"


def build_monthly_main_view(
    workbook,
    *,
    source_sheet: str = "main",
    target_sheet: str = "main_monthly",
    require_timescale: bool = True,
) -> int:
    """Build a formula-driven monthly view from the weekly ``main`` sheet.

    ``main`` remains the editable/source schedule. The monthly sheet keeps the
    same row hierarchy and Activity Data columns, but every monthly timescale
    cell derives from the weekly cells in ``main``. This means edits to weekly
    Actual progress automatically flow through when Excel recalculates.
    """
    if source_sheet not in workbook.sheetnames:
        raise ValueError(f"Monthly view source worksheet was not found: {source_sheet}")

    source = workbook[source_sheet]
    timescale_cols = _timescale_columns(source)
    if not timescale_cols:
        if require_timescale:
            raise ValueError("Weekly timescale columns were not found on the main worksheet.")
        return 0

    if target_sheet in workbook.sheetnames:
        del workbook[target_sheet]

    first_timescale_col = timescale_cols[0]
    buckets = _month_buckets(source, timescale_cols)
    if not buckets:
        raise ValueError("No monthly buckets could be derived from the main weekly timescale.")

    # copy_worksheet keeps the exact Activity Data layout, row outline levels,
    # styles and formulas. We then replace only the weekly timescale.
    monthly = workbook.copy_worksheet(source)
    monthly.title = target_sheet
    source_index = workbook.worksheets.index(source)
    workbook._sheets.remove(monthly)
    workbook._sheets.insert(source_index + 1, monthly)
    monthly.cell(1, 1).value = "Activity Data — Monthly View"

    # Snapshot weekly row formatting before removing those columns.
    row_styles = {
        row: copy(monthly.cell(row, first_timescale_col)._style)
        for row in range(FIRST_DATA_ROW, monthly.max_row + 1)
    }
    row_number_formats = {
        row: monthly.cell(row, first_timescale_col).number_format
        for row in range(FIRST_DATA_ROW, monthly.max_row + 1)
    }

    _remove_timescale_merges(monthly, first_timescale_col)
    monthly.delete_cols(first_timescale_col, monthly.max_column - first_timescale_col + 1)

    source_ref = _quote_sheet(source_sheet)
    for month_index, ((year, month), weekly_cols) in enumerate(buckets, start=1):
        target_col = first_timescale_col + month_index - 1
        first_week_col = get_column_letter(weekly_cols[0])
        last_week_col = get_column_letter(weekly_cols[-1])
        last_reporting_date = _as_date(source.cell(HEADER_ROW, weekly_cols[-1]).value)

        # Preserve the four-row timescale grammar: Year / Month / Period / Date.
        month_cell = monthly.cell(2, target_col)
        month_cell.value = date(year, month, 1).strftime("%B")
        month_cell.fill = MONTH_FILL
        month_cell.font = Font(color="000000", bold=True)
        month_cell.alignment = Alignment(horizontal="center", vertical="center")
        month_cell.border = HEADER_BORDER

        period_cell = monthly.cell(3, target_col)
        period_cell.value = f"M{month_index}"
        period_cell.fill = WEEK_FILL
        period_cell.font = Font(color="000000", bold=True)
        period_cell.alignment = Alignment(horizontal="center", vertical="center")
        period_cell.border = HEADER_BORDER

        date_cell = monthly.cell(HEADER_ROW, target_col)
        date_cell.value = last_reporting_date
        date_cell.fill = DATE_FILL
        date_cell.font = Font(color="000000", bold=True)
        date_cell.alignment = Alignment(horizontal="center", vertical="center")
        date_cell.border = HEADER_BORDER
        date_cell.number_format = "dd/mm/yy"
        monthly.column_dimensions[get_column_letter(target_col)].width = 12

        for row in range(FIRST_DATA_ROW, monthly.max_row + 1):
            cell = monthly.cell(row, target_col)
            cell._style = copy(row_styles[row])
            cell.number_format = row_number_formats[row]

            pa = str(monthly.cell(row, 4).value or "").strip().upper()
            row_type = _normalize(monthly.cell(row, 1).value)
            source_range = f"{source_ref}!{first_week_col}{row}:{last_week_col}{row}"
            if row_type == "s-curve" and pa in {"AP", "AA"}:
                # Cumulative S-curve rows use the last reporting value in month.
                cell.value = f"={source_ref}!{last_week_col}{row}"
            elif row_type == "activity" and pa == "P":
                # Activity Plan distributions are generated numeric inputs in the
                # weekly main sheet. Freeze their monthly sum as a value so the
                # monthly view does not duplicate thousands of static formulas.
                weekly_values = [source.cell(row, col).value for col in weekly_cols]
                if all(value in (None, "") or isinstance(value, (int, float)) for value in weekly_values):
                    populated = [float(value) for value in weekly_values if value not in (None, "")]
                    cell.value = sum(populated) if populated else ""
                else:
                    cell.value = f'=IF(COUNT({source_range})=0,"",SUM({source_range}))'
            else:
                # Actual and summary rows remain formula-driven so edits to weekly
                # progress/Amount continue to flow through immediately.
                cell.value = f'=IF(COUNT({source_range})=0,"",SUM({source_range}))'

    _write_year_headers(monthly, first_timescale_col, buckets)

    headers = _header_map(monthly)
    required = {
        "row type": headers.get("row type"),
        "p/a": headers.get("p/a"),
        "activity id": headers.get("activity id"),
        "outline level": headers.get("outline level"),
    }
    if any(value is None for value in required.values()):
        missing = [name for name, value in required.items() if value is None]
        raise ValueError("Monthly main view is missing required columns: " + ", ".join(missing))

    monthly_timescale_cols = list(
        range(first_timescale_col, first_timescale_col + len(buckets))
    )
    progress_rows = [
        row
        for row in range(FIRST_DATA_ROW, monthly.max_row + 1)
        if _normalize(monthly.cell(row, required["row type"]).value)
        in {"project summary", "wbs", "activity"}
        or str(monthly.cell(row, required["p/a"]).value or "").strip().upper() == "A"
    ]
    last_progress_row = max(progress_rows) if progress_rows else FIRST_DATA_ROW

    # Monthly is calculated-only: no input validation is copied. Rebuild the
    # progress color bands on the new timescale and preserve the simple filters.
    try:
        monthly.data_validations.dataValidation = []
    except AttributeError:
        pass
    clear_timescale_direct_fills(monthly, monthly_timescale_cols, last_progress_row)
    clear_progress_conditional_formatting(monthly, monthly_timescale_cols)
    add_progress_conditional_formatting(
        monthly,
        monthly_timescale_cols,
        required["row type"],
        required["activity id"],
        required["p/a"],
        required["outline level"],
        last_progress_row,
    )

    monthly.freeze_panes = monthly.cell(FIRST_DATA_ROW, first_timescale_col)
    configure_filter_buttons(
        monthly,
        header_row=HEADER_ROW,
        last_row=last_progress_row,
        last_col=monthly.max_column,
        visible_columns={required["row type"], required["p/a"]},
    )
    monthly.sheet_properties.outlinePr.summaryBelow = False
    monthly.sheet_properties.outlinePr.applyStyles = True
    monthly.sheet_view.showGridLines = source.sheet_view.showGridLines

    configure_incremental_excel_recalculation(workbook)
    return len(buckets)
