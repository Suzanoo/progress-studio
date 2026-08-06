from __future__ import annotations

import math
from datetime import date, datetime
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.datetime import from_excel
    from openpyxl.formatting.rule import FormulaRule
except ImportError as exc:
    raise SystemExit(
        "openpyxl is not installed. Run: pip install openpyxl"
    ) from exc

from progress_studio.infrastructure.excel.styles import (
    ACTUAL_FILL,
    ACTUAL_FONT,
    BORDER_COLOR,
    HEADER_FILL,
    HEADER_FONT,
    PLAN_FILL,
    PLAN_FONT,
)
from progress_studio.domain import ActivityWbsSequencer


SCRIPT_VERSION = "1.9.1-interactive-accplan-project-wbs"
DEFAULT_SOURCE_SHEET = "main"
HEADER_ROW = 4
WEEK_ROW = 3

OKD_PROGRESS_SHEET = "progress"
OKD_TABLE_SHEET = "progress_table"


class OKDExportError(RuntimeError):
    pass


def normalize(value: object) -> str:
    return " ".join(str(value or "").strip().lower().split())


def as_number(value: object) -> float | None:
    if value is None or value == "" or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        number = float(value)
        return number if math.isfinite(number) else None
    try:
        number = float(str(value).replace(",", "").strip())
        return number if math.isfinite(number) else None
    except (TypeError, ValueError):
        return None


def as_date(value: object, epoch) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        try:
            converted = from_excel(value, epoch=epoch)
            return converted.date() if isinstance(converted, datetime) else converted
        except (TypeError, ValueError, OverflowError):
            return None

    text = str(value).strip()
    for fmt in (
        "%d/%m/%Y",
        "%d/%m/%y",
        "%Y-%m-%d",
        "%d-%m-%Y",
        "%m/%d/%Y",
    ):
        try:
            parsed = datetime.strptime(text, fmt).date()
            return (
                parsed.replace(year=parsed.year - 543)
                if parsed.year > 2400
                else parsed
            )
        except ValueError:
            continue
    return None


def get_headers(ws) -> dict[str, int]:
    result: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        key = normalize(ws.cell(HEADER_ROW, col).value)
        if key:
            result[key] = col
    return result


def find_column(
    headers: dict[str, int],
    aliases: tuple[str, ...],
    *,
    required: bool = False,
) -> int | None:
    for alias in aliases:
        key = normalize(alias)
        if key in headers:
            return headers[key]
    if required:
        raise OKDExportError(
            "Missing required column. Expected one of: "
            + ", ".join(aliases)
        )
    return None


def find_week_columns(ws, epoch) -> list[tuple[int, date]]:
    result: list[tuple[int, date]] = []

    for col in range(1, ws.max_column + 1):
        label = str(ws.cell(WEEK_ROW, col).value or "").strip().upper()
        if not (label.startswith("W") and label[1:].isdigit()):
            continue

        cutoff = as_date(ws.cell(HEADER_ROW, col).value, epoch)
        if cutoff is not None:
            result.append((col, cutoff))

    if not result:
        raise OKDExportError(
            "Weekly timescale not found. Expected W1, W2, ... in row 3 "
            "and dates in row 4."
        )

    return result


def find_actual_row(
    ws,
    plan_row: int,
    activity_id: str,
    activity_id_col: int,
    pa_col: int,
) -> int | None:
    # Normal structure from Script 02: Actual row is immediately below Plan.
    candidate = plan_row + 1
    if candidate <= ws.max_row:
        candidate_pa = normalize(ws.cell(candidate, pa_col).value)
        candidate_id = str(
            ws.cell(candidate, activity_id_col).value or ""
        ).strip()
        if candidate_pa == "a" and candidate_id == activity_id:
            return candidate

    # Small fallback search for manually rearranged rows.
    for row in range(plan_row + 1, min(plan_row + 6, ws.max_row) + 1):
        row_pa = normalize(ws.cell(row, pa_col).value)
        row_id = str(ws.cell(row, activity_id_col).value or "").strip()
        if row_pa == "a" and row_id == activity_id:
            return row

    return None


def is_wbs_row_type(value: object) -> bool:
    row_type = normalize(value)
    return (
        "wbs" in row_type
        or row_type in {"project summary", "project"}
    )


def belongs_to_wbs(activity_wbs: str, parent_wbs: str) -> bool:
    activity_key = activity_wbs.strip()
    parent_key = parent_wbs.strip()
    if not activity_key or not parent_key:
        return False
    return (
        activity_key == parent_key
        or activity_key.startswith(parent_key + ".")
        or activity_key.startswith(parent_key + "-")
        or activity_key.startswith(parent_key + "/")
    )



def excel_sheet_ref(sheet_name: str) -> str:
    """Return a safely quoted Excel worksheet reference."""
    return "'" + sheet_name.replace("'", "''") + "'"


def find_actual_summary_row(
    ws,
    plan_row: int,
    pa_col: int,
    row_type_col: int,
    wbs_col: int | None,
    description_col: int,
) -> int | None:
    """Find the Actual row paired with a WBS or Project summary Plan row.

    Prefer the immediate following Actual row. Then try matching WBS,
    description, or row type within a small local range.
    """
    plan_type = normalize(ws.cell(plan_row, row_type_col).value)
    plan_wbs = (
        str(ws.cell(plan_row, wbs_col).value or "").strip()
        if wbs_col is not None
        else ""
    )
    plan_description = normalize(ws.cell(plan_row, description_col).value)

    candidate = plan_row + 1
    if candidate <= ws.max_row:
        if normalize(ws.cell(candidate, pa_col).value) == "a":
            return candidate

    fallback_same_type: int | None = None

    for row in range(plan_row + 1, min(plan_row + 8, ws.max_row) + 1):
        row_pa = normalize(ws.cell(row, pa_col).value)

        # Stop when the next unrelated Plan block begins.
        if row_pa == "p":
            break
        if row_pa != "a":
            continue

        row_type = normalize(ws.cell(row, row_type_col).value)
        row_wbs = (
            str(ws.cell(row, wbs_col).value or "").strip()
            if wbs_col is not None
            else ""
        )
        row_description = normalize(ws.cell(row, description_col).value)

        if plan_wbs and row_wbs == plan_wbs:
            return row
        if plan_description and row_description == plan_description:
            return row
        if row_type == plan_type and fallback_same_type is None:
            fallback_same_type = row

    return fallback_same_type


def find_scurve_cumulative_rows(
    ws,
    headers: dict[str, int],
) -> tuple[int, int]:
    """Return the Acc. Plan and Acc. Actual rows from the main sheet."""
    row_type_col = find_column(headers, ("Row Type",), required=True)
    description_col = find_column(
        headers,
        ("Description", "Activity Name", "Activity Description"),
        required=True,
    )
    pa_col = find_column(headers, ("P/A",), required=True)

    acc_plan_row: int | None = None
    acc_actual_row: int | None = None

    for row in range(HEADER_ROW + 1, ws.max_row + 1):
        row_type = normalize(ws.cell(row, row_type_col).value)
        description = normalize(ws.cell(row, description_col).value)
        pa_value = normalize(ws.cell(row, pa_col).value)

        if row_type != "s-curve":
            continue

        if (
            description in {"acc. plan", "acc plan", "accumulate plan"}
            or pa_value == "ap"
        ):
            acc_plan_row = row
        elif (
            description in {"acc. actual", "acc actual", "accumulate actual"}
            or pa_value == "aa"
        ):
            acc_actual_row = row

    if acc_plan_row is None:
        raise OKDExportError("S-Curve Acc. Plan row was not found.")
    if acc_actual_row is None:
        raise OKDExportError("S-Curve Acc. Actual row was not found.")

    return acc_plan_row, acc_actual_row


def find_project_summary_rows(ws, headers: dict[str, int]) -> tuple[int, int, int, int]:
    row_type_col = find_column(headers, ("Row Type",), required=True)
    pa_col = find_column(headers, ("P/A",), required=True)

    plan_row: int | None = None
    actual_row: int | None = None

    for row in range(HEADER_ROW + 1, ws.max_row + 1):
        if normalize(ws.cell(row, row_type_col).value) != "project summary":
            continue

        pa_value = normalize(ws.cell(row, pa_col).value)
        if pa_value == "p" and plan_row is None:
            plan_row = row
        elif pa_value == "a" and actual_row is None:
            actual_row = row

    if plan_row is None:
        raise OKDExportError("Project Summary Plan row was not found.")

    if actual_row is None:
        actual_row = plan_row + 1
        if actual_row > ws.max_row:
            raise OKDExportError("Project Summary Actual row was not found.")

    return plan_row, actual_row


def read_schedule_items(ws, headers: dict[str, int], weeks, epoch):
    row_type_col = find_column(headers, ("Row Type",), required=True)
    pa_col = find_column(headers, ("P/A",), required=True)
    activity_id_col = find_column(headers, ("Activity ID",), required=True)
    description_col = find_column(
        headers,
        ("Description", "Activity Name", "Activity Description"),
        required=True,
    )
    wbs_col = find_column(
        headers,
        (
            "WBS",
            "WBS Code",
            "WBS Path",
            "WBS-1",
            "WBS 1",
        ),
    )
    amount_col = find_column(headers, ("Amount",), required=True)
    plan_start_col = find_column(headers, ("Plan Start",), required=True)
    plan_finish_col = find_column(headers, ("Plan Finish",), required=True)

    activities: list[dict[str, object]] = []
    table_items: list[dict[str, object]] = []
    current_wbs = ""
    wbs_sequencer = ActivityWbsSequencer()

    # Add the whole-project summary as the top row in progress_table.
    project_plan_row, project_actual_row = find_project_summary_rows(ws, headers)
    project_description = str(
        ws.cell(project_plan_row, description_col).value or "Project"
    ).strip()
    table_items.append(
        {
            "kind": "project",
            "wbs": "PROJECT",
            "display_wbs": "PROJECT",
            "activity_name": project_description,
            "plan_row": project_plan_row,
            "actual_row": project_actual_row,
        }
    )

    for row in range(HEADER_ROW + 1, ws.max_row + 1):
        row_type = ws.cell(row, row_type_col).value
        pa_value = normalize(ws.cell(row, pa_col).value)
        if pa_value != "p":
            continue

        source_wbs = (
            str(ws.cell(row, wbs_col).value or "").strip()
            if wbs_col is not None
            else ""
        )
        description = str(
            ws.cell(row, description_col).value or ""
        ).strip()

        if is_wbs_row_type(row_type):
            # WBS rows normally carry the WBS code. Remember it so that
            # following Activity rows with a blank WBS cell inherit it.
            if source_wbs:
                current_wbs = source_wbs

            if current_wbs:
                actual_row = find_actual_summary_row(
                    ws,
                    row,
                    pa_col,
                    row_type_col,
                    wbs_col,
                    description_col,
                )
                table_items.append(
                    {
                        "kind": "wbs",
                        "wbs": current_wbs,
                        "activity_name": description,
                        "plan_row": row,
                        "actual_row": actual_row,
                    }
                )
            continue

        if normalize(row_type) != "activity":
            continue

        activity_id = str(
            ws.cell(row, activity_id_col).value or ""
        ).strip()
        if not activity_id:
            continue

        activity_wbs = source_wbs or current_wbs

        # Activity rows from some P6 exports contain WBS=0 or a blank value.
        # Generate a readable child code under the current WBS so Plan and
        # Actual rows share the same stable display code (e.g. 1.1.1).
        parent_wbs = current_wbs or activity_wbs
        display_wbs = wbs_sequencer.next_code(
            parent_wbs,
            fallback=activity_id,
        )

        actual_row = find_actual_row(
            ws,
            row,
            activity_id,
            activity_id_col,
            pa_col,
        )

        plan_values: list[float | None] = []
        actual_values: list[float | None] = []

        for col, _ in weeks:
            plan_values.append(as_number(ws.cell(row, col).value))
            actual_values.append(
                as_number(ws.cell(actual_row, col).value)
                if actual_row is not None
                else None
            )

        activity = {
            "kind": "activity",
            "activity_id": activity_id,
            "activity_name": description,
            "wbs": activity_wbs,
            "display_wbs": display_wbs,
            "amount": as_number(ws.cell(row, amount_col).value) or 0.0,
            "plan_start": as_date(
                ws.cell(row, plan_start_col).value,
                epoch,
            ),
            "plan_finish": as_date(
                ws.cell(row, plan_finish_col).value,
                epoch,
            ),
            "plan": plan_values,
            "actual": actual_values,
            "plan_row": row,
            "actual_row": actual_row,
        }
        activities.append(activity)
        table_items.append(activity)

    if not activities:
        raise OKDExportError("No Activity Plan rows found.")

    return activities, table_items


def aggregate_wbs_item(
    item: dict[str, object],
    activities: list[dict[str, object]],
    week_count: int,
) -> dict[str, object]:
    parent_wbs = str(item["wbs"])
    children = [
        activity
        for activity in activities
        if belongs_to_wbs(str(activity["wbs"]), parent_wbs)
    ]

    if not children:
        return {
            **item,
            "amount": 0.0,
            "plan": [0.0] * week_count,
            "actual": [0.0] * week_count,
        }

    weights = [
        max(0.0, float(activity["amount"]))
        for activity in children
    ]
    total_amount = sum(weights)
    if total_amount <= 0:
        weights = [1.0] * len(children)
        divisor = float(len(children))
    else:
        divisor = total_amount

    plan_values: list[float] = []
    actual_values: list[float] = []

    for week_index in range(week_count):
        weighted_plan = 0.0
        weighted_actual = 0.0

        for activity, weight in zip(children, weights):
            plan_value = activity["plan"][week_index]
            actual_value = activity["actual"][week_index]

            if plan_value is not None:
                weighted_plan += weight * float(plan_value)
            if actual_value is not None:
                weighted_actual += weight * float(actual_value)

        plan_values.append(weighted_plan / divisor)
        actual_values.append(weighted_actual / divisor)

    return {
        **item,
        "amount": total_amount,
        "plan": plan_values,
        "actual": actual_values,
    }


def build_table_rows(
    table_items: list[dict[str, object]],
    activities: list[dict[str, object]],
    week_count: int,
) -> list[dict[str, object]]:
    """Keep source WBS and Activity rows in their original order.

    WBS totals already exist as formulas in the main worksheet. The OKD
    worksheet must reference those source rows directly so future edits in
    the main worksheet flow through automatically.
    """
    rows: list[dict[str, object]] = []
    seen_wbs: set[str] = set()

    for item in table_items:
        if item["kind"] == "project":
            rows.append(item)
        elif item["kind"] == "wbs":
            wbs_code = str(item["wbs"])
            if wbs_code in seen_wbs:
                continue
            seen_wbs.add(wbs_code)
            rows.append(
                aggregate_wbs_item(item, activities, week_count)
            )
        else:
            rows.append(item)

    return rows


def project_dates(ws, headers, activities, epoch) -> tuple[date, date]:
    row_type_col = find_column(headers, ("Row Type",), required=True)
    pa_col = find_column(headers, ("P/A",), required=True)
    plan_start_col = find_column(headers, ("Plan Start",), required=True)
    plan_finish_col = find_column(headers, ("Plan Finish",), required=True)

    for row in range(HEADER_ROW + 1, ws.max_row + 1):
        if (
            normalize(ws.cell(row, row_type_col).value) == "project summary"
            and normalize(ws.cell(row, pa_col).value) == "p"
        ):
            start = as_date(ws.cell(row, plan_start_col).value, epoch)
            finish = as_date(ws.cell(row, plan_finish_col).value, epoch)
            if start is not None and finish is not None:
                return start, finish

    starts = [
        item["plan_start"]
        for item in activities
        if isinstance(item["plan_start"], date)
    ]
    finishes = [
        item["plan_finish"]
        for item in activities
        if isinstance(item["plan_finish"], date)
    ]
    if not starts or not finishes:
        raise OKDExportError("Unable to determine project dates.")

    return min(starts), max(finishes)


def activity_weights(activities) -> tuple[list[float], float]:
    raw = [
        max(0.0, float(item["amount"]))
        for item in activities
    ]
    total = sum(raw)

    # Placeholder or actual BOQ amount can be used directly.
    # If every Amount is zero, fall back to equal activity weights.
    if total <= 0:
        raw = [1.0] * len(activities)
        total = float(len(activities))

    return raw, total


def cumulative_project_series(activities, week_count: int):
    weights, total_weight = activity_weights(activities)

    weekly_plan: list[float] = []
    weekly_actual: list[float | None] = []

    for index in range(week_count):
        plan_weighted = 0.0
        actual_weighted = 0.0
        actual_has_data = False

        for item, weight in zip(activities, weights):
            plan_value = item["plan"][index]
            actual_value = item["actual"][index]

            if plan_value is not None:
                plan_weighted += weight * plan_value

            if actual_value is not None:
                actual_weighted += weight * actual_value
                actual_has_data = True

        weekly_plan.append(plan_weighted / total_weight)
        weekly_actual.append(
            actual_weighted / total_weight
            if actual_has_data
            else None
        )

    plan_cumulative: list[float] = []
    running_plan = 0.0
    for value in weekly_plan:
        running_plan += value
        plan_cumulative.append(running_plan * 100.0)

    actual_indices = [
        index
        for index, value in enumerate(weekly_actual)
        if value is not None
    ]
    first_actual = min(actual_indices) if actual_indices else None
    last_actual = max(actual_indices) if actual_indices else None

    actual_cumulative: list[float | None] = []
    running_actual = 0.0
    for index, value in enumerate(weekly_actual):
        if first_actual is None or index < first_actual or index > last_actual:
            actual_cumulative.append(None)
            continue

        if value is not None:
            running_actual += value
        actual_cumulative.append(running_actual * 100.0)

    # Avoid display noise such as 100.0000000003.
    plan_cumulative = [round(value, 6) for value in plan_cumulative]
    actual_cumulative = [
        round(value, 6) if value is not None else None
        for value in actual_cumulative
    ]

    return plan_cumulative, actual_cumulative


def remove_existing_sheet(wb, name: str) -> None:
    if name in wb.sheetnames:
        del wb[name]


def style_header(ws, cell_range: str) -> None:
    fill = PatternFill("solid", fgColor=HEADER_FILL)
    font = Font(bold=True, color=HEADER_FONT)
    for row in ws[cell_range]:
        for cell in row:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )


def build_progress_sheet(
    wb,
    source_ws,
    headers: dict[str, int],
    weeks,
    activities,
) -> None:
    """Build an interactive progress sheet using live Excel formulas."""
    remove_existing_sheet(wb, OKD_PROGRESS_SHEET)
    ws = wb.create_sheet(OKD_PROGRESS_SHEET)

    ws.append([
        "project_start",
        "project_finish",
        "week_start",
        "plan",
        "actual",
    ])

    project_plan_row, _ = find_project_summary_rows(source_ws, headers)
    acc_plan_row, acc_actual_row = find_scurve_cumulative_rows(
        source_ws,
        headers,
    )
    plan_start_col = find_column(headers, ("Plan Start",), required=True)
    plan_finish_col = find_column(headers, ("Plan Finish",), required=True)
    source_ref = excel_sheet_ref(source_ws.title)

    for output_row, (source_col, _) in enumerate(weeks, start=2):
        source_col_letter = get_column_letter(source_col)
        plan_start_addr = f"{get_column_letter(plan_start_col)}{project_plan_row}"
        plan_finish_addr = f"{get_column_letter(plan_finish_col)}{project_plan_row}"

        ws.cell(output_row, 1, f"={source_ref}!{plan_start_addr}")
        ws.cell(output_row, 2, f"={source_ref}!{plan_finish_addr}")
        ws.cell(output_row, 3, f"={source_ref}!{source_col_letter}{HEADER_ROW}")
        ws.cell(
            output_row,
            4,
            f'=IF({source_ref}!{source_col_letter}{acc_plan_row}="","",'
            f'ROUND({source_ref}!{source_col_letter}{acc_plan_row}*100,6))',
        )
        ws.cell(
            output_row,
            5,
            f'=IF({source_ref}!{source_col_letter}{acc_actual_row}="","",'
            f'ROUND({source_ref}!{source_col_letter}{acc_actual_row}*100,6))',
        )

    style_header(ws, "A1:E1")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:E{ws.max_row}"

    for row in range(2, ws.max_row + 1):
        for col in (1, 2, 3):
            ws.cell(row, col).number_format = "yyyy-mm-dd"
        ws.cell(row, 4).number_format = "0.00"
        ws.cell(row, 5).number_format = "0.00"

    for letter, width in {
        "A": 16, "B": 16, "C": 14, "D": 12, "E": 12
    }.items():
        ws.column_dimensions[letter].width = width


def verify_progress_table_links(
    ws,
    weeks,
    table_rows,
) -> tuple[int, int]:
    """Verify that every item has Plan and Actual rows and live formulas."""
    expected_rows = len(table_rows) * 2
    if ws.max_row - 1 != expected_rows:
        raise OKDExportError(
            f"Row verification failed: expected {expected_rows}, "
            f"found {ws.max_row - 1}"
        )

    checked = 0
    for row in range(2, ws.max_row + 1):
        if ws.cell(row, 4).value not in {"P", "A"}:
            raise OKDExportError(
                f"Invalid P/A value at {OKD_TABLE_SHEET}!D{row}"
            )

        for col in range(3, 6 + len(weeks)):
            value = ws.cell(row, col).value
            if col != 4 and not (
                isinstance(value, str) and value.startswith("=")
            ):
                raise OKDExportError(
                    f"Live formula missing at "
                    f"{OKD_TABLE_SHEET}!{get_column_letter(col)}{row}"
                )
            checked += 1

    return expected_rows, checked


def build_progress_table_sheet(
    wb,
    source_ws,
    headers: dict[str, int],
    weeks,
    table_rows,
) -> tuple[int, int]:
    """Build an interactive progress_table with Plan and Actual for all rows."""
    remove_existing_sheet(wb, OKD_TABLE_SHEET)
    ws = wb.create_sheet(OKD_TABLE_SHEET)

    ws.append(["WBS", "Activities", "Amount", "P/A", "%Progress"] + [
        f"={excel_sheet_ref(source_ws.title)}!"
        f"{get_column_letter(source_col)}{HEADER_ROW}"
        for source_col, _ in weeks
    ])

    # Header cells are live links to the date headers in the main worksheet.
    # Without an explicit date number format, Excel displays their serial
    # values (for example 46045) instead of readable dates.
    for output_col in range(6, 6 + len(weeks)):
        ws.cell(1, output_col).number_format = "dd/mm/yyyy"

    amount_col = find_column(headers, ("Amount",), required=True)
    description_col = find_column(
        headers,
        ("Description", "Activity Name", "Activity Description"),
        required=True,
    )
    source_ref = excel_sheet_ref(source_ws.title)

    output_row = 2
    for item in table_rows:
        display_wbs = str(
            item.get("display_wbs") or item.get("wbs") or ""
        )
        plan_row = int(item["plan_row"])
        actual_row = item.get("actual_row")
        actual_source_row = (
            int(actual_row) if actual_row is not None else plan_row
        )

        for pa_value, source_row in (
            ("P", plan_row),
            ("A", actual_source_row),
        ):
            ws.cell(output_row, 1, display_wbs)
            ws.cell(
                output_row,
                2,
                f"={source_ref}!{get_column_letter(description_col)}{plan_row}",
            )
            ws.cell(
                output_row,
                3,
                f"={source_ref}!{get_column_letter(amount_col)}{plan_row}",
            )
            ws.cell(output_row, 4, pa_value)

            first_week_letter = get_column_letter(6)
            last_week_letter = get_column_letter(5 + len(weeks))
            ws.cell(
                output_row,
                5,
                f'=IF(COUNT({first_week_letter}{output_row}:'
                f'{last_week_letter}{output_row})=0,"",'
                f'ROUND(SUM({first_week_letter}{output_row}:'
                f'{last_week_letter}{output_row}),6))',
            )

            for week_offset, (source_col, _) in enumerate(weeks, start=6):
                source_col_letter = get_column_letter(source_col)
                ws.cell(
                    output_row,
                    week_offset,
                    f'=IF({source_ref}!{source_col_letter}{source_row}="","",'
                    f'ROUND({source_ref}!{source_col_letter}{source_row}*100,6))',
                )

            ws.cell(output_row, 3).number_format = '#,##0.00'
            ws.cell(output_row, 5).number_format = "0.00"
            for output_col in range(6, 6 + len(weeks)):
                ws.cell(output_row, output_col).number_format = "0.00"
            output_row += 1

    visible_last_col = get_column_letter(5 + len(weeks))
    first_timescale_col = get_column_letter(6)
    last_row = ws.max_row

    header_fill = PatternFill("solid", fgColor=HEADER_FILL)
    header_font = Font(bold=True, color=HEADER_FONT)
    thin_side = Side(style="thin", color=BORDER_COLOR)
    thin_border = Border(
        left=thin_side, right=thin_side, top=thin_side, bottom=thin_side
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for row in ws.iter_rows(
        min_row=2, max_row=last_row, min_col=1, max_col=5 + len(weeks)
    ):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(
                vertical="center", wrap_text=(cell.column == 2)
            )

    timescale_range = f"{first_timescale_col}2:{visible_last_col}{last_row}"
    ws.conditional_formatting.add(
        timescale_range,
        FormulaRule(
            formula=[f'=AND($D2="P",{first_timescale_col}2<>"")'],
            fill=PatternFill("solid", fgColor=PLAN_FILL),
            font=Font(color=PLAN_FONT),
            stopIfTrue=True,
        ),
    )
    ws.conditional_formatting.add(
        timescale_range,
        FormulaRule(
            formula=[f'=AND($D2="A",{first_timescale_col}2<>"")'],
            fill=PatternFill("solid", fgColor=ACTUAL_FILL),
            font=Font(color=ACTUAL_FONT),
            stopIfTrue=True,
        ),
    )

    ws.freeze_panes = "F2"
    ws.auto_filter.ref = f"A1:{visible_last_col}{last_row}"
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 46
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 14
    for output_col in range(6, 6 + len(weeks)):
        ws.column_dimensions[get_column_letter(output_col)].width = 12
    ws.row_dimensions[1].height = 24

    return verify_progress_table_links(ws, weeks, table_rows)


def build_progress_views_from_source(wb, source_ws) -> tuple[int, int, int, int]:
    """Build both ``progress`` and ``progress_table`` from the current main sheet.

    This is the shared pre-mapping workbook contract.  The final OKD step may
    call the same builders again as a deterministic refresh after distribution.
    """
    headers = get_headers(source_ws)
    weeks = find_week_columns(source_ws, wb.epoch)
    activities, table_items = read_schedule_items(
        source_ws, headers, weeks, wb.epoch
    )
    table_rows = build_table_rows(table_items, activities, len(weeks))
    build_progress_sheet(wb, source_ws, headers, weeks, activities)
    table_rows_count, checked_links = build_progress_table_sheet(
        wb, source_ws, headers, weeks, table_rows
    )
    return len(activities), len(weeks), table_rows_count, checked_links


def build_progress_table_from_source(wb, source_ws) -> tuple[int, int]:
    """Backward-compatible helper that rebuilds only ``progress_table``."""
    headers = get_headers(source_ws)
    weeks = find_week_columns(source_ws, wb.epoch)
    activities, table_items = read_schedule_items(
        source_ws, headers, weeks, wb.epoch
    )
    table_rows = build_table_rows(table_items, activities, len(weeks))
    return build_progress_table_sheet(wb, source_ws, headers, weeks, table_rows)


def update_info_sheet(wb, activities_count: int, weeks_count: int) -> None:
    ws = wb["Info"] if "Info" in wb.sheetnames else wb.create_sheet("Info")
    start = ws.max_row + 2 if ws.max_row else 1

    rows = [
        ("Interactive Export", "READY"),
        ("OKD Sheets", "progress, progress_table"),
        ("OKD Activities", activities_count),
        ("OKD Weeks", weeks_count),
        ("Value Scale", "Formula result 0-100 (not Excel %)"),
        ("OKD WBS Rollup", "Included"),
        ("Calculation", "Live formulas linked to main"),
        ("WBS Rows", "Plan and Actual created for every WBS/activity"),
        ("OKD Child Codes", "Generated under parent WBS"),
    ]
    for offset, (label, value) in enumerate(rows):
        ws.cell(start + offset, 1, label)
        ws.cell(start + offset, 2, value)

    ws.column_dimensions["A"].width = max(
        ws.column_dimensions["A"].width or 0,
        28,
    )
    ws.column_dimensions["B"].width = max(
        ws.column_dimensions["B"].width or 0,
        58,
    )


def build_okd_sheets(
    input_file: Path,
    output_file: Path,
    *,
    source_sheet: str,
) -> tuple[int, int]:
    wb = load_workbook(input_file, data_only=False)

    if source_sheet not in wb.sheetnames:
        wb.close()
        raise OKDExportError(
            f"Source worksheet not found: {source_sheet}"
        )

    ws = wb[source_sheet]
    headers = get_headers(ws)
    weeks = find_week_columns(ws, wb.epoch)
    activities, table_items = read_schedule_items(
        ws, headers, weeks, wb.epoch
    )
    table_rows = build_table_rows(
        table_items, activities, len(weeks)
    )
    build_progress_sheet(
        wb,
        ws,
        headers,
        weeks,
        activities,
    )
    table_rows_count, checked_links = build_progress_table_sheet(
        wb,
        ws,
        headers,
        weeks,
        table_rows,
    )
    update_info_sheet(wb, len(activities), len(weeks))

    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True

    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_file)
    wb.close()

    return len(activities), len(weeks), table_rows_count, checked_links
