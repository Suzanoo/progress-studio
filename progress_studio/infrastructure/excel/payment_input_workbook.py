from __future__ import annotations

from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from progress_studio.domain.payment_models import PaymentInputData, PaymentInputResult, PaymentInputValidation
from progress_studio.infrastructure.excel.payment_input_reader import PaymentInputSparseReader
from progress_studio.infrastructure.excel.payment_workbook import PaymentWorkbookError, PaymentWorkbookSnapshotter


class PaymentInputWorkbook:
    """Create/validate the small user-editable Payment Requirement payload."""

    SHEET = "Payment Input"
    HEADER_ROW = 6
    DATE_ROW = 7
    FIRST_ACTIVITY_ROW = 8

    def __init__(
        self,
        snapshotter: PaymentWorkbookSnapshotter | None = None,
        reader: PaymentInputSparseReader | None = None,
    ) -> None:
        self.snapshotter = snapshotter or PaymentWorkbookSnapshotter()
        self.reader = reader or PaymentInputSparseReader()

    def create(self, source: Path, output: Path, periods: int) -> PaymentInputResult:
        source = Path(source)
        output = Path(output)
        if periods < 1 or periods > 120:
            raise PaymentWorkbookError("Payment periods must be between 1 and 120.")

        validation = self.snapshotter.validate(source)
        if validation.project_start is None or validation.project_finish is None:
            raise PaymentWorkbookError("Project Start / Finish could not be read from the main sheet.")

        tree_rows = self.snapshotter.payment_tree_rows(source)
        activity_rows = [row for row in tree_rows if row["row_type"] == "ACT"]
        if not activity_rows:
            raise PaymentWorkbookError("No Activity rows were found in the main sheet.")
        payment_dates = self._spread_dates(validation.project_start, validation.project_finish, periods)

        wb = Workbook()
        ws = wb.active
        ws.title = self.SHEET
        ws.sheet_view.showGridLines = False
        ws.sheet_properties.outlinePr.summaryBelow = False

        ws["A1"] = "Payment Requirement Input"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"] = "Project Start"
        ws["B2"] = validation.project_start
        ws["A3"] = "Project Finish"
        ws["B3"] = validation.project_finish
        ws["A4"] = "Payment Periods"
        ws["B4"] = periods
        ws["B2"].number_format = ws["B3"].number_format = "dd-mmm-yy"

        fixed_headers = ("Type", "WBS", "Activity ID", "Activity Name")
        for col, label in enumerate(fixed_headers, start=1):
            ws.cell(self.HEADER_ROW, col, label)
        ws.cell(self.DATE_ROW, 1, "Payment Date")

        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        for idx, payment_date in enumerate(payment_dates, start=1):
            col = idx + len(fixed_headers)
            ws.cell(self.HEADER_ROW, col, f"P{idx:02d}")
            ws.cell(self.DATE_ROW, col, payment_date)
            ws.cell(self.DATE_ROW, col).number_format = "dd-mmm-yy"
            ws.column_dimensions[get_column_letter(col)].width = 11
        for cell in ws[self.HEADER_ROW]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        ws.cell(self.DATE_ROW, 1).font = Font(bold=True)

        wbs_fills = {
            1: PatternFill("solid", fgColor="F4B183"),
            2: PatternFill("solid", fgColor="F8CBAD"),
            3: PatternFill("solid", fgColor="FCE4D6"),
            4: PatternFill("solid", fgColor="FFF2CC"),
        }
        default_wbs_fill = PatternFill("solid", fgColor="F2F2F2")

        out_row = self.FIRST_ACTIVITY_ROW
        for item in tree_rows:
            level = max(int(item.get("outline_level") or 0), 0)
            ws.row_dimensions[out_row].outlineLevel = min(level, 7)
            if item["row_type"] == "WBS":
                ws.cell(out_row, 1, "WBS")
                ws.cell(out_row, 2, item["wbs"])
                ws.cell(out_row, 4, item["activity_name"])
                fill = wbs_fills.get(level, default_wbs_fill)
                for col in range(1, periods + len(fixed_headers) + 1):
                    cell = ws.cell(out_row, col)
                    cell.fill = fill
                    cell.font = Font(bold=True)
                ws.cell(out_row, 4).alignment = Alignment(indent=max(level - 1, 0))
            else:
                ws.cell(out_row, 1, "ACT")
                ws.cell(out_row, 2, item["wbs"])
                ws.cell(out_row, 3, item["activity_id"])
                ws.cell(out_row, 4, item["activity_name"])
                ws.cell(out_row, 4).alignment = Alignment(indent=max(level - 1, 0))
                fake_values = self._fake_requirements(item["weekly_plan"], payment_dates, validation.project_start)
                for idx, value in enumerate(fake_values, start=1):
                    cell = ws.cell(out_row, idx + len(fixed_headers))
                    cell.number_format = "0%"
                    if value is not None:
                        cell.value = value
            out_row += 1

        ws.freeze_panes = "E8"
        # Intentionally no AutoFilter: hierarchy rows should stay visually attached to their children.
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 42
        output.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output)
        wb.close()

        return PaymentInputResult(
            source_workbook=source,
            output_workbook=output,
            payment_periods=periods,
            activity_rows=len(activity_rows),
            project_start=validation.project_start,
            project_finish=validation.project_finish,
        )

    def embed(
        self,
        workbook,
        preserved: PaymentInputData | None = None,
        periods: int | None = None,
    ) -> dict[str, int]:
        """Rebuild the embedded Payment Input from current ``main``.

        Existing user-entered requirements are preserved by Activity ID + period.
        New activities receive suggested fake requirements; deleted activities are
        dropped and counted in the reconciliation note.
        """
        if "main" not in workbook.sheetnames:
            raise PaymentWorkbookError("Worksheet 'main' was not found.")

        main = workbook["main"]
        model = self._model_from_main_sheet(main)
        project_start = model["project_start"]
        project_finish = model["project_finish"]
        tree_rows = model["tree_rows"]
        if project_start is None or project_finish is None:
            raise PaymentWorkbookError("Project Start / Finish could not be read from the main sheet.")

        activity_ids = [row["activity_id"] for row in tree_rows if row["row_type"] == "ACT"]
        if not activity_ids:
            raise PaymentWorkbookError("No Activity rows were found in the main sheet.")

        preserved_ids = set(preserved.activity_ids) if preserved is not None else set()
        if periods is None:
            periods = len(preserved.periods) if preserved is not None else max(
                (project_finish.year - project_start.year) * 12 + project_finish.month - project_start.month,
                1,
            )
        periods = max(min(int(periods), 120), 1)

        preserved_values: dict[tuple[str, str], float] = {}
        if preserved is not None:
            for period in preserved.periods:
                for req in period.requirements:
                    preserved_values[(req.activity_id, period.period_id)] = req.required_fraction

        if self.SHEET in workbook.sheetnames:
            del workbook[self.SHEET]
        main_index = workbook.sheetnames.index("main")
        ws = workbook.create_sheet(self.SHEET, index=main_index + 1)
        ws.sheet_view.showGridLines = False
        ws.sheet_properties.outlinePr.summaryBelow = False

        ws["A1"] = "Payment Requirement Input"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"] = "Project Start"
        ws["B2"] = project_start
        ws["A3"] = "Project Finish"
        ws["B3"] = project_finish
        ws["A4"] = "Payment Periods"
        ws["B4"] = periods
        ws["B2"].number_format = ws["B3"].number_format = "dd-mmm-yy"
        # Payment Date is intentionally no longer an input. Keep row 7 reserved
        # so older Payment Input files remain readable by the sparse reader.
        ws["A5"] = "Eligible dates are calculated from the latest required Activity point."
        ws["A5"].font = Font(italic=True, color="666666")

        fixed_headers = ("Type", "WBS", "Activity ID", "Activity Name")
        for col, label in enumerate(fixed_headers, start=1):
            ws.cell(self.HEADER_ROW, col, label)
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        for idx in range(1, periods + 1):
            col = idx + len(fixed_headers)
            ws.cell(self.HEADER_ROW, col, f"P{idx:02d}")
            ws.column_dimensions[get_column_letter(col)].width = 11
        for cell in ws[self.HEADER_ROW]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        wbs_fills = {
            1: PatternFill("solid", fgColor="F4B183"),
            2: PatternFill("solid", fgColor="F8CBAD"),
            3: PatternFill("solid", fgColor="FCE4D6"),
            4: PatternFill("solid", fgColor="FFF2CC"),
        }
        default_wbs_fill = PatternFill("solid", fgColor="F2F2F2")
        out_row = self.FIRST_ACTIVITY_ROW
        new_count = preserved_count = 0
        for item in tree_rows:
            level = max(int(item.get("outline_level") or 0), 0)
            ws.row_dimensions[out_row].outlineLevel = min(level, 7)
            if item["row_type"] == "WBS":
                ws.cell(out_row, 1, "WBS")
                ws.cell(out_row, 2, item["wbs"])
                ws.cell(out_row, 4, item["activity_name"])
                fill = wbs_fills.get(level, default_wbs_fill)
                for col in range(1, periods + len(fixed_headers) + 1):
                    cell = ws.cell(out_row, col)
                    cell.fill = fill
                    cell.font = Font(bold=True)
                ws.cell(out_row, 4).alignment = Alignment(indent=max(level - 1, 0))
            else:
                activity_id = item["activity_id"]
                ws.cell(out_row, 1, "ACT")
                ws.cell(out_row, 2, item["wbs"])
                ws.cell(out_row, 3, activity_id)
                ws.cell(out_row, 4, item["activity_name"])
                ws.cell(out_row, 4).alignment = Alignment(indent=max(level - 1, 0))
                if activity_id in preserved_ids:
                    preserved_count += 1
                    values = [
                        preserved_values.get((activity_id, f"P{idx:02d}"))
                        for idx in range(1, periods + 1)
                    ]
                else:
                    new_count += 1
                    payment_cuts = self._spread_dates(project_start, project_finish, periods)
                    values = self._fake_requirements(item["weekly_plan"], payment_cuts, project_start)
                for idx, value in enumerate(values, start=1):
                    cell = ws.cell(out_row, idx + len(fixed_headers))
                    cell.number_format = "0%"
                    if value is not None:
                        cell.value = value
            out_row += 1

        removed_count = len(preserved_ids - set(activity_ids))
        ws["D5"] = (
            f"Reconcile: {preserved_count} preserved • {new_count} new • "
            f"{removed_count} removed"
        )
        ws["D5"].font = Font(italic=True, color="666666")
        ws.freeze_panes = "E8"
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 42
        return {
            "periods": periods,
            "activities": len(activity_ids),
            "preserved": preserved_count,
            "new": new_count,
            "removed": removed_count,
        }

    @staticmethod
    def _model_from_main_sheet(ws: Worksheet) -> dict:
        header_row = None
        headers: dict[str, int] = {}
        for row in range(1, min(ws.max_row, 30) + 1):
            current = {
                str(ws.cell(row, col).value or "").strip().lower(): col
                for col in range(1, ws.max_column + 1)
                if str(ws.cell(row, col).value or "").strip()
            }
            if {"row type", "wbs", "description", "p/a", "activity id", "outline level", "plan start", "plan finish"}.issubset(current):
                header_row = row
                headers = current
                break
        if header_row is None:
            raise PaymentWorkbookError("Required main headers were not found.")

        fixed_end = headers.get("xml amount", max(headers.values()))
        timescale = []
        for col in range(fixed_end + 1, ws.max_column + 1):
            value = ws.cell(header_row, col).value
            if isinstance(value, date):
                timescale.append((col, value.date() if hasattr(value, "date") else value))

        project_start = project_finish = None
        tree_rows = []
        for row in range(header_row + 1, ws.max_row + 1):
            row_type = str(ws.cell(row, headers["row type"]).value or "").strip().lower()
            pa = str(ws.cell(row, headers["p/a"]).value or "").strip().upper()
            if pa != "P":
                continue
            if row_type == "project summary":
                start = ws.cell(row, headers["plan start"]).value
                finish = ws.cell(row, headers["plan finish"]).value
                if isinstance(start, date):
                    project_start = start.date() if hasattr(start, "date") else start
                if isinstance(finish, date):
                    project_finish = finish.date() if hasattr(finish, "date") else finish
                continue
            if row_type not in {"wbs", "activity"}:
                continue
            item = {
                "row_type": "WBS" if row_type == "wbs" else "ACT",
                "wbs": str(ws.cell(row, headers["wbs"]).value or "").strip(),
                "activity_id": str(ws.cell(row, headers["activity id"]).value or "").strip(),
                "activity_name": str(ws.cell(row, headers["description"]).value or "").strip(),
                "outline_level": int(ws.cell(row, headers["outline level"]).value or 0),
                "weekly_plan": (),
            }
            if row_type == "activity":
                weekly = []
                for col, week_start in timescale:
                    value = ws.cell(row, col).value
                    if isinstance(value, (int, float)) and value != 0:
                        weekly.append((week_start, float(value)))
                item["weekly_plan"] = tuple(weekly)
            tree_rows.append(item)

        return {
            "project_start": project_start,
            "project_finish": project_finish,
            "tree_rows": tree_rows,
        }

    def validate(self, payment_workbook: Path, progress_workbook: Path | None = None) -> PaymentInputValidation:
        data = self.reader.read(Path(payment_workbook))

        matched = len(data.activity_ids)
        missing = 0
        if progress_workbook is not None:
            progress_ids = set(self.snapshotter.activity_ids(Path(progress_workbook)))
            matched = sum(1 for activity_id in data.activity_ids if activity_id in progress_ids)
            missing = len(data.activity_ids) - matched

        return PaymentInputValidation(
            workbook=data.workbook,
            payment_sheet=data.sheet,
            payment_periods=len(data.periods),
            activity_rows=len(data.activity_ids),
            matched_activities=matched,
            missing_activities=missing,
            populated_requirements=data.populated_requirements,
        )

    @staticmethod
    def _fake_requirements(weekly_plan, payment_dates: list[date], project_start: date) -> list[float | None]:
        """Suggested cumulative requirements, but only for periods with planned movement.

        A period gets a value only when the activity has incremental Plan progress
        after the previous payment cut and on/before the current cut. This keeps
        the generated workbook sparse while ensuring short activities are not lost.
        """
        points = sorted((d, float(v)) for d, v in weekly_plan if v is not None)
        result: list[float | None] = []
        cumulative = 0.0
        previous = project_start - timedelta(days=1)
        point_index = 0
        for cut in payment_dates:
            moved = False
            while point_index < len(points) and points[point_index][0] <= cut:
                d, value = points[point_index]
                cumulative += value
                if d > previous and value != 0:
                    moved = True
                point_index += 1
            result.append(min(max(cumulative, 0.0), 1.0) if moved else None)
            previous = cut
        return result

    @staticmethod
    def _spread_dates(start: date, finish: date, periods: int) -> list[date]:
        total_days = max((finish - start).days, 1)
        return [start + timedelta(days=round(total_days * idx / periods)) for idx in range(1, periods + 1)]
