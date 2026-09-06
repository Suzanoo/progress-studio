from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime

from openpyxl.chart import LineChart, Reference
from openpyxl.chart.error_bar import ErrorBars
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.legend import LegendEntry
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter


from progress_studio.domain.earned_value import (
    ActivityEarnedValue,
    BOQEarnedValue,
    EarnedValuePoint,
    EarnedValueResult,
)
from progress_studio.infrastructure.excel.dashboard_workbook import (
    DATA_SHEET as DASHBOARD_DATA_SHEET,
    _COLORS as DASHBOARD_COLORS,
    _FONT as DASHBOARD_FONT,
    _LAYOUT as DASHBOARD_LAYOUT,
    _date_axis_for_line_chart,
)


EARNED_VALUE_SHEET = "Earned Value"
EV_TABLE_SHEET = "EV Table"
EV_DATA_SHEET = "EV_Data"
_MAPPING_SHEET = "BOQ Activity Mapping"

_FONT = DASHBOARD_FONT
_NAVY = DASHBOARD_COLORS["navy"]
_BLUE = DASHBOARD_COLORS["blue"]
_GREEN = DASHBOARD_COLORS["green"]
_RED = DASHBOARD_COLORS["red"]
_AMBER = DASHBOARD_COLORS["amber"]
_LIGHT_BLUE = DASHBOARD_COLORS["light_blue"]
_LIGHT_GREEN = DASHBOARD_COLORS["light_green"]
_LIGHT_RED = DASHBOARD_COLORS["light_red"]
_LIGHT_AMBER = DASHBOARD_COLORS["light_amber"]
_LIGHT_GRAY = DASHBOARD_COLORS["light_gray"]
_BORDER = DASHBOARD_COLORS["border"]
_TEXT = DASHBOARD_COLORS["text"]
_MUTED = DASHBOARD_COLORS["muted"]
_WHITE = DASHBOARD_COLORS["white"]


@dataclass(frozen=True, slots=True)
class _ChartPoint:
    reporting_date: datetime
    planned_value: float | None
    earned_value: float | None
    point_type: str


@dataclass(frozen=True, slots=True)
class _PerformanceRow:
    label: str
    bac: float
    planned_value: float | None
    earned_value: float | None
    schedule_variance: float | None
    schedule_performance_index: float | None
    name: str = ""
    activity_id: str = ""
    order: int = 0


@dataclass(frozen=True, slots=True)
class _MainWBS:
    code: str
    name: str
    order: int


def _solid(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def _thin_border() -> Border:
    side = Side(style="thin", color=_BORDER)
    return Border(left=side, right=side, top=side, bottom=side)


def _as_date(value: datetime | date | None) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    return value


def _as_datetime(value: datetime | date | None) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    return None


def _point_on_or_before(
    points: tuple[EarnedValuePoint, ...],
    cutoff: datetime | date | None,
    *,
    require_ev: bool = False,
) -> EarnedValuePoint | None:
    cutoff_date = _as_date(cutoff)
    latest: EarnedValuePoint | None = None
    for point in points:
        reporting_date = _as_date(point.reporting_date)
        if reporting_date is None:
            continue
        if cutoff_date is not None and reporting_date > cutoff_date:
            continue
        if require_ev and point.earned_value is None:
            continue
        if latest is None or reporting_date >= _as_date(latest.reporting_date):
            latest = point
    return latest


def _monthly_chart_points(result: EarnedValueResult) -> tuple[_ChartPoint, ...]:
    """Sample EV-1 weekly cumulative points for a monthly management view.

    No interpolation occurs. For each calendar month we retain the latest source
    point in that month. The cutoff month also receives an exact Status Date
    point whose values are the latest proven weekly cumulative values at or
    before cutoff. This preserves the weekly source contract while giving the
    chart a true management reporting marker.
    """
    source = tuple(
        sorted(
            (
                point
                for point in result.project_points
                if _as_datetime(point.reporting_date) is not None
            ),
            key=lambda point: _as_datetime(point.reporting_date),
        )
    )
    if not source:
        return ()

    cutoff = _as_datetime(result.cutoff_date)
    cutoff_date = cutoff.date() if cutoff is not None else None

    by_month: dict[tuple[int, int], list[EarnedValuePoint]] = defaultdict(list)
    for point in source:
        reporting = _as_datetime(point.reporting_date)
        by_month[(reporting.year, reporting.month)].append(point)

    sampled: list[_ChartPoint] = []
    cutoff_month = None if cutoff is None else (cutoff.year, cutoff.month)

    for month_key in sorted(by_month):
        month_points = by_month[month_key]
        latest = month_points[-1]

        if month_key != cutoff_month:
            sampled.append(
                _ChartPoint(
                    reporting_date=_as_datetime(latest.reporting_date),
                    planned_value=latest.planned_value,
                    earned_value=latest.earned_value,
                    point_type="Monthly",
                )
            )
            continue

        status_source = _point_on_or_before(
            tuple(month_points),
            cutoff,
            require_ev=True,
        )
        if status_source is not None:
            sampled.append(
                _ChartPoint(
                    reporting_date=cutoff,
                    planned_value=status_source.planned_value,
                    earned_value=status_source.earned_value,
                    point_type="Status Date",
                )
            )

        latest_date = _as_date(latest.reporting_date)
        if (
            cutoff_date is not None
            and latest_date is not None
            and latest_date > cutoff_date
        ):
            sampled.append(
                _ChartPoint(
                    reporting_date=_as_datetime(latest.reporting_date),
                    planned_value=latest.planned_value,
                    earned_value=None,
                    point_type="Monthly",
                )
            )
        elif status_source is None:
            sampled.append(
                _ChartPoint(
                    reporting_date=_as_datetime(latest.reporting_date),
                    planned_value=latest.planned_value,
                    earned_value=latest.earned_value,
                    point_type="Monthly",
                )
            )

    sampled.sort(key=lambda point: point.reporting_date)
    return tuple(sampled)


def _cutoff_options(result: EarnedValueResult) -> tuple[datetime, ...]:
    """Return valid EV Status Dates up to the authoritative reporting cutoff.

    Reuse the same monthly reporting convention as Dashboard/main_monthly: the
    final real weekly reporting point inside each calendar month.  EV may move
    retrospectively to an earlier Status Date in Excel; a future Status Date must
    first become the authoritative project cutoff and then be rebuilt by Python.
    """
    cutoff = _as_datetime(result.cutoff_date)
    by_month: dict[tuple[int, int], datetime] = {}
    for point in result.project_points:
        reporting = _as_datetime(point.reporting_date)
        if reporting is None:
            continue
        if cutoff is not None and reporting > cutoff:
            continue
        by_month[(reporting.year, reporting.month)] = reporting

    values = set(by_month.values())
    if cutoff is not None:
        values.add(cutoff)
    return tuple(sorted(values))


def _dashboard_monthly_cutoff_range(workbook, result: EarnedValueResult) -> tuple[str, int] | None:
    """Return the complete canonical Dashboard_Data monthly cutoff list.

    Dashboard_Data!K is already the monthly reporting calendar used by the
    existing Progress Dashboard/main_monthly views. EV consumes that calendar
    as-is; its selected Status Date must not truncate or redefine the list.
    """
    if DASHBOARD_DATA_SHEET not in workbook.sheetnames:
        return None
    ws = workbook[DASHBOARD_DATA_SHEET]
    last_row = 1
    for row in range(2, ws.max_row + 1):
        if _as_datetime(ws.cell(row, 11).value) is not None:
            last_row = row
    if last_row < 2:
        return None
    return (DASHBOARD_DATA_SHEET, last_row)

def _add_cutoff_dropdown(workbook, ws, data_ws, result: EarnedValueResult) -> None:
    """Add the EV Status Date selector using the existing monthly cutoff source."""
    options = _cutoff_options(result)

    ws["L3"] = "Status Date"
    ws["L3"].fill = _solid(_NAVY)
    ws["L3"].font = Font(name=_FONT, size=10, bold=True, color=_WHITE)
    ws["L3"].alignment = Alignment(horizontal="right", vertical="center")
    ws["M3"] = result.cutoff_date
    ws["M3"].number_format = "dd-mmm-yyyy"
    ws["M3"].fill = _solid(_LIGHT_BLUE)
    ws["M3"].font = Font(name=_FONT, size=10, bold=True, color=_TEXT)
    ws["M3"].alignment = Alignment(horizontal="center", vertical="center")
    ws["M3"].comment = Comment(
        "This Earned Value Status Date updates KPI, chart, WBS and variance views in Excel. "
        "Choose a later project reporting cutoff in the normal Progress Studio controls and rebuild EV first.",
        "Progress Studio",
    )

    dashboard_range = _dashboard_monthly_cutoff_range(workbook, result)
    if dashboard_range is not None:
        sheet_name, last_row = dashboard_range
        formula1 = f'=INDIRECT("{sheet_name}!$K$2:$K${last_row}")'
    elif options:
        formula1 = f'=INDIRECT("{EV_DATA_SHEET}!$H$2:$H${len(options) + 1}")'
    else:
        return

    validation = DataValidation(type="list", formula1=formula1, allow_blank=False)
    validation.promptTitle = "Status Date"
    validation.prompt = "Choose the Earned Value reporting cutoff date."
    validation.errorTitle = "Invalid Status Date"
    validation.error = "Select a date from the list."
    validation.errorStyle = "stop"
    validation.showInputMessage = True
    validation.showErrorMessage = True
    ws.add_data_validation(validation)
    validation.add(ws["M3"])

def _remove_owned_sheets(workbook) -> None:
    for sheet_name in (EARNED_VALUE_SHEET, EV_TABLE_SHEET, EV_DATA_SHEET):
        if sheet_name in workbook.sheetnames:
            del workbook[sheet_name]


@dataclass(frozen=True, slots=True)
class _EVDataLayout:
    chart_last_row: int
    status_top: float
    wbs_last_row: int
    wbs_display_rows: int
    negative_last_row: int
    boq_snapshot_last_row: int


def _main_wbs_by_activity(workbook, result: EarnedValueResult) -> dict[str, _MainWBS]:
    """Resolve each EV activity to the authoritative WBS row in ``main``.

    The visible main sheet is the workbook source of truth.  Activity display
    codes such as ``6.1`` are not assumed to be WBS codes; instead we walk the
    main rows, remember the current WBS row, and bind Plan Activity rows by
    their stable Activity ID.  Matching the ID anywhere in the row keeps this
    helper tolerant of presentation-column shifts.
    """
    if "main" not in workbook.sheetnames:
        return {}
    ws = workbook["main"]
    wanted = {activity.activity_id for activity in result.activities if activity.activity_id}
    if not wanted:
        return {}

    current: _MainWBS | None = None
    resolved: dict[str, _MainWBS] = {}
    wbs_order = 0
    for row in ws.iter_rows(values_only=True):
        values = ["" if value is None else str(value).strip() for value in row]
        if not values:
            continue
        row_type = values[0].casefold() if values else ""
        if row_type == "wbs":
            nonempty = [value for value in values[1:] if value]
            if not nonempty:
                current = None
                continue
            wbs_order += 1
            code = nonempty[0]
            name = nonempty[1] if len(nonempty) > 1 else ""
            current = _MainWBS(code=code, name=name, order=wbs_order)
            continue
        if row_type != "activity" or current is None:
            continue
        # Only the Plan row owns the Activity identity; the following Actual row
        # may repeat fields and must not alter the WBS binding.
        upper = {value.upper() for value in values if value}
        if "P" not in upper:
            continue
        for value in values:
            if value in wanted:
                resolved.setdefault(value, current)
                break
    return resolved




@dataclass(frozen=True, slots=True)
class _MainLiveContract:
    header_row: int
    first_period_col: int
    last_period_col: int
    project_plan_row: int | None
    project_actual_row: int | None
    activity_rows: dict[str, tuple[int | None, int | None]]


def _main_live_contract(workbook, result: EarnedValueResult) -> _MainLiveContract | None:
    """Resolve the minimal direct-to-main formula contract used by live EV."""
    if "main" not in workbook.sheetnames:
        return None
    ws = workbook["main"]
    header_row = None
    headers: dict[str, int] = {}
    for row in range(1, min(ws.max_row, 25) + 1):
        row_headers = {
            str(ws.cell(row, col).value or "").strip().casefold(): col
            for col in range(1, ws.max_column + 1)
            if str(ws.cell(row, col).value or "").strip()
        }
        if "row type" in row_headers and "p/a" in row_headers:
            header_row = row
            headers = row_headers
            break
    if header_row is None:
        return None
    period_cols = [
        col
        for col in range(1, ws.max_column + 1)
        if _as_datetime(ws.cell(header_row, col).value) is not None
    ]
    if not period_cols:
        return None
    row_type_col = headers["row type"]
    pa_col = headers["p/a"]
    activity_col = headers.get("activity id")
    wanted = {activity.activity_id for activity in result.activities if activity.activity_id}
    activity_rows: dict[str, list[int | None]] = {key: [None, None] for key in wanted}
    project_plan_row = None
    project_actual_row = None
    project_summary_seen = False
    for row in range(header_row + 1, ws.max_row + 1):
        row_type = str(ws.cell(row, row_type_col).value or "").strip().casefold()
        pa = str(ws.cell(row, pa_col).value or "").strip().upper()
        activity_id = (
            str(ws.cell(row, activity_col).value or "").strip()
            if activity_col is not None
            else ""
        )
        if row_type == "project summary" and pa == "P":
            project_plan_row = row
            project_summary_seen = True
            continue
        if project_summary_seen and project_actual_row is None and pa == "A" and not activity_id:
            project_actual_row = row
        if activity_id in activity_rows:
            if pa == "P":
                activity_rows[activity_id][0] = row
            elif pa == "A":
                activity_rows[activity_id][1] = row
    return _MainLiveContract(
        header_row=header_row,
        first_period_col=min(period_cols),
        last_period_col=max(period_cols),
        project_plan_row=project_plan_row,
        project_actual_row=project_actual_row,
        activity_rows={key: (rows[0], rows[1]) for key, rows in activity_rows.items()},
    )


def _main_progress_formula(contract: _MainLiveContract, source_row: int | None, date_ref: str) -> str:
    if source_row is None:
        return "0"
    first = get_column_letter(contract.first_period_col)
    last = get_column_letter(contract.last_period_col)
    return (
        f'IFERROR(SUMIFS(main!${first}${source_row}:${last}${source_row},'
        f'main!${first}${contract.header_row}:${last}${contract.header_row},'
        f'"<="&{date_ref}),0)'
    )


def _mapping_allocations(workbook) -> tuple[tuple[str, str, float], ...]:
    """Return frozen BOQ→Activity allocated amounts from the embedded mapping."""
    if _MAPPING_SHEET not in workbook.sheetnames:
        return ()
    ws = workbook[_MAPPING_SHEET]
    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        return ()
    headers = {
        str(value or "").strip().casefold(): index
        for index, value in enumerate(header)
        if str(value or "").strip()
    }
    required = ("boq key", "activity id", "allocated amount")
    if any(name not in headers for name in required):
        return ()
    result: list[tuple[str, str, float]] = []
    for values in rows:
        def raw(name: str):
            index = headers[name]
            return values[index] if index < len(values) else None
        boq_key = str(raw("boq key") or "").strip()
        activity_id = str(raw("activity id") or "").strip()
        try:
            allocated = float(raw("allocated amount") or 0.0)
        except (TypeError, ValueError):
            allocated = 0.0
        if boq_key and activity_id:
            result.append((boq_key, activity_id, allocated))
    return tuple(result)


def _activity_performance_at(
    result: EarnedValueResult,
    cutoff: datetime | date | None,
    *,
    wbs_by_activity: dict[str, _MainWBS] | None = None,
) -> tuple[_PerformanceRow, ...]:
    grouped: dict[str, dict[str, float | bool | str | int]] = defaultdict(
        lambda: {
            "bac": 0.0, "pv": 0.0, "ev": 0.0, "pv_has": False, "ev_has": False,
            "name": "", "order": 10**9,
        }
    )
    for activity in result.activities:
        status = _point_on_or_before(activity.points, cutoff, require_ev=True)
        meta = None if wbs_by_activity is None else wbs_by_activity.get(activity.activity_id)
        if meta is None:
            # Backward-compatible fallback for isolated/unit workbooks that do
            # not contain main.  Real rebuilt workbooks use authoritative main.
            label = activity.wbs.strip() or "(Unassigned WBS)"
            name = ""
            order = 10**9
        else:
            label = meta.code
            name = meta.name
            order = meta.order
        bucket = grouped[label]
        bucket["bac"] += float(activity.bac or 0.0)
        if not bucket["name"] and name:
            bucket["name"] = name
        bucket["order"] = min(int(bucket["order"]), order)
        if status is not None and status.planned_value is not None:
            bucket["pv"] += float(status.planned_value)
            bucket["pv_has"] = True
        if status is not None and status.earned_value is not None:
            bucket["ev"] += float(status.earned_value)
            bucket["ev_has"] = True

    rows: list[_PerformanceRow] = []
    for label, bucket in grouped.items():
        pv = float(bucket["pv"]) if bucket["pv_has"] else None
        ev = float(bucket["ev"]) if bucket["ev_has"] else None
        sv = None if pv is None or ev is None else ev - pv
        spi = None if pv in (None, 0.0) or ev is None else ev / pv
        rows.append(
            _PerformanceRow(
                label, float(bucket["bac"]), pv, ev, sv, spi,
                name=str(bucket["name"]), order=int(bucket["order"]),
            )
        )
    rows.sort(key=lambda row: (row.order, row.label.lower()))
    return tuple(rows)


def _ranked_wbs_performance_at(
    result: EarnedValueResult,
    cutoff: datetime | date | None,
    *,
    wbs_by_activity: dict[str, _MainWBS] | None = None,
) -> tuple[_PerformanceRow, ...]:
    """Return every WBS active at cutoff, preserving authoritative main order."""
    return tuple(
        row
        for row in _activity_performance_at(
            result, cutoff, wbs_by_activity=wbs_by_activity
        )
        if row.planned_value is not None and row.planned_value > 0.0
    )


def _negative_boq_variance_at(
    result: EarnedValueResult,
    cutoff: datetime | date | None,
    *,
    activity_ids_by_boq: dict[str, str] | None = None,
) -> tuple[_PerformanceRow, ...]:
    rows: list[_PerformanceRow] = []
    for boq in result.boq_items:
        status = _point_on_or_before(boq.points, cutoff, require_ev=True)
        if status is None or status.schedule_variance is None or status.schedule_variance >= 0.0:
            continue
        rows.append(
            _PerformanceRow(
                boq.description.strip() or boq.stable_id or boq.boq_key,
                float(boq.bac or 0.0),
                status.planned_value,
                status.earned_value,
                status.schedule_variance,
                status.schedule_performance_index,
                activity_id=(activity_ids_by_boq or {}).get(boq.boq_key, ""),
            )
        )
    rows.sort(key=lambda row: (row.schedule_variance if row.schedule_variance is not None else float("inf"), row.label.lower()))
    return tuple(rows)


def _activity_ids_by_boq_from_mapping(workbook) -> dict[str, str]:
    """Return stable Activity IDs linked to each BOQ key from embedded mapping."""
    if _MAPPING_SHEET not in workbook.sheetnames:
        return {}
    ws = workbook[_MAPPING_SHEET]
    rows = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration:
        return {}
    headers = {
        str(value or "").strip().lower(): index
        for index, value in enumerate(header_row)
        if str(value or "").strip()
    }
    if "boq key" not in headers or "activity id" not in headers:
        return {}
    grouped: dict[str, set[str]] = defaultdict(set)
    for values in rows:
        boq_index = headers["boq key"]
        activity_index = headers["activity id"]
        boq_key = str(values[boq_index] if boq_index < len(values) and values[boq_index] is not None else "").strip()
        activity_id = str(values[activity_index] if activity_index < len(values) and values[activity_index] is not None else "").strip()
        if boq_key and activity_id:
            grouped[boq_key].add(activity_id)
    return {key: ", ".join(sorted(values)) for key, values in grouped.items()}


def _selectable_cutoffs(workbook, result: EarnedValueResult) -> tuple[datetime, ...]:
    """Reuse the complete canonical Dashboard_Data monthly cutoff list."""
    dashboard_range = _dashboard_monthly_cutoff_range(workbook, result)
    if dashboard_range is None:
        return _cutoff_options(result)
    sheet_name, last_row = dashboard_range
    ws = workbook[sheet_name]
    return tuple(
        value
        for row in range(2, last_row + 1)
        if (value := _as_datetime(ws.cell(row, 11).value)) is not None
    )


def _boq_metadata_from_mapping(workbook) -> dict[str, tuple[str, str, str, str, str]]:
    """Reuse embedded BOQ provenance for EV Table labels and WBS columns.

    ``BOQ Activity Mapping`` is already the workbook source for BOQ identity and
    hierarchy.  EV-6 reads those labels only; calculation remains owned by the
    existing EarnedValueResult.
    """
    if _MAPPING_SHEET not in workbook.sheetnames:
        return {}
    ws = workbook[_MAPPING_SHEET]
    rows = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration:
        return {}
    headers = {
        str(value or "").strip().lower(): index
        for index, value in enumerate(header_row)
        if str(value or "").strip()
    }
    required = ("boq key", "boq id", "wbs-2", "wbs-3", "wbs-4", "boq description")
    if any(name not in headers for name in required):
        return {}

    def cell(values, name: str) -> str:
        index = headers[name]
        return str(values[index] if index < len(values) and values[index] is not None else "").strip()

    metadata: dict[str, tuple[str, str, str, str, str]] = {}
    for values in rows:
        boq_key = cell(values, "boq key")
        if not boq_key or boq_key in metadata:
            continue
        metadata[boq_key] = (
            cell(values, "boq id"),
            cell(values, "wbs-2"),
            cell(values, "wbs-3"),
            cell(values, "wbs-4"),
            cell(values, "boq description"),
        )
    return metadata


def _render_ev_table(workbook, result: EarnedValueResult, data_layout: _EVDataLayout) -> None:
    """Render the EV-6 BOQ table driven by the dashboard Status Date.

    The visible table contains one row per BOQ item and reads the live EV_Data
    formula layer. Filtering/sorting remains native worksheet behavior.
    """
    ws = workbook.create_sheet(EV_TABLE_SHEET)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A6"
    ws.sheet_properties.tabColor = _NAVY
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    ws.merge_cells("A1:J2")
    ws["A1"] = "EARNED VALUE TABLE"
    ws["A1"].font = Font(name=_FONT, size=20, bold=True, color=_WHITE)
    ws["A1"].fill = _solid(_NAVY)
    ws["A1"].alignment = Alignment(vertical="center")
    for row in ws["A1:J2"]:
        for cell in row:
            cell.fill = _solid(_NAVY)

    ws["A3"] = "Status Date"
    ws["B3"] = f"='{EARNED_VALUE_SHEET}'!$M$3"
    ws["B3"].number_format = "dd-mmm-yyyy"
    ws["D3"] = "BOQ Items"
    ws["E3"] = len(result.boq_items)
    ws["G3"] = "Basis"
    ws["H3"] = "Mapped BOQ"
    for coord in ("A3", "D3", "G3"):
        ws[coord].font = Font(name=_FONT, size=10, bold=True, color=_MUTED)
    for coord in ("B3", "E3", "H3"):
        ws[coord].font = Font(name=_FONT, size=10, bold=True, color=_TEXT)
    for row in ws["A3:J3"]:
        for cell in row:
            cell.fill = _solid(_LIGHT_GRAY)
            cell.border = _thin_border()
    ws["J3"] = "← EV Dashboard"
    ws["J3"].font = Font(name=_FONT, size=9, bold=True, color=_BLUE, underline="single")
    ws["J3"].alignment = Alignment(horizontal="right")
    ws["J3"].hyperlink = f"#'{EARNED_VALUE_SHEET}'!A1"

    header_row = 5
    headers = ("BOQ ID", "WBS-2", "WBS-3", "WBS-4", "BOQ / WORK", "BAC", "PV", "EV", "SV", "SPI")
    for column, header in enumerate(headers, start=1):
        cell = ws.cell(header_row, column, header)
        cell.fill = _solid(_NAVY)
        cell.font = Font(name=_FONT, size=9, bold=True, color=_WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _thin_border()

    metadata = _boq_metadata_from_mapping(workbook)
    rows = []
    for boq in result.boq_items:
        mapped = metadata.get(boq.boq_key, ("", "", "", "", ""))
        boq_id = mapped[0] or boq.stable_id or boq.boq_key
        description = mapped[4] or boq.description or boq_id
        rows.append((mapped[1], mapped[2], mapped[3], description, boq_id, boq))
    rows.sort(key=lambda item: (item[0].lower(), item[1].lower(), item[2].lower(), item[3].lower(), item[4].lower()))

    snapshot_last = data_layout.boq_snapshot_last_row
    for offset, (wbs2, wbs3, wbs4, description, boq_id, boq) in enumerate(rows, start=1):
        row = header_row + offset
        values = (boq_id, wbs2, wbs3, wbs4, description, float(boq.bac or 0.0))
        for column, value in enumerate(values, start=1):
            ws.cell(row, column, value)
        # Hidden stable key keeps the visible table readable while formulas use
        # the same BOQ identity as EV-1 reverse aggregation.
        ws.cell(row, 11, boq.boq_key)
        pv = (
            f'=IFERROR(SUMIFS({EV_DATA_SHEET}!$AG$2:$AG${snapshot_last},'
            f'{EV_DATA_SHEET}!$AF$2:$AF${snapshot_last},$K{row},'
            f'{EV_DATA_SHEET}!$AE$2:$AE${snapshot_last},$B$3),0)'
        )
        ev = (
            f'=IFERROR(SUMIFS({EV_DATA_SHEET}!$AH$2:$AH${snapshot_last},'
            f'{EV_DATA_SHEET}!$AF$2:$AF${snapshot_last},$K{row},'
            f'{EV_DATA_SHEET}!$AE$2:$AE${snapshot_last},$B$3),0)'
        )
        ws.cell(row, 7, pv)
        ws.cell(row, 8, ev)
        ws.cell(row, 9, f'=H{row}-G{row}')
        ws.cell(row, 10, f'=IF(G{row}=0,0,H{row}/G{row})')
        for column in range(1, 11):
            cell = ws.cell(row, column)
            cell.border = _thin_border()
            if offset % 2 == 0:
                cell.fill = _solid(_LIGHT_BLUE)
        for column in (6, 7, 8, 9):
            ws.cell(row, column).number_format = "#,##0.00"
        ws.cell(row, 10).number_format = "0.00"

    last_row = max(header_row + 1, header_row + len(rows))
    ws.auto_filter.ref = f"A{header_row}:J{last_row}"
    ws.column_dimensions["K"].hidden = True
    widths = {
        "A": 18, "B": 16, "C": 16, "D": 16, "E": 46,
        "F": 16, "G": 16, "H": 16, "I": 16, "J": 12,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 25
    ws.row_dimensions[5].height = 22

    if rows:
        first = header_row + 1
        last = header_row + len(rows)
        ws.conditional_formatting.add(
            f"I{first}:I{last}",
            FormulaRule(formula=[f'$I{first}<0'], fill=_solid(_LIGHT_RED), font=Font(name=_FONT, color=_RED)),
        )
        ws.conditional_formatting.add(
            f"I{first}:I{last}",
            FormulaRule(formula=[f'$I{first}>0'], fill=_solid(_LIGHT_GREEN), font=Font(name=_FONT, color=_GREEN)),
        )
        ws.conditional_formatting.add(
            f"J{first}:J{last}",
            FormulaRule(formula=[f'AND($G{first}>0,$J{first}<1)'], fill=_solid(_LIGHT_RED), font=Font(name=_FONT, color=_RED)),
        )


def _write_ev_data(workbook, result: EarnedValueResult) -> _EVDataLayout:
    """Build the EV-owned live formula layer over the editable ``main`` sheet.

    BAC and BOQ allocation topology remain frozen from the successful EV rebuild.
    Plan/Actual progress is read directly from ``main`` with Excel formulas so a
    user edit followed by F9/Save updates PV/EV/SV/SPI without another rebuild.
    """
    ws = workbook.create_sheet(EV_DATA_SHEET)
    ws.sheet_state = "hidden"
    ws.sheet_view.showGridLines = False

    live = _main_live_contract(workbook, result)
    allocations = _mapping_allocations(workbook)

    # Reporting context is independent from EV!M3. Keep it live when the normal
    # Progress Dashboard exists; standalone workbooks retain the rebuild cutoff.
    ws["G1"] = "Status Top"
    ws["G3"] = "Reporting Cutoff"
    if "Dashboard" in workbook.sheetnames:
        ws["G4"] = "=Dashboard!$K$5"
    else:
        ws["G4"] = result.cutoff_date
    ws["G4"].number_format = "dd-mmm-yyyy"

    headers = ("Reporting Date", "PV", "EV Source", "EV", "Status Date", "Point Type")
    for column, header in enumerate(headers, start=1):
        ws.cell(1, column, header)

    chart_points = _monthly_chart_points(result)
    status_top = max(float(result.project_bac or 0.0) * 1.05, 1.0)
    ws["G2"] = status_top
    ws["G2"].number_format = "#,##0.00"
    ws.cell(1, 29, "PV @ Status Date")
    ws.cell(1, 30, "EV @ Status Date")

    for row_index, point in enumerate(chart_points, start=2):
        ws.cell(row_index, 1, point.reporting_date)
        if live is not None and live.project_plan_row is not None:
            plan_progress = _main_progress_formula(live, live.project_plan_row, f"A{row_index}")
            ws.cell(row_index, 2, f"={float(result.project_bac or 0.0)}*({plan_progress})")
        else:
            ws.cell(row_index, 2, point.planned_value)
        if live is not None and live.project_actual_row is not None:
            actual_progress = _main_progress_formula(
                live,
                live.project_actual_row,
                f"MIN(A{row_index},$G$4)",
            )
            ws.cell(row_index, 3, f"={float(result.project_bac or 0.0)}*({actual_progress})")
        else:
            ws.cell(row_index, 3, point.earned_value)
        # Render EV through the selected Status Date. When M3 is later than the
        # reporting cutoff, column C is already frozen at the latest permitted
        # Actual date, so the plotted line carries forward flat through M3.
        ws.cell(
            row_index,
            4,
            f'=IF(A{row_index}="",NA(),IF(A{row_index}>'
            f"'{EARNED_VALUE_SHEET}'!$M$3,NA(),C{row_index}))",
        )
        next_row = row_index + 1
        status_ref = f"'{EARNED_VALUE_SHEET}'!$M$3"
        ws.cell(
            row_index,
            5,
            f'=IF(A{row_index}="",NA(),IF(AND(A{row_index}<={status_ref},'
            f'OR(A{next_row}="",A{next_row}>{status_ref})),$G$2,NA()))',
        )
        ws.cell(row_index, 6, point.point_type)
        ws.cell(row_index, 29, f'=IF(ISNUMBER(E{row_index}),B{row_index},NA())')
        ws.cell(row_index, 30, f'=IF(ISNUMBER(E{row_index}),D{row_index},NA())')
        ws.cell(row_index, 1).number_format = "dd-mmm-yyyy"
        for column in (2, 3, 4, 5, 29, 30):
            ws.cell(row_index, column).number_format = "#,##0.00"
    chart_last_row = max(2, len(chart_points) + 1)

    options = _selectable_cutoffs(workbook, result)
    if _dashboard_monthly_cutoff_range(workbook, result) is None:
        ws["H1"] = "Cutoff Options"
        for row, value in enumerate(options, start=2):
            ws.cell(row, 8, value)
            ws.cell(row, 8).number_format = "dd-mmm-yyyy"

    # One live Activity row per Activity. This deliberately mirrors the direct-
    # to-main Activity Progress pattern: no Activity×time Python snapshot cache.
    activity_start_col = 53  # BA
    activity_headers = (
        "Status Date", "Activity ID", "WBS", "WBS Name", "BAC",
        "Plan Progress", "Actual Progress", "PV", "EV",
    )
    for col, header in enumerate(activity_headers, start=activity_start_col):
        ws.cell(1, col, header)
    wbs_by_activity = _main_wbs_by_activity(workbook, result)
    activity_first_row = 2
    for offset, activity in enumerate(result.activities):
        row = activity_first_row + offset
        meta = wbs_by_activity.get(activity.activity_id)
        wbs_code = meta.code if meta is not None else (activity.wbs.strip() or "(Unassigned WBS)")
        wbs_name = meta.name if meta is not None else ""
        ws.cell(row, activity_start_col, f"='{EARNED_VALUE_SHEET}'!$M$3")
        ws.cell(row, activity_start_col + 1, activity.activity_id)
        ws.cell(row, activity_start_col + 2, wbs_code)
        ws.cell(row, activity_start_col + 3, wbs_name)
        ws.cell(row, activity_start_col + 4, float(activity.bac or 0.0))
        source_rows = None if live is None else live.activity_rows.get(activity.activity_id)
        if source_rows is not None:
            plan_row, actual_row = source_rows
            plan_formula = _main_progress_formula(live, plan_row, f"${get_column_letter(activity_start_col)}{row}")
            actual_formula = _main_progress_formula(
                live,
                actual_row,
                f"MIN(${get_column_letter(activity_start_col)}{row},$G$4)",
            )
            ws.cell(row, activity_start_col + 5, f"={plan_formula}")
            ws.cell(row, activity_start_col + 6, f"={actual_formula}")
            ws.cell(row, activity_start_col + 7, f"={get_column_letter(activity_start_col + 4)}{row}*{get_column_letter(activity_start_col + 5)}{row}")
            ws.cell(row, activity_start_col + 8, f"={get_column_letter(activity_start_col + 4)}{row}*{get_column_letter(activity_start_col + 6)}{row}")
        else:
            status = _point_on_or_before(activity.points, result.cutoff_date, require_ev=True)
            ws.cell(row, activity_start_col + 5, 0 if activity.bac in (None, 0) or status is None or status.planned_value is None else status.planned_value / activity.bac)
            ws.cell(row, activity_start_col + 6, 0 if activity.bac in (None, 0) or status is None or status.earned_value is None else status.earned_value / activity.bac)
            ws.cell(row, activity_start_col + 7, None if status is None else status.planned_value)
            ws.cell(row, activity_start_col + 8, None if status is None else status.earned_value)
        ws.cell(row, activity_start_col).number_format = "dd-mmm-yyyy"
        for col in range(activity_start_col + 4, activity_start_col + 9):
            ws.cell(row, col).number_format = "#,##0.00"
    activity_last_row = max(2, activity_first_row + len(result.activities) - 1)

    # WBS interface J:S is kept stable for the existing dashboard, but its rows
    # are now live formulas over the one-row-per-Activity helper above.
    for col, header in enumerate(
        ("Lookup Key", "Snapshot Date", "Rank", "WBS", "WBS Name", "BAC", "PV", "EV", "SV", "SPI"),
        start=10,
    ):
        ws.cell(1, col, header)
    unique_wbs: dict[str, _MainWBS] = {}
    for activity in result.activities:
        meta = wbs_by_activity.get(activity.activity_id)
        if meta is None:
            code = activity.wbs.strip() or "(Unassigned WBS)"
            unique_wbs.setdefault(code, _MainWBS(code, "", 10**9))
        else:
            unique_wbs.setdefault(meta.code, meta)
    wbs_items = sorted(unique_wbs.values(), key=lambda item: (item.order, item.code.casefold()))
    act_wbs_col = get_column_letter(activity_start_col + 2)
    act_bac_col = get_column_letter(activity_start_col + 4)
    act_pv_col = get_column_letter(activity_start_col + 7)
    act_ev_col = get_column_letter(activity_start_col + 8)
    for offset, meta in enumerate(wbs_items, start=0):
        row = 2 + offset
        ws.cell(row, 11, f"='{EARNED_VALUE_SHEET}'!$M$3")
        ws.cell(row, 13, meta.code)
        ws.cell(row, 14, meta.name)
        ws.cell(row, 15, f'=SUMIFS(${act_bac_col}$2:${act_bac_col}${activity_last_row},${act_wbs_col}$2:${act_wbs_col}${activity_last_row},M{row})')
        ws.cell(row, 16, f'=SUMIFS(${act_pv_col}$2:${act_pv_col}${activity_last_row},${act_wbs_col}$2:${act_wbs_col}${activity_last_row},M{row})')
        ws.cell(row, 17, f'=SUMIFS(${act_ev_col}$2:${act_ev_col}${activity_last_row},${act_wbs_col}$2:${act_wbs_col}${activity_last_row},M{row})')
        ws.cell(row, 18, f'=Q{row}-P{row}')
        ws.cell(row, 19, f'=IF(P{row}=0,0,Q{row}/P{row})')
        ws.cell(row, 12, f'=IF(P{row}<=0,"",COUNTIF($P$2:P{row},">0"))')
        ws.cell(row, 10, f'=IF(L{row}="","",TEXT(K{row},"yyyymmdd")&"|"&L{row})')
        ws.cell(row, 11).number_format = "dd-mmm-yyyy"
        for col in (15, 16, 17, 18):
            ws.cell(row, col).number_format = "#,##0.00"
        ws.cell(row, 19).number_format = "0.00"
    wbs_last_row = max(2, len(wbs_items) + 1)
    wbs_display_rows = max(1, len(wbs_items))

    # Mapping helper: allocation amounts remain structural/BAC authority; only
    # the linked Activity Plan/Actual progress is live.
    mapping_start_col = 62  # BJ
    mapping_headers = ("BOQ Key", "Activity ID", "Allocated BAC", "PV", "EV")
    for col, header in enumerate(mapping_headers, start=mapping_start_col):
        ws.cell(1, col, header)
    activity_id_col = get_column_letter(activity_start_col + 1)
    activity_plan_col = get_column_letter(activity_start_col + 5)
    activity_actual_col = get_column_letter(activity_start_col + 6)
    mapping_first_row = 2
    for offset, (boq_key, activity_id, allocated) in enumerate(allocations):
        row = mapping_first_row + offset
        ws.cell(row, mapping_start_col, boq_key)
        ws.cell(row, mapping_start_col + 1, activity_id)
        ws.cell(row, mapping_start_col + 2, allocated)
        lookup_range = f'${activity_id_col}$2:${get_column_letter(activity_start_col + 8)}${activity_last_row}'
        ws.cell(row, mapping_start_col + 3, f'=IFERROR({get_column_letter(mapping_start_col + 2)}{row}*VLOOKUP({get_column_letter(mapping_start_col + 1)}{row},{lookup_range},5,FALSE),0)')
        ws.cell(row, mapping_start_col + 4, f'=IFERROR({get_column_letter(mapping_start_col + 2)}{row}*VLOOKUP({get_column_letter(mapping_start_col + 1)}{row},{lookup_range},6,FALSE),0)')
        for col in range(mapping_start_col + 2, mapping_start_col + 5):
            ws.cell(row, col).number_format = "#,##0.00"
    mapping_last_row = max(2, mapping_first_row + len(allocations) - 1)

    # BOQ interface AE:AK: one row per BOQ, selected Status Date only. This is
    # live and O(BOQ), not BOQ×time. Existing EV Table SUMIFS remain compatible.
    boq_headers = ("BOQ Snapshot Date", "BOQ Key", "BOQ PV", "BOQ EV", "BOQ SV", "BOQ SPI", "Negative Rank")
    for col, header in enumerate(boq_headers, start=31):
        ws.cell(1, col, header)
    mapping_boq_col = get_column_letter(mapping_start_col)
    mapping_pv_col = get_column_letter(mapping_start_col + 3)
    mapping_ev_col = get_column_letter(mapping_start_col + 4)
    boq_first_row = 2
    for offset, boq in enumerate(result.boq_items):
        row = boq_first_row + offset
        ws.cell(row, 31, f"='{EARNED_VALUE_SHEET}'!$M$3")
        ws.cell(row, 32, boq.boq_key)
        if allocations:
            ws.cell(row, 33, f'=SUMIFS(${mapping_pv_col}$2:${mapping_pv_col}${mapping_last_row},${mapping_boq_col}$2:${mapping_boq_col}${mapping_last_row},AF{row})')
            ws.cell(row, 34, f'=SUMIFS(${mapping_ev_col}$2:${mapping_ev_col}${mapping_last_row},${mapping_boq_col}$2:${mapping_boq_col}${mapping_last_row},AF{row})')
        else:
            status = _point_on_or_before(boq.points, result.cutoff_date, require_ev=True)
            ws.cell(row, 33, None if status is None else status.planned_value)
            ws.cell(row, 34, None if status is None else status.earned_value)
        ws.cell(row, 35, f'=AH{row}-AG{row}')
        ws.cell(row, 36, f'=IF(AG{row}=0,0,AH{row}/AG{row})')
        # Unique ascending negative rank; row-order breaks exact SV ties.
        ws.cell(row, 37, f'=IF(AI{row}>=0,"",COUNTIF($AI$2:$AI${max(2, len(result.boq_items)+1)},"<"&AI{row})+COUNTIF($AI$2:AI{row},AI{row}))')
        ws.cell(row, 31).number_format = "dd-mmm-yyyy"
        for col in (33, 34, 35):
            ws.cell(row, col).number_format = "#,##0.00"
        ws.cell(row, 36).number_format = "0.00"
    boq_snapshot_last_row = max(2, len(result.boq_items) + 1)

    # Top-10 interface T:Z stays stable, but now looks up the live BOQ ranking.
    activity_ids_by_boq = _activity_ids_by_boq_from_mapping(workbook)
    metadata = _boq_metadata_from_mapping(workbook)
    for col, header in enumerate(
        ("Lookup Key", "Snapshot Date", "Rank", "Activity ID", "BOQ / Work", "SV", "SPI"),
        start=20,
    ):
        ws.cell(1, col, header)
    boq_rank_range = f'$AK$2:$AK${boq_snapshot_last_row}'
    for rank in range(1, 11):
        row = rank + 1
        ws.cell(row, 21, f"='{EARNED_VALUE_SHEET}'!$M$3")
        ws.cell(row, 22, rank)
        ws.cell(row, 20, f'=TEXT(U{row},"yyyymmdd")&"|"&V{row}')
        # Store static fallback labels beside the live lookup through INDEX/MATCH.
        ws.cell(row, 23, f'=IFERROR(INDEX($BO$2:$BO${boq_snapshot_last_row},MATCH(V{row},{boq_rank_range},0)),"")')
        ws.cell(row, 24, f'=IFERROR(INDEX($BP$2:$BP${boq_snapshot_last_row},MATCH(V{row},{boq_rank_range},0)),"")')
        ws.cell(row, 25, f'=IFERROR(INDEX($AI$2:$AI${boq_snapshot_last_row},MATCH(V{row},{boq_rank_range},0)),"")')
        ws.cell(row, 26, f'=IFERROR(INDEX($AJ$2:$AJ${boq_snapshot_last_row},MATCH(V{row},{boq_rank_range},0)),"")')
        ws.cell(row, 21).number_format = "dd-mmm-yyyy"
        ws.cell(row, 25).number_format = "#,##0.00"
        ws.cell(row, 26).number_format = "0.00"

    # BO:BP provide BOQ Activity ID / display label aligned with AE:AK rows.
    ws["BO1"] = "Activity ID"
    ws["BP1"] = "BOQ / Work"
    for offset, boq in enumerate(result.boq_items):
        row = boq_first_row + offset
        mapped = metadata.get(boq.boq_key, ("", "", "", "", ""))
        ws.cell(row, 67, activity_ids_by_boq.get(boq.boq_key, ""))
        ws.cell(row, 68, mapped[4] or boq.description or boq.stable_id or boq.boq_key)

    negative_last_row = 11
    return _EVDataLayout(
        chart_last_row,
        status_top,
        wbs_last_row,
        wbs_display_rows,
        negative_last_row,
        boq_snapshot_last_row,
    )

def render_earned_value_sheet(workbook, result: EarnedValueResult, *, include_chart: bool = True) -> None:
    """Render the EV-5 management dashboard from an EV-1 result.

    Presentation contract:
    - five project KPIs at Status Date,
    - monthly PV/EV management curve plus exact Status Date marker,
    - WBS performance summary,
    - top negative BOQ variance,
    - hidden ``EV_Data`` live formula source,
    - EV-6 BOQ table with native Excel filter/sort.
    """
    _remove_owned_sheets(workbook)
    data_layout = _write_ev_data(workbook, result)

    ws = workbook.create_sheet(EARNED_VALUE_SHEET)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A6"
    ws.sheet_properties.tabColor = _NAVY
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    # Overall dashboard canvas.
    for column in range(1, 16):
        ws.column_dimensions[chr(64 + column)].width = 12
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 18
    # EV dashboard uses A:O as a 15-column layout grid.  The five KPI cards
    # continue to consume three columns each (A:C, D:F, G:I, J:L, M:O), while
    # the management tables merge cells across the same grid for long labels.
    # This keeps the KPI geometry balanced without resorting to drawing-layer
    # shapes and gives BOQ/WBS descriptions substantially more room.
    grid_widths = {
        "A": 10, "B": 15, "C": 12,
        "D": 12, "E": 12, "F": 13,
        "G": 12, "H": 12, "I": 13,
        "J": 11, "K": 13, "L": 13,
        "M": 12, "N": 12, "O": 13,
    }
    for column, width in grid_widths.items():
        ws.column_dimensions[column].width = width

    # Header.
    ws.merge_cells("A1:O2")
    title = ws["A1"]
    title.value = "EARNED VALUE"
    title.font = Font(name=_FONT, size=20, bold=True, color=_WHITE)
    title.fill = _solid(_NAVY)
    title.alignment = Alignment(vertical="center")
    for row in ws["A1:O2"]:
        for cell in row:
            cell.fill = _solid(_NAVY)

    ws["A3"] = "PROJECT PERFORMANCE"
    ws["A3"].font = Font(name=_FONT, size=11, bold=True, color=_NAVY)
    _add_cutoff_dropdown(workbook, ws, workbook[EV_DATA_SHEET], result)

    chart_last_row = data_layout.chart_last_row
    # KPI values use the same selected Status Date as the chart.  This mirrors
    # Dashboard's formula-driven cutoff contract and prevents a changed dropdown
    # from leaving stale Python snapshot values on screen.
    pv_formula = f'=IFERROR(SUMIFS({EV_DATA_SHEET}!$B$2:$B${chart_last_row},{EV_DATA_SHEET}!$A$2:$A${chart_last_row},$M$3),0)'
    ev_formula = f'=IFERROR(SUMIFS({EV_DATA_SHEET}!$D$2:$D${chart_last_row},{EV_DATA_SHEET}!$A$2:$A${chart_last_row},$M$3),0)'
    kpis = (
        ("BAC", result.project_bac, "#,##0.00", "Budget"),
        ("PV", pv_formula, "#,##0.00", "Planned"),
        ("EV", ev_formula, "#,##0.00", "Earned"),
        ("SV", "=G6-D6", "#,##0.00", '=IF(ROUND(J6,2)=0,"On Plan",IF(J6<0,"Behind","Ahead"))'),
        ("SPI", '=IF(D6=0,0,G6/D6)', "0.00", '=IF(D6=0,"N/A",IF(ROUND(M6,2)=1,"On Plan",IF(M6<1,"Behind","Ahead")))'),
    )
    starts = (1, 4, 7, 10, 13)
    for (label, value, number_format, note), col in zip(kpis, starts):
        end_col = col + 2
        ws.merge_cells(start_row=5, start_column=col, end_row=5, end_column=end_col)
        ws.merge_cells(start_row=6, start_column=col, end_row=7, end_column=end_col)
        ws.merge_cells(start_row=8, start_column=col, end_row=8, end_column=end_col)
        label_cell = ws.cell(5, col)
        value_cell = ws.cell(6, col)
        note_cell = ws.cell(8, col)
        label_cell.value = label
        value_cell.value = value
        note_cell.value = note
        card_style = {
            "BAC": (_LIGHT_GRAY, _TEXT),
            "PV": (_LIGHT_BLUE, _BLUE),
            "EV": (_LIGHT_GREEN, _GREEN),
            "SV": (_LIGHT_AMBER, _AMBER),
            "SPI": (_LIGHT_AMBER, _AMBER),
        }
        fill_color, accent_color = card_style[label]
        label_cell.font = Font(name=_FONT, size=10, bold=True, color=accent_color)
        value_cell.font = Font(name=_FONT, size=17, bold=True, color=accent_color)
        note_cell.font = Font(name=_FONT, size=9, bold=label in {"SV", "SPI"}, color=_MUTED)
        label_cell.alignment = Alignment(horizontal="center", vertical="center")
        value_cell.alignment = Alignment(horizontal="center", vertical="center")
        note_cell.alignment = Alignment(horizontal="center", vertical="center")
        value_cell.number_format = number_format
        for row in range(5, 9):
            for c in range(col, end_col + 1):
                cell = ws.cell(row, c)
                cell.border = _thin_border()
                cell.fill = _solid(fill_color)

    # Reuse Dashboard status-card language: red = behind, green = ahead, blue = on plan.
    ws.conditional_formatting.add(
        "J5:L8",
        FormulaRule(formula=['$J$6<0'], fill=_solid(_LIGHT_RED), font=Font(name=_FONT, bold=True, color=_RED)),
    )
    ws.conditional_formatting.add(
        "J5:L8",
        FormulaRule(formula=['$J$6>0'], fill=_solid(_LIGHT_GREEN), font=Font(name=_FONT, bold=True, color=_GREEN)),
    )
    ws.conditional_formatting.add(
        "J5:L8",
        FormulaRule(formula=['ROUND($J$6,2)=0'], fill=_solid(_LIGHT_BLUE), font=Font(name=_FONT, bold=True, color=_BLUE)),
    )
    ws.conditional_formatting.add(
        "M5:O8",
        FormulaRule(formula=['AND($D$6>0,$M$6<1)'], fill=_solid(_LIGHT_RED), font=Font(name=_FONT, bold=True, color=_RED)),
    )
    ws.conditional_formatting.add(
        "M5:O8",
        FormulaRule(formula=['AND($D$6>0,$M$6>1)'], fill=_solid(_LIGHT_GREEN), font=Font(name=_FONT, bold=True, color=_GREEN)),
    )
    ws.conditional_formatting.add(
        "M5:O8",
        FormulaRule(formula=['AND($D$6>0,ROUND($M$6,2)=1)'], fill=_solid(_LIGHT_BLUE), font=Font(name=_FONT, bold=True, color=_BLUE)),
    )

    # Project performance chart.
    ws["A10"] = "PROJECT PERFORMANCE"
    ws["A10"].font = Font(name=_FONT, size=11, bold=True, color=_NAVY)
    # The chart already owns the PV / EV / Status Date legend.  Do not repeat a
    # worksheet-level legend on the right; it creates dead space and visually
    # disconnects the labels from the chart.

    if data_layout.chart_last_row >= 2:
        data_ws = workbook[EV_DATA_SHEET]
        chart = LineChart()
        chart.title = None
        chart.style = int(DASHBOARD_LAYOUT.get("chart_style", 2))
        chart.height = float(DASHBOARD_LAYOUT.get("chart_height", 9.2))
        # Use the full A:O dashboard canvas.  S-curves benefit from horizontal
        # space, and the internal legend now sits inside the chart itself.
        chart.width = max(float(DASHBOARD_LAYOUT.get("chart_width", 25.0)), 38.0)
        chart.y_axis.title = "Value (Million)"
        chart.y_axis.numFmt = '#,##0,,"M"'
        chart.display_blanks = "gap"
        # Reuse Progress Studio's established Excel-safe LineChart DateAxis
        # contract. It keeps reciprocal axId/crossAx references so Excel does
        # not repair/remove the chart when the workbook is opened.  The date
        # ticks already explain the horizontal dimension, so no generic
        # ``Period`` title is needed.
        _date_axis_for_line_chart(chart, title=None)
        chart.legend.position = "t"
        # Match the existing Dashboard chart treatment instead of inventing a
        # second visual language for the same workbook.
        grid = GraphicalProperties()
        grid.line.solidFill = DASHBOARD_COLORS.get("chart_grid", "E5E7EB")
        grid.line.width = 9000
        chart.y_axis.majorGridlines.graphicalProperties = grid
        axis_line = GraphicalProperties()
        axis_line.line.solidFill = DASHBOARD_COLORS.get("chart_axis", "B8C2CC")
        axis_line.line.width = 9000
        chart.y_axis.spPr = axis_line
        chart.x_axis.spPr = axis_line

        data = Reference(
            data_ws,
            min_col=2,
            max_col=2,
            min_row=1,
            max_row=data_layout.chart_last_row,
        )
        dates = Reference(
            data_ws,
            min_col=1,
            min_row=2,
            max_row=data_layout.chart_last_row,
        )
        chart.add_data(data, titles_from_data=True)
        chart.add_data(Reference(data_ws, min_col=4, max_col=4, min_row=1, max_row=data_layout.chart_last_row), titles_from_data=True)
        chart.add_data(Reference(data_ws, min_col=5, max_col=5, min_row=1, max_row=data_layout.chart_last_row), titles_from_data=True)
        chart.add_data(Reference(data_ws, min_col=29, max_col=29, min_row=1, max_row=data_layout.chart_last_row), titles_from_data=True)
        chart.add_data(Reference(data_ws, min_col=30, max_col=30, min_row=1, max_row=data_layout.chart_last_row), titles_from_data=True)
        chart.set_categories(dates)

        if len(chart.series) >= 1:
            chart.series[0].graphicalProperties.line.solidFill = _BLUE
            chart.series[0].graphicalProperties.line.width = int(DASHBOARD_LAYOUT.get("plan_line_width", 26000))
        if len(chart.series) >= 2:
            chart.series[1].graphicalProperties.line.solidFill = _GREEN
            chart.series[1].graphicalProperties.line.width = int(DASHBOARD_LAYOUT.get("actual_line_width", 26000))
        if len(chart.series) >= 3:
            status_series = chart.series[2]
            status_series.graphicalProperties.line.noFill = True
            status_series.marker.symbol = "none"
            status_series.errBars = ErrorBars(
                errDir="y",
                errBarType="minus",
                errValType="fixedVal",
                noEndCap=True,
                val=data_layout.status_top,
                spPr=GraphicalProperties(solidFill=_RED),
            )
            # Show the selected date directly above the redline.  Category-name
            # labels keep the text live with the dropdown and avoid duplicating
            # a second floating worksheet label.
            status_series.dLbls = DataLabelList(
                showCatName=True,
                showVal=False,
                showSerName=False,
                dLblPos="t",
            )
            chart.y_axis.scaling.min = 0
            chart.y_axis.scaling.max = data_layout.status_top

        # Cutoff markers are marker-only series: PV and EV remain clean S-curves
        # with no per-period markers, while the selected comparison point is
        # visually explicit.
        if len(chart.series) >= 4:
            pv_marker = chart.series[3]
            pv_marker.graphicalProperties.line.noFill = True
            pv_marker.marker.symbol = "circle"
            pv_marker.marker.size = 7
            pv_marker.marker.graphicalProperties.solidFill = _BLUE
            pv_marker.marker.graphicalProperties.line.solidFill = _BLUE
        if len(chart.series) >= 5:
            ev_marker = chart.series[4]
            ev_marker.graphicalProperties.line.noFill = True
            ev_marker.marker.symbol = "circle"
            ev_marker.marker.size = 7
            ev_marker.marker.graphicalProperties.solidFill = _GREEN
            ev_marker.marker.graphicalProperties.line.solidFill = _GREEN

        # Keep the management legend compact: PV, EV and Status Date only.
        # Marker-only helper series are deliberately hidden from the legend.
        if chart.legend is not None:
            chart.legend.legendEntry = [
                LegendEntry(idx=3, delete=True),
                LegendEntry(idx=4, delete=True),
            ]

        if include_chart:
            ws.add_chart(chart, "A11")

    ws.merge_cells("A29:O29")
    ws["A29"] = "PV curve = full baseline    •    EV curve = selected Status Date"
    ws["A29"].font = Font(name=_FONT, size=9, italic=True, color=_MUTED)
    ws["A29"].alignment = Alignment(horizontal="left", vertical="center")

    # Management tables.  Both tables share the A:O layout grid; long text
    # fields use merged cells instead of changing column widths that would
    # distort the KPI cards above.
    table_row = 31
    ws["A30"] = "ACTIVE WBS PERFORMANCE"
    ws["A30"].font = Font(name=_FONT, size=11, bold=True, color=_NAVY)
    wbs_headers = (
        (1, 1, "WBS"),
        (2, 3, "WBS NAME"),
        (4, 4, "BAC"),
        (5, 5, "PV"),
        (6, 6, "EV"),
        (7, 7, "SV"),
        (8, 8, "SPI"),
    )
    for start_col, end_col, header in wbs_headers:
        if end_col > start_col:
            ws.merge_cells(start_row=table_row, start_column=start_col, end_row=table_row, end_column=end_col)
        cell = ws.cell(table_row, start_col, header)
        cell.fill = _solid(_NAVY)
        cell.font = Font(name=_FONT, size=9, bold=True, color=_WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        for column in range(start_col, end_col + 1):
            ws.cell(table_row, column).fill = _solid(_NAVY)
            ws.cell(table_row, column).border = _thin_border()

    wbs_lookup_range = f'{EV_DATA_SHEET}!$J$2:$S${data_layout.wbs_last_row}'
    wbs_rows = max(1, data_layout.wbs_display_rows)
    for offset in range(1, wbs_rows + 1):
        excel_row = table_row + offset
        key = f'TEXT($M$3,"yyyymmdd")&"|{offset}"'
        # WBS code, WBS name, BAC, PV, EV, SV, SPI.  WBS name spans B:C.
        lookup_targets = (
            (1, 1, 4),
            (2, 3, 5),
            (4, 4, 6),
            (5, 5, 7),
            (6, 6, 8),
            (7, 7, 9),
            (8, 8, 10),
        )
        for start_col, end_col, lookup_index in lookup_targets:
            if end_col > start_col:
                ws.merge_cells(start_row=excel_row, start_column=start_col, end_row=excel_row, end_column=end_col)
            ws.cell(
                excel_row, start_col,
                f'=IFERROR(VLOOKUP({key},{wbs_lookup_range},{lookup_index},FALSE),"")',
            )
            for column in range(start_col, end_col + 1):
                cell = ws.cell(excel_row, column)
                cell.border = _thin_border()
                if offset % 2 == 0:
                    cell.fill = _solid(_LIGHT_BLUE)
        for column in (4, 5, 6, 7):
            ws.cell(excel_row, column).number_format = "#,##0.00"
        ws.cell(excel_row, 8).number_format = "0.00"

    last_wbs_row = table_row + wbs_rows
    ws.conditional_formatting.add(
        f"G{table_row + 1}:H{last_wbs_row}",
        FormulaRule(formula=[f'$G{table_row + 1}<0'], fill=_solid(_LIGHT_RED), font=Font(name=_FONT, color=_RED)),
    )
    ws.conditional_formatting.add(
        f"G{table_row + 1}:H{last_wbs_row}",
        FormulaRule(formula=[f'$G{table_row + 1}>0'], fill=_solid(_LIGHT_GREEN), font=Font(name=_FONT, color=_GREEN)),
    )

    ws["I30"] = "TOP 10 NEGATIVE VARIANCE"
    ws["I30"].font = Font(name=_FONT, size=11, bold=True, color=_NAVY)
    variance_headers = (
        (9, 9, "ACTIVITY ID"),
        (10, 12, "BOQ / WORK"),
        (13, 14, "SV"),
        (15, 15, "SPI"),
    )
    for start_col, end_col, header in variance_headers:
        if end_col > start_col:
            ws.merge_cells(start_row=table_row, start_column=start_col, end_row=table_row, end_column=end_col)
        cell = ws.cell(table_row, start_col, header)
        cell.fill = _solid(_NAVY)
        cell.font = Font(name=_FONT, size=9, bold=True, color=_WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        for column in range(start_col, end_col + 1):
            ws.cell(table_row, column).fill = _solid(_NAVY)
            ws.cell(table_row, column).border = _thin_border()

    lookup_range = f'{EV_DATA_SHEET}!$T$2:$Z${data_layout.negative_last_row}'
    for offset in range(1, 11):
        excel_row = table_row + offset
        key = f'TEXT($M$3,"yyyymmdd")&"|{offset}"'
        # Activity ID | BOQ/Work (J:L) | SV (M:N) | SPI (O)
        lookup_targets = (
            (9, 9, 4),
            (10, 12, 5),
            (13, 14, 6),
            (15, 15, 7),
        )
        for start_col, end_col, lookup_index in lookup_targets:
            if end_col > start_col:
                ws.merge_cells(start_row=excel_row, start_column=start_col, end_row=excel_row, end_column=end_col)
            ws.cell(
                excel_row, start_col,
                f'=IFERROR(VLOOKUP({key},{lookup_range},{lookup_index},FALSE),"")',
            )
            for column in range(start_col, end_col + 1):
                cell = ws.cell(excel_row, column)
                cell.border = _thin_border()
                if offset % 2 == 0:
                    cell.fill = _solid(_LIGHT_RED)
        ws.cell(excel_row, 13).number_format = "#,##0.00"
        ws.cell(excel_row, 15).number_format = "0.00"

    detail_row = max(last_wbs_row, table_row + 10) + 2
    ws.cell(detail_row, 1, "Detailed BOQ analysis → EV Table")
    ws.cell(detail_row, 1).font = Font(name=_FONT, size=9, italic=True, color=_BLUE, underline="single")
    ws.cell(detail_row, 1).hyperlink = f"#'{EV_TABLE_SHEET}'!A1"

    _render_ev_table(workbook, result, data_layout)

    final_dashboard_row = max(last_wbs_row, table_row + 10) + 3
    for row in range(1, final_dashboard_row + 1):
        ws.row_dimensions[row].height = 20
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 25
    ws.row_dimensions[3].height = 24
    ws.row_dimensions[5].height = 22
    ws.row_dimensions[6].height = 30
    ws.row_dimensions[7].height = 30
    ws.row_dimensions[8].height = 22
    ws.row_dimensions[10].height = 24
    ws.row_dimensions[30].height = 24
