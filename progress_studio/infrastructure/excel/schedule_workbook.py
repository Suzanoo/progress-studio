from __future__ import annotations

import argparse
import copy
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError as exc:
    raise SystemExit(
        "openpyxl was not found.\nInstall it with: pip install openpyxl"
    ) from exc


DEFAULT_SHEET = "main"
DATE_FORMAT = "dd/mm/yyyy"
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
VERY_LIGHT_GRAY_FILL = PatternFill("solid", fgColor="F2F2F2")


def normalize_header(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip().lower()


def resolve_input_files(input_path: Path) -> list[Path]:
    if not input_path.exists():
        raise FileNotFoundError(f"Input not found: {input_path}")

    if input_path.is_file():
        if input_path.suffix.lower() != ".xlsx":
            raise ValueError(f"Input file must be .xlsx: {input_path}")
        return [input_path]

    if input_path.is_dir():
        files = sorted(
            [
                path
                for path in input_path.iterdir()
                if path.is_file()
                and path.suffix.lower() == ".xlsx"
                and not path.name.startswith("~$")
            ],
            key=lambda path: path.name.lower(),
        )
        if not files:
            raise ValueError(f"No .xlsx files found in folder: {input_path}")
        return files

    raise ValueError(f"Input is not a supported file or folder: {input_path}")


def resolve_output_folder(output_path: Path) -> Path:
    if output_path.exists() and not output_path.is_dir():
        raise ValueError(f"--output must be a folder: {output_path}")
    output_path.mkdir(parents=True, exist_ok=True)
    return output_path


def find_sheet(workbook, sheet_name: str):
    if sheet_name in workbook.sheetnames:
        return workbook[sheet_name]
    if len(workbook.sheetnames) == 1:
        return workbook[workbook.sheetnames[0]]
    raise ValueError(
        f"Worksheet not found: '{sheet_name}' "
        f"(available worksheets: {', '.join(workbook.sheetnames)})"
    )


def get_header_map(ws) -> dict[str, int]:
    result: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        header = normalize_header(ws.cell(1, col).value)
        if header:
            result[header] = col
    return result


def require_columns(header_map: dict[str, int], required: list[str]) -> None:
    missing = [name for name in required if name.lower() not in header_map]
    if missing:
        raise ValueError("Required columns not found: " + ", ".join(missing))


def copy_row(ws, source_row: int, target_row: int, max_column: int) -> None:
    for col in range(1, max_column + 1):
        source = ws.cell(source_row, col)
        target = ws.cell(target_row, col)
        target.value = source.value
        target._style = copy.copy(source._style)
        target.font = copy.copy(source.font)
        target.fill = copy.copy(source.fill)
        target.border = copy.copy(source.border)
        target.alignment = copy.copy(source.alignment)
        target.protection = copy.copy(source.protection)
        target.number_format = source.number_format

    source_dim = ws.row_dimensions[source_row]
    target_dim = ws.row_dimensions[target_row]
    target_dim.height = source_dim.height
    target_dim.hidden = source_dim.hidden
    target_dim.outlineLevel = source_dim.outlineLevel
    target_dim.collapsed = source_dim.collapsed


def insert_pa_column(ws, description_col: int) -> int:
    pa_col = description_col + 1
    ws.insert_cols(pa_col, 1)

    source_header = ws.cell(1, description_col)
    header = ws.cell(1, pa_col)
    header.value = "P/A"
    header._style = copy.copy(source_header._style)
    header.font = copy.copy(source_header.font)
    header.fill = copy.copy(source_header.fill)
    header.border = copy.copy(source_header.border)
    header.alignment = Alignment(horizontal="center", vertical="center")

    ws.column_dimensions[get_column_letter(pa_col)].width = 7
    return pa_col


def style_pa_cell(cell, value: str) -> None:
    """P/A uses the row background and borders without special coloring."""
    original_fill = copy.copy(cell.fill)
    original_border = copy.copy(cell.border)
    cell.value = value
    cell.fill = original_fill
    cell.border = original_border
    cell.font = Font(bold=True, color="000000")
    cell.alignment = Alignment(horizontal="center", vertical="center")


def clear_cell(ws, row: int, col: int | None) -> None:
    if col is not None:
        ws.cell(row, col).value = None


def apply_row_fill(ws, row: int, max_column: int, fill: PatternFill) -> None:
    for col in range(1, max_column + 1):
        ws.cell(row, col).fill = copy.copy(fill)


def force_date_format(ws, date_columns: list[int]) -> None:
    for row in range(2, ws.max_row + 1):
        for col in date_columns:
            ws.cell(row, col).number_format = DATE_FORMAT



def rebuild_outline_grouping(
    ws,
    row_type_col: int,
    outline_level_col: int | None,
    pa_col: int,
) -> None:
    """Rebuild WBS grouping after inserting P/A rows to keep RowDimension aligned."""

    # Clear stale or shifted RowDimension data after insert_rows().
    for row in range(2, ws.max_row + 1):
        dim = ws.row_dimensions[row]
        dim.outlineLevel = 0
        dim.hidden = False
        dim.collapsed = False

    if outline_level_col is None:
        return

    current_level = 0

    for row in range(2, ws.max_row + 1):
        row_type = normalize_header(ws.cell(row, row_type_col).value)
        pa_value = normalize_header(ws.cell(row, pa_col).value)

        if row_type in {"wbs", "activity"}:
            raw_level = ws.cell(row, outline_level_col).value
            try:
                current_level = max(0, min(8, int(raw_level)))
            except (TypeError, ValueError):
                current_level = 0

            ws.row_dimensions[row].outlineLevel = current_level

        elif pa_value == "a":
            # Use the same outline level for the Actual row as its Plan row.
            ws.row_dimensions[row].outlineLevel = current_level

        else:
            current_level = 0
            ws.row_dimensions[row].outlineLevel = 0

    # Place +/- controls above groups to match Primavera-style WBS structure.
    ws.sheet_properties.outlinePr.summaryBelow = False
    ws.sheet_properties.outlinePr.showOutlineSymbols = True


def split_plan_actual_values(
    ws,
    plan_row: int,
    actual_row: int,
    metadata_columns: list[int],
    description_col: int,
    plan_start_col: int,
    plan_finish_col: int,
    actual_start_col: int,
    actual_finish_col: int,
    percent_col: int | None,
    physical_col: int | None,
    total_float_col: int | None,
    repeat_description: bool,
) -> None:
    # The Plan row stores item identity and planned schedule data.
    clear_cell(ws, plan_row, actual_start_col)
    clear_cell(ws, plan_row, actual_finish_col)
    clear_cell(ws, plan_row, percent_col)
    clear_cell(ws, plan_row, physical_col)

    # The Actual row displays actual data:
    # Keep only P/A, Actual Start, Actual Finish, and Progress.
    clear_cell(ws, actual_row, plan_start_col)
    clear_cell(ws, actual_row, plan_finish_col)
    clear_cell(ws, actual_row, total_float_col)

    for col in metadata_columns:
        clear_cell(ws, actual_row, col)

    if repeat_description:
        ws.cell(actual_row, description_col).value = ws.cell(plan_row, description_col).value


def inspect_file(input_file: Path, sheet_name: str) -> tuple[str, int, int, int]:
    wb = load_workbook(input_file, read_only=True, data_only=False)
    ws = find_sheet(wb, sheet_name)
    header_map = get_header_map(ws)
    require_columns(
        header_map,
        ["row type", "description", "plan start", "plan finish"],
    )

    row_type_col = header_map["row type"]
    activity_count = 0
    wbs_count = 0
    for row in range(2, ws.max_row + 1):
        row_type = normalize_header(ws.cell(row, row_type_col).value)
        if row_type == "activity":
            activity_count += 1
        elif row_type == "wbs":
            wbs_count += 1

    result = (ws.title, ws.max_row - 1, wbs_count, activity_count)
    wb.close()
    return result


def transform_file(
    input_file: Path,
    output_file: Path,
    sheet_name: str,
    repeat_description: bool,
) -> tuple[int, int, int]:
    wb = load_workbook(input_file)
    ws = find_sheet(wb, sheet_name)

    header_map = get_header_map(ws)
    require_columns(
        header_map,
        [
            "row type",
            "description",
            "plan start",
            "plan finish",
            "actual start",
            "actual finish",
        ],
    )

    pa_col = insert_pa_column(ws, header_map["description"])
    header_map = get_header_map(ws)

    row_type_col = header_map["row type"]
    wbs_col = header_map.get("wbs")
    description_col = header_map["description"]
    task_id_col = header_map.get("task id")
    uid_col = header_map.get("uid")
    outline_level_col = header_map.get("outline level")
    plan_start_col = header_map["plan start"]
    plan_finish_col = header_map["plan finish"]
    actual_start_col = header_map["actual start"]
    actual_finish_col = header_map["actual finish"]
    percent_col = header_map.get("% complete")
    physical_col = header_map.get("physical %")
    total_float_col = header_map.get("total float (hr)")

    original_last_row = ws.max_row
    activity_count = 0
    wbs_count = 0

    # Process bottom-up so every WBS and Activity receives a P/A pair.
    for row in range(original_last_row, 1, -1):
        row_type = normalize_header(ws.cell(row, row_type_col).value)
        if row_type not in {"wbs", "activity"}:
            continue

        if row_type == "wbs":
            wbs_count += 1
        else:
            activity_count += 1

        ws.insert_rows(row + 1, 1)
        copy_row(ws, row, row + 1, ws.max_column)

        plan_row = row
        actual_row = row + 1

        metadata_columns = [
            col
            for col in (
                row_type_col,
                wbs_col,
                description_col,
                task_id_col,
                uid_col,
                outline_level_col,
            )
            if col is not None
        ]

        split_plan_actual_values(
            ws=ws,
            plan_row=plan_row,
            actual_row=actual_row,
            metadata_columns=metadata_columns,
            description_col=description_col,
            plan_start_col=plan_start_col,
            plan_finish_col=plan_finish_col,
            actual_start_col=actual_start_col,
            actual_finish_col=actual_finish_col,
            percent_col=percent_col,
            physical_col=physical_col,
            total_float_col=total_float_col,
            repeat_description=repeat_description,
        )

        if row_type == "activity":
            # Readable zebra pattern: white Plan rows and very light gray Actual rows.
            apply_row_fill(ws, plan_row, ws.max_column, WHITE_FILL)
            apply_row_fill(ws, actual_row, ws.max_column, VERY_LIGHT_GRAY_FILL)
        # Both WBS P/A rows retain the WBS level color from Script 01.

        # Set P/A after row styling so the P/A cell inherits the row color.
        style_pa_cell(ws.cell(plan_row, pa_col), "P")
        style_pa_cell(ws.cell(actual_row, pa_col), "A")

    force_date_format(
        ws,
        [plan_start_col, plan_finish_col, actual_start_col, actual_finish_col],
    )

    # insert_rows() may shift RowDimension or OutlineLevel incorrectly.
    # Rebuild WBS grouping from the Outline Level column after all P/A rows are created.
    rebuild_outline_grouping(
        ws=ws,
        row_type_col=row_type_col,
        outline_level_col=outline_level_col,
        pa_col=pa_col,
    )

    if ws.auto_filter.ref:
        last_col_letter = get_column_letter(ws.max_column)
        ws.auto_filter.ref = f"A1:{last_col_letter}{ws.max_row}"

    output_file.parent.mkdir(parents=True, exist_ok=True)
    final_rows = ws.max_row - 1
    wb.save(output_file)
    wb.close()

    return wbs_count, activity_count, final_rows


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="02_insert_plan_actual_rows.py",
        description=(
            "Add the P/A column and create Plan/Actual rows "
            "for every WBS and Activity."
        ),
    )
    parser.add_argument(
        "--input",
        required=True,
        type=Path,
        help="An .xlsx file or a folder containing .xlsx files.",
    )
    parser.add_argument(
        "--output",
        required=True,
        type=Path,
        help="Output folder.",
    )
    parser.add_argument(
        "--sheet",
        default=DEFAULT_SHEET,
        help=f"Source worksheet name (default: {DEFAULT_SHEET})",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Validate files and count rows without creating Excel output.",
    )
    parser.add_argument(
        "--repeat-description",
        action="store_true",
        help="Repeat Description on Actual rows while leaving other metadata blank.",
    )
    return parser


def main() -> int:
    args = build_parser().parse_args()
    input_path = args.input.expanduser().resolve()
    output_path = args.output.expanduser().resolve()

    try:
        input_files = resolve_input_files(input_path)
        output_folder = (
            output_path if args.dry_run else resolve_output_folder(output_path)
        )
    except (FileNotFoundError, OSError, ValueError) as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1

    print(f"INPUT  : {input_path}")
    print(f"OUTPUT : {output_folder}")
    print(f"FILES  : {len(input_files)}")
    print(f"MODE   : {'DRY RUN' if args.dry_run else 'EXPORT'}")
    print()

    errors = 0
    for index, input_file in enumerate(input_files, start=1):
        print(f"[{index}/{len(input_files)}] {input_file.name}")
        try:
            if args.dry_run:
                sheet, source_rows, wbs_count, activity_count = inspect_file(
                    input_file, args.sheet
                )
                result_rows = (wbs_count + activity_count) * 2
                print(f"    SHEET           : {sheet}")
                print(f"    SOURCE ROWS     : {source_rows}")
                print(f"    WBS ROWS        : {wbs_count}")
                print(f"    ACTIVITY ROWS   : {activity_count}")
                print(f"    NEW WBS A ROWS  : {wbs_count}")
                print(f"    NEW ACTUAL ROWS : {activity_count}")
                print(f"    RESULT ROWS     : {result_rows}")
            else:
                output_file = output_folder / f"{input_file.stem}_plan_actual.xlsx"
                wbs_count, activity_count, result_rows = transform_file(
                    input_file,
                    output_file,
                    args.sheet,
                    args.repeat_description,
                )
                print(f"    OK              : {output_file}")
                print(f"    WBS PAIRS       : {wbs_count}")
                print(f"    ACTIVITY PAIRS  : {activity_count}")
                print(f"    RESULT ROWS     : {result_rows}")
        except Exception as exc:
            errors += 1
            print(f"ERROR: {input_file.name}: {exc}", file=sys.stderr)
        print()

    if errors:
        print(f"Completed with errors in {errors} file(s)", file=sys.stderr)
        return 1

    if args.dry_run:
        print("DRY RUN passed: no Excel files were created.")
    else:
        print(f"SUCCESS {output_folder}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
