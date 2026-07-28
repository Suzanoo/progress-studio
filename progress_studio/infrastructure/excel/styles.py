
from __future__ import annotations

import re
from openpyxl.styles import Border, Font, PatternFill, Side

_HEX_COLOR = re.compile(r"^[0-9A-F]{8}$")


def normalize_argb(value: str) -> str:
    color = str(value).strip().replace("#", "").upper()
    if len(color) == 6:
        color = "FF" + color
    if not _HEX_COLOR.fullmatch(color):
        raise ValueError(
            "Invalid Excel color. Expected #RRGGBB, RRGGBB, "
            f"or AARRGGBB; received {value!r}"
        )
    return color


def solid_fill(color: str) -> PatternFill:
    argb = normalize_argb(color)
    return PatternFill(fill_type="solid", fgColor=argb, bgColor=argb)


def theme_font(
    *,
    color: str = "000000",
    bold: bool = False,
    size: float | None = None,
) -> Font:
    return Font(color=normalize_argb(color), bold=bold, size=size)


def theme_side(color: str = "B7B7B7", style: str = "thin") -> Side:
    return Side(style=style, color=normalize_argb(color))


def thin_border(color: str = "B7B7B7") -> Border:
    side = theme_side(color)
    return Border(left=side, right=side, top=side, bottom=side)


HEADER_FILL = normalize_argb("1F4E78")
HEADER_FONT = normalize_argb("FFFFFF")
PROJECT_FILL = normalize_argb("B4C6E7")
PROJECT_FONT = normalize_argb("000000")
WBS_FILL = normalize_argb("D9E1F2")
WBS_FONT = normalize_argb("1F1F1F")
ACTIVITY_PLAN_FILL = normalize_argb("DDEBF7")
ACTIVITY_ACTUAL_FILL = normalize_argb("E2F0D9")
PLAN_FILL = normalize_argb("D9EAF7")
PLAN_FONT = normalize_argb("1F1F1F")
ACTUAL_FILL = normalize_argb("E2F0D9")
ACTUAL_FONT = normalize_argb("1F1F1F")
BORDER_COLOR = normalize_argb("B7B7B7")
WARNING_FILL = normalize_argb("FFC7CE")
WARNING_FONT = normalize_argb("9C0006")
