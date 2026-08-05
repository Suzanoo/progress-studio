from __future__ import annotations

from dataclasses import asdict
from datetime import date, datetime, time, timezone
import hashlib
import json
import os
from pathlib import Path
import tempfile
from typing import Any, Callable

from openpyxl import load_workbook

from progress_studio.domain.mapping_models import ActivityRow, AllocationRecord, SupplementalWBS
from progress_studio.domain.working_tree import (
    WorkingScheduleTree, WorkingTreeNode, WorkingNodeKind, WorkingNodeOrigin,
)
from progress_studio.infrastructure.platform_paths import user_data_dir

from progress_studio.domain.mapping_session import (
    MappingSessionData,
    SESSION_FORMAT,
    SESSION_VERSION,
    WorkbookFingerprint,
)


class SessionValidationError(ValueError):
    """Raised when a saved mapping session cannot be safely restored."""


def _atomic_json_write(path: Path, payload: object) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fd, temp_name = tempfile.mkstemp(
        prefix=f".{path.name}.", suffix=".tmp", dir=str(path.parent)
    )
    temp_path = Path(temp_name)
    try:
        with os.fdopen(fd, "w", encoding="utf-8", newline="\n") as stream:
            json.dump(payload, stream, ensure_ascii=False, indent=2)
            stream.write("\n")
            stream.flush()
            os.fsync(stream.fileno())
        os.replace(temp_path, path)
    except Exception:
        temp_path.unlink(missing_ok=True)
        raise


def _normalise_cell_value(value: object) -> str:
    """Return a deterministic representation for workbook identity hashing."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(timespec="microseconds")
    if isinstance(value, (date, time)):
        return value.isoformat()
    if isinstance(value, float):
        return format(value, ".17g")
    if isinstance(value, bytes):
        return value.hex()
    return str(value)


def semantic_fingerprint(path: Path) -> str:
    """Hash workbook meaning rather than the XLSX/ZIP byte stream.

    Excel commonly rewrites package metadata, style tables and ZIP ordering even
    when worksheet data is unchanged. Those changes must not break project
    relinking. This digest intentionally covers sheet order/names and every
    non-empty cell's coordinate, data type, formula or value. Formatting, file
    timestamps, calculation caches and document metadata are ignored.
    """
    keep_vba = path.suffix.lower() == ".xlsm"
    try:
        workbook = load_workbook(
            path, read_only=True, data_only=False, keep_links=False, keep_vba=keep_vba
        )
    except Exception as exc:
        raise SessionValidationError(
            f"Workbook cannot be opened as an Excel file: {path}"
        ) from exc

    digest = hashlib.sha256()
    try:
        digest.update(b"progress-studio-workbook-identity-v1\0")
        for sheet in workbook.worksheets:
            digest.update(b"S\0")
            digest.update(sheet.title.encode("utf-8", errors="surrogatepass"))
            digest.update(b"\0")
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    digest.update(b"C\0")
                    digest.update(cell.coordinate.encode("ascii"))
                    digest.update(b"\0")
                    digest.update(str(cell.data_type or "").encode("ascii", errors="ignore"))
                    digest.update(b"\0")
                    digest.update(
                        _normalise_cell_value(cell.value).encode(
                            "utf-8", errors="surrogatepass"
                        )
                    )
                    digest.update(b"\0")
    finally:
        workbook.close()
    return digest.hexdigest()


def fingerprint(path: Path) -> WorkbookFingerprint:
    path = Path(path).expanduser().resolve()
    if not path.is_file():
        raise SessionValidationError(f"Workbook was not found: {path}")
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    stat = path.stat()

    semantic_sha256 = ""
    identity_kind = "binary-sha256"
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        try:
            semantic_sha256 = semantic_fingerprint(path)
        except SessionValidationError:
            # Preserve compatibility with legacy tests and corrupted/non-Excel
            # files carrying an Excel suffix. The binary fingerprint remains
            # available and validation stays strict.
            semantic_sha256 = ""
        else:
            identity_kind = "excel-semantic-v1"

    return WorkbookFingerprint(
        path=str(path),
        filename=path.name,
        size=stat.st_size,
        modified_ns=stat.st_mtime_ns,
        sha256=digest.hexdigest(),
        semantic_sha256=semantic_sha256,
        identity_kind=identity_kind,
    )


def _migrate_v1_to_v2(payload: dict[str, Any]) -> dict[str, Any]:
    """Add explicit workbook filenames while preserving v1 compatibility."""
    migrated = dict(payload)
    for key in ("progress", "boq"):
        workbook = dict(migrated.get(key) or {})
        workbook.setdefault("filename", Path(str(workbook.get("path", ""))).name)
        migrated[key] = workbook
    migrated["version"] = 2
    return migrated


Migration = Callable[[dict[str, Any]], dict[str, Any]]
def _migrate_v2_to_v3(payload: dict[str, Any]) -> dict[str, Any]:
    migrated = dict(payload)
    migrated.setdefault("supplemental_activities", [])
    migrated["version"] = 3
    return migrated


def _migrate_v3_to_v4(payload: dict[str, Any]) -> dict[str, Any]:
    migrated = dict(payload)
    migrated.setdefault("supplemental_wbs", [])
    migrated["version"] = 4
    return migrated


def _migrate_v4_to_v5(payload: dict[str, Any]) -> dict[str, Any]:
    """Assign stable editor identities to previously created nodes."""
    migrated = dict(payload)
    activities = []
    for item in migrated.get("supplemental_activities", []):
        record = dict(item)
        path = "/".join(part[0] for part in record.get("wbs_path", []))
        record.setdefault(
            "node_id",
            WorkingScheduleTree.legacy_created_node_id(
                "activity", f"{record.get('activity_id', '')}:{path}"
            ),
        )
        activities.append(record)
    wbs_nodes = []
    for item in migrated.get("supplemental_wbs", []):
        record = dict(item)
        parent = "/".join(part[0] for part in record.get("parent_path", []))
        record.setdefault(
            "node_id",
            WorkingScheduleTree.legacy_created_node_id(
                "wbs", f"{parent}:{record.get('code', '')}"
            ),
        )
        wbs_nodes.append(record)
    migrated["supplemental_activities"] = activities
    migrated["supplemental_wbs"] = wbs_nodes
    migrated["version"] = 5
    return migrated




def _migrate_v5_to_v6(payload: dict[str, Any]) -> dict[str, Any]:
    migrated = dict(payload)
    migrated.setdefault("working_tree_nodes", [])
    migrated["version"] = 6
    return migrated


def _migrate_v6_to_v7(payload: dict[str, Any]) -> dict[str, Any]:
    """Keep legacy binary fingerprints; new saves gain semantic identities."""
    migrated = dict(payload)
    for key in ("progress", "boq"):
        workbook = dict(migrated.get(key) or {})
        workbook.setdefault("semantic_sha256", "")
        workbook.setdefault("identity_kind", "binary-sha256")
        migrated[key] = workbook
    migrated["version"] = 7
    return migrated


_MIGRATIONS: dict[int, Migration] = {
    1: _migrate_v1_to_v2,
    2: _migrate_v2_to_v3,
    3: _migrate_v3_to_v4,
    4: _migrate_v4_to_v5,
    5: _migrate_v5_to_v6,
    6: _migrate_v6_to_v7,
}


def _migrate_payload(payload: dict[str, Any]) -> dict[str, Any]:
    try:
        version = int(payload.get("version"))
    except (TypeError, ValueError) as exc:
        raise SessionValidationError("The mapping session version is missing or invalid.") from exc

    if version > SESSION_VERSION:
        raise SessionValidationError(
            f"Unsupported mapping session version: {version}. "
            f"This application supports up to version {SESSION_VERSION}."
        )

    migrated = dict(payload)
    while version < SESSION_VERSION:
        migration = _MIGRATIONS.get(version)
        if migration is None:
            raise SessionValidationError(
                f"No migration path is available from mapping session version {version}."
            )
        migrated = migration(migrated)
        try:
            next_version = int(migrated.get("version"))
        except (TypeError, ValueError) as exc:
            raise SessionValidationError("A mapping session migration produced an invalid version.") from exc
        if next_version <= version:
            raise SessionValidationError("A mapping session migration did not advance the version.")
        version = next_version
    return migrated


class MappingSessionRepository:
    """Persist mapping allocations as a small, atomic JSON sidecar."""

    @staticmethod
    def fingerprint(path: Path) -> WorkbookFingerprint:
        """Compute a workbook identity at an explicit load/relink boundary."""
        return fingerprint(path)

    def create(
        self,
        progress_file: Path,
        boq_file: Path,
        boq_sheet: str,
        allocations: list[AllocationRecord],
        supplemental_activities: list[ActivityRow] | None = None,
        supplemental_wbs: list[SupplementalWBS] | None = None,
        working_tree_nodes: list[WorkingTreeNode] | None = None,
        progress_fingerprint: WorkbookFingerprint | None = None,
        boq_fingerprint: WorkbookFingerprint | None = None,
    ) -> MappingSessionData:
        """Build session data, reusing identities already verified in memory.

        Workbook hashing is intentionally optional here. Interactive callers cache
        identities when a workbook is loaded or relinked, so high-frequency
        autosaves only serialize the small project JSON. Non-GUI callers retain
        the safe legacy behaviour by omitting the cached fingerprints.
        """
        return MappingSessionData(
            progress=progress_fingerprint or fingerprint(progress_file),
            boq=boq_fingerprint or fingerprint(boq_file),
            boq_sheet=boq_sheet,
            allocations=tuple(allocations),
            saved_at=datetime.now(timezone.utc).isoformat(),
            supplemental_activities=tuple(supplemental_activities or ()),
            supplemental_wbs=tuple(supplemental_wbs or ()),
            working_tree_nodes=tuple(working_tree_nodes or ()),
        )

    def save(self, path: Path, session: MappingSessionData) -> Path:
        path = Path(path)
        payload = {
            "format": session.format,
            "version": session.version,
            "saved_at": session.saved_at,
            "progress": asdict(session.progress),
            "boq": asdict(session.boq),
            "boq_sheet": session.boq_sheet,
            "allocations": [asdict(record) for record in session.allocations],
            "supplemental_activities": [asdict(record) for record in session.supplemental_activities],
            "supplemental_wbs": [asdict(record) for record in session.supplemental_wbs],
            "working_tree_nodes": [asdict(record) for record in session.working_tree_nodes],
        }
        _atomic_json_write(path, payload)
        return path

    def load(self, path: Path) -> MappingSessionData:
        path = Path(path)
        try:
            raw_payload = json.loads(path.read_text(encoding="utf-8"))
        except FileNotFoundError as exc:
            raise SessionValidationError(f"Mapping session was not found: {path}") from exc
        except (OSError, json.JSONDecodeError) as exc:
            raise SessionValidationError(f"Mapping session cannot be read: {path}") from exc

        if not isinstance(raw_payload, dict):
            raise SessionValidationError("The mapping session root must be a JSON object.")
        if raw_payload.get("format") != SESSION_FORMAT:
            raise SessionValidationError("This file is not a Progress Studio mapping session.")
        payload = _migrate_payload(raw_payload)
        try:
            progress = WorkbookFingerprint(**payload["progress"])
            boq = WorkbookFingerprint(**payload["boq"])
            allocations = tuple(AllocationRecord(**item) for item in payload.get("allocations", []))
            supplemental_activities = tuple(
                ActivityRow(**{**item, "wbs_path": tuple(tuple(part) for part in item.get("wbs_path", ()))})
                for item in payload.get("supplemental_activities", [])
            )
            supplemental_wbs = tuple(
                SupplementalWBS(**{**item, "parent_path": tuple(tuple(part) for part in item.get("parent_path", ()))})
                for item in payload.get("supplemental_wbs", [])
            )
            working_tree_nodes = tuple(
                WorkingTreeNode(
                    **{
                        **item,
                        "kind": WorkingNodeKind(item["kind"]),
                        "origin": WorkingNodeOrigin(item["origin"]),
                        "source_path": tuple(tuple(part) for part in item.get("source_path", ())),
                    }
                )
                for item in payload.get("working_tree_nodes", [])
            )
            boq_sheet = str(payload["boq_sheet"]).strip()
            saved_at = str(payload["saved_at"])
        except (KeyError, TypeError, ValueError) as exc:
            raise SessionValidationError("The mapping session is incomplete or malformed.") from exc
        if not boq_sheet:
            raise SessionValidationError("The mapping session does not specify a BOQ worksheet.")
        return MappingSessionData(
            progress=progress,
            boq=boq,
            boq_sheet=boq_sheet,
            allocations=allocations,
            saved_at=saved_at,
            supplemental_activities=supplemental_activities,
            supplemental_wbs=supplemental_wbs,
            working_tree_nodes=working_tree_nodes,
        )

    @staticmethod
    def validate_workbook(saved: WorkbookFingerprint, candidate: Path | None = None) -> Path:
        """Return a verified workbook path.

        The saved absolute path is tried by default. A user-selected candidate may be
        supplied when the workbook was moved or renamed. New sessions use a stable
        Excel semantic identity, so harmless re-saves and formatting changes are
        accepted while worksheet data changes are rejected. Legacy sessions retain
        strict SHA-256 verification until they are saved again.
        """
        path = Path(candidate).expanduser().resolve() if candidate else saved.saved_path
        try:
            current = fingerprint(path)
        except SessionValidationError as exc:
            if candidate is None:
                raise SessionValidationError(
                    f"{saved.filename} was not found at its saved location. "
                    "Browse for the moved or renamed workbook to continue."
                ) from exc
            raise
        if saved.has_semantic_identity and current.has_semantic_identity:
            matches = current.semantic_sha256 == saved.semantic_sha256
        else:
            # Sessions created before v7 do not contain enough information to
            # verify a changed binary safely, so they retain strict SHA-256
            # behaviour. Re-saving the project upgrades it for future relinks.
            matches = current.sha256 == saved.sha256

        if not matches:
            source = "selected workbook" if candidate else "workbook at the saved location"
            raise SessionValidationError(
                f"The {source} does not match the project workbook: {saved.filename}. "
                "Progress Studio will not merge workbook data from another version "
                "or project automatically."
            )
        return path


class RecentSessionRepository:
    """Keep a compact list of recently opened mapping-session files."""

    MAX_ITEMS = 10

    def __init__(self, path: Path | None = None) -> None:
        self.path = path or (user_data_dir() / "recent_sessions.json")

    def list(self) -> list[Path]:
        try:
            payload = json.loads(self.path.read_text(encoding="utf-8"))
        except (FileNotFoundError, OSError, json.JSONDecodeError):
            return []
        result: list[Path] = []
        for raw in payload if isinstance(payload, list) else []:
            path = Path(str(raw))
            if path.is_file() and path not in result:
                result.append(path)
        return result[: self.MAX_ITEMS]

    def remember(self, session_path: Path) -> None:
        session_path = Path(session_path).expanduser().resolve()
        items = [path for path in self.list() if path != session_path]
        items.insert(0, session_path)
        _atomic_json_write(self.path, [str(path) for path in items[: self.MAX_ITEMS]])
