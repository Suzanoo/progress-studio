
from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook


def count_formulas(path: Path, sheet: str = "main_monthly") -> tuple[int, int, int]:
    wb = load_workbook(path, read_only=False, data_only=False)
    try:
        ws = wb[sheet]
        formula_count = 0
        populated = 0
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if value not in (None, ""):
                    populated += 1
                if isinstance(value, str) and value.startswith("="):
                    formula_count += 1
        return formula_count, populated, path.stat().st_size
    finally:
        wb.close()


def main() -> int:
    parser = argparse.ArgumentParser(description="LW-10 Monthly baseline metrics")
    parser.add_argument("workbook", type=Path)
    args = parser.parse_args()
    formulas, populated, size = count_formulas(args.workbook)
    print(f"monthly_formula_cells={formulas}")
    print(f"monthly_populated_cells={populated}")
    print(f"file_bytes={size}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
