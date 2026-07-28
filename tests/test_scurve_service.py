from datetime import datetime
from pathlib import Path
import tempfile
import unittest

from openpyxl import Workbook

from progress_studio.services.scurve_service import SCurveService


class SCurveServiceTests(unittest.TestCase):
    def test_builds_weighted_cumulative_plan_and_actual(self) -> None:
        with tempfile.TemporaryDirectory() as folder:
            path = Path(folder) / "progress.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "main"
            ws.cell(3, 5, "W1")
            ws.cell(3, 6, "W2")
            ws.cell(4, 1, "Row Type")
            ws.cell(4, 2, "P/A")
            ws.cell(4, 3, "Amount")
            ws.cell(4, 5, datetime(2026, 1, 2))
            ws.cell(4, 6, datetime(2026, 1, 9))

            ws.append(["Activity", "P", 100, None, 0.5, 0.5])
            ws.append([None, "A", None, None, 0.25, None])
            ws.append(["Activity", "P", 300, None, 0.0, 1.0])
            ws.append([None, "A", None, None, 0.0, 0.5])
            wb.save(path)
            wb.close()

            data = SCurveService().read(path)

        self.assertEqual(tuple(round(value, 2) for value in data.plan), (12.5, 100.0))
        self.assertEqual(tuple(None if value is None else round(value, 2) for value in data.actual), (6.25, 43.75))

    def test_actual_stops_after_last_entered_week(self) -> None:
        with tempfile.TemporaryDirectory() as folder:
            path = Path(folder) / "progress.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "main"
            for col, week in enumerate(("W1", "W2", "W3"), start=5):
                ws.cell(3, col, week)
                ws.cell(4, col, datetime(2026, 1, col))
            ws.cell(4, 1, "Row Type")
            ws.cell(4, 2, "P/A")
            ws.cell(4, 3, "Amount")
            ws.append(["Activity", "P", 100, None, 0.2, 0.3, 0.5])
            ws.append([None, "A", None, None, 0.1, None, None])
            wb.save(path)
            wb.close()

            data = SCurveService().read(path)

        self.assertEqual(data.actual[0], 10.0)
        self.assertIsNone(data.actual[1])
        self.assertIsNone(data.actual[2])


if __name__ == "__main__":
    unittest.main()
