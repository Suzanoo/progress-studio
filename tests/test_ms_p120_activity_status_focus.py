from __future__ import annotations

from datetime import date

from openpyxl import Workbook

from progress_studio.infrastructure.excel.dashboard_workbook import (
    DASHBOARD_SHEET,
    build_dashboard,
)


def _workbook() -> Workbook:
    wb = Workbook()
    progress = wb.active
    progress.title = "progress"
    progress.append(["project_start", "project_finish", "week_start", "plan", "actual"])
    progress.append([date(2026,1,1), date(2026,2,28), date(2026,1,2), 20, 10])
    progress.append([date(2026,1,1), date(2026,2,28), date(2026,1,9), 60, 25])

    table = wb.create_sheet("progress_table")
    table.append([
        "WBS", "Activities", "Amount", "P/A", "%Progress",
        date(2026,1,2), date(2026,1,9), "_Kind",
    ])
    table.append(["1.1", "Activity A", 1000, "P", 60, 20, 40, "activity"])
    table.append(["1.1", "Activity A", 1000, "A", 25, 10, 15, "activity"])
    table.append(["1.2", "Activity B", 500, "P", 100, 50, 50, "activity"])
    table.append(["1.2", "Activity B", 500, "A", 100, 50, 50, "activity"])
    return wb


def test_activity_progress_keeps_two_rows_and_adds_variance_status() -> None:
    wb = _workbook()
    build_dashboard(wb)
    ws = wb[DASHBOARD_SHEET]

    assert ws["N38"].value == "Variance"
    assert ws["P38"].value == "Status"

    assert ws["F39"].value == "Plan"
    assert ws["F40"].value == "Actual"
    assert ws["N39"].value == ""
    assert '"Behind"' in ws["P39"].value
    assert ws["N40"].value == "=IFERROR(L40-L39,0)"
    assert '"Behind"' in ws["P40"].value
    assert '"Complete"' in ws["P40"].value

    # Outline grouping remains paired.
    assert ws.row_dimensions[39].outlineLevel == ws.row_dimensions[40].outlineLevel


def test_status_uses_native_excel_filter_and_keeps_plan_actual_pairs() -> None:
    wb = _workbook()
    build_dashboard(wb)
    ws = wb[DASHBOARD_SHEET]

    assert ws["N37"].value is None
    assert ws["P37"].value is None
    assert ws.auto_filter.ref == "P38:P42"

    # Plan and Actual rows carry the same Status logic so native filtering keeps
    # a pair together.  Plan-row status is visually hidden, not blank.
    assert isinstance(ws["P39"].value, str) and ws["P39"].value.startswith("=IF(")
    assert isinstance(ws["P40"].value, str) and ws["P40"].value.startswith("=IF(")
    assert 'L40<L39' in ws["P39"].value
    assert 'L40<L39' in ws["P40"].value
    assert ws["P39"].font.color.type == "rgb"

    # The old Status Focus dropdown/conditional-format layer is gone.
    status_validations = [
        v for v in ws.data_validations.dataValidation
        if v.formula1 == '"All,Behind,On Track,Complete,Not Started"'
    ]
    assert status_validations == []


def test_variance_status_columns_do_not_change_okd_progress_table_contract() -> None:
    wb = _workbook()
    original_headers = [cell.value for cell in wb["progress_table"][1]]
    build_dashboard(wb)

    assert [cell.value for cell in wb["progress_table"][1]] == original_headers
