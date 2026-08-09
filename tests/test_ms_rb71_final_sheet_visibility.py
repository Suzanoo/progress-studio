from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from progress_studio.infrastructure.excel.workbook_visibility import (
    VISIBLE_SHEETS,
    apply_final_sheet_visibility,
)
from progress_studio.infrastructure.excel.payment_input_workbook import PaymentInputWorkbook
from progress_studio.services.payment_service import PaymentService
from progress_studio.services.rebuild_service import WorkbookRebuildEngine
from tests.test_ms_rb2_progress_rebuild_engine import _full_rebuild_fixture
from tests.test_ms_pay6_payment_line_renderer import _progress_workbook


def test_rb71_visibility_policy_exposes_only_final_user_sheets() -> None:
    wb = Workbook()
    wb.active.title = "main"
    for name in (
        "main_monthly",
        "Payment Input",
        "Payment",
        "Dashboard",
        "progress",
        "progress_table",
        "Dashboard_Data",
        "Info",
        "User Notes",
    ):
        wb.create_sheet(name)

    visible, hidden, very_hidden = apply_final_sheet_visibility(wb)

    assert visible == tuple(name for name in VISIBLE_SHEETS if name in wb.sheetnames)
    for name in ("main", "main_monthly", "Payment Input", "Payment", "Dashboard"):
        assert wb[name].sheet_state == "visible"
    for name in ("progress", "progress_table"):
        assert wb[name].sheet_state == "hidden"
    for name in ("Dashboard_Data", "Info", "User Notes"):
        assert wb[name].sheet_state == "veryHidden"
    assert "progress" in hidden


def test_rb71_progress_rebuild_applies_final_visibility(tmp_path: Path) -> None:
    source = _full_rebuild_fixture(tmp_path / "source.xlsx")
    output = tmp_path / "rebuilt.xlsx"

    WorkbookRebuildEngine().rebuild_progress(source, output)

    wb = load_workbook(output)
    try:
        for name in ("main", "main_monthly", "Payment Input", "Payment", "Dashboard"):
            assert wb[name].sheet_state == "visible"
        for name in ("progress", "progress_table"):
            assert wb[name].sheet_state == "hidden"
        for name in ("Dashboard_Data", "User Notes"):
            assert wb[name].sheet_state == "veryHidden"
    finally:
        wb.close()


def test_rb71_payment_input_and_payment_rebuild_keep_same_visibility_contract(tmp_path: Path) -> None:
    source = _progress_workbook(tmp_path / "source.xlsx")
    prepared = tmp_path / "prepared.xlsx"

    PaymentService().prepare_embedded_payment_input(source, prepared, periods=3)

    wb = load_workbook(prepared)
    try:
        assert wb["main"].sheet_state == "visible"
        assert wb["Payment Input"].sheet_state == "visible"
        assert "Payment" not in wb.sheetnames
    finally:
        wb.close()

    # Add enough generated support sheets to prove they are hidden after Payment rebuild.
    wb = load_workbook(prepared)
    for name in ("main_monthly", "progress", "progress_table", "Dashboard_Data", "Dashboard"):
        if name not in wb.sheetnames:
            ws = wb.create_sheet(name)
            ws["A1"] = name
    wb.save(prepared)
    wb.close()

    output = tmp_path / "payment_rebuilt.xlsx"
    WorkbookRebuildEngine().rebuild_payment(prepared, output)

    wb = load_workbook(output)
    try:
        for name in ("main", "main_monthly", "Payment Input", "Payment", "Dashboard"):
            assert wb[name].sheet_state == "visible"
        for name in ("progress", "progress_table"):
            assert wb[name].sheet_state == "hidden"
        assert wb["Dashboard_Data"].sheet_state == "veryHidden"
    finally:
        wb.close()
