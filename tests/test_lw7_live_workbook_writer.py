
from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

from progress_studio.services.rebuild_service import WorkbookRebuildEngine


def _fixture(path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "main"
    headers = [
        "Row Type", "WBS", "Description", "P/A", "% Complete",
        "Amount", "Activity ID", "Outline Level", "Plan Start", "Plan Finish",
    ]
    for col, value in enumerate(headers, 1):
        ws.cell(4, col, value)
    for col, dt in enumerate(
        (datetime(2026,3,6), datetime(2026,3,13), datetime(2026,4,3)),
        start=11,
    ):
        ws.cell(3, col, f"W{col-10}")
        ws.cell(4, col, dt)
    rows = [
        ["Activity", "1.1", "Concrete", "P", "", 1000, "A1000", 2, datetime(2026,3,1), datetime(2026,4,3), .4, .4, .2],
        ["Activity", "1.1", "Concrete", "A", "", 0, "A1000", 2, datetime(2026,3,1), datetime(2026,4,3), .2, .3, None],
    ]
    for row in rows:
        ws.append(row)

    pin = wb.create_sheet("Payment Input")
    pin["A1"] = "KEEP"
    payment = wb.create_sheet("Payment")
    payment["A1"] = "KEEP_PAYMENT"
    for name in ("main_monthly", "progress", "progress_table", "Dashboard_Data", "Dashboard"):
        if name not in wb.sheetnames:
            stale = wb.create_sheet(name)
            stale["A1"] = "STALE"
    wb.save(path)
    return path


def test_lw7_live_progress_rebuild_is_one_pass_output_contract(tmp_path: Path) -> None:
    source = _fixture(tmp_path / "source.xlsx")
    output = tmp_path / "live.xlsx"

    result = WorkbookRebuildEngine().rebuild_live_progress(source, output, project_name="Demo")

    assert result.activity_count == 1
    assert result.week_count == 3
    assert result.monthly_periods == 2
    assert output.exists()

    wb = load_workbook(output, data_only=False)
    try:
        assert "main" in wb.sheetnames
        assert "main_monthly" in wb.sheetnames
        assert "Dashboard" in wb.sheetnames
        assert "Dashboard_Data" in wb.sheetnames
        assert "progress" in wb.sheetnames
        assert "progress_table" not in wb.sheetnames
        assert wb["Payment Input"]["A1"].value == "KEEP"
        assert wb["Payment"]["A1"].value == "KEEP_PAYMENT"
        assert wb["Dashboard"]["B6"].value == "Project Start"
        assert wb["Dashboard"]["F6"].value == "Project Finish"
        assert wb["Dashboard"]["J6"].value == "Project Value"
    finally:
        wb.close()


def test_lw7_monthly_view_remains_present_in_live_writer(tmp_path: Path) -> None:
    source = _fixture(tmp_path / "source.xlsx")
    output = tmp_path / "live.xlsx"
    WorkbookRebuildEngine().rebuild_live_progress(source, output)

    wb = load_workbook(output, data_only=False)
    try:
        ws = wb["main_monthly"]
        # LW-10.0 later evolves the monthly timescale from cache values to
        # full-live direct formulas while preserving the same worksheet position.
        assert isinstance(ws["K5"].value, str) and ws["K5"].value.startswith("=")
        assert isinstance(ws["L5"].value, str) and ws["L5"].value.startswith("=")
    finally:
        wb.close()


def test_lw7_live_service_does_not_open_data_only_workbook() -> None:
    source = Path("progress_studio/services/rebuild_service.py").read_text(encoding="utf-8")
    start = source.index("    def rebuild_live_progress(")
    end = source.index("    def rebuild_live_payment(", start)
    block = source[start:end]
    assert "data_only=True" not in block
    assert block.count("load_workbook(") == 1
    assert block.count(".save(") == 1


def test_lw7_ui_keeps_explicit_live_progress_routing() -> None:
    source = Path("progress_studio/presentation/gui/rebuild.py").read_text(encoding="utf-8")
    assert "rebuild_live_progress" in source
    assert 'output_mode == "live"' in source
    # LW-9 later activates Payment through its own explicit route.
    assert "rebuild_live_payment" in source
