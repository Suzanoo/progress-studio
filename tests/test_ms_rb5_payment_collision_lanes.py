from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from progress_studio.domain.payment_models import PaymentResolvedPeriod, PaymentResolvedPoint
from progress_studio.infrastructure.excel.payment_line_renderer import PaymentLineRenderer
from tests.test_ms_pay6_payment_line_renderer import _progress_workbook


def _point(period_id: str, row: int, col: int, activity_id: str) -> PaymentResolvedPoint:
    return PaymentResolvedPoint(
        period_id=period_id,
        activity_id=activity_id,
        required_fraction=1.0,
        activity_row=row,
        timescale_column=col,
        timescale_column_letter="R",
        boundary_edge="left",
        week_start=date(2026, 5, 1),
        reached_cumulative=1.0,
    )


def test_rb5_allocator_separates_same_true_boundary() -> None:
    renderer = PaymentLineRenderer()
    periods = (
        PaymentResolvedPeriod(
            "P08", None, (_point("P08", 11, 18, "A1000"),),
            date(2026, 5, 1), ("A1000",)
        ),
        PaymentResolvedPeriod(
            "P09", None, (_point("P09", 13, 18, "A1010"),),
            date(2026, 5, 1), ("A1010",)
        ),
    )
    timeline = tuple((col, date(2026, 1, 1)) for col in range(18, 25))

    lanes = renderer._allocate_visual_lanes(periods, timeline)

    assert lanes["P08"][0] == 18
    assert lanes["P09"][0] == 18
    assert lanes["P08"][1] != lanes["P09"][1]
    assert lanes["P08"][2] == 0
    assert lanes["P09"][2] == 1


def test_rb5_collision_render_has_distinct_backbones_and_staggered_labels(tmp_path: Path) -> None:
    source = _progress_workbook(tmp_path / "progress.xlsx")
    output = tmp_path / "collision.xlsx"

    periods = (
        PaymentResolvedPeriod(
            "P08", None, (_point("P08", 11, 18, "A1000"),),
            date(2026, 5, 1), ("A1000",)
        ),
        PaymentResolvedPeriod(
            "P09", None, (_point("P09", 13, 18, "A1010"),),
            date(2026, 5, 1), ("A1010",)
        ),
    )

    PaymentLineRenderer().render_periods(source, output, periods)

    wb = load_workbook(output, data_only=False)
    try:
        ws = wb["Payment"]
        assert len(ws._images) == 2
        anchors = [
            (img.anchor._from.col + 1, img.anchor._from.row + 1)
            for img in ws._images
        ]
        assert anchors[0] != anchors[1]
        assert ws._images[0].width == 145
        assert ws._images[0].height == 26
        assert ws._images[1].width == 145
        assert ws._images[1].height == 26

        # P08 keeps the true boundary, P09 takes a nearby visual lane.
        p08_left = ws.cell(11, 18).border.left
        p09_left = ws.cell(13, 19).border.left
        assert p08_left.style is not None
        assert p09_left.style is not None

        # The shifted marker explicitly records that this is display-only.
        assert ws.cell(4, 19).comment is not None
        assert "Visual lane offset" in ws.cell(4, 19).comment.text
    finally:
        wb.close()
