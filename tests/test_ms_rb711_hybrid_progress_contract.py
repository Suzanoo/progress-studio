from __future__ import annotations

from openpyxl import load_workbook

from progress_studio.services.rebuild_service import WorkbookRebuildEngine
from tests.test_ms_rb2_progress_rebuild_engine import _full_rebuild_fixture


def test_rb711_progress_plan_is_snapshot_but_actual_is_live(tmp_path) -> None:
    source = _full_rebuild_fixture(tmp_path / "source.xlsx")
    output = tmp_path / "rebuilt.xlsx"

    WorkbookRebuildEngine().rebuild_progress(source, output)

    wb = load_workbook(output, data_only=False)
    try:
        progress = wb["progress"]

        assert progress["D2"].value == 50.0
        assert progress["D3"].value == 100.0
        assert not isinstance(progress["D2"].value, str)

        assert isinstance(progress["E2"].value, str)
        assert isinstance(progress["E3"].value, str)
        assert "main" in progress["E2"].value.lower()
        assert "main" in progress["E3"].value.lower()

        # Dashboard weekly Actual reads progress, not main directly.
        data = wb["Dashboard_Data"]
        assert isinstance(data["C2"].value, str)
        assert "'progress'!E2" in data["C2"].value
        assert "main" not in data["C2"].value.lower()

        # Monthly Actual is also live through a contiguous progress range.
        assert isinstance(data["F2"].value, str)
        assert "'progress'!E2:E3" in data["F2"].value
    finally:
        wb.close()


def test_rb711_large_support_snapshots_remain_formula_free(tmp_path) -> None:
    source = _full_rebuild_fixture(tmp_path / "source.xlsx")
    output = tmp_path / "rebuilt.xlsx"

    WorkbookRebuildEngine().rebuild_progress(source, output)

    wb = load_workbook(output, data_only=False)
    try:
        for name in ("main_monthly", "progress_table"):
            assert not any(
                isinstance(cell.value, str) and cell.value.startswith("=")
                for row in wb[name].iter_rows()
                for cell in row
            )
    finally:
        wb.close()
