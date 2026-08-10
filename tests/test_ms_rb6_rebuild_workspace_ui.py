from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from progress_studio.services.payment_service import PaymentService
from tests.test_ms_pay6_payment_line_renderer import _progress_workbook


def test_rb6_sidebar_has_standalone_rebuild_after_export() -> None:
    source = (
        Path(__file__).parents[1]
        / "progress_studio"
        / "presentation"
        / "gui"
        / "app.py"
    ).read_text(encoding="utf-8")

    export_pos = source.index('("export", "⇧", "Export")')
    rebuild_pos = source.index('("rebuild", "↻", "Rebuild")')
    settings_pos = source.index('("settings", "⚙", "Settings")')
    assert export_pos < rebuild_pos < settings_pos
    assert "RebuildFrame" in source
    assert "self._build_rebuild_workspace()" in source


def test_rb6_export_workspace_has_no_rebuild_controls() -> None:
    source = (
        Path(__file__).parents[1]
        / "progress_studio"
        / "presentation"
        / "gui"
        / "app.py"
    ).read_text(encoding="utf-8")

    start = source.index("    def _build_export_workspace")
    end = source.index("    def _build_rebuild_workspace", start)
    export_block = source[start:end]

    assert "Export Mapped Workbook" in export_block
    assert "Rebuild Latest Workbook" not in export_block
    assert "Rebuild from Edited Workbook" not in export_block
    assert "rebuild_workbook" not in export_block
    assert "rebuild_from_edited_workbook" not in export_block


def test_rb6_rebuild_ui_has_only_workbook_and_mode_inputs() -> None:
    source = (
        Path(__file__).parents[1]
        / "progress_studio"
        / "presentation"
        / "gui"
        / "rebuild.py"
    ).read_text(encoding="utf-8")

    assert "Progress Workbook" in source
    assert "Payment" in source
    assert "main is authoritative" in source
    assert "rebuild_progress" in source
    assert "rebuild_payment" in source

    # No old project/session input contract is present.
    assert "progressstudio" not in source.lower()
    assert "boqstudio" not in source.lower()
    assert "WorkingScheduleTree" not in source
    assert "mapping_store" not in source


def test_rb6_payment_workspace_is_input_only() -> None:
    source = (
        Path(__file__).parents[1]
        / "progress_studio"
        / "presentation"
        / "gui"
        / "payment.py"
    ).read_text(encoding="utf-8")

    assert "Prepare Payment Input" in source
    assert "prepare_embedded_payment_input" in source
    assert "Prepare / Rebuild Payment Workbook" not in source
    assert "render_payment_backbones" not in source
    assert "rebuild_embedded_workbook" not in source


def test_rb6_payment_input_preparation_does_not_create_payment(tmp_path: Path) -> None:
    source = _progress_workbook(tmp_path / "source.xlsx")
    output = tmp_path / "prepared.xlsx"

    stats = PaymentService().prepare_embedded_payment_input(source, output, periods=3)

    assert stats["periods"] == 3
    wb = load_workbook(output)
    try:
        assert "Payment Input" in wb.sheetnames
        assert "Payment" not in wb.sheetnames
        assert "main" in wb.sheetnames
    finally:
        wb.close()
