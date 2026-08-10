from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from progress_studio.domain.rebuild_models import RebuildMode
from progress_studio.services.rebuild_service import (
    DEFAULT_REBUILD_CONTRACT,
    RebuildContractError,
    WorkbookRebuildEngine,
)


def _workbook(path: Path, *, payment_input: bool = False) -> Path:
    wb = Workbook()
    main = wb.active
    main.title = "main"

    headers = ("Row Type", "WBS", "Description", "P/A", "Activity ID")
    for col, value in enumerate(headers, start=1):
        main.cell(4, col, value)

    main.cell(5, 1, "WBS")
    main.cell(5, 2, "1")
    main.cell(5, 3, "Structure")
    main.cell(5, 4, "P")

    main.cell(6, 1, "Activity")
    main.cell(6, 2, "1")
    main.cell(6, 3, "Test Activity")
    main.cell(6, 4, "P")
    main.cell(6, 5, "A1000")

    for name in ("main_monthly", "progress", "progress_table", "Dashboard_Data", "Dashboard"):
        wb.create_sheet(name)
    wb.create_sheet("Info")
    wb.create_sheet("User Notes")
    if payment_input:
        wb.create_sheet("Payment Input")
        wb.create_sheet("Payment")

    wb.save(path)
    wb.close()
    return path


def test_rb1_sheet_contract_is_explicit_and_mode_specific() -> None:
    contract = DEFAULT_REBUILD_CONTRACT

    assert contract.preserve == ("main", "Payment Input")
    assert contract.generated_for(RebuildMode.PROGRESS) == (
        "main_monthly",
        "progress",
        "progress_table",
        "Dashboard_Data",
        "Dashboard",
    )
    assert contract.generated_for(RebuildMode.PAYMENT) == ("Payment",)


def test_rb1_progress_analyze_uses_main_without_payment_or_sessions(tmp_path: Path) -> None:
    path = _workbook(tmp_path / "edited.xlsx", payment_input=False)
    analysis = WorkbookRebuildEngine().analyze(path, RebuildMode.PROGRESS)

    assert analysis.ready
    assert analysis.main_sheet == "main"
    assert analysis.activity_count == 1
    assert analysis.payment_input_present is False
    assert analysis.existing_generated_sheets == (
        "main_monthly",
        "progress",
        "progress_table",
        "Dashboard_Data",
        "Dashboard",
    )
    assert analysis.unknown_sheets == ("User Notes",)


def test_rb1_payment_requires_embedded_payment_input(tmp_path: Path) -> None:
    path = _workbook(tmp_path / "edited.xlsx", payment_input=False)

    with pytest.raises(RebuildContractError, match="Payment Input"):
        WorkbookRebuildEngine().analyze(path, RebuildMode.PAYMENT)


def test_rb1_payment_analyze_only_owns_payment_sheet(tmp_path: Path) -> None:
    path = _workbook(tmp_path / "edited.xlsx", payment_input=True)
    analysis = WorkbookRebuildEngine().analyze(path, "payment")

    assert analysis.ready
    assert analysis.payment_input_present
    assert analysis.existing_generated_sheets == ("Payment",)
    assert analysis.missing_generated_sheets == ()
    assert "main" in analysis.preserve_sheets_present
    assert "Payment Input" in analysis.preserve_sheets_present


def test_rb1_core_imports_no_mapping_session_or_tree_runtime() -> None:
    import ast

    source = (
        Path(__file__).parents[1]
        / "progress_studio"
        / "services"
        / "rebuild_service.py"
    ).read_text(encoding="utf-8")
    tree = ast.parse(source)
    imported = []
    for node in ast.walk(tree):
        if isinstance(node, ast.Import):
            imported.extend(alias.name for alias in node.names)
        elif isinstance(node, ast.ImportFrom):
            imported.append(node.module or "")

    joined = " ".join(imported).lower()
    assert "mapping_store" not in joined
    assert "working_tree" not in joined
    assert "mapped_workbook_exporter" not in joined
    assert "xml" not in joined
