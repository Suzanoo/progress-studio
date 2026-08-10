
from __future__ import annotations

from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook

from progress_studio.infrastructure.excel.rebuild_workbook_reader import RebuildWorkbookReader
from progress_studio.services.activity_table_deriver import ActivityTableDeriver


def _fixture(path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "main"
    headers = [
        "Row Type", "WBS", "Description", "P/A", "% Complete",
        "Amount", "Activity ID", "Outline Level", "Plan Start", "Plan Finish",
    ]
    for col, value in enumerate(headers, 1):
        ws.cell(4, col, value)
    for col, dt in enumerate((datetime(2026,3,6), datetime(2026,3,13)), start=11):
        ws.cell(3, col, f"W{col-10}")
        ws.cell(4, col, dt)

    rows = [
        ["Project Summary", "", "Project", "P", "", 1000, "", 0, datetime(2026,3,1), datetime(2026,3,31), .5, .5],
        ["Project Summary", "", "Project", "A", "", 0, "", 0, datetime(2026,3,1), datetime(2026,3,31), .2, .3],
        ["WBS", "1", "Structure", "P", "", 1000, "", 1, datetime(2026,3,1), datetime(2026,3,31), .5, .5],
        ["WBS", "1", "Structure", "A", "", 0, "", 1, datetime(2026,3,1), datetime(2026,3,31), .2, .3],
        ["Activity", "1.1", "Concrete", "P", 1.0, 1000, "A1000", 2, datetime(2026,3,1), datetime(2026,3,13), .5, .5],
        ["Activity", "1.1", "Concrete", "A", .5, 0, "A1000", 2, datetime(2026,3,1), datetime(2026,3,13), .2, .3],
    ]
    for row in rows:
        ws.append(row)
    wb.save(path)
    return path


def test_lw3_derives_two_row_activity_contract_from_main_dataset(tmp_path: Path) -> None:
    data = RebuildWorkbookReader().read_main_dataset(_fixture(tmp_path / "p.xlsx"))
    model = ActivityTableDeriver().derive(data)

    assert model.pair_count == 3
    plan, actual = model.rows[-2:]
    assert plan.type_label == "Plan"
    assert actual.type_label == "Actual"
    assert plan.activity == "Concrete"
    assert plan.total == 1000.0
    assert plan.progress == 1.0
    assert plan.amount == 1000.0
    assert actual.progress == 0.5
    assert actual.amount == 500.0
    assert actual.variance == -0.5
    assert actual.status == "Behind"
    assert plan.outline_level == actual.outline_level == 2


def test_lw3_cutoff_limits_progress_without_progress_table(tmp_path: Path) -> None:
    data = RebuildWorkbookReader().read_main_dataset(_fixture(tmp_path / "p.xlsx"))
    model = ActivityTableDeriver().derive(data, cutoff=date(2026,3,6))

    plan, actual = model.rows[-2:]
    assert plan.progress == 0.5
    assert actual.progress == 0.2
    assert actual.amount == 200.0
    assert actual.variance == -0.3
    assert actual.status == "Behind"


def test_lw3_deriver_has_no_openpyxl_or_progress_table_dependency() -> None:
    source = Path("progress_studio/services/activity_table_deriver.py").read_text(encoding="utf-8")
    assert "openpyxl" not in source
    assert "load_workbook" not in source
    # The only occurrence is the explanatory contract text; there is no sheet access.
    assert 'workbook["progress_table"]' not in source
    assert "TABLE_SHEET" not in source


def test_lw3_keeps_summary_hierarchy_for_outline(tmp_path: Path) -> None:
    data = RebuildWorkbookReader().read_main_dataset(_fixture(tmp_path / "p.xlsx"))
    model = ActivityTableDeriver().derive(data)
    assert [model.rows[i].outline_level for i in (0, 2, 4)] == [0, 1, 2]
    assert [model.rows[i].row_type for i in (0, 2, 4)] == ["project summary", "wbs", "activity"]
