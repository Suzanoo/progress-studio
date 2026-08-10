from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from progress_studio.services.rebuild_service import WorkbookRebuildEngine
from tests.test_ms_rb2_progress_rebuild_engine import _full_rebuild_fixture


def _formula_count(ws) -> int:
    return sum(
        1
        for row in ws.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=")
    )


def test_rb3_monthly_progress_table_snapshot_and_progress_actual_hybrid(tmp_path: Path) -> None:
    source = _full_rebuild_fixture(tmp_path / "source.xlsx")
    output = tmp_path / "rebuilt.xlsx"

    WorkbookRebuildEngine().rebuild_progress(source, output)

    wb = load_workbook(output, data_only=False)
    try:
        assert _formula_count(wb["main_monthly"]) == 0
        assert _formula_count(wb["progress_table"]) == 0

        progress = wb["progress"]
        # Only Actual is live: one lightweight formula per reporting week.
        assert _formula_count(progress) == 2
        assert not isinstance(progress["D2"].value, str)
        assert not isinstance(progress["D3"].value, str)
        assert isinstance(progress["E2"].value, str) and progress["E2"].value.startswith("=")
        assert isinstance(progress["E3"].value, str) and progress["E3"].value.startswith("=")
        assert "main" in progress["E2"].value.lower()

        # Large snapshots stay disconnected from main. Dashboard_Data may link
        # only through the tiny progress adapter, never directly to main.
        for sheet_name in ("main_monthly", "progress_table", "Dashboard_Data"):
            for row in wb[sheet_name].iter_rows():
                for cell in row:
                    value = cell.value
                    if isinstance(value, str) and value.startswith("="):
                        assert "'main'!" not in value
                        assert "main!" not in value
    finally:
        wb.close()


def test_rb3_snapshot_keeps_expected_monthly_values(tmp_path: Path) -> None:
    source = _full_rebuild_fixture(tmp_path / "source.xlsx")
    output = tmp_path / "rebuilt.xlsx"

    WorkbookRebuildEngine().rebuild_progress(source, output)

    wb = load_workbook(output, data_only=False)
    try:
        monthly = wb["main_monthly"]
        # Activity Plan and Actual rows aggregate the two March weeks.
        assert monthly.cell(11, 12).value == 1.0
        assert monthly.cell(12, 12).value == 0.5

        progress = wb["progress"]
        assert progress["D2"].value == 50.0
        assert progress["D3"].value == 100.0
        assert isinstance(progress["E2"].value, str)
        assert isinstance(progress["E3"].value, str)
    finally:
        wb.close()
