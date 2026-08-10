from datetime import datetime

from openpyxl import Workbook

from progress_studio.infrastructure.excel.okd_workbook import (
    HEADER_ROW,
    build_progress_table_sheet,
)


def test_progress_table_is_value_only_snapshot() -> None:
    wb = Workbook()
    main = wb.active
    main.title = "main"
    headers = [
        "Row Type", "Description", "P/A", "Activity ID", "Amount",
        "Plan Start", "Plan Finish",
    ]
    for col, value in enumerate(headers, start=1):
        main.cell(HEADER_ROW, col, value)
    cutoff = datetime(2026, 1, 23)
    main.cell(HEADER_ROW, 8, cutoff)
    main.cell(HEADER_ROW + 1, 1, "Activity")
    main.cell(HEADER_ROW + 1, 2, "Test Activity")
    main.cell(HEADER_ROW + 1, 3, "P")
    main.cell(HEADER_ROW + 1, 4, "A1000")
    main.cell(HEADER_ROW + 1, 5, 1000.0)
    main.cell(HEADER_ROW + 1, 6, datetime(2026, 1, 1))
    main.cell(HEADER_ROW + 1, 7, datetime(2026, 1, 31))
    main.cell(HEADER_ROW + 1, 8, 0.25)
    main.cell(HEADER_ROW + 2, 3, "A")
    main.cell(HEADER_ROW + 2, 8, 0.10)

    header_map = {
        str(main.cell(HEADER_ROW, col).value).strip().lower(): col
        for col in range(1, main.max_column + 1)
    }
    weeks = [(8, cutoff.date())]
    table_rows = [{
        "kind": "activity",
        "display_wbs": "1.1.1",
        "activity_name": "Test Activity",
        "amount": 1000.0,
        "plan": [0.25],
        "actual": [0.10],
        "plan_row": HEADER_ROW + 1,
        "actual_row": HEADER_ROW + 2,
    }]

    build_progress_table_sheet(wb, main, header_map, weeks, table_rows)
    table = wb["progress_table"]

    assert table["F1"].value == cutoff.date()
    assert table["B2"].value == "Test Activity"
    assert table["C2"].value == 1000.0
    assert table["E2"].value == 25.0
    assert table["F2"].value == 25.0
    assert table["E3"].value == 10.0
    assert table["F3"].value == 10.0
    assert not any(
        isinstance(cell.value, str) and cell.value.startswith("=")
        for row in table.iter_rows()
        for cell in row
    )
