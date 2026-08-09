from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

from progress_studio.services.rebuild_service import WorkbookRebuildEngine


def _full_rebuild_fixture(path: Path) -> Path:
    wb = Workbook()
    main = wb.active
    main.title = "main"

    fixed = [
        "Row Type", "WBS", "Description", "P/A", "% Complete",
        "Amount", "Activity ID", "Outline Level", "Plan Start",
        "Plan Finish", "XML Amount",
    ]
    for col, value in enumerate(fixed, start=1):
        main.cell(4, col, value)

    # Weekly grammar: Wn in row 3, reporting date in row 4.
    weeks = [datetime(2026, 3, 6), datetime(2026, 3, 13)]
    for idx, week in enumerate(weeks, start=12):
        main.cell(3, idx, f"W{idx - 11}")
        main.cell(4, idx, week)

    rows = [
        (5, "Project Summary", "", "Project", "P", "", 1000.0, "", 0, datetime(2026,3,1), datetime(2026,3,31)),
        (6, "Project Summary", "", "Project", "A", "", 0.0, "", 0, datetime(2026,3,1), datetime(2026,3,31)),
        (7, "S-Curve", "", "Acc. Plan", "AP", "", "", "", 0, None, None),
        (8, "S-Curve", "", "Acc. Actual", "AA", "", "", "", 0, None, None),
        (9, "WBS", "1", "Structure", "P", "", 1000.0, "", 1, datetime(2026,3,1), datetime(2026,3,31)),
        (10, "WBS", "1", "Structure", "A", "", 0.0, "", 1, datetime(2026,3,1), datetime(2026,3,31)),
        (11, "Activity", "1.1", "Concrete", "P", 1.0, 1000.0, "A1000", 2, datetime(2026,3,1), datetime(2026,3,13)),
        (12, "Activity", "1.1", "Concrete", "A", 0.5, 0.0, "A1000", 2, datetime(2026,3,1), datetime(2026,3,13)),
    ]
    for row, row_type, wbs, desc, pa, pct, amount, aid, level, start, finish in rows:
        values=[row_type,wbs,desc,pa,pct,amount,aid,level,start,finish,amount if row_type=="Activity" and pa=="P" else ""]
        for col,value in enumerate(values,start=1):
            main.cell(row,col,value)

    # Project/WBS/activity weekly values and cumulative rows.
    main.cell(5,12,0.5); main.cell(5,13,0.5)
    main.cell(6,12,0.2); main.cell(6,13,0.3)
    main.cell(7,12,0.5); main.cell(7,13,1.0)
    main.cell(8,12,0.2); main.cell(8,13,0.5)
    main.cell(9,12,0.5); main.cell(9,13,0.5)
    main.cell(10,12,0.2); main.cell(10,13,0.3)
    main.cell(11,12,0.5); main.cell(11,13,0.5)
    main.cell(12,12,0.2); main.cell(12,13,0.3)

    # Persistent Payment/user sheets that RB2 must not touch.
    pin = wb.create_sheet("Payment Input")
    pin["A1"] = "KEEP_PAYMENT_INPUT"
    payment = wb.create_sheet("Payment")
    payment["A1"] = "KEEP_PAYMENT"
    notes = wb.create_sheet("User Notes")
    notes["A1"] = "KEEP_USER_NOTES"

    # Stale generated sheets must be replaced.
    for name in ("main_monthly", "progress", "progress_table", "Dashboard_Data", "Dashboard"):
        ws = wb.create_sheet(name)
        ws["A1"] = "STALE"

    wb.save(path)
    wb.close()
    return path


def test_rb2_rebuild_progress_replaces_only_progress_owned_sheets(tmp_path: Path) -> None:
    source = _full_rebuild_fixture(tmp_path / "source.xlsx")
    output = tmp_path / "rebuilt.xlsx"

    result = WorkbookRebuildEngine().rebuild_progress(source, output, project_name="Demo")

    assert result.activity_count == 1
    assert result.week_count == 2
    assert result.monthly_periods == 1
    assert result.preserved_payment_sheet
    assert result.preserved_payment_input_sheet
    assert result.rebuilt_sheets == (
        "main_monthly", "progress", "progress_table", "Dashboard_Data", "Dashboard"
    )

    wb = load_workbook(output, data_only=False)
    try:
        assert wb["Payment Input"]["A1"].value == "KEEP_PAYMENT_INPUT"
        assert wb["Payment"]["A1"].value == "KEEP_PAYMENT"
        assert wb["User Notes"]["A1"].value == "KEEP_USER_NOTES"

        # main remains the source; user-entered actual stays unchanged.
        assert wb["main"]["L12"].value == 0.2
        assert wb["main"]["M12"].value == 0.3

        for name in result.rebuilt_sheets:
            assert name in wb.sheetnames
            assert wb[name]["A1"].value != "STALE"

        assert wb["progress_table"].sheet_state == "veryHidden"
        assert wb["Dashboard_Data"].sheet_state == "veryHidden"

        # Snapshot contract remains intact.
        assert not any(
            isinstance(cell.value, str) and cell.value.startswith("=")
            for row in wb["progress_table"].iter_rows()
            for cell in row
        )

        calc = wb.calculation
        assert calc.calcMode == "auto"
        assert calc.fullCalcOnLoad is False
        assert calc.forceFullCalc is False
    finally:
        wb.close()


def test_rb2_atomic_rebuild_can_replace_same_path(tmp_path: Path) -> None:
    path = _full_rebuild_fixture(tmp_path / "same.xlsx")
    result = WorkbookRebuildEngine().rebuild_progress(path, path)

    assert result.output_workbook == path.resolve()
    wb = load_workbook(path)
    try:
        assert wb["Payment"]["A1"].value == "KEEP_PAYMENT"
        assert wb["main"]["L12"].value == 0.2
        assert wb["progress"]["A1"].value == "project_start"
    finally:
        wb.close()
