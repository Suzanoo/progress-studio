from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

from progress_studio.infrastructure.excel.payment_input_workbook import PaymentInputWorkbook
from progress_studio.infrastructure.excel.payment_workbook import PaymentWorkbookSnapshotter


def _progress_workbook(path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "main"
    headers = ["Row Type", "WBS", "Description", "P/A", "Activity ID", "Task ID", "UID", "Outline Level", "Plan Start", "Plan Finish"]
    for col, value in enumerate(headers, start=1):
        ws.cell(4, col, value)
    ws.append(["Project Summary", None, "Demo", "P", None, None, None, 0, datetime(2026, 2, 23), datetime(2027, 5, 31)])
    ws.append(["Activity", "1.1", "Mobilization", "P", "A1000", 1, 1, 1, datetime(2026, 2, 23), datetime(2026, 4, 23)])
    ws.append(["Activity", "1.2", "Foundation", "P", "A1010", 2, 2, 1, datetime(2026, 4, 1), datetime(2026, 8, 1)])
    ws.append(["Activity", "1.3", "Structure", "P", "A1020", 3, 3, 1, datetime(2026, 7, 1), datetime(2027, 2, 1)])
    wb.save(path)
    wb.close()
    return path


def test_progress_validation_calculates_default_periods_from_project_dates(tmp_path: Path) -> None:
    source = _progress_workbook(tmp_path / "progress.xlsx")
    result = PaymentWorkbookSnapshotter().validate(source)
    assert result.project_start.isoformat() == "2026-02-23"
    assert result.project_finish.isoformat() == "2027-05-31"
    assert result.default_payment_periods == 15


def test_generate_fake_payment_input_is_single_sheet_and_lightweight(tmp_path: Path) -> None:
    source = _progress_workbook(tmp_path / "progress.xlsx")
    output = tmp_path / "payment_input.xlsx"
    result = PaymentInputWorkbook().create(source, output, 15)

    assert result.payment_periods == 15
    assert result.activity_rows == 3

    wb = load_workbook(output, data_only=False)
    try:
        assert wb.sheetnames == ["Payment Input"]
        ws = wb["Payment Input"]
        assert ws["A6"].value == "Activity ID"
        assert ws["B6"].value == "Activity Name"
        assert ws["C6"].value == "P01"
        assert ws["Q6"].value == "P15"
        assert ws["A8"].value == "A1000"
        assert ws["B8"].value == "Mobilization"
        assert ws["A10"].value == "A1020"
        assert ws["Q7"].value is not None
        assert ws["C8"].value is None  # fixture has no weekly plan payload
        assert ws["C8"].number_format == "0%"
        assert ws.auto_filter.ref is None
    finally:
        wb.close()


def test_validate_payment_input_matches_progress_activity_ids(tmp_path: Path) -> None:
    source = _progress_workbook(tmp_path / "progress.xlsx")
    payment = tmp_path / "payment_input.xlsx"
    PaymentInputWorkbook().create(source, payment, 15)

    result = PaymentInputWorkbook().validate(payment, source)
    assert result.payment_periods == 15
    assert result.activity_rows == 3
    assert result.matched_activities == 3
    assert result.missing_activities == 0


def test_validate_payment_input_reports_missing_activity(tmp_path: Path) -> None:
    source = _progress_workbook(tmp_path / "progress.xlsx")
    payment = tmp_path / "payment_input.xlsx"
    PaymentInputWorkbook().create(source, payment, 15)

    wb = load_workbook(payment)
    ws = wb["Payment Input"]
    ws["A9"] = "A9999"
    wb.save(payment)
    wb.close()

    result = PaymentInputWorkbook().validate(payment, source)
    assert result.matched_activities == 2
    assert result.missing_activities == 1


def test_fake_payment_uses_activity_name_and_sparse_plan_suggestions(tmp_path: Path) -> None:
    source = tmp_path / "progress_with_plan.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "main"
    headers = ["Row Type", "WBS", "Description", "P/A", "Activity ID", "Task ID", "UID", "Outline Level", "Plan Start", "Plan Finish",
               "Actual Start", "Actual Finish", "% Complete", "Physical %", "Amount", "Total Float (hr)", "XML Amount"]
    for col, value in enumerate(headers, start=1):
        ws.cell(4, col, value)
    for offset, d in enumerate((datetime(2026, 3, 6), datetime(2026, 3, 13), datetime(2026, 3, 20), datetime(2026, 3, 27)), start=18):
        ws.cell(4, offset, d)
    ws.cell(5, 1, "Project Summary"); ws.cell(5, 3, "Demo"); ws.cell(5, 4, "P")
    ws.cell(5, 9, datetime(2026, 3, 1)); ws.cell(5, 10, datetime(2026, 3, 31))
    ws.cell(6, 1, "Activity"); ws.cell(6, 3, "Mobilization"); ws.cell(6, 4, "P"); ws.cell(6, 5, "A1000")
    ws.cell(6, 9, datetime(2026, 3, 1)); ws.cell(6, 10, datetime(2026, 3, 31))
    for col, value in zip(range(18, 22), (0.25, 0.25, 0.25, 0.25)):
        ws.cell(6, col, value)
    wb.save(source); wb.close()

    output = tmp_path / "payment.xlsx"
    PaymentInputWorkbook().create(source, output, 1)
    wb = load_workbook(output)
    try:
        ws = wb["Payment Input"]
        assert ws["B8"].value == "Mobilization"
        assert ws["C8"].value == 1.0
        assert ws["C8"].number_format == "0%"
        assert ws.auto_filter.ref is None
    finally:
        wb.close()
