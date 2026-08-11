
from pathlib import Path

from openpyxl import Workbook

from progress_studio.config.workbook_protection import WORKBOOK_SHEET_PASSWORD
from progress_studio.infrastructure.excel.workbook_guide import build_workbook_guide
from progress_studio.infrastructure.excel.workbook_visibility import apply_final_sheet_visibility


def test_internal_password_is_okmd() -> None:
    assert WORKBOOK_SHEET_PASSWORD == "okmd"


def test_readme_is_first_visible_sheet() -> None:
    wb = Workbook()
    wb.active.title = "main"
    wb.create_sheet("Dashboard")
    build_workbook_guide(wb)
    apply_final_sheet_visibility(wb)
    assert wb.sheetnames[0] == "README"
    assert wb["README"].sheet_state == "visible"
    assert wb["README"]["B3"].value.startswith("Normal progress update:")


def test_monthly_writer_uses_main_timescale_color_contract() -> None:
    source = Path("progress_studio/infrastructure/excel/live_monthly_workbook.py").read_text(encoding="utf-8")
    assert "add_progress_conditional_formatting" in source
    assert "clear_timescale_direct_fills" in source
    assert '"F7FBFF"' not in source
    assert '"EEF5FA"' not in source
