
from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook

from progress_studio.infrastructure.excel.rebuild_workbook_reader import RebuildWorkbookReader
from progress_studio.services.monthly_cache_deriver import MonthlyArchitectureEvaluator, MonthlyCacheDeriver


def _fixture(path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "main"
    headers = ["Row Type", "WBS", "Description", "P/A", "Activity ID", "Outline Level", "Amount"]
    for col, value in enumerate(headers, 1):
        ws.cell(4, col, value)
    dates = [
        datetime(2026,1,30),
        datetime(2026,2,6),
        datetime(2026,2,13),
        datetime(2026,2,27),
    ]
    for col, dt in enumerate(dates, start=8):
        ws.cell(3, col, f"W{col-7}")
        ws.cell(4, col, dt)
    ws.append(["Activity","1.1","Activity 1","P","A1000",2,1000,.10,.20,.30,.40])
    ws.append(["Activity","1.1","Activity 1","A","A1000",2,0,.05,.10,.15,.20])
    ws.append(["S-Curve","","Acc. Plan","AP","",0,None,.10,.30,.60,1.00])
    wb.save(path)
    return path


def test_lw6_monthly_cache_groups_weekly_values_by_month(tmp_path: Path) -> None:
    dataset = RebuildWorkbookReader().read_main_dataset(_fixture(tmp_path / "p.xlsx"))
    cache = MonthlyCacheDeriver().derive(dataset)
    assert [p.key for p in cache.periods] == ["M1", "M2"]
    assert cache.periods[0].reporting_date.date().isoformat() == "2026-01-30"
    assert cache.periods[1].reporting_date.date().isoformat() == "2026-02-27"
    assert cache.rows[0].values == (0.10, 0.90)
    assert cache.rows[1].values == (0.05, 0.45)


def test_lw6_cumulative_rows_take_month_end_value(tmp_path: Path) -> None:
    dataset = RebuildWorkbookReader().read_main_dataset(_fixture(tmp_path / "p.xlsx"))
    cache = MonthlyCacheDeriver().derive(dataset)
    scurve = cache.rows[2]
    assert scurve.values == (0.10, 1.00)


def test_lw6_decision_gate_selects_cache_not_formula(tmp_path: Path) -> None:
    dataset = RebuildWorkbookReader().read_main_dataset(_fixture(tmp_path / "p.xlsx"))
    decision = MonthlyArchitectureEvaluator().evaluate(dataset)
    assert decision.winner == "cache"
    assert decision.formula_cells == decision.cache_value_cells
    assert decision.direct_render_cells < decision.cache_value_cells
    assert any("dependency" in reason.lower() for reason in decision.rationale)


def test_lw6_monthly_deriver_has_no_openpyxl_dependency() -> None:
    source = Path("progress_studio/services/monthly_cache_deriver.py").read_text(encoding="utf-8")
    assert "openpyxl" not in source
    assert "load_workbook" not in source
