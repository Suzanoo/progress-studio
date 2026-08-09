from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from progress_studio.infrastructure.excel.payment_input_reader import PaymentInputSparseReader
from progress_studio.infrastructure.excel.payment_input_workbook import PaymentInputWorkbook
from progress_studio.services.payment_service import PaymentService
from tests.test_ms_pay6_payment_line_renderer import _progress_workbook


def test_embedded_payment_input_has_no_payment_date_row(tmp_path: Path) -> None:
    source = _progress_workbook(tmp_path / "progress.xlsx")
    wb = load_workbook(source)
    stats = PaymentInputWorkbook().embed(wb, periods=3)
    wb.save(source)
    wb.close()

    assert stats["periods"] == 3
    assert stats["activities"] == 2

    wb = load_workbook(source)
    try:
        ws = wb["Payment Input"]
        assert ws["A6"].value == "Type"
        assert ws["E6"].value == "P01"
        assert ws["A7"].value is None
        assert ws["E7"].value is None
        assert "Eligible dates are calculated" in ws["A5"].value
    finally:
        wb.close()


def test_embedded_payment_input_reconcile_preserves_user_percentages(tmp_path: Path) -> None:
    source = _progress_workbook(tmp_path / "progress.xlsx")
    wb = load_workbook(source)
    PaymentInputWorkbook().embed(wb, periods=3)
    wb.save(source)
    wb.close()

    wb = load_workbook(source)
    ws = wb["Payment Input"]
    row = next(r for r in range(8, ws.max_row + 1) if ws.cell(r, 3).value == "A1000")
    ws.cell(row, 5).value = 0.37
    wb.save(source)
    wb.close()

    preserved = PaymentInputSparseReader().read(source)
    wb = load_workbook(source)
    stats = PaymentInputWorkbook().embed(wb, preserved=preserved)
    wb.save(source)
    wb.close()

    assert stats["preserved"] == 2
    assert stats["new"] == 0
    wb = load_workbook(source)
    try:
        ws = wb["Payment Input"]
        row = next(r for r in range(8, ws.max_row + 1) if ws.cell(r, 3).value == "A1000")
        assert ws.cell(row, 5).value == 0.37
    finally:
        wb.close()


def test_one_workbook_rebuild_replaces_generated_views(tmp_path: Path) -> None:
    source = _progress_workbook(tmp_path / "progress.xlsx")
    wb = load_workbook(source)
    ws = wb["main"]
    for offset in range(5):
        ws.cell(3, 18 + offset, f"W{offset + 1}")
    wb.save(source)
    wb.close()

    # First pass embeds Payment Input + generates Payment in one workbook.
    output = tmp_path / "rebuilt.xlsx"
    result = PaymentService().rebuild_embedded_workbook(source, output, periods=3)
    assert result is not None
    assert result.rendered_periods >= 1

    wb = load_workbook(output)
    try:
        assert "Payment Input" in wb.sheetnames
        assert "Payment" in wb.sheetnames
        assert "progress_table" in wb.sheetnames
        assert wb["progress_table"].sheet_state == "veryHidden"
        assert wb.sheetnames.index("Payment") == wb.sheetnames.index("Payment Input") + 1
    finally:
        wb.close()
