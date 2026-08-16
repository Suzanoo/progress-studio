from __future__ import annotations

import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import Workbook, load_workbook
from openpyxl.utils.protection import hash_password

from progress_studio.config.workbook_protection import WORKBOOK_SHEET_PASSWORD
from progress_studio.infrastructure.excel.final_workbook_policy import finalize_workbook
from progress_studio.services.rebuild_service import WorkbookRebuildEngine
from tests.test_lw9_live_payment_integration import _fixture as _live_payment_fixture
from tests.test_ms_rb2_progress_rebuild_engine import _full_rebuild_fixture
from tests.test_ms_rb4_payment_only_rebuild import _embedded_payment_workbook


NS = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}


def _minimal_final_workbook(path: Path) -> Path:
    wb = Workbook()
    main = wb.active
    main.title = "main"
    headers = [
        "Row Type", "WBS", "Description", "P/A", "% Complete", "Amount",
        "Activity ID", "Outline Level", "Plan Start", "Plan Finish",
        "Actual Start", "Actual Finish", "Physical %", "XML Amount", "2026-01-02",
    ]
    for col, value in enumerate(headers, 1):
        main.cell(4, col, value)
    main.append(["Activity", "1", "Demo", "P", 0, 100, "A1000", 1,
                 "2026-01-01", "2026-01-02", None, None, 0, 100, "=1+1"])
    main.append(["Activity", "", "", "A", 0, None, "A1000", 1,
                 None, None, None, None, 0, None, 0])

    monthly = wb.create_sheet("main_monthly")
    monthly["L2"] = "Cutoff Date"
    monthly["M2"] = "2026-01-31"
    dashboard = wb.create_sheet("Dashboard")
    dashboard["G5"] = "Weekly"
    dashboard["K5"] = "2026-01-02"
    dashboard["P37"] = "All"
    wb.create_sheet("Dashboard_Data")

    finalize_workbook(wb, mode="snapshot", include_guide=True)
    wb.save(path)
    wb.close()
    return path


def test_final_policy_persists_structure_and_sheet_protection_to_ooxml(tmp_path: Path) -> None:
    path = _minimal_final_workbook(tmp_path / "final.xlsx")

    wb = load_workbook(path, data_only=False)
    try:
        assert wb.security.lockStructure is True
        assert wb.security.lockWindows is False
        assert wb.security.workbookPassword == hash_password(WORKBOOK_SHEET_PASSWORD)
        for ws in wb.worksheets:
            assert ws.protection.sheet is True
            assert ws.protection.password == hash_password(WORKBOOK_SHEET_PASSWORD)

        # User controls remain editable while formulas stay protected.
        assert wb["Dashboard"]["G5"].protection.locked is False
        assert wb["Dashboard"]["K5"].protection.locked is False
        assert wb["Dashboard"]["P37"].protection.locked is False
        assert wb["main_monthly"]["M2"].protection.locked is False
        assert wb["main"]["O5"].protection.locked is True
    finally:
        wb.close()

    with zipfile.ZipFile(path) as package:
        root = ET.fromstring(package.read("xl/workbook.xml"))
        protection = root.find("x:workbookProtection", NS)
        assert protection is not None
        assert protection.attrib["lockStructure"] == "1"
        assert protection.attrib["lockWindows"] == "0"
        assert protection.attrib["workbookPassword"] == hash_password(WORKBOOK_SHEET_PASSWORD)


def test_final_policy_persists_f9_save_formula_contract_to_ooxml(tmp_path: Path) -> None:
    path = _minimal_final_workbook(tmp_path / "calc.xlsx")
    wb = load_workbook(path, data_only=False)
    try:
        calc = wb.calculation
        assert calc.calcMode == "manual"
        assert calc.calcOnSave is True
        assert calc.fullCalcOnLoad is False
        assert calc.forceFullCalc is False
        assert calc.calcId != 0
    finally:
        wb.close()

    with zipfile.ZipFile(path) as package:
        root = ET.fromstring(package.read("xl/workbook.xml"))
        calc = root.find("x:calcPr", NS)
        assert calc is not None
        assert calc.attrib["calcMode"] == "manual"
        assert calc.attrib["calcOnSave"] == "1"
        assert calc.attrib["fullCalcOnLoad"] == "0"
        assert calc.attrib["forceFullCalc"] == "0"
        assert int(calc.attrib["calcId"]) > 0


def test_rebuild_can_rebuild_an_already_structure_protected_workbook(tmp_path: Path) -> None:
    source = _full_rebuild_fixture(tmp_path / "source.xlsx")
    protected = tmp_path / "protected.xlsx"
    rebuilt = tmp_path / "rebuilt.xlsx"

    WorkbookRebuildEngine().rebuild_progress(source, protected)
    first = load_workbook(protected)
    try:
        assert first.security.lockStructure is True
    finally:
        first.close()

    # OOXML protection is an Excel UI guard; Progress Studio/openpyxl still owns rebuild.
    WorkbookRebuildEngine().rebuild_progress(protected, rebuilt)
    second = load_workbook(rebuilt)
    try:
        assert second.security.lockStructure is True
        assert second["main"].protection.sheet is True
        assert second["Dashboard"].protection.sheet is True
    finally:
        second.close()


def test_snapshot_payment_rebuild_finalizes_once_and_protects_new_payment(tmp_path: Path) -> None:
    source = _embedded_payment_workbook(tmp_path / "source.xlsx")
    output = tmp_path / "payment.xlsx"

    WorkbookRebuildEngine().rebuild_payment(source, output)

    wb = load_workbook(output, data_only=False)
    try:
        assert wb.security.lockStructure is True
        assert wb["Payment"].protection.sheet is True
        assert wb.calculation.calcMode == "manual"
        assert wb.calculation.calcOnSave is True
    finally:
        wb.close()

    source_text = Path("progress_studio/services/rebuild_service.py").read_text(encoding="utf-8")
    start = source_text.index("    def rebuild_payment(")
    end = source_text.index("    def generated_sheets_for(", start)
    block = source_text[start:end]
    assert "finalize_workbook(" not in block
    assert "render_payment_backbones" in block


def test_live_payment_finalizes_after_render_so_new_payment_is_protected(tmp_path: Path) -> None:
    source = _live_payment_fixture(tmp_path / "source.xlsx")
    output = tmp_path / "live_payment.xlsx"

    WorkbookRebuildEngine().rebuild_live_payment(source, output)

    wb = load_workbook(output, data_only=False)
    try:
        assert wb["Payment"].protection.sheet is True
        assert wb.security.lockStructure is True
        assert wb.calculation.calcMode == "manual"
        assert wb.calculation.calcOnSave is True
    finally:
        wb.close()

    source_text = Path("progress_studio/services/rebuild_service.py").read_text(encoding="utf-8")
    start = source_text.index("    def rebuild_live_payment(")
    end = source_text.index("    def rebuild_payment(", start)
    block = source_text[start:end]
    assert "finalize_workbook(" not in block
    assert 'finalize_mode="live"' in block


def test_readme_states_exact_f9_save_vs_rebuild_ownership(tmp_path: Path) -> None:
    path = _minimal_final_workbook(tmp_path / "guide.xlsx")
    wb = load_workbook(path, data_only=False)
    try:
        readme = wb["README"]
        text = " ".join(str(cell.value or "") for row in readme.iter_rows() for cell in row)
        assert "F9 / Save" in text
        assert "recalculate Excel formulas" in text
        assert "Generated snapshots/caches are not rebuilt by Excel" in text
        assert "Rebuild" in text
    finally:
        wb.close()
