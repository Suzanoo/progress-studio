from datetime import date, datetime

from openpyxl import Workbook
from openpyxl.styles import PatternFill

from progress_studio.app.desktop import build_desktop_pipeline
from progress_studio.infrastructure.excel.monthly_main_workbook import build_monthly_main_view


def _workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = "main"
    headers = [
        "Row Type", "WBS", "Description", "P/A", "Activity ID", "Task ID", "UID",
        "Outline Level", "Plan Start", "Plan Finish", "Actual Start", "Actual Finish",
        "% Complete", "Physical %", "Amount", "Total Float (hr)", "XML Amount",
    ]
    for col, value in enumerate(headers, start=1):
        ws.cell(4, col).value = value
    dates = [
        datetime(2026, 1, 30),
        datetime(2026, 2, 6),
        datetime(2026, 2, 13),
        datetime(2026, 2, 27),
    ]
    for index, value in enumerate(dates, start=18):
        ws.cell(4, index).value = value
        ws.cell(3, index).value = f"W{index - 17}"
    rows = [
        ["Project Summary", "", "Project", "P", "", "", "", 0],
        ["", "", "", "A", "", "", "", 0],
        ["WBS", "1", "WBS 1", "P", "", "", "", 1],
        ["", "", "", "A", "", "", "", 1],
        ["Activity", "1.1", "Activity 1", "P", "A1000", "", "", 2],
        ["", "", "", "A", "A1000", "", "", 2],
        ["S-Curve", "", "Plan", "P", "", "", "", 0],
        ["S-Curve", "", "Acc. Plan", "AP", "", "", "", 0],
    ]
    for r, values in enumerate(rows, start=5):
        for c, value in enumerate(values, start=1):
            ws.cell(r, c).value = value
        ws.row_dimensions[r].outlineLevel = int(values[7])
    for r in range(5, 13):
        for c in range(18, 22):
            ws.cell(r, c).number_format = "0.00%"
            ws.cell(r, c).fill = PatternFill("solid", fgColor="FFFFFF")
    ws.cell(9, 18).value = 0.10
    ws.cell(9, 19).value = 0.20
    ws.cell(9, 20).value = 0.30
    ws.cell(9, 21).value = 0.40
    ws.cell(12, 18).value = 0.10
    ws.cell(12, 19).value = 0.30
    ws.cell(12, 20).value = 0.60
    ws.cell(12, 21).value = 1.00
    return wb


def test_monthly_main_freezes_static_activity_plan_and_keeps_live_summary_rows() -> None:
    wb = _workbook()
    count = build_monthly_main_view(wb)
    ws = wb["main_monthly"]

    assert count == 2
    assert wb.sheetnames[:2] == ["main", "main_monthly"]
    assert ws.cell(1, 1).value == "Activity Data — Monthly View"
    assert ws.cell(4, 18).value == date(2026, 1, 30)
    assert ws.cell(4, 19).value == date(2026, 2, 27)
    assert ws.cell(9, 18).value == 0.10
    assert ws.cell(9, 19).value == 0.90
    # Project/WBS/Actual rows stay live to weekly edits.
    assert isinstance(ws.cell(5, 18).value, str) and ws.cell(5, 18).value.startswith('=')
    assert isinstance(ws.cell(10, 18).value, str) and ws.cell(10, 18).value.startswith('=')


def test_monthly_cumulative_scurve_uses_last_nonblank_reporting_value() -> None:
    wb = _workbook()
    build_monthly_main_view(wb)
    ws = wb["main_monthly"]
    assert ws.cell(12, 18).value == '=IFERROR(LOOKUP(2,1/(\'main\'!R12:R12<>""),\'main\'!R12:R12),"")'
    assert ws.cell(12, 19).value == '=IFERROR(LOOKUP(2,1/(\'main\'!S12:U12<>""),\'main\'!S12:U12),"")'


def test_monthly_acc_plan_keeps_final_100_when_later_weeks_in_same_month_are_blank() -> None:
    wb = _workbook()
    ws = wb["main"]
    # February's real Plan finishes on 13-Feb; the later 27-Feb weekly column is
    # display margin.  Acc.Plan must therefore retain 100% for the February
    # monthly bucket instead of inheriting the trailing blank.
    ws.cell(12, 19).value = 0.60
    ws.cell(12, 20).value = 1.00
    ws.cell(12, 21).value = ""

    build_monthly_main_view(wb)
    monthly = wb["main_monthly"]

    assert monthly.cell(12, 19).value == '=IFERROR(LOOKUP(2,1/(\'main\'!S12:U12<>""),\'main\'!S12:U12),"")'


def test_monthly_snapshot_acc_plan_uses_last_numeric_value_before_margin_blank() -> None:
    wb = _workbook()
    value_source = _workbook()["main"]
    value_source.cell(12, 19).value = 0.60
    value_source.cell(12, 20).value = 1.00
    value_source.cell(12, 21).value = None

    build_monthly_main_view(wb, snapshot=True, value_source=value_source)
    monthly = wb["main_monthly"]

    assert monthly.cell(12, 19).value == 1.00


def test_monthly_view_preserves_outline_and_only_rowtype_pa_filter_buttons() -> None:
    wb = _workbook()
    build_monthly_main_view(wb)
    ws = wb["main_monthly"]

    assert ws.row_dimensions[7].outlineLevel == 1
    assert ws.row_dimensions[9].outlineLevel == 2
    assert ws.freeze_panes == "R5"
    hidden = {item.colId for item in ws.auto_filter.filterColumn if item.showButton is False}
    assert 0 not in hidden  # Row Type
    assert 3 not in hidden  # P/A
    assert 1 in hidden
    assert len(ws.data_validations.dataValidation) == 0


def test_desktop_xml_pipeline_has_monthly_as_real_eighth_step() -> None:
    pipeline = build_desktop_pipeline("flat")
    assert len(pipeline.steps) == 8
    assert pipeline.steps[-1].name == "build-monthly-main-view"
