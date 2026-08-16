from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from progress_studio.infrastructure.excel.final_workbook_policy import (
    FinalWorkbookMode,
    finalize_workbook,
)


def _policy_fixture() -> Workbook:
    wb = Workbook()
    main = wb.active
    main.title = "main"

    # Minimal main contract with one Plan/Actual Activity pair and one weekly cell.
    headers = [
        "Row Type", "WBS", "Description", "P/A", "% Complete", "Amount",
        "Activity ID", "Outline Level", "Plan Start", "Plan Finish",
        "Actual Start", "Actual Finish", "Physical %", "XML Amount",
        "2026-01-02",
    ]
    for col, value in enumerate(headers, start=1):
        main.cell(4, col, value)
    main.append(["Activity", "1", "Demo", "P", 0, 100, "A1000", 1, "2026-01-01", "2026-01-02", None, None, 0, 100, 1])
    main.append(["Activity", "", "", "A", 0, None, "A1000", 1, None, None, None, None, 0, None, 0])

    monthly = wb.create_sheet("main_monthly")
    monthly["L2"] = "Cutoff Date"
    monthly["M2"] = "2026-01-31"

    dashboard = wb.create_sheet("Dashboard")
    dashboard["G5"] = "Weekly"
    dashboard["K5"] = "2026-01-02"
    dashboard["P37"] = "All"

    for name in ("progress", "progress_table", "Dashboard_Data", "Info"):
        wb.create_sheet(name)
    return wb


def test_wp2_snapshot_policy_is_one_finalization_contract() -> None:
    wb = _policy_fixture()

    result = finalize_workbook(wb, mode=FinalWorkbookMode.SNAPSHOT)

    assert result.mode is FinalWorkbookMode.SNAPSHOT
    assert wb.sheetnames[0] == "README"
    assert wb["README"].sheet_state == "visible"
    assert wb["main"].sheet_state == "visible"
    assert wb["main_monthly"].sheet_state == "visible"
    assert wb["Dashboard"].sheet_state == "visible"
    assert wb["progress"].sheet_state == "hidden"
    assert wb["progress_table"].sheet_state == "hidden"
    assert wb["Dashboard_Data"].sheet_state == "hidden"
    assert wb["Info"].sheet_state == "veryHidden"

    for ws in wb.worksheets:
        assert ws.protection.sheet is True

    # Intended controls stay editable after the shared protection pass.
    assert wb["Dashboard"]["G5"].protection.locked is False
    assert wb["Dashboard"]["K5"].protection.locked is False
    assert wb["Dashboard"]["P37"].protection.locked is False
    assert wb["main_monthly"]["M2"].protection.locked is False

    # Snapshot workbooks preserve the current incremental/automatic policy.
    assert wb.calculation.calcMode == "auto"
    assert wb.calculation.fullCalcOnLoad is False
    assert wb.calculation.forceFullCalc is False
    assert wb.calculation.calcOnSave is True


def test_wp2_live_policy_preserves_live_recalculation_mode() -> None:
    wb = _policy_fixture()

    result = finalize_workbook(wb, mode="live")

    assert result.mode is FinalWorkbookMode.LIVE
    assert wb.calculation.calcMode == "manual"
    assert wb.calculation.fullCalcOnLoad is False
    assert wb.calculation.forceFullCalc is False
    assert wb.calculation.calcOnSave is True


def test_wp2_finalizer_is_wired_into_all_final_workbook_pipelines() -> None:
    expected = {
        Path("progress_studio/services/workbook_generation_service.py"),
        Path("progress_studio/infrastructure/excel/mapped_workbook_exporter.py"),
        Path("progress_studio/services/rebuild_service.py"),
        Path("progress_studio/services/payment_service.py"),
    }
    for path in expected:
        source = path.read_text(encoding="utf-8")
        assert "finalize_workbook" in source, path

    create_source = Path("progress_studio/services/workbook_generation_service.py").read_text(encoding="utf-8")
    assert 'finalize_workbook(final_wb, mode="snapshot", include_guide=True)' in create_source
