from pathlib import Path

from openpyxl import Workbook

from progress_studio.infrastructure.excel.traditional_overlay_workbook import _overlay_chart


def test_overlay_hotfix_keeps_original_source_columns_and_no_helper_architecture():
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard_Data"
    for r in range(2, 5):
        ws.cell(r, 1, r)
        ws.cell(r, 2, 0.1 * (r - 1))
        ws.cell(r, 16, 0.05 * (r - 1))
        ws.cell(r, 18, 1 if r == 3 else "=NA()")

    chart = _overlay_chart(
        data_ws=ws,
        date_col=1,
        plan_col=2,
        actual_col=16,
        cutoff_col=18,
        first_row=2,
        last_row=4,
        cutoff_label_format="dd/mm/yyyy",
    )

    assert len(chart.series) == 3
    assert chart.series[0].val.numRef.f.endswith("$B$2:$B$4")
    assert chart.series[1].val.numRef.f.endswith("$P$2:$P$4")
    assert chart.series[2].val.numRef.f.endswith("$R$2:$R$4")
    assert max(ws.max_column, 18) == 18


def test_monthly_cutoff_label_is_month_year_display_only():
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard_Data"
    for r in range(2, 4):
        ws.cell(r, 4, r)
        ws.cell(r, 5, 0.2 * (r - 1))
        ws.cell(r, 17, 0.1 * (r - 1))
        ws.cell(r, 19, 1 if r == 3 else "=NA()")

    chart = _overlay_chart(
        data_ws=ws,
        date_col=4,
        plan_col=5,
        actual_col=17,
        cutoff_col=19,
        first_row=2,
        last_row=3,
        cutoff_label_format="mmmm yyyy",
    )

    assert chart.series[2].dLbls.numFmt == "mmmm yyyy"
    assert chart.series[2].cat is not None
