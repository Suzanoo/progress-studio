from __future__ import annotations

from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook

from progress_studio.infrastructure.excel.payment_input_workbook import PaymentInputWorkbook
from progress_studio.services.payment_service import PaymentService


def _progress_workbook(path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "main"
    headers = [
        "Row Type", "WBS", "Description", "P/A", "Activity ID", "Task ID", "UID", "Outline Level",
        "Plan Start", "Plan Finish", "Actual Start", "Actual Finish", "% Complete", "Physical %",
        "Amount", "Total Float (hr)", "XML Amount",
    ]
    for col, value in enumerate(headers, start=1):
        ws.cell(4, col, value)
    week0 = datetime(2026, 3, 2)
    for offset in range(5):
        ws.cell(4, 18 + offset, week0 + timedelta(days=7 * offset))

    ws.cell(5, 1, "Project Summary")
    ws.cell(5, 3, "Demo")
    ws.cell(5, 4, "P")
    ws.cell(5, 9, datetime(2026, 3, 2))
    ws.cell(5, 10, datetime(2026, 4, 3))

    ws.cell(6, 1, "Activity")
    ws.cell(6, 2, "1.1")
    ws.cell(6, 3, "Foundation")
    ws.cell(6, 4, "P")
    ws.cell(6, 5, "A1000")
    ws.cell(6, 9, datetime(2026, 3, 2))
    ws.cell(6, 10, datetime(2026, 4, 3))
    for col, value in zip(range(18, 23), [0.10, 0.15, 0.20, 0.20, 0.35]):
        ws.cell(6, col, value)

    # Keep a spacer/WBS-like row between points so the vertical line must cross rows.
    ws.cell(7, 1, "WBS")
    ws.cell(7, 2, "2")
    ws.cell(7, 3, "Structure")
    ws.cell(7, 4, "P")

    ws.cell(8, 1, "Activity")
    ws.cell(8, 2, "2.1")
    ws.cell(8, 3, "Structure L1")
    ws.cell(8, 4, "P")
    ws.cell(8, 5, "A1010")
    ws.cell(8, 9, datetime(2026, 3, 16))
    ws.cell(8, 10, datetime(2026, 3, 30))
    for col, value in zip(range(20, 23), [0.25, 0.25, 0.50]):
        ws.cell(8, col, value)

    wb.save(path)
    wb.close()
    return path


def _payment_input(progress: Path, path: Path) -> Path:
    PaymentInputWorkbook().create(progress, path, 3)
    wb = load_workbook(path)
    ws = wb["Payment Input"]
    # P01 begins at column E in the lightweight tree layout. Clear generated
    # suggestions first, then make P01 intentionally sparse with two points.
    for row in range(8, ws.max_row + 1):
        if ws.cell(row, 1).value == "ACT":
            for col in range(5, 8):
                ws.cell(row, col).value = None
    activity_rows = {
        ws.cell(row, 3).value: row
        for row in range(8, ws.max_row + 1)
        if ws.cell(row, 1).value == "ACT"
    }
    ws.cell(activity_rows["A1000"], 5, 0.60)
    ws.cell(activity_rows["A1010"], 5, 0.25)
    wb.save(path)
    wb.close()
    return path


def test_ms_pay6_single_period_uses_vertical_backbone_without_shapes(tmp_path: Path) -> None:
    progress = _progress_workbook(tmp_path / "progress.xlsx")
    payment = _payment_input(progress, tmp_path / "payment_input.xlsx")
    output = tmp_path / "payment_p01.xlsx"

    result = PaymentService().render_single_payment_line(progress, payment, output, "P01")

    assert result.period_id == "P01"
    assert result.rendered_points == 2
    assert result.color == "C00000"

    wb = load_workbook(output)
    try:
        assert "Payment" in wb.sheetnames
        main = wb["main"]
        sheet = wb["Payment"]
        assert len(sheet._images) == 1  # one lightweight label badge for P01
        assert not sheet._charts

        # Backbone is no longer driven by the input Payment Date. P01 becomes
        # eligible at the latest resolved requirement: A1000 at right edge of U,
        # i.e. boundary before V. The backbone crosses the spacer/WBS row.
        assert sheet["V6"].border.left.style == "thick"  # target cap overlaps backbone on controlling row
        assert sheet["V7"].border.left.style == "medium"
        assert sheet["V8"].border.left.style == "medium"
        assert sheet["V7"].border.left.color.rgb.endswith("C00000")

        # Header note reports the calculated eligible date and keeps input date
        # only as reference metadata.
        assert sheet["V4"].comment is not None
        assert "Planned Eligible Date: 2026-03-23" in sheet["V4"].comment.text
        assert "Controlling Activity: A1000" in sheet["V4"].comment.text
        assert "reference only" in sheet["V4"].comment.text

        # A1010 target is one bucket earlier (right edge of T => before U),
        # so its branch runs back from the eligible backbone to U.
        assert sheet["U8"].border.bottom.style == "medium"
        assert sheet["U8"].border.left.style == "thick"

        # Source of truth is untouched.
        assert main["V7"].border.left.style != "medium"
        assert main["U8"].border.bottom.style != "medium"
    finally:
        wb.close()


def test_ms_pay61_renders_three_colored_vertical_backbones(tmp_path: Path) -> None:
    progress = _progress_workbook(tmp_path / "progress.xlsx")
    payment = _payment_input(progress, tmp_path / "payment_input.xlsx")

    wb = load_workbook(payment)
    ws = wb["Payment Input"]
    activity_rows = {
        ws.cell(row, 3).value: row
        for row in range(8, ws.max_row + 1)
        if ws.cell(row, 1).value == "ACT"
    }
    # Tree layout: E=P01, F=P02, G=P03. Keep all three periods sparse.
    ws.cell(activity_rows["A1000"], 5, 0.25)
    ws.cell(activity_rows["A1000"], 6, 0.60)
    ws.cell(activity_rows["A1000"], 7, 1.00)
    ws.cell(activity_rows["A1010"], 5).value = None
    ws.cell(activity_rows["A1010"], 6, 0.25)
    ws.cell(activity_rows["A1010"], 7, 1.00)
    wb.save(payment)
    wb.close()

    output = tmp_path / "payment_three.xlsx"
    result = PaymentService().render_payment_backbones(
        progress, payment, output, ("P01", "P02", "P03")
    )

    assert result.period_ids == ("P01", "P02", "P03")
    assert result.rendered_periods == 3
    assert result.rendered_points == 5
    assert dict(result.colors) == {"P01": "C00000", "P02": "0070C0", "P03": "548235"}

    wb = load_workbook(output)
    try:
        sheet = wb["Payment"]
        assert len(sheet._images) == 3  # one lightweight label badge per Payment
        assert not sheet._charts
        # Eligible boundaries come from the latest requirement point, not
        # from the dates stored in the Payment Input row:
        # P01=S-right => T, P02=U-right => V, P03=V-right => W.
        assert sheet["T6"].border.left.color.rgb.endswith("C00000")
        assert sheet["V6"].border.left.color.rgb.endswith("0070C0")
        assert sheet["V6"].border.right.style in {"medium", "thick"}
        assert sheet["V6"].border.right.color.rgb.endswith("548235")
        assert "Planned Eligible Date: 2026-03-09" in sheet["T4"].comment.text
        assert "Planned Eligible Date: 2026-03-23" in sheet["V4"].comment.text
        assert "Planned Eligible Date: 2026-03-30" in sheet["V4"].comment.text
    finally:
        wb.close()



def test_ms_pay63_eligible_date_ignores_input_payment_date(tmp_path: Path) -> None:
    progress = _progress_workbook(tmp_path / "progress.xlsx")
    payment = _payment_input(progress, tmp_path / "payment_input.xlsx")
    prepared = PaymentService().prepare_payment_input(progress, payment)
    p01 = next(period for period in prepared.positions.periods if period.period_id == "P01")

    assert p01.payment_date.isoformat() == "2026-03-13"  # legacy/reference input
    assert p01.planned_eligible_date.isoformat() == "2026-03-23"
    assert p01.controlling_activity_ids == ("A1000",)


def test_ms_pay64_adds_only_one_label_drawing_per_payment(tmp_path: Path) -> None:
    progress = _progress_workbook(tmp_path / "progress.xlsx")
    payment = _payment_input(progress, tmp_path / "payment_input.xlsx")
    output = tmp_path / "payment_labeled.xlsx"

    PaymentService().render_single_payment_line(progress, payment, output, "P01")

    wb = load_workbook(output)
    try:
        sheet = wb["Payment"]
        assert len(sheet._images) == 1
        image = sheet._images[0]
        assert image.width == 116
        assert image.height == 22
        assert not sheet._charts
    finally:
        wb.close()


def test_ms_pay65_default_render_includes_every_populated_payment(tmp_path: Path) -> None:
    progress = _progress_workbook(tmp_path / "progress.xlsx")
    payment = _payment_input(progress, tmp_path / "payment_input.xlsx")

    wb = load_workbook(payment)
    ws = wb["Payment Input"]
    activity_rows = {
        ws.cell(row, 3).value: row
        for row in range(8, ws.max_row + 1)
        if ws.cell(row, 1).value == "ACT"
    }
    # Populate all three available periods in this fixture.
    ws.cell(activity_rows["A1000"], 5, 0.25)
    ws.cell(activity_rows["A1000"], 6, 0.60)
    ws.cell(activity_rows["A1000"], 7, 1.00)
    ws.cell(activity_rows["A1010"], 5, 0.25)
    ws.cell(activity_rows["A1010"], 6, 0.50)
    ws.cell(activity_rows["A1010"], 7, 1.00)
    wb.save(payment)
    wb.close()

    output = tmp_path / "payment_all.xlsx"
    result = PaymentService().render_payment_backbones(progress, payment, output)

    assert result.period_ids == ("P01", "P02", "P03")
    assert result.rendered_periods == 3
    wb = load_workbook(output)
    try:
        assert len(wb["Payment"]._images) == 3
    finally:
        wb.close()


def test_ms_pay65_payment_theme_config_controls_color_and_label_size(tmp_path: Path) -> None:
    import json

    from progress_studio.config.payment_theme import load_payment_line_theme
    from progress_studio.infrastructure.excel.payment_line_renderer import PaymentLineRenderer

    config = tmp_path / "payment_lines.json"
    config.write_text(json.dumps({
        "line": {"style": "thin", "endpoint_style": "medium", "fallback_color": "999999"},
        "label": {
            "width_px": 90,
            "height_px": 18,
            "font_size": 9,
            "corner_radius_px": 3,
            "text_color": "FFFFFF",
            "anchor_column_offset": -1,
            "anchor_row": 1
        },
        "colors": {"P01": "112233"}
    }), encoding="utf-8")

    progress = _progress_workbook(tmp_path / "progress.xlsx")
    payment = _payment_input(progress, tmp_path / "payment_input.xlsx")
    output = tmp_path / "payment_custom_theme.xlsx"

    service = PaymentService(
        line_renderer=PaymentLineRenderer(load_payment_line_theme(config))
    )
    result = service.render_single_payment_line(progress, payment, output, "P01")
    assert result.color == "112233"

    wb = load_workbook(output)
    try:
        sheet = wb["Payment"]
        assert len(sheet._images) == 1
        assert sheet._images[0].width == 90
        assert sheet._images[0].height == 18
    finally:
        wb.close()
