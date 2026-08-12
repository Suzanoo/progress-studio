from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

from progress_studio.infrastructure.excel.live_dashboard_workbook import build_live_dashboard
from progress_studio.infrastructure.excel.live_scurve_workbook import build_live_progress_contract
from progress_studio.infrastructure.excel.rebuild_workbook_reader import RebuildWorkbookReader


def _fixture(path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "main"
    headers = [
        "Row Type", "WBS", "Description", "P/A", "Activity ID",
        "Outline Level", "Plan Start", "Plan Finish", "Amount",
    ]
    for col, value in enumerate(headers, 1):
        ws.cell(4, col, value)
    dates = [
        datetime(2026, 7, 10), datetime(2026, 7, 17),
        datetime(2026, 7, 24), datetime(2026, 7, 31),
        datetime(2026, 8, 7),
    ]
    for col, dt in enumerate(dates, start=10):
        ws.cell(3, col, f"W{col-9}")
        ws.cell(4, col, dt)

    ws.append(["Project Summary","","Project","P","",0,datetime(2026,7,1),datetime(2026,8,31),1000,.20,.30,.20,.20,.10])
    ws.append(["Project Summary","","Project","A","",0,None,None,0,.10,.05,.05,None,None])
    ws.append(["S-Curve","","Plan","P","",0,None,None,None,.20,.30,.20,.20,.10])
    ws.append(["S-Curve","","Acc. Plan","AP","",0,None,None,None,.20,.50,.70,.90,1.00])
    ws.append(["S-Curve","","Actual","A","",0,None,None,None,.10,.05,.05,None,None])
    ws.append(["S-Curve","","Acc. Actual","AA","",0,None,None,None,.10,.15,.20,None,None])
    wb.save(path)
    return path


def test_lw11_progress_keeps_raw_actual_history(tmp_path: Path) -> None:
    path = _fixture(tmp_path / "p.xlsx")
    dataset = RebuildWorkbookReader().read_main_dataset(path)
    wb = load_workbook(path)
    wb.create_sheet("Dashboard")["K5"] = datetime(2026, 7, 24)

    count = build_live_progress_contract(wb, dataset)
    assert count == 5
    assert wb["progress"]["A1"].value == "Date"
    assert "Dashboard!$K$5" not in wb["main"]["J10"].value
    assert "Dashboard!$K$5" not in wb["progress"]["C2"].value
    assert "Dashboard!$K$5" not in wb["progress"]["C6"].value
    assert "Dashboard!$K$5" not in wb["progress"]["B6"].value
    wb.close()


def test_lw11_dashboard_data_is_progress_renderer(tmp_path: Path) -> None:
    path = _fixture(tmp_path / "p.xlsx")
    dataset = RebuildWorkbookReader().read_main_dataset(path)
    wb = load_workbook(path)
    build_live_dashboard(wb, dataset, cutoff=datetime(2026, 7, 24))

    data = wb["Dashboard_Data"]
    assert data["B2"].value == "='progress'!B2"
    assert data["C2"].value == "='progress'!C2"
    # July monthly point is the final July progress row (31-Jul, progress row 5).
    assert data["E2"].value == "='progress'!B5"
    assert data["F2"].value == "='progress'!C5"
    # Dashboard selector owns the tiny chart-only Actual cutoff mask.
    assert "Dashboard!$K$5" in data["I2"].value
    assert "NA()" in data["I2"].value
    # Empty monthly slots must stay blank instead of becoming Excel serial date 0.
    assert 'IF(D6="","",D6)' in data["G6"].value
    assert data.sheet_state == "hidden"
    wb.close()


def test_lw1133_kpis_use_error_free_raw_actual_and_cutoff_markers(tmp_path: Path) -> None:
    path = _fixture(tmp_path / "p.xlsx")
    dataset = RebuildWorkbookReader().read_main_dataset(path)
    wb = load_workbook(path)
    build_live_dashboard(wb, dataset, cutoff=datetime(2026, 7, 24))

    data = wb["Dashboard_Data"]
    dashboard = wb["Dashboard"]
    assert data["L1"].value == "Selected Actual Raw"
    assert "NA()" not in data["L2"].value
    assert data["M1"].value == "Marker Date"
    assert data["N1"].value == "Cutoff Plan Marker"
    assert data["O1"].value == "Cutoff Actual Marker"
    assert data["M2"].value == "=Dashboard!$K$5"
    assert "Dashboard!$K$5" in data["N2"].value
    assert "Dashboard!$K$5" in data["O2"].value
    # Marker sources are physically one row only: Excel cannot create labels
    # for the rest of the weekly/monthly range.
    assert data["N3"].value is None
    assert data["O3"].value is None

    assert "SUMIFS" in dashboard["B10"].value
    assert "Dashboard_Data!$H$2" in dashboard["B10"].value
    assert "SUMIFS" in dashboard["E10"].value
    assert "Dashboard_Data!$L$2" in dashboard["E10"].value
    assert "#N/A" not in dashboard["E10"].value

    chart = dashboard._charts[0]
    assert len(chart.series) == 4
    assert chart.y_axis.majorUnit == 0.25
    assert chart.y_axis.title is None
    assert chart.x_axis.title is None
    assert chart.series[2].marker.symbol == "circle"
    assert chart.series[3].marker.symbol == "circle"
    assert chart.series[2].cat.numRef.f.endswith("$M$2:$M$2")
    assert chart.series[3].cat.numRef.f.endswith("$M$2:$M$2")
    assert chart.series[2].val.numRef.f.endswith("$N$2")
    assert chart.series[3].val.numRef.f.endswith("$O$2")
    assert chart.series[2].dLbls.showVal is True
    assert chart.series[3].dLbls.showVal is True
    assert chart.series[2].dLbls.position == "t"
    assert chart.series[3].dLbls.position == "b"
    assert [(entry.idx, entry.delete) for entry in chart.legend.legendEntry] == [(2, True), (3, True)]

    # LW-11.3.4: schedule card keeps both status and signed progress gap.
    assert "CHAR(10)" in dashboard["H10"].value
    assert "+0.00%;-0.00%;0.00%" in dashboard["H10"].value
    assert dashboard["H10"].alignment.wrap_text is True
    # Coincident Plan/Actual cutoff markers collapse to one visible marker.
    assert "ABS(SUMIFS" in data["O2"].value
    # KPI icons are restored in the live dashboard path as well.
    assert len(dashboard._images) == 4
    wb.close()
