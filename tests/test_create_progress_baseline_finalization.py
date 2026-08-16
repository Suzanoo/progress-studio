from pathlib import Path

from openpyxl import load_workbook

from progress_studio.services.monthly_main_service import MonthlyMainService
from progress_studio.services.rebuild_service import WorkbookRebuildEngine
from tests.test_ms_rb2_progress_rebuild_engine import _full_rebuild_fixture


def _ready_create_input(tmp_path: Path) -> Path:
    """Use the proven rebuild builders only to make a realistic pre-finalization fixture."""
    source = _full_rebuild_fixture(tmp_path / "source.xlsx")
    ready = tmp_path / "ready.xlsx"
    WorkbookRebuildEngine().rebuild_progress(source, ready, project_name="Create Boundary")
    return ready


def test_create_progress_final_step_moves_proven_rebuild_features_to_create_boundary(tmp_path: Path) -> None:
    source = _ready_create_input(tmp_path)
    output = tmp_path / "created.xlsx"

    MonthlyMainService().build(source, output)

    wb = load_workbook(output, data_only=False)
    try:
        # Proven pre-normalizer Snapshot recalc contract.
        assert wb.calculation.calcMode == "auto"
        assert wb.calculation.calcOnSave is True
        assert wb.calculation.fullCalcOnLoad is False
        assert wb.calculation.forceFullCalc is False

        # Proven lightweight protection: sheets protected, structure left manageable.
        assert wb.security.lockStructure is False
        assert wb["main"].protection.sheet is True
        assert wb["main_monthly"].protection.sheet is True
        assert wb["Dashboard"].protection.sheet is True

        # Proven visibility contract is present on the first Create output.
        assert wb["progress"].sheet_state == "hidden"
        assert wb["progress_table"].sheet_state == "hidden"
        assert wb["Dashboard_Data"].sheet_state == "hidden"

        # Traditional cutoff overlays are present immediately; no Rebuild required.
        assert len(wb["main"]._charts) >= 1
        assert len(wb["main_monthly"]._charts) >= 1
        for sheet_name in ("main", "main_monthly"):
            ws = wb[sheet_name]
            cutoff_row = next(
                row for row in range(1, ws.max_row + 1)
                if ws.cell(row, 12).value == "Cutoff Date"
            )
            assert ws.cell(cutoff_row, 13).protection.locked is False
    finally:
        wb.close()


def test_create_final_step_uses_one_openpyxl_load_and_one_save(monkeypatch, tmp_path: Path) -> None:
    source = _ready_create_input(tmp_path)
    output = tmp_path / "created.xlsx"

    import progress_studio.services.monthly_main_service as module

    real_load = module.load_workbook
    calls = {"load": 0, "save": 0}

    def counted_load(*args, **kwargs):
        calls["load"] += 1
        wb = real_load(*args, **kwargs)
        real_save = wb.save

        def counted_save(*save_args, **save_kwargs):
            calls["save"] += 1
            return real_save(*save_args, **save_kwargs)

        wb.save = counted_save
        return wb

    monkeypatch.setattr(module, "load_workbook", counted_load)
    MonthlyMainService().build(source, output)

    assert calls == {"load": 1, "save": 1}
