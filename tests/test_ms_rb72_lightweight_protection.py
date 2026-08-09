from __future__ import annotations

from openpyxl import load_workbook

from progress_studio.config.workbook_protection import WORKBOOK_SHEET_PASSWORD
from progress_studio.services.payment_service import PaymentService
from progress_studio.services.rebuild_service import WorkbookRebuildEngine
from tests.test_ms_rb2_progress_rebuild_engine import _full_rebuild_fixture
from tests.test_ms_pay6_payment_line_renderer import _progress_workbook


def test_rb72_progress_rebuild_applies_lightweight_sheet_protection(tmp_path) -> None:
    source = _full_rebuild_fixture(tmp_path / "source.xlsx")
    output = tmp_path / "rebuilt.xlsx"

    WorkbookRebuildEngine().rebuild_progress(source, output)

    wb = load_workbook(output, data_only=False)
    try:
        assert wb.security.lockStructure is False
        assert wb.security.lockWindows is False

        for ws in wb.worksheets:
            assert ws.protection.sheet is True
            assert ws.protection.password

        main = wb["main"]
        # Headers stay locked.
        assert main["A4"].protection.locked is True

        # Activity Plan identity/schedule/value fields remain editable.
        assert main["B11"].protection.locked is False  # WBS
        assert main["C11"].protection.locked is False  # Description
        assert main["F11"].protection.locked is False  # Amount
        assert main["G11"].protection.locked is False  # Activity ID
        assert main["I11"].protection.locked is False  # Plan Start
        assert main["J11"].protection.locked is False  # Plan Finish

        # Activity weekly Plan/Actual values are editable.
        assert main["L11"].protection.locked is False
        assert main["L12"].protection.locked is False

        # WBS/project roll-up areas stay locked.
        assert main["L9"].protection.locked is True
        assert main["L5"].protection.locked is True

        # Row operations remain available on main.
        assert main.protection.insertRows is False
        assert main.protection.deleteRows is False

        # Generated/public data sheets are protected.
        for name in ("main_monthly", "progress", "progress_table", "Dashboard_Data", "Dashboard"):
            assert wb[name].protection.sheet is True
    finally:
        wb.close()


def test_rb72_payment_input_unlocks_only_activity_percentages(tmp_path) -> None:
    source = _progress_workbook(tmp_path / "source.xlsx")
    output = tmp_path / "payment_input.xlsx"

    PaymentService().prepare_embedded_payment_input(source, output, periods=3)

    wb = load_workbook(output, data_only=False)
    try:
        ws = wb["Payment Input"]
        assert ws.protection.sheet is True

        activity_row = next(
            row for row in range(8, ws.max_row + 1)
            if ws.cell(row, 1).value == "ACT"
        )
        wbs_row = next(
            row for row in range(8, ws.max_row + 1)
            if ws.cell(row, 1).value == "WBS"
        )

        # Tree identity is protected.
        for col in range(1, 5):
            assert ws.cell(activity_row, col).protection.locked is True

        # Payment percentages are editable for ACT rows only.
        for col in range(5, 8):
            assert ws.cell(activity_row, col).protection.locked is False
            assert ws.cell(wbs_row, col).protection.locked is True
    finally:
        wb.close()


def test_rb72_visibility_keeps_public_data_easy_to_unhide(tmp_path) -> None:
    source = _full_rebuild_fixture(tmp_path / "source.xlsx")
    output = tmp_path / "rebuilt.xlsx"

    WorkbookRebuildEngine().rebuild_progress(source, output)

    wb = load_workbook(output)
    try:
        assert wb["progress"].sheet_state == "hidden"
        assert wb["progress_table"].sheet_state == "hidden"
        assert wb["Dashboard_Data"].sheet_state == "veryHidden"
    finally:
        wb.close()


def test_rb72_password_config_is_nonempty_internal_guard() -> None:
    assert isinstance(WORKBOOK_SHEET_PASSWORD, str)
    assert WORKBOOK_SHEET_PASSWORD
