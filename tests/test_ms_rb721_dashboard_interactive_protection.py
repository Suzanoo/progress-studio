from __future__ import annotations

from openpyxl import load_workbook

from progress_studio.services.rebuild_service import WorkbookRebuildEngine
from tests.test_ms_rb2_progress_rebuild_engine import _full_rebuild_fixture


def test_rb721_dashboard_controls_remain_editable_when_sheet_is_protected(tmp_path) -> None:
    source = _full_rebuild_fixture(tmp_path / "source.xlsx")
    output = tmp_path / "rebuilt.xlsx"

    WorkbookRebuildEngine().rebuild_progress(source, output)

    wb = load_workbook(output, data_only=False)
    try:
        ws = wb["Dashboard"]

        assert ws.protection.sheet is True

        # Interactive dashboard controls.
        assert ws["G5"].protection.locked is False
        assert ws["K5"].protection.locked is False
        assert ws["P37"].protection.locked is False

        # Labels, formulas and KPI cells remain protected.
        assert ws["F5"].protection.locked is True
        assert ws["J5"].protection.locked is True
        assert ws["B2"].protection.locked is True
        assert ws["C10"].protection.locked is True
    finally:
        wb.close()
