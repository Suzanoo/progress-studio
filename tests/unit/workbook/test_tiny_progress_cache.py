
from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook

from progress_studio.infrastructure.excel.rebuild_workbook_reader import RebuildWorkbookReader
from progress_studio.services.progress_cache_deriver import ProgressCacheDeriver


def _fixture(path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "main"
    headers = [
        "Row Type", "WBS", "Description", "P/A", "% Complete",
        "Amount", "Activity ID", "Outline Level", "Plan Start", "Plan Finish",
    ]
    for col, value in enumerate(headers, 1):
        ws.cell(4, col, value)

    for col, dt in enumerate(
        (datetime(2026,3,6), datetime(2026,3,13), datetime(2026,3,20), datetime(2026,3,27)),
        start=11,
    ):
        ws.cell(3, col, f"W{col-10}")
        ws.cell(4, col, dt)

    # A1000 = 75% of project amount; A2000 = 25%.
    rows = [
        ["Activity", "1.1", "Concrete", "P", "", 750, "A1000", 2, None, None, .4, .4, .2, None],
        ["Activity", "1.1", "Concrete", "A", "", 0, "A1000", 2, None, None, .2, None, .3, None],
        ["Activity", "2.1", "MEP", "P", "", 250, "A2000", 2, None, None, .2, .4, .4, None],
        ["Activity", "2.1", "MEP", "A", "", 0, "A2000", 2, None, None, None, .2, .2, None],
    ]
    for row in rows:
        ws.append(row)
    wb.save(path)
    return path


def test_lw4_builds_one_cache_point_per_reporting_period(tmp_path: Path) -> None:
    data = RebuildWorkbookReader().read_main_dataset(_fixture(tmp_path / "p.xlsx"))
    cache = ProgressCacheDeriver().derive(data)

    assert cache.total_amount == 1000.0
    assert cache.period_count == 4
    assert [p.period_key for p in cache.points] == ["W1", "W2", "W3", "W4"]


def test_lw4_uses_full_plan_amount_weighting(tmp_path: Path) -> None:
    data = RebuildWorkbookReader().read_main_dataset(_fixture(tmp_path / "p.xlsx"))
    cache = ProgressCacheDeriver().derive(data)

    # W1 plan = .75*.4 + .25*.2 = .35
    # W2 plan = .75*.4 + .25*.4 = .40 => cumulative .75
    assert cache.points[0].plan_weekly == 0.35
    assert cache.points[0].plan_cumulative == 0.35
    assert cache.points[1].plan_weekly == 0.4
    assert cache.points[1].plan_cumulative == 0.75


def test_lw4_actual_cumulative_plateaus_inside_range_and_stops_after_latest_actual(tmp_path: Path) -> None:
    data = RebuildWorkbookReader().read_main_dataset(_fixture(tmp_path / "p.xlsx"))
    cache = ProgressCacheDeriver().derive(data)

    # W1 actual = .75*.2 = .15
    assert cache.points[0].actual_weekly == 0.15
    assert cache.points[0].actual_cumulative == 0.15
    # W2 actual = .25*.2 = .05
    assert cache.points[1].actual_cumulative == 0.2
    # W3 actual = .75*.3 + .25*.2 = .275 -> cumulative .475
    assert abs(cache.points[2].actual_cumulative - 0.475) < 1e-12
    # Preserve legacy graph semantics: no Actual line beyond the last reported period.
    assert cache.points[3].actual_weekly is None
    assert cache.points[3].actual_cumulative is None


def test_lw4_progress_cache_has_no_openpyxl_or_progress_table_dependency() -> None:
    source = Path("progress_studio/services/progress_cache_deriver.py").read_text(encoding="utf-8")
    assert "openpyxl" not in source
    assert "load_workbook" not in source
    assert 'workbook["progress_table"]' not in source
    assert "TABLE_SHEET" not in source
