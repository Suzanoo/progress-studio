
from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook

from progress_studio.infrastructure.excel.rebuild_workbook_reader import RebuildWorkbookReader


def _fixture(path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "main"
    headers = [
        "Row Type", "WBS", "Description", "P/A", "% Complete",
        "Amount", "Activity ID", "Outline Level", "Plan Start", "Plan Finish",
    ]
    for col, value in enumerate(headers, 1):
        ws.cell(4, col, value)
    for col, date in enumerate((datetime(2026, 3, 6), datetime(2026, 3, 13)), start=11):
        ws.cell(3, col, f"W{col - 10}")
        ws.cell(4, col, date)
    rows = [
        ["Project Summary", "", "Project", "P", "", 1000, "", 0, datetime(2026,3,1), datetime(2026,3,31), 0.5, 0.5],
        ["Activity", "1.1", "Concrete", "P", 1.0, 1000, "A1000", 2, datetime(2026,3,1), datetime(2026,3,13), 0.5, 0.5],
        ["Activity", "1.1", "Concrete", "A", 0.5, 0, "A1000", 2, datetime(2026,3,1), datetime(2026,3,13), 0.2, 0.3],
    ]
    for row in rows:
        ws.append(row)
    wb.save(path)
    return path


def test_lw2_parses_main_once_into_immutable_dataset(tmp_path: Path) -> None:
    path = _fixture(tmp_path / "project.xlsx")
    data = RebuildWorkbookReader().read_main_dataset(path)

    assert data.workbook_name == "project.xlsx"
    assert data.header_row == 4
    assert len(data.periods) == 2
    assert data.periods[0].key == "W1"
    assert data.periods[0].reporting_date.date().isoformat() == "2026-03-06"
    assert len(data.activities) == 1

    activity = data.activities[0]
    assert activity.activity_id == "A1000"
    assert activity.wbs == "1.1"
    assert activity.description == "Concrete"
    assert activity.amount == 1000.0
    assert activity.outline_level == 2
    assert activity.plan_start.date().isoformat() == "2026-03-01"
    assert activity.period_value(data.periods[0].column) == 0.5


def test_lw2_reader_path_remains_openpyxl_free() -> None:
    source = Path(
        "progress_studio/infrastructure/excel/rebuild_workbook_reader.py"
    ).read_text(encoding="utf-8")
    assert "load_workbook" not in source
    assert "import openpyxl" not in source
    assert "from openpyxl" not in source


def test_lw2_dataset_keeps_plan_and_actual_rows(tmp_path: Path) -> None:
    path = _fixture(tmp_path / "project.xlsx")
    data = RebuildWorkbookReader().read_main_dataset(path)

    activity_rows = [r for r in data.rows if r.activity_id == "A1000"]
    assert [r.pa for r in activity_rows] == ["P", "A"]
    assert activity_rows[1].period_values[1][1] == 0.3
