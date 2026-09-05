from datetime import datetime

import pytest
from openpyxl import Workbook

from progress_studio.domain.main_dataset import MainPeriod
from progress_studio.infrastructure.excel.payment_breakdown_workbook import (
    PAYMENT_BREAKDOWN_SHEET,
    render_payment_breakdown,
)
from progress_studio.services.payment_breakdown_adapter import (
    PaymentBreakdownDatasetSnapshot,
)
from progress_studio.services.payment_breakdown_service import (
    PaymentBreakdownService,
    PaymentBreakdownSourceActivity,
)


def _derived():
    return PaymentBreakdownService().derive_activity(
        (
            PaymentBreakdownSourceActivity(
                "A1",
                "First Fixed",
                100.0,
                (0.0, 0.5, 0.5, 0.0),
                "3.2",
            ),
            PaymentBreakdownSourceActivity(
                "A2",
                "First Fixed",
                300.0,
                (0.0, 0.0, 1.0, 0.0),
                "3.3",
            ),
        )
    )


def _snapshot():
    periods = (
        MainPeriod(6, "W1", datetime(2026, 1, 2)),
        MainPeriod(7, "W2", datetime(2026, 1, 9)),
        MainPeriod(8, "W3", datetime(2026, 1, 16)),
        MainPeriod(9, "W4", datetime(2026, 1, 23)),
    )
    return PaymentBreakdownDatasetSnapshot(
        periods=periods,
        activities=(_derived(),),
        eligible_source_count=2,
        skipped_activity_ids=(),
    )


def _rgb(cell):
    color = cell.font.color
    return None if color is None else color.rgb


def test_pb41_renderer_replaces_only_target_sheet_and_keeps_block_structure():
    wb = Workbook()
    main = wb.active
    main.title = "main"
    old = wb.create_sheet(PAYMENT_BREAKDOWN_SHEET)
    old["A1"] = "old"

    ws = render_payment_breakdown(wb, _snapshot())
    assert wb.sheetnames == ["main", PAYMENT_BREAKDOWN_SHEET]
    assert ws["A1"].value == "Payment Breakdown"
    assert ws["A4"].value == "Activity Name"
    assert ws["A5"].value == "First Fixed"
    assert ws["E5"].value == "Activity Progress"
    assert ws["E6"].value == "Activity Cumulative"
    assert ws["E9"].value == "Combined Progress"
    assert ws["E10"].value == "Combined Cumulative"


def test_pb41_progress_rows_hide_zero_but_keep_nonzero_period_values():
    wb = Workbook()
    wb.active.title = "main"
    ws = render_payment_breakdown(wb, _snapshot())

    # A1 period progress: 0%, 50%, 50%, 0%
    assert ws["F5"].value is None
    assert ws["G5"].value == pytest.approx(0.5)
    assert ws["H5"].value == pytest.approx(0.5)
    assert ws["I5"].value is None
    # A2 has a genuine one-period 100% progress value; keep it visible.
    assert ws["F7"].value is None
    assert ws["G7"].value is None
    assert ws["H7"].value == pytest.approx(1.0)
    assert ws["I7"].value is None
    assert not (_rgb(ws["H7"]) or "").endswith("FF0000")


def test_pb41_cumulative_rows_show_first_100_only():
    wb = Workbook()
    wb.active.title = "main"
    ws = render_payment_breakdown(wb, _snapshot())
    # A1 cumulative: 0%, 50%, 100%, 100%
    assert ws["F6"].value is None
    assert ws["G6"].value == pytest.approx(0.5)
    assert ws["H6"].value == pytest.approx(1.0)
    assert ws["I6"].value is None

    # A2 cumulative: 0%, 0%, 100%, 100%
    assert ws["F8"].value is None
    assert ws["G8"].value is None
    assert ws["H8"].value == pytest.approx(1.0)
    assert ws["I8"].value is None


def test_pb41_red_text_is_only_for_materially_in_progress_values():
    wb = Workbook()
    wb.active.title = "main"
    ws = render_payment_breakdown(wb, _snapshot())

    assert (_rgb(ws["G5"]) or "").endswith("FF0000")
    assert not (_rgb(ws["H6"]) or "").endswith("FF0000")
    # Combined cumulative: blank, 12.5%, 100%, blank.
    assert ws["F10"].value is None
    assert ws["G10"].value == pytest.approx(0.125)
    assert (_rgb(ws["G10"]) or "").endswith("FF0000")
    assert ws["H10"].value == pytest.approx(1.0)
    assert not (_rgb(ws["H10"]) or "").endswith("FF0000")
    assert ws["I10"].value is None


def test_pb41_renderer_keeps_amount_weighted_calculation_unchanged():
    wb = Workbook()
    wb.active.title = "main"
    ws = render_payment_breakdown(wb, _snapshot())
    assert ws["D9"].value == pytest.approx(400.0)
    assert ws["F9"].value is None
    assert ws["G9"].value == pytest.approx(0.125)
    assert ws["H9"].value == pytest.approx(0.875)
    assert ws["I9"].value is None

    # Rendering is compact, but the derived engine still reaches and carries 100%.
    derived = _derived()
    assert derived.cumulative_progress == pytest.approx(
        (0.0, 0.125, 1.0, 1.0)
    )


def test_pb42_source_activity_rows_are_grouped_and_expanded_by_default():
    wb = Workbook()
    wb.active.title = "main"
    ws = render_payment_breakdown(wb, _snapshot())

    # Source Activity detail is collapsible.
    for row in range(5, 9):
        assert ws.row_dimensions[row].outlineLevel == 1
        assert ws.row_dimensions[row].hidden is False

    # Header and Combined summary rows remain outside the detail outline.
    assert ws.row_dimensions[4].outlineLevel == 0
    assert ws.row_dimensions[9].outlineLevel == 0
    assert ws.row_dimensions[10].outlineLevel == 0

    # Excel should place the outline summary below the detail rows.
    assert ws.sheet_properties.outlinePr.summaryBelow is True
