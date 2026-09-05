from datetime import datetime

from openpyxl import Workbook, load_workbook

from progress_studio.services.payment_service import PaymentService


def _source_workbook(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "main"

    ws["A3"] = "X"
    ws["F3"] = "W1"
    ws["G3"] = "W2"
    ws["H3"] = "W3"

    headers = [
        "Row Type",
        "WBS",
        "Description",
        "P/A",
        "Activity ID",
        "W1",
        "W2",
        "W3",
        "Outline Level",
        "Plan Start",
        "Plan Finish",
        "Amount",
        "% Complete",
    ]
    for col, value in enumerate(headers, start=1):
        ws.cell(4, col, value)

    ws["F4"] = datetime(2026, 1, 2)
    ws["G4"] = datetime(2026, 1, 9)
    ws["H4"] = datetime(2026, 1, 16)

    rows = [
        (
            "Activity",
            "3.2.5.1",
            "First Fixed",
            "P",
            "A1",
            0.50,
            0.50,
            0.00,
            4,
            datetime(2026, 1, 1),
            datetime(2026, 1, 16),
            100.0,
            0.0,
        ),
        (
            "Activity",
            "3.3.5.1",
            "First Fixed",
            "P",
            "A2",
            0.00,
            0.50,
            0.50,
            4,
            datetime(2026, 1, 1),
            datetime(2026, 1, 16),
            300.0,
            0.0,
        ),
        (
            "Activity",
            "3.2.5.2",
            "Second Fixed",
            "P",
            "B1",
            0.25,
            0.50,
            0.25,
            4,
            datetime(2026, 1, 1),
            datetime(2026, 1, 16),
            200.0,
            0.0,
        ),
        (
            "Activity",
            "3.3.5.2",
            "Second Fixed",
            "P",
            "B2",
            0.00,
            0.25,
            0.75,
            4,
            datetime(2026, 1, 1),
            datetime(2026, 1, 16),
            200.0,
            0.0,
        ),
    ]
    for row_index, values in enumerate(rows, start=5):
        for col, value in enumerate(values, start=1):
            ws.cell(row_index, col, value)

    wb.save(path)
    wb.close()


def test_pb4_payment_workspace_builds_visible_breakdown_without_touching_source(
    tmp_path,
):
    source = tmp_path / "source.xlsx"
    output = tmp_path / "output.xlsx"
    _source_workbook(source)

    snapshot = PaymentService().prepare_payment_breakdown(
        source,
        output,
    )

    assert [item.activity_name for item in snapshot.activities] == [
        "First Fixed",
        "Second Fixed",
    ]
    assert snapshot.eligible_source_count == 4

    source_wb = load_workbook(source)
    try:
        assert "Payment-Breakdown" not in source_wb.sheetnames
    finally:
        source_wb.close()

    output_wb = load_workbook(output)
    try:
        assert "Payment-Breakdown" in output_wb.sheetnames
        assert output_wb["Payment-Breakdown"].sheet_state == "visible"
        assert output_wb["Payment-Breakdown"]["A1"].value == "Payment Breakdown"
        assert "main" in output_wb.sheetnames
    finally:
        output_wb.close()
