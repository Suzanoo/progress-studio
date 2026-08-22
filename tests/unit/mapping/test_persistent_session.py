from __future__ import annotations

import json
from dataclasses import asdict
from pathlib import Path

import pytest

from progress_studio.domain.mapping_models import ActivityRow, AllocationRecord, BOQRow
from progress_studio.infrastructure.session import (
    MappingSessionRepository,
    RecentSessionRepository,
    SessionValidationError,
)
from progress_studio.services.mapping_store import MappingStore


def make_store() -> MappingStore:
    store = MappingStore()
    store.load_activities([
        ActivityRow("A1000", "1", "1.1", "Floor demolition"),
        ActivityRow("A2000", "1", "1.2", "Wall demolition"),
    ])
    store.load_boq([
        BOQRow("Project|10|10", "Project", 10, "Demo", "Floor", "", "Remove floor", 1000.0),
        BOQRow("Project|11|11", "Project", 11, "Demo", "Wall", "", "Remove wall", 500.0),
    ])
    return store


def write_workbook_placeholder(path: Path, content: bytes) -> Path:
    path.write_bytes(content)
    return path


def test_session_round_trip_and_workbook_validation(tmp_path: Path) -> None:
    progress = write_workbook_placeholder(tmp_path / "progress.xlsx", b"progress-v1")
    boq = write_workbook_placeholder(tmp_path / "boq.xlsx", b"boq-v1")
    target = tmp_path / "project.mapping.json"
    repository = MappingSessionRepository()

    session = repository.create(
        progress,
        boq,
        "Project",
        [AllocationRecord("Project|10|10", "A1000", 35.0)],
    )
    repository.save(target, session)
    loaded = repository.load(target)

    assert loaded.boq_sheet == "Project"
    assert loaded.allocations == (
        AllocationRecord("Project|10|10", "A1000", 35.0),
    )
    assert repository.validate_workbook(loaded.progress) == progress.resolve()
    assert repository.validate_workbook(loaded.boq) == boq.resolve()


def test_session_write_is_json_and_has_expected_contract(tmp_path: Path) -> None:
    progress = write_workbook_placeholder(tmp_path / "progress.xlsx", b"progress")
    boq = write_workbook_placeholder(tmp_path / "boq.xlsx", b"boq")
    target = tmp_path / "project.mapping.json"
    repository = MappingSessionRepository()
    repository.save(target, repository.create(progress, boq, "NKC2", []))

    payload = json.loads(target.read_text(encoding="utf-8"))
    assert payload["format"] == "progress-studio-mapping-session"
    assert payload["version"] == 8
    assert payload["boq_sheet"] == "NKC2"
    assert payload["allocations"] == []
    assert not list(tmp_path.glob("*.tmp"))


def test_changed_workbook_is_rejected(tmp_path: Path) -> None:
    progress = write_workbook_placeholder(tmp_path / "progress.xlsx", b"progress-v1")
    boq = write_workbook_placeholder(tmp_path / "boq.xlsx", b"boq-v1")
    repository = MappingSessionRepository()
    session = repository.create(progress, boq, "Project", [])

    progress.write_bytes(b"progress-v2")
    with pytest.raises(SessionValidationError, match="does not match"):
        repository.validate_workbook(session.progress)


def test_store_restores_allocations_and_clears_undo_history() -> None:
    store = make_store()
    change = store.restore_allocations([
        AllocationRecord("Project|10|10", "A1000", 40.0),
        AllocationRecord("Project|10|10", "A2000", 60.0),
        AllocationRecord("Project|11|11", "A2000", 100.0),
    ])

    assert set(change.boq_keys) == {"Project|10|10", "Project|11|11"}
    assert store.activity_amount("A1000") == 400.0
    assert store.activity_amount("A2000") == 1100.0
    assert store.undo() is None


def test_store_rejects_invalid_session_records() -> None:
    store = make_store()
    with pytest.raises(ValueError, match="Activity was not found"):
        store.restore_allocations([
            AllocationRecord("Project|10|10", "MISSING", 50.0),
        ])
    with pytest.raises(ValueError, match="exceeds 100%"):
        store.restore_allocations([
            AllocationRecord("Project|10|10", "A1000", 60.0),
            AllocationRecord("Project|10|10", "A2000", 60.0),
        ])


def test_clear_all_is_one_undoable_command() -> None:
    store = make_store()
    store.restore_allocations([
        AllocationRecord("Project|10|10", "A1000", 100.0),
        AllocationRecord("Project|11|11", "A2000", 100.0),
    ])

    change = store.clear_all()
    assert store.mapped_amount == 0.0
    assert set(change.activity_ids) == {"A1000", "A2000"}

    store.undo()
    assert store.mapped_amount == 1500.0


def test_recent_sessions_keeps_latest_first_and_ignores_missing(tmp_path: Path) -> None:
    recent_file = tmp_path / "recent.json"
    repository = RecentSessionRepository(recent_file)
    first = tmp_path / "first.mapping.json"
    second = tmp_path / "second.mapping.json"
    first.write_text("{}", encoding="utf-8")
    second.write_text("{}", encoding="utf-8")

    repository.remember(first)
    repository.remember(second)
    repository.remember(first)

    assert repository.list() == [first.resolve(), second.resolve()]
    first.unlink()
    assert repository.list() == [second.resolve()]


def test_v1_session_is_migrated_to_current_version(tmp_path: Path) -> None:
    progress = write_workbook_placeholder(tmp_path / "progress.xlsx", b"progress")
    boq = write_workbook_placeholder(tmp_path / "boq.xlsx", b"boq")
    repository = MappingSessionRepository()
    current = repository.create(progress, boq, "Project", [])
    payload = {
        "format": current.format,
        "version": 1,
        "saved_at": current.saved_at,
        "progress": {
            "path": current.progress.path,
            "size": current.progress.size,
            "modified_ns": current.progress.modified_ns,
            "sha256": current.progress.sha256,
        },
        "boq": {
            "path": current.boq.path,
            "size": current.boq.size,
            "modified_ns": current.boq.modified_ns,
            "sha256": current.boq.sha256,
        },
        "boq_sheet": "Project",
        "allocations": [],
    }
    target = tmp_path / "legacy.mapping.json"
    target.write_text(json.dumps(payload), encoding="utf-8")

    loaded = repository.load(target)

    assert loaded.version == 8
    assert loaded.progress.filename == "progress.xlsx"
    assert loaded.boq.filename == "boq.xlsx"


def test_moved_or_renamed_workbook_can_be_relinked_by_hash(tmp_path: Path) -> None:
    original = write_workbook_placeholder(tmp_path / "progress.xlsx", b"same-content")
    repository = MappingSessionRepository()
    saved = repository.create(original, original, "Project", []).progress
    moved = tmp_path / "archive" / "renamed-progress.xlsx"
    moved.parent.mkdir()
    original.replace(moved)

    with pytest.raises(SessionValidationError, match="Browse for the moved"):
        repository.validate_workbook(saved)
    assert repository.validate_workbook(saved, moved) == moved.resolve()


def test_relink_rejects_different_workbook_content(tmp_path: Path) -> None:
    original = write_workbook_placeholder(tmp_path / "progress.xlsx", b"original")
    replacement = write_workbook_placeholder(tmp_path / "replacement.xlsx", b"different")
    repository = MappingSessionRepository()
    saved = repository.create(original, original, "Project", []).progress

    with pytest.raises(SessionValidationError, match="does not match"):
        repository.validate_workbook(saved, replacement)


def test_future_session_version_is_rejected_with_clear_message(tmp_path: Path) -> None:
    target = tmp_path / "future.mapping.json"
    target.write_text(
        json.dumps({"format": "progress-studio-mapping-session", "version": 999}),
        encoding="utf-8",
    )
    with pytest.raises(SessionValidationError, match="supports up to version"):
        MappingSessionRepository().load(target)


def test_excel_resave_and_formatting_change_keeps_semantic_identity(tmp_path: Path) -> None:
    from openpyxl import Workbook, load_workbook

    progress = tmp_path / "progress.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "main"
    sheet.append(["Activity ID", "Activity Name", "Amount"])
    sheet.append(["A1000", "Foundation", 1000.0])
    workbook.save(progress)

    repository = MappingSessionRepository()
    saved = repository.create(progress, progress, "main", []).progress

    # Excel/openpyxl rewrites the package and formatting, but the worksheet data
    # used by the project remains identical.
    workbook = load_workbook(progress)
    workbook["main"].column_dimensions["B"].width = 42
    workbook["main"].freeze_panes = "A2"
    workbook.save(progress)

    assert repository.validate_workbook(saved) == progress.resolve()


def test_excel_cell_data_change_is_rejected_by_semantic_identity(tmp_path: Path) -> None:
    from openpyxl import Workbook, load_workbook

    progress = tmp_path / "progress.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "main"
    sheet.append(["Activity ID", "Activity Name"])
    sheet.append(["A1000", "Foundation"])
    workbook.save(progress)

    repository = MappingSessionRepository()
    saved = repository.create(progress, progress, "main", []).progress

    workbook = load_workbook(progress)
    workbook["main"]["B2"] = "Changed foundation"
    workbook.save(progress)

    with pytest.raises(SessionValidationError, match="does not match"):
        repository.validate_workbook(saved)


def test_excel_workbook_can_be_moved_renamed_and_resaved(tmp_path: Path) -> None:
    from openpyxl import Workbook, load_workbook

    original = tmp_path / "progress.xlsx"
    workbook = Workbook()
    workbook.active["A1"] = "Project Alpha"
    workbook.save(original)

    repository = MappingSessionRepository()
    saved = repository.create(original, original, "Sheet", []).progress
    moved = tmp_path / "archive" / "renamed-progress.xlsx"
    moved.parent.mkdir()
    original.replace(moved)

    workbook = load_workbook(moved)
    workbook.active.sheet_view.showGridLines = False
    workbook.save(moved)

    assert repository.validate_workbook(saved, moved) == moved.resolve()


def test_v6_session_migrates_with_legacy_strict_fingerprint(tmp_path: Path) -> None:
    progress = write_workbook_placeholder(tmp_path / "progress.xlsx", b"legacy-not-real-xlsx")
    repository = MappingSessionRepository()
    current = repository.create(progress, progress, "Project", [])
    payload = {
        "format": current.format,
        "version": 6,
        "saved_at": current.saved_at,
        "progress": {
            "path": current.progress.path,
            "filename": current.progress.filename,
            "size": current.progress.size,
            "modified_ns": current.progress.modified_ns,
            "sha256": current.progress.sha256,
        },
        "boq": {
            "path": current.boq.path,
            "filename": current.boq.filename,
            "size": current.boq.size,
            "modified_ns": current.boq.modified_ns,
            "sha256": current.boq.sha256,
        },
        "boq_sheet": "Project",
        "allocations": [],
        "supplemental_activities": [],
        "supplemental_wbs": [],
        "working_tree_nodes": [],
    }
    target = tmp_path / "legacy-v6.progressstudio"
    target.write_text(json.dumps(payload), encoding="utf-8")

    loaded = repository.load(target)

    assert loaded.version == 8
    assert loaded.progress.semantic_sha256 == ""
    assert repository.validate_workbook(loaded.progress) == progress.resolve()


def test_create_reuses_cached_workbook_fingerprints(tmp_path: Path, monkeypatch) -> None:
    progress = write_workbook_placeholder(tmp_path / "progress.xlsx", b"progress")
    boq = write_workbook_placeholder(tmp_path / "boq.xlsx", b"boq")
    repository = MappingSessionRepository()
    progress_identity = repository.fingerprint(progress)
    boq_identity = repository.fingerprint(boq)

    import progress_studio.infrastructure.session.mapping_session_repository as module

    def unexpected_fingerprint(_path: Path):
        raise AssertionError("cached autosave must not hash workbooks again")

    monkeypatch.setattr(module, "fingerprint", unexpected_fingerprint)
    session = repository.create(
        progress,
        boq,
        "Project",
        [],
        progress_fingerprint=progress_identity,
        boq_fingerprint=boq_identity,
    )

    assert session.progress == progress_identity
    assert session.boq == boq_identity


def test_v8_project_embeds_workbooks_and_restores_them_without_original_files(tmp_path: Path) -> None:
    progress = write_workbook_placeholder(tmp_path / "progress.xlsx", b"progress-source")
    boq = write_workbook_placeholder(tmp_path / "boq.xlsx", b"boq-source")
    repository = MappingSessionRepository()
    project = tmp_path / "standalone.progressstudio"
    repository.save(project, repository.create(progress, boq, "Project", []))

    loaded = repository.load(project)
    assert loaded.progress_snapshot is not None
    assert loaded.boq_snapshot is not None

    progress.unlink()
    boq.unlink()
    restored_progress = repository.materialize_snapshot(loaded.progress_snapshot)
    restored_boq = repository.materialize_snapshot(loaded.boq_snapshot)

    assert restored_progress.read_bytes() == b"progress-source"
    assert restored_boq.read_bytes() == b"boq-source"


def test_v7_project_migrates_but_requires_one_resave_before_standalone_rebuild(tmp_path: Path) -> None:
    progress = write_workbook_placeholder(tmp_path / "progress.xlsx", b"progress")
    boq = write_workbook_placeholder(tmp_path / "boq.xlsx", b"boq")
    repository = MappingSessionRepository()
    current = repository.create(progress, boq, "Project", [])
    payload = {
        "format": current.format,
        "version": 7,
        "saved_at": current.saved_at,
        "progress": asdict(current.progress),
        "boq": asdict(current.boq),
        "boq_sheet": "Project",
        "allocations": [],
        "supplemental_activities": [],
        "supplemental_wbs": [],
        "working_tree_nodes": [],
    }
    target = tmp_path / "legacy-v7.progressstudio"
    target.write_text(json.dumps(payload), encoding="utf-8")

    loaded = repository.load(target)

    assert loaded.version == 8
    assert loaded.progress_snapshot is None
    assert loaded.boq_snapshot is None
