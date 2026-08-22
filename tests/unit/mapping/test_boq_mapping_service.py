from pathlib import Path
from openpyxl import Workbook, load_workbook

from progress_studio.domain.mapping_models import AllocationRecord
from progress_studio.services.boq_mapping_service import BOQMappingService


def add_main_sheet(wb, activities):
    ws = wb.create_sheet("main")
    ws.append(["Progress Studio test workbook"])
    ws.append([])
    ws.append([])
    ws.append(["Row Type", "Activity ID", "Description", "P/A", "Outline Level", "Amount"])
    ws.append(["Project Summary", "", "Project", "P", 0, 0])
    for activity_id, description in activities:
        ws.append(["Activity", activity_id, description, "P", 1, 0])
        ws.append(["Activity", activity_id, description, "A", 1, None])



def test_mapping_read_and_export(tmp_path: Path):
    progress = tmp_path / "progress.xlsx"
    wb = Workbook(); ws = wb.active; ws.title = "Amount Mapping"
    ws.append(["Activity ID", "WBS", "Description", "Amount", "Status"])
    ws.append(["A1000", "1.1", "Foundation", 0, "UNMAPPED"])
    add_main_sheet(wb, [("A1000", "Foundation")])
    wb.save(progress); wb.close()

    boq = tmp_path / "boq.xlsx"
    wb = Workbook(); ws = wb.active; ws.title = "Project"
    ws.append(["WBS-1", "WBS-2", "WBS-3", "WBS-4", "Description", "Amount", "Source Sheet", "Source Row"])
    ws.append(["Project", "Structure", "Foundation", "", "Concrete", 1000, "Structure", 8])
    wb.save(boq); wb.close()

    service = BOQMappingService()
    activities = service.read_activities(progress)
    assert service.list_boq_sheets(boq) == ["Project"]
    rows = service.read_boq(boq, "Project")
    assert activities[0].activity_id == "A1000"
    assert rows[0].amount == 1000

    output = tmp_path / "mapped.xlsx"
    service.export(progress, output, rows, [AllocationRecord(rows[0].key, "A1000", 100.0)])
    wb = load_workbook(output, data_only=False)
    assert wb["Amount Mapping"]["D2"].value == 1000
    assert wb["Amount Mapping"]["E2"].value == "MAPPED"
    assert "BOQ Activity Mapping" in wb.sheetnames
    mapping_ws = wb["BOQ Activity Mapping"]
    assert mapping_ws["J2"].value == 1.0
    assert mapping_ws["K2"].value == 1000
    wb.close()


def test_progress_reader_keeps_wbs_hierarchy_names(tmp_path: Path):
    progress = tmp_path / "progress_hierarchy.xlsx"
    wb = Workbook(); ws = wb.active; ws.title = "Amount Mapping"
    ws.append(["Activity ID", "WBS", "Description", "Amount", "Status"])
    ws.append([None, "1", "Preparation Works", None, "Waiting"])
    ws.append([None, "1.1", "Mobilization", None, "Waiting"])
    ws.append(["A1000", "1.1.1", "Site Mobilization", 0, "UNMAPPED"])
    add_main_sheet(wb, [("A1000", "Site Mobilization")])
    wb.save(progress); wb.close()

    activity = BOQMappingService().read_activities(progress)[0]
    assert activity.wbs_path == (("1", "Preparation Works"), ("1.1", "Mobilization"))
    assert "preparation works" in activity.search_text
