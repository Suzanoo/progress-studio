from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from tests._paths import REPO_ROOT, FIXTURES_ROOT

from openpyxl import load_workbook

from progress_studio.infrastructure.excel import ImportWorkbookWriter
from progress_studio.infrastructure.schedule_xml import ScheduleXmlReader, ScheduleXmlValidationError
from progress_studio.services.import_service import ImportService
from progress_studio.services.schedule_service import ScheduleService

FIXTURES = FIXTURES_ROOT / "xml"
EXAMPLE_XML = REPO_ROOT / "example" / "example.xml"


class GenericXmlImportTests(unittest.TestCase):
    def setUp(self) -> None:
        self.reader = ScheduleXmlReader()

    def test_existing_p6_msp_export_remains_supported(self) -> None:
        project, rows = self.reader.read(EXAMPLE_XML)
        self.assertEqual(project, "NKC2")
        self.assertGreater(sum(not row.is_summary for row in rows), 0)

    def test_generic_flat_xml_generates_deterministic_ids(self) -> None:
        project, rows = self.reader.read(FIXTURES / "generic_flat.xml")
        leaves = [row for row in rows if not row.is_summary]
        self.assertEqual(project, "Generic Flat Project")
        self.assertEqual([row.activity_id for row in leaves], ["ACT-000001", "ACT-000002"])
        self.assertTrue(all(row.outline_level == 1 for row in leaves))

    def test_partial_and_duplicate_ids_are_made_unique(self) -> None:
        _, rows = self.reader.read(FIXTURES / "generic_partial_ids.xml")
        self.assertEqual([row.activity_id for row in rows], ["A1000", "ACT-000002", "A1000__2"])

    def test_namespace_and_unicode_are_supported(self) -> None:
        project, rows = self.reader.read(FIXTURES / "namespace_prefixed.xml")
        self.assertEqual(project, "Namespaced")
        self.assertEqual(rows[0].name, "Steel work")
        project, rows = self.reader.read(FIXTURES / "unicode_activities.xml")
        self.assertEqual(project, "โครงการทดสอบ")
        self.assertIn("งานฐานราก", rows[0].name)

    def test_required_contract_failures_stop_import(self) -> None:
        for filename, field in (
            ("missing_name.xml", "Activity Name"),
            ("missing_start.xml", "Plan Start"),
            ("missing_finish.xml", "Plan Finish"),
            ("finish_before_start.xml", "Schedule Window"),
            ("empty_schedule.xml", "Activities"),
            ("malformed.xml", "Activities"),
        ):
            with self.subTest(filename=filename):
                with self.assertRaises(ScheduleXmlValidationError) as caught:
                    self.reader.read(FIXTURES / filename)
                self.assertIn(field, str(caught.exception))

    def test_validation_collects_multiple_errors(self) -> None:
        with self.assertRaises(ScheduleXmlValidationError) as caught:
            self.reader.read(FIXTURES / "multiple_validation_errors.xml")
        self.assertGreaterEqual(len(caught.exception.issues), 3)
        self.assertIn("No workbook was created", str(caught.exception))

    def test_end_to_end_import_writes_workbook_only_after_validation(self) -> None:
        service = ImportService(self.reader, ScheduleService(), ImportWorkbookWriter())
        with tempfile.TemporaryDirectory() as folder:
            output = Path(folder) / "imported.xlsx"
            project, summaries, activities = service.import_xml(FIXTURES / "generic_flat.xml", output)
            self.assertEqual((project, summaries, activities), ("Generic Flat Project", 0, 2))
            self.assertTrue(output.exists())
            workbook = load_workbook(output, data_only=False)
            self.assertIn("main", workbook.sheetnames)
            workbook.close()

            rejected = Path(folder) / "rejected.xlsx"
            with self.assertRaises(ScheduleXmlValidationError):
                service.import_xml(FIXTURES / "missing_finish.xml", rejected)
            self.assertFalse(rejected.exists())


if __name__ == "__main__":
    unittest.main()
