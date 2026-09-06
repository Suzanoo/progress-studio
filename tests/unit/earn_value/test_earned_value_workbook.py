from __future__ import annotations

from datetime import datetime

import pytest
from openpyxl import Workbook

from progress_studio.domain.earned_value import (
    ActivityEarnedValue,
    BOQEarnedValue,
    EarnedValuePoint,
    EarnedValueResult,
)
from progress_studio.infrastructure.excel.earned_value_workbook import (
    EARNED_VALUE_SHEET,
    EV_DATA_SHEET,
    render_earned_value_sheet,
)


def _point(key: str, reporting_date: datetime, pv: float | None, ev: float | None) -> EarnedValuePoint:
    return EarnedValuePoint(
        period_key=key,
        reporting_date=reporting_date,
        planned_value=pv,
        earned_value=ev,
        schedule_variance=None if pv is None or ev is None else ev - pv,
        schedule_performance_index=None if pv in (None, 0.0) or ev is None else ev / pv,
    )


def _result() -> EarnedValueResult:
    project_points = (
        _point("W1", datetime(2026, 1, 30), 2_000_000.0, 1_500_000.0),
        _point("W2", datetime(2026, 2, 6), 4_000_000.0, 3_000_000.0),
        _point("W3", datetime(2026, 2, 13), 6_000_000.0, 4_500_000.0),
        _point("W4", datetime(2026, 2, 20), 8_000_000.0, None),
        _point("W5", datetime(2026, 2, 27), 9_000_000.0, None),
        _point("W6", datetime(2026, 3, 6), 10_000_000.0, None),
    )
    a1_points = tuple(
        _point(
            p.period_key,
            p.reporting_date,
            (p.planned_value or 0) * 0.6,
            None if p.earned_value is None else p.earned_value * 0.6,
        )
        for p in project_points
    )
    a2_points = tuple(
        _point(
            p.period_key,
            p.reporting_date,
            (p.planned_value or 0) * 0.4,
            None if p.earned_value is None else p.earned_value * 0.4,
        )
        for p in project_points
    )
    boq1 = (
        _point("W1", datetime(2026, 1, 30), 1_000_000.0, 700_000.0),
        _point("W2", datetime(2026, 2, 6), 2_000_000.0, 1_400_000.0),
        _point("W3", datetime(2026, 2, 13), 3_000_000.0, 2_000_000.0),
        _point("W4", datetime(2026, 2, 20), 4_000_000.0, None),
        _point("W5", datetime(2026, 2, 27), 4_500_000.0, None),
        _point("W6", datetime(2026, 3, 6), 5_000_000.0, None),
    )
    boq2 = (
        _point("W1", datetime(2026, 1, 30), 1_000_000.0, 800_000.0),
        _point("W2", datetime(2026, 2, 6), 2_000_000.0, 1_600_000.0),
        _point("W3", datetime(2026, 2, 13), 3_000_000.0, 2_500_000.0),
        _point("W4", datetime(2026, 2, 20), 4_000_000.0, None),
        _point("W5", datetime(2026, 2, 27), 4_500_000.0, None),
        _point("W6", datetime(2026, 3, 6), 5_000_000.0, None),
    )
    return EarnedValueResult(
        cutoff_date=datetime(2026, 2, 13),
        project_bac=10_000_000.0,
        project_points=project_points,
        activities=(
            ActivityEarnedValue("A1", "Structure Activity", "Structure", 6_000_000.0, a1_points),
            ActivityEarnedValue("A2", "MEP Activity", "MEP", 4_000_000.0, a2_points),
        ),
        boq_items=(
            BOQEarnedValue("B1", "BOQ-001", "Concrete", 5_000_000.0, boq1),
            BOQEarnedValue("B2", "BOQ-002", "Electrical", 5_000_000.0, boq2),
        ),
    )


@pytest.mark.unit
def test_ev5_dashboard_kpis_are_formula_driven_by_status_date() -> None:
    workbook = Workbook()
    render_earned_value_sheet(workbook, _result())

    ws = workbook[EARNED_VALUE_SHEET]
    assert ws["A1"].value == "EARNED VALUE"
    assert ws["M3"].value == datetime(2026, 2, 13)
    assert ws["A6"].value == pytest.approx(10_000_000.0)
    assert ws["D6"].value.startswith("=IFERROR(SUMIFS(EV_Data!$B$")
    assert ws["G6"].value.startswith("=IFERROR(SUMIFS(EV_Data!$D$")
    assert ws["J6"].value == "=G6-D6"
    assert ws["M6"].value == "=IF(D6=0,0,G6/D6)"
    assert ws["J8"].value.startswith("=IF(")
    assert ws["M8"].value.startswith("=IF(")


@pytest.mark.unit
def test_ev5_chart_keeps_full_pv_and_uses_cutoff_masked_ev_and_status_line() -> None:
    workbook = Workbook()
    render_earned_value_sheet(workbook, _result())

    data = workbook[EV_DATA_SHEET]
    assert data.sheet_state == "hidden"
    rows = [
        tuple(data.cell(row, col).value for col in range(1, 7))
        for row in range(2, 6)
    ]
    assert [row[0] for row in rows] == [
        datetime(2026, 1, 30),
        datetime(2026, 2, 13),
        datetime(2026, 2, 27),
        datetime(2026, 3, 6),
    ]
    assert [row[5] for row in rows] == ["Monthly", "Status Date", "Monthly", "Monthly"]
    assert rows[1][1] == pytest.approx(6_000_000.0)
    assert rows[1][2] == pytest.approx(4_500_000.0)
    assert rows[2][1] == pytest.approx(9_000_000.0)
    assert rows[2][2] is None
    assert "'Earned Value'!$M$3" in rows[0][3]
    assert "'Earned Value'!$M$3" in rows[0][4]

    dashboard = workbook[EARNED_VALUE_SHEET]
    chart = dashboard._charts[0]
    assert len(chart.series) == 5
    assert chart.legend is not None
    assert chart.legend.position == "t"
    assert chart.x_axis.axId == 10
    assert chart.x_axis.crossAx == 100
    assert chart.y_axis.axId == 100
    assert chart.y_axis.crossAx == 10
    assert chart.series[2].errBars is not None
    assert chart.series[2].errBars.errBarType == "minus"
    assert chart.series[2].dLbls.showCatName is True
    assert chart.series[3].marker.symbol == "circle"
    assert chart.series[4].marker.symbol == "circle"
    assert data["AC1"].value == "PV @ Status Date"
    assert data["AD1"].value == "EV @ Status Date"
    assert "ISNUMBER(E2)" in data["AC2"].value
    assert "ISNUMBER(E2)" in data["AD2"].value


@pytest.mark.unit
def test_ev5_midweek_cutoff_does_not_interpolate_source_values() -> None:
    result = _result()
    shifted = EarnedValueResult(
        cutoff_date=datetime(2026, 2, 16),
        project_bac=result.project_bac,
        project_points=result.project_points,
        activities=result.activities,
        boq_items=result.boq_items,
    )
    workbook = Workbook()
    render_earned_value_sheet(workbook, shifted)

    data = workbook[EV_DATA_SHEET]
    status_row = next(
        row
        for row in range(2, 10)
        if data.cell(row, 6).value == "Status Date"
    )
    assert data.cell(status_row, 1).value == datetime(2026, 2, 16)
    assert data.cell(status_row, 2).value == pytest.approx(6_000_000.0)
    assert data.cell(status_row, 3).value == pytest.approx(4_500_000.0)


@pytest.mark.unit
def test_ev5_wbs_and_negative_variance_tables_follow_status_date() -> None:
    workbook = Workbook()
    render_earned_value_sheet(workbook, _result())

    ws = workbook[EARNED_VALUE_SHEET]
    data = workbook[EV_DATA_SHEET]

    # WBS labels and values are both keyed by selected Status Date + rank.
    assert ws["A32"].value.startswith("=IFERROR(VLOOKUP(")
    assert 'TEXT($M$3,"yyyymmdd")&"|1"' in ws["A32"].value
    assert ws["B32"].value.startswith("=IFERROR(VLOOKUP(")
    assert ws["D32"].value.startswith("=IFERROR(VLOOKUP(")
    assert ws["E32"].value.startswith("=IFERROR(VLOOKUP(")
    assert ws["F32"].value.startswith("=IFERROR(VLOOKUP(")
    assert ws["G32"].value.startswith("=IFERROR(VLOOKUP(")
    assert ws["H32"].value.startswith("=IFERROR(VLOOKUP(")

    # Hidden snapshots are keyed independently per cutoff.  This isolated
    # fixture has no main sheet, so the compatibility fallback is label order.
    structure = next(
        row for row in range(2, data.max_row + 1)
        if data.cell(row, 10).value == "20260213|1"
    )
    assert data.cell(structure, 11).value == datetime(2026, 2, 13)
    assert data.cell(structure, 12).value == 1
    assert data.cell(structure, 13).value == "MEP"
    assert data.cell(structure, 16).value == pytest.approx(2_400_000.0)
    assert data.cell(structure, 17).value == pytest.approx(1_800_000.0)

    assert ws["I32"].value.startswith("=IFERROR(VLOOKUP(")
    assert ws["J32"].value.startswith("=IFERROR(VLOOKUP(")
    concrete = next(
        row for row in range(2, data.max_row + 1)
        if data.cell(row, 20).value == "20260213|1"
    )
    assert data.cell(concrete, 24).value == "Concrete"
    assert data.cell(concrete, 25).value == pytest.approx(-1_000_000.0)


@pytest.mark.unit
def test_ev5_reuses_dashboard_data_monthly_cutoff_list() -> None:
    workbook = Workbook()
    dashboard_data = workbook.create_sheet("Dashboard_Data")
    dashboard_data["K1"] = "Monthly Cutoff"
    dashboard_data["K2"] = datetime(2026, 1, 30)
    dashboard_data["K3"] = datetime(2026, 2, 13)
    # Canonical monthly date remains selectable even when later than EV reporting cutoff.
    dashboard_data["K4"] = datetime(2026, 2, 27)

    render_earned_value_sheet(workbook, _result())

    ws = workbook[EARNED_VALUE_SHEET]
    validation = ws.data_validations.dataValidation[0]
    assert validation.formula1 == '=INDIRECT("Dashboard_Data!$K$2:$K$4")'


@pytest.mark.unit
def test_ev5_cutoff_fallback_contains_only_historical_dates() -> None:
    workbook = Workbook()
    render_earned_value_sheet(workbook, _result())

    ws = workbook[EARNED_VALUE_SHEET]
    data = workbook[EV_DATA_SHEET]
    validation = ws.data_validations.dataValidation[0]
    assert "EV_Data!$H$2:$H$" in validation.formula1
    options = [
        data.cell(row, 8).value
        for row in range(2, data.max_row + 1)
        if isinstance(data.cell(row, 8).value, datetime)
    ]
    assert options == [datetime(2026, 1, 30), datetime(2026, 2, 13)]


@pytest.mark.unit
def test_ev5_replaces_only_owned_ev_sheets() -> None:
    workbook = Workbook()
    main = workbook.active
    main.title = "main"
    main["A1"] = "keep"
    workbook.create_sheet(EARNED_VALUE_SHEET)["A1"] = "old"
    workbook.create_sheet(EV_DATA_SHEET)["A1"] = "old-data"

    render_earned_value_sheet(workbook, _result())

    assert workbook["main"]["A1"].value == "keep"
    assert workbook.sheetnames.count(EARNED_VALUE_SHEET) == 1
    assert workbook.sheetnames.count(EV_DATA_SHEET) == 1
    assert workbook[EV_DATA_SHEET].sheet_state == "hidden"


@pytest.mark.unit
def test_ev52_reuses_dashboard_visual_language_and_avoids_duplicate_cutoff_list() -> None:
    workbook = Workbook()
    dashboard_data = workbook.create_sheet("Dashboard_Data")
    dashboard_data["K1"] = "Monthly Cutoff"
    dashboard_data["K2"] = datetime(2026, 1, 30)
    dashboard_data["K3"] = datetime(2026, 2, 13)

    render_earned_value_sheet(workbook, _result())

    ws = workbook[EARNED_VALUE_SHEET]
    data = workbook[EV_DATA_SHEET]
    chart = ws._charts[0]
    assert data["H1"].value is None
    assert chart.style == 2
    assert chart.y_axis.numFmt.formatCode == '#,##0,,"M"'
    assert chart.y_axis.title.tx.rich.p[0].r[0].t == "Value (Million)"
    assert chart.x_axis.title is None
    assert chart.series[0].graphicalProperties.line.width == 26000
    assert chart.series[1].graphicalProperties.line.width == 26000
    assert ws["A29"].value == "PV curve = full baseline    •    EV curve = selected Status Date"


@pytest.mark.unit
def test_ev52_status_cards_use_dashboard_condition_language() -> None:
    workbook = Workbook()
    render_earned_value_sheet(workbook, _result())

    ws = workbook[EARNED_VALUE_SHEET]
    ranges = {str(key.sqref) for key in ws.conditional_formatting._cf_rules}
    assert "J5:L8" in ranges
    assert "M5:O8" in ranges
    assert any(str(key.sqref).startswith("G32:H") for key in ws.conditional_formatting._cf_rules)

@pytest.mark.unit
def test_ev6_renders_live_boq_table_with_native_filter_and_mapping_metadata() -> None:
    workbook = Workbook()
    mapping = workbook.create_sheet("BOQ Activity Mapping")
    mapping.append([
        "Activity ID", "BOQ Key", "Source Sheet", "Source Row", "WBS-2", "WBS-3", "WBS-4",
        "BOQ Description", "BOQ Amount", "Share %", "Allocated Amount", "Mapping ID", "BOQ ID",
    ])
    mapping.append(["A1", "B1", "BOQ", 2, "CSA", "Structure", "Concrete", "Concrete mapped", 5_000_000, 1.0, 5_000_000, "M1", "BOQ-001"])
    mapping.append(["A2", "B2", "BOQ", 3, "MEP", "Electrical", "Power", "Electrical mapped", 5_000_000, 1.0, 5_000_000, "M2", "BOQ-002"])

    render_earned_value_sheet(workbook, _result())

    ws = workbook["EV Table"]
    data = workbook[EV_DATA_SHEET]
    assert ws["A1"].value == "EARNED VALUE TABLE"
    assert ws["B3"].value == "='Earned Value'!$M$3"
    assert [ws.cell(5, col).value for col in range(1, 11)] == [
        "BOQ ID", "WBS-2", "WBS-3", "WBS-4", "BOQ / WORK", "BAC", "PV", "EV", "SV", "SPI"
    ]
    assert ws.auto_filter.ref == "A5:J7"
    assert ws.column_dimensions["K"].hidden is True
    assert ws["A6"].value == "BOQ-001"
    assert ws["B6"].value == "CSA"
    assert ws["C6"].value == "Structure"
    assert ws["D6"].value == "Concrete"
    assert ws["E6"].value == "Concrete mapped"
    assert ws["F6"].value == pytest.approx(5_000_000.0)
    assert ws["G6"].value.startswith("=IFERROR(SUMIFS(EV_Data!$AG$")
    assert "$B$3" in ws["G6"].value
    assert ws["H6"].value.startswith("=IFERROR(SUMIFS(EV_Data!$AH$")
    assert ws["I6"].value == "=H6-G6"
    assert ws["J6"].value == "=IF(G6=0,0,H6/G6)"

    snapshots = [
        tuple(data.cell(row, col).value for col in range(31, 35))
        for row in range(2, data.max_row + 1)
        if data.cell(row, 31).value is not None
    ]
    # Two selectable cutoffs x two BOQ items; compact helper stores no repeated labels/WBS/BAC.
    assert len(snapshots) == 4
    assert snapshots[0][0] == datetime(2026, 1, 30)
    assert snapshots[0][1] == "B1"
    assert snapshots[-1][0] == datetime(2026, 2, 13)


@pytest.mark.unit
def test_ev6_refresh_replaces_ev_table_with_other_ev_owned_sheets() -> None:
    workbook = Workbook()
    main = workbook.active
    main.title = "main"
    main["A1"] = "keep"
    workbook.create_sheet(EARNED_VALUE_SHEET)["A1"] = "old-dashboard"
    workbook.create_sheet("EV Table")["A1"] = "old-table"
    workbook.create_sheet(EV_DATA_SHEET)["A1"] = "old-data"

    render_earned_value_sheet(workbook, _result())

    assert workbook["main"]["A1"].value == "keep"
    assert workbook.sheetnames.count(EARNED_VALUE_SHEET) == 1
    assert workbook.sheetnames.count("EV Table") == 1
    assert workbook.sheetnames.count(EV_DATA_SHEET) == 1
    assert workbook["EV Table"]["A1"].value == "EARNED VALUE TABLE"
    dashboard_links = [
        cell.hyperlink.target
        for row in workbook[EARNED_VALUE_SHEET].iter_rows()
        for cell in row
        if cell.hyperlink is not None
    ]
    assert "#'EV Table'!A1" in dashboard_links

@pytest.mark.unit
def test_evp_reuses_progress_dashboard_theme_for_cards_and_detail_navigation() -> None:
    workbook = Workbook()
    render_earned_value_sheet(workbook, _result())

    ws = workbook[EARNED_VALUE_SHEET]
    table = workbook["EV Table"]

    # Theme values come from dashboard_workbook/_load_dashboard_theme rather
    # than a second EV palette.
    assert ws["A5"].fill.fgColor.rgb.endswith("F3F4F6")  # BAC / light gray
    assert ws["D5"].fill.fgColor.rgb.endswith("EAF2F8")  # PV / light blue
    assert ws["G5"].fill.fgColor.rgb.endswith("EAF4E3")  # EV / light green
    assert ws["J5"].fill.fgColor.rgb.endswith("FFF4D6")  # SV / light amber
    assert ws["M5"].fill.fgColor.rgb.endswith("FFF4D6")  # SPI / light amber
    assert ws.sheet_properties.tabColor.rgb.endswith("17365D")

    assert table["J3"].value == "← EV Dashboard"
    assert table["J3"].hyperlink.target == "#'Earned Value'!A1"
    assert table.sheet_properties.tabColor.rgb.endswith("17365D")

@pytest.mark.unit
def test_ev_wbs_uses_authoritative_main_wbs_code_name_and_all_active_rows() -> None:
    cutoff = datetime(2026, 2, 27)
    activities = []
    project_points = (_point("W1", cutoff, 10_000.0, 8_000.0),)
    workbook = Workbook()
    main = workbook.active
    main.title = "main"

    for index in range(1, 11):
        activity_id = f"A{index}"
        wbs_code = str(index)
        main.append(["WBS", wbs_code, f"WBS Name {index}", "P"])
        main.append(["Activity", f"{wbs_code}.1", f"Activity {index}", "P", activity_id])
        activities.append(
            ActivityEarnedValue(
                activity_id, f"Activity {index}", f"NOT-WBS-{index}", 1_000.0,
                (_point("W1", cutoff, 1_000.0, 800.0),),
            )
        )

    result = EarnedValueResult(
        cutoff_date=cutoff,
        project_bac=10_000.0,
        project_points=project_points,
        activities=tuple(activities),
        boq_items=(),
    )
    render_earned_value_sheet(workbook, result)
    data = workbook[EV_DATA_SHEET]
    dashboard = workbook[EARNED_VALUE_SHEET]

    snapshot_rows = [
        row for row in range(2, data.max_row + 1)
        if data.cell(row, 10).value
    ]
    assert len(snapshot_rows) == 10
    assert data.cell(snapshot_rows[0], 13).value == "1"
    assert data.cell(snapshot_rows[0], 14).value == "WBS Name 1"
    assert data.cell(snapshot_rows[-1], 13).value == "10"
    assert data.cell(snapshot_rows[-1], 14).value == "WBS Name 10"
    assert dashboard["A31"].value == "WBS"
    assert dashboard["B31"].value == "WBS NAME"
    assert dashboard["A41"].value.startswith("=IFERROR(VLOOKUP(")


@pytest.mark.unit
def test_ev_wbs_excludes_future_zero_pv_work_from_active_rows() -> None:
    cutoff = datetime(2026, 4, 24)
    result = EarnedValueResult(
        cutoff_date=cutoff,
        project_bac=13_000_000.0,
        project_points=(_point("W1", cutoff, 35_152.54, 0.0),),
        activities=(
            ActivityEarnedValue(
                "A1", "Active", "fallback-active", 100_000.0,
                (_point("W1", cutoff, 35_152.54, 0.0),),
            ),
            ActivityEarnedValue(
                "A2", "Future", "fallback-future", 12_770_728.24,
                (_point("W1", cutoff, 0.0, 0.0),),
            ),
        ),
        boq_items=(),
    )
    workbook = Workbook()
    main = workbook.active
    main.title = "main"
    main.append(["WBS", "1", "Current Works", "P"])
    main.append(["Activity", "1.1", "Active", "P", "A1"])
    main.append(["WBS", "6", "Future Works", "P"])
    main.append(["Activity", "6.2", "Future", "P", "A2"])

    render_earned_value_sheet(workbook, result)
    data = workbook[EV_DATA_SHEET]

    labels = [
        data.cell(row, 13).value
        for row in range(2, data.max_row + 1)
        if data.cell(row, 10).value
    ]
    assert labels == ["1"]
    assert "6" not in labels


@pytest.mark.unit
def test_ev_top10_negative_variance_includes_activity_ids_from_mapping() -> None:
    workbook = Workbook()
    mapping = workbook.create_sheet("BOQ Activity Mapping")
    mapping.append([
        "Activity ID", "BOQ Key", "Source Sheet", "Source Row", "WBS-2", "WBS-3", "WBS-4",
        "BOQ Description", "BOQ Amount", "Share %", "Allocated Amount", "Mapping ID", "BOQ ID",
    ])
    mapping.append(["A1", "B1", "BOQ", 2, "CSA", "Structure", "Concrete", "Concrete", 5_000_000, 1.0, 5_000_000, "M1", "BOQ-001"])
    mapping.append(["A2", "B2", "BOQ", 3, "MEP", "Electrical", "Power", "Electrical", 5_000_000, 1.0, 5_000_000, "M2", "BOQ-002"])

    render_earned_value_sheet(workbook, _result())
    dashboard = workbook[EARNED_VALUE_SHEET]
    data = workbook[EV_DATA_SHEET]

    assert dashboard["I30"].value == "TOP 10 NEGATIVE VARIANCE"
    assert dashboard["I31"].value == "ACTIVITY ID"
    assert dashboard["J31"].value == "BOQ / WORK"
    assert dashboard["M31"].value == "SV"
    assert dashboard["O31"].value == "SPI"
    first = next(
        row for row in range(2, data.max_row + 1)
        if data.cell(row, 20).value == "20260213|1"
    )
    assert data.cell(first, 23).value == "A1"
    assert data.cell(first, 24).value == "Concrete"
    assert dashboard["I41"].value.startswith("=IFERROR(VLOOKUP(")



@pytest.mark.unit
def test_ev_dashboard_uses_balanced_grid_and_merged_management_text_fields() -> None:
    workbook = Workbook()
    render_earned_value_sheet(workbook, _result())

    ws = workbook[EARNED_VALUE_SHEET]

    # KPI cards remain five equal three-column blocks across the A:O canvas.
    merged = {str(rng) for rng in ws.merged_cells.ranges}
    for rng in ("A5:C5", "D5:F5", "G5:I5", "J5:L5", "M5:O5"):
        assert rng in merged

    # Long management labels expand by merging cells, not by distorting KPI columns.
    assert "B31:C31" in merged
    assert "B32:C32" in merged
    assert "J31:L31" in merged
    assert "J32:L32" in merged
    assert "M31:N31" in merged
    assert "M32:N32" in merged

    # Worksheet-level duplicate legend is gone; the chart owns the legend.
    assert ws["K10"].value is None
    assert ws["L10"].value is None
    assert ws["M10"].value is None
    assert ws._charts[0].width >= 38.0

    # WBS metrics and Top-10 metrics land on the expanded grid.
    assert ws["H32"].number_format == "0.00"
    assert ws["M32"].number_format == "#,##0.00"
    assert ws["O32"].number_format == "0.00"
