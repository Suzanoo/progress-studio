from __future__ import annotations

from datetime import datetime
from pathlib import Path
from types import SimpleNamespace

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.chart import LineChart, Reference

from progress_studio.domain.earned_value import EarnedValuePoint, EarnedValueResult
from progress_studio.services.earned_value_rebuild_service import (
    EarnedValueRebuildError,
    EarnedValueRebuildService,
)


CUTOFF = datetime(2026, 8, 28)


class StubInputReader:
    def __init__(self) -> None:
        self.boq_rows = (SimpleNamespace(key="B1"), SimpleNamespace(key="B2"))
        self.allocations = (
            SimpleNamespace(boq_key="B1", activity_id="A1", share_percent=100.0),
            SimpleNamespace(boq_key="B2", activity_id="A2", share_percent=100.0),
        )

    def read(self, path: Path):
        return SimpleNamespace(
            workbook=path,
            boq_rows=self.boq_rows,
            allocations=self.allocations,
        )


class StubMainReader:
    def read_main_dataset(self, path: Path):
        return SimpleNamespace(
            activities=(SimpleNamespace(activity_id="A1"), SimpleNamespace(activity_id="A2")),
            periods=(SimpleNamespace(column=18, reporting_date=CUTOFF),),
            rows=(),
        )


class StubDeriver:
    def derive(self, dataset, boq_rows, allocations, *, cutoff_date):
        assert cutoff_date == CUTOFF
        return EarnedValueResult(
            cutoff_date=cutoff_date,
            project_bac=1000.0,
            project_points=(
                EarnedValuePoint(
                    period_key="W1",
                    reporting_date=CUTOFF,
                    planned_value=600.0,
                    earned_value=500.0,
                    schedule_variance=-100.0,
                    schedule_performance_index=5.0 / 6.0,
                ),
            ),
            activities=(),
            boq_items=(),
        )


def _source_workbook(tmp_path: Path, *, cutoff=CUTOFF, earned_value=False) -> Path:
    workbook = Workbook()
    main = workbook.active
    main.title = "main"
    main["A1"] = "user main data"
    main["A2"] = "Week"
    main["B2"] = "Plan"
    main["A3"] = 1
    main["B3"] = 10
    chart = LineChart()
    chart.add_data(Reference(main, min_col=2, min_row=2, max_row=3), titles_from_data=True)
    chart.set_categories(Reference(main, min_col=1, min_row=3, max_row=3))
    main.add_chart(chart, "D2")

    monthly = workbook.create_sheet("main_monthly")
    monthly["A1"] = "keep monthly"
    monthly["A2"] = "Month"
    monthly["B2"] = "Plan"
    monthly["A3"] = 1
    monthly["B3"] = 10
    monthly_chart = LineChart()
    monthly_chart.add_data(
        Reference(monthly, min_col=2, min_row=2, max_row=3), titles_from_data=True
    )
    monthly_chart.set_categories(Reference(monthly, min_col=1, min_row=3, max_row=3))
    monthly.add_chart(monthly_chart, "D2")

    dashboard = workbook.create_sheet("Dashboard")
    dashboard["K5"] = cutoff
    dashboard["A1"] = "keep dashboard"

    payment = workbook.create_sheet("Payment")
    payment["A1"] = "keep payment"

    if earned_value:
        workbook.create_sheet("Earned Value")["A1"] = "old"
        workbook.create_sheet("EV_Data")["A1"] = "old-data"

    path = tmp_path / "source.xlsx"
    workbook.save(path)
    workbook.close()
    return path


def _service() -> EarnedValueRebuildService:
    return EarnedValueRebuildService(
        input_reader=StubInputReader(),
        main_reader=StubMainReader(),
        deriver=StubDeriver(),
    )


@pytest.mark.unit
def test_ev_rebuild_analysis_reports_readiness_and_existing_sheet(tmp_path: Path) -> None:
    source = _source_workbook(tmp_path, earned_value=True)

    analysis = _service().analyze(source)

    assert analysis.cutoff_date == CUTOFF
    assert analysis.activity_count == 2
    assert analysis.boq_count == 2
    assert analysis.allocation_count == 2
    assert analysis.project_bac == pytest.approx(1000.0)
    assert analysis.existing_earned_value_sheet is True


@pytest.mark.unit
def test_ev_rebuild_generate_adds_native_ev_extension_and_preserves_existing_features(
    tmp_path: Path,
) -> None:
    source = _source_workbook(tmp_path)
    output = tmp_path / "output.xlsx"

    result = _service().generate(source, output)

    assert result.refreshed_sheet == "Earned Value"
    assert output.exists()

    workbook = load_workbook(output, data_only=False)
    try:
        assert workbook["main"]["A1"].value == "user main data"
        assert workbook["main_monthly"]["A1"].value == "keep monthly"
        assert workbook["Dashboard"]["A1"].value == "keep dashboard"
        assert workbook["Payment"]["A1"].value == "keep payment"
        assert len(workbook["main"]._charts) == 1
        assert len(workbook["main_monthly"]._charts) == 1
        assert workbook.sheetnames.count("Earned Value") == 1
        assert workbook.sheetnames.count("EV_Data") == 1
        assert workbook["EV_Data"].sheet_state == "hidden"
        assert len(workbook["Earned Value"]._charts) == 1
    finally:
        workbook.close()


@pytest.mark.unit
def test_ev_rebuild_refresh_replaces_only_ev_owned_sheets(tmp_path: Path) -> None:
    source = _source_workbook(tmp_path, earned_value=True)
    output = tmp_path / "output.xlsx"

    _service().generate(source, output)

    workbook = load_workbook(output, data_only=False)
    try:
        assert workbook.sheetnames.count("Earned Value") == 1
        assert workbook.sheetnames.count("EV_Data") == 1
        assert workbook["main"]["A1"].value == "user main data"
        assert workbook["main_monthly"]["A1"].value == "keep monthly"
        assert workbook["Dashboard"]["A1"].value == "keep dashboard"
        assert workbook["Payment"]["A1"].value == "keep payment"
    finally:
        workbook.close()


@pytest.mark.unit
def test_ev_rebuild_does_not_use_main_or_dashboard_cutoff_as_view_seed(tmp_path: Path) -> None:
    workbook = Workbook()
    main = workbook.active
    main.title = "main"
    main["L5"] = "Cutoff Date"
    main["M5"] = datetime(2026, 4, 24)
    dashboard = workbook.create_sheet("Dashboard")
    dashboard["K5"] = datetime(2026, 5, 29)
    path = tmp_path / "neutral-seed.xlsx"
    workbook.save(path)
    workbook.close()

    analysis = _service().analyze(path)
    assert analysis.cutoff_date == CUTOFF


@pytest.mark.unit
def test_ev_rebuild_without_any_reporting_date_is_blocked(tmp_path: Path) -> None:
    class EmptyMainReader:
        def read_main_dataset(self, path: Path):
            return SimpleNamespace(activities=(SimpleNamespace(activity_id="A1"),), periods=(), rows=())

    source = _source_workbook(tmp_path, cutoff=None)
    service = EarnedValueRebuildService(
        input_reader=StubInputReader(), main_reader=EmptyMainReader(), deriver=StubDeriver()
    )
    with pytest.raises(EarnedValueRebuildError, match="requires at least one valid reporting date"):
        service.analyze(source)


@pytest.mark.unit
def test_ev_rebuild_refuses_in_place_output(tmp_path: Path) -> None:
    source = _source_workbook(tmp_path)

    with pytest.raises(EarnedValueRebuildError, match="must be a new workbook path"):
        _service().generate(source, source)

@pytest.mark.unit
def test_ev_rebuild_refresh_preserves_existing_ev_view(tmp_path: Path) -> None:
    source = _source_workbook(tmp_path, earned_value=True)
    selected_view_date = datetime(2026, 7, 31)
    workbook = load_workbook(source)
    workbook["Earned Value"]["M3"] = selected_view_date
    workbook.save(source)
    workbook.close()

    assert EarnedValueRebuildService._view_date_seed(
        source, StubMainReader().read_main_dataset(source)
    ) == selected_view_date
