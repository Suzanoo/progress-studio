from __future__ import annotations

from datetime import datetime
from pathlib import Path
from types import SimpleNamespace

import pytest
from openpyxl import Workbook, load_workbook

from progress_studio.domain.earned_value import EarnedValuePoint, EarnedValueResult
from progress_studio.infrastructure.excel.earned_value_workbook import (
    EARNED_VALUE_SHEET,
    EV_DATA_SHEET,
    render_earned_value_sheet,
)
from progress_studio.services.earned_value_rebuild_service import (
    EarnedValueRebuildError,
    EarnedValueRebuildService,
)


REPORTING_CUTOFF = datetime(2026, 8, 28)
STALE_EV_VIEW = datetime(2026, 4, 24)


class _InputReader:
    def read(self, path: Path):
        return SimpleNamespace(
            workbook=path,
            boq_rows=(SimpleNamespace(key="B1"),),
            allocations=(SimpleNamespace(
                boq_key="B1", activity_id="A1", share_percent=100.0
            ),),
        )


class _MainReader:
    def read_main_dataset(self, path: Path):
        return SimpleNamespace(
            activities=(SimpleNamespace(activity_id="A1"),)
        )


class _Deriver:
    def derive(self, dataset, boq_rows, allocations, *, cutoff_date):
        assert cutoff_date == REPORTING_CUTOFF
        return EarnedValueResult(
            cutoff_date=cutoff_date,
            project_bac=1000.0,
            project_points=(
                EarnedValuePoint(
                    period_key="W1",
                    reporting_date=cutoff_date,
                    planned_value=600.0,
                    earned_value=500.0,
                    schedule_variance=-100.0,
                    schedule_performance_index=5.0 / 6.0,
                ),
            ),
            activities=(),
            boq_items=(),
        )


def _service() -> EarnedValueRebuildService:
    return EarnedValueRebuildService(
        input_reader=_InputReader(),
        main_reader=_MainReader(),
        deriver=_Deriver(),
    )


def _workbook_with_dashboard(tmp_path: Path, cutoff=REPORTING_CUTOFF) -> Path:
    workbook = Workbook()
    workbook.active.title = "main"
    dashboard = workbook.create_sheet("Dashboard")
    dashboard["K5"] = cutoff
    ev = workbook.create_sheet(EARNED_VALUE_SHEET)
    ev["M3"] = STALE_EV_VIEW
    path = tmp_path / "reporting-context.xlsx"
    workbook.save(path)
    workbook.close()
    return path


@pytest.mark.unit
def test_bf1_stale_ev_view_never_overrides_dashboard_reporting_cutoff(
    tmp_path: Path,
) -> None:
    path = _workbook_with_dashboard(tmp_path)
    assert EarnedValueRebuildService._read_cutoff(path) == REPORTING_CUTOFF
    assert _service().analyze(path).cutoff_date == REPORTING_CUTOFF


@pytest.mark.unit
def test_bf1_stale_ev_view_is_not_used_when_no_saved_view_seed_exists(
    tmp_path: Path,
) -> None:
    path = _workbook_with_dashboard(tmp_path, cutoff=None)
    assert EarnedValueRebuildService._read_cutoff(path) is None
    with pytest.raises(
        EarnedValueRebuildError, match="requires at least one valid reporting date"
    ):
        _service().analyze(path)


@pytest.mark.unit
def test_bf1_main_cutoff_remains_fallback_when_dashboard_is_absent(
    tmp_path: Path,
) -> None:
    workbook = Workbook()
    main = workbook.active
    main.title = "main"
    main["L5"] = "Cutoff Date"
    main["M5"] = REPORTING_CUTOFF
    ev = workbook.create_sheet(EARNED_VALUE_SHEET)
    ev["M3"] = STALE_EV_VIEW
    path = tmp_path / "main-fallback.xlsx"
    workbook.save(path)
    workbook.close()

    assert EarnedValueRebuildService._read_cutoff(path) == REPORTING_CUTOFF
    assert _service().analyze(path).cutoff_date == REPORTING_CUTOFF


def _point(key: str, when: datetime, pv: float, ev: float | None) -> EarnedValuePoint:
    return EarnedValuePoint(
        period_key=key,
        reporting_date=when,
        planned_value=pv,
        earned_value=ev,
        schedule_variance=None if ev is None else ev - pv,
        schedule_performance_index=None if ev is None or pv == 0 else ev / pv,
    )


def _ev_result() -> EarnedValueResult:
    return EarnedValueResult(
        cutoff_date=datetime(2026, 2, 13),
        project_bac=1000.0,
        project_points=(
            _point("W1", datetime(2026, 1, 30), 200.0, 150.0),
            _point("W2", datetime(2026, 2, 13), 400.0, 300.0),
            _point("W3", datetime(2026, 2, 27), 600.0, None),
            _point("W4", datetime(2026, 3, 27), 800.0, None),
        ),
        activities=(),
        boq_items=(),
    )


@pytest.mark.unit
def test_bf2_ev_dropdown_reuses_complete_dashboard_monthly_calendar() -> None:
    workbook = Workbook()
    dashboard_data = workbook.create_sheet("Dashboard_Data")
    dashboard_data["K1"] = "Monthly Cutoff"
    for row, value in enumerate((
        datetime(2026, 1, 30),
        datetime(2026, 2, 13),
        datetime(2026, 2, 27),
        datetime(2026, 3, 27),
    ), start=2):
        dashboard_data.cell(row, 11, value)

    render_earned_value_sheet(workbook, _ev_result(), include_chart=False)

    ev = workbook[EARNED_VALUE_SHEET]
    data = workbook[EV_DATA_SHEET]
    validation = ev.data_validations.dataValidation[0]
    assert validation.formula1 == '=INDIRECT("Dashboard_Data!$K$2:$K$5")'
    assert data["H1"].value is None


@pytest.mark.unit
def test_bf2_standalone_workbook_keeps_ev_data_fallback_calendar() -> None:
    workbook = Workbook()
    render_earned_value_sheet(workbook, _ev_result(), include_chart=False)

    ev = workbook[EARNED_VALUE_SHEET]
    data = workbook[EV_DATA_SHEET]
    validation = ev.data_validations.dataValidation[0]

    assert "EV_Data!$H$2:$H$" in validation.formula1
    assert data["H1"].value == "View Date Options"
    options = [
        data.cell(row, 8).value
        for row in range(2, data.max_row + 1)
        if isinstance(data.cell(row, 8).value, datetime)
    ]
    assert options == [datetime(2026, 1, 30), datetime(2026, 2, 27), datetime(2026, 3, 27)]

@pytest.mark.unit
def test_ev_initial_view_prefers_latest_actual_month_over_stale_dashboard_cutoff(
    tmp_path: Path,
) -> None:
    class _Period:
        def __init__(self, column: int, reporting_date: datetime) -> None:
            self.column = column
            self.reporting_date = reporting_date

    class _Row:
        pa = "A"
        activity_id = "A1"

        def __init__(self, values: dict[int, float | None]) -> None:
            self._values = values

        def period_value(self, column: int):
            return self._values.get(column)

    path = _workbook_with_dashboard(tmp_path, cutoff=datetime(2026, 4, 24))
    dataset = SimpleNamespace(
        periods=(
            _Period(18, datetime(2026, 4, 24)),
            _Period(19, datetime(2026, 8, 14)),
            _Period(20, datetime(2026, 8, 28)),
        ),
        rows=(_Row({18: 0.10, 19: 0.20, 20: None}),),
    )

    # Actual exists in August even though Dashboard still shows April. The EV
    # initial view uses the final August reporting point, not Dashboard cutoff.
    assert EarnedValueRebuildService._initial_view_date(path, dataset) == datetime(2026, 8, 28)
