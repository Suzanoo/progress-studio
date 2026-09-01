from pathlib import Path

import pytest
from openpyxl import Workbook

from progress_studio.infrastructure.excel.earned_value_input_reader import (
    EarnedValueInputWorkbookReader,
    EarnedValueWorkbookInputError,
)


def _write_summary(
    workbook: Workbook,
    *,
    boq_count: int = 1,
    full_count: int = 1,
    partial_count: int = 0,
    unmapped_count: int = 0,
    allocated_fraction: float = 1.0,
) -> None:
    ws = workbook.create_sheet("Mapping Summary")
    ws.append(["Progress Studio Mapping Reconciliation", None])
    ws.append([None, None])
    ws.append(["BOQ items", boq_count])
    ws.append(["Fully allocated BOQ items", full_count])
    ws.append(["Partially allocated BOQ items", partial_count])
    ws.append(["Unmapped BOQ items", unmapped_count])
    ws.append(["Allocated percent", allocated_fraction])


def _write_mapping(workbook: Workbook, rows: list[list[object]]) -> None:
    ws = workbook.create_sheet("BOQ Activity Mapping")
    ws.append([
        "Activity ID",
        "BOQ Key",
        "Source Sheet",
        "Source Row",
        "WBS-2",
        "WBS-3",
        "WBS-4",
        "BOQ Description",
        "BOQ Amount",
        "Share %",
        "Allocated Amount",
        "Mapping ID",
        "BOQ ID",
    ])
    for row in rows:
        ws.append(row)


def _save(workbook: Workbook, tmp_path: Path) -> Path:
    path = tmp_path / "mapped.xlsx"
    workbook.save(path)
    workbook.close()
    return path


@pytest.mark.unit
def test_reader_reconstructs_boq_and_converts_excel_share_fraction(
    tmp_path: Path,
) -> None:
    workbook = Workbook()
    del workbook[workbook.sheetnames[0]]
    _write_summary(workbook)
    _write_mapping(
        workbook,
        [
            [
                "A1010", "B1", "BOQ", 8, "CSA", "Structure", "Rebar",
                "Reinforcement", 100.0, 0.30, 30.0, "MAP-000001", "BOQ-001",
            ],
            [
                "A1020", "B1", "BOQ", 8, "CSA", "Structure", "Rebar",
                "Reinforcement", 100.0, 0.70, 70.0, "MAP-000002", "BOQ-001",
            ],
        ],
    )

    result = EarnedValueInputWorkbookReader().read(_save(workbook, tmp_path))

    assert len(result.boq_rows) == 1
    assert result.boq_rows[0].key == "B1"
    assert result.boq_rows[0].stable_id == "BOQ-001"
    assert result.boq_rows[0].amount == pytest.approx(100.0)
    assert [allocation.activity_id for allocation in result.allocations] == [
        "A1010",
        "A1020",
    ]
    # Embedded 0.30 / 0.70 fractions become EV domain 30.0 / 70.0 percent.
    assert [allocation.share_percent for allocation in result.allocations] == pytest.approx(
        [30.0, 70.0]
    )


@pytest.mark.unit
def test_reader_hard_stops_when_mapping_summary_is_not_complete(
    tmp_path: Path,
) -> None:
    workbook = Workbook()
    del workbook[workbook.sheetnames[0]]
    _write_summary(
        workbook,
        boq_count=10,
        full_count=8,
        partial_count=1,
        unmapped_count=1,
        allocated_fraction=0.80,
    )
    _write_mapping(
        workbook,
        [[
            "A1", "B1", "BOQ", 1, "", "", "", "Item", 100.0,
            1.0, 100.0, "MAP-000001", "BOQ-001",
        ]],
    )

    with pytest.raises(
        EarnedValueWorkbookInputError,
        match=r"requires 100% BOQ mapping: 8/10 fully allocated, 1 partial, 1 unmapped, 80\.00% allocated",
    ):
        EarnedValueInputWorkbookReader().read(_save(workbook, tmp_path))


@pytest.mark.unit
def test_reader_rejects_mapping_rows_that_do_not_reconcile_to_100_percent(
    tmp_path: Path,
) -> None:
    workbook = Workbook()
    del workbook[workbook.sheetnames[0]]
    _write_summary(workbook)
    _write_mapping(
        workbook,
        [[
            "A1", "B1", "BOQ", 1, "", "", "", "Item", 100.0,
            0.80, 80.0, "MAP-000001", "BOQ-001",
        ]],
    )

    with pytest.raises(
        EarnedValueWorkbookInputError,
        match=r"BOQ Key 'B1': 80\.00%",
    ):
        EarnedValueInputWorkbookReader().read(_save(workbook, tmp_path))


@pytest.mark.unit
def test_reader_rejects_conflicting_duplicate_boq_metadata(
    tmp_path: Path,
) -> None:
    workbook = Workbook()
    del workbook[workbook.sheetnames[0]]
    _write_summary(workbook)
    _write_mapping(
        workbook,
        [
            [
                "A1", "B1", "BOQ", 1, "", "", "", "Item", 100.0,
                0.50, 50.0, "MAP-000001", "BOQ-001",
            ],
            [
                "A2", "B1", "BOQ", 1, "", "", "", "Changed Item", 100.0,
                0.50, 50.0, "MAP-000002", "BOQ-001",
            ],
        ],
    )

    with pytest.raises(
        EarnedValueWorkbookInputError,
        match=r"Conflicting embedded BOQ metadata",
    ):
        EarnedValueInputWorkbookReader().read(_save(workbook, tmp_path))


@pytest.mark.unit
def test_reader_requires_mapping_summary_for_unmapped_boq_detection(
    tmp_path: Path,
) -> None:
    workbook = Workbook()
    del workbook[workbook.sheetnames[0]]
    _write_mapping(
        workbook,
        [[
            "A1", "B1", "BOQ", 1, "", "", "", "Item", 100.0,
            1.0, 100.0, "MAP-000001", "BOQ-001",
        ]],
    )

    with pytest.raises(
        EarnedValueWorkbookInputError,
        match=r"Mapping Summary.*not found",
    ):
        EarnedValueInputWorkbookReader().read(_save(workbook, tmp_path))


@pytest.mark.unit
def test_reader_rejects_missing_mapping_contract_column(
    tmp_path: Path,
) -> None:
    workbook = Workbook()
    del workbook[workbook.sheetnames[0]]
    _write_summary(workbook)
    ws = workbook.create_sheet("BOQ Activity Mapping")
    ws.append(["Activity ID", "BOQ Key"])

    with pytest.raises(
        EarnedValueWorkbookInputError,
        match=r"missing required columns",
    ):
        EarnedValueInputWorkbookReader().read(_save(workbook, tmp_path))
