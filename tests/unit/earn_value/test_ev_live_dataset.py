from __future__ import annotations

from datetime import datetime

import pytest
from openpyxl import Workbook

from progress_studio.domain.earned_value import (
    ActivityEarnedValue,
    BOQEarnedValue,
    EarnedValuePoint,
    EarnedValueResult,
)
from progress_studio.infrastructure.excel.earned_value_workbook import (
    EARNED_VALUE_SHEET,
    EV_DATA_SHEET,
    EV_TABLE_SHEET,
    render_earned_value_sheet,
)


def _point(key: str, when: datetime, pv: float, ev: float | None) -> EarnedValuePoint:
    return EarnedValuePoint(
        period_key=key,
        reporting_date=when,
        planned_value=pv,
        earned_value=ev,
        schedule_variance=None if ev is None else ev - pv,
        schedule_performance_index=None if ev is None or pv == 0 else ev / pv,
    )


def _result() -> EarnedValueResult:
    jan = datetime(2026, 1, 30)
    apr = datetime(2026, 4, 24)
    aug = datetime(2026, 8, 28)
    project = (
        _point("JAN", jan, 100.0, 80.0),
        _point("APR", apr, 400.0, 300.0),
        _point("AUG", aug, 800.0, None),
    )
    activity = (
        _point("JAN", jan, 100.0, 80.0),
        _point("APR", apr, 400.0, 300.0),
        _point("AUG", aug, 800.0, None),
    )
    boq = (
        _point("JAN", jan, 100.0, 80.0),
        _point("APR", apr, 400.0, 300.0),
        _point("AUG", aug, 800.0, None),
    )
    return EarnedValueResult(
        cutoff_date=apr,
        project_bac=1000.0,
        project_points=project,
        activities=(ActivityEarnedValue("A1", "Activity 1", "1.1", 1000.0, activity),),
        boq_items=(BOQEarnedValue("B1", "BOQ-001", "Concrete", 1000.0, boq),),
    )


def _live_workbook() -> Workbook:
    workbook = Workbook()
    main = workbook.active
    main.title = "main"
    main.append(["Activity Data"])
    main.append([])
    main.append([])
    headers = [
        "Row Type", "WBS", "Description", "P/A", "Activity ID",
        "Task ID", "UID", "Outline Level", "Plan Start", "Plan Finish",
        "Actual Start", "Actual Finish", "% Complete", "Physical %",
        "Amount", "Total Float (hr)", "XML Amount",
        datetime(2026, 1, 30), datetime(2026, 4, 24), datetime(2026, 8, 28),
    ]
    main.append(headers)
    main.append(["Project Summary", None, "Project", "P"] + [None] * 13 + [0.1, 0.3, 0.4])
    main.append([None, None, None, "A"] + [None] * 13 + [0.08, 0.22, None])
    main.append(["WBS", "1.1", "Structure", "P"])
    main.append([None, None, None, "A"])
    main.append(["Activity", "1.1.1", "Activity 1", "P", "A1"] + [None] * 12 + [0.1, 0.3, 0.4])
    main.append([None, None, None, "A", "A1"] + [None] * 12 + [0.08, 0.22, None])

    dashboard = workbook.create_sheet("Dashboard")
    dashboard["K5"] = datetime(2026, 4, 24)

    dashboard_data = workbook.create_sheet("Dashboard_Data")
    dashboard_data["K1"] = "Monthly Cutoff"
    dashboard_data["K2"] = datetime(2026, 1, 30)
    dashboard_data["K3"] = datetime(2026, 4, 24)
    dashboard_data["K4"] = datetime(2026, 8, 28)

    mapping = workbook.create_sheet("BOQ Activity Mapping")
    mapping.append([
        "Activity ID", "BOQ Key", "Source Sheet", "Source Row", "WBS-2",
        "WBS-3", "WBS-4", "BOQ Description", "BOQ Amount", "Share %",
        "Allocated Amount", "Mapping ID", "BOQ ID",
    ])
    mapping.append([
        "A1", "B1", "Project", 1, "W2", "W3", "W4", "Concrete",
        1000.0, 1.0, 1000.0, "MAP-1", "BOQ-001",
    ])
    return workbook


@pytest.mark.unit
def test_ev_l3_dataset_reads_plan_actual_live_from_main() -> None:
    workbook = _live_workbook()
    render_earned_value_sheet(workbook, _result(), include_chart=False)

    data = workbook[EV_DATA_SHEET]
    assert data["G4"].value == "=Dashboard!$K$5"

    # Project chart values are formulas over current main progress, not EV-1 values.
    assert "SUMIFS(main!" in data["B2"].value
    assert "SUMIFS(main!" in data["C2"].value
    assert "MIN(A2,$G$4)" in data["C2"].value

    # BA:BI is one row per Activity and derives selected-date Plan/Actual from main.
    assert data["BA2"].value == "='Earned Value'!$M$3"
    assert data["BB2"].value == "A1"
    assert "SUMIFS(main!" in data["BF2"].value
    assert "SUMIFS(main!" in data["BG2"].value
    assert "MIN($BA2,$G$4)" in data["BG2"].value
    assert data["BH2"].value == "=BE2*BF2"
    assert data["BI2"].value == "=BE2*BG2"


@pytest.mark.unit
def test_ev_l3_wbs_boq_and_table_use_live_selected_date_layer() -> None:
    workbook = _live_workbook()
    render_earned_value_sheet(workbook, _result(), include_chart=False)

    data = workbook[EV_DATA_SHEET]
    ev = workbook[EARNED_VALUE_SHEET]
    table = workbook[EV_TABLE_SHEET]

    # WBS interface remains compatible with the dashboard but is no longer a
    # Python date snapshot: it follows M3 and aggregates live Activity formulas.
    assert data["K2"].value == "='Earned Value'!$M$3"
    assert "SUMIFS(" in data["P2"].value
    assert "SUMIFS(" in data["Q2"].value
    assert "TEXT(K2" in data["J2"].value

    # AE:AK is one live row per BOQ; mapping BAC is static while PV/EV are live.
    assert data["AE2"].value == "='Earned Value'!$M$3"
    assert data["AF2"].value == "B1"
    assert "SUMIFS(" in data["AG2"].value
    assert "SUMIFS(" in data["AH2"].value
    assert data["AI2"].value == "=AH2-AG2"

    # Existing visible views remain wired to the stable EV_Data interfaces.
    assert 'TEXT($M$3,"yyyymmdd")' in ev["A32"].value
    assert table["B3"].value == "='Earned Value'!$M$3"
    assert "EV_Data!$AG$2:$AG$" in table["G6"].value
    assert "EV_Data!$AH$2:$AH$" in table["H6"].value


@pytest.mark.unit
def test_ev_l3_future_view_freezes_actual_at_reporting_cutoff() -> None:
    workbook = _live_workbook()
    render_earned_value_sheet(workbook, _result(), include_chart=False)

    data = workbook[EV_DATA_SHEET]
    ev = workbook[EARNED_VALUE_SHEET]
    ev["M3"] = datetime(2026, 8, 28)

    # Excel evaluates these formulas. The contract is encoded by MIN(M3, G4):
    # Plan reads through M3; Actual freezes at Dashboard!K5 when M3 is later.
    assert "<=\"&$BA2" in data["BF2"].value
    assert "MIN($BA2,$G$4)" in data["BG2"].value
    assert data["G4"].value == "=Dashboard!$K$5"

@pytest.mark.unit
def test_ev_l4_all_visible_views_share_live_m3_dataset() -> None:
    workbook = _live_workbook()
    render_earned_value_sheet(workbook, _result(), include_chart=False)

    data = workbook[EV_DATA_SHEET]
    ev = workbook[EARNED_VALUE_SHEET]
    table = workbook[EV_TABLE_SHEET]

    # KPI cards read the live chart interface selected by M3.
    assert "EV_Data!$B$2:$B$" in ev["D6"].value
    assert "$M$3" in ev["D6"].value
    assert "EV_Data!$D$2:$D$" in ev["G6"].value
    assert "$M$3" in ev["G6"].value
    assert ev["J6"].value == "=G6-D6"
    assert ev["M6"].value == '=IF(D6=0,0,G6/D6)'

    # WBS dashboard cells still use the stable J:S interface, whose values are
    # formulas over the live Activity layer rather than Python date snapshots.
    assert 'TEXT($M$3,"yyyymmdd")' in ev["A32"].value
    assert "SUMIFS(" in data["P2"].value
    assert "SUMIFS(" in data["Q2"].value

    # Top Negative is ranked from the live BOQ SV/SPI layer and selected by M3.
    assert data["T2"].value == '=TEXT(U2,"yyyymmdd")&"|"&V2'
    assert "INDEX($AI$2:$AI$" in data["Y2"].value
    assert "INDEX($AJ$2:$AJ$" in data["Z2"].value
    assert 'TEXT($M$3,"yyyymmdd")' in ev["I32"].value

    # EV Table shares the same selected-date BOQ interface.
    assert table["B3"].value == "='Earned Value'!$M$3"
    assert "EV_Data!$AG$2:$AG$" in table["G6"].value
    assert "EV_Data!$AH$2:$AH$" in table["H6"].value
    assert table["I6"].value == "=H6-G6"
    assert table["J6"].value == "=IF(G6=0,0,H6/G6)"


@pytest.mark.unit
def test_ev_l4_status_date_is_presented_as_view_state_not_cutoff() -> None:
    workbook = _live_workbook()
    render_earned_value_sheet(workbook, _result(), include_chart=False)

    ev = workbook[EARNED_VALUE_SHEET]
    validation = next(iter(ev.data_validations.dataValidation))

    assert validation.prompt == "Choose the Earned Value view date."
    assert "view selector only" in ev["M3"].comment.text
    assert "without changing the project reporting cutoff" in ev["M3"].comment.text
    assert workbook[EV_DATA_SHEET]["G4"].value == "=Dashboard!$K$5"
