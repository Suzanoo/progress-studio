
from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

from progress_studio.infrastructure.excel.rebuild_workbook_reader import RebuildWorkbookReader
from progress_studio.services.activity_table_deriver import ActivityTableDeriver
from progress_studio.services.progress_cache_deriver import ProgressCacheDeriver
from progress_studio.services.rebuild_service import WorkbookRebuildEngine


def _fixture(path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "main"
    headers = [
        "Row Type", "WBS", "Description", "P/A", "% Complete",
        "Amount", "Activity ID", "Outline Level", "Plan Start", "Plan Finish",
    ]
    for col, value in enumerate(headers, 1):
        ws.cell(4, col, value)

    dates = [
        datetime(2026,1,30),
        datetime(2026,2,6),
        datetime(2026,2,13),
        datetime(2026,2,27),
    ]
    for col, dt in enumerate(dates, start=11):
        ws.cell(3, col, f"W{col-10}")
        ws.cell(4, col, dt)

    # Actual row deliberately keeps Row Type/Description blank, matching legacy main grammar.
    ws.append(["Activity","1.1","Concrete","P","",1000,"A1000",2,datetime(2026,1,1),datetime(2026,2,27),.10,.20,.30,.40])
    ws.append(["","","","A","",0,"A1000",2,datetime(2026,1,1),datetime(2026,2,27),.05,.10,.15,.20])

    pin = wb.create_sheet("Payment Input"); pin["A1"] = "KEEP"
    payment = wb.create_sheet("Payment"); payment["A1"] = "KEEP_PAYMENT"
    wb.save(path)
    return path


def test_lw8_legacy_blank_actual_row_pairs_correctly(tmp_path: Path) -> None:
    dataset = RebuildWorkbookReader().read_main_dataset(_fixture(tmp_path / "p.xlsx"))
    activity = ActivityTableDeriver().derive(dataset)
    assert activity.rows[-1].type_label == "Actual"
    assert abs(activity.rows[-1].progress - 0.50) < 1e-12

    cache = ProgressCacheDeriver().derive(dataset)
    assert abs(cache.points[-1].actual_cumulative - 0.50) < 1e-12


def test_lw8_dashboard_has_weekly_monthly_and_dynamic_cutoff_lists(tmp_path: Path) -> None:
    source = _fixture(tmp_path / "source.xlsx")
    output = tmp_path / "live.xlsx"
    WorkbookRebuildEngine().rebuild_live_progress(source, output)

    wb = load_workbook(output, data_only=False)
    try:
        dash = wb["Dashboard"]
        data = wb["Dashboard_Data"]
        assert dash["G5"].value == "Weekly"
        assert dash["K5"].value is not None
        formulas = [dv.formula1 for dv in dash.data_validations.dataValidation]
        assert '"Weekly,Monthly"' in formulas
        assert any("INDIRECT" in str(value) and "Dashboard_Data" in str(value) for value in formulas)

        assert [data.cell(1,c).value for c in range(1,12)] == [
            "Weekly Date","Weekly Plan","Weekly Actual",
            "Monthly Date","Monthly Plan","Monthly Actual",
            "Selected Date","Selected Plan","Selected Actual",
            "Weekly Cutoff","Monthly Cutoff",
        ]
        # Four weekly reporting points and two month-end points.
        assert sum(1 for r in range(2,data.max_row+1) if data.cell(r,10).value) == 4
        assert sum(1 for r in range(2,data.max_row+1) if data.cell(r,11).value) == 2
    finally:
        wb.close()


def test_lw8_activity_table_links_directly_to_main_without_progress_table(tmp_path: Path) -> None:
    source = _fixture(tmp_path / "source.xlsx")
    output = tmp_path / "live.xlsx"
    WorkbookRebuildEngine().rebuild_live_progress(source, output)

    wb = load_workbook(output, data_only=False)
    try:
        assert "progress_table" not in wb.sheetnames
        dash = wb["Dashboard"]
        # First Plan/Actual pair starts at 39/40.
        assert isinstance(dash["L39"].value, str) and "SUMIFS(main!" in dash["L39"].value
        assert isinstance(dash["L40"].value, str) and "SUMIFS(main!" in dash["L40"].value
        assert dash["N40"].value == '=IFERROR(L40-L39,0)'
        assert "Behind" in dash["P40"].value
    finally:
        wb.close()


def test_lw8_live_workbook_uses_manual_calculation_with_calc_on_save(tmp_path: Path) -> None:
    source = _fixture(tmp_path / "source.xlsx")
    output = tmp_path / "live.xlsx"
    WorkbookRebuildEngine().rebuild_live_progress(source, output)

    wb = load_workbook(output, data_only=False)
    try:
        calc = wb.calculation
        assert calc.calcMode == "manual"
        assert calc.calcOnSave is True
        assert calc.fullCalcOnLoad is False
        assert calc.forceFullCalc is False
        # Monthly remains a cache: no weekly formulas in its timescale.
        monthly = wb["main_monthly"]
        assert not isinstance(monthly["K5"].value, str)
        assert not isinstance(monthly["L5"].value, str)
    finally:
        wb.close()


def test_lw8_formula_payload_is_linear_not_activity_times_period_matrix(tmp_path: Path) -> None:
    source = _fixture(tmp_path / "source.xlsx")
    output = tmp_path / "live.xlsx"
    WorkbookRebuildEngine().rebuild_live_progress(source, output)

    wb = load_workbook(output, data_only=False)
    try:
        formula_cells = 0
        for sheet_name in ("Dashboard", "Dashboard_Data"):
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        formula_cells += 1
        # One activity + tiny selector cache should stay well below an Nxperiod matrix.
        assert formula_cells < 40
    finally:
        wb.close()
