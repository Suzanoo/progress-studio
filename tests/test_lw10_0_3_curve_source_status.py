
from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

from progress_studio.infrastructure.excel.live_scurve_workbook import (
    apply_weekly_scurve_cutoff_contract,
)
from progress_studio.infrastructure.excel.rebuild_workbook_reader import RebuildWorkbookReader
from progress_studio.services.activity_table_deriver import _status


def _fixture(path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "main"
    headers = [
        "Row Type", "WBS", "Description", "P/A", "Activity ID",
        "Outline Level", "Plan Start", "Plan Finish", "Amount",
    ]
    for col, value in enumerate(headers, 1):
        ws.cell(4, col, value)
    dates = [
        datetime(2026, 7, 10), datetime(2026, 7, 17),
        datetime(2026, 7, 24), datetime(2026, 7, 31),
    ]
    for col, dt in enumerate(dates, start=10):
        ws.cell(3, col, f"W{col-9}")
        ws.cell(4, col, dt)

    ws.append(["Project Summary","","Project","P","",0,None,None,1000,.20,.30,.30,.20])
    ws.append(["Project Summary","","Project","A","",0,None,None,0,.10,.05,None,None])
    ws.append(["S-Curve","","Plan","P","",0,None,None,None,.20,.30,.30,.20])
    ws.append(["S-Curve","","Acc. Plan","AP","",0,None,None,None,.20,.50,.80,1.00])
    ws.append(["S-Curve","","Actual","A","",0,None,None,None,.10,.05,None,None])
    ws.append(["S-Curve","","Acc. Actual","AA","",0,None,None,None,.10,.15,None,None])
    wb.save(path)
    return path


def test_weekly_scurve_actual_is_source_cutoff_aware(tmp_path: Path) -> None:
    path = _fixture(tmp_path / "p.xlsx")
    dataset = RebuildWorkbookReader().read_main_dataset(path)
    wb = load_workbook(path)
    wb.create_sheet("Dashboard")["K5"] = datetime(2026, 7, 24)
    ws = wb["main"]

    assert apply_weekly_scurve_cutoff_contract(wb, dataset)
    assert "Dashboard!$K$5" in ws["J9"].value
    assert "Dashboard!$K$5" in ws["M9"].value
    assert "Dashboard!$K$5" in ws["J10"].value
    assert "Dashboard!$K$5" in ws["M10"].value
    assert "Dashboard!$K$5" not in ws["M7"].value
    assert "Dashboard!$K$5" not in ws["M8"].value
    wb.close()


def test_status_semantics_distinguish_not_due_and_no_progress() -> None:
    assert _status(0.0, 0.0) == "Not Due"
    assert _status(0.25, 0.0) == "No Progress"
    assert _status(0.50, 0.25) == "Behind"
    assert _status(0.50, 0.50) == "On Track"
    assert _status(1.00, 1.00) == "Complete"


def test_monthly_scurve_actual_formula_is_cutoff_aware() -> None:
    source = Path(
        "progress_studio/infrastructure/excel/live_monthly_workbook.py"
    ).read_text(encoding="utf-8")
    assert 'pa == "AA"' in source
    assert 'pa == "A"' in source
    assert "Dashboard!$K$5" in source
    assert "monthly_date_ref" in source


def test_dashboard_chart_is_adapter_not_second_scurve_deriver() -> None:
    source = Path(
        "progress_studio/infrastructure/excel/live_dashboard_workbook.py"
    ).read_text(encoding="utf-8")
    block = source[source.index("def _build_live_data_sheet"):source.index("def _kpi_box")]
    assert "_find_scurve_rows(main, dataset)" in block
    assert "_find_scurve_rows(monthly_ws, dataset)" in block
    assert "Selected Actual stops at the chosen cutoff" in block
