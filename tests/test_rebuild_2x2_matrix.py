from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path
from zipfile import ZipFile

import pytest
from openpyxl import load_workbook

from progress_studio.services.rebuild_service import WorkbookRebuildEngine
from tests.test_ms_rb2_progress_rebuild_engine import _full_rebuild_fixture


def _matrix_source(path: Path) -> Path:
    """Build one valid source for all four Rebuild workspace combinations."""
    base = _full_rebuild_fixture(path.with_name("matrix_base.xlsx"))
    wb = load_workbook(base)
    try:
        pin = wb["Payment Input"]
        for row in pin.iter_rows():
            for cell in row:
                cell.value = None
        pin["A6"] = "Type"
        pin["B6"] = "WBS"
        pin["C6"] = "Activity ID"
        pin["D6"] = "Activity Name"
        pin["E6"] = "P1"
        pin["E7"] = datetime(2026, 3, 13)
        pin["A8"] = "ACT"
        pin["B8"] = "1.1"
        pin["C8"] = "A1000"
        pin["D8"] = "Concrete"
        pin["E8"] = 0.50
        wb.save(base)
    finally:
        wb.close()

    # Start the matrix from a workbook that already has the full traditional
    # overlay presentation. Payment-only modes must preserve it byte-semantically.
    WorkbookRebuildEngine().rebuild_live_progress(base, path, project_name="2x2 Matrix")
    return path


def _plot_area_nofill_count(path: Path) -> int:
    pattern = re.compile(r"<plotArea>.*?<spPr>\s*<a:noFill", re.S)
    with ZipFile(path) as zf:
        chart_xmls = [name for name in zf.namelist() if name.startswith("xl/charts/chart")]
        return sum(bool(pattern.search(zf.read(name).decode("utf-8"))) for name in chart_xmls)


def _overlay_signature(path: Path) -> tuple:
    wb = load_workbook(path, data_only=False)
    try:
        result = []
        for sheet_name in ("main", "main_monthly"):
            ws = wb[sheet_name]
            assert len(ws._charts) == 1
            chart = ws._charts[0]
            result.append(
                (
                    sheet_name,
                    chart.series[0].val.numRef.f,
                    chart.series[1].val.numRef.f,
                    chart.anchor._from.col,
                    chart.anchor.to.col,
                    chart.anchor.editAs,
                )
            )
        return tuple(result)
    finally:
        wb.close()


def _assert_final_policy(path: Path) -> None:
    wb = load_workbook(path, data_only=False)
    try:
        assert wb.calculation.calcMode == "manual"
        assert wb.calculation.calcOnSave is True
        assert wb["main"].protection.sheet is True
        assert wb["main_monthly"].protection.sheet is True
        assert wb["Dashboard"].protection.sheet is True
    finally:
        wb.close()


def _assert_overlay_series_contract(path: Path) -> None:
    wb = load_workbook(path, data_only=False)
    try:
        data = wb["Dashboard_Data"]
        for sheet_name, plan_col, actual_col in (
            ("main", 21, 22),
            ("main_monthly", 25, 26),
        ):
            chart = wb[sheet_name]._charts[0]
            assert chart.anchor.editAs == "twoCell"
            plan_formula = chart.series[0].val.numRef.f
            actual_formula = chart.series[1].val.numRef.f
            plan_rows = [int(v) for v in re.findall(r"\$(\d+)", plan_formula)]
            actual_rows = [int(v) for v in re.findall(r"\$(\d+)", actual_formula)]
            assert plan_rows[0] == actual_rows[0] == 2
            assert plan_rows[-1] == actual_rows[-1]
            last_row = plan_rows[-1]
            assert data.cell(2, plan_col).value == 0
            assert data.cell(2, actual_col).value == 0
            # The explicit zero belongs only at the left anchor. The series must
            # terminate at its real project source, with no trailing margin zeros.
            assert data.cell(last_row + 1, plan_col).value is None
            assert data.cell(last_row + 1, actual_col).value is None
            assert data.cell(last_row, plan_col).value not in (0, "=0")
    finally:
        wb.close()


@pytest.mark.parametrize("output_mode", ["snapshot", "live"])
def test_rebuild_2x2_progress_modes_rebuild_only_progress_and_keep_overlay_contract(
    tmp_path: Path, output_mode: str
) -> None:
    source = _matrix_source(tmp_path / "source.xlsx")
    output = tmp_path / f"progress_{output_mode}.xlsx"
    engine = WorkbookRebuildEngine()

    if output_mode == "snapshot":
        engine.rebuild_progress(source, output, project_name="2x2 Matrix")
    else:
        engine.rebuild_live_progress(source, output, project_name="2x2 Matrix")

    wb = load_workbook(output, data_only=False)
    try:
        assert wb["Payment Input"]["C8"].value == "A1000"
        # Progress scope must not rebuild Payment.
        assert "Payment" in wb.sheetnames
        if output_mode == "snapshot":
            assert "progress_table" in wb.sheetnames
        else:
            assert "progress_table" not in wb.sheetnames
        assert len(wb["main"]._charts) == 1
        assert len(wb["main_monthly"]._charts) == 1
    finally:
        wb.close()

    _assert_final_policy(output)
    _assert_overlay_series_contract(output)
    assert _plot_area_nofill_count(output) >= 2


@pytest.mark.parametrize("output_mode", ["snapshot", "live"])
def test_rebuild_2x2_payment_modes_replace_payment_only_and_preserve_progress_presentation(
    tmp_path: Path, output_mode: str
) -> None:
    source = _matrix_source(tmp_path / "source.xlsx")
    before_overlay = _overlay_signature(source)
    before_transparent = _plot_area_nofill_count(source)
    assert before_transparent >= 2

    output = tmp_path / f"payment_{output_mode}.xlsx"
    engine = WorkbookRebuildEngine()
    if output_mode == "snapshot":
        engine.rebuild_payment(source, output)
    else:
        engine.rebuild_live_payment(source, output)

    wb = load_workbook(output, data_only=False)
    try:
        assert "Payment" in wb.sheetnames
        assert wb["Payment Input"]["C8"].value == "A1000"
        # Payment scope must preserve all Progress-owned views.
        for name in ("main", "main_monthly", "progress", "Dashboard_Data", "Dashboard"):
            assert name in wb.sheetnames
    finally:
        wb.close()

    _assert_final_policy(output)
    assert _overlay_signature(output) == before_overlay
    assert _plot_area_nofill_count(output) >= 2


def test_rebuild_workspace_routes_all_four_2x2_combinations_explicitly() -> None:
    source = Path("progress_studio/presentation/gui/rebuild.py").read_text(encoding="utf-8")
    worker = source[source.index("    def _worker_rebuild("):source.index("    def _done(")]
    assert 'if mode is RebuildMode.PROGRESS:' in worker
    assert 'if output_mode == "live":' in worker
    assert 'self.engine.rebuild_progress(' in worker
    assert 'self.engine.rebuild_live_progress(' in worker
    assert 'self.engine.rebuild_payment(' in worker
    assert 'self.engine.rebuild_live_payment(' in worker


def test_rebuild_2x2_finalization_ownership_has_no_duplicate_policy_pass() -> None:
    service = Path("progress_studio/services/rebuild_service.py").read_text(encoding="utf-8")
    payment_renderer = Path(
        "progress_studio/infrastructure/excel/payment_line_renderer.py"
    ).read_text(encoding="utf-8")

    progress = service[service.index("    def rebuild_progress("):service.index("    def rebuild_live_progress(")]
    live_progress = service[service.index("    def rebuild_live_progress("):service.index("    def rebuild_live_payment(")]
    live_payment = service[service.index("    def rebuild_live_payment("):service.index("    def rebuild_payment(")]
    snapshot_payment = service[service.index("    def rebuild_payment("):service.index("    def generated_sheets_for(")]
    standalone_payment = payment_renderer[
        payment_renderer.index("    def render_periods("):
        payment_renderer.index("    def render_periods_into_workbook(")
    ]

    assert progress.count("finalize_workbook(") == 1
    assert progress.count("wb.save(") == 1
    assert live_progress.count("finalize_workbook(") == 1
    assert live_progress.count("wb.save(") == 1
    assert live_payment.count("finalize_workbook(") == 1
    assert live_payment.count("wb.save(") == 1

    # Snapshot Payment delegates the workflow boundary to the standalone Payment
    # renderer and therefore must not finalize a second time in Rebuild.
    assert "render_payment_backbones" in snapshot_payment
    assert "finalize_workbook(" not in snapshot_payment
    assert standalone_payment.count("finalize_workbook(") == 1
    assert standalone_payment.count("wb.save(") == 1


def test_rebuild_live_dashboard_masks_margin_plan_as_na_instead_of_zero(tmp_path: Path) -> None:
    source = _matrix_source(tmp_path / "source.xlsx")
    output = tmp_path / "live_progress_margin.xlsx"
    WorkbookRebuildEngine().rebuild_live_progress(source, output, project_name="2x2 Matrix")

    wb = load_workbook(output, data_only=False)
    try:
        data = wb["Dashboard_Data"]
        # Selected Plan must explicitly mask blank Weekly/Monthly source cells.
        formula = str(data["H2"].value)
        assert 'IF(B2="",NA(),B2)' in formula
        assert 'IF(E2="",NA(),E2)' in formula
        assert wb["Dashboard"]._charts[0].x_axis.tagname == "dateAx"
    finally:
        wb.close()


def test_rebuild_progress_monthly_overlay_finishes_on_last_nonblank_monthly_plan(tmp_path: Path) -> None:
    source = _matrix_source(tmp_path / "source.xlsx")
    for mode in ("snapshot", "live"):
        output = tmp_path / f"monthly_end_{mode}.xlsx"
        engine = WorkbookRebuildEngine()
        if mode == "snapshot":
            engine.rebuild_progress(source, output, project_name="2x2 Matrix")
        else:
            engine.rebuild_live_progress(source, output, project_name="2x2 Matrix")
        wb = load_workbook(output, data_only=False)
        try:
            data = wb["Dashboard_Data"]
            chart = wb["main_monthly"]._charts[0]
            rows = [int(v) for v in re.findall(r"\$(\d+)", chart.series[0].val.numRef.f)]
            last_row = rows[-1]
            # Explicit anchor is row 2; the last chart point must be a real source
            # formula, never a synthetic/trailing zero.
            assert data.cell(last_row, 25).value not in (0, "=0", None)
            assert data.cell(last_row + 1, 25).value is None
        finally:
            wb.close()
