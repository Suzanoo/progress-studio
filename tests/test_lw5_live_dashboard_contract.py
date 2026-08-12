
from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook

from progress_studio.infrastructure.excel.live_dashboard_workbook import build_live_dashboard
from progress_studio.infrastructure.excel.rebuild_workbook_reader import RebuildWorkbookReader


def _fixture(path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "main"
    headers = [
        "Row Type", "WBS", "Description", "P/A", "% Complete",
        "Amount", "Activity ID", "Outline Level", "Plan Start", "Plan Finish",
    ]
    for col, value in enumerate(headers, 1):
        ws.cell(4, col, value)
    for col, dt in enumerate((datetime(2026,3,6), datetime(2026,3,13)), start=11):
        ws.cell(3, col, f"W{col-10}")
        ws.cell(4, col, dt)
    rows = [
        ["Activity", "1.1", "Concrete", "P", "", 1000, "A1000", 2, datetime(2026,3,1), datetime(2026,3,13), .5, .5],
        ["Activity", "1.1", "Concrete", "A", "", 0, "A1000", 2, datetime(2026,3,1), datetime(2026,3,13), .2, .3],
    ]
    for row in rows:
        ws.append(row)
    wb.save(path)
    return path


def test_lw5_live_dashboard_renders_with_progress_contract_without_progress_table(tmp_path: Path) -> None:
    source = _fixture(tmp_path / "source.xlsx")
    dataset = RebuildWorkbookReader().read_main_dataset(source)

    wb = Workbook()
    wb.active.title = "main"
    build_live_dashboard(wb, dataset, project_name="Demo")

    assert "Dashboard" in wb.sheetnames
    assert "Dashboard_Data" in wb.sheetnames
    assert "progress" in wb.sheetnames
    assert "progress_table" not in wb.sheetnames
    assert wb["Dashboard"]["C6"].value.startswith("Live curve: progress")
    assert wb["Dashboard"]["C39"].value == "Concrete"
    assert wb["Dashboard"]["F39"].value == "Plan"
    assert wb["Dashboard"]["F40"].value == "Actual"


def test_lw5_dashboard_data_is_tiny_period_level_cache(tmp_path: Path) -> None:
    dataset = RebuildWorkbookReader().read_main_dataset(_fixture(tmp_path / "source.xlsx"))
    wb = Workbook()
    wb.active.title = "main"
    build_live_dashboard(wb, dataset)

    data = wb["Dashboard_Data"]
    assert data.max_row == 3
    # LW-8 evolves the tiny cache to include Weekly/Monthly selector columns.
    assert data.max_column == 11
    assert [data.cell(1, c).value for c in range(1, 4)] == [
        "Weekly Date", "Weekly Plan", "Weekly Actual"
    ]
    assert data.sheet_state == "hidden"


def test_lw5_live_dashboard_boundary_uses_openpyxl_only_for_rendering() -> None:
    service_sources = [
        Path("progress_studio/services/activity_table_deriver.py").read_text(encoding="utf-8"),
        Path("progress_studio/services/progress_cache_deriver.py").read_text(encoding="utf-8"),
    ]
    assert all("openpyxl" not in source for source in service_sources)

    renderer = Path(
        "progress_studio/infrastructure/excel/live_dashboard_workbook.py"
    ).read_text(encoding="utf-8")
    assert "openpyxl" in renderer
    assert 'workbook["progress_table"]' not in renderer
    assert 'workbook["progress"]' in renderer
