import contextlib
import io
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from progress_studio import __version__
from progress_studio.app import build_application
from progress_studio.app.context import PipelineContext
from progress_studio.infrastructure.excel import ImportWorkbookWriter
from progress_studio.infrastructure.primavera import PrimaveraXmlReader
from progress_studio.pipeline.amount_step import AmountStep
from progress_studio.pipeline.distribution_step import DistributionStep
from progress_studio.pipeline.import_step import ImportStep
from progress_studio.pipeline.okd_step import OkdStep
from progress_studio.pipeline.progress_step import ProgressStep
from progress_studio.pipeline.schedule_step import ScheduleStep
from progress_studio.pipeline.timescale_step import TimescaleStep
from progress_studio.services.amount_service import AmountService
from progress_studio.services.distribution_service import DistributionService
from progress_studio.services.import_service import ImportService
from progress_studio.services.okd_service import OkdService
from progress_studio.services.progress_service import ProgressService
from progress_studio.services.schedule_service import ScheduleService
from progress_studio.services.schedule_workbook_service import ScheduleWorkbookService
from progress_studio.services.timescale_service import TimescaleService

ROOT = Path(__file__).resolve().parents[1]
XML = ROOT / "example" / "example.xml"


class FixedPrompt:
    def choose_method(self):
        return "flat"

    def review(self, output_file, method):
        return "exit"


def run_example(output_folder: Path) -> PipelineContext:
    context = PipelineContext(
        source_xml=XML,
        cutoff_day="5",
        amount_per_activity=1000.0,
        project_folder=output_folder,
        working_folder=output_folder,
    )
    schedule_service = ScheduleService()
    steps = [
        ImportStep(ImportService(PrimaveraXmlReader(), schedule_service, ImportWorkbookWriter())),
        ScheduleStep(ScheduleWorkbookService()),
        TimescaleStep(TimescaleService()),
        AmountStep(AmountService()),
        ProgressStep(ProgressService()),
        DistributionStep(DistributionService(), FixedPrompt()),
        OkdStep(OkdService()),
    ]
    for step in steps:
        step.execute(context)
    return context


class Ms7ReleaseTests(unittest.TestCase):
    def test_release_version(self):
        self.assertEqual(__version__, "2.3.0")

    def test_cli_help(self):
        stdout = io.StringIO()
        with self.assertRaises(SystemExit) as raised, contextlib.redirect_stdout(stdout):
            build_application().run(["--help"])
        self.assertEqual(raised.exception.code, 0)
        text = stdout.getvalue()
        self.assertIn("Build a progress workbook", text)
        self.assertFalse(any("\u0e00" <= ch <= "\u0e7f" for ch in text))

    def test_invalid_input_returns_error(self):
        stdout = io.StringIO()
        with contextlib.redirect_stdout(stdout):
            code = build_application().run([
                "--input", str(ROOT / "missing.xml"),
                "--cutoff-day", "5",
            ])
        self.assertEqual(code, 1)
        self.assertIn("Input XML file not found", stdout.getvalue())

    def test_example_end_to_end_regression_manifest(self):
        with tempfile.TemporaryDirectory() as td:
            with contextlib.redirect_stdout(io.StringIO()):
                context = run_example(Path(td))
            self.assertIsNotNone(context.output_workbook)
            self.assertTrue(context.output_workbook.is_file())

            wb = load_workbook(context.output_workbook, data_only=False)
            try:
                required = {"main", "Amount Mapping", "Distribution Report", "progress", "progress_table"}
                self.assertTrue(required.issubset(wb.sheetnames))

                main = wb["main"]
                progress = wb["progress"]
                table = wb["progress_table"]
                report = wb["Distribution Report"]

                self.assertEqual(main.max_row, 519)
                self.assertEqual(progress.max_row, 77)
                self.assertEqual(table.max_row, 511)
                self.assertEqual(report.max_row, 179)
                self.assertEqual(context.metadata["progress"].weekly_columns, 76)
                self.assertEqual(context.metadata["activity_count"], 172)
                self.assertEqual(context.metadata["wbs_count"], 82)
                self.assertEqual(context.metadata["okd"].table_rows, 510)
                self.assertEqual(context.metadata["okd"].checked_links, 40290)

                formulas = sum(
                    1
                    for row in table.iter_rows()
                    for cell in row
                    if isinstance(cell.value, str) and cell.value.startswith("=")
                )
                self.assertGreaterEqual(formulas, 40000)
            finally:
                wb.close()

    def test_release_documents_are_current(self):
        readme = (ROOT / "README.md").read_text(encoding="utf-8")
        roadmap = (ROOT / "README_ROADMAP.md").read_text(encoding="utf-8")
        self.assertIn("2.1.3", readme)
        self.assertIn("MS-7 — Mapping Workspace UX", roadmap)
        self.assertNotIn("legacy/", readme)
        self.assertNotIn("Scripts 05-07 remain", readme)

    def test_no_thai_in_application_or_release_docs(self):
        paths = list((ROOT / "progress_studio").rglob("*.py"))
        paths.extend([ROOT / "README.md", ROOT / "README_ROADMAP.md", ROOT / "docs/acceptance/MS6_ACCEPTANCE.md"])
        for path in paths:
            text = path.read_text(encoding="utf-8")
            self.assertFalse(any("\u0e00" <= ch <= "\u0e7f" for ch in text), str(path))


if __name__ == "__main__":
    unittest.main()
