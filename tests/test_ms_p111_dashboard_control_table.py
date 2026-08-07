from datetime import date
from pathlib import Path

from openpyxl import Workbook

from progress_studio.infrastructure.excel.dashboard_workbook import build_dashboard


def _workbook():
    wb = Workbook()
    progress = wb.active
    progress.title = "progress"
    progress.append(["project_start", "project_finish", "week_start", "plan", "actual"])
    progress.append([date(2026, 1, 1), date(2026, 2, 28), date(2026, 1, 2), 10, 5])
    progress.append([date(2026, 1, 1), date(2026, 2, 28), date(2026, 1, 9), 25, 12])
    table = wb.create_sheet("progress_table")
    table.append(["WBS", "Activities", "Amount", "P/A", "%Progress", date(2026, 1, 2), date(2026, 1, 9)])
    table.append(["1.1", "Activity A", 1000, "P", 25, 10, 15])
    table.append(["1.1", "Activity A", 1000, "A", 12, 5, 7])
    return wb


def test_dashboard_control_table_is_one_row_per_activity():
    wb = _workbook()
    build_dashboard(wb)
    ws = wb["Dashboard"]
    assert ws["B38"].value == "WBS"
    assert ws["H38"].value == "Plan %"
    assert ws["I38"].value == "Actual %"
    assert ws["J38"].value == "Variance"
    assert ws["K38"].value == "Gap Amount"
    assert ws["M38"].value is None
    assert ws["B39"].value == "='progress_table'!A2"
    assert ws["B40"].value is None


def test_pipeline_steps_do_not_hardcode_obsolete_step_counts():
    root = Path(__file__).resolve().parents[1] / "progress_studio" / "pipeline"
    text = "\n".join(path.read_text(encoding="utf-8") for path in root.glob("*_step.py"))
    assert "/8]" not in text
    assert "/7]" not in text


def test_monthly_cutoff_uses_last_real_reporting_date_and_dynamic_dropdown():
    wb = Workbook()
    progress = wb.active
    progress.title = "progress"
    progress.append(["project_start", "project_finish", "week_start", "plan", "actual"])
    progress.append([date(2026, 8, 1), date(2026, 9, 30), date(2026, 8, 7), 10, 5])
    progress.append([date(2026, 8, 1), date(2026, 9, 30), date(2026, 8, 28), 30, 20])
    progress.append([date(2026, 8, 1), date(2026, 9, 30), date(2026, 9, 4), 40, 25])
    progress.append([date(2026, 8, 1), date(2026, 9, 30), date(2026, 9, 25), 70, 50])
    table = wb.create_sheet("progress_table")
    table.append(["WBS", "Activities", "Amount", "P/A", "%Progress", date(2026, 8, 7)])
    table.append(["1.1", "Activity A", 1000, "P", 30, 10])
    table.append(["1.1", "Activity A", 1000, "A", 20, 5])

    build_dashboard(wb)
    data = wb["Dashboard_Data"]
    ws = wb["Dashboard"]

    assert data["K2"].value == date(2026, 8, 28)
    assert data["K3"].value == date(2026, 9, 25)
    assert ws["K5"].value == date(2026, 9, 25)
    cutoff_validations = [v for v in ws.data_validations.dataValidation if "INDIRECT" in str(v.formula1)]
    assert len(cutoff_validations) == 1
    assert '$G$5="Weekly"' in cutoff_validations[0].formula1
    assert "Dashboard_Data!$J$2:$J$5" in cutoff_validations[0].formula1
    assert "Dashboard_Data!$K$2:$K$3" in cutoff_validations[0].formula1


def test_activity_exception_table_only_shows_wbs_filter_button():
    wb = _workbook()
    build_dashboard(wb)
    ws = wb["Dashboard"]

    assert ws.auto_filter.ref.startswith("B38:M")
    hidden_button_columns = {fc.colId for fc in ws.auto_filter.filterColumn if fc.showButton is False}
    assert hidden_button_columns == set(range(1, 12))


def test_activity_exception_table_is_not_capped_at_eight_rows():
    wb = Workbook()
    progress = wb.active
    progress.title = "progress"
    progress.append(["project_start", "project_finish", "week_start", "plan", "actual"])
    progress.append([date(2026, 1, 1), date(2026, 2, 28), date(2026, 1, 2), 10, 5])
    table = wb.create_sheet("progress_table")
    table.append(["WBS", "Activities", "Amount", "P/A", "%Progress", date(2026, 1, 2)])
    for index in range(12):
        table.append([f"1.{index+1}", f"Activity {index+1}", 1000, "P", 10, 10])
        table.append([f"1.{index+1}", f"Activity {index+1}", 1000, "A", 5, 5])

    build_dashboard(wb)
    ws = wb["Dashboard"]

    assert ws["B50"].value == "='progress_table'!A24"
    assert ws.auto_filter.ref == "B38:M50"


def test_activity_progress_uses_native_outline_levels_like_main_sheet():
    wb = Workbook()
    progress = wb.active
    progress.title = "progress"
    progress.append(["project_start", "project_finish", "week_start", "plan", "actual"])
    progress.append([date(2026, 1, 1), date(2026, 2, 28), date(2026, 1, 2), 10, 5])

    table = wb.create_sheet("progress_table")
    table.append(["WBS", "Activities", "Amount", "P/A", "%Progress", date(2026, 1, 2), "_Kind"])
    rows = [
        ["PROJECT", "Project", 1000, "P", 10, 10, "project"],
        ["PROJECT", "Project", 1000, "A", 5, 5, "project"],
        ["1", "WBS 1", 1000, "P", 10, 10, "wbs"],
        ["1", "WBS 1", 1000, "A", 5, 5, "wbs"],
        ["1.1", "WBS 1.1", 1000, "P", 10, 10, "wbs"],
        ["1.1", "WBS 1.1", 1000, "A", 5, 5, "wbs"],
        ["1.1.1", "Activity A", 1000, "P", 10, 10, "activity"],
        ["1.1.1", "Activity A", 1000, "A", 5, 5, "activity"],
    ]
    for row in rows:
        table.append(row)

    build_dashboard(wb)
    ws = wb["Dashboard"]

    assert ws.row_dimensions[39].outlineLevel == 0
    assert ws.row_dimensions[40].outlineLevel == 1
    assert ws.row_dimensions[41].outlineLevel == 2
    assert ws.row_dimensions[42].outlineLevel == 3
    assert ws.sheet_properties.outlinePr.summaryBelow is False
