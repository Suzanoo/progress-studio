from datetime import date

from openpyxl import Workbook

from progress_studio.domain.main_dataset import MainDataset, MainPeriod, MainRow
from progress_studio.infrastructure.excel.traditional_overlay_workbook import (
    _overlay_chart,
    _responsive_anchor,
)


def test_overlay_uses_explicit_zero_anchor_without_new_columns():
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard_Data"
    ws["A2"] = date(2026, 2, 20)
    ws["A3"] = date(2026, 2, 27)
    ws["B2"] = 0.0  # explicit Plan (0,0) chart anchor
    ws["B3"] = 0.0
    ws["P2"] = 0.0  # existing Actual Visible column, not a new helper
    ws["P3"] = 0.0

    chart = _overlay_chart(
        data_ws=ws,
        date_col=1,
        plan_col=2,
        actual_col=16,
        first_row=2,
        last_row=3,
    )

    assert chart.display_blanks == "gap"
    assert chart.series[0].val.numRef.f.endswith("$B$2:$B$3")
    assert chart.series[1].val.numRef.f.endswith("$P$2:$P$3")
    assert ws.max_column == 16


def test_responsive_anchor_still_spans_exact_visible_period_columns():
    anchor = _responsive_anchor(first_col=20, last_col=23, top_row=5, bottom_row=30)
    assert anchor._from.col == 19
    assert anchor.to.col == 23
    # 20..23 = exactly four worksheet columns; there is no trailing fifth period.
    assert anchor.to.col - anchor._from.col == 4
