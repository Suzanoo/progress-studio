from datetime import datetime, date

from openpyxl import Workbook

from progress_studio.domain.main_dataset import MainDataset, MainPeriod, MainRow
from progress_studio.infrastructure.excel.monthly_main_workbook import build_monthly_main_view
from progress_studio.infrastructure.excel.progress_workbook import (
    SCURVE_ACTUAL_FILL,
    SCURVE_PLAN_FILL,
    WBS_ACTUAL_FILL,
    WBS_PLAN_FILL,
)
from progress_studio.infrastructure.excel.traditional_overlay_workbook import (
    _build_explicit_overlay_series_sources,
    _monthly_project_window,
    _weekly_project_window,
    ensure_overlay_visible_actual_columns,
)


def _dataset_with_margin() -> MainDataset:
    periods = (
        MainPeriod(18, "2026-03-20", datetime(2026, 3, 20)),
        MainPeriod(19, "2026-03-27", datetime(2026, 3, 27)),
        MainPeriod(20, "2026-04-03", datetime(2026, 4, 3)),
        MainPeriod(21, "2026-04-10", datetime(2026, 4, 10)),
        MainPeriod(22, "2026-04-17", datetime(2026, 4, 17)),
        MainPeriod(23, "2026-04-24", datetime(2026, 4, 24)),
        MainPeriod(24, "2026-05-01", datetime(2026, 5, 1)),
    )
    summary = MainRow(
        row_number=5,
        row_type="Project Summary",
        pa="P",
        wbs="",
        description="Demo",
        activity_id="",
        outline_level=0,
        plan_start=datetime(2026, 4, 17, 8, 0),
        plan_finish=datetime(2026, 5, 1, 17, 0),
        amount=100.0,
        percent_complete=0.0,
        period_values=(),
    )
    # This activity starts later than the authoritative Project Summary. The
    # overlay must still anchor at Project Start, not the first activity date.
    activity = MainRow(
        row_number=7,
        row_type="Activity",
        pa="P",
        wbs="1",
        description="Later Activity",
        activity_id="A1000",
        outline_level=1,
        plan_start=datetime(2026, 4, 24, 8, 0),
        plan_finish=datetime(2026, 5, 1, 17, 0),
        amount=100.0,
        percent_complete=0.0,
        period_values=(),
    )
    return MainDataset(
        workbook_name="demo.xlsx",
        header_row=4,
        headers=(),
        periods=periods,
        rows=(summary, activity),
    )


def test_weekly_overlay_separates_filtered_helper_rows_from_margin_period_columns():
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard_Data"
    ws.append(["Weekly Date", "Weekly Plan"])
    ws.append([date(2026, 4, 17), 0.01])
    ws.append([date(2026, 4, 24), 0.50])
    ws.append([date(2026, 5, 1), 1.00])

    first_row, last_row, first_col, last_col = _weekly_project_window(ws, _dataset_with_margin())

    # Dashboard_Data is already margin-free, so its source starts at row 2.
    assert (first_row, last_row) == (2, 4)
    # Physical main still contains the visible margin. Project Start is V, not W.
    assert (first_col, last_col) == (22, 24)


def test_monthly_overlay_uses_full_timescale_month_offset_for_physical_anchor():
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard_Data"
    ws["D1"] = "Monthly Date"
    ws["E1"] = "Monthly Plan"
    ws["D2"], ws["E2"] = date(2026, 4, 24), 0.50
    ws["D3"], ws["E3"] = date(2026, 5, 1), 1.00

    first_row, last_row, first_col, last_col = _monthly_project_window(ws, _dataset_with_margin())

    assert (first_row, last_row) == (2, 3)
    # Full visible monthly timescale starts in March at R (18), therefore
    # April/May are S/T (19/20), even though Dashboard_Data starts at April.
    assert (first_col, last_col) == (19, 20)


def test_overlay_actual_helpers_show_zero_then_carry_last_value_until_cutoff():
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard_Data"
    ws.append(["Weekly Date", "Plan", "Weekly Actual", "Monthly Date", "Plan", "Monthly Actual"])
    ws.append([date(2026, 4, 17), 0.01, None, date(2026, 4, 24), 0.02, None])
    ws.append([date(2026, 4, 24), 0.02, 0.01, date(2026, 5, 1), 0.04, 0.02])
    ws.append([date(2026, 5, 1), 0.03, None, date(2026, 5, 29), 0.06, None])

    ensure_overlay_visible_actual_columns(
        wb,
        weekly_cutoff_ref="'main'!$M$10",
        monthly_cutoff_ref="'main_monthly'!$M$10",
    )

    assert "LOOKUP" in ws["P2"].value and ",0))))" in ws["P2"].value
    assert "$C$2:C3" in ws["P3"].value
    assert "$F$2:F3" in ws["Q3"].value
    assert "IF(C3<>\"\",C3" in ws["P3"].value
    assert "IF(F3<>\"\",F3" in ws["Q3"].value



def test_weekly_overlay_uses_literal_cutoff_dates_when_dashboard_weekly_dates_are_formulas():
    """Rebuild regression: weekly Dashboard_Data dates can be formula links."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard_Data"
    ws["A1"] = "Weekly Date"
    ws["B1"] = "Weekly Plan"
    ws["J1"] = "Weekly Cutoff"
    ws["A2"], ws["B2"], ws["J2"] = "=progress!A7", "=progress!B7", date(2026, 4, 17)
    ws["A3"], ws["B3"], ws["J3"] = "=progress!A8", "=progress!B8", date(2026, 4, 24)
    ws["A4"], ws["B4"], ws["J4"] = "=progress!A9", "=progress!B9", date(2026, 5, 1)

    first_row, last_row, first_col, last_col = _weekly_project_window(ws, _dataset_with_margin())

    assert (first_row, last_row) == (2, 4)
    assert (first_col, last_col) == (22, 24)


def test_explicit_weekly_zero_anchor_is_dated_before_first_real_formula_linked_period():
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard_Data"
    for cell, value in (("A1", "Weekly Date"), ("B1", "Weekly Plan"), ("J1", "Weekly Cutoff")):
        ws[cell] = value
    for row, cutoff, plan in (
        (2, date(2026, 4, 17), 0.10),
        (3, date(2026, 4, 24), 0.50),
        (4, date(2026, 5, 1), 1.00),
    ):
        ws.cell(row, 1, f"=progress!A{row + 5}")
        ws.cell(row, 2, plan)
        ws.cell(row, 10, cutoff)
        ws.cell(row, 16, 0)
        ws.cell(row, 18, "=NA()")
    ws["D2"], ws["E2"], ws["K2"] = date(2026, 4, 24), 0.5, date(2026, 4, 24)
    ws["D3"], ws["E3"], ws["K3"] = date(2026, 5, 1), 1.0, date(2026, 5, 1)
    ws["Q2"], ws["S2"], ws["Q3"], ws["S3"] = 0, "=NA()", 0, "=NA()"

    weekly, _monthly = _build_explicit_overlay_series_sources(
        ws, weekly_first=2, weekly_last=4, monthly_first=2, monthly_last=3
    )

    assert weekly[:2] == (2, 5)
    # Synthetic 0,0 point is one reporting interval before the first real point.
    assert ws["T2"].value == date(2026, 4, 10)
    assert ws["U2"].value == 0
    assert ws["T3"].value == date(2026, 4, 17)
    assert ws["T5"].value == date(2026, 5, 1)
    assert ws["T6"].value is None

def _fill_rgb(fill) -> str:
    return (fill.fgColor.rgb or fill.fgColor.indexed or "")


def test_monthly_repaints_all_four_scurve_timescale_bands():
    wb = Workbook()
    ws = wb.active
    ws.title = "main"
    headers = ["Row Type", "WBS", "Description", "P/A", "Activity ID", "Task ID", "UID", "Outline Level"]
    for col, value in enumerate(headers, start=1):
        ws.cell(4, col, value)
    ws.cell(4, 9, date(2026, 4, 17))
    ws.cell(4, 10, date(2026, 4, 24))
    ws.cell(4, 11, date(2026, 5, 1))

    for row, (label, pa, fill) in enumerate(
        [
            ("Plan", "P", SCURVE_PLAN_FILL),
            ("Acc. Plan", "AP", WBS_PLAN_FILL),
            ("Actual", "A", SCURVE_ACTUAL_FILL),
            ("Acc. Actual", "AA", WBS_ACTUAL_FILL),
        ],
        start=5,
    ):
        ws.cell(row, 1, "S-Curve")
        ws.cell(row, 3, label)
        ws.cell(row, 4, pa)
        for col in (9, 10, 11):
            ws.cell(row, col, 0.01)
            ws.cell(row, col).fill = fill

    build_monthly_main_view(wb, source_sheet="main", target_sheet="main_monthly")
    monthly = wb["main_monthly"]
    expected = [SCURVE_PLAN_FILL, WBS_PLAN_FILL, SCURVE_ACTUAL_FILL, WBS_ACTUAL_FILL]
    for row, fill in zip(range(5, 9), expected):
        assert _fill_rgb(monthly.cell(row, 9).fill) == _fill_rgb(fill)
        assert _fill_rgb(monthly.cell(row, 10).fill) == _fill_rgb(fill)


def test_monthly_overlay_keeps_final_reporting_month_after_calendar_project_finish():
    """009 regression: Project Finish in May can still report 100% in June."""
    periods = (
        MainPeriod(18, "2027-04-30", datetime(2027, 4, 30)),
        MainPeriod(19, "2027-05-07", datetime(2027, 5, 7)),
        MainPeriod(20, "2027-05-28", datetime(2027, 5, 28)),
        MainPeriod(21, "2027-06-04", datetime(2027, 6, 4)),  # overlaps 31-May finish
        MainPeriod(22, "2027-06-11", datetime(2027, 6, 11)),  # display-only margin
        MainPeriod(23, "2027-07-02", datetime(2027, 7, 2)),   # display-only margin
    )
    summary = MainRow(
        row_number=5, row_type="Project Summary", pa="P", wbs="", description="009",
        activity_id="", outline_level=0,
        plan_start=datetime(2026, 4, 17), plan_finish=datetime(2027, 5, 31),
        amount=100.0, percent_complete=0.0, period_values=(),
    )
    dataset = MainDataset(
        workbook_name="009.xlsx", header_row=4, headers=(), periods=periods, rows=(summary,),
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard_Data"
    ws["D1"] = "Monthly Date"
    ws["E1"] = "Monthly Plan"
    ws["D2"], ws["E2"] = date(2027, 4, 30), 0.80
    ws["D3"], ws["E3"] = date(2027, 5, 28), 0.99
    ws["D4"], ws["E4"] = date(2027, 6, 4), 1.00

    first_row, last_row, first_col, last_col = _monthly_project_window(ws, dataset)

    assert (first_row, last_row) == (2, 4)
    # April/May/June physical months are the first three month columns here.
    assert (first_col, last_col) == (18, 20)
    assert ws.cell(last_row, 5).value == 1.00
