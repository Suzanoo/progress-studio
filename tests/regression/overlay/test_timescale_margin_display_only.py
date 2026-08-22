from __future__ import annotations

from datetime import datetime

from openpyxl import Workbook

from progress_studio.domain.main_dataset import MainDataset, MainPeriod
from progress_studio.infrastructure.excel.dashboard_workbook import _LAYOUT
from progress_studio.infrastructure.excel.live_scurve_workbook import normalize_weekly_scurve_source_contract
from progress_studio.infrastructure.excel.progress_workbook import add_scurve_rows


def _main_sheet_fixture():
    wb = Workbook()
    ws = wb.active
    ws.title = "main"
    headers = ["Row Type", "Description", "P/A"]
    for col, value in enumerate(headers, start=1):
        ws.cell(4, col, value)
    # Weekly display range: W1..W5, where W5 is post-project display margin.
    for offset, col in enumerate(range(4, 9), start=1):
        ws.cell(3, col, f"W{offset}")
    ws.cell(5, 1, "Project Summary")
    ws.cell(5, 3, "P")
    ws.cell(6, 1, "Project Summary")
    ws.cell(6, 3, "A")
    # Plan exists only W2..W4. W1/W5 are display margin.
    ws.cell(5, 5, 0.25)
    ws.cell(5, 6, 0.50)
    ws.cell(5, 7, 0.25)
    return wb, ws


def test_n71_acc_plan_formula_blanks_after_latest_plan_value() -> None:
    wb, ws = _main_sheet_fixture()
    rows = add_scurve_rows(
        ws,
        timescale_cols=[4, 5, 6, 7, 8],
        header_map={"row type": 1, "description": 2, "p/a": 3},
        project_plan_row=5,
    )
    formula = ws.cell(rows["Acc. Plan"], 8).value
    assert 'COUNT(H' in formula
    assert ':$H' in formula
    assert 'SUM($D' in formula
    wb.close()


def test_n71_live_scurve_keeps_plan_and_actual_out_of_post_margin() -> None:
    wb, ws = _main_sheet_fixture()
    # Explicit S-Curve rows expected by the live normalizer.
    for row, label, pa in ((8, "Plan", "P"), (9, "Acc. Plan", "AP"), (10, "Actual", "A"), (11, "Acc. Actual", "AA")):
        ws.cell(row, 1, "S-Curve")
        ws.cell(row, 2, label)
        ws.cell(row, 3, pa)
    dataset = MainDataset(
        workbook_name="fixture.xlsx",
        header_row=4,
        headers=(("row type", 1), ("description", 2), ("p/a", 3)),
        periods=tuple(
            MainPeriod(column=c, key=f"W{c-3}", reporting_date=datetime(2026, 1, c))
            for c in range(4, 9)
        ),
        rows=(),
    )
    assert normalize_weekly_scurve_source_contract(wb, dataset)
    acc_plan = ws.cell(9, 8).value
    acc_actual = ws.cell(11, 8).value
    assert 'COUNT(H8:$H8)' in acc_plan
    assert 'COUNT(H10:$H10)' in acc_actual
    wb.close()


def test_n71_dashboard_chart_has_headroom_above_100_percent() -> None:
    assert float(_LAYOUT.get("chart_y_max", 1.0)) > 1.0
    assert float(_LAYOUT["chart_y_max"]) == 1.1
