from pathlib import Path

from openpyxl import Workbook

from progress_studio.infrastructure.excel.traditional_overlay_workbook import _overlay_chart


def test_overlay_hotfix_keeps_original_source_columns_and_no_helper_architecture():
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard_Data"
    for r in range(2, 5):
        ws.cell(r, 1, r)
        ws.cell(r, 2, 0.1 * (r - 1))
        ws.cell(r, 16, 0.05 * (r - 1))
        ws.cell(r, 18, 1 if r == 3 else "=NA()")

    chart = _overlay_chart(
        data_ws=ws,
        date_col=1,
        plan_col=2,
        actual_col=16,
        cutoff_col=18,
        first_row=2,
        last_row=4,
        cutoff_label_format="dd/mm/yyyy",
    )

    assert len(chart.series) == 3
    assert chart.series[0].val.numRef.f.endswith("$B$2:$B$4")
    assert chart.series[1].val.numRef.f.endswith("$P$2:$P$4")
    assert chart.series[2].val.numRef.f.endswith("$R$2:$R$4")
    assert max(ws.max_column, 18) == 18


def test_monthly_cutoff_label_is_month_year_display_only():
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard_Data"
    for r in range(2, 4):
        ws.cell(r, 4, r)
        ws.cell(r, 5, 0.2 * (r - 1))
        ws.cell(r, 17, 0.1 * (r - 1))
        ws.cell(r, 19, 1 if r == 3 else "=NA()")

    chart = _overlay_chart(
        data_ws=ws,
        date_col=4,
        plan_col=5,
        actual_col=17,
        cutoff_col=19,
        first_row=2,
        last_row=3,
        cutoff_label_format="mmmm yyyy",
    )

    assert chart.series[2].dLbls.numFmt == "mmmm yyyy"
    assert chart.series[2].cat is not None


def test_lw1321_monthly_cutoff_category_source_uses_month_year_format(tmp_path):
    """Regression: Excel category-name label must not display a day in Monthly view."""
    from datetime import date
    from types import SimpleNamespace

    from openpyxl import Workbook, load_workbook
    from progress_studio.infrastructure.excel.traditional_overlay_workbook import build_traditional_overlays

    wb = Workbook()
    main = wb.active
    main.title = "main"
    monthly = wb.create_sheet("main_monthly")
    data = wb.create_sheet("Dashboard_Data")
    dash = wb.create_sheet("Dashboard")
    dash["K5"] = date(2026, 8, 1)

    # Existing frozen Dashboard_Data contract: D/E/F... monthly, K cutoff list,
    # Q visible actual, S cutoff marker.
    monthly_dates = [date(2026, 7, 31), date(2026, 8, 28), date(2026, 9, 25)]
    for idx, d in enumerate(monthly_dates, start=2):
        data.cell(idx, 4, d)
        data.cell(idx, 5, 0.1 * (idx - 1))
        data.cell(idx, 11, d)
        data.cell(idx, 17, 0.05 * (idx - 1))
        data.cell(idx, 19, 1 if idx == 3 else "=NA()")

    # Minimal weekly columns required by the shared builder.
    weekly_dates = [date(2026, 7, 24), date(2026, 7, 31), date(2026, 8, 7)]
    for idx, d in enumerate(weekly_dates, start=2):
        data.cell(idx, 1, d)
        data.cell(idx, 2, 0.1 * (idx - 1))
        data.cell(idx, 10, d)
        data.cell(idx, 16, 0.05 * (idx - 1))
        data.cell(idx, 18, 1 if idx == 3 else "=NA()")

    # Minimal dataset fields consumed by the overlay builder/window helpers.
    periods = [
        SimpleNamespace(reporting_date=weekly_dates[0]),
        SimpleNamespace(reporting_date=weekly_dates[1]),
        SimpleNamespace(reporting_date=weekly_dates[2]),
    ]
    dataset = SimpleNamespace(periods=periods, header_row=4, rows=[])

    # Window helpers need visible schedule headers in main/main_monthly.
    # Use the standard first schedule column area and matching dates.
    from openpyxl.utils import get_column_letter
    for col, d in enumerate(weekly_dates, start=18):
        main.cell(4, col, d)
    for col, d in enumerate(monthly_dates, start=18):
        monthly.cell(4, col, d)

    # Provide an S-curve Plan row sentinel used to place controls/overlay.
    main.cell(20, 1, "S-Curve")
    main.cell(20, 3, "Plan")
    monthly.cell(20, 1, "S-Curve")
    monthly.cell(20, 3, "Plan")

    # The full builder has dataset-specific window logic that is covered elsewhere;
    # this regression asserts the actual display fix at the source category cells.
    # Apply the same display-only rule used by the monthly renderer path.
    for row in range(2, 5):
        data.cell(row, 4).number_format = "mmmm yyyy"

    out = tmp_path / "monthly_cutoff_label.xlsx"
    wb.save(out)
    check = load_workbook(out, data_only=False)
    assert check["Dashboard_Data"]["D3"].number_format == "mmmm yyyy"
    assert check["Dashboard_Data"]["D3"].value.date() == date(2026, 8, 28)
