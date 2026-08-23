from pathlib import Path
from zipfile import ZipFile
from datetime import datetime
from xml.etree import ElementTree as ET
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation

from progress_studio.domain.main_dataset import MainDataset, MainPeriod, MainRow
from progress_studio.infrastructure.excel.traditional_overlay_workbook import (
    _overlay_chart,
    _responsive_anchor,
    build_traditional_overlays,
)


def test_overlay_module_contract_is_wired():
    text = Path('progress_studio/services/rebuild_service.py').read_text()
    assert 'build_traditional_overlays(wb, dataset)' in text


def test_overlay_uses_all_markers_and_cutoff_helpers():
    text = Path('progress_studio/infrastructure/excel/traditional_overlay_workbook.py').read_text()
    assert 'Weekly Actual Visible' in text
    assert 'Monthly Actual Visible' in text
    assert 'series.marker.symbol = "circle"' in text
    assert 'OVERLAY_MARKER_SIZE = 7' in text
    assert 'DataLabelList(' in text
    assert 'showVal=True' in text
    assert 'numFmt=OVERLAY_LABEL_FORMAT' in text
    assert 'PLAN_LABEL_FILL = "DDEBF7"' in text
    assert 'ACTUAL_LABEL_FILL = "E2F0D9"' in text
    assert 'dLblPos=position' in text
    assert 'spPr=_label_graphical_properties' in text
    assert 'txPr=_label_text_properties' in text
    assert 'chart.x_axis.tickLblPos = "none"' in text
    assert 'weekly_cutoff_ref: str' in text
    assert 'monthly_cutoff_ref: str' in text
    assert 'Weekly Overlay Date' in text and 'Weekly Overlay Plan' in text and 'Weekly Overlay Actual' in text
    assert 'Monthly Overlay Date' in text and 'Monthly Overlay Plan' in text and 'Monthly Overlay Actual' in text
    assert 'weekly_date_col' in text and 'weekly_plan_col' in text and 'weekly_actual_col' in text
    assert 'monthly_date_col' in text and 'monthly_plan_col' in text and 'monthly_actual_col' in text


def test_overlay_lw1231_responsive_geometry_and_transparency_contract():
    text = Path('progress_studio/infrastructure/excel/traditional_overlay_workbook.py').read_text()
    assert 'OVERLAY_TOP_ROW = 5' in text
    assert 'TwoCellAnchor' in text
    assert 'editAs="twoCell"' in text
    assert '_weekly_project_window(data_ws, dataset)' in text
    assert '_scurve_plan_row(dataset)' in text
    assert 'scurve_plan_row - 1' in text
    assert 'chart.plot_area.graphicalProperties' in text
    assert 'LineProperties(noFill=True)' in text


def test_overlay_serializes_two_cell_anchor_transparency_and_value_labels(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Dashboard_Data'
    ws.append(['Date', 'Plan', 'Actual'])
    for i in range(1, 5):
        ws.append([i, i / 4, i / 5])

    chart = _overlay_chart(
        data_ws=ws,
        date_col=1,
        plan_col=2,
        actual_col=3,
        first_row=2,
        last_row=5,
    )
    chart.anchor = _responsive_anchor(first_col=4, last_col=7, top_row=5, bottom_row=20)
    ws.add_chart(chart)
    path = tmp_path / 'overlay.xlsx'
    wb.save(path)

    with ZipFile(path) as zf:
        drawing_xml = zf.read('xl/drawings/drawing1.xml').decode('utf-8')
        chart_xml = zf.read('xl/charts/chart1.xml').decode('utf-8')

    assert '<twoCellAnchor editAs="twoCell">' in drawing_xml
    assert '<from><col>3</col>' in drawing_xml
    assert '<to><col>7</col>' in drawing_xml
    # One noFill is the inner plot area and one is the outer chart area.
    assert chart_xml.count('<a:noFill') >= 2
    assert '<dLbls>' in chart_xml
    chart_root = ET.fromstring(chart_xml)
    show_values = [elem.get('val') for elem in chart_root.iter() if elem.tag.endswith('showVal')]
    assert '1' in show_values
    assert '<numFmt formatCode="0.0%"' in chart_xml
    # LW-12.3.2 label tags: Plan is above with pale-blue fill; Actual is below
    # with pale-green fill. Text is compact 7 pt and tinted by series.
    label_positions = [elem.get('val') for elem in chart_root.iter() if elem.tag.endswith('dLblPos')]
    assert 't' in label_positions
    assert 'b' in label_positions
    assert 'val="DDEBF7"' in chart_xml
    assert 'val="E2F0D9"' in chart_xml
    assert 'val="1F4E79"' in chart_xml
    assert 'val="385723"' in chart_xml
    assert 'sz="700"' in chart_xml


def test_overlay_lw1233_project_bounds_and_cutoff_controls_contract():
    text = Path('progress_studio/infrastructure/excel/traditional_overlay_workbook.py').read_text()
    assert 'data_ws.cell(row, 2).value not in (None, "")' in text
    assert 'data_ws.cell(row, 5).value not in (None, "")' in text
    assert '_build_explicit_overlay_series_sources' in text
    assert 'data_ws.cell(2, 21, 0)' in text
    assert 'data_ws.cell(2, 22, 0)' in text
    assert 'data_ws.cell(2, 25, 0)' in text
    assert 'data_ws.cell(2, 26, 0)' in text
    assert 'first_row=weekly_chart_first' in text
    assert 'last_row=weekly_chart_last' in text
    assert 'first_row=monthly_chart_first' in text
    assert 'last_row=monthly_chart_last' in text
    assert 'label.value = "Cutoff Date"' in text
    assert 'value.value = initial_value' in text
    assert 'This cutoff belongs to this sheet only' in text
    assert 'weekly_cutoff_ref=weekly_cutoff_ref' in text
    assert 'monthly_cutoff_ref=monthly_cutoff_ref' in text
    assert 'list_col="J"' in text
    assert 'list_col="K"' in text
    assert 'label_col = 12  # L' in text
    assert 'value_col = 13  # M' in text
    assert 'display_format="dd/mm/yyyy"' in text
    assert 'display_format="mmm yyyy"' in text



def test_lw124_cutoff_red_line_has_date_label_contract():
    text = Path('progress_studio/infrastructure/excel/traditional_overlay_workbook.py').read_text()
    assert 'CUTOFF_RED = "C00000"' in text
    assert 'weekly_cutoff_col' in text
    assert 'monthly_cutoff_col' in text
    assert 'Weekly Overlay Cutoff' in text
    assert 'Monthly Overlay Cutoff' in text
    assert 'ErrorBars(' in text
    assert 'prstDash="dash"' in text
    assert 'showCatName=True' in text
    assert 'showSerName=True' in text
    assert 'separator=" "' in text
    assert 'SeriesLabel(v="Cutoff")' in text
    assert 'CUTOFF_LABEL_BG = "FCE4D6"' in text
    assert 'spPr=_label_graphical_properties(CUTOFF_LABEL_BG, CUTOFF_LABEL_BORDER)' in text
    assert 'CUTOFF_LABEL_FONT_SIZE = 1000' in text
    assert 'txPr=_cutoff_label_text_properties()' in text
    assert 'chart.legend = None' in text


def test_lw124_cutoff_line_serializes_error_bar_and_label(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Dashboard_Data'
    ws.append(['Date', 'Plan', 'Actual', 'Cutoff'])
    for i in range(1, 5):
        ws.append([i, i / 4, i / 5, 1 if i == 3 else '#N/A'])

    chart = _overlay_chart(
        data_ws=ws,
        date_col=1,
        plan_col=2,
        actual_col=3,
        cutoff_col=4,
        first_row=2,
        last_row=5,
    )
    ws.add_chart(chart, 'F2')
    path = tmp_path / 'cutoff_line.xlsx'
    wb.save(path)

    with ZipFile(path) as zf:
        chart_xml = zf.read('xl/charts/chart1.xml').decode('utf-8')

    assert '<errBars>' in chart_xml
    chart_root = ET.fromstring(chart_xml)
    err_bar_types = [elem.get('val') for elem in chart_root.iter() if elem.tag.endswith('errBarType')]
    assert 'minus' in err_bar_types
    assert 'val="C00000"' in chart_xml
    preset_dashes = [elem.get('val') for elem in chart_root.iter() if elem.tag.endswith('prstDash')]
    assert 'dash' in preset_dashes
    show_category_names = [elem.get('val') for elem in chart_root.iter() if elem.tag.endswith('showCatName')]
    assert '1' in show_category_names
    show_series_names = [elem.get('val') for elem in chart_root.iter() if elem.tag.endswith('showSerName')]
    assert '1' in show_series_names
    separators = [elem.get('val') for elem in chart_root.iter() if elem.tag.endswith('separator')]
    assert ' ' in separators
    assert 'val="FCE4D6"' in chart_xml
    assert 'sz="1000"' in chart_xml
    assert 'b="1"' in chart_xml



def test_lw124_three_cutoffs_are_independent_and_overlay_helpers_are_local():
    wb = Workbook()
    main = wb.active
    main.title = 'main'
    monthly = wb.create_sheet('main_monthly')
    dash = wb.create_sheet('Dashboard')
    data = wb.create_sheet('Dashboard_Data')
    dash['K5'] = datetime(2026, 7, 17)

    headers = (
        ('row type', 1), ('wbs', 2), ('description', 3), ('p/a', 4),
        ('activity id', 5), ('outline level', 6), ('plan start', 7),
        ('plan finish', 8), ('amount', 9),
    )
    periods = tuple(
        MainPeriod(10 + i, f'W{i+1}', datetime(2026, 7, 3 + 7 * i))
        for i in range(4)
    )
    activity = MainRow(
        row_number=5, row_type='Activity', pa='P', wbs='1', description='A',
        activity_id='A100', outline_level=1,
        plan_start=datetime(2026, 7, 3), plan_finish=datetime(2026, 7, 24),
        amount=100.0, percent_complete=None, period_values=(),
    )
    scurve = MainRow(
        row_number=20, row_type='S-Curve', pa='P', wbs='', description='Plan',
        activity_id='', outline_level=0, plan_start=None, plan_finish=None,
        amount=None, percent_complete=None, period_values=(),
    )
    dataset = MainDataset('p.xlsx', 4, headers, periods, (activity, scurve))

    # Weekly source A:C + weekly cutoff list J.
    for col, value in enumerate(('Weekly Date', 'Weekly Plan', 'Weekly Actual'), 1):
        data.cell(1, col, value)
    weekly = [datetime(2026, 7, 3), datetime(2026, 7, 10), datetime(2026, 7, 17), datetime(2026, 7, 24)]
    for r, dt in enumerate(weekly, 2):
        data.cell(r, 1, dt); data.cell(r, 2, (r-1)/4); data.cell(r, 3, (r-1)/5); data.cell(r, 10, dt)

    # Monthly source D:F + monthly cutoff list K.
    for col, value in enumerate(('Monthly Date', 'Monthly Plan', 'Monthly Actual'), 4):
        data.cell(1, col, value)
    monthly_dates = [datetime(2026, 7, 31), datetime(2026, 8, 31)]
    for r, dt in enumerate(monthly_dates, 2):
        data.cell(r, 4, dt); data.cell(r, 5, .5 * (r-1)); data.cell(r, 6, .4 * (r-1)); data.cell(r, 11, dt)

    build_traditional_overlays(wb, dataset)

    # Control row is immediately above the S-Curve Plan row; Amount is col I.
    assert main['M19'].value == datetime(2026, 7, 17)
    # Monthly owns a different value/list and does not bind to Dashboard.
    assert monthly['M19'].value == datetime(2026, 8, 31)
    assert main['M19'].number_format == 'dd/mm/yyyy'
    assert monthly['M19'].number_format == 'mmm yyyy'
    assert dash['K5'].value == datetime(2026, 7, 17)

    # Cutoff helpers use workbook defined-name proxies. This preserves the
    # independent main/monthly controls while keeping Dashboard_Data detached
    # from direct main-sheet formulas for snapshot performance.
    assert 'PS_WEEKLY_OVERLAY_CUTOFF' in data['P2'].value
    assert 'PS_MONTHLY_OVERLAY_CUTOFF' in data['Q2'].value
    assert wb.defined_names['PS_WEEKLY_OVERLAY_CUTOFF'].attr_text == "'main'!$M$19"
    assert wb.defined_names['PS_MONTHLY_OVERLAY_CUTOFF'].attr_text == "'main_monthly'!$M$19"
    assert 'Dashboard!$K$5' not in data['P2'].value
    assert 'Dashboard!$K$5' not in data['Q2'].value
    assert 'PS_WEEKLY_OVERLAY_CUTOFF' in data['R2'].value
    assert 'PS_MONTHLY_OVERLAY_CUTOFF' in data['S2'].value
    assert 'AND(A2<=' in data['R2'].value
    assert 'OR(A3="",A3>' in data['R2'].value

    # Simulate independent user edits: no formula links exist between controls.
    main['M19'] = datetime(2026, 7, 10)
    monthly['M19'] = datetime(2026, 7, 31)
    assert dash['K5'].value == datetime(2026, 7, 17)
    assert main['M19'].value != monthly['M19'].value


def test_lw124_local_cutoff_cells_are_unlocked_without_unprotecting_sheet():
    from progress_studio.infrastructure.excel.workbook_protection import apply_final_sheet_protection

    wb = Workbook()
    main = wb.active
    main.title = 'main'
    monthly = wb.create_sheet('main_monthly')

    for ws in (main, monthly):
        ws['A4'] = 'Row Type'
        ws['D4'] = 'P/A'
        ws['L19'] = 'Cutoff Date'
        ws['M19'] = '2026-07-17'
        ws['N19'] = 'protected neighbor'

    apply_final_sheet_protection(wb)

    assert main.protection.sheet is True
    assert monthly.protection.sheet is True
    assert main['M19'].protection.locked is False
    assert monthly['M19'].protection.locked is False
    assert main['N19'].protection.locked is True
    assert monthly['N19'].protection.locked is True


def test_lw1241_removes_legacy_cutoff_dropdown_and_deduplicates_current_validation():
    wb = Workbook()
    main = wb.active
    main.title = 'main'
    wb.create_sheet('main_monthly')
    dash = wb.create_sheet('Dashboard')
    data = wb.create_sheet('Dashboard_Data')
    dash['K5'] = datetime(2026, 7, 17)

    headers = (
        ('row type', 1), ('wbs', 2), ('description', 3), ('p/a', 4),
        ('activity id', 5), ('outline level', 6), ('plan start', 7),
        ('plan finish', 8), ('amount', 9),
    )
    periods = tuple(
        MainPeriod(10 + i, f'W{i+1}', datetime(2026, 7, 3 + 7 * i))
        for i in range(4)
    )
    activity = MainRow(
        row_number=5, row_type='Activity', pa='P', wbs='1', description='A',
        activity_id='A100', outline_level=1,
        plan_start=datetime(2026, 7, 3), plan_finish=datetime(2026, 7, 24),
        amount=100.0, percent_complete=None, period_values=(),
    )
    scurve = MainRow(
        row_number=20, row_type='S-Curve', pa='P', wbs='', description='Plan',
        activity_id='', outline_level=0, plan_start=None, plan_finish=None,
        amount=None, percent_complete=None, period_values=(),
    )
    dataset = MainDataset('p.xlsx', 4, headers, periods, (activity, scurve))

    weekly = [datetime(2026, 7, 3), datetime(2026, 7, 10), datetime(2026, 7, 17), datetime(2026, 7, 24)]
    for r, dt in enumerate(weekly, 2):
        data.cell(r, 1, dt); data.cell(r, 2, (r-1)/4); data.cell(r, 3, (r-1)/5); data.cell(r, 10, dt)
    monthly_dates = [datetime(2026, 7, 31), datetime(2026, 8, 31)]
    for r, dt in enumerate(monthly_dates, 2):
        data.cell(r, 4, dt); data.cell(r, 5, .5 * (r-1)); data.cell(r, 6, .4 * (r-1)); data.cell(r, 11, dt)

    # Simulate the old control: Description column C + Amount column I.
    main['C19'] = 'Cutoff Date'
    main['I19'] = datetime(2026, 7, 10)
    old_validation = DataValidation(type='list', formula1='"x,y"')
    main.add_data_validation(old_validation)
    old_validation.add(main['I19'])

    build_traditional_overlays(wb, dataset)
    build_traditional_overlays(wb, dataset)

    assert main['C19'].value is None
    assert main['I19'].value is None
    assert main['L19'].value == 'Cutoff Date'
    assert main['M19'].value == datetime(2026, 7, 17)
    target_validations = [dv for dv in main.data_validations.dataValidation if 'M19' in dv.cells]
    assert len(target_validations) == 1
    assert all('I19' not in dv.cells for dv in main.data_validations.dataValidation)
