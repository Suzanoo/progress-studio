
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import PatternFill

from progress_studio.config.payment_theme import load_payment_line_theme
from progress_studio.infrastructure.excel.payment_line_renderer import PaymentLineRenderer


def test_monthly_uses_main_timescale_conditional_formatting_contract() -> None:
    source = Path(
        "progress_studio/infrastructure/excel/live_monthly_workbook.py"
    ).read_text(encoding="utf-8")
    assert "add_progress_conditional_formatting" in source
    assert "clear_timescale_direct_fills" in source
    assert "F7FBFF" not in source
    assert "EEF5FA" not in source


def test_payment_label_is_doubled() -> None:
    theme = load_payment_line_theme()
    assert theme.label.width_px == 290
    assert theme.label.height_px == 52
    assert theme.label.font_size == 24
    assert theme.label.corner_radius_px == 12


def test_payment_timescale_copies_activity_data_fill() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Payment"
    ws["A4"] = "Row Type"
    ws["C4"] = "Description"
    ws["D4"] = "P/A"
    from datetime import datetime
    ws["E4"] = datetime(2026, 3, 6)
    ws["F4"] = datetime(2026, 3, 13)
    ws["C5"].fill = PatternFill("solid", fgColor="F4B183")
    ws["C6"].fill = PatternFill("solid", fgColor="E2F0D9")

    PaymentLineRenderer()._paint_timescale_like_activity_data(ws)

    assert ws["E5"].fill.fgColor.rgb.endswith("F4B183")
    assert ws["F5"].fill.fgColor.rgb.endswith("F4B183")
    assert ws["E6"].fill.fgColor.rgb.endswith("E2F0D9")


def test_activity_table_uses_status_autofilter_and_no_status_dropdown() -> None:
    source = Path(
        "progress_studio/infrastructure/excel/live_dashboard_workbook.py"
    ).read_text(encoding="utf-8")
    start = source.index("def _write_activity_section")
    end = source.index("\ndef build_live_dashboard", start)
    block = source[start:end]
    assert 'ws.auto_filter.ref = f"P38:P{last_activity_row}"' in block
    assert '"All,Behind,On Track,Complete,Not Started"' not in block
    assert "status_validation" not in block
    assert "indent=min(max(item.outline_level, 0), 7)" in block
