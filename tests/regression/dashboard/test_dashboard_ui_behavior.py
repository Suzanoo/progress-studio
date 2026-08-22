from datetime import date

from openpyxl import Workbook

from progress_studio.infrastructure.excel.dashboard_workbook import (
    DASHBOARD_SHEET,
    DATA_SHEET,
    build_dashboard,
)


def _workbook():
    wb = Workbook()
    progress = wb.active
    progress.title = "progress"
    progress.append(["project_start", "project_finish", "week_start", "plan", "actual"])
    progress.append([date(2026, 1, 1), date(2026, 2, 28), date(2026, 1, 2), 10, 5])
    progress.append([date(2026, 1, 1), date(2026, 2, 28), date(2026, 1, 9), 25, 12])
    progress.append([date(2026, 1, 1), date(2026, 2, 28), date(2026, 2, 6), 60, None])
    progress.append([date(2026, 1, 1), date(2026, 2, 28), date(2026, 2, 27), 100, None])

    table = wb.create_sheet("progress_table")
    table.append(["WBS", "Activities", "Amount", "P/A", "%Progress", date(2026, 1, 2), date(2026, 1, 9)])
    table.append(["1.1", "Activity A", 1000, "P", 25, 10, 15])
    table.append(["1.1", "Activity A", 1000, "A", 12, 5, 7])
    return wb


def test_ms_p19_has_only_four_kpi_cards_and_weekly_monthly_view():
    wb = _workbook()
    build_dashboard(wb, project_name="Demo")
    ws = wb[DASHBOARD_SHEET]

    assert ws["B9"].value == "PLANNED PROGRESS"
    assert ws["E9"].value == "ACTUAL PROGRESS"
    assert ws["H9"].value == "SCHEDULE STATUS"
    assert ws["K9"].value == "TIME IMPACT"
    values = [cell.value for row in ws.iter_rows(min_row=1, max_row=20) for cell in row]
    assert "PROGRESS GAP" not in values
    assert "REPORTING PERIOD" not in values

    validations = list(ws.data_validations.dataValidation)
    assert any(v.formula1 == '"Weekly,Monthly"' for v in validations)
    assert ws["G5"].value == "Monthly"


def test_ms_p19_plan_curve_is_full_baseline_but_actual_curve_is_cutoff_limited():
    wb = _workbook()
    build_dashboard(wb)
    data = wb[DATA_SHEET]

    # Display Plan does not contain cutoff logic; Display Actual does.
    assert "Dashboard!$K$5" not in data["H2"].value
    assert "Dashboard!$K$5" in data["I2"].value
    assert data["H5"].value is not None


def test_ms_p19_kpis_read_stable_progress_sheet_directly():
    wb = _workbook()
    build_dashboard(wb)
    ws = wb[DASHBOARD_SHEET]

    assert "'progress'!$C$2:$C$5<=$K$5" in ws["B10"].value
    assert "'progress'!$D$2:$D$5" in ws["B10"].value
    assert "'progress'!$E$2:$E$5" in ws["E10"].value
    assert "Dashboard_Data" not in ws["B10"].value
    assert "Dashboard_Data" not in ws["E10"].value


def test_ms_p19_time_impact_uses_project_duration_from_progress_and_rounds_days():
    wb = _workbook()
    build_dashboard(wb)
    ws = wb[DASHBOARD_SHEET]

    formula = ws["K10"].value
    assert "ABS(E10-B10)" in formula
    assert "'progress'!$B$2-'progress'!$A$2" in formula
    assert formula.startswith("=ROUND(")
    assert formula.endswith('&" Days"')
    assert "Dashboard_Data" not in formula


def test_ms_p19_chart_and_theme_layout_are_configurable_and_wide():
    wb = _workbook()
    build_dashboard(wb)
    ws = wb[DASHBOARD_SHEET]

    assert len(ws._charts) == 1
    chart = ws._charts[0]
    assert chart.width >= 24
    assert chart.height >= 9
    assert ws["B35"].value.startswith("Plan curve = full baseline")
    assert ws.print_area == "'Dashboard'!$B$2:$Q$56"
