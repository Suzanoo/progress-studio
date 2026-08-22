from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

from progress_studio.infrastructure.excel.live_dashboard_workbook import build_live_dashboard
from progress_studio.infrastructure.excel.rebuild_workbook_reader import RebuildWorkbookReader
from progress_studio.services.activity_table_deriver import ActivityTableDeriver


def _fixture(path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "main"
    headers = [
        "Row Type", "WBS", "Description", "P/A", "Activity ID",
        "Outline Level", "Plan Start", "Plan Finish", "Amount",
    ]
    for col, value in enumerate(headers, 1):
        ws.cell(4, col, value)
    dates = [datetime(2026, 3, 6), datetime(2026, 3, 13)]
    for col, dt in enumerate(dates, start=10):
        ws.cell(3, col, f"W{col-9}")
        ws.cell(4, col, dt)

    rows = [
        ["Project Summary", "", "Project", "P", "", 0, datetime(2026,3,1), datetime(2026,4,30), None, .20, .30],
        ["Project Summary", "", "", "A", "", 0, None, None, None, .10, .10],
        ["WBS", "1", "Substructure", "P", "", 1, datetime(2026,3,1), datetime(2026,4,30), None, .20, .30],
        ["", "1", "", "A", "", 1, None, None, None, .10, .10],
        ["Activity", "1", "Excavation", "P", "A1000", 2, datetime(2026,3,1), datetime(2026,3,20), 600, .20, .30],
        ["", "1", "", "A", "A1000", 2, None, None, None, .10, .10],
        ["Activity", "1", "Footing", "P", "A1010", 2, datetime(2026,3,10), datetime(2026,4,30), 400, .20, .30],
        ["", "1", "", "A", "A1010", 2, None, None, None, .10, .10],
    ]
    for row in rows:
        ws.append(row)
    wb.save(path)
    return path


def test_lw114_project_information_and_axes(tmp_path: Path) -> None:
    path = _fixture(tmp_path / "project.xlsx")
    dataset = RebuildWorkbookReader().read_main_dataset(path)
    wb = load_workbook(path)
    build_live_dashboard(wb, dataset, cutoff=datetime(2026, 3, 13))

    dash = wb["Dashboard"]
    assert dash["F5"].value == "View"
    assert dash["G5"].value == "Weekly"
    assert dash["J5"].value == "Cutoff Date"
    assert dash["K5"].value == datetime(2026, 3, 13).date()
    assert dash["B6"].value == "Project Start"
    assert dash["C6"].value == datetime(2026, 3, 1).date()
    assert dash["F6"].value == "Project Finish"
    assert dash["G6"].value == datetime(2026, 4, 30).date()
    assert dash["J6"].value == "Project Value"
    assert dash["K6"].value == 1000.0

    chart = dash._charts[0]
    assert chart.x_axis.title is not None
    assert chart.y_axis.title is not None
    wb.close()


def test_lw114_wbs_total_rolls_up_descendant_activities(tmp_path: Path) -> None:
    path = _fixture(tmp_path / "rollup.xlsx")
    dataset = RebuildWorkbookReader().read_main_dataset(path)
    model = ActivityTableDeriver().derive(dataset, cutoff=datetime(2026, 3, 13))

    wbs_rows = [row for row in model.rows if row.row_type == "wbs"]
    assert len(wbs_rows) == 2
    assert wbs_rows[0].type_label == "Plan"
    assert wbs_rows[0].total == 1000.0
    assert wbs_rows[1].type_label == "Actual"
    assert wbs_rows[1].total == 1000.0
    assert wbs_rows[0].amount == 500.0
    assert wbs_rows[1].amount == 200.0

    wb = load_workbook(path)
    build_live_dashboard(wb, dataset, cutoff=datetime(2026, 3, 13))
    dash = wb["Dashboard"]
    # Project/WBS parent rows show Total on both Plan and Actual rows.
    assert dash["H41"].value == 1000.0
    assert dash["H42"].value == 1000.0
    assert dash["J41"].value == "=IFERROR(H41*L41,0)"
    assert dash["J42"].value == "=IFERROR($H$41*L42,0)"
    wb.close()
