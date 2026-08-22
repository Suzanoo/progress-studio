from datetime import date
from pathlib import Path

from tests._paths import REPO_ROOT

from openpyxl import Workbook, load_workbook

from progress_studio.infrastructure.excel.dashboard_workbook import DASHBOARD_SHEET, build_dashboard


def _workbook():
    wb = Workbook()
    progress = wb.active
    progress.title = "progress"
    progress.append(["project_start", "project_finish", "week_start", "plan", "actual"])
    progress.append([date(2026, 1, 1), date(2026, 2, 28), date(2026, 1, 2), 10, 5])
    progress.append([date(2026, 1, 1), date(2026, 2, 28), date(2026, 1, 9), 25, 12])
    table = wb.create_sheet("progress_table")
    table.append(["WBS", "Activities", "Amount", "P/A", "%Progress", date(2026, 1, 2)])
    table.append(["1.1", "Activity A", 1000, "P", 25, 10])
    table.append(["1.1", "Activity A", 1000, "A", 12, 5])
    return wb


def test_ms_p110_dashboard_has_four_embedded_kpi_icons(tmp_path):
    wb = _workbook()
    build_dashboard(wb, project_name="Demo")
    ws = wb[DASHBOARD_SHEET]
    assert len(ws._images) == 4

    output = tmp_path / "dashboard-icons.xlsx"
    wb.save(output)
    wb.close()

    reopened = load_workbook(output, data_only=False)
    try:
        assert len(reopened[DASHBOARD_SHEET]._images) == 4
    finally:
        reopened.close()


def test_ms_p110_icon_assets_exist():
    root = REPO_ROOT / "progress_studio" / "assets" / "dashboard" / "icons"
    for name in ("planned.png", "actual.png", "schedule.png", "time_impact.png"):
        path = root / name
        assert path.exists()
        assert path.stat().st_size > 0
