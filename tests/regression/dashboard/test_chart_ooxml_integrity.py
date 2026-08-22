from __future__ import annotations

from io import BytesIO
from pathlib import Path
from zipfile import ZipFile
import xml.etree.ElementTree as ET

from openpyxl import Workbook, load_workbook
from openpyxl.chart import LineChart, Reference

from progress_studio.infrastructure.excel.dashboard_workbook import _date_axis_for_line_chart

_CHART_NS = "http://schemas.openxmlformats.org/drawingml/2006/chart"


def _chart_axis_refs(path: Path) -> list[tuple[set[int], list[int]]]:
    results: list[tuple[set[int], list[int]]] = []
    with ZipFile(path) as package:
        chart_names = sorted(
            name for name in package.namelist()
            if name.startswith("xl/charts/chart") and name.endswith(".xml")
        )
        assert chart_names, "fixture must contain at least one chart"
        for name in chart_names:
            root = ET.fromstring(package.read(name))
            axis_ids = {
                int(node.attrib["val"])
                for node in root.findall(f".//{{{_CHART_NS}}}axId")
            }
            cross_refs = [
                int(node.attrib["val"])
                for node in root.findall(f".//{{{_CHART_NS}}}crossAx")
            ]
            results.append((axis_ids, cross_refs))
    return results


def _make_date_axis_chart(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard"
    ws.append(["Period", "Plan"])
    ws.append([45200, 0.0])
    ws.append([45231, 0.5])
    ws.append([45261, 1.0])

    chart = LineChart()
    _date_axis_for_line_chart(chart, title="Period")
    chart.add_data(Reference(ws, min_col=2, min_row=1, max_row=4), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=4))
    ws.add_chart(chart, "D2")
    wb.save(path)
    wb.close()


def test_date_axis_serializes_reciprocal_axis_relationships(tmp_path: Path) -> None:
    path = tmp_path / "date_axis.xlsx"
    _make_date_axis_chart(path)

    for axis_ids, cross_refs in _chart_axis_refs(path):
        assert axis_ids == {10, 100}
        assert cross_refs
        assert set(cross_refs) <= axis_ids
        assert 10 in cross_refs
        assert 100 in cross_refs


def test_date_axis_relationships_survive_openpyxl_round_trip(tmp_path: Path) -> None:
    first = tmp_path / "first.xlsx"
    second = tmp_path / "second.xlsx"
    _make_date_axis_chart(first)

    wb = load_workbook(first, data_only=False)
    wb.save(second)
    wb.close()

    for axis_ids, cross_refs in _chart_axis_refs(second):
        assert set(cross_refs) <= axis_ids
        assert axis_ids == {10, 100}


def test_uploaded_009_failure_shape_is_detected_by_semantic_axis_check() -> None:
    """Regression guard for the Excel recovery dialog from 009 output.

    The historical bad shape was dateAx=500/valueAx=100 while valueAx.crossAx
    still pointed to 10.  XML parsing alone accepts it; semantic validation must
    reject the dangling reference.
    """
    axis_ids = {500, 100}
    cross_refs = [100, 10]
    assert not set(cross_refs) <= axis_ids
