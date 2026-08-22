from __future__ import annotations

from datetime import date

from openpyxl import Workbook

from progress_studio.infrastructure.excel.dashboard_workbook import (
    _build_data_sheet,
    _progress_rows,
)


def _progress_fixture():
    wb = Workbook()
    ws = wb.active
    ws.title = "progress"
    ws.append(["project_start", "project_finish", "week_start", "plan", "actual"])
    start = date(2027, 5, 1)
    finish = date(2027, 6, 30)
    # One pre-project margin week, project weeks including the reporting week that
    # overlaps 30-Jun, then two post-project display-margin weeks.
    rows = [
        (date(2027, 4, 23), None, None),
        (date(2027, 4, 30), None, None),
        (date(2027, 5, 7), 0.20, None),
        (date(2027, 5, 28), 0.60, None),
        (date(2027, 6, 25), 1.00, None),
        (date(2027, 7, 2), 1.00, None),  # 26-Jun..02-Jul overlaps Project Finish.
        (date(2027, 7, 9), None, None),
        (date(2027, 7, 16), None, None),
    ]
    for reporting, plan, actual in rows:
        ws.append([start, finish, reporting, plan, actual])
    return wb, ws


def test_n72_reporting_rows_keep_final_overlapping_week_but_drop_margins() -> None:
    wb, ws = _progress_fixture()
    rows = _progress_rows(wb, ws)
    dates = [reporting for _, reporting in rows]
    assert date(2027, 7, 2) in dates
    assert date(2027, 7, 9) not in dates
    assert date(2027, 4, 30) not in dates
    wb.close()


def test_n72_dashboard_plan_does_not_fall_to_zero_and_actual_stops_at_cutoff() -> None:
    wb, ws = _progress_fixture()
    dashboard = wb.create_sheet("Dashboard")
    dashboard["G5"] = "Weekly"
    dashboard["K5"] = date(2027, 6, 25)

    _build_data_sheet(wb, ws)
    data = wb["Dashboard_Data"]

    # Dashboard_Data must end on the final real reporting week, not post-project margin.
    weekly_dates = [data.cell(r, 1).value for r in range(2, data.max_row + 1) if data.cell(r, 1).value]
    assert weekly_dates[-1] == date(2027, 7, 2)
    assert date(2027, 7, 9) not in weekly_dates

    # Chart Plan uses #N/A for a missing source point, never a formula-empty string
    # that Excel can render as a zero-value plunge.
    plan_formula = str(data.cell(2, 8).value)
    assert "NA()" in plan_formula

    # Actual is explicitly cutoff-aware and uses #N/A after the selected cutoff.
    actual_formula = str(data.cell(2, 9).value)
    assert ">Dashboard!$K$5" in actual_formula
    assert "NA()" in actual_formula
    assert '"",0' in actual_formula
    wb.close()


def test_dashboard_chart_uses_date_axis_so_monthly_points_are_not_squeezed_left() -> None:
    from progress_studio.infrastructure.excel.dashboard_workbook import build_dashboard

    wb, ws = _progress_fixture()
    table = wb.create_sheet("progress_table")
    table.append(["WBS", "Activity", "Total", "Type", "Status", date(2027, 5, 7), "_Kind"])
    # Build the production dashboard, which also creates Dashboard_Data.
    build_dashboard(wb, project_name="Date Axis")
    chart = wb["Dashboard"]._charts[0]
    assert chart.x_axis.tagname == "dateAx"
    wb.close()
