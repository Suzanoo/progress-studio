from datetime import date

from openpyxl import Workbook

from progress_studio.infrastructure.excel.progress_workbook import (
    prepare_progress_and_scurve,
)


def _workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = "main"
    ws.append([])
    ws.append([])
    ws.append([None] * 16 + ["W1", "W2"])
    ws.append([
        "Row Type", "WBS", "Description", "P/A", "Activity ID",
        "Task ID", "UID", "Outline Level", "Plan Start", "Plan Finish",
        "Actual Start", "Actual Finish", "% Complete", "Physical %",
        "Amount", "Total Float (hr)", date(2026, 1, 2), date(2026, 1, 9),
    ])
    ws.append(["Project Summary", "", "Project", "P", "", "", "", 0, None, None, None, None, None, None, None, None, None, None])
    ws.append([None, None, None, "A", None, None, None, None, None, None, None, None, None, None, None, None, None, None])
    ws.append(["WBS", "1", "Group", "P", "", "", "", 1, None, None, None, None, None, None, None, None, None, None])
    ws.append([None, None, None, "A", None, None, None, None, None, None, None, None, None, None, None, None, None, None])
    ws.append(["Activity", "1.1", "Task A", "P", "A1000", "", "", 2, None, None, None, None, None, None, 1000, None, 0.5, 0.5])
    ws.append([None, None, None, "A", "A1000", "", "", 2, None, None, None, None, None, None, None, None, 0.1, 0.2])
    return wb, ws


def test_actual_amount_is_earned_value_and_parent_rows_roll_up_actuals():
    wb, ws = _workbook()

    prepare_progress_and_scurve(wb, ws)

    assert ws["M10"].value == '=IF(COUNT(Q10:R10)=0,"",SUM(Q10:R10))'
    assert ws["O10"].value == '=IF(M10="","",O9*M10)'
    assert ws["O8"].value == '=SUMIFS($O$9:$O$10,$E$9:$E$10,"<>",$D$9:$D$10,"A")'
    assert ws["O6"].value == '=SUMIFS($O$7:$O$10,$E$7:$E$10,"<>",$D$7:$D$10,"A")'
    assert ws["O10"].number_format != ";;;"


def test_actual_weekly_rollup_uses_plan_amount_not_earned_actual_amount():
    wb, ws = _workbook()

    prepare_progress_and_scurve(wb, ws)

    # Parent Actual weekly formulas must remain weighted by the full Plan Amount.
    assert "$O$8:$O$9" in ws["Q8"].value
    assert "$O$6:$O$9" in ws["Q6"].value
