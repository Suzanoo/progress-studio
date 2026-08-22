from __future__ import annotations

from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

from progress_studio.infrastructure.excel.monthly_main_workbook import build_monthly_main_view
from progress_studio.infrastructure.excel.okd_workbook import build_okd_sheets
from progress_studio.infrastructure.excel.progress_workbook import (
    find_timescale_columns,
    prepare_progress_and_scurve,
)
from progress_studio.infrastructure.excel.timescale_workbook import add_weekly_timescale


def _raw_schedule(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "main"
    headers = [
        "Row Type", "WBS", "Description", "P/A", "Activity ID", "Task ID", "UID",
        "Outline Level", "Plan Start", "Plan Finish", "Actual Start", "Actual Finish",
        "% Complete", "Physical %", "Total Float (hr)",
    ]
    ws.append(headers)
    ws.append([
        "Activity", "1", "Task", "P", "A1000", "", "", 1,
        date(2026, 2, 23), date(2026, 3, 10), None, None, 0, 0, 0,
    ])
    wb.save(path)
    wb.close()


def test_create_weekly_timescale_marks_margins_x_and_numbers_reporting_from_w1(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    output = tmp_path / "weekly.xlsx"
    _raw_schedule(source)

    add_weekly_timescale(
        source,
        output,
        sheet_name="main",
        cutoff_day=5,  # Friday
        margin_weeks=2,
    )

    wb = load_workbook(output, data_only=False)
    ws = wb["main"]
    timescale = [
        (ws.cell(3, col).value, ws.cell(4, col).value)
        for col in range(1, ws.max_column + 1)
        if isinstance(ws.cell(4, col).value, (date, datetime))
    ]
    labels = [label for label, _ in timescale]

    assert labels == ["X", "X", "X", "W1", "W2", "W3", "X", "X"]
    assert [label for label in labels if str(label).startswith("W")] == ["W1", "W2", "W3"]
    wb.close()


def _weekly_main_with_x_margins() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "main"
    headers = [
        "Row Type", "WBS", "Description", "P/A", "Activity ID", "Task ID", "UID",
        "Outline Level", "Plan Start", "Plan Finish", "Actual Start", "Actual Finish",
        "% Complete", "Physical %", "Amount", "Total Float (hr)", "XML Amount",
    ]
    for col, value in enumerate(headers, start=1):
        ws.cell(4, col).value = value

    # Four distinct monthly buckets make the X/M presentation contract explicit.
    period_data = [
        ("X", datetime(2026, 1, 30)),
        ("W1", datetime(2026, 2, 27)),
        ("W2", datetime(2026, 3, 27)),
        ("X", datetime(2026, 4, 24)),
    ]
    for col, (label, reporting_date) in enumerate(period_data, start=18):
        ws.cell(3, col).value = label
        ws.cell(4, col).value = reporting_date

    rows = [
        ["Project Summary", "", "Project", "P", "", "", "", 0, date(2026, 2, 1), date(2026, 3, 20)],
        ["", "", "", "A", "", "", "", 0],
        ["Activity", "1", "Activity", "P", "A1000", "", "", 1, date(2026, 2, 1), date(2026, 3, 20)],
        ["", "", "", "A", "A1000", "", "", 1],
        ["S-Curve", "", "Plan", "P", "", "", "", 0],
        ["S-Curve", "", "Acc. Plan", "AP", "", "", "", 0],
        ["S-Curve", "", "Actual", "A", "", "", "", 0],
        ["S-Curve", "", "Acc. Actual", "AA", "", "", "", 0],
    ]
    for row_idx, values in enumerate(rows, start=5):
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row_idx, col_idx).value = value

    # Reporting values exist only under W periods; X columns are display canvas.
    ws.cell(7, 19).value = 0.5
    ws.cell(7, 20).value = 0.5
    ws.cell(9, 19).value = 0.5
    ws.cell(9, 20).value = 0.5
    ws.cell(10, 19).value = 0.5
    ws.cell(10, 20).value = 1.0
    return wb


def test_monthly_create_uses_x_for_display_only_months_and_m1_mn_for_reporting_months() -> None:
    wb = _weekly_main_with_x_margins()
    build_monthly_main_view(wb)
    ws = wb["main_monthly"]

    assert [ws.cell(3, col).value for col in range(18, 22)] == ["X", "M1", "M2", "X"]
    # The physical monthly display range remains intact; only identity changes.
    assert [ws.cell(4, col).value for col in range(18, 22)] == [
        date(2026, 1, 30), date(2026, 2, 27), date(2026, 3, 27), date(2026, 4, 24)
    ]
    wb.close()


def test_progress_consumer_ignores_x_but_keeps_x_physical_cells() -> None:
    wb = _weekly_main_with_x_margins()
    ws = wb["main"]
    headers = {str(ws.cell(4, c).value or "").strip().lower(): c for c in range(1, ws.max_column + 1)}

    assert find_timescale_columns(ws, headers) == [19, 20]
    ws.cell(7, 18).value = "PRE-MARGIN"
    ws.cell(7, 21).value = "POST-MARGIN"

    prepare_progress_and_scurve(wb, ws)

    assert ws.cell(7, 18).value == "PRE-MARGIN"
    assert ws.cell(7, 21).value == "POST-MARGIN"
    assert ws.cell(3, 18).value == "X"
    assert ws.cell(3, 21).value == "X"
    wb.close()


def test_okd_consumer_builds_reporting_helpers_from_w_only(tmp_path: Path) -> None:
    source = tmp_path / "x-margin.xlsx"
    output = tmp_path / "okd.xlsx"
    wb = _weekly_main_with_x_margins()
    wb.save(source)
    wb.close()

    result = build_okd_sheets(source, output, source_sheet="main")
    assert result[1] == 2  # week count

    wb = load_workbook(output, data_only=False)
    progress = wb["progress"]
    week_starts = [progress.cell(row, 1).value for row in range(2, progress.max_row + 1)]
    assert len(week_starts) == 2
    # No helper row is sourced from the X margin dates.
    formulas = [str(progress.cell(row, 1).value or "") for row in range(2, progress.max_row + 1)]
    assert all("R4" not in formula and "U4" not in formula for formula in formulas)
    wb.close()


def test_monthly_reporting_cells_inherit_percent_format_from_w_not_x_margin() -> None:
    wb = _weekly_main_with_x_margins()
    ws = wb["main"]

    # X columns intentionally stay General while real reporting W columns carry
    # the percentage presentation contract used by Create Progress.
    for row in (7, 9, 10):
        ws.cell(row, 18).number_format = "General"
        ws.cell(row, 19).number_format = "0.00%"
        ws.cell(row, 20).number_format = "0.00%"
        ws.cell(row, 21).number_format = "General"

    build_monthly_main_view(wb)
    monthly = wb["main_monthly"]

    # February/March are reporting M1/M2 and must display as percentages rather
    # than raw Excel fractions inherited from the leading X margin column.
    assert monthly.cell(7, 19).number_format == "0.00%"
    assert monthly.cell(7, 20).number_format == "0.00%"
    assert monthly.cell(9, 19).number_format == "0.00%"
    assert monthly.cell(10, 20).number_format == "0.00%"
    wb.close()


def test_monthly_overlay_geometry_maps_to_monthly_sheet_after_x_margin() -> None:
    from types import SimpleNamespace

    from progress_studio.infrastructure.excel.traditional_overlay_workbook import _monthly_project_window

    wb = Workbook()
    data = wb.active
    data.title = "Dashboard_Data"
    monthly = wb.create_sheet("main_monthly")

    # Dashboard_Data owns reporting months only: Apr -> Jun.
    for row, (period_date, plan) in enumerate(
        [(date(2026, 4, 24), 0.1), (date(2026, 5, 29), 0.5), (date(2026, 6, 4), 1.0)],
        start=2,
    ):
        data.cell(row, 4).value = period_date
        data.cell(row, 5).value = plan

    # main_monthly keeps a March X display margin before M1.  Its physical
    # columns therefore no longer line up with the first weekly reporting col.
    for col, (label, period_date) in enumerate(
        [
            ("X", date(2026, 3, 27)),
            ("M1", date(2026, 4, 24)),
            ("M2", date(2026, 5, 29)),
            ("M3", date(2026, 6, 25)),
            ("X", date(2026, 7, 31)),
        ],
        start=18,
    ):
        monthly.cell(3, col).value = label
        monthly.cell(4, col).value = period_date

    dataset = SimpleNamespace(
        periods=(
            SimpleNamespace(column=21, reporting_date=date(2026, 4, 24)),
            SimpleNamespace(column=22, reporting_date=date(2026, 5, 1)),
            SimpleNamespace(column=23, reporting_date=date(2026, 6, 4)),
        )
    )

    first_row, last_row, first_col, last_col = _monthly_project_window(data, dataset, monthly)
    assert (first_row, last_row) == (2, 4)
    assert (first_col, last_col) == (19, 21)  # M1 through M3, not shifted to U/W.
    wb.close()


def test_live_rebuild_monthly_collapses_weekly_x_margin_into_month_buckets(tmp_path: Path) -> None:
    from progress_studio.services.rebuild_service import WorkbookRebuildEngine

    source = tmp_path / "live-x-source.xlsx"
    output = tmp_path / "live-x-output.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "main"
    headers = [
        "Row Type", "WBS", "Description", "P/A", "Activity ID", "Task ID", "UID",
        "Outline Level", "Plan Start", "Plan Finish", "Actual Start", "Actual Finish",
        "% Complete", "Physical %", "Amount", "Total Float (hr)", "XML Amount",
    ]
    for col, value in enumerate(headers, start=1):
        ws.cell(4, col).value = value

    # Physical weekly display range: March margin, early-April margin, then W1/W2.
    # Rebuild's MainDataset intentionally reads W periods only, but main_monthly
    # must still regenerate one monthly X bucket for March instead of preserving
    # the copied weekly X columns one-by-one.
    weekly = [
        ("X", datetime(2026, 3, 20)),
        ("X", datetime(2026, 3, 27)),
        ("X", datetime(2026, 4, 3)),
        ("X", datetime(2026, 4, 10)),
        ("W1", datetime(2026, 4, 24)),
        ("W2", datetime(2026, 5, 29)),
    ]
    for col, (label, reporting_date) in enumerate(weekly, start=18):
        ws.cell(3, col).value = label
        ws.cell(4, col).value = reporting_date

    rows = [
        ["Project Summary", "", "Project", "P", "", "", "", 0, date(2026, 4, 17), date(2026, 5, 29), None, None, 1.0, 0.0, 1000],
        ["", "", "", "A", "", "", "", 0, None, None, None, None, 0.0, 0.0, 0],
        ["Activity", "1.1", "Task", "P", "A1000", "", "", 1, date(2026, 4, 17), date(2026, 5, 29), None, None, 1.0, 0.0, 1000],
        ["", "", "", "A", "A1000", "", "", 1, None, None, None, None, 0.0, 0.0, 0],
        ["S-Curve", "", "Plan", "P", "", "", "", 0],
        ["S-Curve", "", "Acc. Plan", "AP", "", "", "", 0],
        ["S-Curve", "", "Actual", "A", "", "", "", 0],
        ["S-Curve", "", "Acc. Actual", "AA", "", "", "", 0],
    ]
    for row_idx, values in enumerate(rows, start=5):
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row_idx, col_idx).value = value

    # Reporting values under W1/W2 only.
    for row in (5, 7, 9):
        ws.cell(row, 22).value = 0.4
        ws.cell(row, 23).value = 0.6
        ws.cell(row, 22).number_format = "0.00%"
        ws.cell(row, 23).number_format = "0.00%"
    for row in (10,):
        ws.cell(row, 22).value = 0.4
        ws.cell(row, 23).value = 1.0
        ws.cell(row, 22).number_format = "0.00%"
        ws.cell(row, 23).number_format = "0.00%"

    wb.save(source)
    wb.close()

    WorkbookRebuildEngine().rebuild_live_progress(source, output)

    wb = load_workbook(output, data_only=False)
    try:
        monthly = wb["main_monthly"]
        # One month = one physical column. March is X-only; April contains W1
        # and therefore becomes M1 even though early-April weekly cells were X.
        assert [monthly.cell(2, col).value for col in range(18, 21)] == [
            "March", "April", "May"
        ]
        assert [monthly.cell(3, col).value for col in range(18, 21)] == [
            "X", "M1", "M2"
        ]
        assert [monthly.cell(4, col).value for col in range(18, 21)] == [
            datetime(2026, 3, 27), datetime(2026, 4, 24), datetime(2026, 5, 29)
        ]
        # No copied weekly X columns survive to the right of the rebuilt monthly range.
        assert monthly.max_column == 20
        assert monthly.cell(5, 18).value in (None, "")
        assert "'main'!V5:V5" in str(monthly.cell(5, 19).value)
        assert "'main'!W5:W5" in str(monthly.cell(5, 20).value)
    finally:
        wb.close()
