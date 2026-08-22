from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from tests._paths import REPO_ROOT

from openpyxl import load_workbook

from progress_studio.infrastructure.excel.import_workbook_writer import ImportWorkbookWriter
from progress_studio.infrastructure.excel.okd_workbook import build_progress_table_sheet, HEADER_ROW
from progress_studio.infrastructure.primavera.xml_reader import PrimaveraXmlReader


PROJECT_ROOT = REPO_ROOT
EXAMPLE_XML = PROJECT_ROOT / "example" / "example.xml"


class TestV202Bugfixes(unittest.TestCase):
    def test_activity_rows_receive_sequential_child_wbs_codes(self) -> None:
        project_name, rows = PrimaveraXmlReader().read(EXAMPLE_XML)
        with tempfile.TemporaryDirectory() as tmp:
            output = Path(tmp) / "import.xlsx"
            ImportWorkbookWriter().write(output, EXAMPLE_XML, project_name, rows)
            wb = load_workbook(output, data_only=False)
            ws = wb["main"]

            headers = {ws.cell(1, col).value: col for col in range(1, ws.max_column + 1)}
            row_type_col = headers["Row Type"]
            wbs_col = headers["WBS"]
            activity_id_col = headers["Activity ID"]

            checked = 0
            for row in range(2, ws.max_row + 1):
                if ws.cell(row, row_type_col).value != "Activity":
                    continue
                self.assertNotIn(ws.cell(row, wbs_col).value, (None, ""))
                activity_id = ws.cell(row, activity_id_col).value
                if activity_id == "A1000":
                    self.assertEqual(ws.cell(row, wbs_col).value, "1.1.1")
                elif activity_id == "A1005":
                    self.assertEqual(ws.cell(row, wbs_col).value, "1.2.1")
                elif activity_id == "A1010":
                    self.assertEqual(ws.cell(row, wbs_col).value, "1.2.2")
                checked += 1

            self.assertEqual(checked, 172)
            wb.close()

    def test_progress_table_week_headers_use_date_format(self) -> None:
        from openpyxl import Workbook
        from datetime import datetime

        wb = Workbook()
        source = wb.active
        source.title = "main"
        headers = ["Row Type", "Description", "P/A", "Activity ID", "Amount", "Plan Start", "Plan Finish"]
        for col, value in enumerate(headers, start=1):
            source.cell(HEADER_ROW, col, value)
        source.cell(HEADER_ROW, 8, datetime(2026, 1, 23))
        source.cell(HEADER_ROW, 8).number_format = "dd/mm/yyyy"
        source.cell(HEADER_ROW + 1, 1, "Activity")
        source.cell(HEADER_ROW + 1, 2, "Test")
        source.cell(HEADER_ROW + 1, 3, "P")
        source.cell(HEADER_ROW + 1, 4, "A1")
        source.cell(HEADER_ROW + 1, 5, 100)
        source.cell(HEADER_ROW + 1, 6, datetime(2026, 1, 1))
        source.cell(HEADER_ROW + 1, 7, datetime(2026, 1, 31))
        source.cell(HEADER_ROW + 1, 8, 0.1)
        source.cell(HEADER_ROW + 2, 3, "A")
        source.cell(HEADER_ROW + 2, 8, 0.05)

        header_map = {str(source.cell(HEADER_ROW, col).value).strip().lower(): col for col in range(1, source.max_column + 1)}
        weeks = [(8, datetime(2026, 1, 23).date())]
        table_rows = [{"display_wbs": "1.1", "plan_row": HEADER_ROW + 1, "actual_row": HEADER_ROW + 2}]
        build_progress_table_sheet(wb, source, header_map, weeks, table_rows)

        table = wb["progress_table"]
        self.assertEqual(table.cell(1, 6).value, datetime(2026, 1, 23).date())
        self.assertEqual(table.cell(1, 6).number_format, "dd/mm/yyyy")
        # progress_table is a rebuild snapshot: both Plan and Actual are values.
        self.assertEqual(table.cell(2, 6).value, 10.0)
        self.assertEqual(table.cell(2, 5).value, 10.0)
        self.assertEqual(table.cell(3, 6).value, 5.0)
        self.assertEqual(table.cell(3, 5).value, 5.0)


if __name__ == "__main__":
    unittest.main()


def test_boq_formula_fallback_reads_linked_project_value(tmp_path):
    from openpyxl import Workbook, load_workbook
    from progress_studio.infrastructure.excel.mapping_reader import _formula_fallback_value

    path = tmp_path / "boq.xlsx"
    workbook = Workbook()
    source = workbook.active
    source.title = "ARCH"
    source["K5"] = 123.45
    project = workbook.create_sheet("Project")
    project["K5"] = '=IF(ARCH!K5="","",ARCH!K5)'
    workbook.save(path)

    values = load_workbook(path, data_only=True)
    formulas = load_workbook(path, data_only=False)
    try:
        assert _formula_fallback_value(
            values["Project"]["K5"].value,
            formulas["Project"]["K5"].value,
            values,
        ) == 123.45
    finally:
        values.close()
        formulas.close()
