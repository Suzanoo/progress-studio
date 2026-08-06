from datetime import date

from openpyxl import Workbook

from progress_studio.infrastructure.excel.activity_data_theme import (
    apply_activity_data_wbs_hierarchy,
)


def _sheet():
    wb = Workbook()
    ws = wb.active
    ws.title = "main"
    ws.append([])
    ws.append([])
    ws.append([])
    ws.append(["Row Type", "WBS", "Description", "P/A", "Outline Level", "Amount", date(2026, 1, 2)])
    ws.append(["Project Summary", "", "Project", "P", 0, 100, None])
    ws.append(["WBS", "1", "Level 1", "P", 1, 100, None])
    ws.append([None, None, None, "A", None, None, None])
    ws.append(["WBS", "1.1", "Level 2", "P", 2, 100, None])
    ws.append([None, None, None, "A", None, None, None])
    ws.append(["WBS", "1.1.1", "Level 3", "P", 3, 100, None])
    ws.append([None, None, None, "A", None, None, None])
    ws.append(["WBS", "1.1.1.1", "Level 4", "P", 4, 100, None])
    ws.append([None, None, None, "A", None, None, None])
    ws.append(["WBS", "1.1.1.1.1", "Level 5", "P", 5, 100, None])
    ws.append([None, None, None, "A", None, None, None])
    ws.append(["Activity", "A1000", "Task", "P", 6, 100, None])
    return ws


def _border_signature(cell):
    return tuple(
        side.style
        for side in (
            cell.border.left,
            cell.border.right,
            cell.border.top,
            cell.border.bottom,
        )
    )


def test_activity_data_hierarchy_supports_four_levels_and_clamps_deeper_levels():
    ws = _sheet()
    project_fill = ws.cell(5, 1).fill.fgColor.rgb
    activity_fill = ws.cell(16, 1).fill.fgColor.rgb
    timescale_fills = [ws.cell(row, 7).fill.fgColor.rgb for row in range(6, 16)]
    borders = {row: _border_signature(ws.cell(row, 1)) for row in range(6, 16)}

    apply_activity_data_wbs_hierarchy(ws)

    expected = {
        6: "F4B183",
        7: "F4B183",
        8: "F8CBAD",
        9: "F8CBAD",
        10: "FADBC8",
        11: "FADBC8",
        12: "FCE8DE",
        13: "FCE8DE",
        14: "FCE8DE",
        15: "FCE8DE",
    }
    for row, color in expected.items():
        assert ws.cell(row, 1).fill.fgColor.rgb.endswith(color)
        assert ws.cell(row, 1).font.bold is True
        assert _border_signature(ws.cell(row, 1)) == borders[row]

    assert ws.cell(5, 1).fill.fgColor.rgb == project_fill
    assert ws.cell(16, 1).fill.fgColor.rgb == activity_fill
    assert [ws.cell(row, 7).fill.fgColor.rgb for row in range(6, 16)] == timescale_fills
