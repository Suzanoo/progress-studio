from pathlib import Path
import shutil

from openpyxl import load_workbook


def test_progress_service_creates_table_theme_and_dashboard_before_mapping(tmp_path, monkeypatch):
    from progress_studio.services.progress_service import ProgressService
    import progress_studio.services.progress_service as module

    source = tmp_path / "source.xlsx"
    output = tmp_path / "progress.xlsx"
    shutil.copyfile(Path("example/golden/progress.xlsx"), source)

    # Reproduce the real lifecycle bug: the progress builder creates only the
    # progress sheet. progress_table must be created by ProgressService itself,
    # not faked by this test and not deferred to the OKD step.
    def fake_prepare(workbook, ws):
        if "progress" in workbook.sheetnames:
            del workbook["progress"]
        if "progress_table" in workbook.sheetnames:
            del workbook["progress_table"]
        progress = workbook.create_sheet("progress")
        progress.append(["project_start", "project_finish", "week_start", "plan", "actual"])
        progress.append([None, None, ws.cell(4, 13).value, 0, 0])
        return (0, 0, 0, 0, 0)

    calls = []
    monkeypatch.setattr(module, "prepare_progress_and_scurve", fake_prepare)
    original_theme = module.apply_activity_data_wbs_hierarchy
    original_dashboard = module.build_dashboard

    def wrapped_theme(ws):
        calls.append("theme")
        return original_theme(ws)

    def wrapped_dashboard(workbook, *, project_name=None):
        calls.append("dashboard")
        return original_dashboard(workbook, project_name=project_name)

    monkeypatch.setattr(module, "apply_activity_data_wbs_hierarchy", wrapped_theme)
    monkeypatch.setattr(module, "build_dashboard", wrapped_dashboard)

    ProgressService().build(source, output)

    assert calls == ["theme", "dashboard"]
    result = load_workbook(output, data_only=False)
    try:
        assert result.sheetnames[0] == "Dashboard"
        assert {"Dashboard_Data", "progress", "progress_table"}.issubset(result.sheetnames)
        assert result["progress_table"].max_row > 1
        assert result["Dashboard"]["C5"].value not in (None, "")
        # WBS hierarchy theme is already present in the pre-mapping main sheet.
        main = result["main"]
        wbs_row = next(
            row for row in range(5, main.max_row + 1)
            if str(main.cell(row, 1).value or "").strip().lower() == "wbs"
        )
        assert main.cell(wbs_row, 1).fill.fill_type == "solid"
    finally:
        result.close()


def test_full_generation_has_no_late_dashboard_step():
    from progress_studio.services.workbook_generation_service import WorkbookGenerationService

    service = WorkbookGenerationService()
    assert not hasattr(service, "dashboard")
