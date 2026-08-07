from datetime import date

from openpyxl import Workbook

from progress_studio.infrastructure.excel.dashboard_workbook import (
    DASHBOARD_SHEET,
    DATA_SHEET,
    build_dashboard,
)


def _workbook():
    wb = Workbook()
    progress = wb.active
    progress.title = "progress"
    progress.append(["project_start", "project_finish", "week_start", "plan", "actual"])
    progress.append([date(2026, 1, 1), date(2026, 2, 1), date(2026, 1, 2), 10, 5])
    progress.append([date(2026, 1, 1), date(2026, 2, 1), date(2026, 1, 9), 25, 12])
    table = wb.create_sheet("progress_table")
    table.append(["WBS", "Activities", "Amount", "P/A", "%Progress", date(2026, 1, 2), date(2026, 1, 9)])
    table.append(["1.1", "Activity A", 1000, "P", 25, 10, 15])
    table.append(["1.1", "Activity A", 1000, "A", 12, 5, 7])
    return wb


def test_dashboard_is_created_as_first_separate_sheet_with_controls_and_chart():
    wb = _workbook()
    build_dashboard(wb, project_name="Demo Project")

    assert wb.sheetnames[0] == DASHBOARD_SHEET
    assert DATA_SHEET in wb.sheetnames
    assert wb[DATA_SHEET].sheet_state == "hidden"
    dashboard = wb[DASHBOARD_SHEET]
    assert dashboard["B2"].value == "PROGRESS STUDIO DASHBOARD"
    assert dashboard["C5"].value == "Demo Project"
    assert dashboard["G5"].value == "Weekly"
    assert dashboard["K5"].value == date(2026, 1, 9)
    assert len(dashboard.data_validations.dataValidation) == 2
    assert len(dashboard._charts) == 1


def test_dashboard_activity_rows_are_formula_linked_to_progress_table():
    wb = _workbook()
    build_dashboard(wb)
    dashboard = wb[DASHBOARD_SHEET]

    assert dashboard["B39"].value == "='progress_table'!A2"
    assert dashboard["C39"].value == "='progress_table'!B2"
    assert dashboard["F39"].value == "Plan"
    assert dashboard["F40"].value == "Actual"
    assert "SUMPRODUCT" in dashboard["K39"].value


def test_dashboard_data_is_populated_from_real_progress_headers_and_string_dates():
    wb = Workbook()
    progress = wb.active
    progress.title = "progress"
    progress.append(["project_start", "project_finish", "week_start", "plan", "actual"])
    progress.append(["2026-02-23", "2027-05-31", "2026-02-27", 0.09, None])
    progress.append(["2026-02-23", "2027-05-31", "2026-03-06", 0.20, 0.05])
    table = wb.create_sheet("progress_table")
    table.append(["WBS", "Activities", "Amount", "P/A", "%Progress", date(2026, 2, 27)])
    table.append(["1", "Test", 100, "P", 0.2, 0.2])
    table.append(["1", "Test", 100, "A", 0.05, 0.05])

    build_dashboard(wb)

    data = wb[DATA_SHEET]
    assert data["A2"].value == date(2026, 2, 27)
    assert data["B2"].value == 0.09
    assert data["A3"].value == date(2026, 3, 6)
    assert data["C3"].value == 0.05
    assert wb[DASHBOARD_SHEET]["K5"].value == date(2026, 3, 6)
