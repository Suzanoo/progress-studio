from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from progress_studio.domain.mapping_models import ActivityRow, BOQRow
from progress_studio.infrastructure.excel.mapping_reader import BOQSheetReader
from progress_studio.services.mapping_store import MappingStore
from progress_studio.services.workbook_export_service import WorkbookExportService


def make_progress(path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Amount Mapping"
    ws.append(["Activity ID", "WBS", "Description", "Amount", "Status"])
    ws.append([None, "1", "Project", None, None])
    ws.append(["A1000", "1.1", "First activity", 99, "OLD"])
    ws.append(["A2000", "1.2", "Second activity", 99, "OLD"])
    main = wb.create_sheet("main")
    main.append(["Progress Studio test workbook"])
    main.append([])
    main.append([])
    main.append(["Row Type", "Activity ID", "Description", "P/A", "Outline Level", "Amount"])
    main.append(["Project Summary", "", "Project", "P", 0, 0])
    main.append(["WBS", "", "Group", "P", 1, 0])
    main.append(["Activity", "A1000", "First activity", "P", 2, 99])
    main.append(["Activity", "A1000", "First activity", "A", 2, None])
    main.append(["Activity", "A2000", "Second activity", "P", 2, 99])
    main.append(["Activity", "A2000", "Second activity", "A", 2, None])
    formula = wb.create_sheet("Formula Check")
    formula["A1"] = "=SUM(main!F7,F9)"
    wb.save(path)
    return path


def make_store() -> MappingStore:
    store = MappingStore()
    store.load_activities([
        ActivityRow("A1000", "1", "1.1", "First activity"),
        ActivityRow("A2000", "1", "1.2", "Second activity"),
    ])
    store.load_boq([
        BOQRow("Sheet|10|2", "Sheet", 10, "W2", "W3", "W4", "Item one", 1000, "BOQ-ONE"),
        BOQRow("Sheet|11|3", "Sheet", 11, "W2", "W3", "W4", "Item two", 500, "BOQ-TWO"),
    ])
    store.selected_activity_ids = {"A1000"}
    store.selected_boq_ids = {"Sheet|10|2"}
    store.map_selected(40)
    store.selected_activity_ids = {"A2000"}
    store.selected_boq_ids = {"Sheet|10|2", "Sheet|11|3"}
    store.map_selected(60)
    return store


def test_export_updates_amount_mapping_and_writes_reconciliation(tmp_path: Path) -> None:
    source = make_progress(tmp_path / "progress.xlsx")
    output = tmp_path / "progress_mapped.xlsx"
    result = WorkbookExportService().export(source, output, make_store())

    assert result.amount_rows_updated == 2
    assert result.mapping_rows_written == 3
    assert result.validation.partial_boq_count == 1
    assert result.validation.full_boq_count == 1
    assert result.validation.allocated_amount == 1300

    wb = load_workbook(output, data_only=False)
    try:
        main = wb["main"]
        assert main["F7"].value == 400
        assert main["F8"].value is None
        assert main["F9"].value == 900
        assert main["F10"].value is None
        assert str(main["F6"].value).startswith("=SUMIFS(")
        assert str(main["F5"].value).startswith("=SUMIFS(")

        amount = wb["Amount Mapping"]
        assert amount["D3"].value == 400
        assert amount["D4"].value == 900
        assert amount["E3"].value == "MAPPED"
        assert wb["Formula Check"]["A1"].value == "=SUM(main!F7,F9)"

        mapping = wb["BOQ Activity Mapping"]
        assert mapping.max_row == 4
        assert mapping["M2"].value in {"BOQ-ONE", "BOQ-TWO"}
        assert mapping.tables

        summary = wb["Mapping Summary"]
        assert summary["B5"].value == "Partial"
        assert summary["B14"].value == 1300
        assert wb.calculation.fullCalcOnLoad is True
    finally:
        wb.close()


def test_export_refuses_source_overwrite_and_existing_output_by_default(tmp_path: Path) -> None:
    source = make_progress(tmp_path / "progress.xlsx")
    service = WorkbookExportService()
    store = make_store()
    with pytest.raises(ValueError, match="different"):
        service.export(source, source, store, overwrite=True)

    output = tmp_path / "existing.xlsx"
    output.write_bytes(b"keep")
    with pytest.raises(FileExistsError):
        service.export(source, output, store)
    assert output.read_bytes() == b"keep"
    assert not list(tmp_path.glob("*.tmp.xlsx"))


def test_validation_reports_complete_mapping() -> None:
    store = make_store()
    store.selected_activity_ids = {"A1000"}
    store.selected_boq_ids = {"Sheet|11|3"}
    store.map_selected(40)
    validation = WorkbookExportService.validate(store)
    assert validation.is_complete
    assert validation.remaining_amount == 0
    assert validation.allocated_percent == 100


def test_boq_reader_creates_repeatable_stable_export_id(tmp_path: Path) -> None:
    path = tmp_path / "boq.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "BOQ"
    ws.append(["WBS-2", "WBS-3", "WBS-4", "Description", "Amount", "Source Sheet", "Source Row"])
    ws.append(["A", "B", "C", "Concrete", 100, "Original", 55])
    wb.save(path)

    first = BOQSheetReader().read(path, "BOQ")[0]
    second = BOQSheetReader().read(path, "BOQ")[0]
    assert first.stable_id.startswith("BOQ-")
    assert first.stable_id == second.stable_id


def test_exported_mapping_table_has_single_table_owned_filter(tmp_path: Path) -> None:
    from zipfile import ZipFile
    import xml.etree.ElementTree as ET

    source = make_progress(tmp_path / "progress.xlsx")
    output = tmp_path / "progress_mapped.xlsx"
    WorkbookExportService().export(source, output, make_store())

    ns = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    with ZipFile(output) as archive:
        table_xml = ET.fromstring(archive.read("xl/tables/table1.xml"))
        assert table_xml.find("x:autoFilter", ns) is not None

        workbook_xml = ET.fromstring(archive.read("xl/workbook.xml"))
        rel_ns = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
        mapping_rel_id = None
        for sheet in workbook_xml.find("x:sheets", ns):
            if sheet.attrib["name"] == "BOQ Activity Mapping":
                mapping_rel_id = sheet.attrib[f"{{{rel_ns}}}id"]
                break
        assert mapping_rel_id is not None

        package_ns = {"r": "http://schemas.openxmlformats.org/package/2006/relationships"}
        rels = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
        target = next(
            rel.attrib["Target"] for rel in rels.findall("r:Relationship", package_ns)
            if rel.attrib["Id"] == mapping_rel_id
        )
        sheet_path = target.lstrip("/")
        if not sheet_path.startswith("xl/"):
            sheet_path = f"xl/{sheet_path}"
        sheet_xml = ET.fromstring(archive.read(sheet_path))
        assert sheet_xml.find("x:autoFilter", ns) is None
        assert sheet_xml.find("x:tableParts", ns) is not None


def test_export_without_allocations_writes_headers_without_table(tmp_path: Path) -> None:
    source = make_progress(tmp_path / "progress.xlsx")
    output = tmp_path / "empty_mapped.xlsx"
    store = MappingStore()
    store.load_activities([
        ActivityRow("A1000", "1", "1.1", "First activity"),
        ActivityRow("A2000", "1", "1.2", "Second activity"),
    ])
    store.load_boq([
        BOQRow("Sheet|10|2", "Sheet", 10, "W2", "W3", "W4", "Item one", 1000, "BOQ-ONE"),
    ])

    result = WorkbookExportService().export(source, output, store)
    assert result.mapping_rows_written == 0
    wb = load_workbook(output)
    try:
        mapping = wb["BOQ Activity Mapping"]
        assert mapping.max_row == 1
        assert not mapping.tables
        assert mapping.auto_filter.ref is None
    finally:
        wb.close()


def test_package_validator_rejects_worksheet_filter_on_table_sheet(tmp_path: Path) -> None:
    from zipfile import ZIP_DEFLATED, ZipFile
    import xml.etree.ElementTree as ET

    from progress_studio.infrastructure.excel.xlsx_package_validator import (
        WorkbookPackageValidationError,
        validate_xlsx_tables,
    )

    source = make_progress(tmp_path / "progress.xlsx")
    valid = tmp_path / "valid.xlsx"
    broken = tmp_path / "broken.xlsx"
    WorkbookExportService().export(source, valid, make_store())

    main_ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    ET.register_namespace("", main_ns)
    with ZipFile(valid) as src, ZipFile(broken, "w", ZIP_DEFLATED) as dst:
        injected = False
        for info in src.infolist():
            payload = src.read(info.filename)
            if info.filename.startswith("xl/worksheets/sheet") and info.filename.endswith(".xml"):
                root = ET.fromstring(payload)
                table_parts = root.find(f"{{{main_ns}}}tableParts")
                if table_parts is not None:
                    auto_filter = ET.Element(f"{{{main_ns}}}autoFilter", {"ref": "A1:M4"})
                    root.insert(list(root).index(table_parts), auto_filter)
                    payload = ET.tostring(root, encoding="utf-8", xml_declaration=True)
                    injected = True
            dst.writestr(info, payload)
        assert injected

    with pytest.raises(WorkbookPackageValidationError, match="AutoFilter overlaps"):
        validate_xlsx_tables(broken)


def test_progress_reader_rejects_missing_main_sheet(tmp_path: Path) -> None:
    path = tmp_path / "renamed_main.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "schedule"
    amount = wb.create_sheet("Amount Mapping")
    amount.append(["Activity ID", "WBS", "Description", "Amount", "Status"])
    amount.append(["A1000", "1.1", "Activity", 0, "UNMAPPED"])
    wb.save(path)
    wb.close()

    with pytest.raises(ValueError, match='rename it back to "main"'):
        from progress_studio.infrastructure.excel.mapping_reader import ProgressActivityReader
        ProgressActivityReader().read(path)


def test_export_rechecks_main_workbook_contract(tmp_path: Path) -> None:
    source = make_progress(tmp_path / "progress.xlsx")
    wb = load_workbook(source)
    wb["main"].title = "renamed-main"
    wb.save(source)
    wb.close()

    with pytest.raises(ValueError, match='Required worksheet "main"'):
        WorkbookExportService().export(source, tmp_path / "out.xlsx", make_store())


def test_export_forces_excel_to_rebuild_formula_results(tmp_path: Path) -> None:
    import zipfile
    from xml.etree import ElementTree as ET

    source = make_progress(tmp_path / "progress.xlsx")
    output = tmp_path / "progress_mapped.xlsx"
    WorkbookExportService().export(source, output, make_store())

    wb = load_workbook(output, data_only=False)
    try:
        assert wb.calculation.calcMode == "auto"
        assert wb.calculation.fullCalcOnLoad is True
        assert wb.calculation.forceFullCalc is True
        assert wb.calculation.calcId == 0
    finally:
        wb.close()

    with zipfile.ZipFile(output) as package:
        assert "xl/calcChain.xml" not in package.namelist()
        root = ET.fromstring(package.read("xl/workbook.xml"))
        namespace = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
        calc = root.find("x:calcPr", namespace)
        assert calc is not None
        assert calc.attrib["calcMode"] == "auto"
        assert calc.attrib["fullCalcOnLoad"] == "1"
        assert calc.attrib["forceFullCalc"] == "1"
        assert calc.attrib["calcId"] == "0"
