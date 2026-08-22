
from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

from progress_studio.infrastructure.excel.rebuild_workbook_reader import RebuildWorkbookReader
from progress_studio.services.payment_progress_adapter import MainDatasetPaymentProgressAdapter
from progress_studio.services.rebuild_service import WorkbookRebuildEngine


def _fixture(path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "main"
    headers = [
        "Row Type", "WBS", "Description", "P/A", "Activity ID",
        "Outline Level", "Plan Start", "Plan Finish", "Amount",
    ]
    for col, value in enumerate(headers, 1):
        ws.cell(4, col, value)
    for col, dt in enumerate(
        (datetime(2026,3,6), datetime(2026,3,13), datetime(2026,3,20)),
        start=10,
    ):
        ws.cell(3, col, f"W{col-9}")
        ws.cell(4, col, dt)

    ws.append([
        "Activity", "1.1", "Concrete", "P", "A1000", 2,
        datetime(2026,3,1), datetime(2026,3,20), 1000,
        .20, .30, .50,
    ])
    # Legacy Actual row may be sparse/blank.
    ws.append([
        "", "", "", "A", "A1000", 2,
        datetime(2026,3,1), datetime(2026,3,20), 0,
        .10, .10, None,
    ])

    pin = wb.create_sheet("Payment Input")
    pin["A6"] = "Type"
    pin["B6"] = "WBS"
    pin["C6"] = "Activity ID"
    pin["D6"] = "Activity Name"
    pin["E6"] = "P1"
    pin["E7"] = datetime(2026,3,20)
    pin["A8"] = "ACT"
    pin["B8"] = "1.1"
    pin["C8"] = "A1000"
    pin["D8"] = "Concrete"
    pin["E8"] = 0.50

    # Existing Progress views must be preserved by Payment-only rebuild.
    dash = wb.create_sheet("Dashboard")
    dash["A1"] = "KEEP_DASHBOARD"
    monthly = wb.create_sheet("main_monthly")
    monthly["A1"] = "KEEP_MONTHLY"

    wb.save(path)
    return path


def test_lw9_payment_progress_adapter_reuses_main_dataset(tmp_path: Path) -> None:
    source = _fixture(tmp_path / "source.xlsx")
    dataset = RebuildWorkbookReader().read_main_dataset(source)
    index = MainDatasetPaymentProgressAdapter().build(dataset, source)

    activity = index.activities["A1000"]
    assert activity.row_number == 5
    assert len(activity.buckets) == 3
    assert abs(activity.buckets[0].cumulative_fraction - 0.20) < 1e-12
    assert abs(activity.buckets[1].cumulative_fraction - 0.50) < 1e-12
    assert activity.buckets[-1].cumulative_fraction == 1.0


def test_lw9_live_payment_rebuild_preserves_progress_views(tmp_path: Path) -> None:
    source = _fixture(tmp_path / "source.xlsx")
    output = tmp_path / "live_payment.xlsx"

    result = WorkbookRebuildEngine().rebuild_live_payment(source, output)

    assert result.rendered_periods == 1
    assert result.rendered_points == 1
    assert result.period_ids == ("P1",)

    wb = load_workbook(output, data_only=False)
    try:
        assert "Payment" in wb.sheetnames
        assert wb["Dashboard"]["A1"].value == "KEEP_DASHBOARD"
        assert wb["main_monthly"]["A1"].value == "KEEP_MONTHLY"
        assert wb["Payment Input"]["C8"].value == "A1000"
        assert wb.calculation.calcMode == "manual"
        assert wb.calculation.calcOnSave is True
    finally:
        wb.close()


def test_lw9_live_payment_writer_has_one_open_and_one_save() -> None:
    source = Path("progress_studio/services/rebuild_service.py").read_text(encoding="utf-8")
    start = source.index("    def rebuild_live_payment(")
    end = source.index("    def rebuild_payment(", start)
    block = source[start:end]
    assert block.count("load_workbook(") == 1
    assert block.count(".save(") == 1
    assert block.count("finalize_workbook(") == 1
    assert "data_only=True" not in block
    assert "render_periods_into_workbook" in block
    assert "save_path=temp_path" not in block
    assert "finalize_mode" not in block


def test_lw9_in_memory_payment_renderer_is_render_only() -> None:
    source = Path(
        "progress_studio/infrastructure/excel/payment_line_renderer.py"
    ).read_text(encoding="utf-8")
    start = source.index("    def render_periods_into_workbook(")
    end = source.index("    def render_single_period(", start)
    block = source[start:end]
    assert "load_workbook" not in block
    assert block.count(".save(") == 0
    assert "finalize_workbook(" not in block


def test_lw9_ui_routes_live_payment_to_live_engine() -> None:
    source = Path("progress_studio/presentation/gui/rebuild.py").read_text(encoding="utf-8")
    assert "rebuild_live_payment" in source
    assert "Live Payment active in LW-9" in source
    assert "Live Payment is not active yet" not in source
