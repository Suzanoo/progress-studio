from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from progress_studio.infrastructure.excel.payment_workbook import PaymentWorkbookSnapshotter


def _workbook(path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "main"
    ws["A1"] = "Project"
    ws["A4"] = "WBS"
    ws["B4"] = "Activity ID"
    ws["C4"] = "Activity Name"
    ws["B5"] = "A1000"
    ws["C5"] = "Mobilization"
    ws["B6"] = "A1010"
    ws["C6"] = "Site Preparation"
    ws.freeze_panes = "D5"
    ws.auto_filter.ref = "A4:C6"
    ws.column_dimensions["C"].width = 28
    ws.row_dimensions[5].height = 24
    ws.merge_cells("D1:F1")
    wb.create_sheet("progress")
    wb.save(path)
    wb.close()
    return path


def test_validate_progress_workbook(tmp_path: Path) -> None:
    source = _workbook(tmp_path / "source.xlsx")
    result = PaymentWorkbookSnapshotter().validate(source)
    assert result.main_sheet == "main"
    assert result.activity_rows == 2
    assert result.max_row == 6


def test_create_payment_snapshot_preserves_main_and_other_sheets(tmp_path: Path) -> None:
    source = _workbook(tmp_path / "source.xlsx")
    output = tmp_path / "output.xlsx"
    result = PaymentWorkbookSnapshotter().create_snapshot(source, output)

    assert result.payment_sheet == "Payment"
    wb = load_workbook(output, data_only=False)
    try:
        assert "main" in wb.sheetnames
        assert "Payment" in wb.sheetnames
        assert "progress" in wb.sheetnames
        main = wb["main"]
        payment = wb["Payment"]
        assert payment["B5"].value == main["B5"].value == "A1000"
        assert payment["C6"].value == "Site Preparation"
        assert payment.freeze_panes == "D5"
        assert payment.auto_filter.ref == "A4:C6"
        assert payment.column_dimensions["C"].width == main.column_dimensions["C"].width
        assert payment.row_dimensions[5].height == main.row_dimensions[5].height
        assert "D1:F1" in {str(rng) for rng in payment.merged_cells.ranges}
    finally:
        wb.close()


def test_existing_payment_sheet_is_replaced(tmp_path: Path) -> None:
    source = _workbook(tmp_path / "source.xlsx")
    wb = load_workbook(source)
    old = wb.create_sheet("Payment")
    old["A1"] = "OLD"
    wb.save(source)
    wb.close()

    output = tmp_path / "output.xlsx"
    result = PaymentWorkbookSnapshotter().create_snapshot(source, output)
    assert result.replaced_existing_sheet is True
    wb = load_workbook(output)
    try:
        assert wb["Payment"]["A1"].value == "Project"
        assert wb.sheetnames.count("Payment") == 1
    finally:
        wb.close()
