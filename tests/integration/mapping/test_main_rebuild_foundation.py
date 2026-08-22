from pathlib import Path

from openpyxl import Workbook, load_workbook

from progress_studio.domain.mapping_models import ActivityRow, BOQRow
from progress_studio.infrastructure.excel.mapping_reader import ProgressActivityReader
from progress_studio.services.mapping_store import MappingStore
from progress_studio.services.workbook_export_service import WorkbookExportService


def make_workbook(path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "main"
    headers = [
        "Row Type", "WBS", "Description", "P/A", "Activity ID", "Task ID", "UID",
        "Outline Level", "Plan Start", "Plan Finish", "Actual Start", "Actual Finish",
        "% Complete", "Physical %", "Amount", "Total Float (hr)", "XML Amount", "W1",
    ]
    ws.append(headers)
    ws.append(["Project Summary", "", "Project", "P", "", "", "", 0, None, None, None, None, None, None, 0])
    ws.append([None, None, None, "A", None, None, None, None, None, None, None, None, None, None, "=O2"])
    ws.append(["WBS", "1", "Structure", "P", "", 1, 1, 1, None, None, None, None, None, None, 0])
    ws.append([None, None, None, "A", None, None, None, None, None, None, None, None, None, None, "=O4"])
    ws.append(["Activity", "1.1", "Existing", "P", "A1000", 2, 2, 2, "2026-01-01", "2026-01-05", None, None, None, None, 0])
    ws.append([None, None, None, "A", "A1000", None, None, None, None, None, None, None, None, None, "=O6"])

    amount = wb.create_sheet("Amount Mapping")
    amount.append(["Activity ID", "WBS", "Description", "Amount", "Status"])
    amount.append([None, "1", "Structure", 0, ""])
    amount.append(["A1000", "1.1", "Existing", 0, ""])
    wb.save(path)
    return path


def test_rebuild_inserts_created_activity_with_blank_plan_dates(tmp_path: Path) -> None:
    source = make_workbook(tmp_path / "source.xlsx")
    store = MappingStore()
    store.load_activities(ProgressActivityReader().read(source))
    store.load_boq([BOQRow("B1", "BOQ", 2, "S", "", "", "Draft work", 125.0)])
    store.add_supplemental_activity(
        parent_path=(("1", "Structure"),),
        wbs_code="1",
        wbs_name="Structure",
        activity_id="A1010",
        description="Draft Activity",
    )
    store.toggle_boq("B1")
    store.map_selected(100)

    output = tmp_path / "rebuilt.xlsx"
    WorkbookExportService().export(source, output, store)

    wb = load_workbook(output, data_only=False)
    try:
        ws = wb["main"]
        rows = {
            ws.cell(row, 5).value: row
            for row in range(2, ws.max_row + 1)
            if ws.cell(row, 4).value == "P" and ws.cell(row, 5).value
        }
        created_row = rows["A1010"]
        assert ws.cell(created_row, 9).value is None
        assert ws.cell(created_row, 10).value is None
        assert ws.cell(created_row, 15).value == 125.0
        existing_row = rows["A1000"]
        assert ws.cell(existing_row, 9).value == "2026-01-01"
        assert ws.cell(existing_row, 10).value == "2026-01-05"
    finally:
        wb.close()
