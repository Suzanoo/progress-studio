from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from progress_studio.infrastructure.excel.payment_input_workbook import PaymentInputWorkbook
from progress_studio.services.rebuild_service import WorkbookRebuildEngine
from tests.integration.payment.test_payment_line_renderer import _progress_workbook


def _embedded_payment_workbook(path: Path) -> Path:
    source = _progress_workbook(path)

    wb = load_workbook(source)
    PaymentInputWorkbook().embed(wb, periods=3)

    # Create generated Progress sheets with unmistakable payloads. RB4 must
    # preserve them byte-logically at the cell contract level.
    for name in ("main_monthly", "progress", "progress_table", "Dashboard_Data", "Dashboard"):
        if name in wb.sheetnames:
            del wb[name]
        ws = wb.create_sheet(name)
        ws["A1"] = f"KEEP::{name}"
        ws["B2"] = 123.456

    stale = wb.create_sheet("Payment")
    stale["A1"] = "STALE_PAYMENT"
    wb.save(source)
    wb.close()
    return source


def test_rb4_rebuild_payment_replaces_payment_only(tmp_path: Path) -> None:
    source = _embedded_payment_workbook(tmp_path / "source.xlsx")
    output = tmp_path / "payment_rebuilt.xlsx"

    result = WorkbookRebuildEngine().rebuild_payment(source, output)

    assert result.rebuilt_sheets == ("Payment",)
    assert result.rendered_periods >= 1
    assert result.rendered_points >= 1
    assert result.period_ids

    wb = load_workbook(output, data_only=False)
    try:
        # Progress generated sheets are untouched.
        for name in ("main_monthly", "progress", "progress_table", "Dashboard_Data", "Dashboard"):
            assert wb[name]["A1"].value == f"KEEP::{name}"
            assert wb[name]["B2"].value == 123.456

        # Persistent inputs are untouched.
        assert "main" in wb.sheetnames
        assert "Payment Input" in wb.sheetnames

        # Payment is actually regenerated, not patched.
        assert wb["Payment"]["A1"].value != "STALE_PAYMENT"
        assert len(wb["Payment"]._images) == result.rendered_periods
    finally:
        wb.close()


def test_rb4_same_path_rebuild_is_atomic_and_preserves_progress(tmp_path: Path) -> None:
    source = _embedded_payment_workbook(tmp_path / "same.xlsx")

    result = WorkbookRebuildEngine().rebuild_payment(source, source)
    assert result.output_workbook == source.resolve()

    wb = load_workbook(source, data_only=False)
    try:
        assert wb["main_monthly"]["A1"].value == "KEEP::main_monthly"
        assert wb["progress"]["A1"].value == "KEEP::progress"
        assert wb["progress_table"]["A1"].value == "KEEP::progress_table"
        assert wb["Dashboard_Data"]["A1"].value == "KEEP::Dashboard_Data"
        assert wb["Dashboard"]["A1"].value == "KEEP::Dashboard"
        assert wb["Payment"]["A1"].value != "STALE_PAYMENT"
    finally:
        wb.close()
