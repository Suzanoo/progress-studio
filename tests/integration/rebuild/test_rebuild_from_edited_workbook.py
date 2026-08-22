from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

from progress_studio.infrastructure.excel.edited_workbook_migrator import migrate_edited_main_into_workbook


def _make_main(path: Path, *, activity_id="A1000", description="Activity One", amount=100.0, plan=(0.4, 0.6), actual=(None, None)):
    wb = Workbook()
    ws = wb.active
    ws.title = "main"
    ws.cell(3, 18).value = "W1"
    ws.cell(3, 19).value = "W2"
    headers = [
        "Row Type", "WBS", "Description", "P/A", "Activity ID", "Task ID", "UID",
        "Outline Level", "Plan Start", "Plan Finish", "Actual Start", "Actual Finish",
        "% Complete", "Physical %", "Amount", "Total Float (hr)", "XML Amount",
        datetime(2026, 1, 30), datetime(2026, 2, 6),
    ]
    for c, value in enumerate(headers, 1):
        ws.cell(4, c).value = value
    for r, pa in ((5, "P"), (6, "A")):
        ws.cell(r, 1).value = "Activity"
        ws.cell(r, 2).value = "1.1"
        ws.cell(r, 3).value = description
        ws.cell(r, 4).value = pa
        ws.cell(r, 5).value = activity_id
        ws.cell(r, 8).value = 2
        ws.cell(r, 9).value = datetime(2026, 1, 1)
        ws.cell(r, 10).value = datetime(2026, 2, 28)
    ws.cell(5, 15).value = amount
    for offset, value in enumerate(plan, 18):
        ws.cell(5, offset).value = value
    for offset, value in enumerate(actual, 18):
        ws.cell(6, offset).value = value
    wb.save(path)
    wb.close()


def test_migrates_amount_plan_actual_by_activity_id(tmp_path):
    edited = tmp_path / "edited.xlsx"
    target = tmp_path / "target.xlsx"
    _make_main(edited, amount=987.5, plan=(0.25, 0.75), actual=(0.1, 0.2))
    _make_main(target, amount=100.0, plan=(0.5, 0.5), actual=(None, None))
    wb = load_workbook(target)
    result = migrate_edited_main_into_workbook(wb, edited)
    ws = wb["main"]
    assert result.matched_activity_count == 1
    assert result.matched_by_activity_id == 1
    assert ws.cell(5, 15).value == 987.5
    assert [ws.cell(5, c).value for c in (18, 19)] == [0.25, 0.75]
    assert [ws.cell(6, c).value for c in (18, 19)] == [0.1, 0.2]
    wb.close()


def test_falls_back_to_description_and_dates_when_activity_id_changed(tmp_path):
    edited = tmp_path / "edited.xlsx"
    target = tmp_path / "target.xlsx"
    _make_main(edited, activity_id="OLD100", amount=321.0, actual=(0.15, None))
    _make_main(target, activity_id="NEW200", amount=100.0)
    wb = load_workbook(target)
    result = migrate_edited_main_into_workbook(wb, edited)
    assert result.matched_by_activity_id == 0
    assert result.matched_by_signature == 1
    assert wb["main"].cell(5, 15).value == 321.0
    assert wb["main"].cell(6, 18).value == 0.15
    wb.close()


def test_only_intersecting_week_dates_are_migrated(tmp_path):
    edited = tmp_path / "edited.xlsx"
    target = tmp_path / "target.xlsx"
    _make_main(edited, plan=(0.2, 0.8), actual=(0.1, 0.4))
    _make_main(target, plan=(0.5, 0.5), actual=(None, None))
    wb = load_workbook(target)
    # Change target W2 date so only W1 intersects.
    wb["main"].cell(4, 19).value = datetime(2026, 2, 13)
    result = migrate_edited_main_into_workbook(wb, edited)
    assert result.plan_cells_migrated == 1
    assert result.actual_cells_migrated == 1
    assert wb["main"].cell(5, 18).value == 0.2
    assert wb["main"].cell(5, 19).value == 0.5
    wb.close()


def test_accepts_legacy_main_sheet_name(tmp_path):
    edited = tmp_path / "edited.xlsx"
    target = tmp_path / "target.xlsx"
    _make_main(edited, amount=444.0, actual=(0.05, 0.15))
    wb_old = load_workbook(edited)
    wb_old["main"].title = "NKC2-R03"
    wb_old.save(edited)
    wb_old.close()
    _make_main(target)
    wb = load_workbook(target)
    result = migrate_edited_main_into_workbook(wb, edited)
    assert result.matched_activity_count == 1
    assert wb["main"].cell(5, 15).value == 444.0
    wb.close()
