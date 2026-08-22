from __future__ import annotations

from openpyxl import Workbook, load_workbook

from progress_studio.infrastructure.excel.final_workbook_policy import finalize_workbook
from progress_studio.services.monthly_main_service import MonthlyMainService
from tests.integration.create_progress.test_create_progress_finalization import _ready_create_input


def test_shared_final_policy_does_not_force_first_open_recalc() -> None:
    """The first-open override belongs only to Create Progress, not shared policy."""
    wb = Workbook()
    finalize_workbook(wb, mode="snapshot", include_guide=False)
    assert wb.calculation.calcMode == "manual"
    assert wb.calculation.calcOnSave is True
    assert wb.calculation.fullCalcOnLoad is False
    assert wb.calculation.forceFullCalc is False


def test_create_progress_requests_first_open_full_calc_but_stays_manual(tmp_path) -> None:
    source = _ready_create_input(tmp_path)
    output = tmp_path / "created.xlsx"

    MonthlyMainService().build(source, output)

    wb = load_workbook(output, data_only=False)
    try:
        calc = wb.calculation
        assert calc.calcMode == "manual"
        assert calc.calcOnSave is True
        assert calc.fullCalcOnLoad is True
        assert calc.forceFullCalc is True
        assert calc.calcId == 0
    finally:
        wb.close()


def test_create_progress_first_open_recalc_flags_are_serialized_to_ooxml(tmp_path) -> None:
    from zipfile import ZipFile
    from xml.etree import ElementTree as ET

    source = _ready_create_input(tmp_path)
    output = tmp_path / "created.xlsx"
    MonthlyMainService().build(source, output)

    with ZipFile(output) as zf:
        root = ET.fromstring(zf.read("xl/workbook.xml"))
    ns = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    calc = root.find("m:calcPr", ns)
    assert calc is not None
    assert calc.attrib.get("calcMode") == "manual"
    assert calc.attrib.get("calcOnSave") == "1"
    assert calc.attrib.get("fullCalcOnLoad") == "1"
    assert calc.attrib.get("forceFullCalc") == "1"
    assert calc.attrib.get("calcId") == "0"
