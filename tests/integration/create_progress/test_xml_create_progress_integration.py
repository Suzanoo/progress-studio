from __future__ import annotations

from datetime import date
from pathlib import Path

from tests._paths import FIXTURES_ROOT

from openpyxl import load_workbook

from progress_studio.infrastructure.excel.distribution_workbook import calculation_week_columns
from progress_studio.infrastructure.excel.import_workbook_writer import ImportWorkbookWriter
from progress_studio.infrastructure.schedule_xml import NormalizedScheduleXmlReader, ScheduleXmlReader
from progress_studio.services.import_service import ImportService
from progress_studio.services.schedule_service import ScheduleService


FIXTURES = FIXTURES_ROOT / "xml"
MSP = FIXTURES / "msp_n3.xml"
P6 = FIXTURES / "p6_n5.xml"


def test_n7_msp_production_bridge_preserves_legacy_reader_rows() -> None:
    project_legacy, legacy = ScheduleXmlReader().read(MSP)
    project_new, normalized = NormalizedScheduleXmlReader().read(MSP)

    assert project_new == project_legacy
    assert len(normalized) == len(legacy)
    for before, after in zip(legacy, normalized, strict=True):
        assert after.source_order == before.source_order
        assert after.is_summary == before.is_summary
        assert after.activity_id == before.activity_id
        assert after.name == before.name
        assert after.wbs == before.wbs
        assert after.outline_level == before.outline_level
        assert after.plan_start == before.plan_start
        assert after.plan_finish == before.plan_finish
        assert after.actual_start == before.actual_start
        assert after.actual_finish == before.actual_finish
        assert after.percent_complete == before.percent_complete
        assert after.physical_percent_complete == before.physical_percent_complete
        # Normalization owns schedule only; source Cost/Amount never crosses this boundary.
        assert after.amount is None


def test_n7_p6_bridge_rebuilds_real_wbs_and_preserves_activity_ids() -> None:
    project, rows = NormalizedScheduleXmlReader().read(P6)
    assert project
    assert any(row.is_summary for row in rows)
    assert any(not row.is_summary for row in rows)

    by_id = {row.activity_id: row for row in rows if not row.is_summary}
    wbs_codes = {row.wbs for row in rows if row.is_summary}

    assert by_id["A1290"].wbs == "2.1.1"
    assert by_id["A1310"].wbs == "2.1.1.1"
    assert "2.1.1" in wbs_codes
    assert "2.1.1.1" in wbs_codes
    assert all(not row.activity_id.startswith("ACT-") for row in rows if not row.is_summary)


def test_n7_import_service_accepts_p6_through_normalized_boundary(tmp_path: Path) -> None:
    output = tmp_path / "p6_imported.xlsx"
    service = ImportService(
        NormalizedScheduleXmlReader(),
        ScheduleService(),
        ImportWorkbookWriter(),
    )

    _, wbs_count, activity_count = service.import_xml(P6, output)
    assert wbs_count > 0
    assert activity_count > 0

    wb = load_workbook(output, data_only=False)
    try:
        ws = wb["main"]
        headers = {str(ws.cell(1, c).value or "").strip(): c for c in range(1, ws.max_column + 1)}
        row_type_col = headers["Row Type"]
        activity_id_col = headers["Activity ID"]
        values = [
            (str(ws.cell(r, row_type_col).value or ""), str(ws.cell(r, activity_id_col).value or ""))
            for r in range(2, ws.max_row + 1)
        ]
        assert any(kind == "WBS" for kind, _ in values)
        assert any(kind == "Activity" and aid == "A1290" for kind, aid in values)
        assert not any(aid.startswith("ACT-") for kind, aid in values if kind == "Activity")
    finally:
        wb.close()


def test_n7_margin_weeks_are_display_only_not_calculation_periods() -> None:
    # Visible weekly timescale includes 4-week margins around a March schedule.
    display_weeks = [
        (20, date(2026, 2, 6)),
        (21, date(2026, 2, 13)),
        (22, date(2026, 2, 20)),
        (23, date(2026, 2, 27)),
        (24, date(2026, 3, 6)),
        (25, date(2026, 3, 13)),
        (26, date(2026, 3, 20)),
        (27, date(2026, 3, 27)),
        (28, date(2026, 4, 3)),
        (29, date(2026, 4, 10)),
        (30, date(2026, 4, 17)),
        (31, date(2026, 4, 24)),
    ]

    calculation = calculation_week_columns(
        display_weeks,
        schedule_start=date(2026, 3, 1),
        schedule_finish=date(2026, 3, 31),
    )

    assert [col for col, _ in calculation] == [24, 25, 26, 27, 28]
    assert {20, 21, 22, 23, 29, 30, 31}.isdisjoint({col for col, _ in calculation})
