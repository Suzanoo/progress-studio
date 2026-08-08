from __future__ import annotations

from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook

from progress_studio.infrastructure.excel.payment_input_reader import PaymentInputSparseReader
from progress_studio.infrastructure.excel.payment_input_workbook import PaymentInputWorkbook
from progress_studio.infrastructure.excel.payment_progress_index import ActivityProgressIndexReader
from progress_studio.services.payment_position_engine import PaymentPositionEngine


def _progress_workbook(path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "main"
    headers = [
        "Row Type", "WBS", "Description", "P/A", "Activity ID", "Task ID", "UID", "Outline Level",
        "Plan Start", "Plan Finish", "Actual Start", "Actual Finish", "% Complete", "Physical %",
        "Amount", "Total Float (hr)", "XML Amount",
    ]
    for col, value in enumerate(headers, start=1):
        ws.cell(4, col, value)
    week0 = datetime(2026, 3, 2)
    for offset in range(5):
        ws.cell(4, 18 + offset, week0 + timedelta(days=7 * offset))

    ws.cell(5, 1, "Project Summary")
    ws.cell(5, 3, "Demo")
    ws.cell(5, 4, "P")
    ws.cell(5, 9, datetime(2026, 3, 2))
    ws.cell(5, 10, datetime(2026, 4, 3))

    # Plan row. Weekly values are incremental and sum to 100%.
    ws.cell(6, 1, "Activity")
    ws.cell(6, 3, "Foundation")
    ws.cell(6, 4, "P")
    ws.cell(6, 5, "A1000")
    ws.cell(6, 9, datetime(2026, 3, 2))
    ws.cell(6, 10, datetime(2026, 4, 3))
    for col, value in zip(range(18, 23), [0.10, 0.15, 0.20, 0.20, 0.35]):
        ws.cell(6, col, value)

    # Actual partner row with same Activity ID must not enter the Plan index.
    ws.cell(7, 4, "A")
    ws.cell(7, 5, "A1000")
    for col, value in zip(range(18, 23), [0.10, 0.10, 0.10, 0.10, 0.10]):
        ws.cell(7, col, value)

    ws.cell(8, 1, "Activity")
    ws.cell(8, 3, "Structure")
    ws.cell(8, 4, "P")
    ws.cell(8, 5, "A1010")
    ws.cell(8, 9, datetime(2026, 3, 16))
    ws.cell(8, 10, datetime(2026, 3, 30))
    for col, value in zip(range(20, 23), [0.25, 0.25, 0.50]):
        ws.cell(8, col, value)

    wb.save(path)
    wb.close()
    return path


def _payment_input(progress: Path, output: Path) -> Path:
    PaymentInputWorkbook().create(progress, output, 3)
    wb = load_workbook(output)
    ws = wb["Payment Input"]
    # P01: explicit 0% must be retained; P02 has a mid-progress target; P03 = 100%.
    ws["B8"] = 0.0
    ws["C8"] = 0.60
    ws["D8"] = 1.0
    # Only one period populated for A1010; all other cells remain physically blank.
    ws["C9"] = 0.25
    wb.save(output)
    wb.close()
    return output


def test_sparse_reader_keeps_only_populated_requirements_and_preserves_zero(tmp_path: Path) -> None:
    progress = _progress_workbook(tmp_path / "progress.xlsx")
    payment = _payment_input(progress, tmp_path / "payment.xlsx")

    data = PaymentInputSparseReader().read(payment)

    assert len(data.periods) == 3
    assert data.populated_requirements == 4
    assert [r.activity_id for r in data.periods[0].requirements] == ["A1000"]
    assert data.periods[0].requirements[0].required_fraction == 0.0
    assert {r.activity_id for r in data.periods[1].requirements} == {"A1000", "A1010"}
    assert data.periods[2].requirements[0].required_fraction == 1.0


def test_activity_progress_index_reads_plan_rows_once_and_normalizes_cumulative(tmp_path: Path) -> None:
    progress = _progress_workbook(tmp_path / "progress.xlsx")
    index = ActivityProgressIndexReader().read(progress)

    assert set(index.activities) == {"A1000", "A1010"}
    activity = index.activities["A1000"]
    assert activity.row_number == 6
    assert len(activity.buckets) == 5
    assert [round(p.cumulative_fraction, 2) for p in activity.buckets] == [0.10, 0.25, 0.45, 0.65, 1.00]


def test_position_engine_uses_cell_boundaries_and_first_week_reaching_requirement(tmp_path: Path) -> None:
    progress_path = _progress_workbook(tmp_path / "progress.xlsx")
    payment_path = _payment_input(progress_path, tmp_path / "payment.xlsx")
    payment = PaymentInputSparseReader().read(payment_path)
    progress = ActivityProgressIndexReader().read(progress_path)

    result = PaymentPositionEngine().resolve(payment, progress)
    by_period = {period.period_id: period for period in result.periods}

    p01 = by_period["P01"].points[0]
    assert p01.activity_id == "A1000"
    assert p01.timescale_column_letter == "R"
    assert p01.boundary_edge == "left"

    p02 = {point.activity_id: point for point in by_period["P02"].points}
    assert p02["A1000"].timescale_column_letter == "U"  # cumulative 65% first reaches 60%
    assert p02["A1000"].boundary_edge == "right"
    assert p02["A1010"].timescale_column_letter == "T"  # first plan bucket reaches 25%

    p03 = by_period["P03"].points[0]
    assert p03.timescale_column_letter == "V"
    assert p03.boundary_edge == "right"
    assert result.requirement_count == 4
    assert result.resolved_count == 4
    assert not result.issues


def test_payment_service_prepares_uploaded_input_with_one_sparse_pipeline(tmp_path: Path) -> None:
    from progress_studio.services.payment_service import PaymentService

    progress_path = _progress_workbook(tmp_path / "progress.xlsx")
    payment_path = _payment_input(progress_path, tmp_path / "payment.xlsx")

    prepared = PaymentService().prepare_payment_input(progress_path, payment_path)

    assert prepared.validation.matched_activities == 2
    assert prepared.validation.missing_activities == 0
    assert prepared.validation.populated_requirements == 4
    assert prepared.positions.requirement_count == 4
    assert prepared.positions.resolved_count == 4
    assert not prepared.positions.issues
