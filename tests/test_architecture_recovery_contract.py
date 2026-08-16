from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from progress_studio.infrastructure.excel.final_workbook_policy import finalize_workbook
from progress_studio.services.monthly_main_service import MonthlyMainService
from progress_studio.services.rebuild_service import WorkbookRebuildEngine
from tests.test_ms_rb2_progress_rebuild_engine import _full_rebuild_fixture
from tests.test_lw9_live_payment_integration import _fixture as _live_payment_fixture


def test_final_policy_is_manual_f9_save_for_snapshot_and_live() -> None:
    from openpyxl import Workbook

    for mode in ("snapshot", "live"):
        wb = Workbook()
        wb.active.title = "main"
        finalize_workbook(wb, mode=mode, include_guide=False)
        calc = wb.calculation
        assert calc.calcMode == "manual"
        assert calc.calcOnSave is True
        assert calc.fullCalcOnLoad is False
        assert calc.forceFullCalc is False
        assert calc.calcId != 0


def test_create_progress_final_boundary_is_manual_and_seeds_explicit_overlay_zero(tmp_path: Path) -> None:
    # Use the established realistic workbook fixture; this exercises the same
    # final Create boundary as the desktop pipeline without adding another open.
    source = _full_rebuild_fixture(tmp_path / "source.xlsx")
    prepared = tmp_path / "prepared.xlsx"
    WorkbookRebuildEngine().rebuild_progress(source, prepared, project_name="Architecture Recovery")
    output = tmp_path / "created.xlsx"
    MonthlyMainService().build(prepared, output)

    wb = load_workbook(output, data_only=False)
    try:
        assert wb.calculation.calcMode == "manual"
        assert wb.calculation.calcOnSave is True
        data = wb["Dashboard_Data"]
        # Each overlay chart includes a synthetic chart-only point immediately
        # before its real project window. Both Plan and Actual must be explicit 0.
        for sheet_name, plan_col, actual_col in (("main", 21, 22), ("main_monthly", 25, 26)):
            chart = wb[sheet_name]._charts[0]
            import re
            plan_formula = chart.series[0].val.numRef.f
            rows = re.findall(r"\$(\d+)", plan_formula)
            assert rows and int(rows[0]) == 2
            assert data.cell(2, plan_col).value == 0
            assert data.cell(2, actual_col).value == 0
            assert chart.display_blanks == "gap"
    finally:
        wb.close()


def test_in_memory_payment_renderer_is_render_only() -> None:
    source = Path("progress_studio/infrastructure/excel/payment_line_renderer.py").read_text(encoding="utf-8")
    start = source.index("    def render_periods_into_workbook(")
    end = source.index("    def render_single_period(", start)
    block = source[start:end]
    assert "load_workbook" not in block
    assert "finalize_workbook(" not in block
    assert ".save(" not in block
    assert "finalize_mode" not in block
    assert "save_path" not in block


def test_live_payment_rebuild_owns_one_finalize_and_one_save(tmp_path: Path) -> None:
    source = _live_payment_fixture(tmp_path / "source.xlsx")
    output = tmp_path / "out.xlsx"
    WorkbookRebuildEngine().rebuild_live_payment(source, output)

    wb = load_workbook(output, data_only=False)
    try:
        assert "Payment" in wb.sheetnames
        assert wb["Payment"].protection.sheet is True
        assert wb.calculation.calcMode == "manual"
        assert wb.calculation.calcOnSave is True
    finally:
        wb.close()

    text = Path("progress_studio/services/rebuild_service.py").read_text(encoding="utf-8")
    start = text.index("    def rebuild_live_payment(")
    end = text.index("    def rebuild_payment(", start)
    block = text[start:end]
    assert block.count("finalize_workbook(") == 1
    assert block.count(".save(") == 1
    assert block.count("load_workbook(") == 1


def test_snapshot_payment_rebuild_does_not_finalize_twice() -> None:
    text = Path("progress_studio/services/rebuild_service.py").read_text(encoding="utf-8")
    start = text.index("    def rebuild_payment(")
    end = text.index("    def generated_sheets_for(", start)
    block = text[start:end]
    # Snapshot Payment delegates to the standalone PaymentService boundary,
    # which performs the one finalization/save. Rebuild itself must not repeat it.
    assert "finalize_workbook(" not in block
    assert "render_payment_backbones" in block
