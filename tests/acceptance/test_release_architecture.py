import contextlib
import io
import tempfile
import unittest
from pathlib import Path

from tests._paths import REPO_ROOT

from openpyxl import load_workbook

from progress_studio.app.context import PipelineContext
from progress_studio.infrastructure.excel import ImportWorkbookWriter
from progress_studio.infrastructure.primavera import PrimaveraXmlReader
from progress_studio.pipeline.amount_step import AmountStep
from progress_studio.pipeline.distribution_step import DistributionStep
from progress_studio.pipeline.import_step import ImportStep
from progress_studio.pipeline.okd_step import OkdStep
from progress_studio.pipeline.progress_step import ProgressStep
from progress_studio.pipeline.schedule_step import ScheduleStep
from progress_studio.pipeline.timescale_step import TimescaleStep
from progress_studio.services.amount_service import AmountService
from progress_studio.services.distribution_service import DistributionService
from progress_studio.services.import_service import ImportService
from progress_studio.services.okd_service import OkdService
from progress_studio.services.progress_service import ProgressService
from progress_studio.services.schedule_service import ScheduleService
from progress_studio.services.schedule_workbook_service import ScheduleWorkbookService
from progress_studio.services.timescale_service import TimescaleService

ROOT = REPO_ROOT
XML = ROOT / "example" / "example.xml"


class FixedPrompt:
    def choose_method(self): return "flat"
    def review(self, output_file, method): return "exit"


class Ms6AcceptanceTests(unittest.TestCase):
    def test_legacy_package_and_scripts_are_removed(self):
        self.assertFalse((ROOT / "progress_studio" / "legacy").exists())
        for number in range(1, 8):
            self.assertFalse(any(ROOT.glob(f"{number:02d}_*.py")))

    def test_bootstrap_uses_okd_step_not_legacy_adapter(self):
        text = (ROOT / "progress_studio/app/bootstrap.py").read_text(encoding="utf-8")
        self.assertIn("OkdStep(OkdService())", text)
        self.assertNotIn("LegacyPipelineStep", text)

    def test_context_exposes_okd_workbook(self):
        self.assertIn("okd_workbook", PipelineContext.__dataclass_fields__)

    def test_no_subprocess_in_application_package(self):
        for path in (ROOT / "progress_studio").rglob("*.py"):
            self.assertNotIn("subprocess", path.read_text(encoding="utf-8"), str(path))

    def test_no_thai_in_source(self):
        for path in (ROOT / "progress_studio").rglob("*.py"):
            text = path.read_text(encoding="utf-8")
            self.assertFalse(any("\u0e00" <= ch <= "\u0e7f" for ch in text), str(path))

    def test_end_to_end_steps_create_okd_sheets(self):
        with tempfile.TemporaryDirectory() as td:
            folder = Path(td)
            context = PipelineContext(XML, "5", 1000.0, project_folder=folder, working_folder=folder)
            with contextlib.redirect_stdout(io.StringIO()):
                ImportStep(ImportService(PrimaveraXmlReader(), ScheduleService(), ImportWorkbookWriter())).execute(context)
                ScheduleStep(ScheduleWorkbookService()).execute(context)
                TimescaleStep(TimescaleService()).execute(context)
                AmountStep(AmountService()).execute(context)
                ProgressStep(ProgressService()).execute(context)
                DistributionStep(DistributionService(), FixedPrompt()).execute(context)
                OkdStep(OkdService()).execute(context)

            self.assertEqual(context.output_workbook, context.distribution_workbook)
            self.assertTrue(context.output_workbook.is_file())
            wb = load_workbook(context.output_workbook, data_only=False)
            try:
                self.assertIn("main", wb.sheetnames)
                self.assertIn("progress", wb.sheetnames)
                self.assertIn("progress_table", wb.sheetnames)
                self.assertGreater(wb["progress"].max_row, 1)
                self.assertGreater(wb["progress_table"].max_row, 1)
            finally:
                wb.close()


if __name__ == "__main__":
    unittest.main()
