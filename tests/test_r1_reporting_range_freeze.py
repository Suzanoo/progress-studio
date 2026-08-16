from __future__ import annotations

from datetime import date, datetime
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from progress_studio.domain.main_dataset import MainDataset, MainPeriod, MainRow
from progress_studio.infrastructure.excel.dashboard_workbook import _progress_rows
from progress_studio.infrastructure.excel.live_dashboard_workbook import _build_live_data_sheet
from progress_studio.infrastructure.excel.rebuild_workbook_reader import RebuildWorkbookReader
from progress_studio.services.progress_cache_deriver import ProgressCacheDeriver
from progress_studio.services.rebuild_service import WorkbookRebuildEngine
from progress_studio.services.reporting_period_selector import select_reporting_periods
from tests.test_rebuild_2x2_matrix import _matrix_source


def _dataset_with_display_margin() -> MainDataset:
    periods = tuple(
        MainPeriod(20 + i, f"W{i + 1}", datetime.combine(value, datetime.min.time()))
        for i, value in enumerate(
            (
                date(2026, 4, 3),   # pre-project margin
                date(2026, 4, 10),  # pre-project margin
                date(2026, 4, 17),  # project start reporting week
                date(2026, 4, 24),
                date(2026, 5, 1),
                date(2026, 5, 8),
                date(2026, 6, 4),   # 29-May..04-Jun overlaps 31-May finish
                date(2026, 6, 11),  # post-project margin
                date(2026, 6, 18),  # post-project margin
            )
        )
    )
    plan = MainRow(
        row_number=10,
        row_type="Activity",
        pa="P",
        wbs="1",
        description="Test Activity",
        activity_id="A1000",
        outline_level=1,
        plan_start=datetime(2026, 4, 17, 8),
        plan_finish=datetime(2026, 5, 31, 17),
        amount=100.0,
        percent_complete=0.0,
        period_values=tuple((period.column, None) for period in periods),
    )
    actual = MainRow(
        row_number=11,
        row_type="Activity",
        pa="A",
        wbs="1",
        description="Test Activity",
        activity_id="A1000",
        outline_level=1,
        plan_start=datetime(2026, 4, 17, 8),
        plan_finish=datetime(2026, 5, 31, 17),
        amount=100.0,
        percent_complete=0.0,
        period_values=tuple((period.column, None) for period in periods),
    )
    return MainDataset(
        workbook_name="margin.xlsx",
        header_row=4,
        headers=(("row type", 1), ("p/a", 2), ("activity id", 3)),
        periods=periods,
        rows=(plan, actual),
    )


def test_r1_shared_selector_excludes_both_margins_and_keeps_final_overlap() -> None:
    dataset = _dataset_with_display_margin()
    selected = select_reporting_periods(dataset)
    dates = [ref.period.reporting_date.date() for ref in selected]
    assert dates == [
        date(2026, 4, 17),
        date(2026, 4, 24),
        date(2026, 5, 1),
        date(2026, 5, 8),
        date(2026, 6, 4),
    ]


def test_r1_create_dashboard_and_live_dashboard_share_the_same_weekly_boundary() -> None:
    dataset = _dataset_with_display_margin()
    selected = select_reporting_periods(dataset)
    selected_dates = [ref.period.reporting_date.date() for ref in selected]

    # Create/Snapshot dashboard path: progress carries display margin, selector drops it.
    wb = Workbook()
    progress = wb.active
    progress.title = "progress"
    progress.append(["project_start", "project_finish", "week_start", "plan", "actual"])
    for period in dataset.periods:
        progress.append([
            date(2026, 4, 17),
            date(2026, 5, 31),
            period.reporting_date.date(),
            None,
            None,
        ])
    assert [value for _, value in _progress_rows(wb, progress)] == selected_dates

    # Live dashboard path: Dashboard_Data may link to sparse original progress rows,
    # but it must physically store only project reporting periods.
    dashboard = wb.create_sheet("Dashboard")
    dashboard["G5"] = "Weekly"
    dashboard["K5"] = date(2026, 5, 8)
    del wb["progress"]
    progress = wb.create_sheet("progress")
    progress.append(["Date", "Plan", "Actual"])
    for period in dataset.periods:
        progress.append([period.reporting_date, 0.0, None])
    _build_live_data_sheet(wb, dataset, ProgressCacheDeriver().derive(dataset))
    data = wb["Dashboard_Data"]
    weekly_links = [data.cell(row, 1).value for row in range(2, data.max_row + 1) if data.cell(row, 1).value]
    expected_rows = [ref.index + 2 for ref in selected]
    assert weekly_links == [f"='progress'!A{row}" for row in expected_rows]
    assert len(weekly_links) == len(selected_dates)
    # Monthly must also derive only from the selected project reporting periods.
    monthly_dates = [data.cell(row, 4).value for row in range(2, data.max_row + 1) if data.cell(row, 4).value]
    assert monthly_dates == [date(2026, 4, 24), date(2026, 5, 8), date(2026, 6, 4)]
    wb.close()


@pytest.mark.parametrize("mode", ["snapshot", "live"])
def test_r1_progress_rebuild_dashboard_data_contains_project_periods_only(tmp_path: Path, mode: str) -> None:
    source = _matrix_source(tmp_path / "source.xlsx")
    dataset = RebuildWorkbookReader().read_main_dataset(source)
    expected = select_reporting_periods(dataset)
    assert expected

    output = tmp_path / f"progress_{mode}.xlsx"
    engine = WorkbookRebuildEngine()
    if mode == "snapshot":
        engine.rebuild_progress(source, output, project_name="R1 Reporting Range")
    else:
        engine.rebuild_live_progress(source, output, project_name="R1 Reporting Range")

    wb = load_workbook(output, data_only=False)
    try:
        data = wb["Dashboard_Data"]
        weekly_cells = [data.cell(row, 1).value for row in range(2, data.max_row + 1) if data.cell(row, 1).value]
        assert len(weekly_cells) == len(expected)
        # The next Weekly source row must be physically empty: margin is not stored
        # and therefore cannot leak back into chart/KPI ranges later.
        assert data.cell(len(expected) + 2, 1).value is None
    finally:
        wb.close()


def _dashboard_data_signature(path: Path) -> tuple[tuple[object, ...], ...]:
    wb = load_workbook(path, data_only=False)
    try:
        ws = wb["Dashboard_Data"]
        return tuple(
            tuple(ws.cell(row, col).value for col in range(1, 16))
            for row in range(1, ws.max_row + 1)
        )
    finally:
        wb.close()


@pytest.mark.parametrize("mode", ["snapshot", "live"])
def test_r1_payment_rebuild_does_not_rebuild_or_expand_dashboard_data(tmp_path: Path, mode: str) -> None:
    source = _matrix_source(tmp_path / "source.xlsx")
    before = _dashboard_data_signature(source)
    output = tmp_path / f"payment_{mode}.xlsx"
    engine = WorkbookRebuildEngine()
    if mode == "snapshot":
        engine.rebuild_payment(source, output)
    else:
        engine.rebuild_live_payment(source, output)
    assert _dashboard_data_signature(output) == before
