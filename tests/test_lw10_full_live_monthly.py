
from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

from progress_studio.services.rebuild_service import WorkbookRebuildEngine


def _fixture(path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "main"
    headers = [
        "Row Type", "WBS", "Description", "P/A", "Activity ID",
        "Outline Level", "Plan Start", "Plan Finish", "Amount",
    ]
    for col, value in enumerate(headers, 1):
        ws.cell(4, col, value)

    dates = [
        datetime(2026,1,30),
        datetime(2026,2,6),
        datetime(2026,2,13),
        datetime(2026,2,27),
    ]
    for col, dt in enumerate(dates, start=10):
        ws.cell(3, col, f"W{col-9}")
        ws.cell(4, col, dt)

    rows = [
        ["Project Summary", "", "Project", "P", "", 0, None, None, 1000, .10, .20, .30, .40],
        ["", "", "", "A", "", 0, None, None, 0, .05, .10, .15, .20],
        ["WBS", "1", "Structure", "P", "", 1, None, None, 1000, .10, .20, .30, .40],
        ["", "", "", "A", "", 1, None, None, 0, .05, .10, .15, .20],
        ["Activity", "1.1", "Concrete", "P", "A1000", 2, datetime(2026,1,1), datetime(2026,2,27), 1000, .10, .20, .30, .40],
        ["", "", "", "A", "A1000", 2, datetime(2026,1,1), datetime(2026,2,27), 0, .05, .10, .15, .20],
    ]
    for row in rows:
        ws.append(row)
    wb.save(path)
    return path


def test_lw10_full_live_monthly_links_all_plan_actual_wbs_project_rows(tmp_path: Path) -> None:
    source = _fixture(tmp_path / "source.xlsx")
    output = tmp_path / "live.xlsx"
    WorkbookRebuildEngine().rebuild_live_progress(source, output)

    wb = load_workbook(output, data_only=False)
    try:
        ws = wb["main_monthly"]
        # Jan = J only; Feb = K:M. Every visible monthly row is live.
        for row in range(5, 11):
            assert isinstance(ws.cell(row, 10).value, str)
            assert ws.cell(row, 10).value.startswith("=")
            assert isinstance(ws.cell(row, 11).value, str)
            assert ws.cell(row, 11).value.startswith("=")

        assert "'main'!J5:J5" in ws["J5"].value
        assert "'main'!K10:M10" in ws["K10"].value
    finally:
        wb.close()


def test_lw10_monthly_formula_volume_is_full_control_baseline(tmp_path: Path) -> None:
    source = _fixture(tmp_path / "source.xlsx")
    output = tmp_path / "live.xlsx"
    WorkbookRebuildEngine().rebuild_live_progress(source, output)

    wb = load_workbook(output, data_only=False)
    try:
        ws = wb["main_monthly"]
        formulas = 0
        for row in range(5, 11):
            for col in (10, 11):
                if isinstance(ws.cell(row, col).value, str) and ws.cell(row, col).value.startswith("="):
                    formulas += 1
        assert formulas == 12
    finally:
        wb.close()


def test_lw10_full_live_monthly_keeps_manual_calc_on_save(tmp_path: Path) -> None:
    source = _fixture(tmp_path / "source.xlsx")
    output = tmp_path / "live.xlsx"
    WorkbookRebuildEngine().rebuild_live_progress(source, output)

    wb = load_workbook(output, data_only=False)
    try:
        assert wb.calculation.calcMode == "manual"
        assert wb.calculation.calcOnSave is True
    finally:
        wb.close()


def test_lw10_writer_formula_is_deliberately_simple() -> None:
    source = Path(
        "progress_studio/infrastructure/excel/live_monthly_workbook.py"
    ).read_text(encoding="utf-8")
    assert "IF(COUNT(" in source
    assert "SUM(" in source
    assert "SUMIFS(" not in source
    assert "INDEX(" not in source
    assert "MATCH(" not in source
